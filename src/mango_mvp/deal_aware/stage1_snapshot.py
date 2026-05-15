from __future__ import annotations

import csv
import json
import re
import sqlite3
import socket
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from mango_mvp.utils.phone import normalize_phone


SCHEMA_VERSION = "deal_aware_stage1_snapshot_v1"


@dataclass(frozen=True)
class Stage1Paths:
    master_contacts_csv: Path
    master_calls_csv: Path | None
    amo_ready_csv: Path
    calls_csv: Path
    out_root: Path
    current_runtime_json: Path | None = None
    canonical_export_pointer: Path | None = None
    amo_live_snapshot_dir: Path | None = None
    tallanto_students_csv: Path | None = None
    tallanto_writeoff_combined_csv: Path | None = None
    tallanto_writeoff_summary_csv: Path | None = None
    tallanto_schema_json: Path | None = None
    quality_summary_paths: tuple[Path, ...] = ()
    amo_stage_dirs: tuple[Path, ...] = ()
    amo_loss_reason_dirs: tuple[Path, ...] = ()
    tallanto_writeoff_xlsx: tuple[Path, ...] = ()


def build_stage1_snapshot(paths: Stage1Paths) -> dict[str, Any]:
    paths.out_root.mkdir(parents=True, exist_ok=True)

    phone_rollup = build_phone_rollup(paths.master_contacts_csv)
    amo_ready = build_amo_ready_snapshot(paths.amo_ready_csv)
    amo_writebacks = build_amo_writeback_snapshot(paths.amo_stage_dirs)
    amo_loss_reasons = build_amo_loss_reason_snapshot(paths.amo_loss_reason_dirs)
    call_snapshot = build_call_snapshot(paths.master_calls_csv)
    call_rollup = build_call_rollup(paths.calls_csv)
    tallanto_students = build_tallanto_students_snapshot(paths.tallanto_students_csv)
    writeoff_rows = build_tallanto_writeoff_snapshot(
        paths.tallanto_writeoff_xlsx,
        combined_csv=paths.tallanto_writeoff_combined_csv,
    )
    writeoff_summary = build_tallanto_writeoff_summary_snapshot(
        rows=writeoff_rows,
        summary_csv=paths.tallanto_writeoff_summary_csv,
    )
    quality_gates = build_quality_gate_snapshot(paths.quality_summary_paths)
    source_manifest = build_source_manifest(paths)
    runtime_contract = load_json(paths.current_runtime_json) if paths.current_runtime_json else {}

    outputs = {
        "phone_rollup_csv": paths.out_root / "phone_rollup.csv",
        "amo_ready_snapshot_csv": paths.out_root / "amo_ready_snapshot.csv",
        "amo_writeback_snapshot_csv": paths.out_root / "amo_writeback_snapshot.csv",
        "amo_loss_reasons_csv": paths.out_root / "amo_loss_reasons.csv",
        "call_snapshot_csv": paths.out_root / "call_snapshot.csv",
        "call_rollup_csv": paths.out_root / "call_rollup_by_phone.csv",
        "tallanto_students_snapshot_csv": paths.out_root / "tallanto_students_snapshot.csv",
        "tallanto_writeoff_visits_csv": paths.out_root / "tallanto_writeoff_visits.csv",
        "tallanto_writeoff_summary_csv": paths.out_root / "tallanto_writeoff_summary_by_student.csv",
        "quality_gate_snapshot_csv": paths.out_root / "quality_gate_snapshot.csv",
        "source_manifest_csv": paths.out_root / "source_manifest.csv",
        "runtime_contract_json": paths.out_root / "runtime_contract_snapshot.json",
        "sqlite": paths.out_root / "deal_aware_stage1_snapshot.sqlite",
        "summary_json": paths.out_root / "summary.json",
        "report_md": paths.out_root / "README.md",
    }

    write_csv(outputs["phone_rollup_csv"], phone_rollup)
    write_csv(outputs["amo_ready_snapshot_csv"], amo_ready)
    write_csv(outputs["amo_writeback_snapshot_csv"], amo_writebacks)
    write_csv(outputs["amo_loss_reasons_csv"], amo_loss_reasons)
    write_csv(outputs["call_snapshot_csv"], call_snapshot)
    write_csv(outputs["call_rollup_csv"], call_rollup)
    write_csv(outputs["tallanto_students_snapshot_csv"], tallanto_students)
    write_csv(outputs["tallanto_writeoff_visits_csv"], writeoff_rows)
    write_csv(outputs["tallanto_writeoff_summary_csv"], writeoff_summary)
    write_csv(outputs["quality_gate_snapshot_csv"], quality_gates)
    write_csv(outputs["source_manifest_csv"], source_manifest)
    outputs["runtime_contract_json"].write_text(
        json.dumps(runtime_contract, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_sqlite(
        outputs["sqlite"],
        {
            "phone_rollup": phone_rollup,
            "amo_ready_snapshot": amo_ready,
            "amo_writeback_snapshot": amo_writebacks,
            "amo_loss_reasons": amo_loss_reasons,
            "call_snapshot": call_snapshot,
            "call_rollup_by_phone": call_rollup,
            "tallanto_students_snapshot": tallanto_students,
            "tallanto_writeoff_visits": writeoff_rows,
            "tallanto_writeoff_summary_by_student": writeoff_summary,
            "quality_gate_snapshot": quality_gates,
            "source_manifest": source_manifest,
        },
    )

    summary = build_summary(
        paths=paths,
        phone_rollup=phone_rollup,
        amo_ready=amo_ready,
        amo_writebacks=amo_writebacks,
        amo_loss_reasons=amo_loss_reasons,
        call_snapshot=call_snapshot,
        call_rollup=call_rollup,
        tallanto_students=tallanto_students,
        writeoff_rows=writeoff_rows,
        writeoff_summary=writeoff_summary,
        quality_gates=quality_gates,
        source_manifest=source_manifest,
        runtime_contract=runtime_contract,
        outputs=outputs,
    )
    outputs["summary_json"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs["report_md"].write_text(render_report(summary), encoding="utf-8")
    return summary


def build_phone_rollup(path: Path) -> list[dict[str, Any]]:
    rows = []
    for row in read_csv(path):
        phone = normalize_phone(row.get("Телефон клиента", ""))
        if not phone:
            continue
        rows.append(
            {
                "phone": phone,
                "total_calls": int_or_zero(row.get("Всего звонков в истории")),
                "contentful_calls": int_or_zero(row.get("Содержательных звонков в истории")),
                "non_conversation_calls": int_or_zero(row.get("Несодержательных звонков в истории")),
                "first_call_at": safe_text(row.get("Первый звонок")),
                "last_call_at": safe_text(row.get("Последний звонок")),
                "latest_fresh_call_at": safe_text(row.get("Последний свежий звонок")),
                "latest_call_type": safe_text(row.get("Тип последнего свежего звонка")),
                "products": safe_text(row.get("Продукты интереса")),
                "recommended_product": safe_text(row.get("Рекомендуемый продукт")),
                "objections": safe_text(row.get("Возражения")),
                "next_step": safe_text(row.get("Следующий шаг")),
                "priority": safe_text(row.get("Приоритет лида")),
                "probability_percent": safe_text(row.get("Вероятность продажи, %")),
                "tallanto_match_status": safe_text(row.get("Статус матчинга Tallanto")),
                "tallanto_id": safe_text(row.get("ID Tallanto")),
                "amo_contact_ids": safe_text(row.get("AMO contact IDs")),
                "amo_lead_ids": safe_text(row.get("AMO lead IDs")),
                "crm_writeback_policy": safe_text(row.get("CRM writeback policy")),
                "amo_entity_policy": safe_text(row.get("AMO entity policy")),
                "amo_ready": safe_text(row.get("Готово к записи в AMO")),
                "amo_status_reason": safe_text(row.get("Причина статуса AMO")),
            }
        )
    return rows


def build_call_snapshot(path: Path | None) -> list[dict[str, Any]]:
    if not path:
        return []
    rows = []
    for row in read_csv(path):
        rows.append(
            {
                "call_id": safe_text(row.get("ID звонка") or row.get("source_filename")),
                "phone": normalize_phone(row.get("Телефон клиента", "") or row.get("phone", "")),
                "started_at": safe_text(row.get("Дата и время звонка") or row.get("started_at")),
                "manager_name": safe_text(row.get("Менеджер") or row.get("manager_name")),
                "duration_sec": safe_text(row.get("Длительность, сек") or row.get("duration_sec")),
                "source_filename": safe_text(row.get("Имя исходного файла") or row.get("source_filename")),
                "full_pipeline_done": safe_text(row.get("Полная цепочка выполнена")),
                "contentful": safe_text(row.get("Содержательный звонок") or row.get("contentful")),
                "needs_review": safe_text(row.get("Нужна ручная проверка") or row.get("needs_review")),
                "call_summary": safe_text(row.get("Краткое резюме разговора") or row.get("history_summary")),
                "call_type": safe_text(row.get("Тип звонка") or row.get("call_type")),
                "parent_name": safe_text(row.get("ФИО родителя")),
                "student_name": safe_text(row.get("ФИО ребенка")),
                "products": safe_text(row.get("Продукты интереса") or row.get("products")),
                "subjects": safe_text(row.get("Предметы интереса") or row.get("subjects")),
                "objections": safe_text(row.get("Возражения") or row.get("objections")),
                "next_step": safe_text(row.get("Следующий шаг") or row.get("next_step")),
                "priority": safe_text(row.get("Приоритет лида") or row.get("lead_priority")),
                "probability_percent": safe_text(row.get("Вероятность продажи, %")),
                "source_status": safe_text(row.get("Источник лучшего статуса") or row.get("source_db")),
            }
        )
    return rows


def build_amo_ready_snapshot(path: Path) -> list[dict[str, Any]]:
    result = []
    for row in read_csv(path):
        phone = normalize_phone(row.get("Телефон клиента", ""))
        result.append(
            {
                "phone": phone,
                "tallanto_id": safe_text(row.get("ID Tallanto")),
                "tallanto_match_status": safe_text(row.get("Статус матчинга Tallanto")),
                "amo_contact_ids": safe_text(row.get("AMO contact IDs")),
                "amo_lead_ids": safe_text(row.get("AMO lead IDs")),
                "latest_call_at": safe_text(row.get("Дата последнего свежего звонка")),
                "latest_call_type": safe_text(row.get("Тип последнего свежего звонка")),
                "recommended_product": safe_text(row.get("Рекомендуемый продукт")),
                "products": safe_text(row.get("Продукты интереса")),
                "next_step": safe_text(row.get("Следующий шаг")),
                "priority": safe_text(row.get("Приоритет лида")),
                "probability_percent": safe_text(row.get("Вероятность продажи, %")),
                "crm_writeback_policy": safe_text(row.get("CRM writeback policy")),
                "amo_entity_policy": safe_text(row.get("AMO entity policy")),
                "amo_ready": safe_text(row.get("Готово к записи в AMO")),
                "amo_status_reason": safe_text(row.get("Причина статуса AMO")),
                "auto_history_preview": safe_text(row.get("Краткая история общения"))[:1000],
            }
        )
    return result


def build_amo_writeback_snapshot(stage_dirs: Iterable[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for stage_dir in stage_dirs:
        if not stage_dir.exists():
            continue
        candidates = sorted(stage_dir.glob("live*_candidates_ru.csv")) + sorted(stage_dir.glob("repair*_candidates_ru.csv"))
        readback_summary = load_json(stage_dir / "readback_after_live" / "summary.json") or load_json(
            stage_dir / "readback_after_repair" / "summary.json"
        )
        readback_rows = readback_rows_by_phone(stage_dir)
        for candidate_path in candidates:
            for row in read_csv(candidate_path):
                phone = normalize_phone(row.get("Телефон клиента", ""))
                readback = readback_rows.get(phone, {})
                rows.append(
                    {
                        "stage_dir": str(stage_dir),
                        "candidate_file": str(candidate_path),
                        "phone": phone,
                        "amo_contact_ids": safe_text(row.get("AMO contact IDs")),
                        "amo_lead_ids": safe_text(row.get("AMO lead IDs")),
                        "crm_writeback_policy": safe_text(row.get("CRM writeback policy")),
                        "amo_entity_policy": safe_text(row.get("AMO entity policy")),
                        "latest_call_at": safe_text(row.get("Дата последнего свежего звонка") or row.get("Последний свежий звонок")),
                        "recommended_product": safe_text(row.get("Рекомендуемый продукт")),
                        "next_step": safe_text(row.get("Следующий шаг")),
                        "readback_passed": bool(readback_summary.get("passed")) if readback_summary else "",
                        "readback_risk_counts": json.dumps(readback_summary.get("risk_counts", {}), ensure_ascii=False)
                        if readback_summary
                        else "",
                        "readback_decision": safe_text(readback.get("decision")),
                        "readback_risk_types": safe_text(readback.get("risk_types")),
                    }
                )
    return dedupe_dict_rows(rows, key_fields=("stage_dir", "phone", "amo_contact_ids"))


def build_amo_loss_reason_snapshot(loss_reason_dirs: Iterable[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for root in loss_reason_dirs:
        if not root.exists():
            continue
        for filename in ("amo_loss_reason_values.csv", "amo_loss_reason_lead_samples.csv"):
            path = root / filename
            if not path.exists():
                continue
            for row in read_csv(path):
                rows.append({"source_dir": str(root), "source_file": filename, **{k: safe_text(v) for k, v in row.items()}})
    return rows


def build_call_rollup(path: Path) -> list[dict[str, Any]]:
    by_phone: dict[str, dict[str, Any]] = {}
    for row in read_csv(path):
        phone = normalize_phone(row.get("phone", "") or row.get("Телефон клиента", ""))
        if not phone:
            continue
        item = by_phone.setdefault(
            phone,
            {
                "phone": phone,
                "calls": 0,
                "contentful_calls": 0,
                "non_conversation_calls": 0,
                "sales_calls": 0,
                "service_calls": 0,
                "existing_client_progress_calls": 0,
                "technical_calls": 0,
                "first_call_at": "",
                "last_call_at": "",
                "managers": set(),
                "products": set(),
                "subjects": set(),
            },
        )
        item["calls"] += 1
        contentful = safe_text(row.get("contentful") or row.get("Содержательный звонок")).casefold() in {
            "1",
            "true",
            "yes",
            "да",
        }
        if contentful:
            item["contentful_calls"] += 1
        else:
            item["non_conversation_calls"] += 1
        call_type = safe_text(row.get("call_type") or row.get("Тип звонка"))
        if call_type == "sales_call":
            item["sales_calls"] += 1
        elif call_type == "service_call":
            item["service_calls"] += 1
        elif call_type == "existing_client_progress":
            item["existing_client_progress_calls"] += 1
        elif call_type == "technical_call":
            item["technical_calls"] += 1
        started = safe_text(row.get("started_at") or row.get("Дата и время звонка"))
        if started:
            if not item["first_call_at"] or started < item["first_call_at"]:
                item["first_call_at"] = started
            if not item["last_call_at"] or started > item["last_call_at"]:
                item["last_call_at"] = started
        add_pipe_values(item["managers"], row.get("manager_name") or row.get("Менеджер"))
        add_pipe_values(item["products"], row.get("products") or row.get("Продукты интереса"))
        add_pipe_values(item["subjects"], row.get("subjects") or row.get("Предметы интереса"))
    result = []
    for item in by_phone.values():
        row = dict(item)
        row["managers"] = " | ".join(sorted(item["managers"]))
        row["products"] = " | ".join(sorted(item["products"]))
        row["subjects"] = " | ".join(sorted(item["subjects"]))
        result.append(row)
    result.sort(key=lambda item: item["phone"])
    return result


def build_tallanto_students_snapshot(path: Path | None) -> list[dict[str, Any]]:
    if not path or not path.exists():
        return []
    rows = []
    for row in read_delimited(path):
        rows.append(
            {
                "tallanto_id": safe_text(row.get("ID")),
                "first_name": safe_text(row.get("Имя")),
                "last_name": safe_text(row.get("Фамилия")),
                "full_name": " ".join(
                    part
                    for part in (safe_text(row.get("Фамилия")), safe_text(row.get("Имя")))
                    if part
                ),
                "parent_name": safe_text(row.get("ФИО родителя")),
                "email": safe_text(row.get("E-mail")),
                "phone_parent": normalize_phone(row.get("Тел. (родителя)", "")),
                "phone_mobile_digits": safe_text(row.get("Тел. цифровой (моб.)")),
                "barcode": safe_text(row.get("Текстовое значение штрихкода")),
                "amo_contact_id": safe_text(row.get("amoCRM ID")),
                "student_type": safe_text(row.get("Тип ученика")),
                "branch": safe_text(row.get("Филиал")),
                "responsible": safe_text(row.get("Ответственный(ая)")),
                "communication_history": safe_text(row.get("История общения"))[:2000],
                "interests": safe_text(row.get("Интересы"))[:2000],
                "money_spent": safe_text(row.get("Потраченные деньги")),
                "money_topup": safe_text(row.get("Пополнено на сумму")),
                "balance": safe_text(row.get("Баланс")),
                "unconfirmed_payment": safe_text(row.get("Неподтвержденная оплата")),
                "group_id": safe_text(row.get("Группа(ID)")),
                "class_id": safe_text(row.get("Занятие(ID)")),
            }
        )
    return rows


def build_tallanto_writeoff_snapshot(
    paths: Iterable[Path],
    *,
    combined_csv: Path | None = None,
) -> list[dict[str, Any]]:
    if combined_csv and combined_csv.exists():
        return normalize_writeoff_csv_rows(combined_csv)
    rows: list[dict[str, Any]] = []
    for path in paths:
        if not path.exists():
            continue
        rows.extend(read_writeoff_xlsx(path))
    return rows


def build_tallanto_writeoff_summary_snapshot(
    *,
    rows: list[dict[str, Any]],
    summary_csv: Path | None = None,
) -> list[dict[str, Any]]:
    if summary_csv and summary_csv.exists():
        return [{key: safe_text(value) for key, value in row.items()} for row in read_csv(summary_csv)]
    return summarize_writeoffs(rows)


def normalize_writeoff_csv_rows(path: Path) -> list[dict[str, Any]]:
    result = []
    for row in read_csv(path):
        result.append(
            {
                "source_file": safe_text(row.get("source_files")),
                "source_periods": safe_text(row.get("source_periods")),
                "source_count": safe_text(row.get("source_count")),
                "last_name": safe_text(row.get("Фамилия")),
                "first_middle_name": safe_text(row.get("Имя")),
                "barcode": safe_text(row.get("Штрихкод")),
                "subscription": safe_text(row.get("Абонемент")),
                "writeoff_amount": safe_text(row.get("Сумма списания")),
                "writeoff_at": safe_text(row.get("Дата списания")),
                "writeoff_type": safe_text(row.get("Тип списания")),
                "class_title": safe_text(row.get("Занятие")),
                "class_branch": safe_text(row.get("Филиал занятия")),
                "class_at": safe_text(row.get("Дата занятия")),
                "birth_date": safe_text(row.get("День рождения")),
            }
        )
    return result


def read_writeoff_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to read Tallanto write-off reports") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    result: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        try:
            ws.reset_dimensions()
        except Exception:
            pass
        header: list[str] | None = None
        for row_index, raw_row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [safe_text(value) for value in raw_row]
            if header is None:
                if {"Фамилия", "Имя", "Абонемент", "Сумма списания"}.issubset(set(values)):
                    header = [value or f"col_{idx + 1}" for idx, value in enumerate(values)]
                continue
            if not any(values):
                continue
            item = {header[idx]: values[idx] if idx < len(values) else "" for idx in range(len(header))}
            if not safe_text(item.get("Фамилия")) and not safe_text(item.get("Имя")):
                continue
            result.append(
                {
                    "source_file": str(path),
                    "sheet": ws.title,
                    "source_row": row_index,
                    "last_name": safe_text(item.get("Фамилия")),
                    "first_middle_name": safe_text(item.get("Имя")),
                    "barcode": safe_text(item.get("Штрихкод")),
                    "subscription": safe_text(item.get("Абонемент")),
                    "writeoff_amount": safe_text(item.get("Сумма списания")),
                    "writeoff_at": safe_text(item.get("Дата списания")),
                    "writeoff_type": safe_text(item.get("Тип списания")),
                    "class_title": safe_text(item.get("Занятие")),
                    "class_branch": safe_text(item.get("Филиал занятия")),
                    "class_at": safe_text(item.get("Дата занятия")),
                    "birth_date": safe_text(item.get("День рождения")),
                }
            )
    return result


def summarize_writeoffs(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in rows:
        key = (safe_text(row.get("last_name")), safe_text(row.get("first_middle_name")), safe_text(row.get("birth_date")))
        item = grouped.setdefault(
            key,
            {
                "last_name": key[0],
                "first_middle_name": key[1],
                "birth_date": key[2],
                "writeoff_count": 0,
                "total_writeoff_amount": 0.0,
                "first_class_at": "",
                "last_class_at": "",
                "classes": set(),
                "branches": set(),
                "subscriptions": set(),
            },
        )
        item["writeoff_count"] += 1
        item["total_writeoff_amount"] += float_or_zero(row.get("writeoff_amount"))
        class_at = safe_text(row.get("class_at"))
        if class_at:
            if not item["first_class_at"] or class_at < item["first_class_at"]:
                item["first_class_at"] = class_at
            if not item["last_class_at"] or class_at > item["last_class_at"]:
                item["last_class_at"] = class_at
        add_pipe_values(item["classes"], row.get("class_title"), separator="|||")
        add_pipe_values(item["branches"], row.get("class_branch"))
        add_pipe_values(item["subscriptions"], row.get("subscription"))
    result = []
    for item in grouped.values():
        row = dict(item)
        row["total_writeoff_amount"] = round(float(row["total_writeoff_amount"]), 2)
        row["classes"] = " | ".join(sorted(item["classes"])[:20])
        row["branches"] = " | ".join(sorted(item["branches"]))
        row["subscriptions"] = " | ".join(sorted(item["subscriptions"]))
        result.append(row)
    result.sort(key=lambda item: (-int(item["writeoff_count"]), item["last_name"], item["first_middle_name"]))
    return result


def build_summary(
    *,
    paths: Stage1Paths,
    phone_rollup: list[dict[str, Any]],
    amo_ready: list[dict[str, Any]],
    amo_writebacks: list[dict[str, Any]],
    amo_loss_reasons: list[dict[str, Any]],
    call_snapshot: list[dict[str, Any]],
    call_rollup: list[dict[str, Any]],
    tallanto_students: list[dict[str, Any]],
    writeoff_rows: list[dict[str, Any]],
    writeoff_summary: list[dict[str, Any]],
    quality_gates: list[dict[str, Any]],
    source_manifest: list[dict[str, Any]],
    runtime_contract: dict[str, Any],
    outputs: dict[str, Path],
) -> dict[str, Any]:
    phone_set = {row["phone"] for row in phone_rollup if row.get("phone")}
    call_phone_set = {row["phone"] for row in call_rollup if row.get("phone")}
    amo_phone_set = {row["phone"] for row in amo_writebacks if row.get("phone")}
    tallanto_exact = Counter(row.get("tallanto_match_status", "") for row in phone_rollup)
    amo_live_summary = load_json(paths.amo_live_snapshot_dir / "summary.json") if paths.amo_live_snapshot_dir else {}
    amo_live_fetch = amo_live_summary.get("fetch", {}) if isinstance(amo_live_summary.get("fetch"), dict) else {}
    amo_live_ok = bool(
        amo_live_summary
        and amo_live_summary.get("api_read_succeeded", True)
        and int_or_zero(amo_live_fetch.get("contacts_seen")) > 0
        and int_or_zero(amo_live_fetch.get("leads_seen")) > 0
    )
    missing_sources = [
        row["source_key"]
        for row in source_manifest
        if row.get("required") == "true" and row.get("exists") != "true"
    ]
    blocking_gaps = [
        "Full Tallanto payments/groups/API snapshot is not complete in this artifact.",
        "Deal attribution has not been run yet; phone-level context must not be written to deal cards.",
    ]
    known_limitations = [
        "Tallanto write-off reports do not contain phone numbers, so they are linked later through Tallanto barcode/student identity.",
        "Tallanto payments/groups/API module snapshot is not yet complete; current artifact uses exported students and write-off visits.",
        "AMO writeback snapshot covers staged AI-writeback rows, not the full AMO account.",
    ]
    if not amo_live_ok:
        blocking_gaps.insert(
            0,
            "Full live AMO deals/tasks/linked-contacts snapshot is not complete in this artifact.",
        )
        known_limitations.insert(
            0,
            "AMO live contacts/deals/tasks snapshot is not included unless a live AMO tunnel/API read is available.",
        )
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "runtime_contract_schema_version": safe_text(runtime_contract.get("summary", {}).get("schema_version")),
        "sources": {
            "master_contacts_csv": str(paths.master_contacts_csv),
            "master_calls_csv": str(paths.master_calls_csv or ""),
            "amo_ready_csv": str(paths.amo_ready_csv),
            "calls_csv": str(paths.calls_csv),
            "current_runtime_json": str(paths.current_runtime_json or ""),
            "canonical_export_pointer": str(paths.canonical_export_pointer or ""),
            "amo_live_snapshot_dir": str(paths.amo_live_snapshot_dir or ""),
            "tallanto_students_csv": str(paths.tallanto_students_csv or ""),
            "tallanto_writeoff_combined_csv": str(paths.tallanto_writeoff_combined_csv or ""),
            "tallanto_writeoff_summary_csv": str(paths.tallanto_writeoff_summary_csv or ""),
            "tallanto_schema_json": str(paths.tallanto_schema_json or ""),
            "quality_summary_paths": [str(path) for path in paths.quality_summary_paths],
            "amo_stage_dirs": [str(path) for path in paths.amo_stage_dirs],
            "amo_loss_reason_dirs": [str(path) for path in paths.amo_loss_reason_dirs],
            "tallanto_writeoff_xlsx": [str(path) for path in paths.tallanto_writeoff_xlsx],
        },
        "safety": {
            "read_only": True,
            "write_amo": False,
            "write_tallanto": False,
            "run_asr": False,
            "run_resolve_analyze": False,
        },
        "readiness": {
            "local_snapshot_built": True,
            "complete_live_stage1": amo_live_ok,
            "safe_to_use_for_schema_design": True,
            "safe_to_use_for_deal_writeback": False,
            "missing_required_sources": missing_sources,
            "blocking_gaps_for_next_stage": blocking_gaps,
        },
        "live_dependencies": {
            "amocrm_shared_db_tunnel_127_0_0_1_15432": probe_tcp("127.0.0.1", 15432),
        },
        "amo_live_snapshot": {
            "summary_path": str(paths.amo_live_snapshot_dir / "summary.json") if paths.amo_live_snapshot_dir else "",
            "api_read_succeeded": amo_live_summary.get("api_read_succeeded", True) if amo_live_summary else False,
            "preflight_error": safe_text(amo_live_summary.get("preflight_error")),
            "contacts_seen": int_or_zero(amo_live_fetch.get("contacts_seen")),
            "leads_seen": int_or_zero(amo_live_fetch.get("leads_seen")),
            "tasks_seen": int_or_zero(amo_live_fetch.get("tasks_seen")),
        },
        "coverage": {
            "phones_in_master_contacts": len(phone_set),
            "call_snapshot_rows": len(call_snapshot),
            "phones_in_call_rollup": len(call_phone_set),
            "phones_with_call_rollup_not_in_master_contacts": len(call_phone_set - phone_set),
            "tallanto_students_rows": len(tallanto_students),
            "amo_ready_rows": len(amo_ready),
            "amo_writeback_rows": len(amo_writebacks),
            "amo_writeback_unique_phones": len(amo_phone_set),
            "amo_loss_reason_rows": len(amo_loss_reasons),
            "tallanto_writeoff_rows": len(writeoff_rows),
            "tallanto_writeoff_students": len(writeoff_summary),
            "quality_gate_rows": len(quality_gates),
        },
        "policy_counts": {
            "crm_writeback_policy": dict(Counter(row.get("crm_writeback_policy", "") for row in phone_rollup).most_common()),
            "amo_entity_policy": dict(Counter(row.get("amo_entity_policy", "") for row in phone_rollup).most_common()),
            "tallanto_match_status": dict(tallanto_exact.most_common()),
        },
        "known_limitations": known_limitations,
        "source_manifest_counts": dict(Counter(row.get("exists", "") for row in source_manifest).most_common()),
        "quality_gate_counts": dict(Counter(row.get("passed", "") for row in quality_gates).most_common()),
        "outputs": {key: str(path) for key, path in outputs.items()},
    }


def render_report(summary: dict[str, Any]) -> str:
    coverage = summary["coverage"]
    return "\n".join(
        [
            "# Deal-Aware Stage 1 Snapshot",
            "",
            "Read-only snapshot for the first deal-aware implementation stage.",
            "",
            "## Safety",
            "",
            "- No AMO writes.",
            "- No Tallanto writes.",
            "- No ASR / Resolve / Analyze runs.",
            "",
            "## Coverage",
            "",
            f"- phones in master contacts: {coverage['phones_in_master_contacts']}",
            f"- call snapshot rows: {coverage['call_snapshot_rows']}",
            f"- phones in call rollup: {coverage['phones_in_call_rollup']}",
            f"- Tallanto students rows: {coverage['tallanto_students_rows']}",
            f"- AMO-ready rows: {coverage['amo_ready_rows']}",
            f"- staged AMO writeback rows: {coverage['amo_writeback_rows']}",
            f"- staged AMO writeback unique phones: {coverage['amo_writeback_unique_phones']}",
            f"- AMO loss reason rows: {coverage['amo_loss_reason_rows']}",
            f"- Tallanto write-off rows: {coverage['tallanto_writeoff_rows']}",
            f"- Tallanto write-off students: {coverage['tallanto_writeoff_students']}",
            f"- quality gate rows: {coverage['quality_gate_rows']}",
            "",
            "## Readiness",
            "",
            f"- local snapshot built: {summary['readiness']['local_snapshot_built']}",
            f"- complete live Stage 1: {summary['readiness']['complete_live_stage1']}",
            f"- safe for deal writeback: {summary['readiness']['safe_to_use_for_deal_writeback']}",
            "- blocking gaps:",
            *[f"  - {item}" for item in summary["readiness"]["blocking_gaps_for_next_stage"]],
            "",
            "## Known Limitations",
            "",
            *[f"- {item}" for item in summary["known_limitations"]],
            "",
            "## Outputs",
            "",
            *[f"- `{key}`: `{path}`" for key, path in summary["outputs"].items()],
            "",
        ]
    )


def build_quality_gate_snapshot(paths: Iterable[Path]) -> list[dict[str, Any]]:
    rows = []
    for path in paths:
        payload = load_json(path)
        if not payload:
            continue
        rows.append(
            {
                "source_path": str(path),
                "schema_version": safe_text(payload.get("schema_version") or payload.get("gate_version")),
                "generated_at": safe_text(payload.get("generated_at")),
                "input": safe_text(payload.get("input") or payload.get("inputs")),
                "passed": stringify(payload.get("passed")),
                "rows": safe_text(payload.get("rows") or payload.get("rows_scanned")),
                "blocking_rows": safe_text(payload.get("blocking_rows")),
                "failures": safe_text(payload.get("frozen_corpus", {}).get("failures")),
                "risk_counts": stringify(payload.get("risk_counts", {})),
                "tenant_id": safe_text(payload.get("tenant_config", {}).get("tenant_id")),
                "tenant_config_sha256": safe_text(payload.get("tenant_config", {}).get("sha256")),
            }
        )
    return rows


def build_source_manifest(paths: Stage1Paths) -> list[dict[str, Any]]:
    source_paths: list[tuple[str, Path | None, bool]] = [
        ("master_contacts_csv", paths.master_contacts_csv, True),
        ("master_calls_csv", paths.master_calls_csv, True),
        ("amo_ready_csv", paths.amo_ready_csv, True),
        ("calls_csv", paths.calls_csv, True),
        ("current_runtime_json", paths.current_runtime_json, False),
        ("canonical_export_pointer", paths.canonical_export_pointer, False),
        ("amo_live_snapshot_dir", paths.amo_live_snapshot_dir, False),
        ("tallanto_students_csv", paths.tallanto_students_csv, False),
        ("tallanto_writeoff_combined_csv", paths.tallanto_writeoff_combined_csv, False),
        ("tallanto_writeoff_summary_csv", paths.tallanto_writeoff_summary_csv, False),
        ("tallanto_schema_json", paths.tallanto_schema_json, False),
    ]
    for idx, path in enumerate(paths.quality_summary_paths, start=1):
        source_paths.append((f"quality_summary_{idx}", path, False))
    for idx, path in enumerate(paths.amo_stage_dirs, start=1):
        source_paths.append((f"amo_stage_dir_{idx}", path, False))
    for idx, path in enumerate(paths.amo_loss_reason_dirs, start=1):
        source_paths.append((f"amo_loss_reason_dir_{idx}", path, False))
    for idx, path in enumerate(paths.tallanto_writeoff_xlsx, start=1):
        source_paths.append((f"tallanto_writeoff_xlsx_{idx}", path, False))
    return [source_manifest_row(key, path, required) for key, path, required in source_paths]


def source_manifest_row(source_key: str, path: Path | None, required: bool) -> dict[str, Any]:
    exists = bool(path and path.exists())
    stat = path.stat() if exists and path else None
    return {
        "source_key": source_key,
        "path": str(path or ""),
        "required": str(required).lower(),
        "exists": str(exists).lower(),
        "is_dir": str(bool(path and path.is_dir())).lower() if exists else "",
        "bytes": stat.st_size if stat else "",
        "modified_at": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).isoformat(timespec="seconds") if stat else "",
    }


def write_sqlite(path: Path, tables: dict[str, list[dict[str, Any]]]) -> None:
    if path.exists():
        path.unlink()
    con = sqlite3.connect(path)
    try:
        for table, rows in tables.items():
            if not rows:
                con.execute(f'CREATE TABLE "{table}" (empty TEXT)')
                continue
            columns = sorted({key for row in rows for key in row.keys()})
            con.execute(
                f'CREATE TABLE "{table}" ({", ".join(f"{quote_ident(col)} TEXT" for col in columns)})'
            )
            placeholders = ", ".join("?" for _ in columns)
            con.executemany(
                f'INSERT INTO "{table}" ({", ".join(quote_ident(col) for col in columns)}) VALUES ({placeholders})',
                [[stringify(row.get(col)) for col in columns] for row in rows],
            )
        con.commit()
    finally:
        con.close()


def readback_rows_by_phone(stage_dir: Path) -> dict[str, dict[str, str]]:
    candidates = [
        stage_dir / "readback_after_live" / "readback_report.csv",
        stage_dir / "readback_after_repair" / "readback_report.csv",
    ]
    result: dict[str, dict[str, str]] = {}
    for path in candidates:
        if not path.exists():
            continue
        for row in read_csv(path):
            phone = normalize_phone(row.get("phone", "") or row.get("Телефон клиента", ""))
            if phone:
                result[phone] = row
    return result


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def read_delimited(path: Path) -> list[dict[str, str]]:
    encodings = ("utf-8-sig", "cp1251")
    delimiters = ("\t", ",", ";")
    last_error: Exception | None = None
    for encoding in encodings:
        for delimiter in delimiters:
            try:
                with path.open("r", encoding=encoding, newline="") as fh:
                    reader = csv.DictReader(fh, delimiter=delimiter)
                    fieldnames = reader.fieldnames or []
                    if len(fieldnames) <= 1:
                        continue
                    return [dict(row) for row in reader]
            except UnicodeDecodeError as exc:
                last_error = exc
                continue
    if last_error:
        raise last_error
    return []


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = sorted({key for row in rows for key in row.keys()})
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows([{key: stringify(row.get(key)) for key in fieldnames} for row in rows])


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def dedupe_dict_rows(rows: list[dict[str, Any]], *, key_fields: tuple[str, ...]) -> list[dict[str, Any]]:
    seen = set()
    result = []
    for row in rows:
        key = tuple(safe_text(row.get(field)) for field in key_fields)
        if key in seen:
            continue
        seen.add(key)
        result.append(row)
    return result


def add_pipe_values(target: set[str], value: Any, *, separator: str = "|") -> None:
    for part in safe_text(value).replace("\n", f" {separator} ").split(separator):
        text = part.strip(" .;,")
        if text:
            target.add(text)


def quote_ident(value: str) -> str:
    return '"' + value.replace('"', '""') + '"'


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def safe_text(value: Any) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def int_or_zero(value: Any) -> int:
    try:
        return int(float(safe_text(value).replace(",", ".")))
    except ValueError:
        return 0


def float_or_zero(value: Any) -> float:
    text = safe_text(value).replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def probe_tcp(host: str, port: int, *, timeout_sec: float = 0.25) -> dict[str, Any]:
    try:
        with socket.create_connection((host, port), timeout=timeout_sec):
            return {"host": host, "port": port, "open": True, "error": ""}
    except OSError as exc:
        return {"host": host, "port": port, "open": False, "error": str(exc)}


def stage_dirs_default(project_root: Path) -> tuple[Path, ...]:
    stable = project_root / "stable_runtime"
    return tuple(
        path
        for path in (
            stable / "amo_live_stage51_20260512_v1",
            stable / "amo_live_stage51_textarea_repair_20260512_v2",
            stable / "amo_live_stage100_20260512_v1",
            stable / "amo_live_stage100_batch2_20260512_v1",
            stable / "amo_live_stage200_batch3_20260512_v1",
        )
        if path.exists()
    )


def loss_reason_dirs_default(project_root: Path) -> tuple[Path, ...]:
    stable = project_root / "stable_runtime"
    return tuple(sorted(stable.glob("amo_loss_reason_policy_audit_20260513*")))


def writeoff_xlsx_default(project_root: Path) -> tuple[Path, ...]:
    return tuple(sorted(project_root.glob("*write_off_visits_from_class.xlsx")))
