from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from mango_mvp.insights.sanitizers import has_brand_risk, has_money_or_terms_risk, has_personal_data_risk


BOT_READY_STATUSES = {"ready_for_bot_draft", "needs_rop_validation"}
SANITIZER_FLAG_ORDER = (
    "brand_normalized",
    "price_redacted",
    "discount_terms_redacted",
    "percent_redacted",
    "installment_terms_redacted",
    "refund_policy_redacted",
    "deadline_redacted",
    "person_name_redacted",
    "email_redacted",
    "phone_redacted",
)
REQUIRED_AFTER_KB_COLUMNS = {
    "ideal_answer_manager_sanitized",
    "bot_safe_answer",
    "bot_safety_status",
    "sanitizer_flags",
    "brand_risk_flag",
    "money_or_discount_flag",
    "installment_flag",
    "legal_or_refund_flag",
    "deadline_or_promise_flag",
    "personal_data_flag",
}
REQUIRED_AFTER_ROP_COLUMNS = {
    "Идеальный ответ для менеджера",
    "Безопасный ответ для бота",
    "Статус sanitizer",
    "Флаги sanitizer",
    "Риск бренда",
    "Риск цены/скидки",
    "Риск рассрочки",
    "Риск договора/возврата",
    "Риск срока/обещания",
    "Риск персональных данных",
}


@dataclass(frozen=True)
class Stage14ComparisonConfig:
    project_root: Path
    before_kb_root: Path
    after_kb_root: Path
    before_rop_root: Path
    after_rop_root: Path
    before_baseline_root: Path
    after_baseline_root: Path
    out_root: Path
    audit_sample_limit: int = 200
    over_sanitization_limit: int = 250
    residual_sample_limit: int = 100


def build_stage14_quality_comparison(config: Stage14ComparisonConfig) -> dict[str, Any]:
    project_root = config.project_root.resolve()
    out_root = config.out_root.resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    paths = input_paths(config)
    _require_files(paths.values())

    before_baseline = _load_json(paths["before_baseline_summary"])
    after_baseline = _load_json(paths["after_baseline_summary"])
    before_kb_summary = _load_json(paths["before_kb_summary"])
    after_kb_summary = _load_json(paths["after_kb_summary"])
    before_rop_summary = _load_json(paths["before_rop_summary"])
    after_rop_summary = _load_json(paths["after_rop_summary"])

    before_kb_rows = _read_csv(paths["before_kb_enriched"])
    after_kb_rows = _read_csv(paths["after_kb_enriched"])
    before_bot_seed_rows = _read_csv(paths["before_kb_bot_seeds"])
    after_bot_seed_rows = _read_csv(paths["after_kb_bot_seeds"])
    before_rop_rows = _read_csv(paths["before_rop_validation"])
    after_rop_rows = _read_csv(paths["after_rop_validation"])
    before_rop_bot_rows = _read_csv(paths["before_rop_bot_drafts"])
    after_rop_bot_rows = _read_csv(paths["after_rop_bot_drafts"])

    schema = build_schema_report(after_kb_rows, after_rop_rows, after_bot_seed_rows, after_rop_bot_rows)
    metrics = build_metric_delta(before_baseline, after_baseline)
    sanitizer_metrics = build_sanitizer_metrics(
        after_kb_summary=after_kb_summary,
        after_kb_rows=after_kb_rows,
        after_bot_seed_rows=after_bot_seed_rows,
        after_rop_bot_rows=after_rop_bot_rows,
    )
    residual_samples = build_residual_risk_samples(after_bot_seed_rows, after_rop_bot_rows, config.residual_sample_limit)
    over_sanitization_candidates = build_over_sanitization_candidates(after_kb_rows, config.over_sanitization_limit)
    audit_sample = build_audit_sample(after_kb_rows, after_rop_rows, after_rop_bot_rows, config.audit_sample_limit)
    bot_seed_delta_sample = build_bot_seed_delta_sample(before_bot_seed_rows, after_bot_seed_rows, limit=100)
    acceptance = build_acceptance_report(
        metrics=metrics,
        schema=schema,
        sanitizer_metrics=sanitizer_metrics,
        residual_samples=residual_samples,
        audit_sample=audit_sample,
    )

    summary: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "project_root": str(project_root),
        "comparison_version": "transcript_quality_stage14_v1",
        "inputs": {key: str(path) for key, path in paths.items()},
        "input_fingerprints": [_fingerprint(path, project_root) for path in paths.values()],
        "before": {
            "kb_root": str(config.before_kb_root),
            "rop_root": str(config.before_rop_root),
            "baseline_root": str(config.before_baseline_root),
            "kb_summary_key_metrics": summary_key_metrics(before_kb_summary),
            "rop_summary_key_metrics": summary_key_metrics(before_rop_summary),
        },
        "after": {
            "kb_root": str(config.after_kb_root),
            "rop_root": str(config.after_rop_root),
            "baseline_root": str(config.after_baseline_root),
            "kb_summary_key_metrics": summary_key_metrics(after_kb_summary),
            "rop_summary_key_metrics": summary_key_metrics(after_rop_summary),
        },
        "metric_deltas": metrics,
        "schema": schema,
        "sanitizer_metrics": sanitizer_metrics,
        "acceptance": acceptance,
        "audit_sample": {
            "rows": len(audit_sample),
            "by_bucket": dict(Counter(row["audit_bucket"] for row in audit_sample).most_common()),
        },
        "over_sanitization_candidates": {
            "rows": len(over_sanitization_candidates),
            "by_reason": dict(Counter(row["reason"] for row in over_sanitization_candidates).most_common()),
            "interpretation": "Это очередь для выборочной проверки полезности, не список ошибок.",
        },
        "residual_risk_samples": {
            "rows": len(residual_samples),
            "by_source": dict(Counter(row["source"] for row in residual_samples).most_common()),
        },
        "outputs": {},
    }

    outputs = {
        "summary_json": out_root / "summary.json",
        "comparison_report_md": out_root / "STAGE14_QUALITY_COMPARISON_REPORT.md",
        "comparison_xlsx": out_root / "stage14_quality_comparison.xlsx",
        "metric_delta_csv": out_root / "metric_delta.csv",
        "audit_sample_csv": out_root / "audit_sample.csv",
        "over_sanitization_candidates_csv": out_root / "over_sanitization_candidates.csv",
        "residual_risk_sample_csv": out_root / "residual_risk_sample.csv",
        "bot_seed_before_after_sample_csv": out_root / "bot_seed_before_after_sample.csv",
        "audit_prompt_md": out_root / "AUDIT_PROMPT_FOR_CLAUDE_OR_GPT.md",
    }
    _write_csv(outputs["metric_delta_csv"], metric_delta_rows(metrics))
    _write_csv(outputs["audit_sample_csv"], audit_sample)
    _write_csv(outputs["over_sanitization_candidates_csv"], over_sanitization_candidates)
    _write_csv(outputs["residual_risk_sample_csv"], residual_samples)
    _write_csv(outputs["bot_seed_before_after_sample_csv"], bot_seed_delta_sample)
    write_xlsx(
        outputs["comparison_xlsx"],
        summary=summary,
        metric_rows=metric_delta_rows(metrics),
        audit_sample=audit_sample,
        over_sanitization_candidates=over_sanitization_candidates,
        residual_samples=residual_samples,
        bot_seed_delta_sample=bot_seed_delta_sample,
    )
    outputs["comparison_report_md"].write_text(markdown_report(summary), encoding="utf-8")
    outputs["audit_prompt_md"].write_text(audit_prompt(summary), encoding="utf-8")
    summary["outputs"] = {key: str(path) for key, path in outputs.items()}
    outputs["summary_json"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def input_paths(config: Stage14ComparisonConfig) -> dict[str, Path]:
    return {
        "before_baseline_summary": config.before_baseline_root / "summary.json",
        "after_baseline_summary": config.after_baseline_root / "summary.json",
        "before_kb_summary": config.before_kb_root / "summary.json",
        "after_kb_summary": config.after_kb_root / "summary.json",
        "before_rop_summary": config.before_rop_root / "summary.json",
        "after_rop_summary": config.after_rop_root / "summary.json",
        "before_kb_enriched": config.before_kb_root / "enriched_reviews.csv",
        "after_kb_enriched": config.after_kb_root / "enriched_reviews.csv",
        "before_kb_bot_seeds": config.before_kb_root / "bot_knowledge_seeds.csv",
        "after_kb_bot_seeds": config.after_kb_root / "bot_knowledge_seeds.csv",
        "before_rop_validation": config.before_rop_root / "rop_validation.csv",
        "after_rop_validation": config.after_rop_root / "rop_validation.csv",
        "before_rop_bot_drafts": config.before_rop_root / "bot_knowledge_drafts.csv",
        "after_rop_bot_drafts": config.after_rop_root / "bot_knowledge_drafts.csv",
    }


def build_metric_delta(before_baseline: dict[str, Any], after_baseline: dict[str, Any]) -> dict[str, dict[str, Any]]:
    keys = [
        "kb_no_live_revenue_risk",
        "kb_bot_ready_money_or_terms",
        "kb_ideal_answer_brand_risk",
        "kb_bot_safe_answer_brand_risk",
        "kb_bot_safe_answer_personal_data_risk",
        "rop_p0_no_live_or_artifact",
        "rop_revenue_risk_no_live_or_artifact",
        "rop_bot_candidate_money_or_terms",
        "rop_bot_safe_answer_brand_risk",
        "rop_bot_safe_answer_personal_data_risk",
    ]
    before = before_baseline.get("baseline_risks", {})
    after = after_baseline.get("baseline_risks", {})
    deltas: dict[str, dict[str, Any]] = {}
    for key in keys:
        before_value = _to_number(before.get(key))
        after_value = _to_number(after.get(key))
        deltas[key] = {
            "before": before_value,
            "after": after_value,
            "delta": None if before_value is None or after_value is None else after_value - before_value,
            "improved": bool(before_value is not None and after_value is not None and after_value < before_value),
            "regressed": bool(before_value is not None and after_value is not None and after_value > before_value),
        }
    after_kb = after_baseline.get("kb_metrics", {})
    for key in ("raw_ideal_answer_brand_risk", "raw_ideal_answer_money_or_terms"):
        deltas[f"kb_{key}"] = {"before": None, "after": _to_number(after_kb.get(key)), "delta": None, "improved": False, "regressed": False}
    return deltas


def build_schema_report(
    after_kb_rows: list[dict[str, str]],
    after_rop_rows: list[dict[str, str]],
    after_bot_seed_rows: list[dict[str, str]],
    after_rop_bot_rows: list[dict[str, str]],
) -> dict[str, Any]:
    kb_columns = set(after_kb_rows[0].keys()) if after_kb_rows else set()
    rop_columns = set(after_rop_rows[0].keys()) if after_rop_rows else set()
    bot_seed_columns = set(after_bot_seed_rows[0].keys()) if after_bot_seed_rows else set()
    rop_bot_columns = set(after_rop_bot_rows[0].keys()) if after_rop_bot_rows else set()
    return {
        "kb_required_columns_present": sorted(REQUIRED_AFTER_KB_COLUMNS & kb_columns),
        "kb_required_columns_missing": sorted(REQUIRED_AFTER_KB_COLUMNS - kb_columns),
        "rop_required_columns_present": sorted(REQUIRED_AFTER_ROP_COLUMNS & rop_columns),
        "rop_required_columns_missing": sorted(REQUIRED_AFTER_ROP_COLUMNS - rop_columns),
        "bot_seed_uses_safe_column": "Безопасный ответ для бота" in bot_seed_columns,
        "rop_bot_drafts_use_safe_column": "Безопасный ответ для бота" in rop_bot_columns,
        "kb_rows": len(after_kb_rows),
        "rop_rows": len(after_rop_rows),
        "bot_seed_rows": len(after_bot_seed_rows),
        "rop_bot_draft_rows": len(after_rop_bot_rows),
    }


def build_sanitizer_metrics(
    *,
    after_kb_summary: dict[str, Any],
    after_kb_rows: list[dict[str, str]],
    after_bot_seed_rows: list[dict[str, str]],
    after_rop_bot_rows: list[dict[str, str]],
) -> dict[str, Any]:
    flag_counter = Counter()
    bot_statuses = Counter()
    safety_statuses = Counter()
    bot_safe_missing = 0
    for row in after_kb_rows:
        flag_counter.update(parse_flags(row.get("sanitizer_flags")))
        bot_status = clean(row.get("bot_seed_status"))
        bot_statuses[bot_status or "unknown"] += 1
        safety_statuses[clean(row.get("bot_safety_status")) or "unknown"] += 1
        if bot_status in BOT_READY_STATUSES and not clean(row.get("bot_safe_answer")):
            bot_safe_missing += 1
    kb_summary_sanitizer = after_kb_summary.get("sanitizer", {})
    return {
        "kb_summary_sanitizer": kb_summary_sanitizer,
        "flag_counts_from_enriched": dict(flag_counter.most_common()),
        "bot_seed_status_counts_from_enriched": dict(bot_statuses.most_common()),
        "bot_safety_status_counts_from_enriched": dict(safety_statuses.most_common()),
        "bot_ready_missing_safe_answer": bot_safe_missing,
        "after_bot_seed_rows": len(after_bot_seed_rows),
        "after_rop_bot_draft_rows": len(after_rop_bot_rows),
        "bot_seed_safe_answer_risk_counts": risk_counts_for_rows(after_bot_seed_rows, "Безопасный ответ для бота"),
        "rop_bot_safe_answer_risk_counts": risk_counts_for_rows(after_rop_bot_rows, "Безопасный ответ для бота"),
    }


def risk_counts_for_rows(rows: list[dict[str, str]], field: str) -> dict[str, int]:
    counts = {"brand": 0, "money_or_terms": 0, "personal_data": 0}
    for row in rows:
        text = clean(row.get(field)) or clean(row.get("Черновик идеального ответа")) or clean(row.get("Идеальный ответ"))
        counts["brand"] += int(has_brand_risk(text))
        counts["money_or_terms"] += int(has_money_or_terms_risk(text))
        counts["personal_data"] += int(has_personal_data_risk(text))
    return counts


def build_residual_risk_samples(
    after_bot_seed_rows: list[dict[str, str]],
    after_rop_bot_rows: list[dict[str, str]],
    limit: int,
) -> list[dict[str, Any]]:
    samples: list[dict[str, Any]] = []
    for source, rows, field in (
        ("kb_bot_knowledge_seeds", after_bot_seed_rows, "Безопасный ответ для бота"),
        ("rop_bot_knowledge_drafts", after_rop_bot_rows, "Безопасный ответ для бота"),
    ):
        for row in rows:
            text = clean(row.get(field))
            brand = has_brand_risk(text)
            money = has_money_or_terms_risk(text)
            personal = has_personal_data_risk(text)
            if not (brand or money or personal):
                continue
            samples.append(
                {
                    "source": source,
                    "moment_id": clean(row.get("ID момента") or row.get("moment_id")),
                    "risk_brand": str(brand),
                    "risk_money_or_terms": str(money),
                    "risk_personal_data": str(personal),
                    "text": text,
                }
            )
            if len(samples) >= limit:
                return samples
    return samples


def build_over_sanitization_candidates(rows: list[dict[str, str]], limit: int) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for row in rows:
        flags = parse_flags(row.get("sanitizer_flags"))
        if not flags:
            continue
        raw = clean(row.get("ideal_answer_example"))
        manager = clean(row.get("ideal_answer_manager_sanitized"))
        bot = clean(row.get("bot_safe_answer"))
        if not raw or not bot:
            continue
        ratio = len(bot) / max(1, len(raw))
        generic_markers = len(re.findall(r"актуальн|менеджер|подтвердит|уточнит|правил|услови|вариант", bot, re.I))
        reason = ""
        if ratio < 0.45:
            reason = "bot_answer_much_shorter_than_raw"
        elif len(bot) < 80 and len(raw) >= 150:
            reason = "bot_answer_short_after_sanitizer"
        elif generic_markers >= 4:
            reason = "bot_answer_many_generic_markers"
        elif "brand_normalized" in flags and "Фотон" not in manager:
            reason = "manager_brand_normalization_needs_check"
        if not reason:
            continue
        candidates.append(
            {
                "reason": reason,
                "moment_id": clean(row.get("moment_id")),
                "source_filename": clean(row.get("source_filename")),
                "signal": clean(row.get("signal_ru")),
                "stage": clean(row.get("stage_ru")),
                "bot_seed_status": clean(row.get("bot_seed_status")),
                "quality_score": clean(row.get("overall_quality_score")),
                "sanitizer_flags": clean(row.get("sanitizer_flags")),
                "raw_len": len(raw),
                "bot_len": len(bot),
                "bot_to_raw_ratio": round(ratio, 3),
                "generic_marker_count": generic_markers,
                "customer_question": clean(row.get("customer_question_sanitized") or row.get("customer_question")),
                "raw_ideal_answer": raw,
                "manager_ideal_answer": manager,
                "bot_safe_answer": bot,
            }
        )
    return sorted(candidates, key=lambda row: (str(row["reason"]), float(row["bot_to_raw_ratio"]), str(row["moment_id"])))[:limit]


def build_audit_sample(
    after_kb_rows: list[dict[str, str]],
    after_rop_rows: list[dict[str, str]],
    after_rop_bot_rows: list[dict[str, str]],
    limit: int,
) -> list[dict[str, Any]]:
    buckets: list[tuple[str, list[dict[str, str]], int]] = []
    buckets.append(("brand_sanitized", [row for row in after_kb_rows if "brand_normalized" in parse_flags(row.get("sanitizer_flags"))], 25))
    buckets.append(("money_terms_sanitized", [row for row in after_kb_rows if any(flag in parse_flags(row.get("sanitizer_flags")) for flag in ("price_redacted", "discount_terms_redacted", "percent_redacted"))], 35))
    buckets.append(("legal_deadline_sanitized", [row for row in after_kb_rows if any(flag in parse_flags(row.get("sanitizer_flags")) for flag in ("refund_policy_redacted", "deadline_redacted"))], 35))
    buckets.append(("installment_sanitized", [row for row in after_kb_rows if "installment_terms_redacted" in parse_flags(row.get("sanitizer_flags"))], 20))
    buckets.append(("personal_data_sanitized", [row for row in after_kb_rows if any(flag in parse_flags(row.get("sanitizer_flags")) for flag in ("person_name_redacted", "email_redacted", "phone_redacted"))], 25))
    buckets.append(("bot_ready_clean_no_changes", [row for row in after_kb_rows if clean(row.get("bot_seed_status")) == "ready_for_bot_draft" and not parse_flags(row.get("sanitizer_flags"))], 20))
    buckets.append(("rop_revenue_risk", [row for row in after_rop_rows if "Риск потери выручки" in clean(row.get("Категория проверки"))], 20))
    buckets.append(("rop_top_answer", [row for row in after_rop_rows if "ТОП ответ" in clean(row.get("Категория проверки"))], 20))
    buckets.append(("rop_bot_draft", after_rop_bot_rows, 20))

    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for bucket, rows, cap in buckets:
        for row in stable_sample(rows, cap):
            key = clean(row.get("moment_id") or row.get("ID момента")) or clean(row.get("source_filename") or row.get("Файл звонка"))
            if not key or key in seen:
                continue
            seen.add(key)
            selected.append(audit_row(bucket, row))
            if len(selected) >= limit:
                return selected
    if len(selected) < limit:
        existing_moments = {clean(row.get("moment_id")) for row in selected}
        for row in stable_sample(after_kb_rows, len(after_kb_rows)):
            moment_id = clean(row.get("moment_id"))
            if moment_id in existing_moments:
                continue
            selected.append(audit_row("coverage_filler", row))
            existing_moments.add(moment_id)
            if len(selected) >= limit:
                break
    return selected


def audit_row(bucket: str, row: dict[str, str]) -> dict[str, Any]:
    bot_safe = clean(row.get("bot_safe_answer") or row.get("Безопасный ответ для бота"))
    manager_ideal = clean(row.get("ideal_answer_manager_sanitized") or row.get("Идеальный ответ для менеджера") or row.get("Идеальный ответ"))
    raw_ideal = clean(row.get("ideal_answer_example"))
    text_for_risks = bot_safe or clean(row.get("Черновик идеального ответа"))
    return {
        "audit_bucket": bucket,
        "review_question": review_question_for_bucket(bucket),
        "moment_id": clean(row.get("moment_id") or row.get("ID момента")),
        "source_filename": clean(row.get("source_filename") or row.get("Файл звонка")),
        "manager_name": clean(row.get("manager_name") or row.get("Менеджер")),
        "started_at": clean(row.get("started_at") or row.get("Дата звонка")),
        "signal": clean(row.get("signal_ru") or row.get("Сигнал клиента")),
        "stage": clean(row.get("stage_ru") or row.get("Стадия")),
        "answer_pattern": clean(row.get("answer_pattern_ru") or row.get("Паттерн ответа")),
        "outcome": clean(row.get("final_outcome_ru") or row.get("Итог сделки")),
        "category": clean(row.get("Категория проверки")),
        "quality_score": clean(row.get("overall_quality_score") or row.get("Оценка")),
        "bot_seed_status": clean(row.get("bot_seed_status_ru") or row.get("Статус для бота")),
        "bot_safety_status": clean(row.get("bot_safety_status_ru") or row.get("Статус sanitizer")),
        "sanitizer_flags": clean(row.get("sanitizer_flags") or row.get("Флаги sanitizer")),
        "customer_question": clean(row.get("customer_question_sanitized") or row.get("customer_question") or row.get("Вопрос клиента")),
        "manager_answer": clean(row.get("manager_answer") or row.get("Ответ менеджера")),
        "raw_ideal_answer": raw_ideal,
        "manager_ideal_answer": manager_ideal,
        "bot_safe_answer": bot_safe,
        "risk_brand_in_bot_safe": str(has_brand_risk(text_for_risks)),
        "risk_money_or_terms_in_bot_safe": str(has_money_or_terms_risk(text_for_risks)),
        "risk_personal_data_in_bot_safe": str(has_personal_data_risk(text_for_risks)),
        "what_to_check": "Оценить: 1) безопасно ли для бота, 2) не потерян ли смысл, 3) понятно ли РОПу, 4) не нужен ли ручной rewrite.",
    }


def build_bot_seed_delta_sample(before_rows: list[dict[str, str]], after_rows: list[dict[str, str]], limit: int) -> list[dict[str, Any]]:
    before_by_key = {bot_seed_key(row): row for row in before_rows if bot_seed_key(row)}
    rows: list[dict[str, Any]] = []
    for after in stable_sample(after_rows, len(after_rows)):
        key = bot_seed_key(after)
        before = before_by_key.get(key, {})
        before_answer = clean(before.get("Черновик идеального ответа")) or clean(before.get("Идеальный ответ"))
        after_answer = clean(after.get("Безопасный ответ для бота")) or clean(after.get("Черновик идеального ответа"))
        if not after_answer:
            continue
        rows.append(
            {
                "moment_id": clean(after.get("ID момента")) or clean(before.get("ID момента")),
                "signal": clean(after.get("Сигнал клиента")) or clean(before.get("Сигнал клиента")),
                "status_after": clean(after.get("Статус для бота")),
                "sanitizer_flags_after": clean(after.get("Флаги sanitizer")),
                "before_bot_draft": before_answer,
                "after_bot_safe_answer": after_answer,
                "before_brand_risk": str(has_brand_risk(before_answer)),
                "before_money_or_terms_risk": str(has_money_or_terms_risk(before_answer)),
                "after_brand_risk": str(has_brand_risk(after_answer)),
                "after_money_or_terms_risk": str(has_money_or_terms_risk(after_answer)),
            }
        )
        if len(rows) >= limit:
            break
    return rows


def build_acceptance_report(
    *,
    metrics: dict[str, dict[str, Any]],
    schema: dict[str, Any],
    sanitizer_metrics: dict[str, Any],
    residual_samples: list[dict[str, Any]],
    audit_sample: list[dict[str, Any]],
) -> dict[str, Any]:
    checks = {
        "required_kb_columns_present": not schema["kb_required_columns_missing"],
        "required_rop_columns_present": not schema["rop_required_columns_missing"],
        "bot_seed_safe_columns_present": bool(schema["bot_seed_uses_safe_column"] and schema["rop_bot_drafts_use_safe_column"]),
        "no_residual_bot_safe_risks": not residual_samples,
        "kb_no_live_revenue_risk_zero": metric_after(metrics, "kb_no_live_revenue_risk") == 0,
        "rop_p0_no_live_or_artifact_zero": metric_after(metrics, "rop_p0_no_live_or_artifact") == 0,
        "rop_revenue_no_live_or_artifact_zero": metric_after(metrics, "rop_revenue_risk_no_live_or_artifact") == 0,
        "kb_bot_ready_money_or_terms_zero": metric_after(metrics, "kb_bot_ready_money_or_terms") == 0,
        "rop_bot_candidate_money_or_terms_zero": metric_after(metrics, "rop_bot_candidate_money_or_terms") == 0,
        "bot_ready_rows_have_safe_answer": sanitizer_metrics["bot_ready_missing_safe_answer"] == 0,
        "audit_sample_built": len(audit_sample) >= 100,
    }
    warnings: list[str] = []
    if len(audit_sample) < 200:
        warnings.append("Audit sample меньше 200 строк; для текущего объема это допустимо только если мало кандидатов в отдельных strata.")
    if not checks["audit_sample_built"]:
        warnings.append("Audit sample слишком маленький для внешнего GPT/Claude-аудита.")
    passed = all(checks.values())
    return {
        "passed": passed,
        "checks": checks,
        "warnings": warnings,
        "interpretation": "Stage 14 acceptance пройден, если все hard checks=true. Warnings требуют внимания, но не блокируют stage 15.",
    }


def markdown_report(summary: dict[str, Any]) -> str:
    lines = [
        "# Stage 14 Quality Comparison Report",
        "",
        f"Generated at: `{summary['generated_at']}`",
        "",
        "## Decision",
        "",
        f"- Acceptance passed: `{summary['acceptance']['passed']}`",
        f"- Audit sample rows: `{summary['audit_sample']['rows']}`",
        f"- Over-sanitization candidates: `{summary['over_sanitization_candidates']['rows']}`",
        f"- Residual risk samples: `{summary['residual_risk_samples']['rows']}`",
        "",
        "## Key Metric Deltas",
        "",
        "| Metric | Before | After | Delta |",
        "|---|---:|---:|---:|",
    ]
    for key, payload in summary["metric_deltas"].items():
        before = payload.get("before")
        after = payload.get("after")
        delta = payload.get("delta")
        lines.append(f"| `{key}` | {display_number(before)} | {display_number(after)} | {display_number(delta)} |")
    lines.extend(
        [
            "",
            "## Acceptance Checks",
            "",
        ]
    )
    for key, value in summary["acceptance"]["checks"].items():
        lines.append(f"- `{key}`: `{value}`")
    if summary["acceptance"].get("warnings"):
        lines.extend(["", "## Warnings", ""])
        for warning in summary["acceptance"]["warnings"]:
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Sanitizer Metrics",
            "",
            f"- Bot-ready rows missing safe answer: `{summary['sanitizer_metrics']['bot_ready_missing_safe_answer']}`",
            f"- Bot seed safe-answer risks: `{summary['sanitizer_metrics']['bot_seed_safe_answer_risk_counts']}`",
            f"- ROP bot safe-answer risks: `{summary['sanitizer_metrics']['rop_bot_safe_answer_risk_counts']}`",
            f"- Top sanitizer flags: `{dict(list(summary['sanitizer_metrics']['flag_counts_from_enriched'].items())[:10])}`",
            "",
            "## Audit Sample Buckets",
            "",
        ]
    )
    for bucket, count in summary["audit_sample"]["by_bucket"].items():
        lines.append(f"- `{bucket}`: {count}")
    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            "Stage 13 removed unsafe bot/ROP leakage without changing raw source fields. Stage 14 does not prove semantic perfection; it proves that safety gates are measurable, residual bot-safe risks are zero, and a stratified review package exists to check usefulness and over-sanitization.",
            "",
            "## Next Step",
            "",
            "Run GPT/Claude or ROP audit on `audit_sample.csv` and `over_sanitization_candidates.csv`. If accepted, proceed to stage 15: wire these gates into the permanent pipeline before KB/ROP/bot/CRM exports.",
        ]
    )
    return "\n".join(lines) + "\n"


def audit_prompt(summary: dict[str, Any]) -> str:
    return f"""# Prompt for Stage 14 External Audit

Тебе передают audit package для проверки качества sanitizer после исправления обработки звонков.

Файлы:

- `audit_sample.csv` — стратифицированная выборка {summary['audit_sample']['rows']} строк.
- `over_sanitization_candidates.csv` — кандидаты, где sanitizer мог сделать bot-safe ответ слишком общим.
- `metric_delta.csv` — сравнение v2/v3 метрик.
- `summary.json` — полный machine-readable отчет.

Задача:

1. Проверь, что `Безопасный ответ для бота` / `bot_safe_answer` не содержит конкретных цен, скидок, сроков, возвратов, рассрочек, неправильного бренда, телефонов/email/ФИО.
2. Проверь, что `Идеальный ответ для менеджера` полезен РОПу и не потерял смысл исходного ответа.
3. Найди over-sanitization: где bot-safe ответ стал слишком общим и не поможет клиенту.
4. Проверь, что no-live/voicemail/IVR/ASR-мусор не попал в bot/revenue/top-answer строки.
5. Верни таблицу решений: `audit_id`, `decision` (`ok`, `needs_rewrite`, `unsafe`, `not_useful`, `manual_review`), `reason`, `suggested_fix`.

Важно: это не финальная база Telegram-бота, а safety- и usefulness-аудит черновиков. Менеджерский ответ и ответ бота должны оцениваться раздельно.
"""


def write_xlsx(
    path: Path,
    *,
    summary: dict[str, Any],
    metric_rows: list[dict[str, Any]],
    audit_sample: list[dict[str, Any]],
    over_sanitization_candidates: list[dict[str, Any]],
    residual_samples: list[dict[str, Any]],
    bot_seed_delta_sample: list[dict[str, Any]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    summary_rows = [
        {"Раздел": "Decision", "Метрика": "Acceptance passed", "Значение": summary["acceptance"]["passed"], "Комментарий": "Все hard gates stage14."},
        {"Раздел": "Decision", "Метрика": "Audit sample rows", "Значение": summary["audit_sample"]["rows"], "Комментарий": "Выборка для GPT/Claude/РОП-аудита."},
        {"Раздел": "Decision", "Метрика": "Residual risk rows", "Значение": summary["residual_risk_samples"]["rows"], "Комментарий": "Должно быть 0."},
        {"Раздел": "Decision", "Метрика": "Over-sanitization candidates", "Значение": summary["over_sanitization_candidates"]["rows"], "Комментарий": "Очередь проверки полезности, не список ошибок."},
    ]
    for key, value in summary["acceptance"]["checks"].items():
        summary_rows.append({"Раздел": "Acceptance", "Метрика": key, "Значение": value, "Комментарий": ""})
    for bucket, count in summary["audit_sample"]["by_bucket"].items():
        summary_rows.append({"Раздел": "Audit buckets", "Метрика": bucket, "Значение": count, "Комментарий": ""})

    add_sheet(workbook, "Summary", summary_rows)
    add_sheet(workbook, "Metric Delta", metric_rows)
    add_sheet(workbook, "Audit Sample", audit_sample)
    add_sheet(workbook, "Over Sanitization", over_sanitization_candidates)
    add_sheet(workbook, "Residual Risks", residual_samples)
    add_sheet(workbook, "Bot Before After", bot_seed_delta_sample)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 300)):
            for cell in row:
                if isinstance(cell.value, str) and len(cell.value) > 70:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in sheet.columns:
            letter = column_cells[0].column_letter
            max_len = 0
            for cell in column_cells[:250]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 90))
            sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 62)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def add_sheet(workbook: Any, name: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(name[:31])
    if not rows:
        sheet.append(["empty"])
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])


def metric_delta_rows(metrics: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "metric": key,
            "before": payload.get("before"),
            "after": payload.get("after"),
            "delta": payload.get("delta"),
            "improved": payload.get("improved"),
            "regressed": payload.get("regressed"),
        }
        for key, payload in metrics.items()
    ]


def summary_key_metrics(summary: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for section in ("totals", "llm_review", "sanitizer"):
        value = summary.get(section)
        if isinstance(value, dict):
            result[section] = value
    return result


def stable_sample(rows: list[dict[str, str]], limit: int) -> list[dict[str, str]]:
    ordered = sorted(rows, key=lambda row: stable_key(row))
    return ordered[: max(0, limit)]


def stable_key(row: dict[str, str]) -> str:
    joined = "|".join(
        [
            clean(row.get("moment_id") or row.get("ID момента")),
            clean(row.get("source_filename") or row.get("Файл звонка")),
            clean(row.get("customer_question") or row.get("Вопрос клиента")),
        ]
    )
    return hashlib.sha256(joined.encode("utf-8")).hexdigest()


def bot_seed_key(row: dict[str, str]) -> str:
    return "|".join([clean(row.get("ID момента")), clean(row.get("Код сигнала") or row.get("Сигнал клиента")), clean(row.get("Пример вопроса клиента"))])


def parse_flags(value: Any) -> tuple[str, ...]:
    flags = []
    for item in clean(value).split("|"):
        item = item.strip()
        if item:
            flags.append(item)
    return tuple(flags)


def review_question_for_bucket(bucket: str) -> str:
    if bucket == "brand_sanitized":
        return "Проверить, что бренд нормализован корректно и не создана ложная связь с МФТИ."
    if bucket == "money_terms_sanitized":
        return "Проверить, что цена/скидка убраны, но ответ остался полезным."
    if bucket == "legal_deadline_sanitized":
        return "Проверить, что сроки/возвраты/договорные обещания убраны безопасно."
    if bucket == "installment_sanitized":
        return "Проверить, что условия рассрочки не обещаются ботом напрямую."
    if bucket == "personal_data_sanitized":
        return "Проверить, что ПДн не попали в bot-safe ответ."
    if bucket == "bot_ready_clean_no_changes":
        return "Проверить, что строка действительно безопасна без sanitizer-замен."
    if bucket == "rop_revenue_risk":
        return "Проверить, что это реальный риск выручки, а не no-live/переписка вне звонка."
    if bucket == "rop_top_answer":
        return "Проверить, что пример полезен для скриптов менеджеров."
    if bucket == "rop_bot_draft":
        return "Проверить, что bot draft безопасен и достаточно полезен."
    return "Общая проверка полезности и безопасности."


def metric_after(metrics: dict[str, dict[str, Any]], key: str) -> float | int | None:
    return metrics.get(key, {}).get("after")


def display_number(value: Any) -> str:
    return "" if value is None else str(value)


def _to_number(value: Any) -> float | int | None:
    if value is None:
        return None
    try:
        number = float(str(value).strip())
    except ValueError:
        return None
    if number.is_integer():
        return int(number)
    return number


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", "").strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return " ".join(text.split())


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _require_files(paths: Iterable[Path]) -> None:
    missing = [str(path) for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError("Missing required stage14 inputs: " + ", ".join(missing))


def _fingerprint(path: Path, project_root: Path) -> dict[str, Any]:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    try:
        rel = path.resolve().relative_to(project_root)
    except ValueError:
        rel = path.resolve()
    return {"path": str(rel), "size_bytes": path.stat().st_size, "sha256": digest.hexdigest()}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build stage 14 transcript quality v2/v3 comparison package.")
    parser.add_argument("--project-root", type=Path, default=Path.cwd())
    parser.add_argument("--before-kb-root", type=Path, required=True)
    parser.add_argument("--after-kb-root", type=Path, required=True)
    parser.add_argument("--before-rop-root", type=Path, required=True)
    parser.add_argument("--after-rop-root", type=Path, required=True)
    parser.add_argument("--before-baseline-root", type=Path, required=True)
    parser.add_argument("--after-baseline-root", type=Path, required=True)
    parser.add_argument("--out-root", type=Path, required=True)
    parser.add_argument("--audit-sample-limit", type=int, default=200)
    parser.add_argument("--over-sanitization-limit", type=int, default=250)
    parser.add_argument("--residual-sample-limit", type=int, default=100)
    return parser.parse_args(argv)


def config_from_args(args: argparse.Namespace) -> Stage14ComparisonConfig:
    project_root = args.project_root.expanduser().resolve()
    return Stage14ComparisonConfig(
        project_root=project_root,
        before_kb_root=(project_root / args.before_kb_root).resolve() if not args.before_kb_root.is_absolute() else args.before_kb_root.resolve(),
        after_kb_root=(project_root / args.after_kb_root).resolve() if not args.after_kb_root.is_absolute() else args.after_kb_root.resolve(),
        before_rop_root=(project_root / args.before_rop_root).resolve() if not args.before_rop_root.is_absolute() else args.before_rop_root.resolve(),
        after_rop_root=(project_root / args.after_rop_root).resolve() if not args.after_rop_root.is_absolute() else args.after_rop_root.resolve(),
        before_baseline_root=(project_root / args.before_baseline_root).resolve() if not args.before_baseline_root.is_absolute() else args.before_baseline_root.resolve(),
        after_baseline_root=(project_root / args.after_baseline_root).resolve() if not args.after_baseline_root.is_absolute() else args.after_baseline_root.resolve(),
        out_root=(project_root / args.out_root).resolve() if not args.out_root.is_absolute() else args.out_root.resolve(),
        audit_sample_limit=args.audit_sample_limit,
        over_sanitization_limit=args.over_sanitization_limit,
        residual_sample_limit=args.residual_sample_limit,
    )


__all__ = [
    "Stage14ComparisonConfig",
    "build_stage14_quality_comparison",
    "config_from_args",
    "parse_args",
]
