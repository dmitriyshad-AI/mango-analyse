#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import shutil
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_REVIEW_ROOT = ROOT / "stable_runtime/deal_aware_stage709_review_20260514_iter01"
DEFAULT_STAGE100_ROOT = ROOT / "stable_runtime/deal_aware_stage100_rop_final_20260514_v1"
DEFAULT_OUT_ROOT = ROOT / "stable_runtime/deal_aware_stage709_all_batches_20260514_v1"

REVIEW_COLUMNS = [
    "sample_id",
    "sample_bucket_ru",
    "review_url_hint",
    "selected_deal_id",
    "selected_deal_name",
    "selected_pipeline_name",
    "selected_status_name",
    "selected_loss_reason",
    "phones",
    "managers",
    "candidate_call_count",
    "candidate_phone_count",
    "tallanto_context_status",
    "AI-фактический статус сделки",
    "AI-приоритет сделки",
    "AI-рекомендованный следующий шаг",
    "AI-сводка по сделке",
    "AI-история по сделке",
    "AI-Tallanto статус по сделке",
    "AI-предупреждение по сделке",
    "risk_flags_ru",
    "risk_classes",
    "stage6_status",
    "stage6_reason",
    "stage6_finding_types",
]

ROP_COLUMNS = [
    ("rop_row_decision", "ready_for_rop / minor_comment / needs_fix_before_rop / block"),
    ("summary_correctness", "ok / minor_inaccuracy / wrong_or_misleading / cannot_judge"),
    ("next_step_quality", "ok_manager_action / too_passive / customer_side_only / wrong_action / missing"),
    ("deal_status_priority_quality", "ok / status_wrong / priority_wrong / amo_tallanto_mismatch_not_handled / cannot_judge"),
    ("tallanto_block_quality", "ok / wrong_student / wrong_finance_or_attendance / too_raw_or_unreadable / not_relevant"),
    ("history_relevance", "ok / irrelevant_calls / important_call_missing / too_verbose / cannot_judge"),
    ("sales_usefulness", "useful_as_is / useful_after_small_edit / not_useful / dangerous"),
    ("issue_type", "wrong_deal_binding / wrong_customer_context / wrong_next_step / payment_conflict / service_feedback_routing / amo_tallanto_mismatch / bad_tenant_terms / too_verbose / duplicate_fields / other"),
    ("severity", "P0_blocker / P1_fix_before_rop / P2_minor / P3_note"),
    ("comment", "Свободный комментарий РОПа; обязателен, если есть любая проблема."),
]

RENAME_COLUMNS = {
    "batch_id": "Батч",
    "batch_sequence": "Номер в батче",
    "global_sequence": "Номер в общем списке",
    "review_scope_ru": "Контур проверки",
    "review_status_ru": "Статус проверки",
    "sample_id": "ID примера",
    "sample_bucket_ru": "Класс риска",
    "review_url_hint": "Ссылка на сделку AMO",
    "selected_deal_id": "ID сделки AMO",
    "selected_deal_name": "Название сделки",
    "selected_pipeline_name": "Воронка",
    "selected_status_name": "Статус сделки",
    "selected_loss_reason": "Причина отказа",
    "phones": "Телефоны",
    "managers": "Менеджеры",
    "candidate_call_count": "Кол-во звонков",
    "candidate_phone_count": "Кол-во телефонов",
    "tallanto_context_status": "Статус Tallanto",
    "AI-фактический статус сделки": "AI: фактический статус",
    "AI-приоритет сделки": "AI: приоритет",
    "AI-рекомендованный следующий шаг": "AI: следующий шаг",
    "AI-сводка по сделке": "AI: сводка по сделке",
    "AI-история по сделке": "AI: история по сделке",
    "AI-Tallanto статус по сделке": "AI: Tallanto статус",
    "AI-предупреждение по сделке": "AI: предупреждение",
    "risk_flags_ru": "Риски простым языком",
    "risk_classes": "Технические классы риска",
    "stage6_status": "Stage6 статус",
    "stage6_reason": "Stage6 причина",
    "stage6_finding_types": "Stage6 находки",
    "rop_row_decision": "Решение РОПа",
    "summary_correctness": "Сводка верна?",
    "next_step_quality": "Следующий шаг верен?",
    "deal_status_priority_quality": "Статус/приоритет верны?",
    "tallanto_block_quality": "Tallanto-блок верен?",
    "history_relevance": "История релевантна?",
    "sales_usefulness": "Полезность для продаж",
    "issue_type": "Тип проблемы",
    "severity": "Серьёзность",
    "comment": "Комментарий РОПа",
}

VALIDATIONS = {
    "Решение РОПа": "ready_for_rop,minor_comment,needs_fix_before_rop,block",
    "Сводка верна?": "ok,minor_inaccuracy,wrong_or_misleading,cannot_judge",
    "Следующий шаг верен?": "ok_manager_action,too_passive,customer_side_only,wrong_action,missing",
    "Статус/приоритет верны?": "ok,status_wrong,priority_wrong,amo_tallanto_mismatch_not_handled,cannot_judge",
    "Tallanto-блок верен?": "ok,wrong_student,wrong_finance_or_attendance,too_raw_or_unreadable,not_relevant",
    "История релевантна?": "ok,irrelevant_calls,important_call_missing,too_verbose,cannot_judge",
    "Полезность для продаж": "useful_as_is,useful_after_small_edit,not_useful,dangerous",
    "Тип проблемы": "wrong_deal_binding,wrong_customer_context,wrong_next_step,payment_conflict,service_feedback_routing,amo_tallanto_mismatch,bad_tenant_terms,too_verbose,duplicate_fields,other",
    "Серьёзность": "P0_blocker,P1_fix_before_rop,P2_minor,P3_note",
}

RISK_WEIGHTS = {
    "blocked_completed_payment_next_step_conflict": 100,
    "blocked_cross_field_duplicate_information": 90,
    "multiple_tallanto_matches": 80,
    "future_loss_reactivation": 70,
    "paid_or_success_context": 55,
    "payment_stage": 45,
    "amo_tallanto_mismatch": 40,
    "service_feedback": 35,
    "no_reliable_tallanto_match": 25,
    "multi_phone_history": 22,
    "long_history": 18,
    "overdue_tasks": 14,
    "review_priority": 10,
    "stage2_confidence_low": 1,
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Split remaining deal-aware AMO candidates into review batches.")
    parser.add_argument("--review-root", default=str(DEFAULT_REVIEW_ROOT))
    parser.add_argument("--stage100-root", default=str(DEFAULT_STAGE100_ROOT))
    parser.add_argument("--out-root", default=str(DEFAULT_OUT_ROOT))
    parser.add_argument("--batch-count", type=int, default=6)
    return parser.parse_args()


def safe_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def int_or_zero(value: Any) -> int:
    try:
        return int(float(safe_text(value).replace(",", ".")))
    except ValueError:
        return 0


def risk_score(row: pd.Series) -> int:
    classes = [item for item in safe_text(row.get("risk_classes")).split("|") if item]
    score = sum(RISK_WEIGHTS.get(cls, 0) for cls in classes)
    score += min(int_or_zero(row.get("candidate_call_count")), 25)
    if safe_text(row.get("stage6_status")) == "blocked":
        score += 100
    return score


def prepare_review_rows(df: pd.DataFrame) -> pd.DataFrame:
    rows = df.copy()
    if "manager_warning_ru" in rows.columns:
        rows["AI-предупреждение по сделке"] = rows["manager_warning_ru"]
    for column in REVIEW_COLUMNS:
        if column not in rows.columns:
            rows[column] = ""
    out_columns = [
        "batch_id",
        "batch_sequence",
        "global_sequence",
        "review_scope_ru",
        "review_status_ru",
        *REVIEW_COLUMNS,
    ]
    out = rows[out_columns].copy()
    for column, instruction in ROP_COLUMNS:
        if column not in out.columns:
            out[column] = ""
        out[f"{column}_allowed_values"] = instruction
    return out


def split_remaining(remaining: pd.DataFrame, batch_count: int) -> pd.DataFrame:
    capacities = balanced_capacities(len(remaining), batch_count)
    working = remaining.copy()
    working["_risk_score"] = working.apply(risk_score, axis=1)
    working["_primary"] = working["primary_risk_class"].fillna("")
    working = working.sort_values(
        ["_risk_score", "_primary", "last_call_at", "selected_deal_id"],
        ascending=[False, True, False, True],
    )
    batches: list[list[int]] = [[] for _ in range(batch_count)]
    batch_scores = [0 for _ in range(batch_count)]
    for idx, row in working.iterrows():
        candidates = [
            batch_index
            for batch_index, batch in enumerate(batches)
            if len(batch) < capacities[batch_index]
        ]
        target = min(candidates, key=lambda batch_index: (batch_scores[batch_index], len(batches[batch_index]), batch_index))
        batches[target].append(idx)
        batch_scores[target] += int(row["_risk_score"])

    assigned = working.copy()
    assigned["batch_id"] = ""
    assigned["batch_sequence"] = 0
    for batch_index, indexes in enumerate(batches, start=1):
        batch_name = f"batch_{batch_index:02d}"
        ordered_indexes = sorted(indexes, key=lambda idx: int_or_zero(working.loc[idx, "stage6_row_index"]))
        for sequence, idx in enumerate(ordered_indexes, start=1):
            assigned.at[idx, "batch_id"] = batch_name
            assigned.at[idx, "batch_sequence"] = sequence
    return assigned.drop(columns=["_risk_score", "_primary"]).sort_values(["batch_id", "batch_sequence"])


def balanced_capacities(total: int, batch_count: int) -> list[int]:
    base = total // batch_count
    remainder = total % batch_count
    return [base + (1 if index < remainder else 0) for index in range(batch_count)]


def write_rubric(path: Path) -> None:
    rows = [{"field": name, "allowed_values": allowed} for name, allowed in ROP_COLUMNS]
    pd.DataFrame(rows).to_csv(path, index=False)


def write_workbook(path: Path, rows: pd.DataFrame, summary_rows: list[list[Any]], rubric: pd.DataFrame) -> None:
    allowed_cols = [column for column in rows.columns if column.endswith("_allowed_values")]
    main = rows[[column for column in rows.columns if column not in allowed_cols]].copy().rename(columns=RENAME_COLUMNS)
    summary = pd.DataFrame(summary_rows, columns=["Показатель", "Значение"])
    rubric_ru = rubric.rename(columns={"field": "Поле", "allowed_values": "Допустимые значения"})

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Сводка")
        main.to_excel(writer, index=False, sheet_name="Проверка РОП")
        rubric_ru.to_excel(writer, index=False, sheet_name="Рубрика")
    style_workbook(path)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D0D7DE")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions

    ws = wb["Проверка РОП"]
    for index in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = 22
    widths = {
        "A": 13,
        "B": 13,
        "C": 13,
        "D": 24,
        "E": 28,
        "F": 12,
        "G": 42,
        "H": 26,
        "I": 13,
        "J": 28,
        "U": 34,
        "V": 54,
        "W": 62,
        "X": 42,
        "Y": 36,
        "Z": 44,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 96
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    for name, values in VALIDATIONS.items():
        if name not in headers:
            continue
        validation = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{headers[name]}2:{headers[name]}{ws.max_row}")

    ws = wb["Сводка"]
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 78
    for row in ws.iter_rows(min_row=2):
        row[0].fill = light_fill
    wb.save(path)


def class_counts(rows: pd.DataFrame) -> dict[str, int]:
    counter: Counter[str] = Counter()
    for value in rows["risk_classes"].fillna(""):
        for cls in safe_text(value).split("|"):
            if cls:
                counter[cls] += 1
    return dict(counter.most_common())


def status_counts(rows: pd.DataFrame) -> dict[str, int]:
    return {str(key): int(value) for key, value in rows["stage6_status"].value_counts(dropna=False).to_dict().items()}


def write_readme(out_root: Path, summary: dict[str, Any]) -> None:
    lines = [
        "# Deal-aware Stage709 All Batches",
        "",
        "Назначение: offline-пакет для последовательной проверки всех 709 кандидатов AMO.",
        "",
        "Что внутри:",
        "",
        "- `all_709_candidates_for_amo_review.xlsx` — общая таблица по всем кандидатам.",
        "- `all_709_candidates_for_amo_review.csv` — та же таблица в CSV.",
        "- `batch_01` ... `batch_06` — шесть батчей из оставшихся 609 строк.",
        "- `already_reviewed_stage100.csv` — 100 строк, которые уже были в финальном Stage100-пакете.",
        "- `batch_manifest.csv` — сводка по размерам и рискам батчей.",
        "",
        "Live-запись в AMO этим пакетом не разрешена. Это только материал для Claude/РОП-проверки.",
        "",
        "Ключевые числа:",
        "",
        f"- Всего строк: {summary['counts']['total_rows']}",
        f"- Уже в Stage100: {summary['counts']['already_reviewed_stage100_rows']}",
        f"- Разложено в новые батчи: {summary['counts']['remaining_rows']}",
        f"- Размеры батчей: {', '.join(str(item['rows']) for item in summary['batch_manifest'])}",
        "",
    ]
    (out_root / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    args = parse_args()
    review_root = Path(args.review_root).resolve()
    stage100_root = Path(args.stage100_root).resolve()
    out_root = Path(args.out_root).resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    full = pd.read_csv(review_root / "deal_stage6_709_classification.csv")
    stage100 = pd.read_csv(stage100_root / "stratified_preview_100_for_rop.csv")
    stage100_ids = set(stage100["sample_id"].astype(str))
    if len(full) != 709:
        raise SystemExit(f"Expected 709 full rows, got {len(full)}")
    if len(stage100_ids) != 100:
        raise SystemExit(f"Expected 100 unique Stage100 sample ids, got {len(stage100_ids)}")

    full = full.copy()
    full["sample_id"] = full["sample_id"].astype(str)
    full["global_sequence"] = full["stage6_row_index"].apply(int_or_zero)
    already = full[full["sample_id"].isin(stage100_ids)].copy()
    remaining = full[~full["sample_id"].isin(stage100_ids)].copy()
    if len(already) != 100 or len(remaining) != 609:
        raise SystemExit(f"Unexpected split: already={len(already)} remaining={len(remaining)}")

    assigned_remaining = split_remaining(remaining, args.batch_count)
    already["batch_id"] = "stage100_final"
    already["batch_sequence"] = already["global_sequence"].rank(method="first").astype(int)
    already["review_scope_ru"] = "Уже проверено в финальном Stage100-пакете"
    already["review_status_ru"] = "Claude PASS_WITH_LIMITATIONS; готово для проверки РОП"
    assigned_remaining["review_scope_ru"] = "Новый батч из оставшихся 609 строк"
    assigned_remaining["review_status_ru"] = "Ожидает Claude/РОП-проверку"

    combined = pd.concat([already, assigned_remaining], ignore_index=True, sort=False)
    combined = combined.sort_values(["batch_id", "batch_sequence"]).reset_index(drop=True)
    review_rows = prepare_review_rows(combined)

    rubric_path = out_root / "rop_rubric.csv"
    write_rubric(rubric_path)
    rubric = pd.read_csv(rubric_path)
    review_rows.to_csv(out_root / "all_709_candidates_for_amo_review.csv", index=False)
    already_rows = review_rows[review_rows["batch_id"].eq("stage100_final")].copy()
    already_rows.to_csv(out_root / "already_reviewed_stage100.csv", index=False)

    manifest_rows: list[dict[str, Any]] = []
    for batch_id, batch in review_rows.groupby("batch_id", sort=True):
        manifest_rows.append(
            {
                "batch_id": batch_id,
                "rows": int(len(batch)),
                "stage6_status_counts": json.dumps(status_counts(batch), ensure_ascii=False),
                "risk_class_counts": json.dumps(class_counts(batch), ensure_ascii=False),
            }
        )
    manifest = pd.DataFrame(manifest_rows)
    manifest.to_csv(out_root / "batch_manifest.csv", index=False)

    for batch_id in [f"batch_{index:02d}" for index in range(1, args.batch_count + 1)]:
        batch_dir = out_root / batch_id
        batch_dir.mkdir(parents=True, exist_ok=True)
        batch = review_rows[review_rows["batch_id"].eq(batch_id)].copy()
        batch.to_csv(batch_dir / f"{batch_id}_rop_review.csv", index=False)
        batch.to_csv(batch_dir / "stratified_preview_100_for_rop.csv", index=False)
        shutil.copy2(rubric_path, batch_dir / "rop_rubric.csv")
        summary_rows = [
            ["Всего строк в Stage6", len(full)],
            ["Уже проверено в Stage100", len(already)],
            ["Строк в этом батче", len(batch)],
            ["Батч", batch_id],
            ["Live-запись разрешена", "Нет"],
            ["Примечание", "Offline review batch; сначала Claude/РОП, затем отдельная приемка."],
        ]
        write_workbook(batch_dir / f"{batch_id}_rop_review.xlsx", batch, summary_rows, rubric)

    summary_rows = [
        ["Всего строк в Stage6", len(full)],
        ["Уже проверено в Stage100", len(already)],
        ["Осталось и разложено по 6 батчам", len(remaining)],
        ["Live-запись разрешена", "Нет"],
        ["Примечание", "Общая таблица по всем 709 кандидатам AMO: Stage100 + 6 новых батчей."],
    ]
    write_workbook(out_root / "all_709_candidates_for_amo_review.xlsx", review_rows, summary_rows, rubric)

    summary = {
        "schema_version": "deal_aware_stage709_all_batches_v1",
        "generated_at": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
        "inputs": {
            "review_root": str(review_root),
            "stage100_root": str(stage100_root),
            "full_classification": str(review_root / "deal_stage6_709_classification.csv"),
            "stage100_sample": str(stage100_root / "stratified_preview_100_for_rop.csv"),
        },
        "counts": {
            "total_rows": int(len(full)),
            "already_reviewed_stage100_rows": int(len(already)),
            "remaining_rows": int(len(remaining)),
            "batch_count": int(args.batch_count),
        },
        "stage6_status_counts_all_709": status_counts(review_rows),
        "risk_class_counts_all_709": class_counts(review_rows),
        "batch_manifest": manifest_rows,
        "live_write_allowed": False,
        "notes_ru": [
            "609 строк разбиты на 6 сбалансированных по risk_score батчей.",
            "Сбалансированная нарезка нужна, чтобы каждый батч содержал разные классы риска, а не только легкие/тяжелые строки.",
            "Пакет предназначен для Claude/РОП проверки, не для live-write.",
        ],
    }
    (out_root / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_readme(out_root, summary)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
