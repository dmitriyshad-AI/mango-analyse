from __future__ import annotations

import csv
import json
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from mango_mvp.question_catalog.builder import (
    CatalogBuildConfig,
    build_answer_quality_audit,
    build_approved_answer_drafts,
    build_customer_question_catalog,
    build_question_classes,
    build_rop_priority_rows,
    build_rop_review_rows,
)
from mango_mvp.question_catalog.contracts import AnswerTemplate, QuestionClass
from mango_mvp.question_catalog.extractors import build_question_item


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    fields = list(rows[0])
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def test_question_classes_use_raw_classification_not_redacted_sample_text() -> None:
    item = build_question_item(
        tenant_id="foton",
        source_channel="telegram",
        source_ref="telegram:test:1",
        question_raw="Итого 71 250 руб.?",
        manager_raw="",
    )

    classes = build_question_classes([item], tenant_id="foton")

    assert classes[0].canonical_question == "Статус платежа и подтверждение"
    assert classes[0].class_key == "theme:003_payment_status"
    assert classes[0].metadata["theme_id"] == "theme:003_payment_status"


def test_build_customer_question_catalog_e2e(tmp_path: Path) -> None:
    project = tmp_path / "project"
    project.mkdir()
    calls = project / "calls.csv"
    _write_csv(
        calls,
        [
            {
                "moment_id": "m1",
                "started_at": "2026-05-01T10:00:00+00:00",
                "customer_question": "Сколько стоит ЕГЭ по математике для 11 класса?",
                "manager_answer": "Стоимость 50000 рублей.",
                "llm_customer_signal_type": "price_question",
            }
        ],
    )
    telegram_dir = project / "telegram"
    telegram_dir.mkdir()
    telegram = telegram_dir / "messages.jsonl"
    telegram.write_text(
        json.dumps(
            {
                "dialog_id": 100,
                "peer_kind": "user",
                "message_id": 1,
                "date": "2026-03-01T10:00:00+00:00",
                "out": False,
                "text": "Где проходят занятия очно?",
                "has_media": False,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    mail_root = project / "mail"
    folder = mail_root / "regru_edu" / "folder_inbox"
    text_dir = folder / "extracted_text"
    text_dir.mkdir(parents=True)
    text_path = text_dir / "m1.txt"
    text_path.write_text("Добрый день! Пришлите расписание группы.", encoding="utf-8")
    db_path = folder / "mail_archive.sqlite"
    with sqlite3.connect(db_path) as con:
        con.execute(
            """
            CREATE TABLE messages (
              sha256 TEXT PRIMARY KEY,
              message_date_iso TEXT,
              subject TEXT,
              from_header TEXT,
              to_header TEXT,
              mailbox TEXT NOT NULL,
              mailbox_raw TEXT NOT NULL,
              message_kind TEXT NOT NULL,
              extracted_text_path TEXT,
              extracted_text_chars INTEGER
            )
            """
        )
        con.execute(
            "INSERT INTO messages VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                "abc",
                "2026-05-01T10:00:00+00:00",
                "Расписание",
                "client@example.com",
                "edu@kmipt.ru",
                "INBOX",
                "INBOX",
                "external",
                str(text_path),
                40,
            ),
        )
    facts = project / "facts"
    facts.mkdir()
    (facts / "Стоимость 2026.txt").write_text("цены требуют ручной проверки", encoding="utf-8")
    out = project / "product_data" / "question_catalog"
    summary = build_customer_question_catalog(
        CatalogBuildConfig(
            project_root=project,
            out_root=out,
            tenant_id="foton",
            since=datetime(2025, 1, 1, tzinfo=timezone.utc),
            calls_enriched_reviews=calls,
            telegram_messages_jsonl=telegram,
            mail_archive_root=mail_root,
            fact_source_roots=(facts,),
        )
    )

    assert summary["totals"]["question_items"] >= 3
    assert summary["totals"]["question_classes"] >= 3
    assert summary["totals"]["fact_sources"] == 1
    assert summary["counts"]["class_size_distribution"]["buckets"]
    assert "manual_or_fallback_classes_50_plus" in summary["counts"]["class_size_distribution"]
    assert summary["counts"]["class_size_distribution"]["largest_classes"]
    assert (out / "customer_question_items.jsonl").exists()
    assert (out / "customer_question_classes.xlsx").exists()
    assert (out / "rop_question_review_pack.csv").exists()
    assert (out / "rop_question_review_pack.xlsx").exists()
    assert (out / "approved_question_answers_draft.xlsx").exists()
    assert (out / "approved_question_answers_draft.csv").exists()
    assert (out / "approved_question_answers_draft.json").exists()
    assert (out / "rop_review_priority_top100.xlsx").exists()
    assert (out / "answer_quality_check_report.json").exists()
    assert (out / "channel_preview_approved_context_pack.json").exists()
    assert (out / "current_fact_source_registry.json").exists()
    assert "50000" not in (out / "customer_question_items.jsonl").read_text(encoding="utf-8")

    classes_wb = load_workbook(out / "customer_question_classes.xlsx", read_only=True, data_only=True)
    classes_ws = classes_wb[classes_wb.sheetnames[0]]
    classes_headers = [cell.value for cell in next(classes_ws.iter_rows(min_row=1, max_row=1))]
    assert "Класс вопроса" in classes_headers
    assert "Крупный класс вопроса" in classes_headers
    assert "Подкласс вопроса" in classes_headers
    assert "Статус ответа" in classes_headers
    assert "Можно ли боту отвечать" in classes_headers
    assert "Черновик шаблона ответа" in classes_headers
    assert "Примеры для РОПа" in classes_headers
    assert "Безопасные примеры для авто-проверок" in classes_headers
    assert classes_headers.index("Черновик шаблона ответа") < classes_headers.index("Примеры для РОПа")
    assert "canonical_question" not in classes_headers
    assert "answer_status" not in classes_headers

    rop_wb = load_workbook(out / "rop_question_review_pack.xlsx", read_only=True, data_only=True)
    rop_ws = rop_wb[rop_wb.sheetnames[0]]
    rop_headers = [cell.value for cell in next(rop_ws.iter_rows(min_row=1, max_row=1))]
    rop_row = dict(zip(rop_headers, next(rop_ws.iter_rows(min_row=2, max_row=2, values_only=True))))
    assert "Проверять РОПу" in rop_headers
    assert "Крупный класс вопроса" in rop_headers
    assert "Подкласс вопроса" in rop_headers
    assert "Статус ответа" in rop_headers
    assert "Черновик шаблона ответа" in rop_headers
    assert "Примеры для РОПа" in rop_headers
    assert "Безопасные примеры" in rop_headers
    assert rop_headers.index("Черновик шаблона ответа") < rop_headers.index("Примеры для РОПа")
    assert rop_row["Черновик шаблона ответа"]
    assert not str(rop_row["Черновик шаблона ответа"]).startswith("По вопросу")
    assert str(rop_row["Проверять РОПу"]).startswith("да")
    assert rop_row["Статус ответа"] in {
        "шаблон готов, нужен актуальный факт",
        "есть черновик, нужна проверка",
        "только менеджер",
        "нужен ответ РОПа",
    }
    assert rop_row["Можно ли боту"] in {
        "можно после проверки актуальных фактов",
        "только черновик, нужна проверка",
        "только менеджер",
        "нельзя",
    }

    approved = json.loads((out / "approved_question_answers_draft.json").read_text(encoding="utf-8"))
    assert approved["records"]
    assert all(record["auto_approved"] is False for record in approved["records"])

    quality = json.loads((out / "answer_quality_check_report.json").read_text(encoding="utf-8"))
    assert quality["verdict"] == "pass"
    assert quality["auto_approval_errors"] == []


def test_wide_classes_are_blocked_until_split_in_rop_tables() -> None:
    wide = QuestionClass(
        tenant_id="foton",
        canonical_question="общий вопрос",
        narrow_scope="Вопрос клиента про общий вопрос.",
        class_key="service:S5_general_consultation",
        count_total=50,
        examples_redacted=("Можно подробнее?",),
        metadata={
            "parent_question_class": "общий вопрос",
            "question_subclass": "без уточненного подкласса",
        },
    )
    narrow = QuestionClass(
        tenant_id="foton",
        canonical_question="стоимость / что входит в стоимость",
        narrow_scope="Вопрос клиента про стоимость; подкласс: что входит в стоимость.",
        class_key="theme:001_pricing",
        count_total=50,
        examples_redacted=("Что входит в стоимость курса?",),
        required_fact_keys=("price.current",),
        metadata={
            "parent_question_class": "стоимость",
            "question_subclass": "что входит в стоимость",
            "question_subclass_key": "included_services",
        },
    )
    broad_base = QuestionClass(
        tenant_id="foton",
        canonical_question="расписание / общее расписание",
        narrow_scope="Вопрос клиента про расписание; подкласс: общее расписание.",
        class_key="theme:013_schedule",
        count_total=150,
        examples_redacted=("Когда будут занятия?",),
        required_fact_keys=("schedule.current",),
        metadata={
            "parent_question_class": "расписание",
            "question_subclass": "общее расписание",
            "question_subclass_key": "base_schedule",
        },
    )
    small_base = QuestionClass(
        tenant_id="foton",
        canonical_question="стоимость / базовая стоимость",
        narrow_scope="Вопрос клиента про стоимость; подкласс: базовая стоимость.",
        class_key="theme:001_pricing",
        count_total=12,
        examples_redacted=("Сколько стоит пробное занятие?",),
        required_fact_keys=("price.current",),
        metadata={
            "parent_question_class": "стоимость",
            "question_subclass": "базовая стоимость",
            "question_subclass_key": "base_price",
        },
    )
    thematic_fallback = QuestionClass(
        tenant_id="foton",
        canonical_question="программа курса / общий вопрос по программе",
        narrow_scope="Вопрос клиента про программу курса; подкласс: общий вопрос по программе.",
        class_key="theme:016_program",
        count_total=50,
        examples_redacted=("Расскажите про курс по математике.",),
        required_fact_keys=("program.current",),
        metadata={
            "parent_question_class": "программа курса",
            "question_subclass": "общий вопрос по программе",
            "question_subclass_key": "general_program_question",
        },
    )
    broad_manual = QuestionClass(
        tenant_id="foton",
        canonical_question="общий вопрос / ручной разбор: без уточненного подкласса",
        narrow_scope="Остаточная ручная корзина после дробления.",
        class_key="service:S2_unclear",
        count_total=60,
        examples_redacted=("Или вообще не будут?",),
        metadata={
            "parent_question_class": "общий вопрос",
            "question_subclass": "ручной разбор: без уточненного подкласса",
            "question_subclass_key": "other_manual_review",
        },
    )
    small_manual = QuestionClass(
        tenant_id="foton",
        canonical_question="общий вопрос / короткий обрывок внутри темы",
        narrow_scope="Маленькая остаточная ручная корзина.",
        class_key="service:S2_unclear",
        count_total=8,
        examples_redacted=("Что-то срочно?",),
        metadata={
            "parent_question_class": "общий вопрос",
            "question_subclass": "короткий обрывок внутри темы",
            "question_subclass_key": "other_short_context",
        },
    )
    templates = [
        AnswerTemplate(tenant_id="foton", question_class_id=wide.question_class_id, template_text="Черновик широкого класса"),
        AnswerTemplate(tenant_id="foton", question_class_id=narrow.question_class_id, template_text="Черновик узкого класса"),
        AnswerTemplate(tenant_id="foton", question_class_id=broad_base.question_class_id, template_text="Черновик базовой корзины"),
        AnswerTemplate(tenant_id="foton", question_class_id=small_base.question_class_id, template_text="Черновик малой базовой корзины"),
        AnswerTemplate(tenant_id="foton", question_class_id=thematic_fallback.question_class_id, template_text="Черновик общего тематического класса"),
        AnswerTemplate(tenant_id="foton", question_class_id=broad_manual.question_class_id, template_text="Черновик ручной корзины"),
        AnswerTemplate(tenant_id="foton", question_class_id=small_manual.question_class_id, template_text="Черновик малой ручной корзины"),
    ]

    classes = [wide, narrow, broad_base, small_base, thematic_fallback, broad_manual, small_manual]
    review_rows = build_rop_review_rows(classes, templates)
    priority_rows = build_rop_priority_rows(classes, templates, limit=10)
    drafts = build_approved_answer_drafts(classes, templates)
    audit = build_answer_quality_audit(classes, templates, drafts)

    wide_review = next(row for row in review_rows if row["Класс вопроса"] == "общий вопрос")
    narrow_review = next(row for row in review_rows if row["Класс вопроса"] == "стоимость / что входит в стоимость")
    broad_base_review = next(row for row in review_rows if row["Класс вопроса"] == "расписание / общее расписание")
    small_base_review = next(row for row in review_rows if row["Класс вопроса"] == "стоимость / базовая стоимость")
    thematic_review = next(row for row in review_rows if row["Класс вопроса"] == "программа курса / общий вопрос по программе")
    broad_manual_review = next(row for row in review_rows if row["Класс вопроса"] == "общий вопрос / ручной разбор: без уточненного подкласса")
    small_manual_review = next(row for row in review_rows if row["Класс вопроса"] == "общий вопрос / короткий обрывок внутри темы")
    wide_priority = next(row for row in priority_rows if row["Класс вопроса"] == "общий вопрос")
    broad_base_priority = next(row for row in priority_rows if row["Класс вопроса"] == "расписание / общее расписание")
    thematic_priority = next(row for row in priority_rows if row["Класс вопроса"] == "программа курса / общий вопрос по программе")
    broad_manual_priority = next(row for row in priority_rows if row["Класс вопроса"] == "общий вопрос / ручной разбор: без уточненного подкласса")
    wide_draft = next(row for row in drafts if row.question_class_id == wide.question_class_id)
    broad_base_draft = next(row for row in drafts if row.question_class_id == broad_base.question_class_id)
    thematic_draft = next(row for row in drafts if row.question_class_id == thematic_fallback.question_class_id)
    broad_manual_draft = next(row for row in drafts if row.question_class_id == broad_manual.question_class_id)

    assert wide_review["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert wide_review["Причина блокировки утверждения"] == "wide_class_no_subclass"
    assert wide_review["Примеры подсценариев для дробления"] == ""
    assert wide_priority["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert thematic_review["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert thematic_review["Причина блокировки утверждения"] == "thematic_fallback_needs_split"
    assert thematic_priority["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert broad_base_review["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert broad_base_review["Причина блокировки утверждения"] == "thematic_fallback_needs_split"
    assert broad_base_priority["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert broad_manual_review["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert broad_manual_review["Причина блокировки утверждения"] == "thematic_fallback_needs_split"
    assert broad_manual_priority["Можно ли утверждать сейчас"] == "нет, сначала дробить"
    assert small_manual_review["Можно ли утверждать сейчас"] == "да, проверить черновик"
    assert small_base_review["Можно ли утверждать сейчас"] == "да, проверить черновик"
    assert narrow_review["Можно ли утверждать сейчас"] == "да, проверить черновик"
    assert narrow_review["Причина блокировки утверждения"] == ""
    assert wide_draft.rop_decision == "block_until_split"
    assert broad_base_draft.rop_decision == "block_until_split"
    assert thematic_draft.rop_decision == "block_until_split"
    assert broad_manual_draft.rop_decision == "block_until_split"
    assert audit["verdict"] == "blocked"
    assert any(item["code"] == "wide_class_block_until_split" for item in audit["findings"])
    assert any(item["code"] == "thematic_fallback_needs_split" for item in audit["findings"])
