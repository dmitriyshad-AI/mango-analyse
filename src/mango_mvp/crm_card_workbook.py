from __future__ import annotations

import csv
import json
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Mapping, Sequence

from mango_mvp.crm_card_aggregator import build_crm_card_projection
from mango_mvp.customer_timeline.read_api import CustomerTimelineReadApi, CustomerTimelineReadApiConfig


CRM_CARD_WORKBOOK_SCHEMA_VERSION = "crm_card_workbook_v1"
DEFAULT_SAMPLE_SIZE = 200
MAX_SAMPLE_SIZE = 300
DEFAULT_AMO_BASE_URL = "https://educent.amocrm.ru"

CLIENT_HEADERS = (
    "Имя",
    "Телефон",
    "Бренд",
    "Открыть в AMO",
    "Запрос",
    "Статус сделки",
    "Возражения",
    "Следующий шаг",
    "Последняя сводка",
    "Tallanto",
    "Предупреждения",
    "История общения",
    "Готово",
    "Блокеры",
    "Вердикт",
    "Комментарий",
    "customer_id",
)


@dataclass(frozen=True)
class CrmCardWorkbookConfig:
    timeline_db: Path
    allowed_root: Path
    out_xlsx: Path
    tenant_id: str = "foton"
    sample_size: int = DEFAULT_SAMPLE_SIZE
    manager_facts_csv: Path | None = None
    amo_base_url: str = DEFAULT_AMO_BASE_URL
    generated_at: datetime | None = None


def build_crm_card_workbook(config: CrmCardWorkbookConfig) -> Mapping[str, Any]:
    sample_size = min(max(1, int(config.sample_size)), MAX_SAMPLE_SIZE)
    out_xlsx = config.out_xlsx.expanduser().resolve(strict=False)
    if "stable_runtime" in out_xlsx.parts:
        raise ValueError("CRM card workbook must not be written under stable_runtime")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    preview_csv = out_xlsx.with_suffix(".csv")
    summary_json = out_xlsx.with_suffix(".summary.json")
    facts_by_customer_id = _load_manager_facts(config.manager_facts_csv)

    read_config = CustomerTimelineReadApiConfig(
        timeline_db=config.timeline_db,
        allowed_root=config.allowed_root,
    )
    rows: list[dict[str, Any]] = []
    profiles_seen = 0
    with CustomerTimelineReadApi.open(read_config) as api:
        health = api.health()
        customer_ids = _sample_customer_ids(api, config.tenant_id, sample_size)
        for customer_id in customer_ids:
            profile = api.customer_profile(config.tenant_id, customer_id, event_limit=50, bot_context_limit=1)
            profiles_seen += 1
            manager_facts = _facts_for_profile(profile, facts_by_customer_id)
            projection = build_crm_card_projection(profile, manager_facts=manager_facts)
            rows.append(_workbook_row(profile, projection, amo_base_url=config.amo_base_url))

    generated_at = _stable_generated_at(rows, override=config.generated_at)
    summary_rows = _summary_rows(rows, generated_at=generated_at, health=health, profiles_seen=profiles_seen)
    _write_csv(preview_csv, rows, CLIENT_HEADERS)
    _write_workbook(out_xlsx, rows, summary_rows)
    summary = {
        "schema_version": CRM_CARD_WORKBOOK_SCHEMA_VERSION,
        "generated_at": generated_at,
        "timeline_db": str(config.timeline_db),
        "allowed_root": str(config.allowed_root),
        "tenant_id": config.tenant_id,
        "sample_size_requested": config.sample_size,
        "sample_size_effective": sample_size,
        "rows": len(rows),
        "ready_yes": sum(1 for row in rows if row["Готово"] == "да"),
        "ready_no": sum(1 for row in rows if row["Готово"] != "да"),
        "blocker_counts": dict(_blocker_counts(rows).most_common()),
        "outputs": {
            "xlsx": str(out_xlsx),
            "csv": str(preview_csv),
            "summary_json": str(summary_json),
        },
        "safety": {
            "read_only": True,
            "write_amo": False,
            "write_tallanto": False,
            "write_customer_timeline": False,
            "new_storage_table": False,
            "live_network_calls": False,
        },
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def _stable_generated_at(rows: Sequence[Mapping[str, Any]], *, override: datetime | None = None) -> str:
    if override is not None:
        return override.replace(microsecond=0).isoformat(timespec="seconds")
    timestamps = [str(row.get("Дата") or "").strip() for row in rows if str(row.get("Дата") or "").strip()]
    return max(timestamps) if timestamps else ""


def _sample_customer_ids(api: CustomerTimelineReadApi, tenant_id: str, sample_size: int) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    _append_amo_customers(api, tenant_id, result, seen, limit=max(1, sample_size // 2))
    if len(result) >= sample_size:
        return result[:sample_size]
    buckets = ("ambiguous", "partial", "strong", "unmatched")
    per_bucket = max(1, sample_size // len(buckets))
    for status in buckets:
        _append_customers(api, tenant_id, result, seen, limit=per_bucket, identity_status=status)
        if len(result) >= sample_size:
            return result[:sample_size]
    cursor: str | None = None
    while len(result) < sample_size:
        page = api.list_customers(tenant_id, limit=min(200, sample_size - len(result)), cursor=cursor)
        for item in page.get("items") or ():
            customer_id = str(item.get("customer_id") or "").strip()
            if customer_id and customer_id not in seen:
                seen.add(customer_id)
                result.append(customer_id)
        cursor = str(page.get("next_cursor") or "") or None
        if not cursor:
            break
    return result[:sample_size]


def _append_amo_customers(
    api: CustomerTimelineReadApi,
    tenant_id: str,
    result: list[str],
    seen: set[str],
    *,
    limit: int,
) -> None:
    cursor: str | None = None
    scanned_pages = 0
    start_count = len(result)
    while len(result) - start_count < limit and scanned_pages < 30:
        page = api.list_customers(tenant_id, limit=200, cursor=cursor)
        scanned_pages += 1
        for item in page.get("items") or ():
            summary = item.get("summary") if isinstance(item.get("summary"), Mapping) else {}
            has_amo = int(summary.get("amo_contact_id_count") or 0) > 0 or int(summary.get("amo_lead_id_count") or 0) > 0
            customer_id = str(item.get("customer_id") or "").strip()
            if has_amo and customer_id and customer_id not in seen:
                seen.add(customer_id)
                result.append(customer_id)
                if len(result) >= limit:
                    return
        cursor = str(page.get("next_cursor") or "") or None
        if not cursor:
            return


def _append_customers(
    api: CustomerTimelineReadApi,
    tenant_id: str,
    result: list[str],
    seen: set[str],
    *,
    limit: int,
    identity_status: str,
) -> None:
    page = api.list_customers(tenant_id, identity_status=identity_status, limit=min(limit, 200))
    for item in page.get("items") or ():
        customer_id = str(item.get("customer_id") or "").strip()
        if customer_id and customer_id not in seen:
            seen.add(customer_id)
            result.append(customer_id)


def _load_manager_facts(path: Path | None) -> dict[str, Mapping[str, Any]]:
    if path is None:
        return {}
    resolved = path.expanduser().resolve(strict=False)
    if not resolved.exists():
        raise FileNotFoundError(f"manager facts CSV not found: {resolved}")
    with resolved.open("r", encoding="utf-8-sig", newline="") as fh:
        rows = [dict(row) for row in csv.DictReader(fh)]
    result: dict[str, Mapping[str, Any]] = {}
    for row in rows:
        for key in ("customer_id", "new_customer_id", "old_customer_id"):
            value = str(row.get(key) or "").strip()
            if value:
                result[value] = row
    return result


def _facts_for_profile(profile: Mapping[str, Any], facts_by_customer_id: Mapping[str, Mapping[str, Any]]) -> Mapping[str, Any]:
    customer_id = str(profile.get("customer_id") or "").strip()
    if customer_id in facts_by_customer_id:
        return facts_by_customer_id[customer_id]
    for mapping in profile.get("customer_id_mappings") or ():
        if not isinstance(mapping, Mapping):
            continue
        for key in ("new_customer_id", "old_customer_id"):
            value = str(mapping.get(key) or "").strip()
            if value in facts_by_customer_id:
                return facts_by_customer_id[value]
    return {}


def _workbook_row(profile: Mapping[str, Any], projection: Mapping[str, Any], *, amo_base_url: str) -> dict[str, Any]:
    customer = profile.get("customer") if isinstance(profile.get("customer"), Mapping) else {}
    workbook = projection.get("workbook") if isinstance(projection.get("workbook"), Mapping) else {}
    contact_card = projection.get("contact_card") if isinstance(projection.get("contact_card"), Mapping) else {}
    deal_card = projection.get("deal_card") if isinstance(projection.get("deal_card"), Mapping) else {}
    contact_fields = contact_card.get("fields") if isinstance(contact_card.get("fields"), Mapping) else {}
    deal_fields = deal_card.get("fields") if isinstance(deal_card.get("fields"), Mapping) else {}
    lead_id = str(deal_card.get("selected_amo_lead_id") or "").strip()
    return {
        "Имя": str(customer.get("display_name") or ""),
        "Телефон": str(workbook.get("phone") or customer.get("primary_phone") or ""),
        "Бренд": _brand(customer),
        "Открыть в AMO": _amo_url(lead_id, amo_base_url=amo_base_url),
        "Запрос": str(contact_fields.get("Запрос") or ""),
        "Статус сделки": str(deal_fields.get("Статус сделки") or ""),
        "Возражения": str(deal_fields.get("Возражения") or ""),
        "Следующий шаг": str(deal_fields.get("Следующий шаг") or ""),
        "Последняя сводка": str(contact_fields.get("Последняя сводка") or ""),
        "Tallanto": str(deal_fields.get("Tallanto") or ""),
        "Предупреждения": str(deal_fields.get("Предупреждения") or ""),
        "История общения": str(contact_fields.get("История общения") or ""),
        "Дата": str(projection.get("snapshot_as_of") or ""),
        "Готово": str(workbook.get("ready") or "нет"),
        "Блокеры": str(workbook.get("blockers") or ""),
        "Вердикт": "",
        "Комментарий": "",
        "customer_id": str(projection.get("customer_id") or profile.get("customer_id") or ""),
        "crm_card_contact_payload_json": json.dumps(contact_card.get("fields") or {}, ensure_ascii=False, sort_keys=True),
        "crm_card_deal_payload_json": json.dumps(deal_card.get("fields") or {}, ensure_ascii=False, sort_keys=True),
    }


def _brand(customer: Mapping[str, Any]) -> str:
    summary = customer.get("summary") if isinstance(customer.get("summary"), Mapping) else {}
    brands = summary.get("brands")
    if isinstance(brands, Sequence) and not isinstance(brands, (str, bytes, bytearray)):
        return " | ".join(str(item) for item in brands if str(item).strip())
    return str(summary.get("brand") or "")


def _amo_url(lead_id: str, *, amo_base_url: str) -> str:
    if not lead_id.isdigit():
        return ""
    return f"{amo_base_url.rstrip('/')}/leads/detail/{lead_id}"


def _summary_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    generated_at: str,
    health: Mapping[str, Any],
    profiles_seen: int,
) -> list[dict[str, Any]]:
    ready_counts = Counter(str(row.get("Готово") or "нет") for row in rows)
    blocker_counts = _blocker_counts(rows)
    summary: list[dict[str, Any]] = [
        {"Метрика": "generated_at", "Значение": generated_at},
        {"Метрика": "read_api_status", "Значение": str(health.get("status"))},
        {"Метрика": "read_only", "Значение": str(health.get("read_only"))},
        {"Метрика": "rows", "Значение": len(rows)},
        {"Метрика": "profiles_seen", "Значение": profiles_seen},
        {"Метрика": "ready_yes", "Значение": ready_counts.get("да", 0)},
        {"Метрика": "ready_no", "Значение": len(rows) - ready_counts.get("да", 0)},
    ]
    for blocker, count in blocker_counts.most_common():
        summary.append({"Метрика": f"blocker:{blocker}", "Значение": count})
    return summary


def _blocker_counts(rows: Sequence[Mapping[str, Any]]) -> Counter[str]:
    counts: Counter[str] = Counter()
    for row in rows:
        blockers = str(row.get("Блокеры") or "")
        for blocker in [item.strip() for item in blockers.split("|") if item.strip()]:
            counts[blocker] += 1
    return counts


def _write_csv(path: Path, rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(headers) + ["crm_card_contact_payload_json", "crm_card_deal_payload_json"])
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in writer.fieldnames or ()})


def _write_workbook(path: Path, rows: Sequence[Mapping[str, Any]], summary_rows: Sequence[Mapping[str, Any]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - dependency is in project requirements.
        raise RuntimeError("openpyxl is required to build CRM card workbook") from exc

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    ws_clients = wb.create_sheet("Клиенты")
    _fill_sheet(ws_summary, ("Метрика", "Значение"), summary_rows)
    _fill_sheet(ws_clients, CLIENT_HEADERS, rows)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for ws in (ws_summary, ws_clients):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column in range(1, ws.max_column + 1):
            letter = get_column_letter(column)
            header = str(ws.cell(row=1, column=column).value or "")
            ws.column_dimensions[letter].width = min(max(len(header) + 4, 14), 48)
    for letter in ("E", "F", "G", "H", "I", "J", "K"):
        ws_clients.column_dimensions[letter].width = 36
    ws_clients.column_dimensions["L"].width = 70
    ws_clients.column_dimensions["N"].width = 42
    wb.save(path)


def _fill_sheet(ws: Any, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([row.get(header, "") for header in headers])


__all__ = [
    "CLIENT_HEADERS",
    "CRM_CARD_WORKBOOK_SCHEMA_VERSION",
    "CrmCardWorkbookConfig",
    "build_crm_card_workbook",
]
