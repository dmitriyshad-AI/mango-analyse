from __future__ import annotations

import csv
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook

from mango_mvp.question_catalog.rop_policy_import import (
    ROP_DECISION_DRAFT_FOR_MANAGER,
    ROP_DECISION_FACT_CHECK,
    apply_rop_policies_to_taxonomy,
    build_final_v2_rows,
    load_approved_rop_policies,
    write_final_v2_csv,
    write_final_v2_xlsx,
)


def test_load_approved_rop_policies_maps_theme_numbers(tmp_path: Path) -> None:
    workbook_path = _write_rop_workbook(tmp_path)

    policies = load_approved_rop_policies(workbook_path)

    assert policies["theme:001_pricing"].rop_decision == ROP_DECISION_FACT_CHECK
    assert policies["theme:019a_positive_feedback"].rop_approved_phrasing == "Спасибо за обратную связь."


def test_apply_rop_policies_updates_theme_fields_and_permission(tmp_path: Path) -> None:
    policies = load_approved_rop_policies(_write_rop_workbook(tmp_path))
    taxonomy = {
        "schema_version": "test",
        "source_documents": {},
        "themes": [
            {
                "theme_id": "theme:001_pricing",
                "theme_name": "Стоимость обучения",
                "business_block": "Финансы",
                "default_bot_permission": "manager_only",
                "required_facts": ["prices.current"],
                "forbidden_promises": [],
                "escalation_rule": "Менеджер",
            },
            {
                "theme_id": "theme:019a_positive_feedback",
                "theme_name": "Положительная обратная связь / благодарность",
                "business_block": "Сервис",
                "default_bot_permission": "draft_for_manager",
                "required_facts": [],
                "forbidden_promises": [],
                "escalation_rule": "Менеджер",
            },
        ],
    }

    updated = apply_rop_policies_to_taxonomy(taxonomy, policies, require_all_themes=False)

    pricing = updated["themes"][0]
    assert pricing["rop_approved_phrasing"] == "Стоимость: {цена}."
    assert pricing["rop_mandatory_data"] == "Продукт, класс, формат."
    assert pricing["default_bot_permission"] == "answer_after_fact_check"
    positive = updated["themes"][1]
    assert positive["default_bot_permission"] == "bot_self"
    assert "allowed_rop_decisions" in updated


def test_final_v2_export_contains_one_row_per_theme(tmp_path: Path) -> None:
    taxonomy = {
        "themes": [
            {
                "theme_id": "theme:001_pricing",
                "theme_name": "Стоимость обучения",
                "business_block": "Финансы",
                "default_bot_permission": "answer_after_fact_check",
                "required_facts": ["prices.current"],
                "forbidden_promises": ["Не обещать старую цену"],
                "escalation_rule": "Менеджер",
                "rop_decision": ROP_DECISION_FACT_CHECK,
                "rop_approved_phrasing": "Стоимость: {цена}.",
                "rop_mandatory_data": "Продукт, класс, формат.",
                "rop_forbids_extra": "Не называть цену без прайса.",
                "rop_handoff_target": "Менеджер группы.",
                "rop_comment": "",
            }
        ]
    }
    rows = build_final_v2_rows(taxonomy)
    csv_path = tmp_path / "final.csv"
    xlsx_path = tmp_path / "final.xlsx"

    write_final_v2_csv(csv_path, rows)
    write_final_v2_xlsx(xlsx_path, rows)

    with csv_path.open(encoding="utf-8-sig", newline="") as file:
        loaded = list(csv.DictReader(file))
    assert loaded[0]["theme_id"] == "theme:001_pricing"
    workbook = load_workbook(xlsx_path, read_only=True)
    assert workbook["Question Catalog v2"].max_row == 2


def test_current_taxonomy_has_rop_fields_after_import() -> None:
    taxonomy = yaml.safe_load(Path("src/mango_mvp/question_catalog/themes_taxonomy.yaml").read_text(encoding="utf-8"))
    themes = taxonomy["themes"]

    assert len(themes) == 32
    assert all(theme.get("rop_decision") for theme in themes)
    by_id = {theme["theme_id"]: theme for theme in themes}
    assert by_id["theme:008_tax_deduction"]["default_bot_permission"] == "draft_for_manager"


def _write_rop_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "rop.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Опросник РОПа"
    sheet.append(
        [
            "№",
            "Тема",
            "✏️ Решение РОПа",
            "✏️ Утверждённая фраза для бота",
            "✏️ Что бот обязан спросить/проверить",
            "✏️ Что нельзя обещать (помимо системных)",
            "✏️ Кому передавать, если бот не сам",
            "✏️ Комментарий",
        ]
    )
    sheet.append(
        [
            "01",
            "Стоимость обучения",
            ROP_DECISION_FACT_CHECK,
            "Стоимость: {цена}.",
            "Продукт, класс, формат.",
            "Не называть цену без прайса.",
            "Менеджер группы.",
            "",
        ]
    )
    sheet.append(
        [
            "19a",
            "Положительная обратная связь / благодарность",
            "Бот отвечает сам",
            "Спасибо за обратную связь.",
            "",
            "Не обещать подарки.",
            "Менеджер группы.",
            "",
        ]
    )
    sheet.append(
        [
            "08",
            "Налоговый вычет",
            ROP_DECISION_DRAFT_FOR_MANAGER,
            "Передам менеджеру.",
            "ФИО, год оплаты.",
            "Не консультировать по налогам.",
            "Менеджер группы.",
            "",
        ]
    )
    workbook.save(path)
    return path
