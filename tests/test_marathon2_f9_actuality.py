from __future__ import annotations

import importlib.util
import json
import sqlite3
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook


def _load_script(name: str):
    path = Path(__file__).resolve().parents[1] / "scripts" / f"{name}.py"
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def test_reconcile_stalled_deals_checks_open_deals_by_customer(tmp_path: Path) -> None:
    mod = _load_script("reconcile_stalled_deals_with_amo")
    db = _f9_db(tmp_path)
    client = FakeAmoClient({"lead-1": {"id": 1, "status_id": 142, "pipeline_id": 10, "closed_at": 123}})

    report = mod.reconcile(
        timeline_db=db,
        limit=10,
        mcp_env=tmp_path / "missing.env",
        page_limit=50,
        sleep_sec=0,
        client=client,
    )

    assert report["status"] == "checked"
    assert report["customers_selected"] == 1
    assert report["open_opportunities_checked"] == 1
    assert report["customers_changed"] == 1
    assert report["snapshot_stale"] is True
    assert report["rows"][0]["reason"] == "live_lead_closed"
    assert report["rows"][0]["lead_hash"] != "lead-1"
    assert report["source_open_mode"] == "sqlite_mode_ro"
    assert client.calls == 1


def test_reconcile_readonly_connection_sees_uncheckpointed_wal(tmp_path: Path) -> None:
    mod = _load_script("reconcile_stalled_deals_with_amo")
    db = tmp_path / "timeline.sqlite"
    writer = sqlite3.connect(db)
    try:
        assert writer.execute("PRAGMA journal_mode=WAL").fetchone()[0] == "wal"
        writer.execute("CREATE TABLE marker (value INTEGER)")
        writer.execute("INSERT INTO marker VALUES (1)")
        writer.commit()
        with mod.connect_ro(db) as reader:
            assert reader.execute("SELECT value FROM marker").fetchone()[0] == 1
    finally:
        writer.close()


def test_build_manager_dossier_cli_enforces_outreach_eligibility(tmp_path: Path, monkeypatch) -> None:
    mod = _load_script("build_manager_dossier")
    captured = {}

    def fake_build(**kwargs):
        captured.update(kwargs)
        return {"customers": 0}

    monkeypatch.setattr(mod, "build_manager_dossier_workbook", fake_build)
    assert mod.main(
        [
            "--timeline-db",
            str(tmp_path / "timeline.sqlite"),
            "--out-xlsx",
            str(tmp_path / ".codex_local" / "dossier.xlsx"),
        ]
    ) == 0
    assert captured["enforce_outreach_eligibility"] is True


def test_reconcile_stalled_deals_is_unavailable_without_env(tmp_path: Path) -> None:
    mod = _load_script("reconcile_stalled_deals_with_amo")
    db = _f9_db(tmp_path)

    report = mod.reconcile(
        timeline_db=db,
        limit=10,
        mcp_env=tmp_path / "missing.env",
        page_limit=50,
        sleep_sec=0,
    )

    assert report["status"] == "unavailable"
    assert report["reason"] == "mcp_env_missing"
    assert report["customers_selected"] == 1
    assert report["customers_checked"] == 0


def test_reconcile_stalled_deals_is_tenant_scoped(tmp_path: Path) -> None:
    mod = _load_script("reconcile_stalled_deals_with_amo")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT INTO customer_identities VALUES (?,?,?,?,?,?,?)",
            ("other", "customer:other", "strong", "Other", None, None, "{}"),
        )
        con.execute(
            "INSERT INTO derived_signals VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "signal:other", "other", "customer:other", None, None, "deal_stalling", "critical",
                "active", None, 1.0, 1, "2026-07-01T00:00:00+00:00", "h-other", "{}",
            ),
        )
        con.execute(
            "INSERT INTO customer_opportunities VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "opp:other", "other", "customer:other", "amo_deal", "amocrm_snapshot", "lead-other",
                "Other", "12345", "2026-01-01", "", 0.8, "h-other", "{}",
            ),
        )
        con.row_factory = sqlite3.Row
        customer_ids = mod.selected_stalled_customers(con, limit=10)
        opportunities = mod.open_opportunities_for_customers(con, customer_ids)

    assert customer_ids == ["customer:1"]
    assert {item.customer_id for item in opportunities} == {"customer:1"}


def test_wave0_manager_lists_write_actuality_header(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    reconcile = tmp_path / ".codex_local" / "reconcile.json"
    reconcile.parent.mkdir(parents=True)
    reconcile.write_text(
        json.dumps(
            {
                "status": "checked",
                "generated_at": "2026-07-03T12:00:00+00:00",
                "customers_checked": 10,
                "customers_changed": 2,
                "snapshot_stale": True,
                "rows": [
                    {
                        "customer_hash": mod._mask_id("customer:1"),
                        "lead_hash": mod._mask_id("lead-1"),
                        "changed": False,
                        "live_found": True,
                        "reason": "unchanged_or_snapshot_status_name",
                    }
                ],
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    out = tmp_path / ".codex_local" / "wave0.xlsx"

    summary = mod.build_workbook(
        timeline_db=db,
        out_xlsx=out,
        allowed_root=tmp_path,
        reconcile_json=reconcile,
        limit_per_sheet=50,
        enforce_freshness=False,
    )

    assert summary["quick_check"] == "ok"
    assert summary["rows_by_sheet"]["Зависшие факт LTV"] == 1
    wb = load_workbook(out, read_only=True)
    ws = wb["Зависшие факт LTV"]
    assert "2 расхождений из 10" in ws["A2"].value
    assert "cursor_at по источникам: amocrm_snapshot=2026-07-03T04:13:25+00:00" in ws["A2"].value
    assert "imported_at отдельно: amocrm_snapshot=2026-07-03T05:00:00+00:00" in ws["A2"].value
    assert "max event_at отдельно: amocrm_snapshot=2026-07-03T04:18:25+00:00" in ws["A2"].value
    headers = [cell.value for cell in ws[4]]
    first = dict(zip(headers, [cell.value for cell in ws[5]]))
    assert first["action_today"] is False
    assert first["action_due_at"] is None


def test_wave0_prioritizes_all_returned_customers_before_deals_and_seasonal(
    tmp_path: Path, monkeypatch
) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    returned = [{"customer_id": f"returned:{index}", "signal_type": "client_returned"} for index in range(5)]
    deals = [{"customer_id": "returned:0", "signal_type": "deal_stalling"}] + [
        {"customer_id": f"deal:{index}", "signal_type": "deal_stalling"} for index in range(49)
    ]
    seasonal = [{"customer_id": f"seasonal:{index}", "signal_type": "season_return_candidate"} for index in range(5)]

    def fake_rows(_con, signal_types, **_kwargs):
        if "client_returned" in signal_types:
            return returned
        if "deal_stalling" in signal_types:
            return deals
        return seasonal

    monkeypatch.setattr(mod, "build_rows", fake_rows)
    monkeypatch.setattr(mod, "_reconcile_deal_rows", lambda rows, _reconcile: (rows, {}))
    summary = mod.build_workbook(
        timeline_db=db,
        out_xlsx=tmp_path / ".codex_local" / "wave0_priority.xlsx",
        allowed_root=tmp_path,
        reconcile_json=None,
        limit_per_sheet=50,
        enforce_freshness=False,
    )

    assert summary["selected_customer_ids"][:5] == [row["customer_id"] for row in returned]
    assert set(summary["selected_customer_ids"][:50]) >= {row["customer_id"] for row in returned}
    assert summary["duplicate_signal_suppressed"] == {"deal_stalling": 1}


def test_wave0_seasonal_candidate_excludes_active_access(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT INTO timeline_events VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "event:active-access", "access", "foton", "customer:1", None, "tallanto_abonement",
                "2026-07-20T00:00:00+00:00", "tallanto_snapshot", "access-1", "", "system",
                "strong", 1.0, 1, "", "", "Активный абонемент", "2026-07-20T00:00:00+00:00",
                "ha", json.dumps({"record": {"visits_left": 5}}), "", None,
            ),
        )
        con.row_factory = sqlite3.Row
        rows = mod.build_rows(con, ("season_return_candidate",), limit=50, tenant_id="foton")

    assert rows == []


def test_wave0_excludes_expired_and_other_tenant_signals(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.execute("UPDATE derived_signals SET expires_at='2020-01-01T00:00:00+00:00'")
        con.execute(
            "INSERT INTO customer_identities VALUES (?,?,?,?,?,?,?)",
            ("other", "customer:other", "strong", "Other", None, None, "{}"),
        )
        con.execute(
            "INSERT INTO derived_signals VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "signal:other", "other", "customer:other", None, None, "deal_stalling", "critical",
                "active", None, 1.0, 1, "2026-07-01T00:00:00+00:00", "h-other", "{}",
            ),
        )
        con.row_factory = sqlite3.Row
        rows = mod.build_rows(con, ("deal_stalling",), limit=50, tenant_id="foton")

    assert rows == []


def test_wave0_excludes_customer_when_any_live_amo_read_failed(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    rows = [{"customer_id": "customer:1", "open_amo_lead_ids": "lead-1,lead-2"}]
    reconcile = {
        "status": "checked",
        "rows": [
            {
                "customer_hash": mod._mask_id("customer:1"),
                "lead_hash": mod._mask_id("lead-1"),
                "changed": False,
                "live_found": True,
                "reason": "unchanged_or_snapshot_status_name",
            },
            {
                "customer_hash": mod._mask_id("customer:1"),
                "lead_hash": mod._mask_id("lead-2"),
                "changed": False,
                "live_found": None,
                "reason": "live_read_error:TimeoutError",
            },
        ],
    }

    kept, excluded = mod._reconcile_deal_rows(rows, reconcile)

    assert kept == []
    assert excluded == {"live_read_error:TimeoutError": 1}


def test_wave0_reconcile_keeps_rows_without_open_amo_and_blocks_unchecked_open_deal() -> None:
    mod = _load_script("build_wave0_manager_lists")
    rows = [
        {"customer_id": "customer:safe", "open_amo_lead_ids": ""},
        {"customer_id": "customer:open", "open_amo_lead_ids": "lead-1"},
    ]

    kept, excluded = mod._reconcile_deal_rows(rows, {"status": "checked", "rows": []})

    assert kept == [
        {
            "customer_id": "customer:safe",
            "open_amo_lead_ids": "",
            "row_reconcile_status": "not_needed",
            "reconciled_at": None,
        }
    ]
    assert excluded == {"not_reconciled": 1}


def test_source_freshness_does_not_treat_future_event_as_fresh_capture(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        con.execute("UPDATE timeline_events SET event_at='2026-12-29T09:00:00+00:00'")
        con.execute(
            "INSERT INTO ingestion_runs VALUES (?,?,?,?,?)",
            ("foton", "amocrm_snapshot", "timeline_import", "failed", "2026-07-09T06:00:00+00:00"),
        )
        con.execute(
            "UPDATE ingestion_cursors SET updated_at='2026-07-09T06:05:00+00:00'"
        )
        row = mod.source_freshness(con)[0]

    assert row["cursor_at"] == "2026-07-03T04:13:25+00:00"
    assert row["cursor_updated_at"] == "2026-07-09T06:05:00+00:00"
    assert row["imported_at"] == "2026-07-03T05:00:00+00:00"
    assert row["max_event_at"] == "2026-12-29T09:00:00+00:00"


def test_source_freshness_ignores_non_import_processing_runs(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        con.execute(
            "INSERT INTO ingestion_runs VALUES (?,?,?,?,?)",
            ("foton", "amocrm_snapshot", "mail_link_enrich", "completed", "2026-07-09T06:00:00+00:00"),
        )
        row = mod.source_freshness(con)[0]

    assert row["imported_at"] == "2026-07-03T05:00:00+00:00"


def test_source_freshness_lists_missing_required_sources(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        rows = {row["source_system"]: row for row in mod.source_freshness(con)}

    assert rows["amocrm_snapshot"]["missing"] is False
    assert rows["wappi_max"]["expected"] is True
    assert rows["wappi_max"]["missing"] is True


def test_wave0_evidence_date_comes_from_signal_event(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.execute("UPDATE derived_signals SET event_id='event:1' WHERE signal_id='signal:1'")
        con.execute(
            "INSERT INTO timeline_events VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "event:newer", "newer", "foton", "customer:1", None, "message",
                "2026-07-10T00:00:00+00:00", "mail_archive_stage2", "newer", "", "inbound",
                "strong", 1.0, 1, "", "", "newer", "2026-07-10T00:00:00+00:00",
                "h2", "{}", "", None,
            ),
        )
        con.row_factory = sqlite3.Row
        row = mod.build_rows(con, ("deal_stalling",), limit=1)[0]

    assert row["latest_event_at"] == "2026-07-10T00:00:00+00:00"
    assert row["evidence_event_at"] == "2026-07-03T04:18:25+00:00"


def test_wave0_season_evidence_date_comes_from_confirmed_purchase(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        row = mod.build_rows(con, ("season_return_candidate",), limit=1)[0]

    assert row["evidence_event_at"] == "2026-01-01T00:00:00+00:00"
    assert row["evidence_ref_masked"]


def test_wave0_manager_lists_pii_output_must_stay_under_codex_local(tmp_path: Path) -> None:
    mod = _load_script("build_wave0_manager_lists")
    db = _f9_db(tmp_path)

    with pytest.raises(ValueError, match=".codex_local"):
        mod.build_workbook(
            timeline_db=db,
            out_xlsx=tmp_path / "docs" / "wave0.xlsx",
            allowed_root=tmp_path,
            reconcile_json=None,
            limit_per_sheet=50,
        )


def test_refresh_manager_views_is_off_by_default(tmp_path: Path) -> None:
    mod = _load_script("refresh_manager_views")

    report = mod.refresh_manager_views(
        timeline_db=tmp_path / "missing.sqlite",
        allowed_root=tmp_path,
        out_dir=tmp_path / ".codex_local" / "review" / "manager_views",
        reconcile_json=None,
        customer_ids_file=None,
        canonical_calls_db=None,
        limit=10,
        run=False,
    )

    assert report["status"] == "skipped_off_by_default"
    assert report["writes"] == []


def test_refresh_manager_views_run_calls_both_builders(tmp_path: Path, monkeypatch) -> None:
    mod = _load_script("refresh_manager_views")
    calls = []

    def fake_wave0(**kwargs):
        calls.append(("wave0", kwargs))
        return {
            "out_xlsx": str(kwargs["out_xlsx"]),
            "rows_by_sheet": {},
            "selected_customer_ids": ["customer:1", "customer:2"],
        }

    def fake_dossier(**kwargs):
        calls.append(("dossier", kwargs))
        return {"out_xlsx": str(kwargs["out_xlsx"]), "customers": 0}

    monkeypatch.setattr(mod, "build_wave0_workbook", fake_wave0)
    monkeypatch.setattr(mod, "build_manager_dossier_workbook", fake_dossier)

    report = mod.refresh_manager_views(
        timeline_db=tmp_path / "timeline.sqlite",
        allowed_root=tmp_path,
        out_dir=tmp_path / ".codex_local" / "review" / "manager_views",
        reconcile_json=None,
        customer_ids_file=None,
        canonical_calls_db=None,
        limit=10,
        run=True,
    )

    assert report["status"] == "built"
    assert [item[0] for item in calls] == ["wave0", "dossier"]
    assert calls[0][1]["limit_per_sheet"] == 10
    assert calls[1][1]["customer_ids"] == ("customer:1", "customer:2")
    assert calls[1][1]["enforce_outreach_eligibility"] is True
    assert ".codex_local" in report["wave0"]["out_xlsx"]
    assert ".codex_local" in report["dossier"]["out_xlsx"]
    assert Path(report["manifest_path"]).is_file()
    manifest = json.loads(Path(report["manifest_path"]).read_text(encoding="utf-8"))
    assert manifest["provenance"]["git_head"]
    assert manifest["provenance"]["timeline_db_sha256"] is None
    assert manifest["artifact_sha256"]["wave0_xlsx"] is None
    assert manifest["artifact_sha256"]["wave1_xlsx"] is None
    assert manifest["safety_gate_counts"]["durable_opt_out"] == 0


def test_refresh_manager_views_empty_wave0_does_not_widen_wave1(tmp_path: Path, monkeypatch) -> None:
    mod = _load_script("refresh_manager_views")
    captured = {}
    monkeypatch.setattr(
        mod,
        "build_wave0_workbook",
        lambda **kwargs: {"out_xlsx": str(kwargs["out_xlsx"]), "selected_customer_ids": []},
    )

    def fake_dossier(**kwargs):
        captured.update(kwargs)
        return {"out_xlsx": str(kwargs["out_xlsx"]), "customers": 0}

    monkeypatch.setattr(mod, "build_manager_dossier_workbook", fake_dossier)
    mod.refresh_manager_views(
        timeline_db=tmp_path / "timeline.sqlite",
        allowed_root=tmp_path,
        out_dir=tmp_path / ".codex_local" / "review" / "manager_views",
        reconcile_json=None,
        customer_ids_file=None,
        canonical_calls_db=None,
        limit=10,
        run=True,
    )

    assert captured["customer_ids"] == ()
    assert captured["enforce_outreach_eligibility"] is True


def test_refresh_manager_views_run_rejects_non_codex_local_out_dir_before_mkdir(tmp_path: Path) -> None:
    mod = _load_script("refresh_manager_views")
    unsafe = tmp_path / "docs" / "manager_views"

    with pytest.raises(ValueError, match=".codex_local"):
        mod.refresh_manager_views(
            timeline_db=tmp_path / "timeline.sqlite",
            allowed_root=tmp_path,
            out_dir=unsafe,
            reconcile_json=None,
            customer_ids_file=None,
            canonical_calls_db=None,
            limit=10,
            run=True,
        )

    assert not (tmp_path / "docs").exists()


class FakeAmoClient:
    def __init__(self, leads: dict[str, dict]):
        self.leads = leads
        self.calls = 0

    def amo_api_get(self, *, path, params=None, limit=50):
        self.calls += 1
        lead_id = str((params or {}).get("filter[id]") or "")
        lead = self.leads.get(lead_id)
        return {"_embedded": {"leads": [lead] if lead else []}}


def _f9_db(tmp_path: Path) -> Path:
    db = tmp_path / "timeline.sqlite"
    with sqlite3.connect(db) as con:
        con.executescript(
            """
            CREATE TABLE derived_signals (
              signal_id TEXT,
              tenant_id TEXT,
              customer_id TEXT,
              opportunity_id TEXT,
              event_id TEXT,
              signal_type TEXT,
              severity TEXT,
              status TEXT,
              expires_at TEXT,
              confidence REAL,
              requires_manager_review INTEGER,
              created_at TEXT,
              record_hash TEXT,
              record_json TEXT
            );
            CREATE TABLE customer_purchases_v1 (
              tenant_id TEXT,
              customer_id TEXT,
              period TEXT,
              money_kind TEXT,
              total_in REAL,
              total_out REAL,
              deals_cnt INTEGER,
              last_purchase_at TEXT,
              sources_json TEXT,
              computability TEXT,
              code_version TEXT
            );
            CREATE TABLE customer_identities (
              tenant_id TEXT,
              customer_id TEXT,
              identity_status TEXT,
              display_name TEXT,
              primary_phone TEXT,
              primary_email TEXT,
              record_json TEXT
            );
            CREATE TABLE customer_opportunities (
              opportunity_id TEXT,
              tenant_id TEXT,
              customer_id TEXT,
              opportunity_type TEXT,
              source_system TEXT,
              source_id TEXT,
              title TEXT,
              status TEXT,
              opened_at TEXT,
              closed_at TEXT,
              confidence REAL,
              record_hash TEXT,
              record_json TEXT
            );
            CREATE TABLE timeline_events (
              event_id TEXT,
              dedupe_key TEXT,
              tenant_id TEXT,
              customer_id TEXT,
              opportunity_id TEXT,
              event_type TEXT,
              event_at TEXT,
              source_system TEXT,
              source_id TEXT,
              source_ref TEXT,
              direction TEXT,
              match_status TEXT,
              confidence REAL,
              importance INTEGER,
              subject TEXT,
              text_preview TEXT,
              summary TEXT,
              created_at TEXT,
              record_hash TEXT,
              record_json TEXT,
              content_key TEXT,
              superseded_by TEXT
            );
            CREATE TABLE ingestion_cursors (
              tenant_id TEXT,
              source_system TEXT,
              last_cursor_ts TEXT,
              updated_at TEXT,
              metadata_json TEXT
            );
            CREATE TABLE ingestion_runs (
              tenant_id TEXT,
              source_system TEXT,
              run_kind TEXT,
              status TEXT,
              finished_at TEXT
            );
            CREATE TABLE timeline_conflicts (
              tenant_id TEXT,
              status TEXT,
              record_json TEXT
            );
            CREATE TABLE family_links_v1 (
              tenant_id TEXT,
              customer_id TEXT,
              status TEXT,
              confidence TEXT
            );
            """
        )
        con.execute(
            "INSERT INTO ingestion_runs VALUES (?,?,?,?,?)",
            ("foton", "amocrm_snapshot", "timeline_import", "completed", "2026-07-03T05:00:00+00:00"),
        )
        con.executemany(
            "INSERT INTO ingestion_cursors VALUES (?,?,?,?,?)",
            [
                ("foton", "amo_leads_updated_at", "2026-07-03T04:13:25+00:00", "2026-07-03T05:00:00+00:00", "{}"),
                ("foton", "amo_contacts_updated_at", "2026-07-03T04:15:00+00:00", "2026-07-03T05:00:01+00:00", "{}"),
            ],
        )
        con.execute(
            "INSERT INTO customer_identities VALUES (?,?,?,?,?,?,?)",
            (
                "foton",
                "customer:1",
                "strong",
                "Клиент",
                "+70000000000",
                "client@example.com",
                json.dumps({"metadata": {"brands": ["foton"]}}),
            ),
        )
        con.execute(
            "INSERT INTO customer_purchases_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            ("foton", "customer:1", "all_time", "fact", 100000, 0, 1, "2026-01-01", "[]", "computed", "v1"),
        )
        con.execute(
            "INSERT INTO derived_signals VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "signal:1",
                "foton",
                "customer:1",
                None,
                "event:1",
                "deal_stalling",
                "medium",
                "active",
                None,
                0.8,
                1,
                "2026-07-01T00:00:00+00:00",
                "h",
                json.dumps(
                    {
                        "event_id": "event:1",
                        "recommended_action": "Проверить сделку",
                        "evidence_text": "Пауза 14+ дней",
                    },
                    ensure_ascii=False,
                ),
            ),
        )
        con.execute(
            "INSERT INTO derived_signals VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "signal:2",
                "foton",
                "customer:1",
                None,
                None,
                "season_return_candidate",
                "medium",
                "active",
                None,
                0.8,
                1,
                "2026-07-01T00:00:00+00:00",
                "h",
                json.dumps(
                    {
                        "recommended_action": "Вернуться к клиенту",
                        "evidence_text": "Сезон",
                        "metadata": {"last_purchase_at": "2026-01-01T00:00:00+00:00"},
                    },
                    ensure_ascii=False,
                ),
            ),
        )
        con.execute(
            "INSERT INTO customer_opportunities VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            ("opp:1", "foton", "customer:1", "amo_deal", "amocrm_snapshot", "lead-1", "Сделка", "12345", "2026-01-01", "", 0.8, "h", "{}"),
        )
        con.execute(
            "INSERT INTO customer_opportunities VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            ("opp:2", "foton", "customer:1", "amo_deal", "amocrm_snapshot", "lead-closed", "Закрытая", "143", "2026-01-01", "", 0.8, "h", "{}"),
        )
        con.execute(
            "INSERT INTO timeline_events VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "event:1",
                "d",
                "foton",
                "customer:1",
                None,
                "amo_snapshot",
                "2026-07-03T04:18:25+00:00",
                "amocrm_snapshot",
                "1",
                "",
                "system",
                "strong",
                1.0,
                1,
                "",
                "",
                "summary",
                "2026-07-03T04:18:25+00:00",
                "h",
                "{}",
                "",
                None,
            ),
        )
    return db
