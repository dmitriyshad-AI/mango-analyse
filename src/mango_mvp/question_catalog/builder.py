from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from dataclasses import dataclass, replace
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from mango_mvp.question_catalog.contracts import (
    ANSWER_STATUS_DRAFT_NEEDS_REVIEW,
    ANSWER_STATUS_MANAGER_ONLY,
    ANSWER_STATUS_NEEDS_ROP_ANSWER,
    ANSWER_STATUS_NOT_CUSTOMER_QUESTION,
    ANSWER_STATUS_TEMPLATE_NEEDS_CURRENT_FACT,
    BOT_PERMISSION_ALLOWED_AFTER_FACT_CHECK,
    BOT_PERMISSION_DRAFT_ONLY,
    BOT_PERMISSION_MANAGER_ONLY,
    BOT_PERMISSION_NOT_ALLOWED,
    ApprovedQuestionAnswerDraft,
    CurrentFactSource,
    QuestionClass,
    QuestionItem,
    AnswerTemplate,
    normalize_key,
    question_catalog_contract_inventory,
    question_catalog_safety_contract,
    stable_answer_template_id,
)
from mango_mvp.question_catalog.extractors import (
    SINCE_DEFAULT,
    extract_call_questions,
    extract_mail_questions,
    extract_telegram_questions,
)
from mango_mvp.question_catalog.normalization import infer_question_metadata
from mango_mvp.question_catalog.safety import (
    assert_public_text_safe,
    guard_question_catalog_output_path,
)


CATALOG_SCHEMA_VERSION = "customer_question_catalog_v1"


@dataclass(frozen=True)
class CatalogBuildConfig:
    project_root: Path
    out_root: Path
    tenant_id: str = "foton"
    since: datetime = SINCE_DEFAULT
    calls_enriched_reviews: Path | None = None
    telegram_messages_jsonl: Path | None = None
    mail_archive_root: Path | None = None
    fact_source_roots: Sequence[Path] = ()


def default_config(project_root: Path, out_root: Path | None = None) -> CatalogBuildConfig:
    return CatalogBuildConfig(
        project_root=project_root,
        out_root=out_root or project_root / "product_data" / "question_catalog",
        calls_enriched_reviews=project_root
        / "stable_runtime"
        / "sales_insight_knowledge_base_after_quality_backfill_20260510_v11_frozen_gate"
        / "enriched_reviews.csv",
        telegram_messages_jsonl=project_root / "telegram_exports (2)" / "local_vm_2024-04-01" / "messages.jsonl",
        mail_archive_root=project_root / "_external_handoffs" / "mail_archive_2026-05-12",
        fact_source_roots=(
            project_root / ".codex_local" / "kc_source_extract_20260513" / "texts",
            project_root / "docs",
        ),
    )


def build_customer_question_catalog(config: CatalogBuildConfig) -> dict[str, Any]:
    out_root = guard_question_catalog_output_path(config.out_root, project_root=config.project_root)
    out_root.mkdir(parents=True, exist_ok=True)

    source_reports: list[Mapping[str, Any]] = []
    items: list[QuestionItem] = []

    call_items, call_report = extract_call_questions(
        config.calls_enriched_reviews or Path("__missing_calls__"),
        tenant_id=config.tenant_id,
        since=config.since,
    )
    items.extend(call_items)
    source_reports.append(call_report)

    telegram_items, telegram_report = extract_telegram_questions(
        config.telegram_messages_jsonl or Path("__missing_telegram__"),
        tenant_id=config.tenant_id,
        since=config.since,
    )
    items.extend(telegram_items)
    source_reports.append(telegram_report)

    mail_items, mail_report = extract_mail_questions(
        config.mail_archive_root or Path("__missing_mail__"),
        tenant_id=config.tenant_id,
        since=config.since,
    )
    items.extend(mail_items)
    source_reports.append(mail_report)

    items = _dedupe_items(items)
    classes = build_question_classes(items, tenant_id=config.tenant_id)
    templates = build_answer_templates(classes, tenant_id=config.tenant_id)
    classes = attach_template_ids(classes, templates)
    fact_sources = discover_fact_sources(config.fact_source_roots)
    summary = build_summary(
        config=config,
        items=items,
        classes=classes,
        templates=templates,
        fact_sources=fact_sources,
        source_reports=source_reports,
    )

    outputs = write_outputs(
        out_root,
        items=items,
        classes=classes,
        templates=templates,
        fact_sources=fact_sources,
        source_reports=source_reports,
        summary=summary,
    )
    summary["outputs"] = {key: str(path) for key, path in outputs.items()}
    (out_root / "question_catalog_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return summary


def build_question_classes(items: Sequence[QuestionItem], *, tenant_id: str) -> list[QuestionClass]:
    grouped: dict[str, list[QuestionItem]] = defaultdict(list)
    for item in items:
        grouped[item.question_class_id].append(item)
    classes: list[QuestionClass] = []
    for class_id, class_items in sorted(grouped.items(), key=lambda pair: (-len(pair[1]), pair[0])):
        sample = class_items[0]
        metadata = infer_question_metadata(
            sample.customer_text_redacted,
            fallback_signal=str(sample.metadata.get("signal") or ""),
        )
        source_counts = Counter(item.source_channel for item in class_items)
        dates = sorted(item.occurred_at for item in class_items if item.occurred_at)
        examples = _unique([item.customer_text_redacted for item in class_items if item.customer_text_redacted], limit=5)
        products = _unique([item.product for item in class_items if item.product], limit=8)
        grades = _unique([item.grade for item in class_items if item.grade], limit=8)
        subjects = _unique([item.subject for item in class_items if item.subject], limit=8)
        answer_status = metadata.answer_status
        bot_permission = metadata.bot_permission
        if answer_status == ANSWER_STATUS_MANAGER_ONLY:
            bot_permission = BOT_PERMISSION_MANAGER_ONLY
        elif metadata.dynamic_fact_types:
            answer_status = ANSWER_STATUS_TEMPLATE_NEEDS_CURRENT_FACT
            bot_permission = BOT_PERMISSION_ALLOWED_AFTER_FACT_CHECK
        elif any(item.manager_text_redacted for item in class_items):
            answer_status = ANSWER_STATUS_DRAFT_NEEDS_REVIEW
            bot_permission = BOT_PERMISSION_DRAFT_ONLY
        else:
            answer_status = ANSWER_STATUS_NEEDS_ROP_ANSWER
            bot_permission = BOT_PERMISSION_DRAFT_ONLY
        priority = _priority_for_class(class_items, metadata.required_fact_keys)
        classes.append(
            QuestionClass(
                tenant_id=tenant_id,
                question_class_id=class_id,
                canonical_question=metadata.canonical_question,
                narrow_scope=metadata.narrow_scope,
                class_key=metadata.class_key,
                exclusions=metadata.exclusions,
                examples_redacted=examples,
                count_total=len(class_items),
                count_calls=source_counts.get("call", 0),
                count_telegram=source_counts.get("telegram", 0),
                count_email=source_counts.get("email", 0),
                first_seen_at=dates[0] if dates else None,
                last_seen_at=dates[-1] if dates else None,
                products=products,
                grades=grades,
                subjects=subjects,
                answer_status=answer_status,
                required_fact_keys=metadata.required_fact_keys,
                fact_source_refs=_fact_source_refs(metadata.required_fact_keys),
                fact_freshness_policy=metadata.fact_freshness_policy,
                fallback_when_fact_missing=metadata.fallback_when_fact_missing,
                bot_permission=bot_permission,
                manager_handoff_reason=metadata.manager_handoff_reason,
                rop_review_priority=priority,
                metadata={
                    "intents": sorted(Counter(item.intent for item in class_items)),
                    "dynamic_fact_types": sorted({fact for item in class_items for fact in item.dynamic_fact_types}),
                },
            )
        )
    return classes


def build_answer_templates(classes: Sequence[QuestionClass], *, tenant_id: str) -> list[AnswerTemplate]:
    templates: list[AnswerTemplate] = []
    for item in classes:
        template = _template_text_for_class(item)
        permission = BOT_PERMISSION_ALLOWED_AFTER_FACT_CHECK if item.required_fact_keys else BOT_PERMISSION_DRAFT_ONLY
        if item.answer_status == ANSWER_STATUS_MANAGER_ONLY:
            permission = BOT_PERMISSION_MANAGER_ONLY
        template_id = stable_answer_template_id(
            tenant_id=tenant_id,
            question_class_id=item.question_class_id,
            template_text=template,
        )
        templates.append(
            AnswerTemplate(
                tenant_id=tenant_id,
                question_class_id=item.question_class_id,
                answer_template_id=template_id,
                template_text=template,
                required_fact_keys=item.required_fact_keys,
                approval_status=item.answer_status,
                bot_permission_if_facts_fresh=permission,
                fallback_when_fact_missing=item.fallback_when_fact_missing
                or "Если нет утвержденного ответа, передать менеджеру.",
            )
        )
    return templates


def attach_template_ids(classes: Sequence[QuestionClass], templates: Sequence[AnswerTemplate]) -> list[QuestionClass]:
    by_class = {template.question_class_id: template.answer_template_id for template in templates}
    return [replace(item, answer_template_id=by_class.get(item.question_class_id)) for item in classes]


def discover_fact_sources(roots: Sequence[Path]) -> list[CurrentFactSource]:
    sources: list[CurrentFactSource] = []
    seen: set[str] = set()
    for root in roots:
        if not root.exists():
            continue
        for path in sorted(root.rglob("*")):
            if not path.is_file() or path.suffix.lower() not in {".txt", ".md", ".csv", ".xlsx", ".json"}:
                continue
            fact_types = _fact_types_from_path(path)
            if not fact_types:
                continue
            source_id = normalize_key(f"fact_{len(sources) + 1}", "source_id")
            rel = str(path)
            if rel in seen:
                continue
            seen.add(rel)
            sources.append(
                CurrentFactSource(
                    source_id=source_id,
                    fact_types=fact_types,
                    path=rel,
                    owner="manual_owner_required",
                    last_updated_at=None,
                    freshness_policy="Нужна ручная проверка свежести перед автономным ответом.",
                    usable_for_bot=False,
                    notes="Источник найден автоматически; значения можно использовать только после утверждения владельцем.",
                )
            )
    return sources


def write_outputs(
    out_root: Path,
    *,
    items: Sequence[QuestionItem],
    classes: Sequence[QuestionClass],
    templates: Sequence[AnswerTemplate],
    fact_sources: Sequence[CurrentFactSource],
    source_reports: Sequence[Mapping[str, Any]],
    summary: Mapping[str, Any],
) -> Mapping[str, Path]:
    outputs = {
        "items_jsonl": out_root / "customer_question_items.jsonl",
        "classes_csv": out_root / "customer_question_classes.csv",
        "classes_xlsx": out_root / "customer_question_classes.xlsx",
        "templates_csv": out_root / "answer_templates.csv",
        "fact_requirements_csv": out_root / "fact_requirements.csv",
        "fact_registry_json": out_root / "current_fact_source_registry.json",
        "rop_review_xlsx": out_root / "rop_question_review_pack.xlsx",
        "approved_answers_xlsx": out_root / "approved_question_answers_draft.xlsx",
        "approved_answers_csv": out_root / "approved_question_answers_draft.csv",
        "approved_answers_json": out_root / "approved_question_answers_draft.json",
        "rop_priority_top100_xlsx": out_root / "rop_review_priority_top100.xlsx",
        "rop_priority_top100_csv": out_root / "rop_review_priority_top100.csv",
        "answer_quality_audit_json": out_root / "answer_quality_check_report.json",
        "answer_quality_audit_md": out_root / "answer_quality_check_report.md",
        "channel_preview_context_pack_json": out_root / "channel_preview_approved_context_pack.json",
        "unanswered_csv": out_root / "unanswered_questions.csv",
        "source_coverage_md": out_root / "source_coverage_report.md",
        "summary_json": out_root / "question_catalog_summary.json",
    }
    approval_drafts = build_approved_answer_drafts(classes, templates)
    priority_rows = build_rop_priority_rows(classes, templates, limit=100)
    quality_audit = build_answer_quality_audit(classes, templates, approval_drafts)
    write_jsonl(outputs["items_jsonl"], [item.to_json_dict() for item in items])
    write_csv(outputs["classes_csv"], [flatten_class(item) for item in classes])
    write_csv(outputs["templates_csv"], [flatten_template(item) for item in templates])
    write_csv(outputs["fact_requirements_csv"], fact_requirement_rows(classes))
    outputs["fact_registry_json"].write_text(
        json.dumps(
            {
                "schema_version": CATALOG_SCHEMA_VERSION,
                "safety_note": "Найденные источники фактов требуют ручной проверки свежести перед автономным ответом.",
                "sources": [item.to_json_dict() for item in fact_sources],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    unanswered = [
        flatten_class(item)
        for item in classes
        if item.answer_status
        in {
            ANSWER_STATUS_NEEDS_ROP_ANSWER,
            ANSWER_STATUS_TEMPLATE_NEEDS_CURRENT_FACT,
            ANSWER_STATUS_MANAGER_ONLY,
        }
    ]
    write_csv(outputs["unanswered_csv"], unanswered)
    write_classes_xlsx(outputs["classes_xlsx"], classes)
    write_rop_review_xlsx(outputs["rop_review_xlsx"], classes)
    write_csv(outputs["approved_answers_csv"], [flatten_approval_draft(item) for item in approval_drafts])
    write_xlsx(outputs["approved_answers_xlsx"], "approved_answers_draft", [flatten_approval_draft(item) for item in approval_drafts])
    outputs["approved_answers_json"].write_text(
        json.dumps(
            {
                "schema_version": CATALOG_SCHEMA_VERSION,
                "safety_note": "Черновик для РОПа. В этом файле нет автоматически утвержденных ответов.",
                "records": [item.to_json_dict() for item in approval_drafts],
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    write_csv(outputs["rop_priority_top100_csv"], priority_rows)
    write_xlsx(outputs["rop_priority_top100_xlsx"], "rop_priority_top100", priority_rows)
    outputs["answer_quality_audit_json"].write_text(json.dumps(quality_audit, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs["answer_quality_audit_md"].write_text(render_answer_quality_audit(quality_audit), encoding="utf-8")
    outputs["channel_preview_context_pack_json"].write_text(
        json.dumps(build_channel_preview_context_pack(approval_drafts), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    outputs["source_coverage_md"].write_text(render_source_coverage(source_reports, summary), encoding="utf-8")
    return outputs


def write_jsonl(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def write_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_classes_xlsx(path: Path, classes: Sequence[QuestionClass]) -> None:
    rows = [flatten_class(item) for item in classes]
    write_xlsx(path, "question_classes", rows)


def write_rop_review_xlsx(path: Path, classes: Sequence[QuestionClass]) -> None:
    rows = []
    for item in classes:
        if item.answer_status == ANSWER_STATUS_NOT_CUSTOMER_QUESTION:
            continue
        if item.rop_review_priority in {"critical", "high", "medium"}:
            rows.append(
                {
                    "Приоритет": item.rop_review_priority,
                    "Класс вопроса": item.canonical_question,
                    "Примеры": " | ".join(item.examples_redacted[:3]),
                    "Частота всего": item.count_total,
                    "Звонки": item.count_calls,
                    "Telegram": item.count_telegram,
                    "Почта": item.count_email,
                    "Статус ответа": item.answer_status,
                    "Можно ли боту": item.bot_permission,
                    "Нужные актуальные факты": " | ".join(item.required_fact_keys),
                    "Что делать без факта": item.fallback_when_fact_missing or "",
                    "Идеальный ответ РОПа": "",
                    "Решение РОПа": "",
                    "Комментарий": "",
                }
            )
    write_xlsx(path, "rop_review", rows)


def write_xlsx(path: Path, sheet_name: str, rows: Sequence[Mapping[str, Any]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name[:31]
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True)
    wrap_top = Alignment(wrap_text=True, vertical="top")
    fields = list(rows[0].keys()) if rows else ["empty"]
    for col, field in enumerate(fields, start=1):
        cell = worksheet.cell(row=1, column=col, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_top
        worksheet.column_dimensions[cell.column_letter].width = min(max(len(field) + 2, 14), 42)
    for row_index, row in enumerate(rows, start=1):
        for col, field in enumerate(fields, start=1):
            cell = worksheet.cell(row=row_index + 1, column=col, value=row.get(field, ""))
            cell.alignment = wrap_top
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(path)


def flatten_class(item: QuestionClass) -> Mapping[str, Any]:
    return {
        "question_class_id": item.question_class_id,
        "canonical_question": item.canonical_question,
        "narrow_scope": item.narrow_scope,
        "exclusions": item.exclusions,
        "examples_redacted": " | ".join(item.examples_redacted),
        "count_total": item.count_total,
        "count_calls": item.count_calls,
        "count_telegram": item.count_telegram,
        "count_email": item.count_email,
        "first_seen_at": item.first_seen_at.isoformat() if item.first_seen_at else "",
        "last_seen_at": item.last_seen_at.isoformat() if item.last_seen_at else "",
        "products": " | ".join(item.products),
        "grades": " | ".join(item.grades),
        "subjects": " | ".join(item.subjects),
        "answer_status": item.answer_status,
        "answer_template_id": item.answer_template_id or "",
        "required_fact_keys": " | ".join(item.required_fact_keys),
        "fact_source_refs": " | ".join(item.fact_source_refs),
        "fact_freshness_policy": item.fact_freshness_policy or "",
        "fallback_when_fact_missing": item.fallback_when_fact_missing or "",
        "bot_permission": item.bot_permission,
        "manager_handoff_reason": item.manager_handoff_reason or "",
        "rop_review_priority": item.rop_review_priority,
    }


def flatten_template(item: AnswerTemplate) -> Mapping[str, Any]:
    return {
        "answer_template_id": item.answer_template_id,
        "question_class_id": item.question_class_id,
        "template_text": item.template_text,
        "required_fact_keys": " | ".join(item.required_fact_keys),
        "approval_status": item.approval_status,
        "bot_permission_if_facts_fresh": item.bot_permission_if_facts_fresh,
        "fallback_when_fact_missing": item.fallback_when_fact_missing,
    }


def fact_requirement_rows(classes: Sequence[QuestionClass]) -> list[Mapping[str, Any]]:
    rows: list[Mapping[str, Any]] = []
    for item in classes:
        for fact_key in item.required_fact_keys:
            rows.append(
                {
                    "question_class_id": item.question_class_id,
                    "canonical_question": item.canonical_question,
                    "fact_key": fact_key,
                    "freshness_policy": item.fact_freshness_policy or "",
                    "fallback_when_missing": item.fallback_when_fact_missing or "",
                    "bot_permission": item.bot_permission,
                }
            )
    return rows


def build_approved_answer_drafts(
    classes: Sequence[QuestionClass],
    templates: Sequence[AnswerTemplate],
) -> list[ApprovedQuestionAnswerDraft]:
    templates_by_class = {template.question_class_id: template for template in templates}
    drafts: list[ApprovedQuestionAnswerDraft] = []
    for item in sorted(classes, key=lambda row: (-_priority_score(row), -row.count_total, row.canonical_question)):
        if item.answer_status == ANSWER_STATUS_NOT_CUSTOMER_QUESTION:
            continue
        template = templates_by_class.get(item.question_class_id)
        draft_text = template.template_text if template else _template_text_for_class(item)
        drafts.append(
            ApprovedQuestionAnswerDraft(
                tenant_id=item.tenant_id,
                question_class_id=item.question_class_id,
                canonical_question=item.canonical_question,
                count_total=item.count_total,
                draft_template_text=draft_text,
                required_fact_keys=item.required_fact_keys,
                fact_source_refs=item.fact_source_refs,
                bot_permission=item.bot_permission,
                rop_decision=_default_rop_decision(item),
                final_approved_answer="",
                rop_comment="",
                approved_for_bot=False,
                auto_approved=False,
            )
        )
    return drafts


def flatten_approval_draft(item: ApprovedQuestionAnswerDraft) -> Mapping[str, Any]:
    return {
        "approval_record_id": item.approval_record_id,
        "question_class_id": item.question_class_id,
        "canonical_question": item.canonical_question,
        "count_total": item.count_total,
        "draft_template_text": item.draft_template_text,
        "required_fact_keys": " | ".join(item.required_fact_keys),
        "fact_source_refs": " | ".join(item.fact_source_refs),
        "bot_permission": item.bot_permission,
        "runtime_bot_permission": item.runtime_bot_permission,
        "rop_decision": item.rop_decision,
        "approved_for_bot": "yes" if item.approved_for_bot else "no",
        "auto_approved": "yes" if item.auto_approved else "no",
        "can_autosend": "yes" if item.can_autosend else "no",
        "requires_manager_review": "yes" if item.requires_manager_review else "no",
        "final_approved_answer": item.final_approved_answer,
        "rop_comment": item.rop_comment,
    }


def build_rop_priority_rows(
    classes: Sequence[QuestionClass],
    templates: Sequence[AnswerTemplate],
    *,
    limit: int,
) -> list[Mapping[str, Any]]:
    templates_by_class = {template.question_class_id: template for template in templates}
    rows: list[Mapping[str, Any]] = []
    for rank, item in enumerate(sorted(classes, key=lambda row: (-_priority_score(row), -row.count_total, row.canonical_question))[:limit], start=1):
        if item.answer_status == ANSWER_STATUS_NOT_CUSTOMER_QUESTION:
            continue
        template = templates_by_class.get(item.question_class_id)
        rows.append(
            {
                "rank": rank,
                "priority_score": _priority_score(item),
                "rop_review_priority": item.rop_review_priority,
                "canonical_question": item.canonical_question,
                "count_total": item.count_total,
                "count_calls": item.count_calls,
                "count_telegram": item.count_telegram,
                "count_email": item.count_email,
                "answer_status": item.answer_status,
                "bot_permission": item.bot_permission,
                "required_fact_keys": " | ".join(item.required_fact_keys),
                "why_priority": _priority_reason(item),
                "examples_redacted": " | ".join(item.examples_redacted[:3]),
                "draft_template_text": template.template_text if template else _template_text_for_class(item),
                "rop_decision": "",
                "final_approved_answer": "",
                "rop_comment": "",
            }
        )
    return rows


def build_answer_quality_audit(
    classes: Sequence[QuestionClass],
    templates: Sequence[AnswerTemplate],
    approval_drafts: Sequence[ApprovedQuestionAnswerDraft],
) -> Mapping[str, Any]:
    templates_by_class = {template.question_class_id: template for template in templates}
    findings: list[Mapping[str, Any]] = []
    prompt_checks: list[Mapping[str, Any]] = []
    for item in classes:
        template = templates_by_class.get(item.question_class_id)
        template_text = template.template_text if template else ""
        if item.answer_status == ANSWER_STATUS_MANAGER_ONLY and item.bot_permission != BOT_PERMISSION_MANAGER_ONLY:
            findings.append(_quality_finding(item, "p0", "manager_only_permission_mismatch"))
        if item.required_fact_keys and "свеж" not in template_text.lower():
            findings.append(_quality_finding(item, "p1", "dynamic_fact_template_missing_freshness_warning"))
        if item.required_fact_keys and item.bot_permission == BOT_PERMISSION_DRAFT_ONLY:
            findings.append(_quality_finding(item, "p1", "dynamic_fact_class_left_as_draft_only"))
        if item.bot_permission == BOT_PERMISSION_NOT_ALLOWED and item.answer_status != ANSWER_STATUS_NOT_CUSTOMER_QUESTION:
            findings.append(_quality_finding(item, "p1", "not_allowed_without_noise_status"))
        if not item.examples_redacted:
            findings.append(_quality_finding(item, "p2", "missing_examples"))
        for prompt in _synthetic_prompts_for_class(item)[:3]:
            prompt_checks.append(
                {
                    "question_class_id": item.question_class_id,
                    "canonical_question": item.canonical_question,
                    "test_prompt": prompt,
                    "expected_behavior": _expected_behavior(item),
                    "must_not_do": _must_not_do(item),
                }
            )
    approval_errors = [
        item.approval_record_id
        for item in approval_drafts
        if item.auto_approved or item.approved_for_bot or item.final_approved_answer or item.can_autosend or item.runtime_bot_permission != BOT_PERMISSION_NOT_ALLOWED
    ]
    severity_counts = Counter(str(item["severity"]) for item in findings)
    return {
        "schema_version": CATALOG_SCHEMA_VERSION,
        "verdict": "pass" if not approval_errors and not severity_counts.get("p0") else "blocked",
        "classes_checked": len(classes),
        "templates_checked": len(templates),
        "approval_drafts_checked": len(approval_drafts),
        "auto_approval_errors": approval_errors,
        "findings_by_severity": dict(sorted(severity_counts.items())),
        "findings": findings,
        "synthetic_prompt_checks": prompt_checks[:300],
        "safety_note": "Проверка не вызывает LLM и не отправляет сообщения; это deterministic dry-run.",
    }


def render_answer_quality_audit(audit: Mapping[str, Any]) -> str:
    lines = [
        "# Проверка качества шаблонов ответов",
        "",
        f"Вердикт: {audit.get('verdict')}",
        f"Проверено классов: {audit.get('classes_checked')}",
        f"Проверено шаблонов: {audit.get('templates_checked')}",
        f"Черновиков утверждения: {audit.get('approval_drafts_checked')}",
        f"Ошибок автоутверждения: {len(audit.get('auto_approval_errors') or [])}",
        "",
        "## Находки по серьезности",
        "",
    ]
    for severity, count in dict(audit.get("findings_by_severity") or {}).items():
        lines.append(f"- {severity}: {count}")
    lines.extend(["", "## Ограничения", "", str(audit.get("safety_note") or ""), ""])
    return "\n".join(lines)


def build_channel_preview_context_pack(approval_drafts: Sequence[ApprovedQuestionAnswerDraft]) -> Mapping[str, Any]:
    approved = [item for item in approval_drafts if item.approved_for_bot and item.final_approved_answer and not item.auto_approved]
    return {
        "schema_version": CATALOG_SCHEMA_VERSION,
        "context_pack_type": "channel_preview_approved_answers",
        "approved_answers": [item.to_json_dict() for item in approved],
        "pending_rop_review_count": len(approval_drafts) - len(approved),
        "approved_count": len(approved),
        "safety": {
            "approved_only": True,
            "live_send": False,
            "requires_channel_preview_manager_approval": True,
        },
        "note": "Текущая сборка не содержит автоутвержденных ответов. Предпросмотр может использовать только записи, утвержденные РОПом.",
    }


def build_summary(
    *,
    config: CatalogBuildConfig,
    items: Sequence[QuestionItem],
    classes: Sequence[QuestionClass],
    templates: Sequence[AnswerTemplate],
    fact_sources: Sequence[CurrentFactSource],
    source_reports: Sequence[Mapping[str, Any]],
) -> dict[str, Any]:
    by_source = Counter(item.source_channel for item in items)
    by_status = Counter(item.answer_status for item in classes)
    dynamic_classes = [item for item in classes if item.required_fact_keys]
    return {
        "schema_version": CATALOG_SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": config.tenant_id,
        "since": config.since.isoformat(),
        "safety": question_catalog_safety_contract(),
        "contract_inventory": question_catalog_contract_inventory(),
        "totals": {
            "question_items": len(items),
            "question_classes": len(classes),
            "answer_templates": len(templates),
            "fact_sources": len(fact_sources),
            "dynamic_fact_classes": len(dynamic_classes),
        },
        "counts": {
            "items_by_source": dict(sorted(by_source.items())),
            "classes_by_answer_status": dict(sorted(by_status.items())),
            "top_classes": [
                {
                    "question_class_id": item.question_class_id,
                    "canonical_question": item.canonical_question,
                    "count_total": item.count_total,
                    "answer_status": item.answer_status,
                    "required_fact_keys": list(item.required_fact_keys),
                }
                for item in sorted(classes, key=lambda row: (-row.count_total, row.canonical_question))[:25]
            ],
        },
        "source_reports": list(source_reports),
        "limitations": [
            "Каталог построен только по доступным локальным артефактам.",
            "Почта и Telegram не читаются live в этом проходе.",
            "Цены, расписание, скидки и наборы требуют отдельной проверки свежести фактов перед ответом бота.",
            "Все примеры в выходных файлах должны быть очищены от прямых контактов и персональных данных.",
        ],
    }


def render_source_coverage(source_reports: Sequence[Mapping[str, Any]], summary: Mapping[str, Any]) -> str:
    lines = [
        "# Покрытие источников единого каталога вопросов",
        "",
        f"Дата сборки: {summary.get('generated_at')}",
        f"Период с: {summary.get('since')}",
        "",
        "## Итоги",
        "",
        f"- Отдельных вопросов: {summary['totals']['question_items']}",
        f"- Классов вопросов: {summary['totals']['question_classes']}",
        f"- Классов, требующих актуальных фактов: {summary['totals']['dynamic_fact_classes']}",
        "",
        "## Источники",
        "",
    ]
    for report in source_reports:
        lines.extend(
            [
                f"### {report.get('source_id')}",
                "",
                f"- Статус: {report.get('status')}",
                f"- Путь: `{report.get('path')}`",
                f"- Извлечено вопросов: {report.get('items_extracted', 0)}",
                f"- Всего строк/сообщений: {report.get('rows_total', report.get('archives_total', 'н/д'))}",
                "",
            ]
        )
    lines.extend(
        [
            "## Ограничения",
            "",
            "- Live-чтение почты, Telegram и CRM не выполнялось.",
            "- Повторное распознавание и анализ звонков не запускались.",
            "- Для цен и расписания нужен отдельный утвержденный файл фактов.",
            "",
        ]
    )
    return "\n".join(lines)


def _dedupe_items(items: Sequence[QuestionItem]) -> list[QuestionItem]:
    result: list[QuestionItem] = []
    seen: set[str] = set()
    for item in items:
        key = item.question_item_id
        if key in seen:
            continue
        seen.add(key)
        assert_public_text_safe(item.customer_text_redacted, field_name="customer_text_redacted")
        if item.manager_text_redacted:
            assert_public_text_safe(item.manager_text_redacted, field_name="manager_text_redacted")
        result.append(item)
    return result


def _unique(values: Sequence[Any], *, limit: int) -> tuple[str, ...]:
    result: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if text and text not in result:
            result.append(text)
        if len(result) >= limit:
            break
    return tuple(result)


def _priority_for_class(items: Sequence[QuestionItem], required_fact_keys: Sequence[str]) -> str:
    if len(items) >= 25 or required_fact_keys:
        return "high"
    if len(items) >= 5:
        return "medium"
    return "low"


def _priority_score(item: QuestionClass) -> int:
    score = min(item.count_total, 500)
    if item.required_fact_keys:
        score += 140
    if item.bot_permission == BOT_PERMISSION_MANAGER_ONLY:
        score += 70
    if item.count_telegram:
        score += min(item.count_telegram, 80)
    if item.count_email:
        score += min(item.count_email, 80)
    if item.rop_review_priority == "critical":
        score += 250
    elif item.rop_review_priority == "high":
        score += 120
    elif item.rop_review_priority == "medium":
        score += 40
    return score


def _priority_reason(item: QuestionClass) -> str:
    reasons: list[str] = []
    if item.count_total >= 25:
        reasons.append("частый вопрос")
    if item.required_fact_keys:
        reasons.append("нужен актуальный факт")
    if item.bot_permission == BOT_PERMISSION_MANAGER_ONLY:
        reasons.append("рискованный вопрос для менеджера")
    if item.count_telegram or item.count_email:
        reasons.append("приходит из текстовых каналов")
    return "; ".join(reasons) or "низкая частота"


def _default_rop_decision(item: QuestionClass) -> str:
    if item.answer_status == ANSWER_STATUS_MANAGER_ONLY:
        return "manager_only_pending_confirmation"
    if item.required_fact_keys:
        return "needs_current_fact_approval"
    if item.answer_status == ANSWER_STATUS_NEEDS_ROP_ANSWER:
        return "needs_rop_answer"
    return "pending_rop_review"


def _quality_finding(item: QuestionClass, severity: str, code: str) -> Mapping[str, Any]:
    return {
        "severity": severity,
        "code": code,
        "question_class_id": item.question_class_id,
        "canonical_question": item.canonical_question,
        "count_total": item.count_total,
    }


def _synthetic_prompts_for_class(item: QuestionClass) -> tuple[str, ...]:
    prompts = [
        f"Подскажите, пожалуйста: {item.canonical_question}?",
        f"Хочу уточнить про {item.canonical_question.lower()}",
    ]
    if item.examples_redacted:
        prompts.insert(0, item.examples_redacted[0])
    return tuple(dict.fromkeys(prompts))


def _expected_behavior(item: QuestionClass) -> str:
    if item.bot_permission == BOT_PERMISSION_MANAGER_ONLY:
        return "не отвечать самостоятельно, передать менеджеру"
    if item.required_fact_keys:
        return "дать ответ только после проверки актуального файла фактов"
    return "показать менеджеру черновик ответа, не отправлять live"


def _must_not_do(item: QuestionClass) -> str:
    if item.required_fact_keys:
        return "не выдумывать цену, расписание, адрес, скидку, документы или программу"
    if item.bot_permission == BOT_PERMISSION_MANAGER_ONLY:
        return "не обещать клиенту решение без менеджера"
    return "не отправлять клиенту без утверждения"


def _fact_source_refs(required_fact_keys: Sequence[str]) -> tuple[str, ...]:
    return tuple(f"current_fact_source_registry:{key}" for key in required_fact_keys)


def _template_text_for_class(item: QuestionClass) -> str:
    if item.required_fact_keys:
        placeholders = ", ".join("{" + key.replace(".current", "") + "}" for key in item.required_fact_keys)
        return (
            f"По вопросу «{item.canonical_question}» ориентируемся на актуальные данные: {placeholders}. "
            "Перед ответом нужно проверить свежий файл фактов. Если данные не подтверждены, передаем менеджеру."
        )
    if item.answer_status == ANSWER_STATUS_MANAGER_ONLY:
        return (
            f"По вопросу «{item.canonical_question}» бот не отвечает сам. "
            "Нужно передать менеджеру и зафиксировать, какие данные требуется уточнить."
        )
    return (
        f"По вопросу «{item.canonical_question}» дать короткий ответ по утвержденной базе знаний, "
        "затем задать один уточняющий вопрос и предложить помощь менеджера."
    )


def _fact_types_from_path(path: Path) -> tuple[str, ...]:
    text = str(path).casefold()
    fact_types: list[str] = []
    if any(marker in text for marker in ("стоим", "цен", "price", "абонем")):
        fact_types.append("price")
    if any(marker in text for marker in ("распис", "schedule")):
        fact_types.append("schedule")
    if any(marker in text for marker in ("скид", "discount", "акци")):
        fact_types.append("discount")
    if any(marker in text for marker in ("рассроч", "installment", "долями")):
        fact_types.append("installment")
    if any(marker in text for marker in ("адрес", "location", "филиал")):
        fact_types.append("location")
    if any(marker in text for marker in ("пробн", "trial")):
        fact_types.append("trial")
    if any(marker in text for marker in ("документ", "договор", "правила", "оферт")):
        fact_types.append("documents")
    if any(marker in text for marker in ("программ", "курс", "предмет")):
        fact_types.append("program")
    return tuple(dict.fromkeys(fact_types))


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build read-only customer question catalog")
    parser.add_argument("--project-root", default=".", help="Project root")
    parser.add_argument("--out-root", default=None, help="Output root, defaults to product_data/question_catalog")
    parser.add_argument("--tenant-id", default="foton")
    parser.add_argument("--since", default="2025-01-01T00:00:00+00:00")
    parser.add_argument("--calls-enriched-reviews", default=None)
    parser.add_argument("--telegram-messages-jsonl", default=None)
    parser.add_argument("--mail-archive-root", default=None)
    return parser.parse_args(argv)


def config_from_args(args: argparse.Namespace) -> CatalogBuildConfig:
    project_root = Path(args.project_root).resolve()
    config = default_config(project_root, Path(args.out_root).resolve() if args.out_root else None)
    since = datetime.fromisoformat(str(args.since).replace("Z", "+00:00"))
    if since.tzinfo is None or since.utcoffset() is None:
        since = since.replace(tzinfo=timezone.utc)
    return CatalogBuildConfig(
        project_root=project_root,
        out_root=Path(args.out_root).resolve() if args.out_root else config.out_root,
        tenant_id=args.tenant_id,
        since=since,
        calls_enriched_reviews=Path(args.calls_enriched_reviews).resolve()
        if args.calls_enriched_reviews
        else config.calls_enriched_reviews,
        telegram_messages_jsonl=Path(args.telegram_messages_jsonl).resolve()
        if args.telegram_messages_jsonl
        else config.telegram_messages_jsonl,
        mail_archive_root=Path(args.mail_archive_root).resolve() if args.mail_archive_root else config.mail_archive_root,
        fact_source_roots=config.fact_source_roots,
    )
