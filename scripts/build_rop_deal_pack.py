from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Optional


CANONICAL_EXPORT_POINTER_NAME = "CANONICAL_EXPORT.txt"


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _clip_excel_text(value: Any, *, max_chars: int = 32767) -> str:
    text = _safe_text(value)
    if len(text) <= max_chars:
        return text
    marker = " [обрезано по лимиту Excel]"
    return text[: max_chars - len(marker)].rstrip() + marker


def _looks_artificially_clipped(value: Any) -> bool:
    text = _safe_text(value).rstrip()
    if not text:
        return False
    if text.endswith("[обрезано по лимиту Excel]"):
        return False
    return text.endswith("...") or text.endswith("…")


def _prefer_fresh_text(*candidates: Any) -> str:
    clipped_candidates: list[str] = []
    for value in candidates:
        text = _safe_text(value)
        if not text:
            continue
        if not _looks_artificially_clipped(text):
            return text
        clipped_candidates.append(text)
    if not clipped_candidates:
        return ""
    return max(clipped_candidates, key=len)


def _bootstrap_runtime_env(project_root: Path) -> None:
    src_path = str(project_root / "src")
    if src_path not in sys.path:
        sys.path.insert(0, src_path)
    env_private = project_root / "stable_runtime" / "amocrm_runtime" / ".env.private"
    if env_private.exists():
        try:
            from dotenv import load_dotenv

            load_dotenv(env_private, override=False)
        except Exception:
            pass


def _normalize_phone_digits(value: Any) -> str:
    text = _safe_text(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    return digits


def _parse_dt(value: Any) -> Optional[datetime]:
    if isinstance(value, (int, float)) and value:
        try:
            return datetime.fromtimestamp(float(value))
        except (OverflowError, OSError, ValueError):
            pass

    text = _safe_text(value).replace("T", " ").replace("Z", "")
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M:%S",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _fmt_dt(value: Any) -> str:
    dt = _parse_dt(value)
    if dt is None:
        return _safe_text(value)
    return dt.strftime("%Y-%m-%d %H:%M")


def _fmt_date(value: Any) -> str:
    dt = _parse_dt(value)
    if dt is None:
        return _safe_text(value)
    return dt.strftime("%Y-%m-%d")


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _priority_score(result: dict[str, Any], enrichment: dict[str, Any]) -> float:
    risk = _safe_text(result.get("premature_close_risk") or result.get("AI-риск") or result.get("AI-risk")).lower()
    verdict = _safe_text(result.get("close_verdict") or result.get("AI-вердикт")).lower()
    temp = _safe_text(result.get("current_sales_temperature") or enrichment.get("sales_temperature")).lower()

    risk_score = {
        "critical": 100,
        "high": 80,
        "medium": 50,
        "low": 20,
        "manual_review": 0,
    }.get(risk, 0)
    verdict_score = {
        "reopen_recommended": 30,
        "closed_too_early": 22,
        "follow_up_needed": 18,
        "alternative_offer_needed": 16,
        "manual_review": 0,
    }.get(verdict, 0)
    temp_score = {"hot": 12, "warm": 7, "cold": 2}.get(temp, 0)
    llm_conf = float(result.get("confidence") or 0)
    match_conf = float(result.get("match_confidence") or 0)
    post_close = int(enrichment.get("post_close_calls_count") or 0)
    next_step_bonus = 5 if _safe_text(result.get("recommended_next_step") or enrichment.get("recommended_next_step")) else 0

    return round(
        risk_score
        + verdict_score
        + temp_score
        + llm_conf * 10
        + match_conf * 10
        + min(post_close * 4, 16)
        + next_step_bonus,
        2,
    )


def _summarize_tallanto_live(tallanto_live: dict[str, Any]) -> dict[str, str]:
    if not isinstance(tallanto_live, dict):
        return {}

    contexts = tallanto_live.get("contexts") if isinstance(tallanto_live.get("contexts"), list) else []
    first = contexts[0] if contexts else {}
    contact = first.get("contact") if isinstance(first.get("contact"), dict) else {}
    opps = first.get("opportunities") if isinstance(first.get("opportunities"), list) else []

    opportunity_preview = " | ".join(
        f"{_safe_text(item.get('name'))} [{_safe_text(item.get('sales_stage'))}]"
        for item in opps[:5]
        if isinstance(item, dict) and (_safe_text(item.get("name")) or _safe_text(item.get("sales_stage")))
    )

    return {
        "Статус Tallanto live": _safe_text(tallanto_live.get("status")),
        "Способ матчинга Tallanto": _safe_text(tallanto_live.get("matched_via")),
        "Найдено контактов Tallanto live": str(tallanto_live.get("contacts_found") or 0),
        "Контакт Tallanto live": _safe_text(contact.get("name")),
        "Филиал Tallanto live": _safe_text(contact.get("branch")),
        "Ответственный Tallanto live": _safe_text(contact.get("assigned_user_name")),
        "Возможностей Tallanto": str(first.get("opportunity_count") or 0),
        "Финансовых записей Tallanto": str(first.get("finance_count") or 0),
        "Возможности Tallanto": opportunity_preview,
    }


def _summarize_post_close_calls_from_dossier(result: dict[str, Any], call_history: list[dict[str, Any]]) -> dict[str, Any]:
    lead_info = result.get("dossier") if isinstance(result.get("dossier"), dict) else {}
    lead_data = lead_info.get("lead") if isinstance(lead_info.get("lead"), dict) else {}
    closed_at = _parse_dt(result.get("lead_closed_at") or lead_data.get("closed_at"))
    if closed_at is None:
        return {
            "post_close_calls_count": 0,
            "post_close_last_at": "",
            "post_close_calls_summary": "",
        }

    post_rows: list[dict[str, Any]] = []
    for row in call_history:
        started_at = _parse_dt(row.get("started_at"))
        if started_at and started_at > closed_at:
            post_rows.append(row)
    post_rows.sort(key=lambda item: _parse_dt(item.get("started_at")) or datetime.min)

    preview: list[str] = []
    for row in post_rows[:5]:
        summary = _clip_excel_text(row.get("summary"))
        preview.append(f"{_fmt_dt(row.get('started_at'))} - {_safe_text(row.get('manager_name'))}: {summary}")

    return {
        "post_close_calls_count": len(post_rows),
        "post_close_last_at": _fmt_dt(post_rows[-1].get("started_at")) if post_rows else "",
        "post_close_calls_summary": " | ".join(preview),
    }


def _detect_latest_master_export_dir(project_root: Path) -> Path:
    stable_runtime = project_root / "stable_runtime"
    pointer_path = stable_runtime / CANONICAL_EXPORT_POINTER_NAME
    if pointer_path.exists():
        raw_target = _safe_text(pointer_path.read_text(encoding="utf-8"))
        if raw_target:
            candidate = Path(raw_target)
            if not candidate.is_absolute():
                candidate = stable_runtime / raw_target
            candidate = candidate.resolve()
            if candidate.is_dir() and (candidate / "master_contacts_ru.csv").exists() and (candidate / "master_calls_ru.csv").exists():
                return candidate
    candidates = sorted(
        stable_runtime.glob("sales_master_export_*review_accepted"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No sales_master_export_*review_accepted directory found under stable_runtime.")
    return candidates[0]


def _resolve_lead_snapshot_path(project_root: Path, run_dir: Path, provided: str) -> Path:
    if provided:
        return Path(provided).expanduser().resolve()
    run_local = run_dir / "live_recent_leads.json"
    if run_local.exists():
        return run_local.resolve()
    fallback = project_root / "prod_runtime_transfer" / "data_handoff" / "live_export" / "leads.json"
    if fallback.exists():
        return fallback.resolve()
    raise FileNotFoundError("Lead snapshot JSON not found. Provide --lead-snapshot-json or place live_recent_leads.json in run dir.")


def _resolve_full_lead_snapshot_path(project_root: Path, provided: str, lead_snapshot_path: Path) -> Path:
    if provided:
        return Path(provided).expanduser().resolve()
    fallback = project_root / "prod_runtime_transfer" / "data_handoff" / "live_export" / "leads.json"
    if fallback.exists():
        return fallback.resolve()
    return lead_snapshot_path


def _load_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _load_master_contacts(master_export_dir: Path) -> dict[str, dict[str, Any]]:
    path = master_export_dir / "master_contacts_ru.csv"
    if not path.exists():
        return {}
    rows = _load_csv_rows(path)
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        phone = _normalize_phone_digits(row.get("Телефон клиента"))
        if phone:
            result[phone] = row
    return result


def _load_master_calls(master_export_dir: Path) -> dict[str, list[dict[str, Any]]]:
    path = master_export_dir / "master_calls_ru.csv"
    if not path.exists():
        return {}
    rows = _load_csv_rows(path)
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        phone = _normalize_phone_digits(row.get("Телефон клиента"))
        if phone:
            grouped.setdefault(phone, []).append(row)
    for phone_rows in grouped.values():
        phone_rows.sort(key=lambda item: _parse_dt(item.get("Дата и время звонка")) or datetime.min)
    return grouped


def _extract_lead_snapshot_utm(custom_fields: list[dict[str, Any]]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for field in custom_fields:
        if not isinstance(field, dict):
            continue
        field_name = _safe_text(field.get("field_name")).lower()
        field_code = _safe_text(field.get("field_code")).lower()
        values = field.get("values") if isinstance(field.get("values"), list) else []
        value = ""
        if values:
            first = values[0]
            if isinstance(first, dict):
                value = _safe_text(first.get("value"))
            else:
                value = _safe_text(first)
        if not value:
            continue
        key = field_code or field_name
        if key in {"utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term", "utm_referrer"}:
            mapping[key] = value
    return mapping


def _extract_lead_snapshot_source(embedded: dict[str, Any]) -> str:
    source = embedded.get("source")
    if isinstance(source, dict):
        return _safe_text(source.get("name") or source.get("external_id") or source.get("id"))
    if isinstance(source, list) and source:
        first = source[0]
        if isinstance(first, dict):
            return _safe_text(first.get("name") or first.get("external_id") or first.get("id"))
    return ""


def _extract_loss_reason_value(lead: dict[str, Any], custom_fields: list[dict[str, Any]]) -> str:
    embedded = lead.get("_embedded") if isinstance(lead.get("_embedded"), dict) else {}
    loss_reason = embedded.get("loss_reason")
    if isinstance(loss_reason, list) and loss_reason and isinstance(loss_reason[0], dict):
        value = _safe_text(loss_reason[0].get("name"))
        if value:
            return value
    for field in custom_fields:
        if not isinstance(field, dict):
            continue
        field_name = _safe_text(field.get("field_name")).casefold()
        if field_name != "причина отказа (лид)":
            continue
        values = field.get("values") if isinstance(field.get("values"), list) else []
        if not values:
            continue
        first = values[0]
        if isinstance(first, dict):
            return _safe_text(first.get("value"))
        return _safe_text(first)
    return ""


def _loss_reason_is_active_client(value: Any) -> bool:
    text = _safe_text(value).casefold()
    if not text:
        return False
    markers = (
        "действующий клиент",
        "действующий ученик",
        "текущий клиент",
        "текущий ученик",
        "продолжает обучение",
        "уже учится",
    )
    return any(marker in text for marker in markers)


def _extract_snapshot_contact_id(embedded: dict[str, Any]) -> int:
    contacts = embedded.get("contacts")
    if isinstance(contacts, list) and contacts:
        first = contacts[0]
        if isinstance(first, dict):
            return _safe_int(first.get("id"))
    return 0


def _load_lead_snapshot(path: Path) -> dict[int, dict[str, Any]]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        return {}
    result: dict[int, dict[str, Any]] = {}
    for lead in payload:
        if not isinstance(lead, dict):
            continue
        lead_id = _safe_int(lead.get("id"))
        if not lead_id:
            continue
        embedded = lead.get("_embedded") if isinstance(lead.get("_embedded"), dict) else {}
        custom_fields = lead.get("custom_fields_values") if isinstance(lead.get("custom_fields_values"), list) else []
        result[lead_id] = {
            "created_at": lead.get("created_at"),
            "updated_at": lead.get("updated_at"),
            "closed_at": lead.get("closed_at"),
            "pipeline_id": lead.get("pipeline_id"),
            "status_id": lead.get("status_id"),
            "responsible_user_id": lead.get("responsible_user_id"),
            "contact_id": _extract_snapshot_contact_id(embedded),
            "source_name": _extract_lead_snapshot_source(embedded),
            "utm": _extract_lead_snapshot_utm(custom_fields),
            "loss_reason": _extract_loss_reason_value(lead, custom_fields),
            "name": _safe_text(lead.get("name")),
        }
    return result


def _index_open_snapshot_leads_by_contact(path: Path) -> dict[int, list[dict[str, Any]]]:
    if not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        return {}

    result: dict[int, list[dict[str, Any]]] = {}
    for lead in payload:
        if not isinstance(lead, dict):
            continue
        lead_id = _safe_int(lead.get("id"))
        status_id = _safe_int(lead.get("status_id"))
        if not lead_id or status_id in {142, 143}:
            continue
        embedded = lead.get("_embedded") if isinstance(lead.get("_embedded"), dict) else {}
        contact_id = _extract_snapshot_contact_id(embedded)
        if not contact_id:
            continue
        result.setdefault(contact_id, []).append(
            {
                "id": lead_id,
                "name": _safe_text(lead.get("name")),
                "pipeline_name": str(_safe_int(lead.get("pipeline_id")) or _safe_text(lead.get("pipeline_id"))),
                "status_name": str(status_id),
                "updated_at": lead.get("updated_at"),
            }
        )

    for rows in result.values():
        rows.sort(key=lambda item: _parse_dt(item.get("updated_at")) or datetime.min, reverse=True)
    return result


def _load_live_open_leads_by_contact(contact_ids: set[int], *, enabled: bool) -> dict[int, list[dict[str, Any]]]:
    if not enabled or not contact_ids:
        return {}

    from concurrent.futures import ThreadPoolExecutor, as_completed

    from mango_mvp.amocrm_runtime.amo_integration import fetch_pipelines_with_statuses, fetch_related_leads
    from mango_mvp.amocrm_runtime.db import SessionLocal
    from mango_mvp.amocrm_runtime.deals import (
        TERMINAL_LOST_STATUS_IDS,
        TERMINAL_SUCCESS_STATUS_IDS,
        _build_pipeline_meta,
    )

    session = SessionLocal()
    try:
        pipelines = fetch_pipelines_with_statuses(session)
    finally:
        session.close()
    pipeline_map, status_map = _build_pipeline_meta(pipelines)

    def _fetch(contact_id: int) -> tuple[int, list[dict[str, Any]]]:
        worker = SessionLocal()
        try:
            leads = fetch_related_leads(worker, contact_id=contact_id, limit=250)
        finally:
            worker.close()

        rows: list[dict[str, Any]] = []
        for lead in leads:
            if not isinstance(lead, dict):
                continue
            lead_id = _safe_int(lead.get("id"))
            if not lead_id:
                continue
            status_id = _safe_int(lead.get("status_id"))
            if status_id in TERMINAL_SUCCESS_STATUS_IDS or status_id in TERMINAL_LOST_STATUS_IDS:
                continue
            pipeline_id = _safe_int(lead.get("pipeline_id"))
            rows.append(
                {
                    "id": lead_id,
                    "name": _safe_text(lead.get("name")),
                    "pipeline_name": _safe_text((pipeline_map.get(pipeline_id) or {}).get("name")),
                    "status_name": _safe_text((status_map.get((pipeline_id, status_id)) or {}).get("name")),
                    "updated_at": lead.get("updated_at"),
                }
            )
        rows.sort(key=lambda item: _parse_dt(item.get("updated_at")) or datetime.min, reverse=True)
        return contact_id, rows

    result: dict[int, list[dict[str, Any]]] = {}
    with ThreadPoolExecutor(max_workers=min(6, len(contact_ids))) as executor:
        futures = [executor.submit(_fetch, contact_id) for contact_id in sorted(contact_ids)]
        for future in as_completed(futures):
            contact_id, rows = future.result()
            result[contact_id] = rows
    return result


def _summarize_open_leads(open_leads: list[dict[str, Any]], *, current_lead_id: int) -> tuple[int, str]:
    filtered = [row for row in open_leads if _safe_int(row.get("id")) != current_lead_id]
    preview = " | ".join(
        f"{_safe_int(row.get('id'))}: {_safe_text(row.get('name'))} [{_safe_text(row.get('pipeline_name'))} / {_safe_text(row.get('status_name'))}]"
        for row in filtered[:5]
    )
    return len(filtered), preview


def _summarize_post_close_calls_from_master(*, closed_at_value: Any, master_call_rows: list[dict[str, Any]]) -> dict[str, Any]:
    closed_at = _parse_dt(closed_at_value)
    if closed_at is None:
        return {
            "post_close_calls_count": 0,
            "post_close_last_at": "",
            "post_close_calls_summary": "",
        }

    post_rows: list[dict[str, Any]] = []
    for row in master_call_rows:
        call_dt = _parse_dt(row.get("Дата и время звонка"))
        if call_dt and call_dt > closed_at:
            post_rows.append(row)

    preview: list[str] = []
    for row in post_rows[:5]:
        summary = _clip_excel_text(row.get("Краткое резюме разговора"))
        preview.append(f"{_fmt_dt(row.get('Дата и время звонка'))} - {_safe_text(row.get('Менеджер'))}: {summary}")

    return {
        "post_close_calls_count": len(post_rows),
        "post_close_last_at": _fmt_dt(post_rows[-1].get("Дата и время звонка")) if post_rows else "",
        "post_close_calls_summary": " | ".join(preview),
    }


def _manager_history_from_master(master_call_rows: list[dict[str, Any]]) -> list[str]:
    values: list[str] = []
    seen: set[str] = set()
    for row in reversed(master_call_rows):
        manager = _safe_text(row.get("Менеджер"))
        if not manager or manager in seen:
            continue
        seen.add(manager)
        values.append(manager)
        if len(values) >= 6:
            break
    return values


def _compact_candidate_siblings(current_lead_id: int, candidate_rows: list[dict[str, Any]]) -> tuple[int, str]:
    items: list[str] = []
    count = 0
    for row in candidate_rows:
        lead_id = _safe_int(row.get("matched_lead_id"))
        if not lead_id or lead_id == current_lead_id:
            continue
        count += 1
        if len(items) >= 5:
            continue
        items.append(
            f"{lead_id}: {_safe_text(row.get('lead_name'))} "
            f"[{_safe_text(row.get('pipeline_name'))} / {_safe_text(row.get('status_name'))}]"
        )
    return count, " | ".join(items)


def _instruction_rows_rop() -> list[dict[str, str]]:
    return [
        {
            "Раздел": "Цель",
            "Содержание": "Этот пакет показывает закрытые сделки за последние 30 дней, которые AI считает потенциально преждевременно закрытыми или требующими follow-up. Сделки не изменялись автоматически.",
        },
        {
            "Раздел": "Reopen",
            "Содержание": "Сначала разбирайте сделки с verdict = reopen_recommended и риском critical/high. Это самый вероятный резерв выручки.",
        },
        {
            "Раздел": "Follow-up",
            "Содержание": "Это сделки, где нужен мягкий возврат в диалог или альтернативный оффер. Здесь не всегда нужен немедленный reopen.",
        },
        {
            "Раздел": "Manual review",
            "Содержание": "Это спорные кейсы: слабый phone match, несколько сделок на контакт, конфликт сигналов. Их не надо автоматически писать в AMO без проверки.",
        },
        {
            "Раздел": "Что смотреть",
            "Содержание": "Приоритетные поля: AI-риск, AI-вердикт, Основание, Следующий шаг, Постзакрытые звонки, UTM, Tallanto context, Другие сделки контакта.",
        },
        {
            "Раздел": "Порядок работы",
            "Содержание": "1) Top priorities, 2) Reopen, 3) Follow-up, 4) Manual review. После ручной валидации можно включать controlled write-back только в service fields сделки.",
        },
        {
            "Раздел": "Ограничение источника",
            "Содержание": "Если live AMO OAuth в момент сборки недоступен, пакет строится из последнего успешного 30-дневного run и текущих локальных таблиц Mango analyse. Tallanto при этом тянется live.",
        },
    ]


def _instruction_rows_manager() -> list[dict[str, str]]:
    return [
        {
            "Раздел": "Как использовать",
            "Содержание": "Сделка уже закрыта в CRM, но AI видит признаки, что клиента еще можно вернуть. Работайте не по памяти, а по Основанию, Краткой истории и Последним звонкам после закрытия.",
        },
        {
            "Раздел": "Reopen",
            "Содержание": "Если AI рекомендует reopen, сначала проверьте, был ли реальный жесткий отказ. Если нет, делайте аккуратный возврат в работу по рекомендованному следующему шагу.",
        },
        {
            "Раздел": "Follow-up",
            "Содержание": "Если verdict = follow_up_needed или alternative_offer_needed, не давите. Используйте мягкое касание, опираясь на причину отсрочки или другой интерес клиента.",
        },
        {
            "Раздел": "Что не делать",
            "Содержание": "Не меняйте статус сделки автоматически только на основании AI. Не игнорируйте поле Постзакрытые звонки: оно часто показывает, что клиент продолжал общение уже после закрытия.",
        },
        {
            "Раздел": "Фиксация результата",
            "Содержание": "После контакта фиксируйте результат в сделке: что именно сказал клиент, актуален ли продукт, нужен ли другой оффер, когда возвращаться к клиенту.",
        },
    ]


def _write_sheet(ws: Any, rows: list[dict[str, Any]], *, title_fill: str = "D9EAF7") -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if not rows:
        rows = [{"Пусто": "Нет данных"}]

    headers = list(rows[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor=title_fill)
    border = Border(bottom=Side(style="thin", color="95A5A6"))

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column in ws.columns:
        max_len = 0
        column_letter = column[0].column_letter
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 80))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[column_letter].width = max(12, min(max_len + 2, 60))


def main() -> None:
    parser = argparse.ArgumentParser(description="Build ROP workbook from amo deal shadow-run results.")
    parser.add_argument("--run-dir", required=True)
    parser.add_argument("--out", required=True)
    parser.add_argument("--master-export-dir", default="")
    parser.add_argument("--lead-snapshot-json", default="")
    parser.add_argument("--full-lead-snapshot-json", default="")
    parser.add_argument("--no-live-tallanto", action="store_true")
    parser.add_argument("--live-open-deals", action="store_true")
    args = parser.parse_args()

    project_root = Path.cwd()
    _bootstrap_runtime_env(project_root)

    from openpyxl import Workbook

    build_live_tallanto_context = None
    if not args.no_live_tallanto:
        from mango_mvp.amocrm_runtime.tallanto_context import build_live_tallanto_context as _build_live_tallanto_context
        build_live_tallanto_context = _build_live_tallanto_context

    run_dir = Path(args.run_dir).expanduser().resolve()
    out_path = Path(args.out).expanduser().resolve()
    master_export_dir = (
        Path(args.master_export_dir).expanduser().resolve()
        if args.master_export_dir
        else _detect_latest_master_export_dir(project_root)
    )
    lead_snapshot_path = _resolve_lead_snapshot_path(project_root, run_dir, args.lead_snapshot_json)
    full_lead_snapshot_path = _resolve_full_lead_snapshot_path(project_root, args.full_lead_snapshot_json, lead_snapshot_path)
    results = json.loads((run_dir / "all_results.json").read_text(encoding="utf-8"))

    master_contacts = _load_master_contacts(master_export_dir)
    master_calls = _load_master_calls(master_export_dir)
    lead_snapshots = _load_lead_snapshot(lead_snapshot_path)
    snapshot_open_leads_by_contact = _index_open_snapshot_leads_by_contact(full_lead_snapshot_path)
    live_open_leads_by_contact = _load_live_open_leads_by_contact(
        {
            _safe_int((row or {}).get("matched_contact_id"))
            for row in results
            if _safe_int((row or {}).get("matched_contact_id"))
            and _safe_text((row or {}).get("close_verdict")).lower()
            in {"reopen_recommended", "follow_up_needed", "closed_too_early", "alternative_offer_needed"}
        },
        enabled=bool(args.live_open_deals),
    )

    count_by_contact = Counter(
        _safe_int((row or {}).get("matched_contact_id"))
        for row in results
        if _safe_int((row or {}).get("matched_contact_id"))
    )
    count_by_phone = Counter(
        _normalize_phone_digits((row or {}).get("phone"))
        for row in results
        if _normalize_phone_digits((row or {}).get("phone"))
    )
    candidate_rows_by_contact: dict[int, list[dict[str, Any]]] = {}
    candidate_rows_by_phone: dict[str, list[dict[str, Any]]] = {}
    for row in results:
        contact_id = _safe_int((row or {}).get("matched_contact_id"))
        if contact_id:
            candidate_rows_by_contact.setdefault(contact_id, []).append(dict(row))
        phone_key = _normalize_phone_digits((row or {}).get("phone"))
        if phone_key:
            candidate_rows_by_phone.setdefault(phone_key, []).append(dict(row))

    tallanto_cache: dict[tuple[str, str, str], dict[str, Any]] = {}

    enriched_rows: list[dict[str, Any]] = []
    for raw_result in results:
        result = dict(raw_result)
        dossier = result.get("dossier") if isinstance(result.get("dossier"), dict) else {}
        contact_rollup = dossier.get("contact_rollup") if isinstance(dossier.get("contact_rollup"), dict) else {}
        call_history = dossier.get("call_history") if isinstance(dossier.get("call_history"), list) else []
        contact = dossier.get("contact") if isinstance(dossier.get("contact"), dict) else {}
        lead = dossier.get("lead") if isinstance(dossier.get("lead"), dict) else {}
        tallanto_live = dossier.get("tallanto_live") if isinstance(dossier.get("tallanto_live"), dict) else {}

        matched_lead_id = _safe_int(result.get("matched_lead_id"))
        matched_contact_id = _safe_int(result.get("matched_contact_id"))
        phone = _safe_text(result.get("phone") or dossier.get("phone"))
        phone_key = _normalize_phone_digits(phone)
        master_contact = master_contacts.get(phone_key, {})
        master_call_rows = master_calls.get(phone_key, [])
        lead_snapshot = lead_snapshots.get(matched_lead_id, {})

        if not matched_contact_id:
            matched_contact_id = _safe_int(lead_snapshot.get("contact_id"))
        contact_name = _safe_text(contact.get("name") or master_contact.get("ФИО родителя") or master_contact.get("ФИО ребенка"))

        current_tallanto_id = _safe_text(result.get("tallanto_id") or contact_rollup.get("tallanto_id") or master_contact.get("ID Tallanto"))
        current_tallanto_status = _safe_text(
            result.get("tallanto_match_status")
            or contact_rollup.get("tallanto_match_status")
            or master_contact.get("Статус матчинга Tallanto")
        )

        verdict = _safe_text(result.get("close_verdict")).lower()
        should_fetch_live_tallanto = verdict in {
            "reopen_recommended",
            "follow_up_needed",
            "closed_too_early",
            "alternative_offer_needed",
        }
        has_exact_tallanto_match = current_tallanto_status in {
            "exact_phone_single",
            "manual_confirmed",
            "id_confirmed",
        }
        if (
            not tallanto_live
            and build_live_tallanto_context
            and should_fetch_live_tallanto
            and (current_tallanto_id or has_exact_tallanto_match)
            and (phone or current_tallanto_id)
        ):
            cache_key = (phone_key, current_tallanto_id, current_tallanto_status)
            if cache_key not in tallanto_cache:
                tallanto_cache[cache_key] = build_live_tallanto_context(
                    phone=phone,
                    tallanto_id=current_tallanto_id or None,
                    tallanto_match_status=current_tallanto_status or None,
                    max_related_records=25,
                )
            tallanto_live = tallanto_cache[cache_key]

        lead_custom = lead.get("custom_fields") if isinstance(lead.get("custom_fields"), dict) else {}
        if not lead_custom:
            lead_custom = dict(lead_snapshot.get("utm") or {})
        loss_reason = _safe_text(
            result.get("loss_reason_summary")
            or lead.get("loss_reason")
            or lead_snapshot.get("loss_reason")
        )

        lead_created_at = result.get("lead_created_at") or lead.get("created_at") or lead_snapshot.get("created_at")
        lead_closed_at = result.get("lead_closed_at") or lead.get("closed_at") or lead_snapshot.get("closed_at")
        lead_updated_at = result.get("lead_updated_at") or lead.get("updated_at") or lead_snapshot.get("updated_at")

        if call_history:
            post_close = _summarize_post_close_calls_from_dossier(result, call_history)
        else:
            post_close = _summarize_post_close_calls_from_master(
                closed_at_value=lead_closed_at,
                master_call_rows=master_call_rows,
            )

        sibling_rows = candidate_rows_by_contact.get(matched_contact_id, []) if matched_contact_id else []
        sibling_count, sibling_preview = _compact_candidate_siblings(matched_lead_id, sibling_rows)
        if not matched_contact_id and phone_key:
            sibling_count, sibling_preview = _compact_candidate_siblings(
                matched_lead_id,
                candidate_rows_by_phone.get(phone_key, []),
            )
        open_leads_source = live_open_leads_by_contact.get(matched_contact_id) or snapshot_open_leads_by_contact.get(
            matched_contact_id,
            [],
        )
        open_leads_count, open_leads_preview = _summarize_open_leads(
            open_leads_source,
            current_lead_id=matched_lead_id,
        )

        tallanto_summary = _summarize_tallanto_live(tallanto_live)
        tallanto_opportunities = _safe_int(tallanto_summary.get("Возможностей Tallanto"))
        sales_temperature = _safe_text(
            result.get("current_sales_temperature")
            or contact_rollup.get("current_sales_temperature")
            or master_contact.get("Приоритет лида")
        )
        history_summary = _prefer_fresh_text(
            master_contact.get("Краткая история общения"),
            contact_rollup.get("history_summary"),
            result.get("history_summary"),
        )
        chronology = _prefer_fresh_text(
            master_contact.get("Хронология общения (последние 5 касаний)"),
            contact_rollup.get("chronology"),
            result.get("chronology"),
        )
        latest_call_summary = _prefer_fresh_text(
            master_contact.get("Краткое резюме последнего свежего звонка"),
            contact_rollup.get("latest_call_summary"),
            result.get("latest_call_summary"),
        )
        recommended_next_step = _safe_text(result.get("recommended_next_step") or master_contact.get("Следующий шаг"))
        follow_up_due_at = result.get("follow_up_due_at") or master_contact.get("Рекомендуемая дата следующего контакта")
        manager_history = result.get("manager_history") or _manager_history_from_master(master_call_rows)
        repeated_calls = _safe_int(result.get("call_count_for_lead") or master_contact.get("Всего звонков в истории"))
        run_same_contact_count = count_by_contact.get(matched_contact_id, 0) if matched_contact_id else count_by_phone.get(phone_key, 0)
        effective_verdict = _safe_text(result.get("close_verdict"))
        effective_risk = _safe_text(result.get("premature_close_risk"))
        effective_reason = _safe_text(result.get("close_reason_summary"))
        if _loss_reason_is_active_client(loss_reason):
            effective_verdict = "closed_valid"
            effective_risk = "no_risk"
            if effective_reason:
                effective_reason = (
                    "Сделка закрыта по причине 'Действующий клиент': клиент продолжает обучение, переоткрытие не требуется. "
                    + effective_reason
                )
            else:
                effective_reason = "Сделка закрыта по причине 'Действующий клиент': клиент продолжает обучение, переоткрытие не требуется."

        enrichment = {
            "sales_temperature": sales_temperature,
            "recommended_next_step": recommended_next_step,
            **post_close,
            "related_leads_count": sibling_count + 1 if (matched_contact_id or phone_key) else 0,
            "other_leads_count": sibling_count,
            "other_leads_preview": sibling_preview,
            "run_same_contact_count": run_same_contact_count,
        }
        priority_score = _priority_score(
            {
                **result,
                "close_verdict": effective_verdict,
                "premature_close_risk": effective_risk,
            },
            enrichment,
        )

        enriched_rows.append(
            {
                "Ссылка на сделку": f"https://educent.amocrm.ru/leads/detail/{matched_lead_id}" if matched_lead_id else "",
                "Ссылка на контакт": f"https://educent.amocrm.ru/contacts/detail/{matched_contact_id}" if matched_contact_id else "",
                "Приоритетный балл": priority_score,
                "AI-риск": effective_risk,
                "AI-вердикт": effective_verdict,
                "Телефон": phone,
                "ID сделки amoCRM": matched_lead_id or "",
                "ID контакта amoCRM": matched_contact_id or "",
                "Сделка": _safe_text(result.get("lead_name") or lead.get("name")),
                "Контакт": contact_name,
                "Краткая история": _clip_excel_text(history_summary),
                "Хронология": _clip_excel_text(chronology),
                "Возражения": _safe_text(result.get("objections_summary") or master_contact.get("Возражения")),
                "Интерес": _safe_text(result.get("interest_summary") or master_contact.get("Продукты интереса")),
                "Следующий шаг": recommended_next_step,
                "Основание": _clip_excel_text(effective_reason),
                "Воронка": _safe_text(result.get("pipeline_name") or lead.get("pipeline_name")),
                "Статус": _safe_text(result.get("status_name") or lead.get("status_name")),
                "Ответственный по сделке": _safe_text(result.get("lead_responsible_user_name") or lead.get("responsible_user_name")),
                "Источник анализа": _safe_text(result.get("analysis_source") or "llm_shadow_legacy"),
                "Режим анализа": _safe_text(result.get("analysis_mode") or "legacy_shadow"),
                "Уверенность LLM": result.get("confidence") or "",
                "Уверенность матчинга": result.get("match_confidence") or "",
                "Разрешена запись в AMO": "Да" if result.get("writeback_allowed") else "Нет",
                "Блокеры записи в AMO": " | ".join(result.get("writeback_blockers") or []),
                "Дата создания сделки": _fmt_dt(lead_created_at),
                "Дата закрытия сделки": _fmt_dt(lead_closed_at),
                "Дата обновления сделки": _fmt_dt(lead_updated_at),
                "Дата последнего звонка": _fmt_dt(result.get("last_call_at") or master_contact.get("Последний звонок")),
                "Количество звонков по сделке": repeated_calls,
                "Всего звонков по контакту": _safe_int(contact_rollup.get("total_calls_history") or master_contact.get("Всего звонков в истории")),
                "Постзакрытых звонков": enrichment["post_close_calls_count"],
                "Последний постзакрытый звонок": enrichment["post_close_last_at"],
                "История после закрытия": _clip_excel_text(enrichment["post_close_calls_summary"]),
                "Сделок контакта в этой очереди": enrichment["run_same_contact_count"],
                "Других сделок у контакта в очереди": enrichment["other_leads_count"],
                "Другие сделки контакта в очереди": enrichment["other_leads_preview"],
                "Открытых сделок у контакта": open_leads_count,
                "Открытые сделки контакта": _clip_excel_text(open_leads_preview),
                "Возможностей Tallanto": tallanto_opportunities,
                "Есть несколько сделок на контакт": "Да" if enrichment["run_same_contact_count"] > 1 or tallanto_opportunities > 1 or open_leads_count > 0 else "Нет",
                "Температура продаж": sales_temperature,
                "Дата следующего касания": _fmt_date(follow_up_due_at),
                "Причина отказа (лид)": loss_reason,
                "Менеджеры в истории": " | ".join(manager_history or []),
                "Последний тип звонка": _safe_text(result.get("latest_call_type") or master_contact.get("Тип последнего свежего звонка")),
                "Последнее резюме звонка": _clip_excel_text(latest_call_summary),
                "UTM источник": _safe_text(lead_custom.get("utm_source")),
                "UTM канал": _safe_text(lead_custom.get("utm_medium")),
                "UTM кампания": _safe_text(lead_custom.get("utm_campaign")),
                "UTM контент": _safe_text(lead_custom.get("utm_content")),
                "UTM ключ": _safe_text(lead_custom.get("utm_term")),
                "UTM реферер": _safe_text(lead_custom.get("utm_referrer")),
                "Источник лида": _safe_text(lead_snapshot.get("source_name")),
                "ID Tallanto": current_tallanto_id,
                "Статус матчинга Tallanto": current_tallanto_status,
                **tallanto_summary,
                "ФИО родителя Tallanto": _safe_text(contact_rollup.get("tallanto_parent_fio") or master_contact.get("ФИО родителя Tallanto")),
                "Контакт Tallanto snapshot": _safe_text(contact_rollup.get("tallanto_contact") or master_contact.get("Контакт Tallanto")),
                "Ответственный Tallanto snapshot": _safe_text(contact_rollup.get("tallanto_owner") or master_contact.get("Ответственный Tallanto")),
                "Тип ученика Tallanto": _safe_text(contact_rollup.get("tallanto_student_type") or master_contact.get("Тип ученика Tallanto")),
                "Филиал Tallanto snapshot": _safe_text(contact_rollup.get("tallanto_branch") or master_contact.get("Филиал Tallanto")),
            }
        )

    reopen_rows = [row for row in enriched_rows if row.get("AI-вердикт") == "reopen_recommended"]
    follow_rows = [
        row
        for row in enriched_rows
        if row.get("AI-вердикт") in {"closed_too_early", "follow_up_needed", "alternative_offer_needed"}
    ]
    manual_rows = [row for row in enriched_rows if row.get("AI-вердикт") == "manual_review"]

    def sort_key(row: dict[str, Any]) -> tuple[Any, ...]:
        risk_rank = {"critical": 4, "high": 3, "medium": 2, "low": 1}.get(_safe_text(row.get("AI-риск")).lower(), 0)
        return (row.get("Приоритетный балл") or 0, risk_rank, row.get("Постзакрытых звонков") or 0)

    reopen_rows.sort(key=sort_key, reverse=True)
    follow_rows.sort(key=sort_key, reverse=True)
    manual_rows.sort(key=sort_key, reverse=True)
    top_rows = sorted(
        [row for row in enriched_rows if row.get("AI-вердикт") != "closed_valid"],
        key=sort_key,
        reverse=True,
    )[:60]

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)

    for title, rows, fill in (
        ("Top priorities", top_rows, "FCE4D6"),
        ("Reopen", reopen_rows, "FDE9D9"),
        ("Follow-up", follow_rows, "E2F0D9"),
        ("Manual review", manual_rows, "FFF2CC"),
        ("Инструкция для РОПа", _instruction_rows_rop(), "D9EAF7"),
        ("Инструкция для менеджера", _instruction_rows_manager(), "D9EAF7"),
    ):
        ws = workbook.create_sheet(title=title)
        _write_sheet(ws, rows, title_fill=fill)

    workbook.save(out_path)

    summary = {
        "run_dir": str(run_dir),
        "out": str(out_path),
        "master_export_dir": str(master_export_dir),
        "lead_snapshot_json": str(lead_snapshot_path),
        "full_lead_snapshot_json": str(full_lead_snapshot_path),
        "all_results": len(enriched_rows),
        "reopen": len(reopen_rows),
        "follow_up": len(follow_rows),
        "manual_review": len(manual_rows),
        "top_priorities": len(top_rows),
        "live_tallanto_enabled": not args.no_live_tallanto,
        "live_open_deals_enabled": bool(args.live_open_deals),
    }
    (out_path.with_suffix(".json")).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
