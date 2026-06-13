from __future__ import annotations

import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

from scripts.build_tz19_calls_review_table import BuildConfig, build_calls_review_table, load_review_rows


def test_load_review_rows_excludes_tail_and_masks_sensitive_text(tmp_path: Path) -> None:
    db = seed_calls_db(tmp_path / "calls.sqlite")
    blacklist = tmp_path / "blacklist_77.txt"
    blacklist.write_text("101\n", encoding="utf-8")
    tail = tmp_path / "manifest.json"
    tail.write_text(json.dumps({"calls": [{"canonical_call_id": 103}]}), encoding="utf-8")

    rows, summary = load_review_rows(
        BuildConfig(
            db=db,
            out=tmp_path / "review.xlsx",
            blacklist=blacklist,
            tail_manifest=tail,
        )
    )

    assert summary["current_v7_total_in_db"] == 3
    assert summary["rows"] == 2
    assert [row["ID звонка"] for row in rows] == [101, 102]
    assert rows[0]["blacklist-77"] is True
    assert rows[0]["Resolve класс"] == "нашёл одного"
    assert rows[1]["Resolve класс"] == "неоднозначно"
    raw_text = json.dumps(rows, ensure_ascii=False)
    assert "+79990000000" not in raw_text
    assert "test@example.com" not in raw_text
    assert "Иван Петров" not in raw_text
    assert "[email]" in raw_text
    assert "[phone]" in raw_text or "[masked]" in raw_text


def test_build_calls_review_table_writes_expected_sheets(tmp_path: Path) -> None:
    db = seed_calls_db(tmp_path / "calls.sqlite")
    blacklist = tmp_path / "blacklist_77.txt"
    blacklist.write_text("101\n", encoding="utf-8")
    tail = tmp_path / "manifest.json"
    tail.write_text(json.dumps({"calls": [{"canonical_call_id": 103}]}), encoding="utf-8")
    out = tmp_path / "review.xlsx"
    summary_out = tmp_path / "review.summary.json"

    summary = build_calls_review_table(
        BuildConfig(
            db=db,
            out=out,
            blacklist=blacklist,
            tail_manifest=tail,
            long_threshold=10_000,
            summary_out=summary_out,
        )
    )

    workbook = load_workbook(out, read_only=False)
    assert workbook.sheetnames == ["Все", "blacklist-77", "длинные"]
    assert workbook["Все"].max_row == 3
    assert workbook["blacklist-77"].max_row == 2
    assert workbook["длинные"].max_row == 2
    assert workbook["Все"].freeze_panes == "A2"
    assert summary["rows"] == 2
    assert summary["blacklist_77_rows"] == 1
    assert summary["long_rows"] == 1
    assert summary_out.exists()


def seed_calls_db(path: Path) -> Path:
    con = sqlite3.connect(path)
    con.execute(
        """
        CREATE TABLE canonical_calls (
          canonical_call_id INTEGER PRIMARY KEY,
          started_at TEXT,
          duration_sec REAL,
          transcript_chars INTEGER,
          phone TEXT,
          resolve_status TEXT,
          candidate_count INTEGER,
          resolve_json TEXT,
          amocrm_contact_id INTEGER,
          amocrm_lead_id INTEGER,
          analysis_json TEXT
        )
        """
    )
    con.executemany(
        "INSERT INTO canonical_calls VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        [
            (
                101,
                "2026-06-01 10:00:00",
                120.5,
                12001,
                "+79990000000",
                "done",
                1,
                "{}",
                111,
                222,
                analysis_json("v7", summary="Иван Петров звонил с +79990000000, email test@example.com"),
            ),
            (
                102,
                "2026-06-02 10:00:00",
                60.0,
                5000,
                "+79990000001",
                "done",
                2,
                "{}",
                None,
                None,
                analysis_json("v7", summary="Обсуждали математику"),
            ),
            (
                103,
                "2026-06-03 10:00:00",
                61.0,
                5001,
                "+79990000002",
                "done",
                1,
                "{}",
                None,
                None,
                analysis_json("v7", summary="Tail row"),
            ),
            (
                104,
                "2026-06-04 10:00:00",
                62.0,
                5002,
                "+79990000003",
                "done",
                1,
                "{}",
                None,
                None,
                analysis_json("v6", summary="Old row"),
            ),
        ],
    )
    con.commit()
    con.close()
    return path


def analysis_json(prompt_version: str, *, summary: str) -> str:
    return json.dumps(
        {
            "analysis_meta": {"analysis_prompt_version": prompt_version},
            "summary": summary,
            "structured_fields": {
                "people": {"child_fio": "Иван Петров"},
                "student": {"grade_current": "8"},
                "interests": {"subjects": ["математика"], "format": ["онлайн"], "products": ["годовые курсы"]},
                "commercial": {},
                "next_step": {"action": "Позвонить"},
                "objections": ["цена"],
                "lead_priority": "warm",
            },
            "interests": ["годовые курсы", "математика"],
            "next_step": "Позвонить",
            "objections": ["цена"],
            "target_product": "годовые курсы",
            "student_grade": "8",
            "tags": ["foton"],
            "quality_flags": {"call_type": "sales_call"},
        },
        ensure_ascii=False,
    )
