from __future__ import annotations

import hashlib
import json
import os
import shutil
import sqlite3
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import pytest

from tests.conftest import dual_strict_source, ready_capture_proof
from openpyxl import Workbook, load_workbook

from scripts import export_daily_mango_calls_resolve as exporter
from scripts import evaluate_dialogue_quality as dialogue_quality
from mango_mvp.productization.mango_calls_service_contract import (
    STAGE10_SCHEMA,
    approved_runtime_fingerprint,
)
from mango_mvp.productization.ready_publication import (
    commit_ready_generation,
    inspect_ready_publication,
)


SCHEMA = """
CREATE TABLE call_records (
    id INTEGER PRIMARY KEY,
    source_file TEXT,
    source_filename TEXT,
    source_call_id TEXT,
    duration_sec REAL,
    phone TEXT,
    manager_name TEXT,
    direction TEXT,
    started_at TEXT,
    transcription_status TEXT,
    resolve_status TEXT,
    analysis_status TEXT,
    transcript_text TEXT,
    transcript_variants_json TEXT,
    resolve_json TEXT,
    analysis_json TEXT
)
"""


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _interrupt_ready_publication(ready_db: Path) -> None:
    staged = ready_db.parent / "synthetic-next.sqlite"
    shutil.copy2(ready_db, staged)
    staged.chmod(0o600)
    with sqlite3.connect(staged) as con:
        con.execute("CREATE TABLE synthetic_publication_marker(value INTEGER)")
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update(
        ready_db=str(ready_db),
        sha256=_sha(staged),
        size_bytes=staged.stat().st_size,
        ready_mtime_ns=staged.stat().st_mtime_ns,
        published_at=datetime.now(timezone.utc).isoformat(),
    )

    def crash(stage: str) -> None:
        if stage == "db_replaced":
            raise RuntimeError("synthetic ready publication crash")

    with pytest.raises(RuntimeError, match="synthetic ready publication crash"):
        commit_ready_generation(ready_db, staged, manifest, checkpoint=crash)


def _analysis() -> dict:
    return {
        "history_summary": "Клиент обсудил летний лагерь и попросил договор.",
        "structured_fields": {
            "people": {"parent_fio": "Анна Иванова", "child_fio": "Пётр"},
            "student": {"grade_current": "7", "school": "Лицей"},
            "interests": {"products": ["летний лагерь"], "subjects": ["математика"], "format": ["очно"], "exam_targets": []},
            "commercial": {"budget": "100000 рублей", "price_sensitivity": "medium", "discount_interest": True},
            "contacts": {"preferred_channel": "phone"},
            "objections": ["нужно обсудить договор"],
            "next_step": {"action": "Отправить договор", "due": "завтра"},
        },
        "quality_flags": {"call_type": "sales_call", "transcript_quality_requires_manual_review": False},
        "needs_review": False,
    }


def _insert(db: Path, *, pending: bool = False, call_id: str = "call-ready", started: str = "2026-07-28 08:00:00", audio: Path) -> None:
    dialogue = [
        "[00:01.0] Менеджер (Коршунова Анастасия): Здравствуйте, Анна Иванова.",
        "[00:02.0] Клиент: Добрый день. Ищу сыну Петру, он в седьмом классе, очный летний лагерь с математикой.",
        "[00:03.0] Клиент: Бюджет около ста тысяч рублей, цена важна. Есть скидка? Сначала нужно обсудить договор.",
        "[00:04.0] Клиент: Отправьте договор и свяжитесь со мной завтра по телефону.",
        "[00:05.0] Менеджер (Коршунова Анастасия): Хорошо, отправлю договор и позвоню завтра.",
    ]
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT INTO call_records VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                2 if pending else 1,
                str(audio),
                audio.name,
                call_id,
                125.0,
                "+79990001122" if not pending else "+79990001123",
                "19",
                "inbound" if not pending else "outbound",
                started,
                "done",
                "pending" if pending else "done",
                "pending" if pending else "done",
                "MANAGER:\nЗдравствуйте, Анна Иванова. Хорошо, отправлю договор и позвоню завтра.\nCLIENT:\nДобрый день. Ищу сыну Петру, он в седьмом классе, очный летний лагерь с математикой. Бюджет около ста тысяч рублей, цена важна. Есть скидка? Сначала нужно обсудить договор. Отправьте договор и свяжитесь со мной завтра по телефону.",
                json.dumps({
                    "mode": "stereo",
                    "primary_provider": "mlx",
                    "secondary_provider": "gigaam",
                    "dialogue_lines": dialogue,
                    "call_topology": "simple_two_party",
                    "role_mapping": {
                        "confirmed": True,
                        "manager_quality_allowed": True,
                        "topology": "simple_two_party",
                    },
                    "manager": {
                        "physical_channel": "left",
                        "variant_a": "Здравствуйте. Отправлю договор.",
                        "variant_b": "Здравствуйте. Отправлю договор.",
                    },
                    "client": {
                        "physical_channel": "right",
                        "variant_a": "Нужен лагерь. Пришлите договор.",
                        "variant_b": "Нужен лагерь. Пришлите договор.",
                    },
                }, ensure_ascii=False),
                "{}" if pending else json.dumps({"decision": "automatic"}, ensure_ascii=False),
                "{}" if pending else json.dumps(_analysis(), ensure_ascii=False),
            ),
        )


def _fixture(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> tuple[Path, Path, Path, Path, Path]:
    root = tmp_path / "pipeline"
    audio_root = root / "working/audio"
    audio_root.mkdir(parents=True)
    ready_audio, pending_audio = audio_root / "ready.mp3", audio_root / "pending.mp3"
    ready_audio.write_bytes(b"ready-audio")
    pending_audio.write_bytes(b"pending-audio")
    ready_db, working_db = root / "drop/ready.sqlite", root / "working/working.sqlite"
    ready_db.parent.mkdir(parents=True)
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            con.execute(SCHEMA)
    _insert(ready_db, audio=ready_audio)
    _insert(working_db, audio=ready_audio)
    _insert(working_db, pending=True, call_id="call-pending", audio=pending_audio)
    manifest = {
        "status": "ready",
        "quick_check": "ok",
        "sha256": _sha(ready_db),
        "size_bytes": ready_db.stat().st_size,
        "published_at": "2026-07-29T00:00:00Z",
    }
    ready_db.with_suffix(".manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    users = tmp_path / "users.json"
    users.write_text(json.dumps({"users": [{"general": {"name": "Коршунова Анастасия"}, "telephony": {"extension": "19"}}]}, ensure_ascii=False), encoding="utf-8")
    tallanto = tmp_path / "tallanto.csv"
    tallanto.write_text(
        'ID,Имя,Фамилия,ФИО родителя,Тел. (родителя),Тел. (доп.)\n1,Пётр,Иванов,Анна Иванова,+79990001122,\n',
        encoding="utf-8-sig",
    )
    monkeypatch.setattr(exporter, "PIPELINE_ROOT", root)
    return ready_db, working_db, users, tallanto, tmp_path / "out"


def _seal_ready(
    ready_db: Path,
    *,
    ready_count: int,
    mango_count: int | None = None,
    quarantine_count: int = 0,
) -> Path:
    day = "2026-07-28"
    mango_count = ready_count if mango_count is None else mango_count
    source = {
        "mode": "strict_service",
        "since": "2026-07-27T21:00:00+00:00",
        "rolling_since": "2026-07-27T21:00:00+00:00",
        "until": "2026-07-28T21:00:00+00:00",
        "cursor": "not_applicable_stats_request_result",
        "pages": None,
        "pagination": "not_applicable_stats_request_result",
        "requests": 1,
        "covered_intervals": [
            {
                "since": "2026-07-27T21:00:00+00:00",
                "until": "2026-07-28T21:00:00+00:00",
                "result_complete": True,
                "rows": ready_count,
                "scope": "rolling_authority",
            }
        ],
        "catch_up": False,
    }
    synthetic_call_keys = [f"mango-{index + 1}" for index in range(mango_count)]
    source = dual_strict_source(
        source,
        call_keys=synthetic_call_keys,
        calls_by_day={day: synthetic_call_keys},
    )
    verdict = {
        "schema_version": STAGE10_SCHEMA,
        "day": day,
        "generated_at": "2026-07-29T00:00:00+00:00",
        "mango_enumeration_complete": True,
        "mango_enumeration_source": source,
        "mango_unique": mango_count,
        "ready_unique": ready_count,
        "quarantine_unique": quarantine_count,
        "quarantine_items": [
            {
                "call_key": f"quarantine-{index + 1}",
                "started_at": "2026-07-28T12:00:00+00:00",
                "code": "recording_retry_expired",
                "reason": "Аудиозапись не появилась в Mango в течение 72 часов.",
                "action": (
                    "Проверить запись в Mango и повторить загрузку вручную, "
                    "если файл появился."
                ),
            }
            for index in range(quarantine_count)
        ],
        "pending_unique": 0,
        "unexplained_missing": 0,
        "state_overlap_count": 0,
        "pending_awaiting_recording": 0,
        "pending_over_sla": 0,
        "quarantine_without_reason": 0,
        "ready_without_dual_asr_or_explicit_exception": 0,
        "ready_without_resolve": 0,
        "ready_without_analyze": 0,
        "duplicate_call_keys": 0,
        "oldest_pending_age_minutes": 0,
        "state_not_in_mango_enumeration": 0,
        "independent_zero_enumerations": 0,
        "consistency_ok": True,
        "closure_ok": True,
    }
    capture_proof, capture_proof_sha256 = ready_capture_proof(
        source,
        zero_by_day={day: 0},
    )
    manifest = {
        "schema_version": "mango_calls_ready_v3",
        "created_at_utc": "2026-07-29T00:00:01+00:00",
        "published_at": "2026-07-29T00:00:02+00:00",
        "status": "ready",
        "consistency_ok": True,
        "closure_ok": True,
        "moscow_dates": [day],
        "daily_verdicts": {day: verdict},
        "producer_git_sha": "a" * 40,
        "host_id": "m1-synthetic",
        "run_id": "synthetic-run",
        "mango_window": {"since": source["since"], "until": source["until"]},
        "mango_enumeration_complete": True,
        "mango_enumeration_source": source,
        "capture_proof": capture_proof,
        "capture_proof_sha256": capture_proof_sha256,
        "capture_proof_run_id": source["dual_enumeration"]["proof_run_id"],
        "manifest_snapshot": {"end_offset": 1, "sha256": "b" * 64},
        "provenance_mode": "strict_service",
        "quick_check": "ok",
        "integrity_check": "ok",
        "runtime_fingerprint": approved_runtime_fingerprint(),
        "sha256": _sha(ready_db),
        "size_bytes": ready_db.stat().st_size,
    }
    path = ready_db.with_suffix(".manifest.json")
    path.write_text(json.dumps(manifest), encoding="utf-8")
    return path


class FakeTallantoClient:
    calls = 0

    def get_entry_list(self, **_: object) -> dict[str, object]:
        self.calls += 1
        return {
            "entry_list": [{
                "id": "2", "first_name": "Пётр", "last_name": "Петров",
                "phone_mobile": "+79990001123", "phone_work": "", "date_modified": "2099-01-01 00:00:00",
            }],
            "next_offset": None,
        }


class FailingTallantoClient:
    def get_entry_list(self, **_: object) -> dict[str, object]:
        raise exporter.TallantoApiError("temporary failure")


def test_moscow_calendar_day_has_exact_utc_bounds() -> None:
    assert exporter.day_bounds_utc(date(2026, 7, 28)) == ("2026-07-27 21:00:00", "2026-07-28 21:00:00")


def test_controlled_preview_allows_today_without_tallanto_or_api(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    ready, working, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    _seal_ready(ready, ready_count=1)
    tallanto.unlink()
    real_datetime = exporter.datetime

    class FixedDateTime(real_datetime):
        @classmethod
        def now(cls, tz=None):
            value = real_datetime(2026, 7, 28, 12, tzinfo=timezone.utc)
            return value if tz is None else value.astimezone(tz)

    monkeypatch.setattr(exporter, "datetime", FixedDateTime)

    result = exporter.export_day(
        ready,
        working,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_api_enabled=False,
        controlled_preview=True,
    )
    replay = exporter.export_day(
        ready,
        working,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_api_enabled=False,
        controlled_preview=True,
    )

    assert result["package_status"] == "CONTROLLED_PREVIEW_NOT_FINAL"
    assert result["rows"] == 1
    assert result["tallanto_names_found"] == 0
    assert result["tallanto_freshness"]["mode"] == (
        "client_name_unconfirmed_offline"
    )
    assert Path(result["xlsx"]).is_file()
    assert len(result["transcripts"]) == 1
    assert replay["reused"] is True


def test_daily_export_lock_is_derived_from_exact_ready_database(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready = tmp_path / "pipeline" / "drop" / "ready.sqlite"
    output = tmp_path / "reports"
    monkeypatch.setattr(exporter, "PIPELINE_ROOT", tmp_path / "wrong-global-root")

    with exporter.daily_export_lock(ready, output, date(2026, 7, 28)):
        with pytest.raises(RuntimeError, match="already running"):
            with exporter.daily_export_lock(
                ready, output, date(2026, 7, 28)
            ):
                pass

    assert (ready.parent / ".daily-export-locks").is_dir()
    assert not (tmp_path / "wrong-global-root").exists()


def test_daily_export_rejects_generation_changed_after_coordinator_decision(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    ready_db, working_db, users, _tallanto, out = _fixture(
        tmp_path, monkeypatch
    )
    _seal_ready(ready_db, ready_count=1)

    with pytest.raises(RuntimeError, match="coordinator decision"):
        exporter.export_day(
            ready_db,
            working_db,
            out,
            date(2026, 7, 28),
            users,
            expected_ready_manifest_sha256="0" * 64,
        )

    assert not out.exists()


def test_current_mango_users_override_archived_manager_name(tmp_path: Path) -> None:
    archived = tmp_path / "users.json"
    archived.write_text(json.dumps({"users": [{"general": {"name": "Старое имя"}, "telephony": {"extension": "405"}}]}), encoding="utf-8")
    current = [{"general": {"name": "Новое имя"}, "telephony": {"extension": "405"}}]
    assert exporter.load_manager_map(archived, current)["405"] == "Новое имя"
    assert exporter.manager_name_issue("Ольга") == "В справочнике Mango указано неполное имя менеджера"
    assert exporter.manager_name_issue("Тест 4") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("Тестовый сотрудник") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("Test 4") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("Demo User") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("Демо-аккаунт") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("test_user") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("Служебный аккаунт") == "В Mango указана техническая тестовая учётная запись"
    assert exporter.manager_name_issue("Иван Тестов") == ""
    assert exporter.manager_name_issue("Костин Тест-Мамедов") == ""
    assert exporter.manager_name_issue("Новое имя") == ""


def test_role_blocks_are_not_presented_as_confirmed_chronology(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(exporter, "PIPELINE_ROOT", tmp_path)
    text, confirmed = exporter.ordered_dialogue(tmp_path / "audio/call.mp3", {}, "MANAGER:\nПервый блок\nCLIENT:\nВторой блок")
    assert not confirmed
    assert text.startswith("Порядок реплик не сохранён")
    assert "Менеджер:\nПервый блок" in text and "Клиент:\nВторой блок" in text


def test_neutralize_unconfirmed_roles_handles_case_and_spacing() -> None:
    text = "МЕНЕДЖЕР : Первый\nклиент (Анна) : Второй\nМенеджер (Иван Иванов) : Третий"
    neutral = exporter.neutralize_unconfirmed_roles(text)
    assert neutral.count("Спикер A (роль не подтверждена):") == 2
    assert neutral.count("Спикер B (роль не подтверждена):") == 1
    assert not any(label in neutral.casefold() for label in ("менеджер", "клиент"))


def test_sealed_dialogue_never_reads_mutable_transcript_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(exporter, "PIPELINE_ROOT", tmp_path)
    source = tmp_path / "audio" / "call.mp3"
    external = tmp_path / "working" / "transcripts" / "audio" / "call_text.txt"
    external.parent.mkdir(parents=True)
    external.write_text("[00:01.00] Менеджер: ВНЕШНИЙ ИЗМЕНЯЕМЫЙ ТЕКСТ", encoding="utf-8")

    text, confirmed = exporter.ordered_dialogue(
        source, {}, "MANAGER:\nСохранённый текст\nCLIENT:\nОтвет", allow_file_fallback=False,
    )

    assert confirmed is False and "ВНЕШНИЙ" not in text and "Сохранённый текст" in text


def test_mutable_sidecar_must_match_stored_call_text(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(exporter, "PIPELINE_ROOT", tmp_path)
    source = tmp_path / "audio" / "call.mp3"
    sidecar = tmp_path / "working" / "transcripts" / "audio" / "call_text.txt"
    sidecar.parent.mkdir(parents=True)
    fallback = "MANAGER:\nСохранённый текст.\nCLIENT:\nОтвет клиента."
    sidecar.write_text(
        "[00:01.0] Менеджер: Сохранённый текст.\n[00:02.0] Клиент: Ответ клиента.\n",
        encoding="utf-8",
    )
    text, confirmed = exporter.ordered_dialogue(source, {}, fallback)
    assert not confirmed and exporter.ORDER_WARNING in text and "Ответ клиента" in text

    sidecar.write_text("[00:01.0] Менеджер: Чужой разговор.\n", encoding="utf-8")
    text, confirmed = exporter.ordered_dialogue(source, {}, fallback)
    assert not confirmed and "Чужой разговор" not in text and "Сохранённый текст" in text

    sidecar.write_text("[00:01.0 Менеджер: Непривязанный текст.\n", encoding="utf-8")
    text, confirmed = exporter.ordered_dialogue(source, {}, fallback)
    assert not confirmed and "Непривязанный текст" not in text and "Сохранённый текст" in text

    sidecar.write_text(
        "[00:01.0] Менеджер: Чужой разговор.\n"
        "[00:02.0] Спикер (не определен): Чужая деталь.\n",
        encoding="utf-8",
    )
    text, confirmed = exporter.ordered_dialogue(source, {}, fallback)
    assert not confirmed and "Чужой разговор" not in text and "Сохранённый текст" in text

    sidecar.write_text(
        "[00:01.0] Менеджер: Ответ клиента.\n[00:02.0] Клиент: Сохранённый текст.\n",
        encoding="utf-8",
    )
    _, confirmed = exporter.ordered_dialogue(source, {}, fallback)
    assert not confirmed

    sidecar.write_text(
        "[00:01.0] Менеджер: Сохранённый\n"
        "[00:02.0] Клиент: Ответ\n"
        "[00:03.0] Клиент: клиента.\n"
        "[00:04.0] Менеджер: текст.\n",
        encoding="utf-8",
    )
    _, confirmed = exporter.ordered_dialogue(source, {}, fallback)
    assert not confirmed


@pytest.mark.parametrize("dialogue_source", ["sidecar", "stored"])
def test_role_swapped_dialogue_never_reaches_manager_sheet(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    dialogue_source: str,
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    swapped_lines = [
        "[00:01.0] Менеджер: Ищу сыну Петру очный летний лагерь с математикой. Бюджет около ста тысяч рублей, цена важна. Есть скидка? Сначала нужно обсудить договор. Отправьте договор и свяжитесь со мной завтра по телефону.",
        "[00:02.0] Клиент: Здравствуйте, Анна Иванова. Хорошо, отправлю договор и позвоню завтра.",
    ]
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            raw = con.execute(
                "SELECT transcript_variants_json FROM call_records WHERE source_call_id='call-ready'"
            ).fetchone()[0]
            payload = json.loads(raw)
            if dialogue_source == "sidecar":
                payload.pop("dialogue_lines", None)
            else:
                payload["dialogue_lines"] = swapped_lines
            con.execute(
                "UPDATE call_records SET transcript_variants_json=? WHERE source_call_id='call-ready'",
                (json.dumps(payload, ensure_ascii=False),),
            )
    if dialogue_source == "sidecar":
        sidecar = exporter.PIPELINE_ROOT / "working" / "transcripts" / "audio" / "ready_text.txt"
        sidecar.parent.mkdir(parents=True)
        sidecar.write_text("\n".join(swapped_lines) + "\n", encoding="utf-8")
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    assert wb["Звонки"].max_row == 1
    assert wb["Проблемы данных"].max_row >= 2
    if dialogue_source == "stored":
        problem_text = " ".join(
            str(cell) for row in wb["Проблемы данных"].iter_rows(values_only=True) for cell in row
        )
        assert "Текст с таймкодами не совпадает с итоговой расшифровкой" in problem_text
        assert "Роли менеджера и клиента не подтверждены" not in problem_text
        assert "Спикер A (роль не подтверждена)" in problem_text
    wb.close()


def test_per_role_rewrite_reports_text_alignment_not_false_role_conflict(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            raw = con.execute(
                "SELECT transcript_variants_json FROM call_records WHERE source_call_id='call-ready'"
            ).fetchone()[0]
            payload = json.loads(raw)
            payload["resolve"] = {"mode": "stereo_per_role", "applied": True}
            con.execute(
                "UPDATE call_records SET transcript_text=?, transcript_variants_json=? WHERE source_call_id='call-ready'",
                (
                    "MANAGER:\nЗдравствуйте, Анна Иванова, я слушаю вас. Хорошо, отправлю договор и позвоню завтра.\n\n"
                    "CLIENT:\nИщу сыну Петру очный летний лагерь по математике. Бюджет около ста тысяч рублей, "
                    "цена для меня важна. Есть ли скидка? Сначала нужно обсудить договор. Отправьте его и свяжитесь завтра.",
                    json.dumps(payload, ensure_ascii=False),
                ),
            )
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    problem_text = " ".join(
        str(cell) for row in wb["Проблемы данных"].iter_rows(values_only=True) for cell in row
    )
    assert "Текст с таймкодами не совпадает с итоговой расшифровкой" in problem_text
    assert "Роли менеджера и клиента не подтверждены" not in problem_text
    assert "Спикер A (роль не подтверждена)" in problem_text
    wb.close()


def test_sealed_merge_does_not_open_working_db(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready, working = tmp_path / "ready.sqlite", tmp_path / "working.sqlite"
    calls: list[Path] = []

    def fake_read(path: Path, day: date, *, immutable: bool):
        del day, immutable
        calls.append(path)
        return []

    monkeypatch.setattr(exporter, "read_day", fake_read)
    rows, pending = exporter.merged_day_rows(ready, working, date(2026, 7, 29), {}, sealed_only=True)
    assert rows == [] and pending == 0 and calls == [ready]


def test_estimated_timecodes_are_not_treated_as_confirmed_chronology() -> None:
    lines = ["[~00:01] Менеджер (Иван): Добрый день.", "[~00:05] Клиент: Здравствуйте."]
    _, confirmed = exporter.ordered_dialogue(Path("call.mp3"), {"dialogue_lines": lines}, "")
    assert not confirmed


def test_equal_cross_role_timecodes_are_not_treated_as_confirmed_order() -> None:
    ambiguous = ["[00:10.0] Менеджер: Вопрос.", "[00:10.0] Клиент: Ответ."]
    text, confirmed = exporter.ordered_dialogue(
        Path("call.mp3"), {"dialogue_lines": ambiguous}, "MANAGER:\nНеполный запасной текст",
    )
    assert not confirmed and text.startswith("Порядок реплик не сохранён")
    assert "Вопрос." in text and "Ответ." in text

    same_role = ["[00:10.0] Менеджер: Первая часть.", "[00:10.0] Менеджер: Вторая часть."]
    _, same_role_confirmed = exporter.ordered_dialogue(Path("call.mp3"), {"dialogue_lines": same_role}, "")
    assert same_role_confirmed

    empty = [
        "[00:10.0] Менеджер: \u200c",
        "[00:11.0] Клиент: \u200d",
        "[00:12.0] Менеджер: \ufeff",
    ]
    _, empty_confirmed = exporter.ordered_dialogue(Path("call.mp3"), {"dialogue_lines": empty}, "")
    assert not empty_confirmed


def test_hour_timecode_is_preserved_and_invalid_seconds_are_rejected() -> None:
    for hour_line in (
        "[01:00:00.0] Менеджер: Продолжаем разговор.",
        "[100:00:00.0] Менеджер: Длинная запись.",
    ):
        text, confirmed = exporter.ordered_dialogue(
            Path("call.mp3"), {"dialogue_lines": [hour_line]}, "",
        )
        assert confirmed and text == hour_line

    for invalid_line in (
        "[00:99.9] Клиент: Ошибка секунд.",
        "[01:99:00.0] Клиент: Ошибка минут.",
    ):
        _, invalid_confirmed = exporter.ordered_dialogue(
            Path("call.mp3"), {"dialogue_lines": [invalid_line]}, "",
        )
        assert not invalid_confirmed


def test_dialogue_quality_evaluator_reads_hour_timecodes(tmp_path: Path) -> None:
    transcript = tmp_path / "call_text.txt"
    transcript.write_text(
        "[01:00:00.0] Клиент: Продолжаем разговор.\n"
        "[~01:30:00] Клиент: Приблизительная метка.\n"
        "[100:00:00.0] Менеджер: Длинная запись.\n",
        encoding="utf-8",
    )
    assert dialogue_quality._parse_timed_lines(transcript) == [
        (3600.0, "client", "Продолжаем разговор."),
        (360000.0, "manager", "Длинная запись."),
    ]
    metrics = dialogue_quality.evaluate_text_file(transcript, 24, 0.92)
    assert metrics["approximate_lines"] == 1


def test_dialogue_quality_summary_reports_untrusted_lines(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    (tmp_path / "call_text.txt").write_text(
        "[01:00:00.0] Клиент: Точная строка.\n"
        "[~01:30:00] Клиент: Приблизительная строка.\n"
        "[01:40:00.0] Спикер (не определен): Неизвестная роль.\n"
        "[01:50:00.0 Клиент: Битая строка.\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(sys, "argv", ["evaluate_dialogue_quality.py", "--transcripts-dir", str(tmp_path)])

    assert dialogue_quality.main() == 0
    summary = json.loads(capsys.readouterr().out)["summary"]
    assert summary["approximate_lines_total"] == 1
    assert summary["unknown_speaker_lines_total"] == 1
    assert summary["unparsed_lines_total"] == 1
    assert summary["files_without_timed_format"] == 0


def test_dialogue_quality_separates_files_without_timed_format(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]
) -> None:
    (tmp_path / "call_text.txt").write_text(
        "Менеджер (Иван):\nДобрый день.\n\nКлиент:\nЗдравствуйте.\n",
        encoding="utf-8",
    )
    monkeypatch.setattr(sys, "argv", ["evaluate_dialogue_quality.py", "--transcripts-dir", str(tmp_path)])

    assert dialogue_quality.main() == 0
    summary = json.loads(capsys.readouterr().out)["summary"]
    assert summary["files_without_timed_format"] == 1
    assert summary["unparsed_lines_total"] == 4


def test_persisted_mutable_sidecar_is_never_confirmed() -> None:
    variants = {
        "dialogue_lines_source": "mutable_sidecar",
        "dialogue_lines": [
            "[00:01.0] Менеджер: Первый ответ.",
            "[00:20.0] Клиент: Первый вопрос.",
        ],
        "call_topology": "simple_two_party",
        "role_mapping": {
            "confirmed": True,
            "manager_quality_allowed": True,
            "topology": "simple_two_party",
        },
        "manager": {"physical_channel": "left"},
        "client": {"physical_channel": "right"},
    }
    _, chronology_confirmed = exporter.ordered_dialogue(
        Path("call.mp3"), variants, "MANAGER:\nПервый ответ.\nCLIENT:\nПервый вопрос."
    )
    assert not chronology_confirmed
    assert not exporter.manager_roles_confirmed(variants)


def test_invalid_stored_line_is_preserved_for_manual_review() -> None:
    lines = [
        "[00:01.0] Менеджер: Добрый день.",
        "[00:02.0 Клиент: Важная договорённость.",
    ]
    text, confirmed = exporter.ordered_dialogue(
        Path("call.mp3"), {"dialogue_lines": lines}, "MANAGER:\nЗапасной текст.",
    )
    assert not confirmed and "Важная договорённость" in text


def test_equal_cross_role_timecodes_are_excluded_from_manager_report(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch,
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            raw = con.execute(
                "SELECT transcript_variants_json FROM call_records WHERE source_call_id='call-ready'"
            ).fetchone()[0]
            payload = json.loads(raw)
            payload["dialogue_lines"][1] = "[00:01.0] Клиент: Ответ с неоднозначным порядком."
            con.execute(
                "UPDATE call_records SET transcript_variants_json=? WHERE source_call_id='call-ready'",
                (json.dumps(payload, ensure_ascii=False),),
            )
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users,
        tallanto_export=tallanto, tallanto_client=FakeTallantoClient(),
    )

    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    assert wb["Звонки"].max_row == 1
    assert any(
        "Порядок реплик не подтверждён" in str(cell)
        for row in wb["Проблемы данных"].iter_rows(values_only=True)
        for cell in row
    )
    wb.close()


def test_summary_removes_only_duplicated_technical_preamble() -> None:
    raw = "28.07.2026 12:24 менеджер 202 общался с клиентом. Обсудили лагерь. Приоритет лида: теплый. Итог: Есть согласованный следующий шаг."
    assert exporter.clean_summary(raw) == "Обсудили лагерь. Итог: Есть согласованный следующий шаг."


def test_export_merges_pending_rows_and_preserves_dialogue(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    result = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users,
        tallanto_export=tallanto, tallanto_client=FakeTallantoClient(),
    )
    assert (result["rows"], result["ready_rows"], result["unfinished_rows"], result["working_only_rows"]) == (2, 1, 1, 1)
    assert result["transcripts_copied"] == 2
    assert result["xlsx_sha256"] == _sha(Path(result["xlsx"]))
    assert result["tallanto_match_sources"] == {"выгрузка Tallanto": 1, "Tallanto API": 1}
    files = list(Path(result["transcript_dir"]).glob("*.txt"))
    assert len(files) == 2 and not list(out.rglob("*.mp3"))
    assert all("7999" not in path.name for path in files)

    wb = load_workbook(result["xlsx"], data_only=False)
    assert wb.sheetnames == ["Сводка", "Звонки", "Проблемы данных", "Описание полей"]
    sheet = wb["Звонки"]
    headers = {cell.value: cell.column for cell in sheet[1]}
    transcript = sheet.cell(2, headers["Расшифровка разговора, часть 1"]).value
    assert transcript.index("Менеджер:") < transcript.index("Клиент:") < transcript.rindex("Менеджер:")
    assert "MANAGER:" not in transcript and "CLIENT:" not in transcript
    assert sheet.cell(2, headers["ФИО клиента из Tallanto"]).value == "Анна Иванова"
    assert sheet.cell(2, headers["Предметы"]).value == "математика"
    assert sheet.cell(2, headers["Формат"]).value == "очно"
    assert sheet.cell(2, headers["Класс"]).value == "7"
    assert sheet.cell(2, headers["Озвученный бюджет"]).value == "100000 рублей"
    assert sheet.cell(2, headers["Интерес к скидке"]).value == "Да"
    assert headers["Расшифровка разговора, часть 1"] == headers["Краткое содержание разговора"] + 1
    assert "Статус обработки" not in headers and "Школа" not in headers and "Аудиозапись" not in headers
    phone = sheet.cell(2, headers["Телефон клиента"])
    assert phone.value == "+79990001122" and phone.data_type == "s"
    transcript_link = sheet.cell(2, headers["Файл полной расшифровки"]).hyperlink
    assert transcript_link is not None and not transcript_link.target.startswith("file:")
    assert wb["Проблемы данных"].max_row == 2
    problem_headers = {cell.value: cell.column for cell in wb["Проблемы данных"][1]}
    assert wb["Проблемы данных"].cell(2, problem_headers["ФИО клиента из Tallanto"]).value == "Петров Пётр"
    description = " ".join(str(cell) for row in wb["Описание полей"].iter_rows(values_only=True) for cell in row)
    assert "сверить с полной расшифровкой" in description
    assert "audio" not in json.dumps(result, ensure_ascii=False).casefold()
    wb.close()


def test_missing_second_asr_is_not_manager_ready(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    with sqlite3.connect(ready_db) as con:
        raw = con.execute(
            "SELECT transcript_variants_json FROM call_records "
            "WHERE source_call_id='call-ready'"
        ).fetchone()[0]
        variants = json.loads(raw)
        variants.pop("secondary_provider")
        for role in ("manager", "client"):
            variants[role].pop("variant_b")
        con.execute(
            "UPDATE call_records SET transcript_variants_json=? "
            "WHERE source_call_id='call-ready'",
            (json.dumps(variants, ensure_ascii=False),),
        )
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    assert result["ready_rows"] == 0
    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    problems = " ".join(
        str(cell)
        for row in wb["Проблемы данных"].iter_rows(values_only=True)
        for cell in row
    )
    assert "Вторая расшифровка GigaAM не готова" in problems
    wb.close()


def test_sealed_manifest_cannot_hide_row_without_second_asr(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    manifest_path = _seal_ready(ready_db, ready_count=1)
    with sqlite3.connect(ready_db) as con:
        raw = con.execute(
            "SELECT transcript_variants_json FROM call_records "
            "WHERE source_call_id='call-ready'"
        ).fetchone()[0]
        variants = json.loads(raw)
        variants.pop("secondary_provider")
        for role in ("manager", "client"):
            variants[role].pop("variant_b")
        con.execute(
            "UPDATE call_records SET transcript_variants_json=? "
            "WHERE source_call_id='call-ready'",
            (json.dumps(variants, ensure_ascii=False),),
        )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
        sealed_only=True,
    )

    assert result["ready_rows"] == 0
    assert result["closure_ok"] is False
    assert result["package_status"] == "INCOMPLETE_DO_NOT_USE_AS_FINAL"
    assert Path(result["xlsx"]).name.startswith("НЕПОЛНЫЙ")


def test_cloud_export_rejects_sealed_manifest_with_incomplete_row(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, _ = _fixture(tmp_path, monkeypatch)
    manifest_path = _seal_ready(ready_db, ready_count=1)
    with sqlite3.connect(ready_db) as con:
        raw = con.execute(
            "SELECT transcript_variants_json FROM call_records "
            "WHERE source_call_id='call-ready'"
        ).fetchone()[0]
        variants = json.loads(raw)
        variants.pop("secondary_provider")
        for role in ("manager", "client"):
            variants[role].pop("variant_b")
        con.execute(
            "UPDATE call_records SET transcript_variants_json=? "
            "WHERE source_call_id='call-ready'",
            (json.dumps(variants, ensure_ascii=False),),
        )
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    out = tmp_path / "Yandex.Disk" / "calls"
    evidence_path = tmp_path / "publication-authority.json"
    evidence_path.write_text(
        json.dumps(
            {
                "schema_version": "mango_yandex_publication_authority_v1",
                "private_acl_readback_ok": True,
                "retention_policy_approved": True,
                "confirmation": "PUBLISH_CLOSED_MANGO_DAY",
                "day": "2026-07-28",
                "output_root": str(out.resolve()),
                "source_ready_manifest_sha256": _sha(manifest_path),
                "expires_at": "2099-01-01T00:00:00+00:00",
            }
        ),
        encoding="utf-8",
    )
    evidence_path.chmod(0o600)

    with pytest.raises(RuntimeError, match="только закрытый суточный пакет"):
        exporter.export_day(
            ready_db,
            working_db,
            out,
            date(2026, 7, 28),
            users,
            tallanto_export=tallanto,
            tallanto_client=FakeTallantoClient(),
            sealed_only=True,
            external_publication_evidence=evidence_path,
        )

    assert not out.exists()


def test_cloud_export_fails_closed_when_authority_reader_rejects_file(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, _ = _fixture(tmp_path, monkeypatch)
    _seal_ready(ready_db, ready_count=1)
    out = tmp_path / "Yandex.Disk" / "calls"
    evidence_path = tmp_path / "publication-authority.json"
    evidence_path.write_text("{}", encoding="utf-8")
    evidence_path.chmod(0o600)

    def reject_authority(*_args: object, **_kwargs: object) -> bytes:
        raise RuntimeError("synthetic extended ACL")

    monkeypatch.setattr(exporter, "read_stable_regular_bytes", reject_authority)

    with pytest.raises(RuntimeError, match="доказательство внешней публикации недействительно"):
        exporter.export_day(
            ready_db,
            working_db,
            out,
            date(2026, 7, 28),
            users,
            tallanto_export=tallanto,
            tallanto_client=FakeTallantoClient(),
            sealed_only=True,
            external_publication_evidence=evidence_path,
        )

    assert not out.exists()


def test_incomplete_row_without_specific_issue_remains_visible(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    with sqlite3.connect(ready_db) as con:
        con.execute("UPDATE call_records SET analysis_json='{}' WHERE source_call_id='call-ready'")
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    values = list(wb["Проблемы данных"].iter_rows(values_only=True))
    headers = {value: index for index, value in enumerate(values[0])}
    assert any(row[headers["Телефон клиента"]] == "+79990001122" for row in values[1:])
    wb.close()


def test_daily_export_recovers_interrupted_ready_generation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    _interrupt_ready_publication(ready_db)

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    assert result["rows"] == 2
    assert inspect_ready_publication(ready_db)["recovery_required"] is False


def test_missing_target_day_verdict_can_never_be_final(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    manifest_path = _seal_ready(ready_db, ready_count=1)
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["daily_verdicts"] = {}
    manifest["moscow_dates"] = []
    manifest["closure_ok"] = True
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    assert result["closure_ok"] is False
    assert result["package_status"] == "INCOMPLETE_DO_NOT_USE_AS_FINAL"
    assert Path(result["xlsx"]).name.startswith("НЕПОЛНЫЙ")


def test_repeated_export_reuses_identical_audio(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    xlsx_mtime = Path(first["xlsx"]).stat().st_mtime_ns
    second = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    assert (second["transcripts_copied"], second["transcripts_reused"]) == (0, 2)
    assert second["reused"] is True
    assert Path(second["xlsx"]).stat().st_mtime_ns == xlsx_mtime


def test_export_reuses_unchanged_day_across_ready_manifest_generations(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {
        "tallanto_export": tallanto,
        "tallanto_client": FakeTallantoClient(),
    }
    first = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        **kwargs,
    )
    ready_manifest = ready_db.with_suffix(".manifest.json")
    original_ready_manifest_sha256 = _sha(ready_manifest)
    payload = json.loads(ready_manifest.read_text(encoding="utf-8"))
    payload["enumeration_evidence_sha256"] = "c" * 64
    ready_manifest.write_text(
        json.dumps(payload, ensure_ascii=False),
        encoding="utf-8",
    )
    second_ready_manifest_sha256 = _sha(ready_manifest)

    second = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        **kwargs,
    )
    payload["enumeration_evidence_sha256"] = "d" * 64
    ready_manifest.write_text(
        json.dumps(payload, ensure_ascii=False),
        encoding="utf-8",
    )
    third_ready_manifest_sha256 = _sha(ready_manifest)
    third = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        **kwargs,
    )

    assert first["content_sha256"] == second["content_sha256"] == third[
        "content_sha256"
    ]
    assert second["reused"] is True
    assert third["reused"] is True
    assert second["manifest"] == third["manifest"] == first["manifest"]
    assert second["source_ready_manifest_sha256"] == (
        original_ready_manifest_sha256
    )
    assert third["source_ready_manifest_sha256"] == (
        original_ready_manifest_sha256
    )
    assert second["decision_ready_manifest_sha256"] == (
        second_ready_manifest_sha256
    )
    assert third["decision_ready_manifest_sha256"] == (
        third_ready_manifest_sha256
    )
    assert not list(out.glob("*supplement-*.manifest.json"))


def test_timed_dialogue_without_role_evidence_is_review_only(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            row = con.execute("SELECT transcript_variants_json FROM call_records WHERE source_call_id='call-ready'").fetchone()
            payload = json.loads(row[0])
            payload.pop("role_mapping", None)
            analysis = json.loads(con.execute("SELECT analysis_json FROM call_records WHERE source_call_id='call-ready'").fetchone()[0])
            analysis["structured_fields"]["interests"]["products"] = []
            analysis["target_product"] = "летний курс M9"
            con.execute(
                "UPDATE call_records SET transcript_variants_json=?, analysis_json=? WHERE source_call_id='call-ready'",
                (json.dumps(payload, ensure_ascii=False), json.dumps(analysis, ensure_ascii=False)),
            )
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, tallanto_export=tallanto, tallanto_client=FakeTallantoClient())

    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    assert wb["Звонки"].max_row == 1
    values = list(wb["Проблемы данных"].iter_rows(values_only=True))
    assert any("Роли менеджера и клиента не подтверждены" in str(cell) for row in values for cell in row)
    assert any("Роли не подтверждены; не использовать для оценки сотрудника" in str(cell) for row in values for cell in row)
    headers = {value: index for index, value in enumerate(values[0])}
    row = next(item for item in values[1:] if item[headers["Телефон клиента"]] == "+79990001122")
    assert row[headers["Статус смысловой выжимки"]].startswith("Гипотезы:")
    assert row[headers["Тип звонка по смысловому анализу"]] == "Продажа / подбор обучения"
    assert row[headers["Краткое содержание разговора"]] == "Клиент обсудил летний лагерь и попросил договор."
    assert row[headers["Продукт"]] == "летний курс M9"
    assert row[headers["Возражения и ограничения"]] == "нужно обсудить договор"
    assert row[headers["Следующий шаг"]] == "Отправить договор"
    for column in ("Предметы", "Формат", "Целевые экзамены", "Класс", "Срок следующего шага", "Предпочтительный канал", "Озвученный бюджет", "Чувствительность к цене", "Интерес к скидке"):
        assert row[headers[column]] is None
    transcript = "\n".join(str(row[index] or "") for name, index in headers.items() if str(name).startswith("Расшифровка разговора"))
    assert "Спикер A (роль не подтверждена):" in transcript
    assert "Спикер B (роль не подтверждена):" in transcript
    assert "Менеджер (" not in transcript and "Клиент:" not in transcript
    description = list(wb["Описание полей"].iter_rows(values_only=True))
    assert any("не является фактом для оценки менеджера или KPI" in str(cell) for item in description for cell in item)
    wb.close()


def test_historical_model_speaker_correction_is_review_only(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            raw = con.execute(
                "SELECT transcript_variants_json FROM call_records WHERE source_call_id='call-ready'"
            ).fetchone()[0]
            payload = json.loads(raw)
            payload["dialogue_resolve"] = {"speaker_corrections": 1}
            con.execute(
                "UPDATE call_records SET transcript_variants_json=? WHERE source_call_id='call-ready'",
                (json.dumps(payload, ensure_ascii=False),),
            )
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
    )

    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    assert wb["Звонки"].max_row == 1
    assert any(
        "Роли менеджера и клиента не подтверждены" in str(cell)
        for row in wb["Проблемы данных"].iter_rows(values_only=True)
        for cell in row
    )
    wb.close()


def test_conflicting_topology_is_review_only(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            row = con.execute("SELECT transcript_variants_json FROM call_records WHERE source_call_id='call-ready'").fetchone()
            payload = json.loads(row[0])
            payload["call_topology"] = "conference_or_multi_party"
            con.execute("UPDATE call_records SET transcript_variants_json=? WHERE source_call_id='call-ready'", (json.dumps(payload, ensure_ascii=False),))
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, tallanto_export=tallanto, tallanto_client=FakeTallantoClient())

    assert result["manager_ready_rows"] == 0
    transcript_name = f"call_{hashlib.sha256(b'call-ready').hexdigest()[:20]}.txt"
    transcript = (Path(result["transcript_dir"]) / transcript_name).read_text(encoding="utf-8")
    assert "Менеджер:" not in transcript


def test_call_id_is_hashed_in_transcript_filename(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    sensitive_id = "Иванов+79990001122@example.com"
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
            con.execute("UPDATE call_records SET source_call_id=? WHERE source_call_id='call-ready'", (sensitive_id,))
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    result = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, tallanto_export=tallanto, tallanto_client=FakeTallantoClient())

    names = " ".join(path.name for path in Path(result["transcript_dir"]).glob("*.txt"))
    assert "Иванов" not in names and "79990001122" not in names and "example" not in names
    assert sensitive_id not in Path(result["manifest"]).read_text(encoding="utf-8")


def test_changed_started_at_updates_one_stable_transcript_set(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    names_before = {path.name for path in Path(first["transcript_dir"]).glob("*.txt")}
    with sqlite3.connect(ready_db) as con:
        con.execute("UPDATE call_records SET started_at='2026-07-28 09:01:00' WHERE source_call_id='call-ready'")
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    second = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)

    assert second["reused"] is False
    assert second["transcript_dir"] != first["transcript_dir"]
    assert {path.name for path in Path(second["transcript_dir"]).glob("*.txt")} == names_before


@pytest.mark.parametrize("unexpected", ["unexpected.txt", "unexpected.partial"])
def test_unexpected_transcript_file_blocks_unchanged_reuse(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, unexpected: str) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    (Path(first["transcript_dir"]) / unexpected).write_text("unexpected", encoding="utf-8")

    with pytest.raises(RuntimeError, match="unexpected transcript files"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)


def test_xlsx_failure_publishes_no_transcripts(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    monkeypatch.setattr(exporter, "write_workbook", lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("xlsx failed")))

    with pytest.raises(OSError, match="xlsx failed"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, tallanto_export=tallanto, tallanto_client=FakeTallantoClient())

    assert not list(out.rglob("*.txt")) and not list(out.glob("*.xlsx"))


def test_xlsx_verification_failure_removes_internal_pii_staging(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    monkeypatch.setattr(exporter, "load_workbook", lambda *_args, **_kwargs: (_ for _ in ()).throw(OSError("verify failed")))

    with pytest.raises(OSError, match="verify failed"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, tallanto_export=tallanto, tallanto_client=FakeTallantoClient())

    assert not list(out.rglob("*.xlsx")) and not list(out.rglob("*.txt"))


def test_existing_immutable_transcript_generation_fails_closed_on_corruption(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    next(Path(first["transcript_dir"]).glob("*.txt")).write_text("corrupted", encoding="utf-8")
    with pytest.raises(RuntimeError, match="immutable transcript generation is inconsistent"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)


def test_existing_immutable_xlsx_generation_fails_closed_on_corruption(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    Path(first["xlsx"]).write_bytes(b"corrupted")

    with pytest.raises(RuntimeError, match="immutable XLSX generation is inconsistent"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)


def test_unreferenced_xlsx_generation_is_not_overwritten(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    Path(first["manifest"]).unlink()

    with pytest.raises(RuntimeError, match="unreferenced immutable XLSX"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)


def test_removed_call_uses_new_exact_generation(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    first = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)
    with sqlite3.connect(working_db) as con:
        con.execute("DELETE FROM call_records WHERE source_call_id='call-pending'")

    second = exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users, **kwargs)

    assert second["transcript_dir"] != first["transcript_dir"]
    assert len(list(Path(second["transcript_dir"]).glob("*.txt"))) == 1
    current = json.loads(Path(second["manifest"]).read_text(encoding="utf-8"))
    assert current["transcript_dir"] == Path(second["transcript_dir"]).name


def test_transcript_generation_rename_failure_leaves_no_partial_target(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    target = tmp_path / "generation"
    rows = [{"transcript": "full dialogue", "transcript_file": target / "call_a.txt"}]
    rows[0]["transcript_sha256"] = hashlib.sha256(b"full dialogue\n").hexdigest()
    monkeypatch.setattr(exporter.os, "replace", lambda *_args: (_ for _ in ()).throw(OSError("rename failed")))

    with pytest.raises(OSError, match="rename failed"):
        exporter.publish_transcripts(rows, target)

    assert not target.exists() and not list(tmp_path.glob(".generation.staging-*"))


def test_tallanto_issue_revokes_manager_ready(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    tallanto.write_text('ID,Имя,Фамилия,ФИО родителя,Тел. (родителя),Тел. (доп.)\n', encoding="utf-8-sig")

    result = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users,
        tallanto_export=tallanto, tallanto_client=FakeTallantoClient(),
    )

    assert result["manager_ready_rows"] == 0
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    assert wb["Звонки"].max_row == 1 and wb["Проблемы данных"].max_row == 3
    wb.close()


def test_current_moscow_day_is_rejected_before_reading_sources(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="завершённые сутки"):
        exporter.export_day(tmp_path / "missing-ready", tmp_path / "missing-working", tmp_path / "out", datetime.now(exporter.MOSCOW).date(), None)


def test_ready_database_manifest_mismatch_blocks_export(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, _, out = _fixture(tmp_path, monkeypatch)
    ready_db.write_bytes(ready_db.read_bytes() + b"changed")
    with pytest.raises(RuntimeError, match="контрольный файл"):
        exporter.export_day(ready_db, working_db, out, date(2026, 7, 28), users)


def test_verify_ready_drop_rejects_manifest_swap_during_db_check(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, _working_db, _users, _tallanto, _out = _fixture(
        tmp_path, monkeypatch
    )
    manifest_path = ready_db.with_suffix(".manifest.json")
    real_connect = exporter.sqlite3.connect
    swapped = False

    def swapping_connect(database: object, *args: object, **kwargs: object):
        nonlocal swapped
        if not swapped and isinstance(database, str) and "mode=ro" in database:
            payload = json.loads(manifest_path.read_text(encoding="utf-8"))
            payload["published_at"] = "2026-07-29T00:00:01Z"
            manifest_path.write_text(json.dumps(payload), encoding="utf-8")
            swapped = True
        return real_connect(database, *args, **kwargs)

    monkeypatch.setattr(exporter.sqlite3, "connect", swapping_connect)

    with pytest.raises(RuntimeError, match="поколение изменилось"):
        exporter.verify_ready_drop(ready_db)
    assert swapped


def test_verify_ready_drop_rejects_identical_db_inode_swap(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, _working_db, _users, _tallanto, _out = _fixture(
        tmp_path, monkeypatch
    )
    replacement = tmp_path / "same.sqlite"
    replacement.write_bytes(ready_db.read_bytes())
    real_connect = exporter.sqlite3.connect
    swapped = False

    def swapping_connect(database: object, *args: object, **kwargs: object):
        nonlocal swapped
        if not swapped and isinstance(database, str) and "mode=ro" in database:
            os.replace(replacement, ready_db)
            swapped = True
        return real_connect(database, *args, **kwargs)

    monkeypatch.setattr(exporter.sqlite3, "connect", swapping_connect)

    with pytest.raises(RuntimeError, match="поколение изменилось"):
        exporter.verify_ready_drop(ready_db)
    assert swapped


def test_export_rejects_manifest_only_change_during_projection(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    manifest_path = ready_db.with_suffix(".manifest.json")
    real_apply = exporter.apply_tallanto_names

    def mutate_manifest(*args: object, **kwargs: object):
        result = real_apply(*args, **kwargs)
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        payload["published_at"] = "2026-07-29T00:00:03Z"
        manifest_path.write_text(json.dumps(payload), encoding="utf-8")
        return result

    monkeypatch.setattr(exporter, "apply_tallanto_names", mutate_manifest)

    with pytest.raises(RuntimeError, match="поколение изменилось"):
        exporter.export_day(
            ready_db,
            working_db,
            out,
            date(2026, 7, 28),
            users,
            tallanto_export=tallanto,
            tallanto_client=FakeTallantoClient(),
        )

    assert not list(out.glob("*.xlsx")) and not list(out.glob("*.manifest.json"))


def test_crash_after_xlsx_resumes_across_unrelated_ready_generation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    _seal_ready(ready_db, ready_count=1)
    real_write = exporter.write_private_json

    def crash_before_export_manifest(path: Path, payload: object) -> None:
        if isinstance(payload, dict) and payload.get("schema_version") == exporter.EXPORT_SCHEMA_VERSION:
            raise OSError("synthetic crash after XLSX")
        real_write(path, payload)

    monkeypatch.setattr(exporter, "write_private_json", crash_before_export_manifest)
    kwargs = {
        "tallanto_export": tallanto,
        "tallanto_client": FakeTallantoClient(),
        "sealed_only": True,
    }
    with pytest.raises(OSError, match="synthetic crash"):
        exporter.export_day(
            ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
        )

    xlsx_before = list(out.glob("*.xlsx"))
    assert len(xlsx_before) == 1
    assert len(list(out.glob(".daily_export_*.journal.json"))) == 1
    with sqlite3.connect(ready_db) as con:
        con.execute("CREATE TABLE unrelated_generation(value TEXT)")
    _seal_ready(ready_db, ready_count=1)
    monkeypatch.setattr(exporter, "write_private_json", real_write)

    result = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
    )

    assert Path(result["xlsx"]) == xlsx_before[0]
    assert result["supplement_number"] is None
    assert len(list(out.glob("*.xlsx"))) == 1
    assert not list(out.glob("*supplement-*.manifest.json"))
    assert not list(out.glob(".daily_export_*.journal.json"))


def test_changed_day_content_quarantines_interrupted_generation(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    _seal_ready(ready_db, ready_count=1)
    real_write = exporter.write_private_json

    def crash_before_export_manifest(path: Path, payload: object) -> None:
        if isinstance(payload, dict) and payload.get("schema_version") == exporter.EXPORT_SCHEMA_VERSION:
            raise OSError("synthetic crash after XLSX")
        real_write(path, payload)

    monkeypatch.setattr(exporter, "write_private_json", crash_before_export_manifest)
    kwargs = {
        "tallanto_export": tallanto,
        "tallanto_client": FakeTallantoClient(),
        "sealed_only": True,
    }
    with pytest.raises(OSError, match="synthetic crash"):
        exporter.export_day(
            ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
        )
    old_xlsx = next(out.glob("*.xlsx"))
    with sqlite3.connect(ready_db) as con:
        con.execute("UPDATE call_records SET duration_sec=duration_sec+1")
    _seal_ready(ready_db, ready_count=1)
    monkeypatch.setattr(exporter, "write_private_json", real_write)

    result = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
    )
    quarantines = list(out.glob(".daily_export_*.quarantine.json"))

    assert len(quarantines) == 1
    quarantine = json.loads(quarantines[0].read_text(encoding="utf-8"))
    assert quarantine["status"] == "quarantined_source_content_changed"
    assert quarantine["xlsx"] == old_xlsx.name
    assert Path(result["xlsx"]) != old_xlsx
    assert len(list(out.glob("*.xlsx"))) == 2
    assert not list(out.glob(".daily_export_*.journal.json"))


def test_final_supersedes_immutable_incomplete_package(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {"tallanto_export": tallanto, "tallanto_client": FakeTallantoClient()}
    incomplete = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
    )
    old_paths = [
        Path(incomplete["manifest"]),
        Path(incomplete["xlsx"]),
        *Path(incomplete["transcript_dir"]).glob("*.txt"),
    ]
    old_hashes = {path: _sha(path) for path in old_paths}
    _seal_ready(ready_db, ready_count=1)

    final = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
        sealed_only=True,
    )
    link = final["supersedes_incomplete"]

    assert link["manifest"] == Path(incomplete["manifest"]).name
    assert link["sha256"] == old_hashes[Path(incomplete["manifest"])]
    assert all(_sha(path) == digest for path, digest in old_hashes.items())
    repeated = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
        sealed_only=True,
    )
    assert repeated["reused"] is True
    assert repeated["manifest"] == final["manifest"]


def test_late_closed_call_creates_one_immutable_supplement(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    _seal_ready(ready_db, ready_count=1)
    kwargs = {
        "tallanto_export": tallanto,
        "tallanto_client": FakeTallantoClient(),
        "sealed_only": True,
    }
    base = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
    )
    base_paths = [
        Path(base["manifest"]),
        Path(base["xlsx"]),
        *Path(base["transcript_dir"]).glob("*.txt"),
    ]
    base_hashes = {path: _sha(path) for path in base_paths}
    with sqlite3.connect(working_db) as source:
        pending = source.execute(
            "SELECT * FROM call_records WHERE source_call_id='call-pending'"
        ).fetchone()
    assert pending is not None
    with sqlite3.connect(ready_db) as target:
        target.execute(
            "INSERT INTO call_records VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            pending,
        )
        target.execute(
            "UPDATE call_records SET resolve_status='done', analysis_status='done', "
            "resolve_json=?, analysis_json=? WHERE source_call_id='call-pending'",
            (json.dumps({"decision": "automatic"}), json.dumps(_analysis(), ensure_ascii=False)),
        )
    _seal_ready(ready_db, ready_count=2)

    supplement = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
        sealed_only=True,
    )

    assert supplement["supplement_number"] == 1
    assert supplement["supplement_of"] == Path(base["manifest"]).name
    assert supplement["supplement_of_sha256"] == base_hashes[Path(base["manifest"])]
    assert {
        field: supplement[field]
        for field in (
            "mango_unique",
            "ready_unique",
            "quarantine_unique",
            "pending_unique",
            "unexplained_missing",
        )
    } == {
        "mango_unique": 2,
        "ready_unique": 2,
        "quarantine_unique": 0,
        "pending_unique": 0,
        "unexplained_missing": 0,
    }
    assert all(_sha(path) == digest for path, digest in base_hashes.items())
    repeated = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
        sealed_only=True,
    )
    assert repeated["reused"] is True
    assert repeated["manifest"] == supplement["manifest"]
    assert not list(out.glob("*supplement-2.manifest.json"))

    base_manifest = Path(base["manifest"])
    changed_base = json.loads(base_manifest.read_text(encoding="utf-8"))
    changed_base["generated_at"] = "2099-01-01T00:00:00+00:00"
    base_manifest.write_text(json.dumps(changed_base), encoding="utf-8")
    with pytest.raises(RuntimeError, match="supplement lineage"):
        exporter.export_day(
            ready_db,
            working_db,
            out,
            date(2026, 7, 28),
            users,
            tallanto_export=tallanto,
            tallanto_client=FakeTallantoClient(),
            sealed_only=True,
        )
    assert not list(out.glob("*supplement-2.manifest.json"))


def test_balance_only_change_creates_supplement_with_new_balance(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    kwargs = {
        "tallanto_export": tallanto,
        "tallanto_client": FakeTallantoClient(),
        "sealed_only": True,
    }
    _seal_ready(ready_db, ready_count=1)
    base = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users, **kwargs
    )
    _seal_ready(
        ready_db,
        ready_count=1,
        mango_count=2,
        quarantine_count=1,
    )

    supplement = exporter.export_day(
        ready_db,
        working_db,
        out,
        date(2026, 7, 28),
        users,
        tallanto_export=tallanto,
        tallanto_client=FakeTallantoClient(),
        sealed_only=True,
    )

    assert supplement["supplement_number"] == 1
    assert supplement["content_sha256"] != base["content_sha256"]
    assert supplement["mango_unique"] == 2
    assert supplement["ready_unique"] == 1
    assert supplement["quarantine_unique"] == 1
    assert supplement["quarantine_items"][0]["call_key"] == "quarantine-1"
    wb = load_workbook(supplement["xlsx"], read_only=True, data_only=True)
    try:
        summary = {
            str(row[0]): row[1]
            for row in wb["Сводка"].iter_rows(values_only=True)
            if row and row[0]
        }
        problems = " ".join(
            str(cell)
            for row in wb["Проблемы данных"].iter_rows(values_only=True)
            for cell in row
        )
    finally:
        wb.close()
    assert "Карантин Stage10" in problems
    assert "Аудиозапись не появилась в Mango в течение 72 часов." in problems
    assert "повторить загрузку вручную" in problems
    assert "+7999" not in problems
    assert summary["Всего звонков"] == 2
    assert summary["Строк с доступными данными"] == 1
    assert summary["Требуют проверки"] >= 1


def test_tallanto_multiple_matches_are_not_selected(tmp_path: Path) -> None:
    export = tmp_path / "tallanto.csv"
    export.write_text(
        'ID,Имя,Фамилия,ФИО родителя,Тел. (родителя),Тел. (доп.)\n1,Иван,Иванов,,+79990001122,\n2,Пётр,Петров,,+79990001122,\n',
        encoding="utf-8-sig",
    )
    rows = [{"phone": "+79990001122", "issues": [], "client_fio": "", "tallanto_source": ""}]
    exporter.apply_tallanto_names(rows, export, FakeTallantoClient())  # type: ignore[arg-type]
    assert rows[0]["client_fio"] == ""
    assert "несколькими карточками" in rows[0]["issues"][0]


def test_tallanto_api_is_loaded_once_for_all_missing_phones(tmp_path: Path) -> None:
    export = tmp_path / "tallanto.csv"
    export.write_text('ID,Имя,Фамилия,ФИО родителя,Тел. (родителя),Тел. (доп.)\n', encoding="utf-8-sig")
    rows = [
        {"phone": "+79990001123", "issues": [], "client_fio": "", "tallanto_source": ""},
        {"phone": "+79990001124", "issues": [], "client_fio": "", "tallanto_source": ""},
    ]
    client = FakeTallantoClient()
    exporter.apply_tallanto_names(rows, export, client)  # type: ignore[arg-type]
    assert client.calls == 1
    assert rows[0]["client_fio"] == "Петров Пётр"
    assert rows[1]["client_fio"] == ""


def test_stale_explicit_tallanto_snapshot_never_calls_api_or_claims_fio(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    export = tmp_path / "tallanto.csv"
    export.write_text('ID,Имя,Фамилия,ФИО родителя,Тел. (родителя),Тел. (доп.)\n', encoding="utf-8-sig")
    rows = [{"phone": "+79990001123", "issues": [], "client_fio": "", "tallanto_source": ""}]
    seen: list[str] = []

    def fake_changes(client: object, wanted: set[str], modified_after: str) -> tuple[dict[str, dict[str, str]], bool]:
        seen.append(modified_after)
        return {}, True

    monkeypatch.setattr(exporter, "load_tallanto_api_changes", fake_changes)
    exporter.apply_tallanto_names(
        rows, export, FakeTallantoClient(),
        snapshot_as_of=datetime.fromisoformat("2026-06-20T00:00:00+03:00"),
    )  # type: ignore[arg-type]

    assert seen == []
    assert rows[0]["tallanto_source"] == "ФИО не подтверждено"
    assert rows[0]["manager_ready"] is False


def test_tallanto_api_failure_is_not_reported_as_phone_absent(tmp_path: Path) -> None:
    export = tmp_path / "tallanto.csv"
    export.write_text('ID,Имя,Фамилия,ФИО родителя,Тел. (родителя),Тел. (доп.)\n', encoding="utf-8-sig")
    rows = [{"phone": "+79990001123", "issues": [], "client_fio": "", "tallanto_source": ""}]
    exporter.apply_tallanto_names(rows, export, FailingTallantoClient())  # type: ignore[arg-type]
    assert rows[0]["issues"] == ["Не удалось проверить телефон через Tallanto API"]


def test_missing_tallanto_export_blocks_matching(tmp_path: Path) -> None:
    with pytest.raises(RuntimeError, match="выгрузка Tallanto не найдена"):
        exporter.apply_tallanto_names([], tmp_path / "missing.csv", FakeTallantoClient())  # type: ignore[arg-type]


def test_long_transcript_is_split_without_loss(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    ready_db, working_db, users, tallanto, out = _fixture(tmp_path, monkeypatch)
    manager_text = "а" * 70_000
    lines = [f"[00:01.0] Менеджер (Иван): {manager_text}", "[00:02.0] Клиент: конец"]
    for db in (ready_db, working_db):
        with sqlite3.connect(db) as con:
                con.execute(
                        "UPDATE call_records SET transcript_text=?, transcript_variants_json=? WHERE source_call_id='call-ready'",
                        (f"MANAGER:\n{manager_text}\n\nCLIENT:\nконец", json.dumps({
                            "mode": "stereo",
                            "primary_provider": "mlx",
                            "secondary_provider": "gigaam",
                            "dialogue_lines": lines,
                        "call_topology": "simple_two_party",
                        "role_mapping": {"confirmed": True, "manager_quality_allowed": True, "topology": "simple_two_party"},
                            "manager": {
                                "physical_channel": "left",
                                "variant_a": manager_text,
                                "variant_b": manager_text,
                            },
                            "client": {
                                "physical_channel": "right",
                                "variant_a": "конец",
                                "variant_b": "конец",
                            },
                    }, ensure_ascii=False),),
                )
    manifest_path = ready_db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest.update({"sha256": _sha(ready_db), "size_bytes": ready_db.stat().st_size})
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    result = exporter.export_day(
        ready_db, working_db, out, date(2026, 7, 28), users,
        tallanto_export=tallanto, tallanto_client=FakeTallantoClient(),
    )
    wb = load_workbook(result["xlsx"], read_only=True, data_only=True)
    sheet = wb["Звонки"]
    header = next(sheet.iter_rows(values_only=True))
    transcript_columns = [i for i, value in enumerate(header) if str(value).startswith("Расшифровка разговора, часть")]
    row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    restored = "".join(str(row[i] or "") for i in transcript_columns)
    expected, confirmed = exporter.ordered_dialogue(Path("ignored"), {"dialogue_lines": lines}, "")
    assert confirmed and restored == expected
    assert all(len(str(row[i] or "")) <= exporter.TRANSCRIPT_CHUNK for i in transcript_columns)
    transcript_name = f"call_{hashlib.sha256(b'call-ready').hexdigest()[:20]}.txt"
    txt = Path(result["transcript_dir"]) / transcript_name
    assert txt.read_text(encoding="utf-8").rstrip() == expected
    wb.close()


def test_excel_formula_prefixes_remain_plain_text(tmp_path: Path) -> None:
    wb = Workbook()
    sheet = wb.active
    values = ["=1+1", "+79990000000", "-10", "@служебное"]
    exporter.append_safe(sheet, values)
    path = tmp_path / "safe.xlsx"
    wb.save(path)
    checked = load_workbook(path, data_only=False)
    cells = list(checked.active[1])
    assert [cell.value for cell in cells] == values
    assert all(cell.data_type == "s" for cell in cells)
    checked.close()
