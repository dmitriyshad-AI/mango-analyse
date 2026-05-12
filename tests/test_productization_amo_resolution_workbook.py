from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from mango_mvp.productization.amo_manual_resolution import build_amo_manual_resolution_pack
from mango_mvp.productization.amo_resolution_workbook import (
    build_amo_resolution_review_html,
    build_amo_resolution_review_workbook,
    export_decisions_from_amo_resolution_workbook,
)
from tests.test_productization_amo_manual_resolution import _fixture_queue


def test_amo_resolution_workbook_builds_operator_template(tmp_path: Path) -> None:
    queue_root, source = _fixture_queue(tmp_path)
    pack_root = tmp_path / "manual_resolution"
    build_amo_manual_resolution_pack(queue_root=queue_root, source_csv=source, out_root=pack_root)
    out_xlsx = pack_root / "resolution_decisions_manual_template.xlsx"

    result = build_amo_resolution_review_workbook(pack_root=pack_root, out_xlsx=out_xlsx)

    assert result["rows"] == 4
    assert out_xlsx.exists()
    wb = load_workbook(out_xlsx, data_only=False)
    assert wb.sheetnames == ["Решения", "AMO-кандидаты", "Инструкция", "Справочник", "Audit"]
    ws = wb["Решения"]
    headers = [cell.value for cell in ws[1]]
    assert "resolution_status" in headers
    assert "validation_hint" in headers
    assert "suggested_contact_link" in headers
    assert ws["F2"].value == "needs_human"
    assert str(ws["L2"].value).startswith("=IF(")
    assert "AA2=\"written\"" in str(ws["L2"].value)
    assert ws["P2"].hyperlink is not None
    candidates = wb["AMO-кандидаты"]
    candidate_headers = [cell.value for cell in candidates[1]]
    assert "candidate_contact_id" in candidate_headers
    assert "operator_hint_ru" in candidate_headers
    assert candidates.max_row >= 5


def test_amo_resolution_html_builds_read_only_review_page(tmp_path: Path) -> None:
    queue_root, source = _fixture_queue(tmp_path)
    pack_root = tmp_path / "manual_resolution"
    build_amo_manual_resolution_pack(queue_root=queue_root, source_csv=source, out_root=pack_root)
    out_html = pack_root / "resolution_review_operator.html"

    result = build_amo_resolution_review_html(pack_root=pack_root, out_html=out_html)

    assert result["review_rows"] == 4
    assert result["candidate_rows"] >= 4
    text = out_html.read_text(encoding="utf-8")
    assert "AMO manual-resolution review" in text
    assert "страница ничего не пишет в AMO" in text
    assert "https://educent.amocrm.ru/contacts/detail/111" in text


def test_amo_resolution_workbook_converts_back_to_decisions_csv(tmp_path: Path) -> None:
    queue_root, source = _fixture_queue(tmp_path)
    pack_root = tmp_path / "manual_resolution"
    build_amo_manual_resolution_pack(queue_root=queue_root, source_csv=source, out_root=pack_root)
    out_xlsx = pack_root / "resolution_decisions_manual_template.xlsx"
    build_amo_resolution_review_workbook(pack_root=pack_root, out_xlsx=out_xlsx)

    out_csv = pack_root / "resolution_decisions_from_xlsx.csv"
    result = export_decisions_from_amo_resolution_workbook(workbook_path=out_xlsx, out_csv=out_csv)

    assert result["rows"] == 4
    rows = list(csv.DictReader(out_csv.open(encoding="utf-8-sig")))
    assert rows[0]["resolution_status"] == "needs_human"
    assert rows[0]["allow_contact_id_outside_source"] == "no"
    assert rows[0]["suggested_resolved_contact_id"] == "111"
