from __future__ import annotations

import csv
import html
import json
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Optional


AMO_DUPLICATE_RESOLUTION_SCHEMA_VERSION = "amo_duplicate_resolution_v1"
DEFAULT_MANUAL_PACK_ROOT = Path("stable_runtime/amo_manual_resolution_20260511_v1")
DEFAULT_OUT_ROOT = Path("stable_runtime/amo_duplicate_resolution_20260511_v1")
DEFAULT_CURRENT_RUNTIME_PATH = Path("stable_runtime/CURRENT_RUNTIME.json")
DEFAULT_AMO_BASE_URL = "https://educent.amocrm.ru"

DUPLICATE_BUCKETS = {
    "needs_manager_review_multi_contact": "duplicate_contacts_merge_required",
    "blocked_contact_id_mismatch": "contact_id_mismatch_requires_operator",
}

QUEUE_COLUMNS = [
    "resolution_id",
    "phone",
    "duplicate_resolution_status",
    "queue_bucket",
    "merge_priority",
    "owner_hint",
    "source_amo_contact_ids",
    "dry_run_contact_ids",
    "suggested_keep_contact_id",
    "all_candidate_contact_ids",
    "contact_links",
    "amo_lead_ids",
    "lead_links",
    "fio_parent",
    "fio_child",
    "last_call_manager",
    "latest_call_date",
    "latest_call_type",
    "priority",
    "sale_probability_percent",
    "next_step",
    "crm_last_call_summary",
    "crm_short_history",
    "crm_timeline",
    "crm_objections",
    "crm_recommended_next_contact_date",
    "operator_instruction_ru",
    "post_merge_recheck_required",
]

CANDIDATE_COLUMNS = [
    "resolution_id",
    "phone",
    "candidate_contact_id",
    "candidate_role",
    "contact_link",
    "in_source_amo_ids",
    "in_dry_run_contact_ids",
    "is_suggested_keep",
    "amo_lead_ids",
    "lead_links",
    "risk_note_ru",
]


def build_amo_duplicate_resolution_pack(
    *,
    manual_pack_root: Path = DEFAULT_MANUAL_PACK_ROOT,
    out_root: Path = DEFAULT_OUT_ROOT,
    current_runtime_path: Path = DEFAULT_CURRENT_RUNTIME_PATH,
    amo_base_url: str = DEFAULT_AMO_BASE_URL,
    generated_at: Optional[datetime] = None,
) -> Mapping[str, Any]:
    """Build a read-only AMO duplicate/contact-mismatch resolution pack.

    The pack intentionally does not resolve contacts automatically and does not
    write to AMO. It extracts only rows where AMO has duplicate exact matches or
    where the dry-run contact differs from the source contact.
    """

    manual_pack_root = manual_pack_root.expanduser().resolve(strict=False)
    out_root = out_root.expanduser().resolve(strict=False)
    current_runtime_path = current_runtime_path.expanduser().resolve(strict=False)
    amo_base_url = amo_base_url.rstrip("/")
    now = generated_at or datetime.now(timezone.utc)
    out_root.mkdir(parents=True, exist_ok=True)

    source_context_by_index, source_context_by_phone = _load_source_context(manual_pack_root)
    source_rows = _read_resolution_rows(manual_pack_root)
    for row in source_rows:
        context = source_context_by_index.get(_safe_text(row.get("source_row_index"))) or source_context_by_phone.get(
            _safe_text(row.get("phone"))
        )
        if context:
            _merge_source_context(row, context)
    duplicate_rows = [row for row in source_rows if _safe_text(row.get("queue_bucket")) in DUPLICATE_BUCKETS]
    queue_rows = [_queue_row(row, amo_base_url) for row in duplicate_rows]
    candidate_rows = [candidate for row in duplicate_rows for candidate in _candidate_rows(row, amo_base_url)]
    recheck_rows = [_source_like_recheck_row(row) for row in duplicate_rows]
    bucket_counts = Counter(row["queue_bucket"] for row in queue_rows)
    status_counts = Counter(row["duplicate_resolution_status"] for row in queue_rows)

    outputs = {
        "duplicate_merge_queue_csv": out_root / "duplicate_merge_queue.csv",
        "candidate_contacts_csv": out_root / "candidate_contacts.csv",
        "post_merge_recheck_input_csv": out_root / "post_merge_recheck_input_ru.csv",
        "review_workbook_xlsx": out_root / "duplicate_merge_review.xlsx",
        "review_html": out_root / "duplicate_merge_review.html",
        "next_recheck_command_sh": out_root / "next_recheck_command.sh",
        "summary_json": out_root / "summary.json",
        "readme_md": out_root / "README.md",
    }
    _write_csv(outputs["duplicate_merge_queue_csv"], queue_rows, QUEUE_COLUMNS)
    _write_csv(outputs["candidate_contacts_csv"], candidate_rows, CANDIDATE_COLUMNS)
    _write_csv(outputs["post_merge_recheck_input_csv"], recheck_rows, _recheck_columns(recheck_rows))
    _write_review_html(outputs["review_html"], queue_rows, candidate_rows, now)
    _write_review_workbook(outputs["review_workbook_xlsx"], queue_rows, candidate_rows, now)

    runtime = _read_json(current_runtime_path)
    paths = _mapping(runtime.get("paths"))
    stage15_summary = _safe_text(paths.get("stage15_summary"))
    crm_quality_summary = _safe_text(paths.get("crm_quality_summary"))
    _write_recheck_command(
        outputs["next_recheck_command_sh"],
        project_root=current_runtime_path.parents[1] if current_runtime_path.name == "CURRENT_RUNTIME.json" else Path.cwd(),
        input_csv=outputs["post_merge_recheck_input_csv"],
        stage15_summary=stage15_summary,
        crm_quality_summary=crm_quality_summary,
        rows=len(recheck_rows),
    )

    summary = {
        "schema_version": AMO_DUPLICATE_RESOLUTION_SCHEMA_VERSION,
        "generated_at": now.isoformat(timespec="seconds"),
        "manual_pack_root": str(manual_pack_root),
        "out_root": str(out_root),
        "review_rows": len(queue_rows),
        "candidate_contact_rows": len(candidate_rows),
        "post_merge_recheck_rows": len(recheck_rows),
        "by_queue_bucket": dict(bucket_counts),
        "by_duplicate_resolution_status": dict(status_counts),
        "outputs": {key: str(path) for key, path in outputs.items()},
        "policy": {
            "live_write_executed": False,
            "write_crm": False,
            "fail_closed": True,
            "duplicate_rows_require_amo_merge_before_writeback": True,
            "post_merge_recheck_required": True,
        },
        "next_actions": _next_actions(queue_rows),
    }
    _write_json(outputs["summary_json"], summary)
    _write_readme(outputs["readme_md"], summary)
    return summary


def _read_resolution_rows(manual_pack_root: Path) -> list[dict[str, str]]:
    template = manual_pack_root / "resolution_template.csv"
    if template.exists():
        return _read_csv(template)
    rows: list[dict[str, str]] = []
    for name in ("needs_human.csv", "already_written_review.csv"):
        rows.extend(_read_csv(manual_pack_root / name))
    return rows


def _queue_row(row: Mapping[str, str], amo_base_url: str) -> dict[str, str]:
    source_ids = _split_ids(row.get("source_amo_contact_ids"))
    dry_ids = _split_ids(row.get("dry_run_contact_ids"))
    suggested = _safe_text(row.get("suggested_resolved_contact_id")) or (source_ids[0] if source_ids else "")
    all_ids = _unique([*source_ids, *dry_ids, suggested, _safe_text(row.get("written_contact_id"))])
    bucket = _safe_text(row.get("queue_bucket"))
    return {
        "resolution_id": _safe_text(row.get("resolution_id")),
        "phone": _safe_text(row.get("phone")),
        "duplicate_resolution_status": DUPLICATE_BUCKETS[bucket],
        "queue_bucket": bucket,
        "merge_priority": _merge_priority(row, all_ids),
        "owner_hint": _owner_hint(row),
        "source_amo_contact_ids": " | ".join(source_ids),
        "dry_run_contact_ids": " | ".join(dry_ids),
        "suggested_keep_contact_id": suggested,
        "all_candidate_contact_ids": " | ".join(all_ids),
        "contact_links": "\n".join(_contact_url(amo_base_url, contact_id) for contact_id in all_ids),
        "amo_lead_ids": _safe_text(row.get("amo_lead_ids")),
        "lead_links": "\n".join(_lead_url(amo_base_url, lead_id) for lead_id in _split_ids(row.get("amo_lead_ids"))),
        "fio_parent": _safe_text(row.get("fio_parent")),
        "fio_child": _safe_text(row.get("fio_child")),
        "last_call_manager": _safe_text(row.get("last_call_manager")),
        "latest_call_date": _safe_text(row.get("latest_call_date")),
        "latest_call_type": _safe_text(row.get("latest_call_type")),
        "priority": _safe_text(row.get("priority")),
        "sale_probability_percent": _safe_text(row.get("sale_probability_percent")),
        "next_step": _safe_text(row.get("next_step")),
        "crm_last_call_summary": _safe_text(row.get("crm_last_call_summary")),
        "crm_short_history": _safe_text(row.get("crm_short_history")),
        "crm_timeline": _safe_text(row.get("crm_timeline")),
        "crm_objections": _safe_text(row.get("crm_objections")),
        "crm_recommended_next_contact_date": _safe_text(row.get("crm_recommended_next_contact_date")),
        "operator_instruction_ru": _operator_instruction(row, all_ids),
        "post_merge_recheck_required": "yes",
    }


def _candidate_rows(row: Mapping[str, str], amo_base_url: str) -> list[dict[str, str]]:
    source_ids = _split_ids(row.get("source_amo_contact_ids"))
    dry_ids = _split_ids(row.get("dry_run_contact_ids"))
    suggested = _safe_text(row.get("suggested_resolved_contact_id")) or (source_ids[0] if source_ids else "")
    all_ids = _unique([*source_ids, *dry_ids, suggested, _safe_text(row.get("written_contact_id"))])
    rows: list[dict[str, str]] = []
    for contact_id in all_ids:
        roles: list[str] = []
        if contact_id in source_ids:
            roles.append("source")
        if contact_id in dry_ids:
            roles.append("dry_run_exact_match")
        if contact_id == suggested:
            roles.append("suggested_keep")
        if contact_id == _safe_text(row.get("written_contact_id")):
            roles.append("already_written")
        rows.append(
            {
                "resolution_id": _safe_text(row.get("resolution_id")),
                "phone": _safe_text(row.get("phone")),
                "candidate_contact_id": contact_id,
                "candidate_role": " | ".join(roles),
                "contact_link": _contact_url(amo_base_url, contact_id),
                "in_source_amo_ids": "yes" if contact_id in source_ids else "no",
                "in_dry_run_contact_ids": "yes" if contact_id in dry_ids else "no",
                "is_suggested_keep": "yes" if contact_id == suggested else "no",
                "amo_lead_ids": _safe_text(row.get("amo_lead_ids")),
                "lead_links": "\n".join(_lead_url(amo_base_url, lead_id) for lead_id in _split_ids(row.get("amo_lead_ids"))),
                "risk_note_ru": _candidate_risk(row, contact_id, source_ids, dry_ids),
            }
        )
    return rows


def _source_like_recheck_row(row: Mapping[str, str]) -> dict[str, str]:
    return {
        "Телефон клиента": _safe_text(row.get("phone")),
        "AMO contact IDs": _safe_text(row.get("source_amo_contact_ids")),
        "AMO lead IDs": _safe_text(row.get("amo_lead_ids")),
        "Тип последнего свежего звонка": _safe_text(row.get("latest_call_type")),
        "Дата последнего свежего звонка": _safe_text(row.get("latest_call_date")),
        "Следующий шаг": _safe_text(row.get("next_step")),
        "Приоритет лида": _safe_text(row.get("priority")),
        "Вероятность продажи, %": _safe_text(row.get("sale_probability_percent")),
        "Готово к записи в AMO": "Да",
        "CRM writeback policy": "duplicate_merge_recheck_only",
        "Причина статуса AMO": "post-merge dry-run recheck input; no live write",
        "Manual resolution id": _safe_text(row.get("resolution_id")),
        "Duplicate resolution status": DUPLICATE_BUCKETS.get(_safe_text(row.get("queue_bucket")), ""),
    }


def _load_source_context(manual_pack_root: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    summary = _read_json(manual_pack_root / "summary.json")
    source_csv = _safe_text(_mapping(summary.get("summary")).get("source_csv"))
    if not source_csv:
        return {}, {}
    rows = _read_csv(Path(source_csv).expanduser().resolve(strict=False))
    by_index = {str(index): row for index, row in enumerate(rows, start=1)}
    by_phone = {_safe_text(row.get("Телефон клиента")): row for row in rows if _safe_text(row.get("Телефон клиента"))}
    return by_index, by_phone


def _merge_source_context(target: dict[str, str], source: Mapping[str, str]) -> None:
    mapping = {
        "last_call_manager": "Менеджер последнего свежего звонка",
        "crm_last_call_summary": "Краткое резюме последнего свежего звонка",
        "crm_short_history": "Краткая история общения",
        "crm_timeline": "Хронология общения (последние 5 касаний)",
        "crm_objections": "Возражения",
        "crm_recommended_next_contact_date": "Рекомендуемая дата следующего контакта",
    }
    for target_key, source_key in mapping.items():
        if not _safe_text(target.get(target_key)):
            target[target_key] = _safe_text(source.get(source_key))


def _merge_priority(row: Mapping[str, str], all_ids: list[str]) -> str:
    probability = _int(row.get("sale_probability_percent"))
    priority = _safe_text(row.get("priority")).casefold()
    if len(all_ids) >= 3 or probability >= 60 or priority in {"hot", "warm"}:
        return "high"
    if probability >= 40:
        return "medium"
    return "low"


def _owner_hint(row: Mapping[str, str]) -> str:
    if _safe_text(row.get("fio_parent")) or _safe_text(row.get("fio_child")):
        return "manager_who_owns_client_context"
    return "amo_operator_or_last_call_manager"


def _operator_instruction(row: Mapping[str, str], all_ids: list[str]) -> str:
    bucket = _safe_text(row.get("queue_bucket"))
    ids = ", ".join(all_ids)
    if bucket == "needs_manager_review_multi_contact":
        return (
            f"В AMO найдено несколько контактов по одному телефону ({ids}). "
            "Сотруднику нужно объединить дубли в AMO, оставить основную карточку со сделками/историей и сообщить, что склейка завершена. "
            "После этого нужно запустить post-merge dry-run recheck."
        )
    return (
        f"Live dry-run нашел contact_id, который расходится с source AMO contact IDs ({ids}). "
        "Нужно открыть обе карточки, определить правильную, объединить дубли при необходимости и только затем запускать recheck."
    )


def _candidate_risk(row: Mapping[str, str], contact_id: str, source_ids: list[str], dry_ids: list[str]) -> str:
    risks: list[str] = []
    if _safe_text(row.get("queue_bucket")) == "blocked_contact_id_mismatch":
        risks.append("source/live contact_id mismatch")
    if contact_id not in source_ids:
        risks.append("candidate outside source_amo_contact_ids")
    if len(dry_ids) > 1:
        risks.append("multiple exact AMO contacts for same phone")
    return " | ".join(risks)


def _next_actions(rows: list[Mapping[str, str]]) -> list[Mapping[str, Any]]:
    duplicate = sum(1 for row in rows if row.get("duplicate_resolution_status") == "duplicate_contacts_merge_required")
    mismatch = sum(1 for row in rows if row.get("duplicate_resolution_status") == "contact_id_mismatch_requires_operator")
    actions: list[Mapping[str, Any]] = []
    if duplicate:
        actions.append(
            {
                "action": "merge_duplicate_amo_contacts",
                "rows": duplicate,
                "owner": "sales_ops_or_responsible_manager",
                "description_ru": "Объединить AMO-дубли по одному телефону, затем запустить post-merge recheck.",
            }
        )
    if mismatch:
        actions.append(
            {
                "action": "resolve_contact_id_mismatch",
                "rows": mismatch,
                "owner": "amo_operator",
                "description_ru": "Сверить source/live AMO contact_id, объединить дубли или выбрать правильную карточку.",
            }
        )
    if not actions:
        actions.append({"action": "no_duplicate_resolution_rows", "rows": 0})
    return actions


def _write_review_html(path: Path, queue_rows: list[Mapping[str, str]], candidate_rows: list[Mapping[str, str]], now: datetime) -> None:
    candidates_by_id: dict[str, list[Mapping[str, str]]] = {}
    for candidate in candidate_rows:
        candidates_by_id.setdefault(_safe_text(candidate.get("resolution_id")), []).append(candidate)
    cards = "\n".join(_render_card(row, candidates_by_id.get(_safe_text(row.get("resolution_id")), [])) for row in queue_rows)
    path.write_text(
        f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>AMO duplicate resolution</title>
  <style>
    body {{ margin:0; padding:28px; background:#f8fafc; color:#101828; font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif; }}
    h1 {{ margin:0 0 8px; }}
    .meta {{ color:#667085; margin-bottom:22px; }}
    .card {{ background:#fff; border:1px solid #d0d5dd; border-radius:16px; padding:18px; margin:0 0 16px; box-shadow:0 8px 24px rgba(16,24,40,.05); }}
    .top {{ display:flex; justify-content:space-between; gap:16px; }}
    .phone {{ font-size:21px; font-weight:700; }}
    .status {{ color:#b42318; font-weight:700; }}
    .grid {{ display:grid; grid-template-columns:repeat(3,minmax(0,1fr)); gap:12px; margin:14px 0; }}
    .box {{ background:#f9fafb; border:1px solid #eaecf0; border-radius:12px; padding:10px; }}
    .label {{ color:#667085; font-size:12px; text-transform:uppercase; letter-spacing:.04em; }}
    .text {{ white-space:pre-wrap; line-height:1.35; }}
    table {{ width:100%; border-collapse:collapse; }}
    th,td {{ border:1px solid #d0d5dd; padding:8px; vertical-align:top; text-align:left; }}
    th {{ background:#7f1d1d; color:white; }}
    .risk {{ background:#fffbeb; color:#92400e; }}
    a {{ color:#0b5cab; }}
    @media(max-width:900px) {{ .grid {{ grid-template-columns:1fr; }} body {{ padding:14px; }} }}
  </style>
</head>
<body>
  <h1>AMO duplicate resolution</h1>
  <div class="meta">Сгенерировано: {_escape(now.isoformat(timespec="seconds"))}. Только read-only пакет; он ничего не пишет в AMO.</div>
  {cards}
</body>
</html>
""",
        encoding="utf-8",
    )


def _render_card(row: Mapping[str, str], candidates: list[Mapping[str, str]]) -> str:
    candidate_html = "\n".join(_render_candidate(candidate) for candidate in candidates)
    return f"""<section class="card">
  <div class="top">
    <div><div class="phone">{_escape(row.get("phone"))}</div><div class="status">{_escape(row.get("duplicate_resolution_status"))}</div></div>
    <div>{_escape(row.get("resolution_id"))}</div>
  </div>
  <div class="grid">
    <div class="box"><div class="label">ФИО</div><div class="text">{_escape(_join_non_empty([row.get("fio_parent"), row.get("fio_child")], " / "))}</div></div>
    <div class="box"><div class="label">Менеджер</div><div class="text">{_escape(row.get("last_call_manager"))}</div></div>
    <div class="box"><div class="label">Сделки</div><div class="text">{_links_text(row.get("lead_links"))}</div></div>
    <div class="box"><div class="label">Следующий шаг</div><div class="text">{_escape(row.get("next_step"))}</div></div>
    <div class="box"><div class="label">Приоритет</div><div class="text">{_escape(row.get("priority"))} / {_escape(row.get("sale_probability_percent"))}%</div></div>
    <div class="box"><div class="label">Последний звонок</div><div class="text">{_escape(row.get("latest_call_date"))} · {_escape(row.get("latest_call_type"))}</div></div>
    <div class="box"><div class="label">Что сделать</div><div class="text">{_escape(row.get("operator_instruction_ru"))}</div></div>
  </div>
  <div class="box"><div class="label">Краткая история</div><div class="text">{_escape(row.get("crm_short_history"))}</div></div>
  <div class="box"><div class="label">Последняя сводка</div><div class="text">{_escape(row.get("crm_last_call_summary"))}</div></div>
  <table>
    <thead><tr><th>AMO contact_id</th><th>Роль</th><th>Ссылка</th><th>Риск</th></tr></thead>
    <tbody>{candidate_html}</tbody>
  </table>
</section>"""


def _render_candidate(candidate: Mapping[str, str]) -> str:
    link = _safe_text(candidate.get("contact_link"))
    return (
        "<tr>"
        f"<td>{_escape(candidate.get('candidate_contact_id'))}</td>"
        f"<td>{_escape(candidate.get('candidate_role'))}</td>"
        f'<td><a href="{_escape(link)}">{_escape(link)}</a></td>'
        f"<td class=\"risk\">{_escape(candidate.get('risk_note_ru'))}</td>"
        "</tr>"
    )


def _write_review_workbook(path: Path, queue_rows: list[Mapping[str, str]], candidate_rows: list[Mapping[str, str]], now: datetime) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Merge queue"
    candidates = wb.create_sheet("Candidate contacts")
    instructions = wb.create_sheet("Инструкция")
    _populate_sheet(ws, QUEUE_COLUMNS, queue_rows, Table, TableStyleInfo, Font, PatternFill, Alignment, "7F1D1D")
    _populate_sheet(candidates, CANDIDATE_COLUMNS, candidate_rows, Table, TableStyleInfo, Font, PatternFill, Alignment, "0F766E")
    instructions["A1"] = "AMO duplicate resolution"
    instructions["A1"].font = Font(size=18, bold=True)
    lines = [
        "1. Откройте все contact_links по телефону.",
        "2. Определите, какие карточки являются дублями одного клиента/семьи.",
        "3. Объедините дубли в AMO штатными средствами или передайте ответственному сотруднику.",
        "4. После склейки запустите next_recheck_command.sh: он делает только dry-run, без live-записи.",
        "5. До зеленого recheck эти строки не должны попадать в live writeback.",
        f"Сгенерировано: {now.isoformat(timespec='seconds')}",
    ]
    for idx, line in enumerate(lines, start=3):
        instructions[f"A{idx}"] = line
    instructions.column_dimensions["A"].width = 120
    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _populate_sheet(
    ws: Any,
    headers: list[str],
    rows: list[Mapping[str, str]],
    Table: Any,
    TableStyleInfo: Any,
    Font: Any,
    PatternFill: Any,
    Alignment: Any,
    header_color: str,
) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    header_fill = PatternFill("solid", fgColor=header_color)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            value = str(cell.value or "")
            if value.startswith("http"):
                cell.hyperlink = value.splitlines()[0]
                cell.style = "Hyperlink"
    if rows:
        ref = f"A1:{_excel_col(len(headers))}{len(rows)+1}"
        table = Table(displayName=f"Table{ws.title.replace(' ', '')}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_excel_col(len(headers))}{max(2, len(rows)+1)}"
    for idx, header in enumerate(headers, start=1):
        width = min(max(len(header) + 4, 14), 48)
        ws.column_dimensions[_excel_col(idx)].width = width


def _write_recheck_command(
    path: Path,
    *,
    project_root: Path,
    input_csv: Path,
    stage15_summary: str,
    crm_quality_summary: str,
    rows: int,
) -> None:
    if not rows:
        body = """#!/usr/bin/env bash
set -euo pipefail
echo "No duplicate/contact-mismatch rows to recheck."
exit 0
"""
    else:
        body = f'''#!/usr/bin/env bash
set -euo pipefail
cd "{project_root}"
echo "Running AMO post-merge dry-run recheck for {rows} rows. No live write flags are used."
PYTHONDONTWRITEBYTECODE=1 UV_CACHE_DIR=/private/tmp/uv-cache uv run \\
  --with pandas --with openpyxl --with xlsxwriter \\
  --with sqlalchemy --with requests --with 'psycopg[binary]' \\
  python scripts/write_amo_ready_contacts.py \\
  --input "{input_csv}" \\
  --expected-dry-run {rows} \\
  --quality-gate-summary "{stage15_summary}" \\
  --crm-writeback-quality-summary "{crm_quality_summary}"
'''
    path.write_text(body, encoding="utf-8")
    path.chmod(0o755)


def _write_readme(path: Path, summary: Mapping[str, Any]) -> None:
    lines = [
        "# AMO duplicate resolution pack",
        "",
        "Этот пакет нужен не для записи в AMO, а для устранения дублей контактов перед следующими writeback-этапами.",
        "",
        f"- Review rows: `{summary.get('review_rows')}`",
        f"- Candidate contact rows: `{summary.get('candidate_contact_rows')}`",
        f"- Recheck rows: `{summary.get('post_merge_recheck_rows')}`",
        "",
        "Основные файлы:",
        "",
        "- `duplicate_merge_queue.csv` — очередь телефонов, где нужно склеить дубли или разобрать mismatch.",
        "- `candidate_contacts.csv` — все AMO contact_id-кандидаты по каждому телефону.",
        "- `duplicate_merge_review.xlsx` — удобная таблица для сотрудников.",
        "- `duplicate_merge_review.html` — read-only карточки со ссылками в AMO.",
        "- `post_merge_recheck_input_ru.csv` — вход для dry-run после склейки дублей.",
        "- `next_recheck_command.sh` — команда только для dry-run recheck; live-write флагов нет.",
        "",
        "Что делает сотрудник:",
        "",
        "1. Открывает `duplicate_merge_review.html` или `duplicate_merge_review.xlsx`.",
        "2. По `contact_links` проверяет все AMO-карточки по телефону.",
        "3. Если это дубли одного клиента/семьи — склеивает их в AMO, оставляя карточку со сделками/историей.",
        "4. Передает release owner: `resolution_id`, `phone`, `surviving_contact_id`, `merged_contact_ids`, `merge_done_by`, `merge_done_at`.",
        "5. Если это не дубли — не склеивает; строка остается `blocked` до отдельного решения.",
        "",
        "Как фиксировать результат после зеленого recheck:",
        "",
        "- `resolution_status=accepted_by_manager`",
        "- `resolved_contact_id=<surviving AMO contact_id>`",
        "- `allow_contact_id_outside_source=no`, если surviving id есть в `source_amo_contact_ids`; иначе `yes` + `outside_source_approved` в reason.",
        "- `resolution_reason=duplicate_merge_completed_post_merge_recheck_approved`",
        "- `resolved_by=<имя сотрудника/оператора>`",
        "- `resolution_notes=merged_contact_ids=...; merge_done_by=...; merge_done_at=...; recheck_run_dir=...`",
        "",
        "`next_recheck_command.sh` должен вернуть dry-run по всем recheck rows. Команда использует `--expected-dry-run`, поэтому частичный/stale recheck не считается успешным.",
        "",
        "Правило: пока post-merge recheck не покажет один корректный AMO contact_id на телефон, эти строки не должны попадать в live writeback.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _recheck_columns(rows: list[Mapping[str, str]]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def _write_csv(path: Path, rows: list[Mapping[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def _read_json(path: Path) -> Mapping[str, Any]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, Mapping) else {}


def _write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def _split_ids(value: Any) -> list[str]:
    text = _safe_text(value)
    if not text:
        return []
    return [part.strip() for part in text.replace("|", ",").replace(";", ",").split(",") if part.strip()]


def _unique(values: list[str]) -> list[str]:
    result: list[str] = []
    for value in values:
        value = _safe_text(value)
        if value and value not in result:
            result.append(value)
    return result


def _contact_url(base_url: str, contact_id: str) -> str:
    return f"{base_url}/contacts/detail/{contact_id}"


def _lead_url(base_url: str, lead_id: str) -> str:
    return f"{base_url}/leads/detail/{lead_id}"


def _links_text(value: Any) -> str:
    links = [_safe_text(part) for part in _safe_text(value).splitlines() if _safe_text(part)]
    return "<br>".join(f'<a href="{_escape(link)}">{_escape(link)}</a>' for link in links)


def _join_non_empty(values: list[Any], sep: str) -> str:
    return sep.join(_safe_text(value) for value in values if _safe_text(value))


def _escape(value: Any) -> str:
    return html.escape(_safe_text(value))


def _safe_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _int(value: Any) -> int:
    try:
        return int(float(_safe_text(value).replace(",", ".")))
    except (TypeError, ValueError):
        return 0


def _excel_col(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


__all__ = [
    "AMO_DUPLICATE_RESOLUTION_SCHEMA_VERSION",
    "build_amo_duplicate_resolution_pack",
]
