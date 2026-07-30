from __future__ import annotations

import json
import re
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from functools import lru_cache
from pathlib import Path
from typing import Any, Mapping, Sequence

from mango_mvp.channels.p0_recall_spec import hard_codes_from_text
from mango_mvp.customer_profile.contracts import has_explicit_brand_conflict
from mango_mvp.customer_timeline.derived_signals import (
    _is_access_event,
    _is_active_deal,
    dedupe_customer_payment_rows as dedupe_family_payment_rows,
)
from mango_mvp.customer_timeline.freshness import (
    MANAGER_REQUIRED_SOURCE_SYSTEMS,
    manager_freshness_gate,
    source_freshness_rows,
)
from mango_mvp.customer_timeline.next_step_resolver import (
    NEXT_STEP_STATUS_ACTIVE,
    NEXT_STEP_STATUS_EMPTY,
    _event_text,
    _is_non_closing_service_event,
    resolve_customer_next_step,
)
from mango_mvp.customer_timeline.purchases import is_explicit_refund_direction
from mango_mvp.customer_timeline.safety import guard_customer_timeline_output_path
from mango_mvp.customer_timeline.bot_safe_summary import customer_summary_brands
from mango_mvp.customer_timeline.store import (
    authoritative_exact_identity_rows,
    customer_entity_ref_values,
    customer_timeline_readonly_uri,
    open_family_identity_conflict_customer_ids,
)
from mango_mvp.knowledge_base.price_axes_catalog import extract_price_query_axes, select_price


MANAGER_DOSSIER_SCHEMA_VERSION = "customer_timeline_manager_dossier_v1"
INTEREST_MARKER_RE = re.compile(r"\b(?:интересу\w*|рассматрива\w*|хот(?:им|им\s+бы|им\s+посмотреть))\b", re.I)
PAIN_MARKER_RE = re.compile(r"\b(?:не\s+успева\w*|сложн\w*|провалил\w*|провал\w*|пережива\w*)\b", re.I)
INTEREST_CONTEXT_RE = re.compile(
    r"\b(?:"
    r"математик\w*|физик\w*|информатик\w*|программировани\w*|русск\w+\s+язык\w*|английск\w+\s+язык\w*|"
    r"егэ|огэ|олимпиад\w*|курс\w*|заняти\w*|групп\w*|лагер\w*|школ\w*|смен\w*|интенсив\w*|"
    r"подготовк\w*|очно\w*|онлайн\w*|выездн\w*|летн\w*|годов\w*"
    r")\b",
    re.I,
)
CONTACT_RE = re.compile(
    r"[\w.+-]+@[\w.-]+\.[a-zа-я]{2,}|"
    r"(?<!\d)(?:(?:\+7|8|7)\s*)?\(?\d{3,4}\)?[\s.-]*\d{2,3}[\s.-]*\d{2}[\s.-]*\d{2}(?!\d)",
    re.I,
)
EMAIL_SUMMARY_REVIEW_NEEDED_RE = re.compile(r"^\s*Требуется\s+ручная\s+проверка\s+модельной\s+выжимки\b", re.I)
WHITESPACE_RE = re.compile(r"\s+")
SPEECH_FILLER_RE = re.compile(
    r"^(?:(?:ну|ээ+|э+|эм+|мм+|вот|значит|как\s+бы|то\s+есть|скажем|короче)\b[\s,.;:–—-]*)+",
    re.I,
)
SPEECH_CLAUSE_BOUNDARY_RE = re.compile(
    r"\s+(?:"
    r"Можете|можете|можно|подскажите|скажите|скиньте|пришлите|"
    r"сколько\s+будет|сч[её]т\s+скинуть|как\s+там|сейчас"
    r")\b",
    re.I,
)
PRODUCT_KEYS = {
    "products_of_interest",
    "product_of_interest",
    "продукты интереса",
    "интересы",
    "interest",
    "interests",
}
MANAGER_OUTREACH_SIGNAL_TYPES = (
    "client_returned",
    "callback_due",
    "deal_stalling",
    "season_return_candidate",
)
MANAGER_OUTREACH_RISK_SIGNAL_TYPES = ("paid_no_access", "duplicate_contact")
MANAGER_KNOWN_BRANDS = frozenset({"foton", "unpk"})
OWNER50_BRAND_ALIASES = {
    "foton": "foton",
    "фотон": "foton",
    "цдпо": "foton",
    "unpk": "unpk",
    "унпк": "unpk",
    "мфти": "unpk",
}
OWNER50_BRAND_RE = {
    "foton": re.compile(r"\b(?:foton|фотон|цдпо)\b", re.I),
    "unpk": re.compile(r"\b(?:unpk|унпк|мфти)\b", re.I),
}
OWNER50_SIGNAL_PRIORITY = {
    "callback_due": 0,
    "client_returned": 0,
    "deal_stalling": 1,
    "season_return_candidate": 2,
    # bug-fix owner50_pravki #2: hot_streak раньше не долетал до Owner50 -- отсутствовал
    # и в этом приоритете (который питает SQL-фильтр типов в _owner50_snapshot), и в
    # candidate_cte. Приоритет ниже остальных: hot_streak -- самый слабый из 5
    # канонических сигналов, ранжируется последним.
    "hot_streak": 3,
}
# Эти лимиты только ловят неконтролируемый рост чтения. Универсум семей всегда
# берётся целиком из family_members_v1, до итогового limit=50.
OWNER50_SIGNAL_SCAN_LIMIT = 100_000
OWNER50_EVENT_SCAN_LIMIT = 250_000
OWNER50_RELATED_SCAN_LIMIT = 100_000
OWNER50_STAFF_TEST_RE = re.compile(
    r"\b(?:staff|employee|test|system|сотрудник\w*|тестов\w*|служебн\w*|системн\w*)\b",
    re.I,
)
OWNER50_STAFF_TEST_TEXT_RE = re.compile(
    r"\b(?:тестов\w*|служебн\w*)\s+(?:клиент\w*|контакт\w*|запис\w*|сделк\w*)"
    r"|\b(?:клиент\w*|контакт\w*|запис\w*|сделк\w*)\s+(?:для\s+)?(?:тест\w*|служебн\w*)"
    r"|\bсотрудник\w*",
    re.I,
)
OWNER50_TEST_SOURCE_RE = re.compile(r"(?:^|[:/_-])(?:test|staff|employee|sandbox|fixture)(?:$|[:/_-])", re.I)
OWNER50_OFFER_GRADE_RE = re.compile(r"(?<!\d)([1-9]|10|11)\s*(?:[-–—]?\s*(?:й|го))?\s*(?:класс|кл\.?)\b", re.I)
OWNER50_GRADUATE_RE = re.compile(r"\bвыпуск\w*", re.I)
OWNER50_NEXT_ACTION = {
    "callback_due": "Связаться по согласованному сроку",
    "client_returned": "Ответить на новый входящий запрос",
    "deal_stalling": "Уточнить решение по активной сделке",
    "season_return_candidate": "Проверить интерес к новому сезону",
    "hot_streak": "Продолжить активный диалог, пока клиент отвечает",
}
OWNER50_REASON_TEXT = {
    "active_access_or_learning": "У семьи уже есть активный доступ или обучение",
    "active_deal_missing": "Сигнал зависшей сделки не подтверждён активной сделкой",
    "brand_ambiguous": "Есть конфликтующие бренды Foton и UNPK",
    "brand_unproven": "Ни профиль, ни уверенная детская связь не подтверждают бренд",
    "child_ambiguous": "Есть неоднозначная связь с ребёнком",
    "child_grade_unproven": "Не подтверждён класс ниже 11-го",
    "contact_missing": "У уверенного члена семьи нет телефона или email",
    "durable_opt_out": "В истории есть явный отказ от контакта",
    "durable_p0_history": "В истории есть P0: возврат, спор, серьёзная жалоба или юридический вопрос",
    "family_ambiguous": "Состав семьи неоднозначен",
    "grade_11_or_graduate": "11 класс или выпускник исключён",
    "identity_not_strong": "У сигнала нет уверенно связанного клиента",
    "meaningful_outbound_after_evidence": "После основания уже был содержательный исходящий контакт",
    "no_active_outreach_signal": "Нет актуального доказуемого повода для контакта",
    "open_identity_conflict": "Есть открытый конфликт идентификации",
    "season_purchase_not_confirmed": "Сезонный возврат не подтверждён оплатой",
    "signal_evidence_ambiguous": "Событие-основание связано неоднозначно",
    "signal_evidence_missing": "Для сигнала нет первичного события, сделки или оплаты",
    "signal_evidence_not_owned": "Событие-основание принадлежит другому клиенту",
    "signal_evidence_superseded": "Событие-основание заменено более новой записью",
    "signal_evidence_text_missing": "У сигнала нет понятного менеджеру основания",
    "staff_test_system": "Тестовая, служебная или сотрудническая запись",
    "structured_no_contact": "В профиле установлен запрет контакта",
    # E5 (26.07): READY требует структурного person-contact origin (AMO-контакт или Tallanto
    # person ID), не только identity_status=="strong" -- см. _owner50_has_person_contact_origin.
    "person_origin_unproven": "Нет структурного происхождения контакта (AMO-контакт или Tallanto ID) -- имя не подтверждено как клиент",
    "human_name_unproven": "Имя человека не подтверждено структурным источником",
}
# требование E2 (26.07): три новых поля в самом конце -- НЕ вставлять в середину, позиционные
# индексы существующих колонок (в т.ч. в тестах) на них полагаются.
OWNER50_REQUIRED_COLUMNS = (
    "Ранг", "family_id", "Бренд", "Кому", "ID контакта", "Телефон", "Email",
    "Канал основания", "Дата основания", "Сигнал", "Почему сейчас",
    "Следующий шаг", "Сигнал действует до", "Предложение", "Ребёнок/класс", "Оплаты",
    "Формула ранга", "Действие одной фразой",
    "ID ребёнка (адресат)", "Ребёнок (адресат)", "Класс (адресат)",
    "Члены семьи",
)
MANAGER_OPTOUT_PHRASES = (
    "не пишите",
    "больше не пишите",
    "перестаньте писать",
    "не звоните",
    "больше не звоните",
    "не надо мне звонить",
    "не беспокойте",
    "не связывайтесь",
    "не связываться",
    "удалите номер",
    "отпишите меня",
    "хочу отписаться",
    "не хочу получать рассылку",
)


# ---------------------------------------------------------------------------
# Owner50: classify_family -- классификация READY/CANDIDATE/EXCLUDED (§3-5
# спеки владельца SPEC_tablitsa_50_semey.md), встроена сюда по запросу
# архитектора вместо отдельного модуля. Чистые функции без побочных эффектов,
# без сети/файлов/БД -- работают с уже посчитанными полями, которые собирает
# _owner50_family_rows ниже. Правки бизнес-аудитора (Fable, OWNER50_pravki_
# Fable_i_bagi.md) внедрены прямо в правила:
#   #2 окно свежести сигнала 30 дней -> CANDIDATE ("stale_signal");
#   #3 "не предлагать уже купленный продукт" -- УБРАНО по итогам ревью: реального
#      источника с НАЗВАНИЯМИ купленных продуктов на customer_id сейчас нет
#      (customer_purchases_v1 несёт только суммы) -- фейковая всегда-пустая защита
#      хуже её отсутствия. Вернуть, когда появится источник на уровне продукта.
#   #4 "все дети выпускники" -- решение владельца 25.07: EXCLUDED ("grade_11_or_graduate"),
#      НЕ CANDIDATE -- отменяет более раннюю правку Fable #4. Код общий с SQL-предфильтром
#      _owner50_family_rows (та же причина на обоих слоях, не два разных кода).
#   #6 нет точной цены -> и раньше, и сейчас только CANDIDATE, никогда EXCLUDED;
#   #1 поле "Действие" одной фразой -- owner50_action_text() ниже (тест 5 секунд);
#   #5 возражение-перевес -- свежее основание перекрывает более старое возражение
#      (ужесточено 26.07, см. блок E3 ниже -- client_returned сам по себе больше не позитив).
# Правка #7 (7 условий -> 3 гейта) СОЗНАТЕЛЬНО не применена: она противоречит
# #2 (стейл-сигнал -- это жёсткий гейт, а #7 просит сделать сигнал/свежесть
# атрибутом сортировки, не гейтом) -- решение оставлено архитектору/Кодексу.
#
# Блок E (аудит 26.07, owner50_review): пять точечных фиксов поверх уже прошедшего аудита.
#   E1 brand_unproven (0 брендов) -- убран из жёсткого SQL-предфильтра _owner50_family_rows
#      (там раньше сразу шёл в EXCLUDED, минуя classify_family) -- теперь долетает сюда и
#      получает CANDIDATE, как и требует Г2. Заодно защищены от того же "unproven читается как
#      ambiguous" два соседних сравнения offer/signal-брендов с пустым brands (см. _owner50_family_rows).
#   E2 продукт/цена/ребёнок -- offer/product для READY только из выбранного price_axes_catalog
#      entry (никогда из старых названий сделок AMO -- это evidence интереса, не предложение);
#      добавлен target_child_key/name/grade конкретного адресата -- _owner50_select_target_child
#      ниже НЕ берёт минимальный класс из истории ребёнка (несколько разных чисел = класс не
#      известен), несколько РАЗНЫХ подходящих детей без доказанного адресата -> CANDIDATE
#      "target_child_ambiguous" (кроме client_returned/callback_due с конкретным next_step --
#      там offer/product и так пустые, адресат ребёнка предложению не нужен).
#   E3 client_returned убран из авто-позитива, перекрывающего возражение -- сам факт нового
#      входящего сообщения не доказывает, что клиент передумал после возражения; перекрыть
#      возражение может только более свежая подтверждённая оплата или доказанная interest_quote.
#   E4 evidence "payment"/"child"/"product" были СТРУКТУРНО нерезолвируемыми (event_id всегда
#      пуст либо known_records никогда не содержал их source) -- теперь резолвятся в реальные
#      customer_purchases_v1/family_links_v1/price_axes_catalog записи, когда это честно
#      возможно (один вклад в оплату, конкретный family_links_v1 child_key, найденный entry_id).
#   E5 READY требует структурного person-contact origin (source_ref вида "amocrm:contact:*"
#      или "tallanto:student:*"), не только identity_status=="strong" -- название сделки/шага
#      без такого происхождения не может стать "именем клиента". Проверка -- по структурному
#      полю record.source_ref, НЕ по смыслу текста display_name (никакого нового regex понимания).
# ---------------------------------------------------------------------------

OWNER50_CANONICAL_SIGNAL_TYPES = (
    "client_returned",
    "callback_due",
    "deal_stalling",
    "season_return_candidate",
    "hot_streak",
)
OWNER50_PRODUCT_OPTIONAL_SIGNALS = frozenset({"client_returned", "callback_due"})
OWNER50_GRADUATE_GRADE_THRESHOLD = 11
OWNER50_INTEREST_QUOTE_MAX_AGE_MONTHS_FOR_READY = 12
OWNER50_AVERAGE_DAYS_PER_MONTH = 30.436875
OWNER50_STALE_SIGNAL_MAX_DAYS = 30  # правка Fable #2: сигнал старше -- уже не "сегодня"
_OWNER50_CLARIFY_INTEREST_PREFIX = "уточнить интерес"

# --- требование архитектора #3: продукт только по точному entry_id актуального
# price_axes_catalog (select_price с brand+grade, полученным из child_key). Каталог --
# уже посчитанный артефакт (product_data/knowledge_base/.../price_axes_catalog.json),
# читаем его как данные, не пересобираем. Формат/период для select_price сначала
# пытаемся вытащить из текста предложения (extract_price_query_axes), иначе перебираем
# канонические комбинации -- ни один "додуманный" формат/период не выбрать честно без
# текста, поэтому берём ПЕРВЫЙ, для которого в каталоге есть однозначная цена; предмет
# спора (subject) сознательно не участвует -- цена регулярных курсов от предмета не
# зависит (см. price_axes_catalog.py: regular_course_price_does_not_depend_on).
OWNER50_PRICE_AXES_CATALOG_DEFAULT_PATH = (
    Path(__file__).resolve().parents[3]
    / "product_data"
    / "knowledge_base"
    / "kb_release_20260612_v6_7_staging_r4_1"
    / "price_axes_catalog.json"
)

# --- требование аудиторов BLOCKED #2 (25.07, отменяет прежнее требование архитектора #6):
# эвристика "mention старше порога И после него было хоть какое-то содержательное событие --
# значит закрыт" УБРАНА -- она ничего не доказывает о РЕЗУЛЬТАТЕ жалобы, только о том, что
# переписка продолжилась. Прощать refund-упоминание теперь может ТОЛЬКО явный СТРУКТУРНЫЙ
# статус резолюции на самой записи (event.record/metadata.resolution_status) -- см.
# _owner50_event_p0_is_stale_and_resolved. Сегодня ни один источник ingestion НЕ пишет такое
# поле -- значит функция всегда возвращает False, и это ЧЕСТНО: "нет статуса -> не READY",
# а не "предположим, что закрыто, раз переписка не оборвалась". Серьёзные коды
# (legal/complaint/payment_dispute) эта функция и раньше не смягчала -- только чистый "refund".

OWNER50_TIER_REASON_TEXT: dict[str, str] = {
    "no_payment_or_interest_evidence": "Нет ни оплаты, ни цитаты интереса не старше 12 месяцев",
    "next_step_missing_or_vague": "Следующий шаг не указан, без срока или слишком общий",
    "product_not_confirmed_by_kb": "Продукт не подтверждён действующим прайсом базы знаний",
    "late_objection_no_fresh_positive": "Последнее возражение свежее любого позитивного основания",
    "active_work_recent_manager_touch": "Менеджер уже ведёт эту сделку недавно — не лезть под руку",
    "capture_stale_beyond_sla": "Данные источника устарели относительно SLA свежести",
    "stale_signal": "Сигнал старше 30 дней — уже не повод писать сегодня",
    "signal_date_invalid": "Дата сигнала отсутствует или записана в неверном формате",
    "signal_date_in_future": "Дата сигнала в будущем — данные недостоверны",
    # E2 (26.07): несколько РАЗНЫХ верифицированных детей подходят по классу, а адресат
    # предложения (кому конкретно продукт) ничем не доказан -- не гадаем, кому из них.
    "target_child_ambiguous": "Несколько детей подходят по классу — кому из них предлагать, не доказано",
    "target_child_unproven": "Не доказан конкретный ребёнок с устойчивым child_key",
}


def owner50_tier_reason_text(code: str) -> str:
    """Человекочитаемый текст для кода причины из classify_family (не найден -- код как есть)."""
    if code.startswith("active_risk_signal:"):
        return f"Активен риск-сигнал {code.split(':', 1)[1]}"
    if code.startswith("classification_error:"):
        return f"Внутренняя ошибка классификации ({code.split(':', 1)[1]}) — строка исключена, не подана как готовая"
    return OWNER50_TIER_REASON_TEXT.get(code, code)


def classify_family(family: Mapping[str, Any], *, as_of: datetime | None = None) -> dict[str, Any]:
    """Классифицирует ОДНУ семью в READY / CANDIDATE / EXCLUDED (спека §3-5).

    Возвращает {"status", "reasons", "missing", "assumptions", "action_text"}:
      - status: "READY" | "CANDIDATE" | "EXCLUDED".
      - reasons: почему именно такой статус (для EXCLUDED -- жёсткие коды блокировки,
        готовы для _owner50_control_rows как есть; для CANDIDATE -- те же коды, что и
        missing; для READY -- пустой tuple).
      - missing: ТОЛЬКО для CANDIDATE -- список непройденных пунктов Г2-Г7. Для READY
        и EXCLUDED всегда ().
      - assumptions: короткие пометки (атрибуция по тегу, пересчитанный класс).
      - action_text: одна императивная фраза "что сделать" из next_step (правка
        Fable #1, "тест 5 секунд") -- пусто, если next_step не задан.

    Схема входного family описана в docstring _owner50_classify_family_unsafe ниже
    (перенесена туда, чтобы не дублировать в двух местах). Падает ЗАКРЫТО: любое
    исключение внутри классификации одной семьи превращается в status=EXCLUDED с
    кодом "classification_error:<тип>", а не роняет весь batch на тысячах семей.
    """
    try:
        return _owner50_classify_family_unsafe(family, as_of=as_of)
    except Exception as exc:  # fail closed: одна плохая семья не должна ронять весь batch
        return {
            "status": "EXCLUDED",
            "reasons": (f"classification_error:{type(exc).__name__}",),
            "missing": (),
            "assumptions": (),
            "action_text": "",
        }


def _owner50_classify_family_unsafe(family: Mapping[str, Any], *, as_of: datetime | None) -> dict[str, Any]:
    """Схема входного `family` (все ключи, кроме family_id, необязательны -- отсутствие
    трактуется КОНСЕРВАТИВНО, "пусто лучше догадки", никогда не в пользу READY):

        family_id: str
        identity: {"customer_id": str, "identity_status": str, "display_name": str}
            Г1: identity_status должен быть буквально "strong".
        brands: set[str] -- уже объединённый набор брендов семьи. Г2: len==1 для READY;
            0 -> CANDIDATE "brand_unproven"; >=2 -> EXCLUDED "brand_ambiguous".
        unrecognized_brand_present: bool -- нераспознанный бренд даёт CANDIDATE "brand_unproven".
        family_conflict: bool -- настоящий конфликт семьи -> EXCLUDED "family_ambiguous".
        children: Sequence[{"child_key", "name", "grade_current"|"grade_recorded"+
            "grade_fixed_at", "is_graduate"}] -- ПО ОДНОЙ записи на ребёнка. Пустой
            список = нет проверенных детей -> не влияет на graduate-логику ниже.
        payment: {"total_in", "total_out", "deals_cnt", "last_purchase_at", "above_median"} | None
            FAMILY-level агрегат, уже прогнанный через dedupe_family_payment_rows.
            total_out -- парное списание с баланса Tallanto, а не доказательство возврата.
            last_purchase_at ОБЯЗАТЕЛЕН (требование архитектора #2/#5) -- без даты или
            с датой в будущем оплата не засчитывается вообще.
        interest_quote: {"text", "quoted_at", "event_id", "source_system"} | None -- Г4.
            Засчитывается ТОЛЬКО если event_id+source_system заданы И (при наличии
            events_by_id) разрешаются в событие с СОВПАДАЮЩИМ source_system, и quoted_at не
            в будущем (требование архитектора #2/#5) -- иначе просто не считается интересом.
        signal: {"signal_type", "created_at", "evidence_text", "event_id", "source_system"} | None -- Г5.
            event_id+source_system -- ТОЛЬКО если разрешаются в реальное события events_by_id
            с СОВПАДАЮЩИМ source_system (требование архитектора #2); иначе signal_ok=False.
        next_step: {"action", "due"} | None -- Г6.
        product: {"name", "brand", "verified", "source", "seats_available", "grade_min", "grade_max"} | None -- Г7.
            source, начинающийся с "kb" (в проде -- "kb_price_axes_catalog:<entry_id>" из
            _owner50_select_price_entry/_owner50_product_from_price_entry, требование #3) --
            ЕДИНСТВЕННЫЙ признанный источник "verified" продукта; seats_available=False --
            единственный явный sold_out, None/True/отсутствие -- места есть по умолчанию.
        last_objection: {"text", "at"} | None. Перекрывается только позитивом СВЕЖЕЕ него --
            позитив = подтверждённая оплата или доказанная свежая interest_quote клиента (Г3).
            Требование E3 (26.07, отменяет более раннее правило): client_returned САМ ПО СЕБЕ
            больше НЕ позитив -- то, что клиент написал снова, не доказывает, что он передумал
            после возражения (мог написать и с новой претензией). callback_due/deal_stalling/
            season_return_candidate тоже не позитив (требование архитектора #7, не менялось).
        target_child_ambiguous: bool -- требование E2 (26.07): несколько РАЗНЫХ верифицированных
            детей одинаково подходят под предложение, а адресат (кому конкретно) ничем не
            доказан -- _owner50_family_rows уже решил не выбирать никого наугад. Флаг НЕ ставится,
            когда продукт всё равно не нужен (client_returned/callback_due с конкретным
            next_step) -- см. её комментарий у _owner50_select_target_child.
        events_by_id: Mapping[event_id, {"source_system", ...}] | None -- проверка Г5-ссылки.
        contact_missing / open_p0 / opt_out / identity_conflict /
        recent_meaningful_outbound_after_evidence / active_recent_manager_work /
        active_risk_signals / stale_data -- EXCLUDED-флаги, 1:1 с существующими кодами
        OWNER50_REASON_TEXT (contact_missing, durable_p0_history, durable_opt_out,
        open_identity_conflict, meaningful_outbound_after_evidence,
        active_work_recent_manager_touch, active_risk_signal:<name>, capture_stale_beyond_sla).
    """
    now = _owner50_resolve_as_of(as_of, family)

    identity = _mapping(family.get("identity"))
    raw_brands = {str(item).strip().casefold() for item in (family.get("brands") or ()) if str(item).strip()}
    brands = {OWNER50_BRAND_ALIASES[item] for item in raw_brands if item in OWNER50_BRAND_ALIASES}
    children = _owner50_dedupe_children(family.get("children") or ())
    payment = _mapping(family.get("payment"))

    exclusions: list[str] = []
    precondition_missing: list[str] = []
    if str(identity.get("identity_status") or "").strip().casefold() != "strong":
        precondition_missing.append("identity_not_strong")
    unrecognized_brand_present = bool(raw_brands - OWNER50_BRAND_ALIASES.keys()) or bool(
        family.get("unrecognized_brand_present")
    )
    if len(brands) > 1:
        exclusions.append("brand_ambiguous")
    if family.get("family_conflict"):
        exclusions.append("family_ambiguous")
    if family.get("contact_missing"):
        exclusions.append("contact_missing")
    if family.get("open_p0"):
        exclusions.append("durable_p0_history")
    if family.get("opt_out"):
        exclusions.append("durable_opt_out")
    if family.get("identity_conflict"):
        exclusions.append("open_identity_conflict")
    if family.get("recent_meaningful_outbound_after_evidence"):
        exclusions.append("meaningful_outbound_after_evidence")
    if family.get("active_recent_manager_work"):
        exclusions.append("active_work_recent_manager_touch")
    if family.get("stale_data"):
        exclusions.append("capture_stale_beyond_sla")
    for risk_signal_type in family.get("active_risk_signals") or ():
        exclusions.append(f"active_risk_signal:{risk_signal_type}")
    # решение владельца 25.07 (отменяет более раннюю правку Fable #4): все верифицированные
    # дети семьи -- 11 класс/выпускники -> EXCLUDED, а не CANDIDATE. Код "grade_11_or_graduate"
    # общий с SQL-предфильтром _owner50_family_rows (та же проверка, та же причина, один код).
    if children and all(_owner50_child_is_graduate(child, now) for child in children):
        exclusions.append("grade_11_or_graduate")

    if exclusions:
        return {
            "status": "EXCLUDED",
            "reasons": tuple(dict.fromkeys(exclusions)),
            "missing": (),
            "assumptions": _owner50_family_assumptions(family, children),
            "action_text": "",
        }

    missing: list[str] = [
        *precondition_missing,
        *(str(code) for code in family.get("missing") or () if str(code)),
    ]

    if len(brands) == 0 or unrecognized_brand_present:
        missing.append("brand_unproven")

    has_payment = _owner50_payment_is_confirmed(payment, as_of=now)
    events_by_id = family.get("events_by_id")
    interest = _mapping(family.get("interest_quote"))
    interest_at = _parse_iso_datetime(interest.get("quoted_at")) if interest else None
    # требование архитектора #2: interest засчитывается ТОЛЬКО при event_id+source_system
    # и разрешении в реальное событие с СОВПАДАЮЩИМ source_system -- цитата "из воздуха"
    # (без привязки к событию) не считается доказательством интереса.
    interest_event_id = _clean_text(interest.get("event_id")) if interest else ""
    interest_source_system = _clean_text(interest.get("source_system")) if interest else ""
    interest_direction = _clean_text(interest.get("direction")).casefold() if interest else ""
    interest_provenance_ok = (
        bool(interest_event_id)
        and bool(interest_source_system)
        and isinstance(events_by_id, Mapping)
        and resolve_evidence_source(interest_event_id, events_by_id) == interest_source_system
    )
    has_fresh_interest = (
        bool(interest)
        and interest_direction == "inbound"
        and interest_provenance_ok
        and interest_at is not None
        and interest_at <= now  # требование архитектора #5: будущая дата -- не "свежо"
        and _owner50_months_since(interest_at, now) <= OWNER50_INTEREST_QUOTE_MAX_AGE_MONTHS_FOR_READY
    )
    if not has_payment and not has_fresh_interest:
        missing.append("no_payment_or_interest_evidence")

    signal = _mapping(family.get("signal"))
    signal_type = str(signal.get("signal_type") or "")
    signal_ok = bool(signal) and signal_type in OWNER50_CANONICAL_SIGNAL_TYPES and bool(_clean_text(signal.get("evidence_text")))
    # требование архитектора #2 + требование аудиторов BLOCKED #1 (fail-open доказательств):
    # сигнал засчитывается ТОЛЬКО при event_id+source_system И разрешении в реальное событие
    # с СОВПАДАЮЩИМ source_system -- симметрично с interest_provenance_ok ниже. РАНЬШЕ пустой
    # event_id вообще ПРОПУСКАЛ эту проверку (signal_ok оставался True) -- это и есть fail-open:
    # "нечего проверить" читалось как "доверяю". Теперь пустой event_id/source_system -- это
    # "не разрешилось", а не "пропускаем проверку".
    if signal_ok:
        signal_event_id = _clean_text(signal.get("event_id"))
        signal_source_system = _clean_text(signal.get("source_system"))
        signal_ok = (
            bool(signal_event_id)
            and bool(signal_source_system)
            and isinstance(events_by_id, Mapping)
            and resolve_evidence_source(signal_event_id, events_by_id) == signal_source_system
        )
    signal_created_at = _parse_iso_datetime(signal.get("created_at")) if signal else None
    if not signal_ok:
        missing.append("no_active_outreach_signal")
    elif signal_created_at is None:
        missing.append("signal_date_invalid")
    elif signal_created_at > now:
        # требование архитектора #5 (ужесточено по итогам ревью): _owner50_days_since
        # клэмпит отрицательную разницу к 0.0 -- без этой явной проверки сигнал с датой в
        # будущем читался бы как "0 дней назад", то есть максимально свежий. Будущее -- не
        # свежее, это ошибка данных; отдельная причина, не путать со "старым" сигналом.
        missing.append("signal_date_in_future")
    elif signal_created_at is not None and _owner50_days_since(signal_created_at, now) > OWNER50_STALE_SIGNAL_MAX_DAYS:
        # правка Fable #2: устаревший сигнал (>30 дней) -- уже не повод писать "сегодня".
        missing.append("stale_signal")

    next_step = _mapping(family.get("next_step"))
    next_step_ok = (
        _owner50_is_concrete_next_step(str(next_step.get("action") or ""))
        and _parse_iso_datetime(next_step.get("due")) is not None
    )
    if not next_step_ok:
        missing.append("next_step_missing_or_vague")

    product = _mapping(family.get("product"))
    # ponytail: защита "не предлагать уже купленный продукт" убрана здесь -- см. блок
    # комментариев в начале файла (#3). Вернуть вместе с реальным источником названий
    # купленных продуктов, не как фейковую всегда-пустую проверку.
    product_confirmed = _owner50_product_confirmed(product, brands, children, now)
    next_step_is_clarify_interest = _clean_text(next_step.get("action")).casefold().startswith(_OWNER50_CLARIFY_INTEREST_PREFIX)
    product_waived = signal_type in OWNER50_PRODUCT_OPTIONAL_SIGNALS and next_step_ok
    if not product_confirmed and not next_step_is_clarify_interest and not product_waived:
        # правка Fable #6: нет точной цены -- НЕ повод исключать совсем, только CANDIDATE
        # (этот код всегда идёт в missing, никогда в exclusions).
        missing.append("product_not_confirmed_by_kb")
    if family.get("target_child_ambiguous"):
        # требование E2 (26.07): несколько РАЗНЫХ детей подходят, адресат не доказан -- как и
        # Даже когда продукт не нужен, READY должен называть конкретного ребёнка.
        missing.append("target_child_ambiguous")
    eligible_child_keys = {
        _clean_text(child.get("child_key"))
        for child in children
        if _clean_text(child.get("child_key"))
        and _clean_text(child.get("name"))
        and not _owner50_child_is_graduate(child, now)
    }
    if not family.get("target_child_ambiguous") and len(eligible_child_keys) != 1:
        missing.append("target_child_unproven")

    last_objection = _mapping(family.get("last_objection"))
    if last_objection:
        objection_at = _parse_iso_datetime(last_objection.get("at"))
        # требование E3 (26.07, отменяет более раннее правило #7 в части client_returned):
        # client_returned убран из позитива -- сам факт нового входящего сообщения не
        # доказывает, что клиент передумал после возражения (мог написать и с новой
        # претензией). Позитив, способный перекрыть возражение, -- ТОЛЬКО подтверждённая
        # оплата или доказанная свежая interest_quote клиента. callback_due/deal_stalling/
        # season_return_candidate по-прежнему не позитив (требование архитектора #7).
        positive_candidates = [
            _parse_iso_datetime(payment.get("last_purchase_at")) if has_payment else None,
            interest_at if has_fresh_interest else None,
        ]
        latest_positive = max((value for value in positive_candidates if value is not None), default=None)
        # правка Fable #5 (перевес): позитив СВЕЖЕЕ возражения -- не блокирует; блокирует,
        # только если возражение свежее (или нет позитива вообще).
        if objection_at is not None and (latest_positive is None or objection_at > latest_positive):
            missing.append("late_objection_no_fresh_positive")

    status = "READY" if not missing else "CANDIDATE"
    return {
        "status": status,
        "reasons": tuple(missing),
        "missing": tuple(missing),
        "assumptions": _owner50_family_assumptions(family, children),
        "action_text": owner50_action_text(family),
    }


def build_evidence_record(
    *,
    kind: str,
    text: str,
    event_id: str,
    source_system: str,
    at: Any,
    known_records: Mapping[str, Mapping[str, Any]] | None = None,
    require_at: bool = True,
) -> dict[str, Any]:
    """Нормализует одно доказательство в формат "дата+source_system+event_id" (ТЗ владельца:
    "каждое доказательство с датой+source_system+event_id"). Никогда не бросает исключение --
    если чего-то не хватает, resolvable=False и missing_fields перечисляет, чего именно.

    Требование аудиторов BLOCKED #1 (fail-open доказательств): раньше resolvable значило
    ТОЛЬКО "все поля заполнены" -- запись с выдуманным, но непустым event_id тоже считалась
    resolvable=True, потому что ничего не сверялось с реальными данными. Теперь resolvable=True
    ТОЛЬКО если event_id реально найден в known_records -- индексе, который вызывающая сторона
    обязана построить ИЗ РЕАЛЬНЫХ строк БД/каталога (не из заявленных значений), его
    source_system совпадает буквально, и у найденной записи вообще есть дата (доказывает, что
    это настоящая датированная запись, а не пустая заглушка). known_records=None (не передан) --
    fail CLOSED: resolvable всегда False, "нечем проверить" не значит "верю на слово".

    require_at -- требование E4 (26.07): "у каждого доказательства дата, если есть время"
    (ТЗ владельца) -- для большинства kind дата обязательна (require_at=True, по умолчанию,
    поведение не меняется). Ровно один существующий вызывающий передаёт False -- kind="product"
    (_owner50_family_rows): запись price_axes_catalog -- это действующий факт о цене, а не
    датированное событие; "если есть время" -- условие необязательное для такого kind."""
    at_value = _parse_iso_datetime(at)
    missing = [name for name, value in (("event_id", event_id), ("source_system", source_system)) if not _clean_text(value)]
    if require_at and at_value is None:
        missing.append("at")
    if not _clean_text(text):
        missing.append("text")
    resolvable = False
    if not missing:
        known = (known_records or {}).get(_clean_text(event_id))
        resolvable = (
            isinstance(known, Mapping)
            and _clean_text(known.get("source_system")) == _clean_text(source_system)
            and (not require_at or _parse_iso_datetime(known.get("at")) is not None)
        )
        if not resolvable:
            missing.append("not_found_in_database")
    return {
        "kind": kind,
        "text": _clean_text(text),
        "event_id": _clean_text(event_id),
        "source_system": _clean_text(source_system),
        "at": at_value.isoformat() if at_value else "",
        "resolvable": resolvable,
        "missing_fields": tuple(missing),
    }


def resolve_evidence_source(event_id: str, events_by_id: Mapping[str, Mapping[str, Any]]) -> str | None:
    """Разрешает event_id в его source_system через переданный индекс событий семьи. None --
    ссылка "висит в воздухе" (событие не найдено или у него нет source_system); вызывающая
    сторона должна трактовать это как непройденное доказательство (аналог signal_evidence_not_owned)."""
    event = events_by_id.get(str(event_id))
    if not isinstance(event, Mapping):
        return None
    source_system = _clean_text(event.get("source_system"))
    return source_system or None


def owner50_action_text(family: Mapping[str, Any]) -> str:
    """Правка Fable #1: одна императивная фраза "кому + что + срок" для READY-строки
    ("тест 5 секунд" -- менеджер прочитал и сразу звонит). Пусто, если next_step.action
    не задан -- в этом случае строка и не должна была стать READY (см. Г6 выше)."""
    next_step = _mapping(family.get("next_step"))
    action = _clean_text(next_step.get("action"))
    if not action:
        return ""
    due = _clean_text(next_step.get("due"))
    who = _clean_text(_mapping(family.get("identity")).get("display_name"))
    pieces = [action.rstrip(".")]
    if who and who.casefold() not in action.casefold():
        pieces.append(f"кому: {who}")
    if due:
        pieces.append(f"до {due}")
    return "; ".join(pieces)


# --- classify_family: внутренние хелперы (специфичные для этого блока; общие с
# остальным manager_dossier.py -- _mapping/_clean_text/_parse_iso_datetime --
# сознательно переиспользованы выше, не продублированы) ---


def _owner50_resolve_as_of(explicit: datetime | None, family: Mapping[str, Any]) -> datetime:
    candidate = explicit if explicit is not None else family.get("as_of")
    resolved = _parse_iso_datetime(candidate) if candidate is not None else None
    return resolved or datetime.now(timezone.utc)


def _owner50_months_since(at: datetime, as_of: datetime) -> float:
    return max(0.0, (as_of - at).total_seconds() / 86400.0 / OWNER50_AVERAGE_DAYS_PER_MONTH)


def _owner50_days_since(at: datetime, as_of: datetime) -> float:
    return max(0.0, (as_of - at).total_seconds() / 86400.0)


def _owner50_academic_year(dt: datetime) -> int:
    # Российский учебный год начинается 1 сентября -- до сентября считаем "прошлым" годом.
    return dt.year if dt.month >= 9 else dt.year - 1


def _owner50_int_or_none(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _owner50_resolve_child_grade(child: Mapping[str, Any], as_of: datetime) -> int | None:
    explicit = _owner50_int_or_none(child.get("grade_current"))
    if explicit is not None:
        return explicit
    recorded = _owner50_int_or_none(child.get("grade_recorded"))
    fixed_at = _parse_iso_datetime(child.get("grade_fixed_at"))
    if recorded is None or fixed_at is None:
        return None
    return recorded + (_owner50_academic_year(as_of) - _owner50_academic_year(fixed_at))


def _owner50_child_is_graduate(child: Mapping[str, Any], as_of: datetime) -> bool:
    if child.get("is_graduate"):
        return True
    grade = _owner50_resolve_child_grade(child, as_of)
    return grade is not None and grade >= OWNER50_GRADUATE_GRADE_THRESHOLD


def _owner50_dedupe_children(children: Sequence[Mapping[str, Any]]) -> list[Mapping[str, Any]]:
    """Группирует по child_key (никогда не сливает РАЗНЫХ детей; повтор одного и того же
    child_key -- это одна и та же карточка, последняя запись побеждает)."""
    order: list[str] = []
    by_key: dict[str, Mapping[str, Any]] = {}
    for index, child in enumerate(children):
        if not isinstance(child, Mapping):
            continue
        key = _clean_text(child.get("child_key")) or f"__unkeyed_{index}"
        if key not in by_key:
            order.append(key)
        by_key[key] = child
    return [by_key[key] for key in order]


def _owner50_payment_is_confirmed(payment: Mapping[str, Any], *, as_of: datetime | None = None) -> bool:
    if not payment:
        return False
    # требование архитектора #2: оплата засчитывается только при известной дате (без
    # last_purchase_at это агрегат без "когда" -- не отличить от протухшего снапшота).
    last_purchase_at = _parse_iso_datetime(payment.get("last_purchase_at"))
    if last_purchase_at is None:
        return False
    # требование архитектора #5: дата оплаты в будущем -- не "свежо", это ошибка данных,
    # а не повод считать оплату подтверждённой.
    if as_of is not None and last_purchase_at > as_of:
        return False
    return (
        float(payment.get("total_in") or 0) > 0
        and int(payment.get("deals_cnt") or 0) > 0
    )


def _owner50_product_confirmed(
    product: Mapping[str, Any],
    brands: set[str],
    children: Sequence[Mapping[str, Any]],
    as_of: datetime,
) -> bool:
    if not product or not product.get("verified"):
        return False
    source = _clean_text(product.get("source"))
    # требование аудиторов BLOCKED #6 (продукт только из KB): "kb"-префикс источника САМ по
    # себе ничего не проверяет -- дополнительно требуем непустой entry_id, тот самый точный
    # ключ актуального price_axes_catalog (_owner50_product_from_price_entry пишет entry_id
    # ТОЛЬКО когда select_price реально нашла запись; никакой другой код product не строит).
    if not source.casefold().startswith("kb") or not _clean_text(product.get("entry_id")):
        return False
    if product.get("seats_available") is False:
        return False
    product_brand = _clean_text(product.get("brand")).casefold()
    if brands and product_brand not in brands:
        return False
    grade_min = _owner50_int_or_none(product.get("grade_min"))
    grade_max = _owner50_int_or_none(product.get("grade_max"))
    if grade_min is not None and grade_max is not None and children:
        grades = [g for g in (_owner50_resolve_child_grade(child, as_of) for child in children) if g is not None]
        if grades and not any(grade_min <= g <= grade_max for g in grades):
            return False
    return True


def _owner50_is_concrete_next_step(text: str) -> bool:
    value = _clean_text(text)
    if value.casefold().startswith(_OWNER50_CLARIFY_INTEREST_PREFIX):
        return True  # спека §5.3: "уточнить интерес" -- легальный шаг сам по себе, даже короткий
    return _meaningful_next_step(value)


def _owner50_select_target_child(
    verified_children: Sequence[Any],
    child_grade_sets: Sequence[set[int]],
    child_is_graduate: Sequence[bool],
) -> tuple[Any, int | None, bool]:
    """Требование E2 (26.07): продукт/предложение адресованы ОДНОМУ конкретному ребёнку, не
    "младшему классу из всей истории семьи". child_grade_sets[i] -- это МНОЖЕСТВО ВСЕХ
    когда-либо упомянутых классов ребёнка i (grades_json копится без дат -- family_graph.py
    только добавляет в set, никогда не переписывает) -- если там больше одного разного числа
    1-10, значит текущий класс этого ребёнка НЕИЗВЕСТЕН, и брать минимум из этого набора -- то
    самое запрещённое "гадание по истории". Ребёнок годится в адресаты, только если его
    СОБСТВЕННЫЙ набор классов однозначен (ровно одно число 1-10).

    Среди однозначных кандидатов:
      - нет ни одного -> (None, None, ambiguous=False) -- честно "не знаем", вызывающая сторона
        обязана либо подтвердить класс из текста предложения (offer), либо не подтверждать
        продукт вовсе;
      - ровно один -> он и есть адресат (child_row, его класс, ambiguous=False);
      - несколько РАЗНЫХ детей с однозначным классом каждый -> адресат не доказан
        (None, None, ambiguous=True) -- НИКАКОЙ класс не выбирается, в т.ч. не берётся
        "самый младший из подходящих" (то же запрещённое гадание, просто на уровне детей,
        а не истории одного ребёнка)."""
    eligible: list[tuple[Any, int]] = []
    for child_row, grades, is_graduate in zip(verified_children, child_grade_sets, child_is_graduate):
        if is_graduate:
            continue
        single_grades = {grade for grade in grades if 1 <= grade <= 10}
        if len(single_grades) == 1:
            eligible.append((child_row, next(iter(single_grades))))
    if not eligible:
        return None, None, False
    distinct_children = {_clean_text(child_row["child_key"]) for child_row, _grade in eligible}
    if len(distinct_children) > 1:
        return None, None, True
    child_row, grade = eligible[0]
    return child_row, grade, False


def _owner50_family_assumptions(family: Mapping[str, Any], children: Sequence[Mapping[str, Any]]) -> tuple[str, ...]:
    notes: list[str] = []
    if family.get("family_attribution_by_tag"):
        notes.append("семья: атрибуция по тегу (child_key), не отдельная запись")
    for child in children:
        if _owner50_int_or_none(child.get("grade_current")) is None and child.get("grade_recorded") is not None:
            label = _clean_text(child.get("name")) or _clean_text(child.get("child_key")) or "ребёнок"
            notes.append(f"класс «{label}» — пересчёт от {child.get('grade_recorded')} кл. (§5.5)")
    return tuple(notes)


@dataclass(frozen=True)
class DossierMarker:
    kind: str
    text: str
    source: str


@dataclass(frozen=True)
class DossierRow:
    section: str
    text: str
    source: str


@dataclass(frozen=True)
class CustomerDossier:
    tenant_id: str
    customer_id: str
    display_name: str
    brand: str
    phone: str
    email: str
    actuality_header: str = ""
    family: tuple[DossierRow, ...] = field(default_factory=tuple)
    money: tuple[DossierRow, ...] = field(default_factory=tuple)
    signals: tuple[DossierRow, ...] = field(default_factory=tuple)
    next_step: str = ""
    next_step_source: str = ""
    action_status: str = NEXT_STEP_STATUS_EMPTY
    no_action_reason_code: str = ""
    objections: tuple[DossierRow, ...] = field(default_factory=tuple)
    chronology: tuple[DossierRow, ...] = field(default_factory=tuple)
    interests: tuple[DossierMarker, ...] = field(default_factory=tuple)
    pains: tuple[DossierMarker, ...] = field(default_factory=tuple)


def build_customer_dossier(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_id: str,
    canonical_calls: Mapping[str, str] | None = None,
    actuality_header: str = "",
) -> CustomerDossier:
    con.row_factory = sqlite3.Row
    customer = con.execute(
        """
        SELECT customer_id, tenant_id, display_name, primary_phone, primary_email, record_json
        FROM customer_identities
        WHERE tenant_id = ? AND customer_id = ?
        """,
        (tenant_id, customer_id),
    ).fetchone()
    if customer is None:
        raise ValueError(f"customer not found: {customer_id}")
    customer_record = _safe_json(customer["record_json"])
    identity_brands = [
        str(item).strip().casefold()
        for item in (_mapping(customer_record.get("metadata")).get("brands") or ())
        if str(item).strip().casefold() in MANAGER_KNOWN_BRANDS
    ]
    opportunities = con.execute(
        """
        SELECT opportunity_id, record_json
        FROM customer_opportunities
        WHERE tenant_id = ? AND customer_id = ?
        ORDER BY opened_at DESC, opportunity_id
        """,
        (tenant_id, customer_id),
    ).fetchall()
    opportunity_records = tuple(_safe_json(row["record_json"]) for row in opportunities)
    event_brand_records = tuple(
        _safe_json(row["record_json"])
        for row in con.execute(
            """
            SELECT record_json
            FROM timeline_events
            WHERE tenant_id = ? AND customer_id = ?
            """,
            (tenant_id, customer_id),
        ).fetchall()
    )
    derived_brands = customer_summary_brands(opportunity_records, event_brand_records, ())
    known_identity_brands = set(identity_brands)
    if any(has_explicit_brand_conflict(record) for record in (*opportunity_records, *event_brand_records)):
        brands = []
    elif len(known_identity_brands) == 1:
        brands = [next(iter(known_identity_brands))]
    else:
        brands = [brand for brand in derived_brands if brand in MANAGER_KNOWN_BRANDS]
    events = con.execute(
        """
        SELECT event_id, event_at, source_id, source_ref, event_type, record_json
        FROM timeline_events
        WHERE tenant_id = ?
          AND customer_id = ?
          AND event_type = 'mango_call'
          AND match_status = 'strong_unique'
          AND (superseded_by IS NULL OR superseded_by = '')
          AND COALESCE(json_extract(record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true')
        ORDER BY event_at DESC, event_id DESC
        LIMIT 100
        """,
        (tenant_id, customer_id),
    ).fetchall()
    interests: list[DossierMarker] = []
    pains: list[DossierMarker] = []
    for value in _product_interest_values(customer["record_json"], opportunities):
        interests.append(DossierMarker(kind="interest", text=f"Из данных: {value}", source="products_of_interest"))
    call_texts = canonical_calls or {}
    for event in events:
        client_text = _lookup_canonical_client_text(event, call_texts)
        if not client_text.strip():
            continue
        source = f"mango_call:{event['source_id']}"
        interests.extend(_markers_from_client_text(client_text, INTEREST_MARKER_RE, kind="interest", label="Интерес из звонка", source=source))
        pains.extend(_markers_from_client_text(client_text, PAIN_MARKER_RE, kind="pain", label="Боль из звонка", source=source))
    signals = _signal_rows(con, tenant_id=tenant_id, customer_id=customer_id)
    next_step, next_step_source, action_status, no_action_reason_code = _next_step_for_dossier(
        con, tenant_id=tenant_id, customer_id=customer_id, signals=signals
    )
    return CustomerDossier(
        tenant_id=str(customer["tenant_id"]),
        customer_id=str(customer["customer_id"]),
        display_name=_clean_text(customer["display_name"]),
        brand=brands[0] if len(brands) == 1 else "",
        phone=_clean_text(customer["primary_phone"]),
        email=_clean_text(customer["primary_email"]),
        actuality_header=actuality_header,
        family=tuple(
            _family_rows(
                con,
                tenant_id=tenant_id,
                customer_id=customer_id,
                active_brand=brands[0] if len(brands) == 1 else "",
            )
        ),
        money=tuple(_money_rows(con, tenant_id=tenant_id, customer_id=customer_id)),
        signals=tuple(signals),
        next_step=next_step,
        next_step_source=next_step_source,
        action_status=action_status,
        no_action_reason_code=no_action_reason_code,
        objections=tuple(_objection_rows(con, tenant_id=tenant_id, customer_id=customer_id)),
        chronology=tuple(_chronology_rows(con, tenant_id=tenant_id, customer_id=customer_id, limit=12)),
        interests=tuple(_dedupe_markers(interests, limit=8)),
        pains=tuple(_dedupe_markers(pains, limit=8)),
    )


def build_manager_dossier_workbook(
    *,
    timeline_db: Path | str,
    allowed_root: Path | str,
    out_xlsx: Path | str,
    tenant_id: str = "foton",
    customer_ids: Sequence[str] | None = None,
    canonical_calls_db: Path | str | None = None,
    reconcile_json: Path | str | None = None,
    limit: int = 50,
    enforce_freshness: bool = True,
    enforce_outreach_eligibility: bool = False,
) -> Mapping[str, Any]:
    db = Path(timeline_db).expanduser().resolve(strict=False)
    out = _guard_local_dossier_output_path(out_xlsx, allowed_root)
    out.parent.mkdir(parents=True, exist_ok=True)
    canonical_calls, canonical_warning = _load_canonical_calls_fail_soft(canonical_calls_db)
    reconcile = _read_json(Path(reconcile_json).expanduser()) if reconcile_json else {}
    with _connect_ro(db) as con:
        ids = (
            tuple(_full_dossier_segment_customer_ids(con, tenant_id=tenant_id, limit=limit))
            if customer_ids is None
            else tuple(customer_ids)
        )
        freshness = _source_freshness(con, tenant_id=tenant_id)
        freshness_gate = manager_freshness_gate(freshness)
        if enforce_freshness and not freshness_gate["passed"]:
            reasons = ", ".join(
                f"{item['source_system']}:{item['reason']}" for item in freshness_gate["blockers"]
            )
            raise RuntimeError(f"manager freshness gate failed: {reasons}")
        segment_total = _full_dossier_segment_count(con, tenant_id=tenant_id)
        actuality_header = _actuality_header(freshness, reconcile)
        dossiers: list[CustomerDossier] = []
        missing_customer_ids: list[str] = []
        exclusion_counts: Counter[str] = Counter()
        for customer_id in ids:
            if enforce_outreach_eligibility:
                eligibility = manager_outreach_eligibility(
                    con,
                    tenant_id=tenant_id,
                    customer_id=customer_id,
                )
                if not eligibility["eligible"]:
                    exclusion_counts.update(eligibility["reasons"])
                    continue
            try:
                dossiers.append(
                    build_customer_dossier(
                        con,
                        tenant_id=tenant_id,
                        customer_id=customer_id,
                        canonical_calls=canonical_calls,
                        actuality_header=actuality_header,
                    )
                )
            except ValueError:
                missing_customer_ids.append(customer_id)
    _write_workbook(out, dossiers)
    summary = {
        "schema_version": MANAGER_DOSSIER_SCHEMA_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": tenant_id,
        "requested_customers": len(ids),
        "customers": len(dossiers),
        "missing_customer_ids_count": len(missing_customer_ids),
        "missing_customer_ids_sample": missing_customer_ids[:10],
        "outreach_eligibility_enforced": bool(enforce_outreach_eligibility),
        "outreach_exclusion_counts": dict(exclusion_counts),
        "full_dossier_segment_total": segment_total,
        "interests_total": sum(len(item.interests) for item in dossiers),
        "pains_total": sum(len(item.pains) for item in dossiers),
        "family_rows_total": sum(len(item.family) for item in dossiers),
        "money_rows_total": sum(len(item.money) for item in dossiers),
        "signals_total": sum(len(item.signals) for item in dossiers),
        "objections_total": sum(len(item.objections) for item in dossiers),
        "chronology_rows_total": sum(len(item.chronology) for item in dossiers),
        "next_step_rows_total": sum(1 for item in dossiers if item.next_step),
        "missing_next_step_rows_total": sum(1 for item in dossiers if not item.next_step),
        "action_status_counts": dict(sorted(Counter(item.action_status for item in dossiers).items())),
        "no_action_reason_counts": dict(sorted(Counter(
            item.no_action_reason_code for item in dossiers if item.no_action_reason_code
        ).items())),
        "canonical_calls_loaded": len(canonical_calls),
        "canonical_calls_warning": canonical_warning,
        "actuality_header": actuality_header,
        "source_freshness_top": freshness[:12],
        "freshness_gate": freshness_gate,
        "reconcile_status": reconcile.get("status") if reconcile else "missing",
        "out_xlsx": str(out),
        "safety": {
            "source_open_mode": "sqlite_mode_ro",
            "write_crm": False,
            "write_tallanto": False,
            "send_messages": False,
            "pii_scope": "local_codex_local_only",
        },
    }
    out.with_suffix(".summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def build_owner50_family_workbook(
    *,
    timeline_db: Path | str,
    allowed_root: Path | str,
    out_xlsx: Path | str,
    tenant_id: str = "foton",
    limit: int = 50,
    as_of: datetime | None = None,
    price_axes_catalog: Mapping[str, Any] | Sequence[Any] | Path | str | None = None,
    enforce_freshness: bool = True,
) -> Mapping[str, Any]:
    """Build the owner-only family outreach queue without external writes.

    price_axes_catalog -- опционально: уже загруженный каталог (dict с "entries") или путь
    к price_axes_catalog.json. По умолчанию (None) читает актуальный файл каталога с диска
    (OWNER50_PRICE_AXES_CATALOG_DEFAULT_PATH) -- требование архитектора #3.

    enforce_freshness -- требование аудиторов BLOCKED #3 (устаревшие данные должны
    останавливать ВЕСЬ build, 25.07; отменяет прежнее требование архитектора #10, которое
    само же признавало проблему наполовину): при enforce_freshness=True (прод-дефолт) и
    непройденном manager_freshness_gate функция БРОСАЕТ RuntimeError ДО того, как открыта хоть
    одна семья -- ни один лист не пишется. Раньше здесь стоял stale_data=True, тихо
    протаскиваемый в classify_family на КАЖДУЮ семью -- весь batch становился EXCLUDED
    построчно, а сборка как ни в чём не бывало писала файл. Тот же паттерн (raise, а не тихая
    пометка) уже был у build_manager_dossier_workbook -- теперь применён и здесь, тем же
    способом, а не отдельным путём. Тесты, которые сознательно не сеют
    ingestion_cursors/ingestion_runs, передают enforce_freshness=False (тот же паттерн, что и
    у dossier-тестов)."""
    now = as_of or datetime.now(timezone.utc)
    if now.tzinfo is None:
        raise ValueError("as_of must be timezone-aware")
    out = _guard_local_dossier_output_path(out_xlsx, allowed_root)
    out.parent.mkdir(parents=True, exist_ok=True)
    catalog = _owner50_resolve_price_axes_catalog(price_axes_catalog)
    with _connect_ro(Path(timeline_db).expanduser().resolve(strict=False)) as con:
        freshness = _source_freshness(con, tenant_id=tenant_id)
        freshness_gate = manager_freshness_gate(freshness)
        if enforce_freshness and not freshness_gate["passed"]:
            reasons = ", ".join(f"{item['source_system']}:{item['reason']}" for item in freshness_gate["blockers"])
            raise RuntimeError(f"owner50 freshness gate failed (build stopped, no workbook written): {reasons}")
        candidates, control = _owner50_family_rows(
            con, tenant_id=tenant_id, as_of=now, price_axes_catalog=catalog,
        )
    candidates.sort(key=lambda row: row["rank_key"])
    effective_limit = min(50, max(0, int(limit)))
    selected = candidates[:effective_limit]
    selected_ids = {row["family_id"] for row in selected}
    # требование архитектора #10 (лист кандидатов неполноценный): control несёт те же
    # контакт/дети/сигнал/действие колонки, что и candidate/excluded строки ниже -- данные уже
    # реально посчитаны в row_common (_owner50_family_rows), здесь просто те же поля.
    for rank, row in enumerate(selected, start=1):
        row["rank"] = rank
        control.append(_owner50_control_row_from_ready(row, status="selected", code="selected"))
    control.extend(
        _owner50_control_row_from_ready(row, status="outside_limit", code="outside_limit")
        for row in candidates
        if row["family_id"] not in selected_ids
    )
    excluded_ids = {family_id for family_id, status, *_rest in control if status == "excluded"}
    # требование архитектора #1/#9: READY (candidates/selected) / CANDIDATE (candidate,
    # никогда не добивает READY до limit) / EXCLUDED -- три категории видны раздельно.
    # требование аудиторов BLOCKED #5: теперь они физически разведены по трём листам
    # (READY_50/CANDIDATES/EXCLUDED), а не различаются только колонкой "Статус" одного листа.
    candidate_queue_ids = {family_id for family_id, status, *_rest in control if status == "candidate"}
    exclusion_counts = Counter(code for _, status, code, *_rest in control if status == "excluded")
    candidate_reason_counts = Counter(code for _, status, code, *_rest in control if status == "candidate")
    outside_limit_count = sum(1 for _, status, *_rest in control if status == "outside_limit")
    ready_total = len(selected) + outside_limit_count
    catalog_provenance = _owner50_catalog_provenance(catalog)
    control_meta = {
        "tenant_id": tenant_id,
        "as_of": now.isoformat(),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "ready_50": len(selected),
        "ready_total": ready_total,
        "ready_outside_limit": outside_limit_count,
        "ready_audit_population_complete": outside_limit_count == 0,
        "candidates": len(candidate_queue_ids),
        "excluded": len(excluded_ids),
        "families_classified_total": len(candidates) + len(excluded_ids) + len(candidate_queue_ids),
        "freshness_gate_passed": freshness_gate["passed"],
        "freshness_gate_checked_at": freshness_gate["checked_at"],
        "freshness_rows": freshness,
        "price_axes_catalog_provenance": catalog_provenance,
    }
    _write_owner50_workbook(out, selected, control, control_meta)
    return {
        "families": len(selected),
        "ready_total": ready_total,
        "ready_audit_population_complete": outside_limit_count == 0,
        "candidate_families": len(candidates),
        "candidate_queue_families": len(candidate_queue_ids),
        "candidate_queue_reason_counts": dict(sorted(candidate_reason_counts.items())),
        "excluded_families": len(excluded_ids),
        "exclusion_counts": dict(sorted(exclusion_counts.items())),
        "requested_limit": int(limit),
        "effective_limit": effective_limit,
        "required_business_columns": OWNER50_REQUIRED_COLUMNS,
        "out_xlsx": str(out),
        "sheets": ("READY_50", "CANDIDATES", "EXCLUDED", "EVIDENCE", "CONTROL"),
        "freshness_gate_passed": freshness_gate["passed"],
        "price_axes_catalog_provenance": catalog_provenance,
        "write_external": False,
    }


def load_canonical_call_client_texts(path: Path | str | None) -> Mapping[str, str]:
    if path is None:
        return {}
    db = Path(path).expanduser().resolve(strict=False)
    if not db.exists():
        return {}
    with _connect_ro(db) as con:
        rows = con.execute("SELECT canonical_call_id, transcript_client FROM canonical_calls").fetchall()
    return {str(row[0]): str(row[1] or "") for row in rows}


def _load_canonical_calls_fail_soft(path: Path | str | None) -> tuple[Mapping[str, str], str]:
    if path is None:
        return {}, ""
    db = Path(path).expanduser().resolve(strict=False)
    if not db.exists():
        return {}, f"canonical calls DB not found, continuing without call quotes: {db}"
    try:
        return load_canonical_call_client_texts(db), ""
    except (sqlite3.Error, OSError) as exc:
        return {}, f"canonical calls DB unavailable, continuing without call quotes: {type(exc).__name__}"


def _guard_local_dossier_output_path(path: Path | str, allowed_root: Path | str) -> Path:
    resolved = guard_customer_timeline_output_path(path, allowed_root)
    root = Path(allowed_root).resolve(strict=False)
    relative = resolved.relative_to(root)
    if not relative.parts or relative.parts[0] != ".codex_local":
        raise ValueError("manager dossier output contains PII and must stay under .codex_local")
    return resolved


def _lookup_canonical_client_text(event: sqlite3.Row, canonical_calls: Mapping[str, str]) -> str:
    for key in _canonical_call_candidate_keys(event):
        value = canonical_calls.get(key)
        if value:
            return value
    return ""


def _canonical_call_candidate_keys(event: sqlite3.Row) -> tuple[str, ...]:
    keys: list[str] = []
    record = _safe_json(event["record_json"])
    nested_record = record.get("record") if isinstance(record, Mapping) and isinstance(record.get("record"), Mapping) else {}
    canonical_call_id = _clean_text(
        (record.get("canonical_call_id") or nested_record.get("canonical_call_id")) if isinstance(record, Mapping) else None
    )
    if canonical_call_id:
        keys.append(canonical_call_id)
    for raw in (event["source_id"], event["source_ref"]):
        text = _clean_text(raw)
        if not text:
            continue
        keys.append(text)
        if text.startswith("call:"):
            keys.append(text.removeprefix("call:"))
        if ":" in text:
            keys.append(text.split(":", 1)[0])
    return tuple(_dedupe_texts(keys))


def _connect_ro(path: Path) -> sqlite3.Connection:
    con = sqlite3.connect(customer_timeline_readonly_uri(path), uri=True)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA query_only = ON")
    return con


def manager_outreach_eligibility(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_id: str,
    signal_id: str | None = None,
    as_of: datetime | None = None,
) -> Mapping[str, Any]:
    """Fail closed before a customer reaches a proactive manager list."""
    con.row_factory = sqlite3.Row
    now = as_of or datetime.now(timezone.utc)
    if now.tzinfo is None:
        raise ValueError("as_of must be timezone-aware")
    required_tables = ("customer_identities", "derived_signals", "timeline_events", "timeline_conflicts", "family_links_v1")
    missing = [table for table in required_tables if not _table_exists(con, table)]
    if missing:
        return {"eligible": False, "reasons": tuple(f"safety_table_missing:{table}" for table in missing)}

    identity = con.execute(
        "SELECT identity_status, record_json FROM customer_identities WHERE tenant_id=? AND customer_id=?",
        (tenant_id, customer_id),
    ).fetchone()
    reasons: list[str] = []
    identity_record = _safe_json(identity["record_json"]) if identity else {}
    raw_brands = {
        str(item).strip().casefold()
        for item in (_mapping(identity_record.get("metadata")).get("brands") or ())
        if str(item).strip()
    }
    brands = {OWNER50_BRAND_ALIASES[item] for item in raw_brands if item in OWNER50_BRAND_ALIASES}
    if not identity or str(identity["identity_status"] or "") != "strong":
        reasons.append("identity_not_strong")
    if raw_brands - OWNER50_BRAND_ALIASES.keys() or len(brands) != 1 or not brands.issubset(MANAGER_KNOWN_BRANDS):
        reasons.append("brand_not_exactly_one_known")

    signal_clauses = [
        "tenant_id=?", "customer_id=?", "status='active'",
        f"signal_type IN ({','.join('?' for _ in MANAGER_OUTREACH_SIGNAL_TYPES)})",
        "(expires_at IS NULL OR expires_at='' OR julianday(expires_at)>=julianday(?))",
    ]
    signal_params: list[Any] = [tenant_id, customer_id, *MANAGER_OUTREACH_SIGNAL_TYPES, now.isoformat()]
    if signal_id:
        signal_clauses.append("signal_id=?")
        signal_params.append(signal_id)
    signal = con.execute(
        f"SELECT signal_id, event_id, signal_type, created_at, record_json FROM derived_signals "
        f"WHERE {' AND '.join(signal_clauses)} ORDER BY created_at DESC, signal_id LIMIT 1",
        tuple(signal_params),
    ).fetchone()
    if signal is None:
        reasons.append("no_active_outreach_signal")

    refs = customer_entity_ref_values(customer_id)
    open_conflict = con.execute(
        "SELECT 1 FROM timeline_conflicts c WHERE c.tenant_id=? AND c.status IN ('open','active') "
        "AND json_valid(c.record_json) AND EXISTS (SELECT 1 FROM json_each(c.record_json,'$.entity_refs') r "
        f"WHERE CAST(r.value AS TEXT) IN ({','.join('?' for _ in refs)})) LIMIT 1",
        (tenant_id, *refs),
    ).fetchone()
    if open_conflict:
        reasons.append("open_identity_conflict")
    family_risk = con.execute(
        "SELECT 1 FROM family_links_v1 WHERE tenant_id=? AND customer_id=? "
        "AND (COALESCE(status,'')!='confident' OR COALESCE(confidence,'') NOT IN ('high','medium')) LIMIT 1",
        (tenant_id, customer_id),
    ).fetchone()
    if family_risk:
        reasons.append("family_ambiguous")
    risk_signal = con.execute(
        f"SELECT signal_type FROM derived_signals WHERE tenant_id=? AND customer_id=? AND status='active' "
        f"AND signal_type IN ({','.join('?' for _ in MANAGER_OUTREACH_RISK_SIGNAL_TYPES)}) "
        "AND (expires_at IS NULL OR expires_at='' OR julianday(expires_at)>=julianday(?)) LIMIT 1",
        (tenant_id, customer_id, *MANAGER_OUTREACH_RISK_SIGNAL_TYPES, now.isoformat()),
    ).fetchone()
    if risk_signal:
        reasons.append(f"active_risk_signal:{risk_signal['signal_type']}")

    evidence_at: datetime | None = None
    signal_created_at: datetime | None = None
    if signal is not None:
        signal_created_at = _parse_iso_datetime(signal["created_at"])
        signal_record = _safe_json(signal["record_json"])
        event_id = _clean_text(signal["event_id"] or signal_record.get("event_id"))
        if event_id:
            event = con.execute(
                "SELECT event_at,event_type,match_status,superseded_by FROM timeline_events "
                "WHERE tenant_id=? AND customer_id=? AND event_id=? "
                "AND COALESCE(json_extract(record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true') LIMIT 1",
                (tenant_id, customer_id, event_id),
            ).fetchone()
            if event is None:
                reasons.append("signal_evidence_not_owned")
            elif event["superseded_by"]:
                reasons.append("signal_evidence_superseded")
            elif str(event["event_type"] or "") == "mango_call" and str(event["match_status"] or "") != "strong_unique":
                reasons.append("signal_evidence_ambiguous_call")
            else:
                evidence_at = _parse_iso_datetime(event["event_at"])
        elif str(signal["signal_type"]) == "season_return_candidate":
            evidence_at = _parse_iso_datetime(_mapping(signal_record.get("metadata")).get("last_purchase_at"))
            if evidence_at is None:
                reasons.append("signal_evidence_missing")
            elif not _season_purchase_matches(
                con,
                tenant_id=tenant_id,
                customer_id=customer_id,
                evidence_at=evidence_at,
                as_of=now,
            ):
                reasons.append("season_purchase_not_confirmed")
            elif _has_active_customer_access(con, tenant_id=tenant_id, customer_id=customer_id):
                reasons.append("active_access_or_learning")
        else:
            reasons.append("signal_evidence_missing")

    scan_from = min(filter(None, (signal_created_at, now - timedelta(days=30))), default=now - timedelta(days=30))
    event_rows = con.execute(
        "SELECT event_id,event_at,event_type,source_system,source_id,source_ref,subject,text_preview,summary,direction,record_json "
        "FROM timeline_events WHERE tenant_id=? AND customer_id=? AND (superseded_by IS NULL OR superseded_by='') "
        "AND COALESCE(json_extract(record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true') "
        "AND julianday(event_at)>=julianday(?) ORDER BY event_at,event_id",
        (tenant_id, customer_id, scan_from.isoformat()),
    ).fetchall()
    outbound_cutoff = max(filter(None, (evidence_at, now - timedelta(days=30))), default=now - timedelta(days=30))
    for row in event_rows:
        event = dict(row)
        stored = _safe_json(row["record_json"])
        event["record"] = _mapping(stored.get("record"))
        event["metadata"] = _mapping(stored.get("metadata"))
        text = _event_text(event)
        if (
            str(row["direction"] or "").casefold() == "outbound"
            and (_parse_iso_datetime(row["event_at"]) or now) > outbound_cutoff
            and not _is_non_closing_service_event(event)
        ):
            reasons.append("meaningful_outbound_after_evidence")
    # ponytail: block historical hard risks until a structured resolution/opt-in field exists.
    reasons.extend(_durable_contact_risks(con, tenant_id=tenant_id, customer_id=customer_id))
    unique_reasons = tuple(dict.fromkeys(reasons))
    return {
        "eligible": not unique_reasons,
        "reasons": unique_reasons,
        "signal_id": str(signal["signal_id"]) if signal else None,
        "signal_type": str(signal["signal_type"]) if signal else None,
        "brand": next(iter(brands)) if len(brands) == 1 else None,
    }


def _season_purchase_matches(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_id: str,
    evidence_at: datetime,
    as_of: datetime,
) -> bool:
    if not _table_exists(con, "customer_purchases_v1"):
        return False
    row = con.execute(
        "SELECT SUM(total_in) AS total_in, SUM(total_out) AS total_out, "
        "SUM(deals_cnt) AS deals_cnt, MAX(last_purchase_at) AS last_purchase_at "
        "FROM customer_purchases_v1 WHERE tenant_id=? AND customer_id=? AND money_kind='fact'",
        (tenant_id, customer_id),
    ).fetchone()
    stored_at = _parse_iso_datetime(row["last_purchase_at"]) if row else None
    return bool(
        row
        and float(row["total_in"] or 0) > 0
        and int(row["deals_cnt"] or 0) > 0
        and stored_at
        and stored_at <= as_of
        and evidence_at <= as_of
        and stored_at.date() == evidence_at.date()
    )


def _has_active_customer_access(con: sqlite3.Connection, *, tenant_id: str, customer_id: str) -> bool:
    rows = con.execute(
        "SELECT event_id,event_at,event_type,source_system,source_id,source_ref,subject,text_preview,summary,direction,record_json "
        "FROM timeline_events WHERE tenant_id=? AND customer_id=? "
        "AND COALESCE(json_extract(record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true') "
        "AND (superseded_by IS NULL OR superseded_by='') ORDER BY event_at DESC,event_id DESC",
        (tenant_id, customer_id),
    ).fetchall()
    for row in rows:
        event = dict(row)
        stored = _safe_json(row["record_json"])
        event["record"] = _mapping(stored.get("record"))
        event["metadata"] = _mapping(stored.get("metadata"))
        if _is_access_event(event):
            return True
    return False


def _durable_contact_risks(con: sqlite3.Connection, *, tenant_id: str, customer_id: str) -> tuple[str, ...]:
    customer_ids = list(_family_scope_customer_ids(con, tenant_id=tenant_id, customer_id=customer_id))
    family_link_columns = {str(row[1]) for row in con.execute("PRAGMA table_info(family_links_v1)").fetchall()}
    if not _table_exists(con, "family_members_v1") and "family_id" in family_link_columns:
        related = con.execute(
            "SELECT DISTINCT sibling.customer_id FROM family_links_v1 current "
            "JOIN family_links_v1 sibling ON sibling.tenant_id=current.tenant_id AND sibling.family_id=current.family_id "
            "WHERE current.tenant_id=? AND current.customer_id=? "
            "AND current.status='confident' AND current.confidence IN ('high','medium') "
            "AND sibling.status='confident' AND sibling.confidence IN ('high','medium')",
            (tenant_id, customer_id),
        ).fetchall()
        customer_ids.extend(str(row[0]) for row in related if row[0])
    customer_ids = list(dict.fromkeys(customer_ids))
    rows = con.execute(
        "SELECT event_id,event_at,event_type,source_system,source_id,source_ref,subject,text_preview,summary,direction,record_json "
        f"FROM timeline_events WHERE tenant_id=? AND customer_id IN ({','.join('?' for _ in customer_ids)}) "
        "AND COALESCE(json_extract(record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true') "
        "AND (superseded_by IS NULL OR superseded_by='') ORDER BY event_at DESC,event_id DESC",
        (tenant_id, *customer_ids),
    ).fetchall()
    risks: list[str] = []
    for row in rows:
        event = dict(row)
        stored = _safe_json(row["record_json"])
        event["record"] = _mapping(stored.get("record"))
        event["metadata"] = _mapping(stored.get("metadata"))
        if _owner50_event_is_explicit_refund(event):
            risks.append("durable_p0_history")
        if _clean_text(event.get("direction")).casefold() != "inbound":
            continue
        text = _event_text(event)
        if hard_codes_from_text(text):
            risks.append("durable_p0_history")
        if any(phrase in text for phrase in MANAGER_OPTOUT_PHRASES):
            risks.append("durable_opt_out")
    return tuple(dict.fromkeys(risks))


@lru_cache(maxsize=4)
def _owner50_load_price_axes_catalog(path_str: str) -> Mapping[str, Any]:
    """Читает уже посчитанный price_axes_catalog.json как данные (кэш по пути файла --
    внутри одного процесса каталог не меняется). Отказ читать файл -- fail SOFT: продукт
    просто не подтвердится (CANDIDATE), а не роняет весь batch owner50."""
    try:
        raw = Path(path_str).read_text(encoding="utf-8")
        parsed = json.loads(raw)
    except (OSError, ValueError):
        return {"entries": ()}
    return parsed if isinstance(parsed, Mapping) else {"entries": ()}


def _owner50_resolve_price_axes_catalog(
    catalog: Mapping[str, Any] | Sequence[Any] | Path | str | None,
) -> Mapping[str, Any] | Sequence[Any]:
    if catalog is None:
        return _owner50_load_price_axes_catalog(str(OWNER50_PRICE_AXES_CATALOG_DEFAULT_PATH))
    if isinstance(catalog, (Path, str)):
        return _owner50_load_price_axes_catalog(str(Path(catalog)))
    return catalog


def _owner50_catalog_provenance(catalog: Mapping[str, Any] | Sequence[Any]) -> str:
    """Требование аудиторов BLOCKED #5/#6 (лист CONTROL несёт версию базы; продукт только из
    KB): "версия" каталога -- честно из его собственных данных (source_snapshot -- путь до
    kb_release_v3_snapshot.json, из которого build_kb_price_axes_catalog.py посчитал этот
    price_axes_catalog.json), не выдумка. Каталог может быть как полным словарём (боевой
    price_axes_catalog.json), так и голым списком entries (синтетика в тестах) -- во втором
    случае версии просто нет, это тоже честный ответ, а не догадка."""
    if isinstance(catalog, Mapping):
        return (
            _clean_text(catalog.get("source_snapshot"))
            or _clean_text(catalog.get("schema_version"))
            or "unknown_kb_snapshot"
        )
    return "inline_catalog_no_version_metadata"


def _owner50_select_price_entry(
    catalog: Mapping[str, Any] | Sequence[Any],
    *,
    brand: str,
    grade: int | None,
    offer_texts: Sequence[str],
) -> Mapping[str, Any] | None:
    """Резолвит ОДИН entry_id актуального price_axes_catalog под (brand, grade) -- требование
    архитектора #3. Пробует ТОЛЬКО формат/период, явно упомянутые в тексте предложения (через
    extract_price_query_axes -- та же логика, что и у бота при ответе про цену). Требование
    архитектора #10 (по итогам ревью 25.07): раньше при неизвестном формате/периоде код
    перебирал канонические комбинации и брал ПЕРВУЮ подходящую -- это была догадка о
    формате/цене, а не разбор данных. Теперь без явного текстового сигнала -- None (продукт не
    подтверждён, поля остаются пустыми); вызывающая сторона обязана трактовать None как
    "продукт не подтверждён", а не выдумывать вариант."""
    if grade is None or not brand:
        return None
    for text in offer_texts:
        if not text:
            continue
        axes = extract_price_query_axes(text, active_brand=brand)
        fmt = str(axes.get("format") or "")
        period = str(axes.get("period") or "")
        if not fmt or not period:
            continue
        result = select_price(catalog, brand=brand, grade=grade, format=fmt, period=period)
        if result.get("status") == "exact":
            entry = result.get("entry")
            return entry if isinstance(entry, Mapping) else None
    return None


def _owner50_product_from_price_entry(entry: Mapping[str, Any] | None, *, brand: str) -> Mapping[str, Any] | None:
    """Строит family["product"] для classify_family из найденной строки каталога. seats_available
    сознательно НЕ выставляется в False -- "места есть по умолчанию, если нет явного sold_out"
    (требование #3); явный sold_out в owner50-данных сейчас не отслеживается, поэтому здесь
    всегда None (= доступно), а не выдумка. verified=True и source, начинающийся с "kb_", -- ТОЛЬКО
    для реально найденного entry_id, никогда для угаданного."""
    if entry is None:
        return None
    entry_id = _clean_text(entry.get("entry_id"))
    if not entry_id:
        return None
    return {
        "name": _clean_text(entry.get("client_safe_text")) or entry_id,
        "brand": _clean_text(entry.get("brand")) or brand,
        "verified": True,
        "source": f"kb_price_axes_catalog:{entry_id}",
        "seats_available": None,
        "grade_min": entry.get("grade_min"),
        "grade_max": entry.get("grade_max"),
        "entry_id": entry_id,
        "amount": entry.get("amount"),
        "currency": _clean_text(entry.get("currency")),
        # требование E4 (26.07): дата для evidence "product", когда каталог её несёт (боевой
        # price_axes_catalog всегда её пишет -- KC_SOURCE_UPDATED_AT в price_axes_catalog.py).
        # Синтетические тестовые каталоги её не несут -- это честно (см. require_at=False у
        # _owner50_evidence_item(kind="product", ...): каталожная запись это не дата+событие,
        # а действующий факт, "если есть время" -- необязательное условие резолвируемости).
        "source_document_updated_at": _clean_text(entry.get("source_document_updated_at")),
    }


OWNER50_P0_RESOLVED_STATUSES = frozenset({"resolved", "closed", "refunded", "settled"})


def _owner50_event_p0_is_stale_and_resolved(
    event: Mapping[str, Any],
    *,
    codes: Sequence[str],
    all_events: Sequence[Mapping[str, Any]],
    as_of: datetime,
) -> bool:
    """Требование аудиторов BLOCKED #2 (25.07): refund-упоминание может быть прощено ТОЛЬКО
    ДОКАЗАННЫМ структурным статусом разрешения, найденным на самой записи -- НЕ эвристикой
    "после жалобы было ещё какое-то событие" (это доказывает лишь то, что переписка
    продолжилась, а не то, ЧЕМ закончилась жалоба). Ищем event.record.resolution_status или
    event.metadata.resolution_status (тот же вложенный формат, что несёт _owner50_event) --
    если он буквально "resolved"/"closed"/"refunded"/"settled", refund прощается; ЛЮБОЕ
    другое значение (включая отсутствие поля) -- False, семья остаётся EXCLUDED
    ("нет статуса -> CANDIDATE/EXCLUDED, никогда не READY"). Сегодня ни один источник
    ingestion не пишет resolution_status -- функция ЧЕСТНО всегда возвращает False, а не
    предполагает "наверное закрыто". all_events/as_of сохранены в сигнатуре ради обратной
    совместимости вызывающего кода и на случай будущего структурного источника, который решает
    вопрос "что было раньше/позже" по данным, а не по этой функции. Серьёзные коды (legal,
    complaint, payment_dispute) эта функция и раньше не смягчала -- только чистый "refund"."""
    if set(codes) - {"refund"}:
        return False
    if _parse_iso_datetime(event.get("event_at")) is None:
        return False
    record = _mapping(event.get("record"))
    metadata = _mapping(event.get("metadata"))
    resolution_status = (
        _clean_text(record.get("resolution_status")) or _clean_text(metadata.get("resolution_status"))
    ).casefold()
    return resolution_status in OWNER50_P0_RESOLVED_STATUSES


def _owner50_evidence_item(
    kind: str, text: str, source: str, *, event_id: str, source_system: str, at: Any,
    known_records: Mapping[str, Mapping[str, Any]] | None = None,
    require_at: bool = True,
) -> dict[str, Any]:
    """Требование архитектора #9: каждое доказательство несёт date+source_system+event_id
    (через build_evidence_record), а не только человекочитаемую строку "table:id".
    known_records -- требование аудиторов BLOCKED #1: индекс РЕАЛЬНО существующих записей,
    построенный вызывающей стороной (_owner50_family_rows) из настоящих строк БД, чтобы
    resolvable отражало действительность, а не просто "поля не пустые". require_at -- требование
    E4 (26.07), см. build_evidence_record."""
    record = build_evidence_record(
        kind=kind, text=text, event_id=event_id, source_system=source_system, at=at,
        known_records=known_records, require_at=require_at,
    )
    return {
        "kind": kind,
        "text": text,
        "source": source,
        "event_id": record["event_id"],
        "source_system": record["source_system"],
        "at": record["at"],
        "resolvable": record["resolvable"],
    }


# требование E5 (26.07): "человек" в owner50 -- ТОЛЬКО структурное происхождение, никогда смысл
# текста. Старые AMO/Tallanto identity хранят его в record.source_ref; объединённые identities
# доказывают его strong/manual связью amo_contact_id/tallanto_student_id. Название сделки или
# следующего шага без такого доказательства получает person_origin_unproven независимо от имени.
OWNER50_PERSON_CONTACT_LINK_TYPES = ("amo_contact_id", "tallanto_student_id")


def _owner50_has_person_contact_origin(
    customer_id: str,
    linked_person_origins: frozenset[str],
) -> bool:
    return customer_id in linked_person_origins


def _owner50_family_rows(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    as_of: datetime,
    price_axes_catalog: Mapping[str, Any] | Sequence[Any] | Path | str | None = None,
    family_ids: Sequence[str] = (),
) -> tuple[list[dict[str, Any]], list[tuple[str, ...]]]:
    required = ("family_members_v1", "family_links_v1", "customer_identities", "identity_links",
                "customer_opportunities", "derived_signals", "timeline_events", "timeline_conflicts")
    optional = ("customer_purchases_v1", "customer_objections_v1")
    # required+optional -- ОДНИМ запросом (не отдельными _table_exists) -- держит число
    # запросов на семью константным, см. test_owner50_bulk_selection_has_constant_query_count.
    available = {
        str(row["name"])
        for row in con.execute(
            f"SELECT name FROM sqlite_master WHERE type='table' AND name IN "
            f"({','.join('?' for _ in required + optional)})",
            required + optional,
        )
    }
    missing = [table for table in required if table not in available]
    if missing:
        raise RuntimeError(f"owner50 safety tables missing: {', '.join(missing)}")
    catalog = _owner50_resolve_price_axes_catalog(price_axes_catalog)
    snapshot = _owner50_snapshot(
        con,
        tenant_id=tenant_id,
        as_of=as_of,
        include_purchases="customer_purchases_v1" in available,
        include_objections="customer_objections_v1" in available,
    )
    selected_families = frozenset(str(value) for value in family_ids if str(value))
    grouped: dict[str, dict[str, list[sqlite3.Row]]] = defaultdict(lambda: defaultdict(list))
    for kind in ("signals", "members", "children", "opportunities", "events", "risk_signals", "purchases", "objections"):
        for row in snapshot[kind]:
            family_id = str(row["family_id"])
            if selected_families and family_id not in selected_families:
                continue
            grouped[family_id][kind].append(row)
    blocked_customer_ids = open_family_identity_conflict_customer_ids(con, tenant_id)
    exact_origins = authoritative_exact_identity_rows(
        con, tenant_id, link_types=OWNER50_PERSON_CONTACT_LINK_TYPES,
    )
    safe_origin_keys: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for row in exact_origins:
        if int(row["owner_count"] or 0) == 1 and not int(row["has_open_conflict"] or 0):
            safe_origin_keys[str(row["customer_id"])].add((str(row["link_type"]), str(row["link_value"])))
    linked_person_origins = frozenset(safe_origin_keys)
    conflicted_person_origins = frozenset(
        str(row["customer_id"])
        for row in exact_origins
        if int(row["owner_count"] or 0) > 1 or int(row["has_open_conflict"] or 0)
    )
    identity_names = {
        str(row["customer_id"]): _clean_text(row["display_name"])
        for row in snapshot["members"]
    }
    technical_labels: dict[str, set[str]] = defaultdict(set)
    for row in snapshot["opportunities"]:
        label = _clean_text(row["title"]).casefold()
        if label:
            technical_labels[str(row["customer_id"])].add(label)
    for row in snapshot["signals"]:
        signal_record = _safe_json(row["record_json"])
        for field in ("recommended_action", "next_step", "task_title"):
            label = _clean_text(signal_record.get(field)).casefold()
            if label:
                technical_labels[str(row["customer_id"])].add(label)
    verified_person_names = frozenset(
        str(event["customer_id"])
        for event in snapshot["events"]
        if not _clean_text(event["superseded_by"])
        and bool(identity_names.get(str(event["customer_id"]), ""))
        and _clean_text(event["subject"]) == identity_names.get(str(event["customer_id"]), "")
        and identity_names[str(event["customer_id"])].casefold()
        not in technical_labels.get(str(event["customer_id"]), set())
        and (
            ("amo_contact_id", str(event["source_id"]))
            if str(event["event_type"]) == "amo_contact_snapshot"
            else ("tallanto_student_id", str(event["source_id"]))
            if str(event["event_type"]) == "tallanto_student_snapshot"
            else ("", "")
        ) in safe_origin_keys.get(str(event["customer_id"]), set())
    )
    candidates: list[dict[str, Any]] = []
    control: list[tuple[str, ...]] = []
    for family_id, family in grouped.items():
        # требование аудиторов BLOCKED #4 (полная классификация семей): раньше семья без
        # сигналов, прошедших SQL-предфильтр, тихо пропускалась здесь ("if not signals:
        # continue") -- без единой строки в control. candidate_families в _owner50_snapshot
        # больше не завязан на наличие сигнала (см. её комментарий), поэтому signals ЗАКОННО
        # может быть пустым списком для семьи без активного сигнала вообще -- она всё равно
        # обязана получить статус (см. ветку "else" под циклом "for signal in signals" ниже:
        # signals=[] -> rejected остаётся пустым -> control получает
        # "no_active_outreach_signal", как и для семьи, у которой сигнал был, но не прошёл
        # проверки -- то же самое ужe протестированное поведение, просто теперь достижимое и
        # для пустого списка).
        signals = family["signals"]
        members = family["members"]
        member_by_id = {str(row["customer_id"]): row for row in members}
        member_texts = _dedupe_texts(
            f"{_clean_text(row['display_name']) or 'Имя не подтверждено'} [{row['customer_id']}]"
            for row in members
        )
        hard_reasons: list[str] = []
        missing_reasons: list[str] = []
        identity_brands: set[str] = set()
        unrecognized_brand_present = False
        for member in members:
            if str(member["membership_status"]) not in {"confident", "singleton"} or str(member["confidence"]) not in {"high", "medium"}:
                hard_reasons.append("family_ambiguous")
            if str(member["identity_status"] or "") != "strong":
                missing_reasons.append("identity_not_strong")
            record = _safe_json(member["identity_record_json"])
            metadata = _mapping(record.get("metadata"))
            raw_brands = {
                str(value).strip().casefold()
                for value in _plain_values(metadata.get("brands"))
                if str(value).strip()
            }
            if raw_brands - OWNER50_BRAND_ALIASES.keys():
                unrecognized_brand_present = True
            identity_brands.update(
                OWNER50_BRAND_ALIASES[value]
                for value in raw_brands
                if value in OWNER50_BRAND_ALIASES
            )
            no_contact = (
                record.get("no_contact"), record.get("opt_out"), metadata.get("no_contact"),
                metadata.get("opt_out"), metadata.get("do_not_contact"),
            )
            if any(str(value).strip().casefold() in {"1", "true", "yes", "да"} for value in no_contact) or str(
                metadata.get("contact_allowed", "true")
            ).casefold() in {"0", "false", "no", "нет"}:
                hard_reasons.append("structured_no_contact")
            roles = [member["customer_id"], member["display_name"], *(metadata.get(key) for key in ("role", "kind", "type", "tags"))]
            if OWNER50_STAFF_TEST_RE.search(" ".join(_plain_values(roles))):
                hard_reasons.append("staff_test_system")
            if str(member["customer_id"]) in blocked_customer_ids:
                hard_reasons.append("open_identity_conflict")
            if str(member["customer_id"]) in conflicted_person_origins:
                hard_reasons.append("open_identity_conflict")
            # требование E5 (26.07): identity_status=="strong" одного недостаточно -- READY
            # требует ещё и структурного person-contact origin (стабильный AMO-контакт или
            # Tallanto person ID). Его доказывает source_ref либо общий strong/manual индекс
            # identity_links; НЕ смысл display_name и не новый regex.
            if not _owner50_has_person_contact_origin(
                str(member["customer_id"]), linked_person_origins,
            ):
                missing_reasons.append("person_origin_unproven")
            if str(member["customer_id"]) not in verified_person_names:
                missing_reasons.append("human_name_unproven")

        opportunities = [
            row for row in family["opportunities"]
            if not row["closed_at"] and _is_active_deal(dict(row))
        ]
        if any(_owner50_structured_staff_test(row) for row in opportunities):
            hard_reasons.append("staff_test_system")
        product_offers = _product_interest_values(None, opportunities)
        offers = _dedupe_texts([*product_offers, *(_clean_text(row["title"]) for row in opportunities)])
        children = family["children"]
        if any(str(row["status"]) != "confident" or str(row["confidence"]) not in {"high", "medium"} for row in children):
            hard_reasons.append("child_ambiguous")
        verified_children = [
            row for row in children
            if str(row["status"]) == "confident" and str(row["confidence"]) in {"high", "medium"}
        ]
        child_texts = _dedupe_texts(
            f"{_clean_text(row['canonical_name'])} ({_join_list_json(row['grades_json'])}; {_join_list_json(row['subjects_json'])})"
            for row in verified_children
            if _clean_text(row["canonical_name"])
        )
        child_text_from_graph = bool(child_texts)
        # bug-fix owner50_pravki #1: "все дети выпускники" раньше проверялось по ПЛОСКОМУ
        # множеству классов всей семьи (grade_values) -- семья с 11-классником И
        # 6-классником ошибочно исключалась целиком, потому что 11 просто попадал в общий
        # набор. Правильно: считать выпускника ПО КАЖДОМУ ребёнку отдельно и исключать,
        # только если это истинно для ВСЕХ верифицированных детей (all(), не any()).
        child_grade_sets = [
            {
                int(value)
                for value in re.findall(r"(?<!\d)(?:[1-9]|10|11)(?!\d)", _join_list_json(row["grades_json"]))
            }
            for row in verified_children
        ]
        child_is_graduate = [
            11 in grades or bool(OWNER50_GRADUATE_RE.search(_clean_text(row["canonical_name"])))
            for row, grades in zip(verified_children, child_grade_sets)
        ]
        grade_values = {grade for grades in child_grade_sets for grade in grades}
        grade_values.update(int(match.group(1)) for offer in offers for match in OWNER50_OFFER_GRADE_RE.finditer(offer))
        if verified_children:
            if all(child_is_graduate):
                hard_reasons.append("grade_11_or_graduate")
            elif not any(1 <= grade <= 10 for grade in grade_values):
                missing_reasons.append("child_grade_unproven")
        elif 11 in grade_values or any(OWNER50_GRADUATE_RE.search(offer) for offer in offers):
            hard_reasons.append("grade_11_or_graduate")
        elif not any(1 <= grade <= 10 for grade in grade_values):
            missing_reasons.append("child_grade_unproven")
        grade_offer = next(
            (
                offer
                for offer in offers
                if any(1 <= int(match.group(1)) <= 10 for match in OWNER50_OFFER_GRADE_RE.finditer(offer))
            ),
            "",
        )
        if not child_texts and grade_offer:
            child_texts = [f"Класс из предложения: {grade_offer}"]
        raw_child_brands = {
            str(row["brand"]).strip().casefold()
            for row in verified_children
            if str(row["brand"]).strip()
        }
        if raw_child_brands - OWNER50_BRAND_ALIASES.keys():
            unrecognized_brand_present = True
        child_brands = {
            OWNER50_BRAND_ALIASES[value]
            for value in raw_child_brands
            if value in OWNER50_BRAND_ALIASES
        }
        # требование E1 (26.07): brand_unproven (0 брендов) убран отсюда -- это жёсткий
        # SQL-предфильтр, и "reasons" здесь означает EXCLUDED (см. "if reasons: ... continue"
        # ниже). Правило и classify_family требуют для 0 брендов CANDIDATE ("brand_unproven",
        # Г2), не EXCLUDED -- пустой brands обязан долететь до classify_family через
        # family_mapping["brands"] ниже, а не обрываться здесь. len(brands) > 1 остаётся
        # жёстким EXCLUDED ("brand_ambiguous") -- это настоящий конфликт, не "непроверено".
        brands = identity_brands | child_brands
        if len(brands) > 1:
            hard_reasons.append("brand_ambiguous")
        offer_brands = {
            brand
            for offer in offers
            for brand, pattern in OWNER50_BRAND_RE.items()
            if pattern.search(offer)
        }
        # требование E1 (продолжение): те же грабли, что и выше, но на сравнении с ПУСТЫМ
        # brands -- "оффер упоминает фотон/унпк, а подтверждённого бренда семьи ещё нет" не
        # конфликт (нечего конфликтовать), а тот же brand_unproven; "if brands and ...", как и
        # у _owner50_product_confirmed (та же логика для этого же случая чуть выше по файлу).
        if brands and offer_brands - brands:
            hard_reasons.append("brand_ambiguous")

        all_events = [_owner50_event(row) for row in family["events"]]
        events = [row for row in all_events if not _clean_text(row["superseded_by"])]
        event_by_id = {str(row["event_id"]): row for row in all_events}
        active_opportunities = {
            str(row["opportunity_id"]): row
            for row in opportunities
        }
        signal_quality_reasons: dict[str, list[str]] = defaultdict(list)
        for signal in signals:
            quality_reasons = signal_quality_reasons[str(signal["signal_id"])]
            signal_record = _safe_json(signal["record_json"])
            event_id = _clean_text(signal["event_id"] or signal_record.get("event_id"))
            event = event_by_id.get(event_id) if event_id else None
            signal_text = " ".join(
                filter(
                    None,
                    (
                        _clean_text(signal_record.get("evidence_text")),
                        _clean_text(signal_record.get("recommended_action")),
                    ),
                )
            ).casefold()
            event_text = _event_text(event) if event else ""
            if (
                _owner50_structured_staff_test(signal)
                or (event and _owner50_structured_staff_test(event))
                or OWNER50_STAFF_TEST_TEXT_RE.search(f"{signal_text} {event_text}")
                or (
                    event
                    and OWNER50_TEST_SOURCE_RE.search(
                        " ".join(
                            _clean_text(event.get(key))
                            for key in ("source_system", "source_id", "source_ref")
                        )
                    )
                )
            ):
                hard_reasons.append("staff_test_system")
            event_is_inbound = bool(event) and _clean_text(event.get("direction")).casefold() == "inbound"
            customer_risk_text = " ".join(
                filter(None, (_clean_text(signal_record.get("evidence_text")), event_text))
            ).casefold()
            if event_is_inbound and hard_codes_from_text(customer_risk_text):
                hard_reasons.append("durable_p0_history")
            if event_is_inbound and any(phrase in customer_risk_text for phrase in MANAGER_OPTOUT_PHRASES):
                hard_reasons.append("durable_opt_out")
            signal_brands = {
                brand
                for text in (signal_text, event_text)
                for brand, pattern in OWNER50_BRAND_RE.items()
                if pattern.search(text)
            }
            # требование E1 (продолжение): тот же guard -- пустой brands не "конфликтует" с
            # брендом, упомянутым в тексте сигнала/события, это по-прежнему brand_unproven.
            if brands and signal_brands - brands:
                hard_reasons.append("brand_ambiguous")
            if event_id and event is None:
                quality_reasons.append("signal_evidence_not_owned")
            elif event and str(event["customer_id"]) != str(signal["customer_id"]):
                quality_reasons.append("signal_evidence_not_owned")
            elif event and _clean_text(event["superseded_by"]):
                quality_reasons.append("signal_evidence_superseded")
            elif event and str(event["match_status"] or "") != "strong_unique":
                quality_reasons.append("signal_evidence_ambiguous")
            elif not event_id and str(signal["signal_type"]) != "season_return_candidate":
                quality_reasons.append("signal_evidence_missing")
            if event and not _owner50_event_evidence_text(event):
                quality_reasons.append("signal_evidence_text_missing")
            if str(signal["signal_type"]) == "deal_stalling":
                opportunity = active_opportunities.get(_clean_text(signal["opportunity_id"]))
                if not opportunity or str(opportunity["customer_id"]) != str(signal["customer_id"]):
                    quality_reasons.append("active_deal_missing")
        for event in events:
            if _owner50_event_is_explicit_refund(event):
                hard_reasons.append("durable_p0_history")
            if _clean_text(event.get("direction")).casefold() != "inbound":
                continue
            text = _event_text(event)
            event_p0_codes = hard_codes_from_text(text)
            # требование архитектора #6: старый ЗАКРЫТЫЙ refund не исключает навсегда --
            # см. _owner50_event_p0_is_stale_and_resolved (открытый/недавний/безответный
            # P0, а также legal/complaint/payment_dispute остаются жёстким EXCLUDED).
            if event_p0_codes and not _owner50_event_p0_is_stale_and_resolved(
                event, codes=event_p0_codes, all_events=all_events, as_of=as_of,
            ):
                hard_reasons.append("durable_p0_history")
            if any(phrase in text for phrase in MANAGER_OPTOUT_PHRASES):
                hard_reasons.append("durable_opt_out")
        hard_reasons.extend(f"active_risk_signal:{row['signal_type']}" for row in family["risk_signals"])
        # bug-fix owner50_pravki #3 (continued): dedupe_family_payment_rows схлопывает
        # сырые per-period строки в один аггрегат на customer_id (all_time побеждает
        # более узкие периоды того же клиента, никогда не складываются). Суммирование
        # РЕЗУЛЬТАТА по нескольким customer_id (по членам семьи) -- это НЕ повторное
        # дублирование, так и было задумано раньше.
        raw_purchase_rows = [dict(row) for row in family["purchases"]]
        purchase_by_customer = dedupe_family_payment_rows(raw_purchase_rows)
        total_in = sum(row["total_in"] for row in purchase_by_customer.values())
        total_out = sum(row["total_out"] for row in purchase_by_customer.values())
        deals_cnt = sum(row["deals_cnt"] for row in purchase_by_customer.values())
        last_purchase_dt = max(
            (row["last_purchase_at"] for row in purchase_by_customer.values() if row["last_purchase_at"]),
            default=None,
        )
        last_purchase_at = last_purchase_dt.isoformat() if last_purchase_dt else ""
        periods_by_customer: dict[str, set[str]] = defaultdict(set)
        for row in raw_purchase_rows:
            period = _clean_text(row.get("period"))
            if period:
                periods_by_customer[str(row.get("customer_id") or "")].add(period)
        payment_scope = "; ".join(
            f"{customer_id} [{', '.join(sorted(periods))}]"
            for customer_id, periods in periods_by_customer.items()
        )
        if hard_reasons:
            control.extend(
                _owner50_control_rows(
                    family_id, hard_reasons,
                    brand="; ".join(sorted(brands)) if brands else "",
                    family_members="; ".join(member_texts),
                )
            )
            continue
        payment_history = total_in > 0 and deals_cnt > 0
        specific_offer = bool(product_offers)
        child_tokens = {
            token[:5]
            for row in verified_children
            for token in re.findall(
                r"[a-zа-яё0-9]+",
                f"{_join_list_json(row['grades_json'])} {_join_list_json(row['subjects_json'])}".casefold(),
            )
            if token != "класс"
        }
        offer_tokens = {token[:5] for token in re.findall(r"[a-zа-яё0-9]+", " ".join(product_offers).casefold())}
        child_fit = bool(specific_offer and child_tokens & offer_tokens)
        # требование архитектора #7 (ужесточено по итогам ревью 25.07): самое свежее
        # возражение ВСЕЙ семьи, а не только владельца текущего сигнала -- свежий негатив
        # ДРУГОГО родителя обязан так же откатывать семью в CANDIDATE. family["objections"]
        # уже отсортирован DESC extracted_at единым tenant-wide запросом в _owner50_snapshot,
        # и порядок сохраняется при группировке по family_id -- значит первая строка это
        # самое свежее возражение по семье целиком, кто бы из членов его ни высказал.
        family_last_objection = family["objections"][0] if family["objections"] else None
        # требование архитектора #4: смешанная "11 + младший" СОХРАНЯЕТСЯ (см. graduate-логику
        # выше, bug-fix owner50_pravki #1), но предложение и подбор продукта адресованы ОДНОМУ
        # конкретному не-выпускнику -- требование E2 (26.07) заменило старую эвристику "минимум
        # из истории" (_owner50_select_target_child не выбирает минимальный класс из истории и
        # не гадает между несколькими РАЗНЫМИ подходящими детьми, см. её docstring).
        family_brand = next(iter(brands), "")
        # требование аудиторов BLOCKED #1 (fail-open доказательств): индекс РЕАЛЬНО
        # существующих записей для проверки event_id -- построен ИЗ РЕАЛЬНЫХ строк БД (events,
        # opportunities, платёжный аггрегат), а не из заявленных вызывающей стороной значений.
        # Раньше _owner50_family_rows подмешивал в такой индекс синтетическую запись
        # (events_by_id_for_classify.setdefault(classify_event_id, ...)) ПОСЛЕ того, как сама
        # же решала, какой id использовать -- проверка была самоссылочной (индекс подтверждал
        # то, что сам же и получил). Теперь known_records строится ЗАРАНЕЕ, независимо от
        # того, какой сигнал будет выбран ниже, и передаётся и в classify_family
        # (family_mapping["events_by_id"]), и в каждую строку листа EVIDENCE (единый источник
        # правды для обеих проверок).
        known_records: dict[str, dict[str, Any]] = {
            str(row["event_id"]): {"source_system": _clean_text(row["source_system"]), "at": row["event_at"]}
            for row in all_events
        }
        for opportunity in opportunities:
            known_records.setdefault(
                str(opportunity["opportunity_id"]),
                {"source_system": "customer_opportunities", "at": opportunity["opened_at"]},
            )
        # требование E4 (26.07): раньше композитный платёжный ключ регистрировался ТОЛЬКО для
        # period_used=="all_time" -- любая семья с несколькими вкладчиками/периодами не могла
        # разрешить payment evidence вообще. Теперь регистрируем КАЖДУЮ сырую строку
        # customer_purchases_v1 семьи (это всегда реальная строка БД, не выдумка) -- одиночный
        # вклад резолвится честно, а многострочный агрегат (см. "payment" evidence ниже) как и
        # раньше остаётся без единого event_id, когда вкладов несколько.
        for raw_purchase_row in raw_purchase_rows:
            raw_customer_id = str(raw_purchase_row.get("customer_id") or "")
            raw_period = _clean_text(raw_purchase_row.get("period"))
            if raw_customer_id and raw_period:
                known_records.setdefault(
                    f"customer_purchases_v1:{raw_customer_id}:{raw_period}:fact",
                    {"source_system": "customer_purchases_v1", "at": raw_purchase_row.get("last_purchase_at")},
                )
        # требование E4 (26.07): family_links_v1 -- реальная таблица с PRIMARY KEY (tenant_id,
        # customer_id, child_key) -- регистрируем ключ КАЖДОГО верифицированного ребёнка (не
        # только адресата), чтобы evidence "child" резолвилась в конкретную запись, а не в
        # family_id (запрет E4 -- family_id не является ID ребёнка).
        for verified_child_row in verified_children:
            child_customer_id = str(verified_child_row["customer_id"])
            child_key = _clean_text(verified_child_row["child_key"])
            if child_customer_id and child_key:
                known_records.setdefault(
                    f"family_links_v1:{child_customer_id}:{child_key}",
                    {"source_system": "family_links_v1", "at": verified_child_row["created_at"]},
                )
        rejected: list[str] = []
        signals.sort(key=lambda row: (
            OWNER50_SIGNAL_PRIORITY[str(row["signal_type"])],
            str(row["signal_type"]) != "callback_due",
            -(_parse_iso_datetime(row["created_at"]) or datetime.min.replace(tzinfo=timezone.utc)).timestamp(),
            str(row["signal_id"]),
        ))
        for signal in signals:
            customer_id = str(signal["customer_id"])
            signal_type = str(signal["signal_type"])
            precheck_reasons = signal_quality_reasons.get(str(signal["signal_id"]), ())
            if precheck_reasons:
                rejected.extend(precheck_reasons)
                continue
            if str(signal["identity_status"] or "") != "strong":
                rejected.append("identity_not_strong")
                continue
            signal_record = _safe_json(signal["record_json"])
            evidence_text = _clean_text(signal_record.get("evidence_text"))
            if not evidence_text:
                rejected.append("signal_evidence_text_missing")
                continue
            linked_opportunity = active_opportunities.get(_clean_text(signal["opportunity_id"]))
            if signal_type == "deal_stalling" and (
                not linked_opportunity or str(linked_opportunity["customer_id"]) != customer_id
            ):
                rejected.append("active_deal_missing")
                continue

            event_id = _clean_text(signal["event_id"] or signal_record.get("event_id"))
            event = event_by_id.get(event_id) if event_id else None
            if event and str(event["customer_id"]) != customer_id:
                rejected.append("signal_evidence_not_owned")
                continue
            if event and _clean_text(event["superseded_by"]):
                rejected.append("signal_evidence_superseded")
                continue
            if event and str(event["match_status"] or "") != "strong_unique":
                rejected.append("signal_evidence_ambiguous")
                continue
            evidence_at = _parse_iso_datetime(event["event_at"]) if event else None
            channel = _owner50_event_channel(event) if event else ""
            purchase = purchase_by_customer.get(customer_id)
            if signal_type == "season_return_candidate":
                if any(_is_access_event(item) for item in events):
                    rejected.append("active_access_or_learning")
                    continue
                purchase_at = _parse_iso_datetime(_mapping(signal_record.get("metadata")).get("last_purchase_at"))
                if not _owner50_purchase_matches(purchase, purchase_at, as_of=as_of):
                    rejected.append("season_purchase_not_confirmed")
                    continue
                evidence_at = purchase_at
                channel = "История оплат"
            elif event is None:
                rejected.append("signal_evidence_missing")
                continue
            cutoff = max(filter(None, (evidence_at, as_of - timedelta(days=30))), default=as_of - timedelta(days=30))
            if any(
                str(item["direction"] or "").casefold() == "outbound"
                and (_parse_iso_datetime(item["event_at"]) or as_of) > cutoff
                and not _is_non_closing_service_event(item)
                for item in events
            ):
                rejected.append("meaningful_outbound_after_evidence")
                continue
            contact = member_by_id.get(customer_id)
            if not contact or (not _clean_text(contact["primary_phone"]) and not _clean_text(contact["primary_email"])):
                rejected.append("contact_missing")
                continue

            due = signal_type == "callback_due"
            fresh_intent = signal_type == "client_returned"
            rank_reason = (
                f"tier={OWNER50_SIGNAL_PRIORITY[signal_type]}; due={int(due)}; fresh_intent={int(fresh_intent)}; "
                f"specific_offer={int(specific_offer)}; child_fit={int(child_fit)}; payment_history={int(payment_history)}"
            )

            # требование архитектора #2 (ужесточено по итогам ревью 25.07): signal
            # засчитывается только с event_id+source_system, разрешёнными в РЕАЛЬНУЮ запись.
            # season_return_candidate опирается на факт оплаты, а не на timeline_events;
            # customer_purchases_v1 не имеет суррогатного id, реальный составной ключ --
            # (customer_id, period, money_kind='fact'). Честная ссылка возможна ТОЛЬКО когда
            # purchase_by_customer[customer_id] это ОДНА настоящая строка (period_used=
            # "all_time" -- см. dedupe_family_payment_rows); "sum_of_periods" -- сумма
            # НЕСКОЛЬКИХ строк без единого event_id, выдумывать его нельзя -- честно оставляем
            # пустым (сигнал всё равно пройдёт по базовому evidence_text, просто без
            # дополнительной проверки provenance).
            if signal_type == "season_return_candidate":
                purchase_period = purchase_by_customer.get(customer_id, {}).get("period_used", "")
                if purchase_period == "all_time":
                    classify_event_id = f"customer_purchases_v1:{customer_id}:all_time:fact"
                    classify_source_system = "customer_purchases_v1"
                else:
                    classify_event_id = ""
                    classify_source_system = ""
            else:
                classify_event_id = event_id
                classify_source_system = _clean_text(event.get("source_system")) if event else ""
            # known_records (посчитан один раз на семью выше, до цикла по сигналам) уже несёт
            # все реальные events, платёжный композитный ключ каждой сырой строки и ключ
            # каждого верифицированного ребёнка -- требование аудиторов BLOCKED #1: больше
            # никакой самоссылочной подмешенной записи, только независимо построенный индекс.

            child_indexes = [
                index
                for index, child in enumerate(verified_children)
                if str(child["customer_id"]) == customer_id
            ] or list(range(len(verified_children)))
            target_child, target_grade, target_child_ambiguous = _owner50_select_target_child(
                [verified_children[index] for index in child_indexes],
                [child_grade_sets[index] for index in child_indexes],
                [child_is_graduate[index] for index in child_indexes],
            )
            target_child_key = _clean_text(target_child["child_key"]) if target_child is not None else ""
            target_child_name = _clean_text(target_child["canonical_name"]) if target_child is not None else ""

            # требование архитектора #3: продукт только по точному entry_id актуального
            # price_axes_catalog, адресован конкретному не-выпускнику (требование E2, 26.07).
            product_entry = _owner50_select_price_entry(
                catalog, brand=family_brand, grade=target_grade, offer_texts=offers,
            )
            product_dict = _owner50_product_from_price_entry(product_entry, brand=family_brand)
            if product_dict:
                # требование E4 (26.07): регистрируем НАЙДЕННЫЙ entry_id, чтобы evidence
                # "product" резолвилась в реальную запись каталога, а не оставалась вечно
                # нерезолвируемой (at может быть пуст у синтетических тестовых каталогов --
                # см. require_at=False у самой evidence-строки чуть ниже).
                known_records.setdefault(
                    str(product_dict["entry_id"]),
                    {
                        "source_system": "price_axes_catalog",
                        "at": product_dict.get("source_document_updated_at") or None,
                    },
                )

            # интерес засчитывается только если это цитата, привязанная к реальному событию
            # (требование #2) -- иначе просто не заполняем, а не выдумываем провенанс.
            interest_quote = None
            if (
                event
                and _clean_text(event.get("direction")).casefold() == "inbound"
                and INTEREST_MARKER_RE.search(_event_text(event))
            ):
                interest_quote = {
                    "text": _owner50_event_evidence_text(event),
                    "quoted_at": evidence_at.isoformat() if evidence_at else "",
                    "event_id": classify_event_id,
                    "source_system": classify_source_system,
                    "direction": "inbound",
                }

            last_objection = (
                {
                    "text": _clean_text(family_last_objection["quote_preview"]),
                    "at": family_last_objection["extracted_at"],
                }
                if family_last_objection is not None
                else None
            )

            # expires_at — срок жизни сигнала, а не обещанный срок действия менеджера.
            due_dt = next(
                (
                    parsed
                    for key in ("follow_up_due_at", "manager_followup_deadline", "deadline_at", "due_at")
                    if (parsed := _parse_iso_datetime(signal_record.get(key))) is not None
                ),
                None,
            )
            next_step_action = _clean_text(signal_record.get("recommended_action")) or OWNER50_NEXT_ACTION[signal_type]
            # требование E2 (26.07): classify_family решает "продукт подтверждён именно для
            # этого ребёнка" через _owner50_product_confirmed(product, brands, children, ...) --
            # раньше сюда шли ВСЕ верифицированные дети семьи (min(grades) из истории каждого),
            # из-за чего продукт мог "подтвердиться" по классу ЧУЖОГО ребёнка, не адресата.
            # Теперь -- ровно один выбранный _owner50_select_target_child адресат (его класс уже
            # однозначен по построению), либо пусто, если адресат не выбран (нет кандидата или
            # он неоднозначен -- target_child_ambiguous ниже, а не догадка).
            classify_children = (
                [{
                    "child_key": target_child_key,
                    "name": target_child_name,
                    "grade_current": target_grade,
                    "is_graduate": False,
                }]
                if target_child is not None
                else []
            )

            # Все булевые EXCLUDED-флаги classify_family ниже -- False по построению: если бы
            # хоть один был True, эта семья уже ушла бы через "if reasons:
            # control.extend(...); continue" выше и до этой точки не дошла бы. stale_data тоже
            # всегда False здесь (см. её комментарий чуть ниже -- гейт свежести теперь либо
            # проходит, либо останавливает весь build раньше, чем мы вообще сюда попадаем).
            # classify_family решает то, чего SQL-предфильтр не проверяет: свежесть сигнала,
            # next_step, продукт по прайсу, возражение-перевес (требования #1-2,7,8).
            family_mapping: dict[str, Any] = {
                "family_id": family_id,
                "identity": {
                    "customer_id": customer_id,
                    "identity_status": "strong",
                    "display_name": _clean_text(contact["display_name"]),
                },
                "missing": tuple(dict.fromkeys(missing_reasons)),
                "brands": brands,
                "unrecognized_brand_present": unrecognized_brand_present,
                "family_conflict": False,
                "contact_missing": False,
                "children": classify_children,
                "payment": {
                    "total_in": total_in,
                    "total_out": total_out,
                    "deals_cnt": deals_cnt,
                    "last_purchase_at": last_purchase_dt,
                },
                "interest_quote": interest_quote,
                "signal": {
                    "signal_type": signal_type,
                    "created_at": signal["created_at"],
                    "evidence_text": evidence_text,
                    "event_id": classify_event_id,
                    "source_system": classify_source_system,
                },
                "next_step": {
                    "action": next_step_action,
                    "due": due_dt.date().isoformat() if due_dt else "",
                },
                "product": product_dict,
                "target_child_ambiguous": target_child_ambiguous,
                "last_objection": last_objection,
                "events_by_id": known_records,
                "open_p0": False,
                "opt_out": False,
                "identity_conflict": False,
                "recent_meaningful_outbound_after_evidence": False,
                "active_recent_manager_work": False,
                "active_risk_signals": (),
                # требование аудиторов BLOCKED #3 (стейл -- стоп всей сборки, 25.07): раньше
                # сюда приходил флаг реального manager_freshness_gate (требование архитектора
                # #10), считался на КАЖДУЮ семью -- при непройденном гейте весь batch тихо
                # становился EXCLUDED построчно. Теперь непройденный гейт останавливает ВЕСЬ
                # build ДО вызова этой функции (raise в build_owner50_family_workbook) -- сюда
                # эта функция просто никогда не доходит со stale-данными, поэтому здесь всегда
                # False (не догадка, а факт: раз мы досюда дошли, гейт уже проверен снаружи).
                "stale_data": False,
            }
            classification = classify_family(family_mapping, as_of=as_of)
            status = classification["status"]

            evidence = [
                _owner50_evidence_item(
                    "signal", evidence_text, f"derived_signals:{signal['signal_id']}",
                    event_id=classify_event_id, source_system=classify_source_system, at=signal["created_at"],
                    known_records=known_records,
                )
            ]
            if event:
                evidence.append(_owner50_evidence_item(
                    "event", _owner50_event_evidence_text(event), f"timeline_events:{event_id}",
                    event_id=event_id, source_system=_clean_text(event.get("source_system")), at=event.get("event_at"),
                    known_records=known_records,
                ))
            for opportunity in opportunities:
                offer_evidence = _dedupe_texts([
                    *_product_interest_values(None, (opportunity,)), _clean_text(opportunity["title"])
                ])
                if offer_evidence:
                    evidence.append(_owner50_evidence_item(
                        "offer", "; ".join(offer_evidence), f"customer_opportunities:{opportunity['opportunity_id']}",
                        event_id=str(opportunity["opportunity_id"]), source_system="customer_opportunities",
                        at=opportunity["opened_at"], known_records=known_records,
                    ))
            # требование E4 (26.07): "child" -- одна строка НА КАЖДОГО верифицированного
            # ребёнка, event_id -- конкретный family_links_v1 composite key (customer_id +
            # child_key, реальный PRIMARY KEY этой таблицы), НЕ family_id (family_id -- не ID
            # ребёнка). Раньше была одна агрегированная строка на family_id, которая никогда не
            # резолвилась (family_id не встречался в known_records ни при каких условиях).
            for verified_child_row in verified_children:
                child_customer_id = str(verified_child_row["customer_id"])
                child_key = _clean_text(verified_child_row["child_key"])
                if not child_customer_id or not child_key:
                    continue
                child_event_id = f"family_links_v1:{child_customer_id}:{child_key}"
                evidence.append(_owner50_evidence_item(
                    "child",
                    f"{_clean_text(verified_child_row['canonical_name'])} "
                    f"({_join_list_json(verified_child_row['grades_json'])}; "
                    f"{_join_list_json(verified_child_row['subjects_json'])})",
                    child_event_id,
                    event_id=child_event_id, source_system="family_links_v1",
                    at=verified_child_row["created_at"], known_records=known_records,
                ))
            if raw_purchase_rows:
                # требование архитектора #9 (ужесточено по итогам ревью 25.07) + E4 (26.07):
                # family_id НЕ является ID оплаты -- текст по-прежнему суммирует
                # customer_purchases_v1 по НЕСКОЛЬКИМ customer_id/периодам семьи (payment_scope),
                # у такого агрегата в общем случае нет единого первичного ключа. Честно:
                # РОВНО один вклад (один customer_id, один period) -> используем его настоящий
                # композитный ключ (уже в known_records выше) -> резолвится; больше одного
                # вклада -> event_id пустой, build_evidence_record сам пометит resolvable=False,
                # а не выдуманное "подтверждено".
                single_purchase_row = raw_purchase_rows[0] if len(raw_purchase_rows) == 1 else None
                payment_event_id = (
                    f"customer_purchases_v1:{single_purchase_row.get('customer_id')}:"
                    f"{_clean_text(single_purchase_row.get('period'))}:fact"
                    if single_purchase_row is not None
                    else ""
                )
                evidence.append(_owner50_evidence_item(
                    "payment",
                    f"{payment_scope}; вход: {_format_money(total_in)}; выход: {_format_money(total_out)}; "
                    f"сделок: {deals_cnt}; последнее: {last_purchase_at}",
                    "customer_purchases_v1",
                    event_id=payment_event_id, source_system="customer_purchases_v1", at=last_purchase_dt,
                    known_records=known_records,
                ))
            if product_dict:
                # требование архитектора #3/#9 + E4 (26.07): точный entry_id прайс-каталога --
                # отдельная, проверяемая строка доказательства (не только текст оффера из CRM),
                # теперь резолвируемая (entry_id зарегистрирован в known_records выше).
                # require_at=False: каталожная запись -- действующий факт о цене, не датированное
                # событие ("дата, если есть время" -- ТЗ E4).
                evidence.append(_owner50_evidence_item(
                    "product", str(product_dict["name"]), f"price_axes_catalog:{product_dict['entry_id']}",
                    event_id=str(product_dict["entry_id"]), source_system="price_axes_catalog",
                    at=product_dict.get("source_document_updated_at") or None,
                    known_records=known_records, require_at=False,
                ))

            # требование E2 (26.07): "Предложение" никогда не берётся из старых названий сделок
            # AMO (они остаются evidence интереса -- см. evidence "offer" выше -- но не
            # предложением). Из выбранного KB/price entry -- когда продукт подтверждён; честное
            # "цена требует уточнения", когда строка нуждается в продукте, но КБ его не дал
            # (это же условие держит classification "product_not_confirmed_by_kb" в missing);
            # пусто -- когда продукт вообще не нужен (client_returned/callback_due с конкретным
            # действием, требование E2 "исключение"). next_step_ok_for_offer повторяет
            # классификаторский _owner50_is_concrete_next_step(...)+due -- те же чистые функции
            # модуля, без изменения контракта classify_family.
            next_step_ok_for_offer = _owner50_is_concrete_next_step(next_step_action) and due_dt is not None
            product_not_needed_for_offer = (
                _clean_text(next_step_action).casefold().startswith(_OWNER50_CLARIFY_INTEREST_PREFIX)
                or (signal_type in OWNER50_PRODUCT_OPTIONAL_SIGNALS and next_step_ok_for_offer)
            )
            if product_dict:
                offer_display = str(product_dict["name"])
            elif product_not_needed_for_offer:
                offer_display = ""
            else:
                offer_display = "Цена требует уточнения"

            row_common = {
                "family_id": family_id,
                "contact_customer_id": str(contact["customer_id"]),
                "name": _clean_text(contact["display_name"]),
                "phone": _clean_text(contact["primary_phone"]),
                "email": _clean_text(contact["primary_email"]),
                "brand": family_brand,
                "channel": channel,
                "evidence_at": evidence_at.isoformat() if evidence_at else "",
                "signal_type": signal_type,
                "evidence_text": evidence_text,
                "next_action": next_step_action,
                "expires_at": _clean_text(signal["expires_at"]),
                "offer": offer_display,
                "children": "; ".join(child_texts),
                "payment": (
                    f"Вход: {_format_money(total_in)}; выход: {_format_money(total_out)}; {deals_cnt} сделок"
                    if family["purchases"] else ""
                ),
                "rank_reason": rank_reason,
                "rank_key": (
                    OWNER50_SIGNAL_PRIORITY[signal_type], -int(due), -int(fresh_intent),
                    -int(specific_offer), -int(child_fit), -int(payment_history),
                    -(evidence_at or as_of).timestamp(), family_id,
                ),
                "evidence": evidence,
                "status": status,
                "action_text": classification["action_text"],
                "product_entry_id": str(product_dict["entry_id"]) if product_dict else "",
                "target_child_key": target_child_key,
                "target_child_name": target_child_name,
                "target_child_grade": str(target_grade) if target_grade is not None else "",
                "family_members": "; ".join(member_texts),
            }
            # требование архитектора #1: READY/CANDIDATE/EXCLUDED -- CANDIDATE никогда не
            # попадает в READY_50 (candidates), только в CANDIDATES со статусом candidate.
            # требование архитектора #10 (лист кандидатов неполноценный): контакт/дети/сигнал/
            # действие уже реально посчитаны выше (row_common) -- передаём их же, не догадки.
            # требование аудиторов BLOCKED #5 (пять листов, "рабочие колонки" CANDIDATES/
            # EXCLUDED): row_common уже несёт бренд/канал/дату основания/следующий шаг/
            # предложение/оплаты -- передаём их же в control, а не только family_id+код.
            if status == "READY":
                candidates.append(row_common)
            elif status == "CANDIDATE":
                control.extend(
                    _owner50_control_rows(
                        family_id, classification["missing"], status="candidate",
                        name=row_common["name"], phone=row_common["phone"], email=row_common["email"],
                        children=row_common["children"], signal_type=signal_type,
                        evidence_text=evidence_text, action_text=classification["action_text"],
                        brand=row_common["brand"], channel=row_common["channel"],
                        evidence_at=row_common["evidence_at"], next_action=row_common["next_action"],
                        offer=row_common["offer"], payment=row_common["payment"],
                        family_members=row_common["family_members"],
                    )
                )
            else:
                # защитный, штатно недостижимый путь: все exclusion-флаги family_mapping выше
                # заведомо False, поэтому classify_family может дойти сюда только через
                # fail-closed classification_error (см. classify_family docstring).
                control.extend(
                    _owner50_control_rows(
                        family_id, classification["reasons"] or ("classification_error:unknown",),
                        name=row_common["name"], phone=row_common["phone"], email=row_common["email"],
                        children=row_common["children"], signal_type=signal_type, evidence_text=evidence_text,
                        brand=row_common["brand"], channel=row_common["channel"],
                        evidence_at=row_common["evidence_at"], next_action=row_common["next_action"],
                        offer=row_common["offer"], payment=row_common["payment"],
                        family_members=row_common["family_members"],
                    )
                )
            break
        else:
            # нет ни одного сигнала, прошедшего проверки -- контакт для конкретного члена не
            # выбран (мог отличаться от сигнала к сигналу), но состав детей известен на уровне
            # семьи независимо от исхода сигналов. требование аудиторов BLOCKED #4: если
            # signals был пуст С САМОГО НАЧАЛА (rejected тоже пуст -- цикл не выполнил ни одной
            # итерации), это не "заблокирована", а "пока нечего предлагать" -- статус
            # "candidate" (как решил бы classify_family для одинокого "no_active_outreach_signal",
            # Г5). Когда сигнал БЫЛ, но НЕ прошёл проверку (rejected непустой), поведение то же,
            # что и раньше: "excluded".
            payment_summary = (
                f"Вход: {_format_money(total_in)}; выход: {_format_money(total_out)}; {deals_cnt} сделок"
                if family["purchases"] else ""
            )
            candidate_signal_reasons = {
                "active_deal_missing",
                "season_purchase_not_confirmed",
                "signal_evidence_ambiguous",
                "signal_evidence_missing",
                "signal_evidence_not_owned",
                "signal_evidence_superseded",
                "signal_evidence_text_missing",
            }
            control.extend(
                _owner50_control_rows(
                    family_id,
                    [*dict.fromkeys(missing_reasons), *(rejected or ["no_active_outreach_signal"])],
                    status=(
                        "candidate"
                        if not rejected or set(rejected).issubset(candidate_signal_reasons)
                        else "excluded"
                    ),
                    # требование E2 (26.07): нет выбранного сигнала -- продукт не резолвился
                    # вовсе, "Предложение" не может быть старым названием сделки AMO.
                    children="; ".join(child_texts), brand=family_brand,
                    offer="", payment=payment_summary,
                    family_members="; ".join(member_texts),
                )
            )
    return candidates, control


def _owner50_snapshot(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    as_of: datetime,
    include_purchases: bool,
    include_objections: bool = False,
) -> Mapping[str, list[sqlite3.Row]]:
    types = tuple(OWNER50_SIGNAL_PRIORITY)
    # Полный универсум идёт напрямую из family_members_v1. Активные сигналы только
    # обогащают семьи и не определяют, попадёт ли семья в классификацию.
    candidate_cte = f"""
        WITH candidate_signals AS (
          SELECT signal.*
          FROM derived_signals AS signal
          JOIN family_members_v1 AS member
            ON member.tenant_id=signal.tenant_id AND member.customer_id=signal.customer_id
          WHERE signal.tenant_id=? AND signal.status='active'
            AND signal.signal_type IN ({','.join('?' for _ in types)})
            AND (signal.expires_at IS NULL OR signal.expires_at='' OR julianday(signal.expires_at)>=julianday(?))
          ORDER BY CASE signal_type
                     WHEN 'callback_due' THEN 0 WHEN 'client_returned' THEN 1
                     WHEN 'deal_stalling' THEN 2 ELSE 3
                   END, created_at DESC, signal_id
          LIMIT ?
        )
    """
    base_params: tuple[Any, ...] = (
        tenant_id,
        *types,
        as_of.isoformat(),
        OWNER50_SIGNAL_SCAN_LIMIT + 1,
    )

    def fetch(sql: str, extra: Sequence[Any] = ()) -> list[sqlite3.Row]:
        return con.execute(f"{candidate_cte}\n{sql}", (*base_params, *extra)).fetchall()

    result = {
        "signals": fetch(
            """
            SELECT member.family_id, signal.*, identity.identity_status, identity.display_name,
                   identity.primary_phone, identity.primary_email,
                   identity.record_json AS identity_record_json
            FROM candidate_signals AS signal
            JOIN family_members_v1 AS member
              ON member.tenant_id=signal.tenant_id AND member.customer_id=signal.customer_id
            LEFT JOIN customer_identities AS identity
              ON identity.tenant_id=signal.tenant_id AND identity.customer_id=signal.customer_id
            """
        ),
        "members": fetch(
            """
            SELECT member.*, identity.identity_status, identity.display_name,
                   identity.primary_phone, identity.primary_email,
                   identity.record_json AS identity_record_json
            FROM family_members_v1 AS member
            LEFT JOIN customer_identities AS identity
              ON identity.tenant_id=member.tenant_id AND identity.customer_id=member.customer_id
            WHERE member.tenant_id=?
            LIMIT ?
            """,
            (tenant_id, OWNER50_RELATED_SCAN_LIMIT + 1),
        ),
        "children": fetch(
            """
            SELECT child.*
            FROM family_links_v1 AS child
            WHERE child.tenant_id=?
            LIMIT ?
            """,
            (tenant_id, OWNER50_RELATED_SCAN_LIMIT + 1),
        ),
        "opportunities": fetch(
            """
            SELECT member.family_id, opportunity.*
            FROM customer_opportunities AS opportunity
            JOIN family_members_v1 AS member
              ON member.tenant_id=opportunity.tenant_id AND member.customer_id=opportunity.customer_id
            WHERE opportunity.tenant_id=?
              AND opportunity.opportunity_type='amo_deal'
              AND opportunity.closed_at IS NULL
            ORDER BY opportunity.opened_at DESC
            LIMIT ?
            """,
            (tenant_id, OWNER50_RELATED_SCAN_LIMIT + 1),
        ),
        "events": fetch(
            """
            SELECT member.family_id, event.*
            FROM timeline_events AS event
            JOIN family_members_v1 AS member
              ON member.tenant_id=event.tenant_id AND member.customer_id=event.customer_id
            WHERE event.tenant_id=?
              AND COALESCE(json_extract(event.record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true')
              AND (
                event.direction IN ('inbound','outbound')
                OR event.event_type IN (
                  'amo_contact_snapshot','tallanto_student_snapshot','tallanto_group','tallanto_abonement'
                )
                OR json_extract(event.record_json,'$.record.module')='most_class'
                OR (
                  event.event_type='tallanto_payment'
                  AND lower(COALESCE(
                    json_extract(event.record_json,'$.record.payment_direction'),
                    json_extract(event.record_json,'$.record.direction'), ''
                  )) IN ('refund','return','возврат')
                )
                OR event.event_id IN (
                  SELECT signal.event_id FROM candidate_signals AS signal
                  WHERE signal.event_id IS NOT NULL AND signal.event_id!=''
                )
              )
            LIMIT ?
            """,
            (tenant_id, OWNER50_EVENT_SCAN_LIMIT + 1),
        ),
        "risk_signals": fetch(
            f"""
            SELECT member.family_id, risk.customer_id, risk.signal_type
            FROM derived_signals AS risk
            JOIN family_members_v1 AS member
              ON member.tenant_id=risk.tenant_id AND member.customer_id=risk.customer_id
            WHERE risk.tenant_id=? AND risk.status='active'
              AND risk.signal_type IN ({','.join('?' for _ in MANAGER_OUTREACH_RISK_SIGNAL_TYPES)})
              AND (risk.expires_at IS NULL OR risk.expires_at='' OR julianday(risk.expires_at)>=julianday(?))
            LIMIT ?
            """,
            (
                tenant_id,
                *MANAGER_OUTREACH_RISK_SIGNAL_TYPES,
                as_of.isoformat(),
                OWNER50_RELATED_SCAN_LIMIT + 1,
            ),
        ),
        "conflicts": con.execute(
            "SELECT record_json FROM timeline_conflicts WHERE tenant_id=? AND status IN ('open','active') LIMIT ?",
            (tenant_id, OWNER50_RELATED_SCAN_LIMIT + 1),
        ).fetchall(),
        "purchases": [],
        "objections": [],
    }
    if include_objections:
        # требование архитектора #7: "свежий позитив" перекрывает возражение, только
        # если он ПОЗЖЕ последнего возражения -- нужна и дата, и текст последнего
        # возражения семьи (client-side, самое уверенное сначала).
        result["objections"] = fetch(
            """
            SELECT member.family_id, objection.customer_id, objection.quote_preview,
                   objection.extracted_at, objection.objection_type
            FROM customer_objections_v1 AS objection
            JOIN family_members_v1 AS member
              ON member.tenant_id=objection.tenant_id AND member.customer_id=objection.customer_id
            WHERE objection.tenant_id=? AND objection.speaker='client'
            ORDER BY objection.extracted_at DESC
            LIMIT ?
            """,
            (tenant_id, OWNER50_RELATED_SCAN_LIMIT + 1),
        )
    if include_purchases:
        # bug-fix owner50_pravki #3: раньше здесь суммировалось SUM(total_in) по ВСЕМ
        # периодам без фильтра period -- если у одного customer_id одновременно были
        # строки period='all_time' И более узкого периода (обе money_kind='fact'), они
        # складывались (двойной счёт). Теперь отдаём СЫРЫЕ строки по периодам, а
        # схлопывание в один аггрегат на customer_id (all_time побеждает узкие периоды,
        # никогда не складываются) делает dedupe_family_payment_rows ниже по стеку.
        result["purchases"] = fetch(
            """
            SELECT member.family_id, purchase.customer_id, purchase.period,
                   purchase.total_in, purchase.total_out, purchase.deals_cnt,
                   purchase.last_purchase_at
            FROM customer_purchases_v1 AS purchase
            JOIN family_members_v1 AS member
              ON member.tenant_id=purchase.tenant_id AND member.customer_id=purchase.customer_id
            WHERE purchase.tenant_id=? AND purchase.money_kind='fact'
              AND purchase.computability='computed'
            ORDER BY purchase.customer_id, purchase.period
            LIMIT ?
            """,
            (tenant_id, OWNER50_RELATED_SCAN_LIMIT + 1),
        )
    if len(result["signals"]) > OWNER50_SIGNAL_SCAN_LIMIT:
        raise RuntimeError("owner50 candidate signal budget exceeded")
    if len(result["events"]) > OWNER50_EVENT_SCAN_LIMIT:
        raise RuntimeError("owner50 event budget exceeded")
    for kind in (
        "members", "children", "opportunities", "risk_signals", "conflicts",
        "purchases", "objections",
    ):
        if len(result[kind]) > OWNER50_RELATED_SCAN_LIMIT:
            raise RuntimeError(f"owner50 {kind} budget exceeded")
    return result


def _owner50_event(row: sqlite3.Row) -> dict[str, Any]:
    event = dict(row)
    stored = _safe_json(row["record_json"])
    event["record"] = _mapping(stored.get("record"))
    event["metadata"] = _mapping(stored.get("metadata"))
    return event


def _owner50_event_evidence_text(event: Mapping[str, Any]) -> str:
    return next(
        (
            text
            for text in (
                _clean_text(event.get("summary")),
                _clean_text(event.get("text_preview")),
                _clean_text(event.get("subject")),
            )
            if text
        ),
        "",
    )


def _owner50_structured_staff_test(row: Mapping[str, Any]) -> bool:
    payload = _safe_json(row["record_json"])
    nested = _mapping(payload.get("record"))
    sections = (payload, nested, _mapping(payload.get("metadata")), _mapping(nested.get("metadata")))
    for section in sections:
        if any(
            str(section.get(key, "")).strip().casefold() in {"1", "true", "yes", "да"}
            for key in ("is_test", "test", "is_staff", "staff", "is_system")
        ):
            return True
        values = [section.get(key) for key in ("role", "kind", "type", "tags")]
        if OWNER50_STAFF_TEST_RE.search(" ".join(_plain_values(values))):
            return True
    return False


def _owner50_event_channel(event: Mapping[str, Any] | None) -> str:
    if not event:
        return ""
    event_type = str(event.get("event_type") or "")
    source_system = str(event.get("source_system") or "")
    if event_type == "mango_call":
        return "Звонок"
    if event_type == "email_message":
        return "Email"
    if event_type in {"telegram_message", "telegram_dialog"} or source_system in {"telegram_history", "wappi_telegram"}:
        return "Telegram"
    if source_system == "wappi_max":
        return "MAX"
    return _display_source(source_system)


def _owner50_purchase_matches(
    row: sqlite3.Row | None,
    evidence_at: datetime | None,
    *,
    as_of: datetime,
) -> bool:
    stored_at = _parse_iso_datetime(row["last_purchase_at"]) if row else None
    return bool(
        row
        and evidence_at
        and float(row["total_in"] or 0) > 0
        and int(row["deals_cnt"] or 0) > 0
        and stored_at
        and stored_at <= as_of
        and evidence_at <= as_of
        and stored_at.date() == evidence_at.date()
    )


def _owner50_event_is_explicit_refund(event: Mapping[str, Any]) -> bool:
    if _clean_text(event.get("event_type")) != "tallanto_payment":
        return False
    record = _mapping(event.get("record"))
    return any(
        is_explicit_refund_direction(record.get(field))
        for field in ("payment_direction", "direction")
    )


# требование архитектора #10 (лист кандидатов неполноценный, 25.07) + требование аудиторов
# BLOCKED #5 (пять листов, 25.07): единый порядок "рабочих колонок" для CANDIDATES/EXCLUDED --
# family_id/status/code/reason_text + бренд/контакт/канал/дата основания/дети/сигнал/
# основание/следующий шаг/предложение/оплаты/действие, чтобы менеджер мог работать со строкой,
# не открывая другой лист. Статус остаётся колонкой (а не только именем листа), потому что
# CANDIDATES несёт ДВА статуса -- "candidate" и "outside_limit".
OWNER50_CONTROL_COLUMNS = (
    "family_id", "Статус", "Код причины", "Пояснение",
    "Бренд", "Контакт", "Телефон", "Email", "Канал", "Дата основания",
    "Дети", "Сигнал", "Основание", "Следующий шаг", "Предложение", "Оплаты", "Действие",
    "Члены семьи",
)


def _owner50_control_rows(
    family_id: str,
    reasons: Sequence[str],
    *,
    status: str = "excluded",
    name: str = "",
    phone: str = "",
    email: str = "",
    children: str = "",
    signal_type: str = "",
    evidence_text: str = "",
    action_text: str = "",
    brand: str = "",
    channel: str = "",
    evidence_at: str = "",
    next_action: str = "",
    offer: str = "",
    payment: str = "",
    family_members: str = "",
) -> list[tuple[str, ...]]:
    return [
        (
            family_id, status, reason, _owner50_reason_text(reason),
            brand, name, phone, email, channel, evidence_at,
            children, signal_type, evidence_text, next_action, offer, payment, action_text,
            family_members,
        )
        for reason in dict.fromkeys(reasons)
    ]


def _owner50_control_row_from_ready(row: Mapping[str, Any], *, status: str, code: str) -> tuple[str, ...]:
    """Строка control-листа для READY-семьи (selected/outside_limit) -- те же реальные поля,
    что уже посчитаны в row_common (_owner50_family_rows), не догадки."""
    return (
        row["family_id"], status, code, row["rank_reason"],
        row["brand"], row["name"], row["phone"], row["email"], row["channel"], row["evidence_at"],
        row["children"], row["signal_type"], row["evidence_text"], row["next_action"],
        row["offer"], row["payment"], row["action_text"], row["family_members"],
    )


def _owner50_reason_text(reason: str) -> str:
    # OWNER50_REASON_TEXT покрывает старые SQL-слой коды; owner50_tier_reason_text
    # (classify_family) покрывает новые -- вместе с active_risk_signal:*/
    # classification_error:* и никогда не возвращает пусто (по умолчанию сам код).
    return OWNER50_REASON_TEXT.get(reason) or owner50_tier_reason_text(reason)


def _full_dossier_segment_customer_ids(con: sqlite3.Connection, *, tenant_id: str, limit: int) -> list[str]:
    sql = """
        SELECT e.customer_id
        FROM timeline_events e
        JOIN customer_identities ci
          ON ci.tenant_id=e.tenant_id AND ci.customer_id=e.customer_id
        WHERE e.tenant_id = ?
          AND e.customer_id IS NOT NULL
          AND e.customer_id != ''
          AND (e.superseded_by IS NULL OR e.superseded_by = '')
          AND COALESCE(json_extract(e.record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true')
          AND ci.identity_status='strong'
          AND COALESCE(json_array_length(json_extract(ci.record_json, '$.metadata.brands')), 0)=1
          AND LOWER(json_extract(ci.record_json, '$.metadata.brands[0]')) IN ('foton','unpk')
        GROUP BY e.customer_id
        HAVING SUM(e.event_type = 'mango_call' AND e.match_status = 'strong_unique') > 0
           AND SUM(e.event_type = 'email_message') > 0
        ORDER BY MAX(e.event_at) DESC, e.customer_id
    """
    params: tuple[Any, ...]
    if limit > 0:
        sql += " LIMIT ?"
        params = (tenant_id, int(limit))
    else:
        params = (tenant_id,)
    return [str(row[0]) for row in con.execute(sql, params).fetchall()]


def _full_dossier_segment_count(con: sqlite3.Connection, *, tenant_id: str) -> int:
    row = con.execute(
        """
        SELECT COUNT(*) FROM (
          SELECT e.customer_id
          FROM timeline_events e
          JOIN customer_identities ci
            ON ci.tenant_id=e.tenant_id AND ci.customer_id=e.customer_id
          WHERE e.tenant_id = ?
            AND e.customer_id IS NOT NULL
            AND e.customer_id != ''
            AND (e.superseded_by IS NULL OR e.superseded_by = '')
            AND COALESCE(json_extract(e.record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true')
            AND ci.identity_status='strong'
            AND COALESCE(json_array_length(json_extract(ci.record_json, '$.metadata.brands')), 0)=1
            AND LOWER(json_extract(ci.record_json, '$.metadata.brands[0]')) IN ('foton','unpk')
          GROUP BY e.customer_id
          HAVING SUM(e.event_type = 'mango_call' AND e.match_status = 'strong_unique') > 0
             AND SUM(e.event_type = 'email_message') > 0
        )
        """,
        (tenant_id,),
    ).fetchone()
    return int(row[0] or 0) if row else 0


def _table_exists(con: sqlite3.Connection, table: str) -> bool:
    row = con.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return row is not None


def _read_json(path: Path | None) -> Mapping[str, Any]:
    if path is None:
        return {}
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, Mapping) else {}


def _source_freshness(con: sqlite3.Connection, *, tenant_id: str = "foton") -> list[Mapping[str, Any]]:
    if not _table_exists(con, "timeline_events"):
        return []
    return source_freshness_rows(
        con,
        tenant_id=tenant_id,
        expected_sources=MANAGER_REQUIRED_SOURCE_SYSTEMS,
    )


def _actuality_header(freshness: Sequence[Mapping[str, Any]], reconcile: Mapping[str, Any]) -> str:
    cursor_text = "; ".join(
        f"{_display_freshness_source(row.get('source_system'))}={row.get('cursor_at') or 'нет курсора'}"
        for row in freshness[:8]
    ) or "нет данных"
    cursor_checked_text = "; ".join(
        f"{_display_freshness_source(row.get('source_system'))}={row.get('cursor_updated_at') or 'нет проверки'}"
        for row in freshness[:8]
    ) or "нет данных"
    event_text = "; ".join(
        f"{_display_freshness_source(row.get('source_system'))}={row.get('max_event_at')}"
        for row in freshness[:8]
    ) or "нет данных"
    imported_text = "; ".join(
        f"{_display_freshness_source(row.get('source_system'))}={row.get('imported_at') or 'нет успешного импорта'}"
        for row in freshness[:8]
    ) or "нет данных"
    status = str(reconcile.get("status") or "")
    if status == "checked":
        reconcile_text = (
            f"{reconcile.get('generated_at')}; "
            f"{reconcile.get('customers_changed')} расхождений из {reconcile.get('customers_checked')}; "
            f"snapshot_stale={reconcile.get('snapshot_stale')}"
        )
    elif reconcile:
        reconcile_text = f"не проводилась ({reconcile.get('reason') or status or 'unknown'})"
    else:
        reconcile_text = "не проводилась"
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    return (
        f"Данные: cursor_at по источникам: {cursor_text}; "
        f"cursor checked_at отдельно: {cursor_checked_text}; "
        f"imported_at отдельно: {imported_text}; "
        f"max event_at отдельно: {event_text}; собрано {generated_at}; "
        f"сверка с живым AMO: {reconcile_text}"
    )


def _display_freshness_source(source: Any) -> str:
    mapping = {
        "amocrm_snapshot": "AMO снимок",
        "amocrm_event": "AMO события",
        "amocrm_price_readonly": "AMO цены",
        "mango_processed_summary": "сводки звонков",
        "mail_archive": "архив почты",
        "mail_archive_stage2": "письма",
        "tallanto_crm_call": "Tallanto платежи",
        "master_contacts_snapshot": "сводка контактов",
        "tallanto_snapshot": "Tallanto снимок",
        "telegram_history": "Telegram история",
    }
    return mapping.get(str(source or ""), _clean_text(source) or "неизвестный источник")


def _family_rows(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_id: str,
    active_brand: str,
) -> list[DossierRow]:
    if not _table_exists(con, "family_links_v1"):
        return []
    customer_ids = _family_scope_customer_ids(con, tenant_id=tenant_id, customer_id=customer_id)
    placeholders = ",".join("?" for _ in customer_ids)
    rows = con.execute(
        f"""
        SELECT canonical_name, name_variants_json, grades_json, subjects_json, brand, status, confidence, reason
        FROM family_links_v1
        WHERE tenant_id = ? AND customer_id IN ({placeholders})
        ORDER BY status, confidence DESC, canonical_name
        """,
        (tenant_id, *customer_ids),
    ).fetchall()
    result: list[DossierRow] = []
    for row in rows:
        variants = _join_list_json(row["name_variants_json"])
        grades = _join_list_json(row["grades_json"])
        subjects = _join_list_json(row["subjects_json"])
        quality = f"{row['status']}/{row['confidence']}"
        text = f"{_clean_text(row['canonical_name'])}"
        details = []
        if variants and variants != text:
            details.append(f"варианты: {variants}")
        if grades:
            details.append(f"класс: {grades}")
        if subjects:
            details.append(f"предметы: {subjects}")
        if row["brand"]:
            details.append(f"бренд: {row['brand']}")
            if active_brand and str(row["brand"]).casefold() != active_brand.casefold():
                details.append("исторический другой бренд — не переносить в текущее предложение")
        if str(row["status"]) != "confident" or str(row["confidence"]) not in {"high", "medium"}:
            details.append("уточнить семейную связь")
        if details:
            text += " (" + "; ".join(details) + ")"
        result.append(DossierRow("Семья", text, f"family_links_v1:{quality}:{row['reason']}"))
    return result


def _money_rows(con: sqlite3.Connection, *, tenant_id: str, customer_id: str) -> list[DossierRow]:
    if not _table_exists(con, "customer_purchases_v1"):
        return []
    customer_ids = _family_scope_customer_ids(con, tenant_id=tenant_id, customer_id=customer_id)
    placeholders = ",".join("?" for _ in customer_ids)
    rows = con.execute(
        f"""
        SELECT purchase.customer_id, identity.display_name, purchase.period, purchase.money_kind,
               purchase.total_in, purchase.total_out, purchase.deals_cnt,
               purchase.last_purchase_at, purchase.computability
        FROM customer_purchases_v1 AS purchase
        LEFT JOIN customer_identities AS identity
          ON identity.tenant_id = purchase.tenant_id AND identity.customer_id = purchase.customer_id
        WHERE purchase.tenant_id = ? AND purchase.customer_id IN ({placeholders})
        ORDER BY purchase.customer_id, purchase.period, purchase.money_kind
        """,
        (tenant_id, *customer_ids),
    ).fetchall()
    result: list[DossierRow] = []
    labels = {"fact": "факт оплат", "plan": "план сделок"}
    for row in rows:
        kind = str(row["money_kind"] or "plan")
        label = labels.get(kind, kind)
        text = (
            f"{label}, период {row['period']}: вход {_format_money(row['total_in'])}; "
            f"списания/расход {_format_money(row['total_out'])}; сделок {int(row['deals_cnt'] or 0)}"
        )
        if row["last_purchase_at"]:
            text += f"; последнее событие {row['last_purchase_at']}"
        if row["computability"]:
            text += f"; вычислимость {row['computability']}"
        if len(customer_ids) > 1:
            member = _clean_text(row["display_name"]) or str(row["customer_id"])
            text += f" [карточка: {member}]"
        result.append(DossierRow("Деньги", text, "customer_purchases_v1"))
    return result


def _signal_rows(con: sqlite3.Connection, *, tenant_id: str, customer_id: str) -> list[DossierRow]:
    if not _table_exists(con, "derived_signals"):
        return []
    rows = con.execute(
        """
        SELECT signal_type, severity, expires_at, confidence, requires_manager_review, record_json
        FROM derived_signals
        WHERE tenant_id = ? AND customer_id = ? AND status = 'active'
          AND (expires_at IS NULL OR expires_at = '' OR julianday(expires_at) >= julianday('now'))
        ORDER BY CASE severity
                   WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 WHEN 'low' THEN 3 ELSE 4
                 END,
                 expires_at, signal_type
        LIMIT 12
        """,
        (tenant_id, customer_id),
    ).fetchall()
    result: list[DossierRow] = []
    for row in rows:
        record = _safe_json(row["record_json"])
        action = _clean_text(record.get("recommended_action") or record.get("action") or "")
        evidence = _clean_text(record.get("evidence_text") or record.get("reason") or "")
        label = _signal_label(str(row["signal_type"] or ""))
        parts = [label]
        if row["severity"]:
            parts.append(f"важность: {row['severity']}")
        if row["expires_at"]:
            parts.append(f"до: {row['expires_at']}")
        if action and _meaningful_next_step(action):
            parts.append(f"рекомендация: {action}")
        if evidence:
            parts.append(f"основание: {evidence}")
        result.append(DossierRow("Сигналы", "; ".join(parts), f"derived_signals:{row['signal_type']}"))
    return result


def _next_step_from_signals(signals: Sequence[DossierRow]) -> str:
    for signal in signals:
        match = re.search(r"рекомендация:\s*([^;]+)", signal.text)
        if not match:
            continue
        value = _clean_text(match.group(1))
        if _meaningful_next_step(value):
            return value
    return ""


def _next_step_for_dossier(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_id: str,
    signals: Sequence[DossierRow],
) -> tuple[str, str, str, str]:
    rows = con.execute(
        """
        SELECT event_id, customer_id, event_at, event_type, source_system, source_id,
               source_ref, subject, summary, text_preview, direction, record_json
        FROM timeline_events
        WHERE tenant_id = ? AND customer_id = ?
          AND (superseded_by IS NULL OR superseded_by = '')
          AND COALESCE(json_extract(record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true')
          AND (event_type != 'mango_call' OR match_status = 'strong_unique')
        ORDER BY event_at DESC, event_id DESC
        LIMIT 500
        """,
        (tenant_id, customer_id),
    ).fetchall()
    events: list[Mapping[str, Any]] = []
    for row in rows:
        stored = _safe_json(row["record_json"])
        event = dict(row)
        event["record"] = dict(stored["record"]) if isinstance(stored.get("record"), Mapping) else {}
        event["metadata"] = dict(stored["metadata"]) if isinstance(stored.get("metadata"), Mapping) else {}
        event["stage_before"] = stored.get("stage_before")
        event["stage_after"] = stored.get("stage_after")
        events.append(event)
    conflicts: list[Mapping[str, Any]] = []
    if _table_exists(con, "timeline_conflicts"):
        customer_refs = set(customer_entity_ref_values(customer_id))
        for row in con.execute(
            "SELECT conflict_type, status, record_json FROM timeline_conflicts WHERE tenant_id = ? AND status = 'open'",
            (tenant_id,),
        ).fetchall():
            record = dict(_safe_json(row["record_json"]))
            entity_refs = {str(item) for item in (record.get("entity_refs") or ())}
            if customer_refs.isdisjoint(entity_refs):
                continue
            record.setdefault("conflict_type", row["conflict_type"])
            record.setdefault("status", row["status"])
            conflicts.append(record)
    if customer_id in open_family_identity_conflict_customer_ids(con, tenant_id) and not any(
        "ambiguous_identity" in str(conflict.get("conflict_type") or "").casefold()
        for conflict in conflicts
    ):
        conflicts.append(
            {
                "conflict_type": "ambiguous_identity",
                "status": "open",
                "summary": "canonical customer or family conflict",
            }
        )
    resolved = resolve_customer_next_step(
        events,
        readiness={"open_conflicts": len(conflicts)},
        conflicts=conflicts,
        customer_id=customer_id,
    )
    if resolved.status == NEXT_STEP_STATUS_ACTIVE and _meaningful_next_step(resolved.action):
        return resolved.display_text, "timeline_events", resolved.status, ""
    if resolved.status != NEXT_STEP_STATUS_EMPTY:
        return "", "", resolved.status, resolved.reason_code
    fallback = _next_step_from_signals(signals)
    if fallback:
        return fallback, "derived_signals", NEXT_STEP_STATUS_ACTIVE, ""
    return "", "", resolved.status, resolved.reason_code


def _meaningful_next_step(value: str) -> bool:
    text = value.casefold()
    if not text or text in {"уточнить у менеджера", "связаться с клиентом", "позвонить клиенту"}:
        return False
    if "посмотреть историю" in text:
        return False
    return len(text.split()) >= 3


def _objection_rows(con: sqlite3.Connection, *, tenant_id: str, customer_id: str) -> list[DossierRow]:
    if not _table_exists(con, "customer_objections_v1"):
        return []
    rows = con.execute(
        """
        SELECT source_channel, objection_type, quote_preview, budget_hint_rub, price_sensitivity, confidence, speaker
        FROM customer_objections_v1
        WHERE tenant_id = ?
          AND customer_id = ?
          AND speaker = 'client'
        ORDER BY confidence DESC, extracted_at DESC
        LIMIT 12
        """,
        (tenant_id, customer_id),
    ).fetchall()
    result: list[DossierRow] = []
    for row in rows:
        quote = _safe_marker_phrase(row["quote_preview"], re.compile(r".+"))
        text = f"{row['objection_type']}: {quote}"
        if row["budget_hint_rub"]:
            text += f"; бюджет: {_format_money(row['budget_hint_rub'])}"
        if row["price_sensitivity"]:
            text += f"; чувствительность к цене: {row['price_sensitivity']}"
        result.append(DossierRow("Возражения", text, f"customer_objections_v1:{row['source_channel']}:{row['confidence']}"))
    return result


def _chronology_rows(con: sqlite3.Connection, *, tenant_id: str, customer_id: str, limit: int) -> list[DossierRow]:
    customer_ids = _family_scope_customer_ids(
        con,
        tenant_id=tenant_id,
        customer_id=customer_id,
    )
    placeholders = ",".join("?" for _ in customer_ids)
    rows = con.execute(
        f"""
        SELECT event.event_at, event.event_type, event.source_system, event.subject,
               event.summary, event.text_preview, event.record_json, event.customer_id,
               identity.display_name AS source_customer_name
        FROM timeline_events AS event
        LEFT JOIN customer_identities AS identity
          ON identity.tenant_id = event.tenant_id AND identity.customer_id = event.customer_id
        WHERE event.tenant_id = ?
          AND event.customer_id IN ({placeholders})
          AND (event.superseded_by IS NULL OR event.superseded_by = '')
          AND COALESCE(json_extract(event.record_json,'$.metadata.pending_attribution'),0) NOT IN (1,'true')
          AND (event.event_type != 'mango_call' OR event.match_status = 'strong_unique')
        ORDER BY event.event_at DESC, event.event_id DESC
        LIMIT ?
        """,
        (tenant_id, *customer_ids, int(limit)),
    ).fetchall()
    result: list[DossierRow] = []
    for row in rows:
        summary = _event_summary_for_manager(row)
        if not summary:
            continue
        text = f"{row['event_at']} [{row['event_type']}] {summary}"
        if len(customer_ids) > 1:
            member = _clean_text(row["source_customer_name"]) or str(row["customer_id"])
            text = f"{text} [карточка: {member}]"
        result.append(
            DossierRow(
                "Хронология",
                text,
                f"{row['source_system']}:{row['customer_id']}",
            )
        )
    return result



def _family_scope_customer_ids(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_id: str,
) -> tuple[str, ...]:
    if not _table_exists(con, "family_members_v1"):
        return (customer_id,)
    root = con.execute(
        """
        SELECT family_id, membership_status
        FROM family_members_v1
        WHERE tenant_id = ? AND customer_id = ?
        """,
        (tenant_id, customer_id),
    ).fetchone()
    if root is None or str(root["membership_status"] or "") not in {"confident", "singleton"}:
        return (customer_id,)
    members = tuple(
        str(row["customer_id"])
        for row in con.execute(
            """
            SELECT customer_id
            FROM family_members_v1
            WHERE tenant_id = ? AND family_id = ?
              AND membership_status IN ('confident', 'singleton')
            ORDER BY customer_id
            """,
            (tenant_id, root["family_id"]),
        )
    )
    if customer_id not in members or not 1 <= len(members) <= 8:
        return (customer_id,)
    return members


def _event_summary_for_manager(row: sqlite3.Row) -> str:
    event_type = str(row["event_type"] or "")
    subject = _clean_text(row["subject"])
    summary = _clean_text(row["summary"]) or _clean_text(row["text_preview"])
    if event_type == "email_message":
        if EMAIL_SUMMARY_REVIEW_NEEDED_RE.search(summary):
            summary = f"Письмо «{subject or 'без темы'}»: полный текст в базе."
        elif summary:
            summary = f"{summary} Полный текст в базе."
        elif subject:
            summary = f"Письмо «{subject}»: полный текст в базе."
    elif event_type == "tallanto_attendance" and subject:
        summary = f"Запись Tallanto о занятии: {subject}."
    return summary


def _signal_label(signal_type: str) -> str:
    labels = {
        "client_returned": "клиент вернулся",
        "callback_due": "нужно перезвонить",
        "deal_stalling": "сделка зависла",
        "hot_streak": "горячая серия касаний",
        "season_return_candidate": "сезонный возврат",
    }
    return labels.get(signal_type, signal_type)


def _format_money(value: Any) -> str:
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0.0
    return f"{amount:,.0f} руб.".replace(",", " ")


def _join_list_json(raw: Any) -> str:
    value = _json_any(raw) if isinstance(raw, str) else raw
    if isinstance(value, list):
        return ", ".join(_clean_text(item) for item in value if _clean_text(item))
    return _clean_text(value)


def _json_any(value: str | None) -> Any:
    if not value:
        return None
    try:
        return json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return None


def _product_interest_values(customer_record_json: str | None, opportunities: Sequence[sqlite3.Row]) -> tuple[str, ...]:
    values: list[str] = []
    values.extend(_recursive_product_values(_safe_json(customer_record_json)))
    for row in opportunities:
        payload = _safe_json(row["record_json"])
        values.extend(_recursive_product_values(payload))
        values.extend(_plain_values(payload.get("product_context") if isinstance(payload, Mapping) else None))
    return tuple(_dedupe_texts(_safe_phrase(value) for value in values if _safe_phrase(value)))


def _recursive_product_values(value: Any) -> list[str]:
    result: list[str] = []
    if isinstance(value, Mapping):
        for key, nested in value.items():
            key_text = str(key).strip().casefold()
            if key_text in PRODUCT_KEYS:
                result.extend(_plain_values(nested))
            else:
                result.extend(_recursive_product_values(nested))
    elif isinstance(value, list):
        for item in value:
            result.extend(_recursive_product_values(item))
    return result


def _plain_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [value]
    if isinstance(value, Mapping):
        result: list[str] = []
        for key in (
            "title", "name", "course", "group", "filial", "subject", "subject_name",
            "format", "class", "value",
        ):
            result.extend(_plain_values(value.get(key)))
        return result
    if isinstance(value, (list, tuple)):
        result: list[str] = []
        for item in value:
            result.extend(_plain_values(item))
        return result
    return []


def _markers_from_client_text(text: str, pattern: re.Pattern[str], *, kind: str, label: str, source: str) -> list[DossierMarker]:
    markers: list[DossierMarker] = []
    seen: set[str] = set()
    for sentence in _sentences(text):
        if not pattern.search(sentence):
            continue
        phrase = _safe_marker_phrase(sentence, pattern)
        if kind == "interest" and not INTEREST_CONTEXT_RE.search(phrase):
            continue
        if not phrase or phrase.casefold() in seen:
            continue
        seen.add(phrase.casefold())
        markers.append(DossierMarker(kind=kind, text=f"{label}: {phrase}", source=source))
    return markers


def _sentences(text: str) -> list[str]:
    compact = _clean_text(text)
    if not compact:
        return []
    parts = re.split(r"(?<=[.!?…])\s+|\n+", compact)
    return [part.strip(" -–—\t") for part in parts if part.strip(" -–—\t")]


def _safe_phrase(value: Any) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    text = CONTACT_RE.sub("[contact]", text)
    text = text.strip(" .;:,")
    return text[:220]


def _safe_marker_phrase(value: Any, pattern: re.Pattern[str]) -> str:
    text = CONTACT_RE.sub("[contact]", _clean_text(value))
    if not text:
        return ""
    text = _trim_to_marker_window(text, pattern)
    text = SPEECH_FILLER_RE.sub("", text).strip(" -–—,.;:")
    text = _collapse_repeated_words(text)
    text = _trim_to_first_meaningful_clause(text)
    text = _trim_to_word_boundary(text, 160)
    if not text:
        return ""
    text = text[0].upper() + text[1:]
    if text[-1] not in ".!?…":
        text += "."
    return text


def _trim_to_marker_window(text: str, pattern: re.Pattern[str], *, after: int = 110) -> str:
    match = _select_marker_match(text, pattern)
    if match is None:
        return text
    start = _marker_window_start(text, match.start())
    end = min(len(text), match.end() + after)
    window = text[start:end].strip(" -–—,.;:")
    return window


def _select_marker_match(text: str, pattern: re.Pattern[str]) -> re.Match[str] | None:
    matches = list(pattern.finditer(text))
    if not matches:
        return None
    for index, match in enumerate(matches):
        token = match.group(0).casefold()
        tail = text[match.end() : match.end() + 80].casefold()
        if token.startswith("хотел") and "звон" in tail and index + 1 < len(matches):
            continue
        return match
    return matches[0]


def _marker_window_start(text: str, marker_start: int) -> int:
    prefix = text[:marker_start]
    words = list(re.finditer(r"\b[А-Яа-яA-Za-zЁё]{1,8}\b", prefix))
    if not words:
        return marker_start
    previous = words[-1]
    gap = prefix[previous.end() :].strip(" ,.;:–—-")
    if gap:
        return marker_start
    if previous.group(0).casefold() in {"нас", "нам", "мы", "мне", "меня"}:
        return previous.start()
    return marker_start


def _trim_to_first_meaningful_clause(text: str) -> str:
    boundary = SPEECH_CLAUSE_BOUNDARY_RE.search(text)
    if boundary and boundary.start() >= 12:
        text = text[: boundary.start()]
    text = re.split(r"\s+(?:но|а|и)\s+", text, maxsplit=1, flags=re.I)[0]
    return text.strip(" -–—,.;:")


def _trim_to_word_boundary(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text.strip(" -–—,.;:")
    chunk = text[:limit].rstrip()
    cut = max(chunk.rfind(" "), chunk.rfind(","), chunk.rfind(";"), chunk.rfind("."))
    if cut >= int(limit * 0.55):
        chunk = chunk[:cut]
    return chunk.strip(" -–—,.;:")


def _collapse_repeated_words(text: str) -> str:
    parts = text.split()
    result: list[str] = []
    previous = ""
    for part in parts:
        key = part.strip(" ,.;:!?").casefold()
        if key and key == previous:
            continue
        result.append(part)
        previous = key
    return " ".join(result)


def _clean_text(value: Any) -> str:
    return WHITESPACE_RE.sub(" ", str(value or "").replace("\u00a0", " ")).strip()


def _safe_json(value: str | None) -> Mapping[str, Any]:
    if not value:
        return {}
    try:
        payload = json.loads(value)
    except (TypeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, Mapping) else {}


def _mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def _parse_iso_datetime(value: Any) -> datetime | None:
    text = _clean_text(value)
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed.astimezone(timezone.utc)


def _dedupe_texts(values: Sequence[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = value.casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(value)
    return result


def _dedupe_markers(values: Sequence[DossierMarker], *, limit: int) -> list[DossierMarker]:
    result: list[DossierMarker] = []
    seen: set[str] = set()
    for item in values:
        key = item.text.casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(item)
        if len(result) >= limit:
            break
    return result


def _write_owner50_workbook(
    path: Path,
    rows: Sequence[Mapping[str, Any]],
    control: Sequence[tuple[str, ...]],
    control_meta: Mapping[str, Any],
) -> None:
    """Требование аудиторов BLOCKED #5 (пять отдельных рабочих листов, 25.07): раньше было три
    листа ("Кому писать"/"Доказательства"/"Контроль"), и "Контроль" смешивал CANDIDATE и
    EXCLUDED в одном месте, различимые только колонкой "Статус". Теперь -- пять листов с
    литеральными именами READY_50/CANDIDATES/EXCLUDED/EVIDENCE/CONTROL: READY_50 -- прежнее
    "Кому писать" (топ-50 READY); CANDIDATES -- статусы candidate+outside_limit (READY, не
    попавшие в топ-50, тоже сюда -- они не "исключены", им просто не хватило места); EXCLUDED --
    статус excluded с причиной; EVIDENCE -- прежнее "Доказательства" (event_id/source/дата на
    каждую строку READY_50); CONTROL -- теперь МЕТА-лист сборки (свежесть источников, счётчики
    READY/CANDIDATES/EXCLUDED, версия каталога цен), а не построчные причины -- те переехали на
    CANDIDATES/EXCLUDED."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ready = wb.active
    ready.title = "READY_50"
    candidates_ws = wb.create_sheet("CANDIDATES")
    excluded_ws = wb.create_sheet("EXCLUDED")
    evidence = wb.create_sheet("EVIDENCE")
    control_ws = wb.create_sheet("CONTROL")

    ready.append(OWNER50_REQUIRED_COLUMNS)
    # требование архитектора #9: доказательство несёт date+source_system+event_id отдельными
    # колонками (не только человекочитаемый "Источник") -- см. _owner50_evidence_item.
    evidence.append((
        "family_id", "Ранг", "Тип", "Доказательство", "Источник",
        "Дата", "source_system", "event_id", "Проверяемо",
    ))
    # требование архитектора #10 (лист кандидатов неполноценный): контакт/дети/сигнал/действие
    # видны прямо на строке, не только family_id и код причины -- см. OWNER50_CONTROL_COLUMNS.
    candidates_ws.append(OWNER50_CONTROL_COLUMNS)
    excluded_ws.append(OWNER50_CONTROL_COLUMNS)
    control_ws.append(("Показатель", "Значение"))

    for row in rows:
        ready.append(
            (
                row["rank"],
                row["family_id"],
                row["brand"],
                row["name"],
                row["contact_customer_id"],
                row["phone"],
                row["email"],
                row["channel"],
                row["evidence_at"],
                row["signal_type"],
                row["evidence_text"],
                row["next_action"],
                row["expires_at"],
                row["offer"],
                row["children"],
                row["payment"],
                row["rank_reason"],
                row["action_text"],
                row["target_child_key"],
                row["target_child_name"],
                row["target_child_grade"],
                row["family_members"],
            )
        )
        for item in row["evidence"]:
            evidence.append((
                row["family_id"], row["rank"], item["kind"], item["text"], item["source"],
                item["at"], item["source_system"], item["event_id"], item["resolvable"],
            ))
    # требование аудиторов BLOCKED #5: CANDIDATES (candidate+outside_limit) и EXCLUDED
    # (excluded) -- отдельные листы; "selected" уже полностью на READY_50 и не дублируется.
    for item in sorted(control):
        status = item[1]
        if status in ("candidate", "outside_limit"):
            candidates_ws.append(item)
        elif status == "excluded":
            excluded_ws.append(item)

    for key in (
        "tenant_id", "as_of", "generated_at",
        "ready_50", "ready_total", "ready_outside_limit", "ready_audit_population_complete",
        "candidates", "excluded", "families_classified_total",
    ):
        control_ws.append((key, control_meta.get(key)))
    control_ws.append(("freshness_gate_passed", control_meta.get("freshness_gate_passed")))
    control_ws.append(("freshness_gate_checked_at", control_meta.get("freshness_gate_checked_at")))
    control_ws.append(("price_axes_catalog_version", control_meta.get("price_axes_catalog_provenance")))
    control_ws.append(("", ""))
    control_ws.append(("Свежесть источника", "cursor_at / imported_at / max_event_at / events"))
    for freshness_row in control_meta.get("freshness_rows") or ():
        control_ws.append((
            f"источник: {freshness_row.get('source_system')}",
            f"cursor_at={freshness_row.get('cursor_at')}; imported_at={freshness_row.get('imported_at')}; "
            f"max_event_at={freshness_row.get('max_event_at')}; events={freshness_row.get('events')}",
        ))

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column in ws.columns:
            letter = column[0].column_letter
            ws.column_dimensions[letter].width = min(80, max(12, *(len(str(cell.value or "")) for cell in column)))
    wb.save(path)
    path.chmod(0o600)


def _write_workbook(path: Path, dossiers: Sequence[CustomerDossier]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    overview = wb.active
    overview.title = "Оглавление"
    overview.append(("customer_id", "Имя", "Бренд", "Семья", "Сигналы", "Следующий шаг", "Статус действия", "Код причины бездействия", "Интересов", "Болей", "Возражений", "Хронология"))
    overview.freeze_panes = "A2"
    for cell in overview[1]:
        cell.font = Font(bold=True)
    for index, dossier in enumerate(dossiers, start=1):
        sheet_name = f"Клиент {index}"
        overview.append(
            (
                dossier.customer_id,
                dossier.display_name,
                dossier.brand,
                len(dossier.family),
                len(dossier.signals),
                dossier.next_step,
                dossier.action_status,
                dossier.no_action_reason_code,
                len(dossier.interests),
                len(dossier.pains),
                len(dossier.objections),
                len(dossier.chronology),
            )
        )
        ws = wb.create_sheet(sheet_name)
        ws.append(("Раздел", "Значение", "Откуда"))
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        if dossier.actuality_header:
            ws.append(("Актуальность", dossier.actuality_header, _display_source("timeline_events/reconcile")))
        ws.append(("Кто", dossier.display_name, _display_source("customer_identities")))
        ws.append(("Бренд", dossier.brand or "Не определён однозначно", _display_source("customer_identities")))
        ws.append(("Контакт", f"{dossier.phone} {dossier.email}".strip(), _display_source("customer_identities")))
        for row in dossier.family:
            ws.append((row.section, row.text, _display_source(row.source)))
        for row in dossier.money:
            ws.append((row.section, row.text, _display_source(row.source)))
        for row in dossier.signals:
            ws.append((row.section, row.text, _display_source(row.source)))
        ws.append(
            (
                "Следующий шаг",
                dossier.next_step or "Не определён: менеджеру нужно выбрать действие после проверки истории.",
                _display_source(dossier.next_step_source) if dossier.next_step else "Требует решения менеджера",
            )
        )
        ws.append(("Статус действия", dossier.action_status, dossier.no_action_reason_code or "Активный шаг"))
        for row in dossier.objections:
            ws.append((row.section, row.text, _display_source(row.source)))
        for item in dossier.interests:
            ws.append(("Интересы", item.text, _display_source(item.source)))
        for item in dossier.pains:
            ws.append(("Боли", item.text, _display_source(item.source)))
        for row in dossier.chronology:
            ws.append((row.section, row.text, _display_source(row.source)))
        for column, width in {"A": 18, "B": 90, "C": 28}.items():
            ws.column_dimensions[column].width = width
    wb.save(path)
    path.chmod(0o600)


def _display_source(source: str) -> str:
    text = str(source or "")
    if text.startswith("family_links_v1"):
        return "Семейная карта"
    if text.startswith("customer_purchases_v1"):
        return "Деньги из staging"
    if text.startswith("derived_signals"):
        return "Сигнал Customer Timeline"
    if text.startswith("customer_objections_v1"):
        return "Клиентское возражение"
    if text.startswith("mango_call"):
        return "Клиентская реплика из звонка"
    if text == "products_of_interest":
        return "Данные клиента/сделки"
    mapping = {
        "timeline_events/reconcile": "Шапка актуальности",
        "customer_identities": "Карточка клиента",
        "mango_processed_summary": "Сводка звонка",
        "mail_archive_stage2": "Письмо",
        "mail_archive": "Письмо",
        "amocrm_snapshot": "AMO read-only",
        "amocrm_price_readonly": "AMO read-only",
        "amocrm_event": "AMO read-only",
        "master_contacts_snapshot": "Сводка контакта",
        "tallanto_snapshot": "Tallanto staging",
        "tallanto_crm_call": "Tallanto staging",
        "telegram_history": "Telegram история",
    }
    return mapping.get(text, text or "Источник не указан")
