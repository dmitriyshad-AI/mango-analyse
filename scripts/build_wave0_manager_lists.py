#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from mango_mvp.customer_timeline.freshness import (
    MANAGER_REQUIRED_SOURCE_SYSTEMS,
    manager_freshness_gate,
    source_freshness_rows,
)
from mango_mvp.customer_timeline.manager_dossier import manager_outreach_eligibility
from mango_mvp.customer_timeline.safety import guard_customer_timeline_output_path
from mango_mvp.customer_timeline.store import customer_timeline_readonly_uri


DEFAULT_STAGING_DB = Path(".codex_local/staging/customer_timeline_staging.sqlite")
DEFAULT_RECONCILE_JSON = Path(".codex_local/review/f9_amo_actuality/stalled_deals_amo_reconcile.json")
DEFAULT_OUT_XLSX = Path(".codex_local/review/f9_amo_actuality/2026-07-03_VOLNA0_dengi_na_polu_refresh.xlsx")
SCHEMA_VERSION = "marathon2_f9_wave0_manager_lists_v1"


def connect_ro(db: Path) -> sqlite3.Connection:
    con = sqlite3.connect(customer_timeline_readonly_uri(db), uri=True)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA query_only = ON")
    return con


def source_freshness(con: sqlite3.Connection) -> list[Mapping[str, Any]]:
    return source_freshness_rows(con, expected_sources=MANAGER_REQUIRED_SOURCE_SYSTEMS)


def build_rows(
    con: sqlite3.Connection,
    signal_types: Sequence[str],
    *,
    limit: int,
    tenant_id: str = "foton",
    exclusion_counts: Counter[str] | None = None,
) -> list[Mapping[str, Any]]:
    placeholders = ",".join("?" for _ in signal_types)
    params: list[Any] = [tenant_id, *signal_types]
    sql = f"""
        SELECT
          s.customer_id,
          s.tenant_id,
          s.signal_id,
          ci.display_name,
          ci.primary_phone,
          ci.primary_email,
          ci.identity_status,
          json_extract(ci.record_json, '$.metadata.brands') AS brands_json,
          s.event_id,
          s.signal_type,
          s.severity,
          s.expires_at,
          s.created_at AS signal_created_at,
          s.record_json AS signal_record_json,
          COALESCE(pf.total_in, 0) AS fact_total_in,
          COALESCE(pf.total_out, 0) AS fact_total_out,
          COALESCE(pp.total_in, 0) AS plan_total_in,
          COALESCE(pf.deals_cnt, 0) AS fact_deals_cnt,
          COALESCE(pp.deals_cnt, 0) AS plan_deals_cnt,
          MAX(e.event_at) AS latest_event_at,
          COALESCE(se.event_at, s.created_at) AS evidence_event_at,
          GROUP_CONCAT(DISTINCT o.source_id) AS open_amo_lead_ids,
          GROUP_CONCAT(DISTINCT o.status) AS open_amo_statuses
        FROM derived_signals s
        JOIN customer_identities ci
          ON ci.tenant_id = s.tenant_id
         AND ci.customer_id = s.customer_id
        LEFT JOIN customer_purchases_v1 pf
          ON pf.tenant_id = s.tenant_id
         AND pf.customer_id = s.customer_id
         AND pf.period = 'all_time'
         AND pf.money_kind = 'fact'
        LEFT JOIN customer_purchases_v1 pp
          ON pp.tenant_id = s.tenant_id
         AND pp.customer_id = s.customer_id
         AND pp.period = 'all_time'
         AND pp.money_kind = 'plan'
        LEFT JOIN timeline_events e
          ON e.tenant_id = s.tenant_id
         AND e.customer_id = s.customer_id
         AND (e.superseded_by IS NULL OR e.superseded_by = '')
        LEFT JOIN timeline_events se
          ON se.tenant_id = s.tenant_id
         AND se.event_id = s.event_id
        LEFT JOIN customer_opportunities o
          ON o.tenant_id = s.tenant_id
         AND o.customer_id = s.customer_id
         AND o.opportunity_type = 'amo_deal'
         AND o.source_id IS NOT NULL
         AND o.source_id != ''
         AND (o.closed_at IS NULL OR o.closed_at = '')
         AND COALESCE(o.status, '') NOT IN ('142', '143', 'Закрыто и не реализовано', 'Успешно')
        WHERE s.tenant_id = ?
          AND s.signal_type IN ({placeholders})
          AND s.status = 'active'
          AND ci.identity_status = 'strong'
          AND COALESCE(json_array_length(json_extract(ci.record_json, '$.metadata.brands')), 0) = 1
          AND LOWER(json_extract(ci.record_json, '$.metadata.brands[0]')) IN ('foton', 'unpk')
          AND (s.expires_at IS NULL OR s.expires_at = '' OR julianday(s.expires_at) >= julianday('now'))
        GROUP BY s.signal_id
        ORDER BY CASE s.severity
                   WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 WHEN 'low' THEN 3 ELSE 4
                 END,
                 fact_total_in DESC, plan_total_in DESC, latest_event_at DESC, s.customer_id
    """
    result: list[Mapping[str, Any]] = []
    for row in con.execute(sql, tuple(params)).fetchall():
        eligibility = manager_outreach_eligibility(
            con,
            tenant_id=tenant_id,
            customer_id=str(row["customer_id"]),
            signal_id=str(row["signal_id"]),
        )
        if not eligibility["eligible"]:
            if exclusion_counts is not None:
                exclusion_counts.update(eligibility["reasons"])
            continue
        signal_record = _loads(row["signal_record_json"])
        brands = _loads_list(row["brands_json"])
        recommended_action = str(signal_record.get("recommended_action") or "").strip()
        action_due_at = str(signal_record.get("action_due_at") or "").strip()
        evidence_event_at = row["evidence_event_at"]
        if row["signal_type"] == "season_return_candidate":
            evidence_event_at = (_loads(signal_record.get("metadata")).get("last_purchase_at") or evidence_event_at)
        verified_offer = str(signal_record.get("offer_product") or "").strip() if signal_record.get("offer_verified") is True else ""
        result.append(
            {
                "tenant_id": row["tenant_id"],
                "brand": brands[0] if len(brands) == 1 else "",
                "customer_id": row["customer_id"],
                "display_name": row["display_name"],
                "primary_phone": row["primary_phone"],
                "primary_email": row["primary_email"],
                "signal_type": row["signal_type"],
                "severity": row["severity"],
                "fact_total_in": row["fact_total_in"],
                "fact_total_out": row["fact_total_out"],
                "plan_total_in": row["plan_total_in"],
                "fact_deals_cnt": row["fact_deals_cnt"],
                "plan_deals_cnt": row["plan_deals_cnt"],
                "latest_event_at": row["latest_event_at"],
                "signal_created_at": row["signal_created_at"],
                "expires_at": row["expires_at"],
                "action_today": signal_record.get("action_today") is True,
                "action_due_at": action_due_at,
                "priority_rank": {"critical": 1, "high": 2, "medium": 3, "low": 4}.get(str(row["severity"]), 5),
                "priority_reason": signal_record.get("evidence_text") or row["signal_type"],
                "identity_status": row["identity_status"],
                "signal_is_current": True,
                "offer_product": verified_offer or "сначала уточнить",
                "offer_verified": bool(verified_offer),
                "recommended_action": recommended_action,
                "next_action": recommended_action,
                "evidence_text": signal_record.get("evidence_text") or "",
                "evidence_event_at": evidence_event_at,
                "evidence_ref_masked": _mask_id(row["event_id"] or row["signal_id"]),
                "open_amo_lead_ids": row["open_amo_lead_ids"] or "",
                "open_amo_statuses": row["open_amo_statuses"] or "",
            }
        )
        if limit > 0 and len(result) >= limit:
            break
    return result


def _loads(raw: object) -> Mapping[str, Any]:
    if isinstance(raw, Mapping):
        return raw
    try:
        value = json.loads(str(raw or "{}"))
    except json.JSONDecodeError:
        return {}
    return value if isinstance(value, Mapping) else {}


def _loads_list(raw: object) -> list[str]:
    try:
        value = json.loads(str(raw or "[]"))
    except json.JSONDecodeError:
        return []
    return [str(item) for item in value if str(item).strip()] if isinstance(value, list) else []


def _mask_id(value: object) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8")).hexdigest()[:16] if value else ""


def _reconcile_deal_rows(
    rows: Sequence[Mapping[str, Any]], reconcile: Mapping[str, Any]
) -> tuple[list[Mapping[str, Any]], Mapping[str, int]]:
    by_customer: dict[str, list[Mapping[str, Any]]] = {}
    for item in reconcile.get("rows") or ():
        if isinstance(item, Mapping) and item.get("customer_hash"):
            by_customer.setdefault(str(item["customer_hash"]), []).append(item)
    kept: list[Mapping[str, Any]] = []
    excluded = Counter()
    for row in rows:
        expected_lead_hashes = {
            _mask_id(item.strip())
            for item in str(row.get("open_amo_lead_ids") or "").split(",")
            if item.strip()
        }
        if not expected_lead_hashes:
            kept.append({**dict(row), "row_reconcile_status": "not_needed", "reconciled_at": None})
            continue
        live_rows = by_customer.get(_mask_id(row.get("customer_id")), []) if reconcile.get("status") == "checked" else []
        if not live_rows:
            excluded["not_reconciled"] += 1
            continue
        observed_lead_hashes = {str(item.get("lead_hash") or "") for item in live_rows}
        if not expected_lead_hashes.issubset(observed_lead_hashes):
            excluded["reconcile_rows_missing"] += 1
            continue
        failed_read = next(
            (
                item
                for item in live_rows
                if str(item.get("reason") or "").startswith("live_read_error:")
                or item.get("live_found") is not True
            ),
            None,
        )
        if failed_read is not None:
            excluded[str(failed_read.get("reason") or "live_read_incomplete")] += 1
            continue
        changed = next((item for item in live_rows if item.get("changed") is True), None)
        if changed is not None:
            excluded[str(changed.get("reason") or "live_changed")] += 1
            continue
        kept.append({**dict(row), "row_reconcile_status": "checked", "reconciled_at": reconcile.get("generated_at")})
    return kept, dict(excluded)


def reconcile_header(reconcile: Mapping[str, Any]) -> str:
    if not reconcile:
        return "сверка с живым AMO: не проводилась"
    status = str(reconcile.get("status") or "unknown")
    if status == "checked":
        return (
            "сверка с живым AMO: "
            f"{reconcile.get('generated_at')}; "
            f"{reconcile.get('customers_changed')} расхождений из {reconcile.get('customers_checked')}; "
            f"snapshot_stale={reconcile.get('snapshot_stale')}"
        )
    return f"сверка с живым AMO: не проводилась ({reconcile.get('reason') or status})"


def build_workbook(
    *,
    timeline_db: Path,
    out_xlsx: Path,
    allowed_root: Path,
    reconcile_json: Path | None,
    limit_per_sheet: int,
    tenant_id: str = "foton",
    enforce_freshness: bool = True,
) -> Mapping[str, Any]:
    db = timeline_db.expanduser().resolve(strict=False)
    out = _guard_local_wave0_output_path(out_xlsx, allowed_root)
    out.parent.mkdir(parents=True, exist_ok=True)
    reconcile = _read_json(reconcile_json) if reconcile_json and reconcile_json.exists() else {}
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds")
    with connect_ro(db) as con:
        eligibility_excluded: Counter[str] = Counter()
        freshness = source_freshness(con)
        freshness_gate = manager_freshness_gate(freshness)
        if enforce_freshness and not freshness_gate["passed"]:
            reasons = ", ".join(
                f"{item['source_system']}:{item['reason']}" for item in freshness_gate["blockers"]
            )
            raise RuntimeError(f"manager freshness gate failed: {reasons}")
        returned_rows, returned_excluded = _reconcile_deal_rows(
            build_rows(
                con,
                ("client_returned", "callback_due"),
                limit=limit_per_sheet,
                tenant_id=tenant_id,
                exclusion_counts=eligibility_excluded,
            ),
            reconcile,
        )
        deal_rows, deal_excluded = _reconcile_deal_rows(
            build_rows(
                con,
                ("deal_stalling",),
                limit=limit_per_sheet,
                tenant_id=tenant_id,
                exclusion_counts=eligibility_excluded,
            ),
            reconcile,
        )
        seasonal_rows, seasonal_excluded = _reconcile_deal_rows(
            build_rows(
                con,
                ("season_return_candidate",),
                limit=limit_per_sheet,
                tenant_id=tenant_id,
                exclusion_counts=eligibility_excluded,
            ),
            reconcile,
        )
        sheets = {
            "Вернулись и перезвон": returned_rows,
            "Зависшие факт LTV": deal_rows,
            "Сезонные": seasonal_rows,
        }
        duplicate_signal_suppressed: Counter[str] = Counter()
        seen_customers: set[str] = set()
        for sheet_name, rows in sheets.items():
            kept: list[Mapping[str, Any]] = []
            for row in rows:
                customer_id = str(row["customer_id"])
                if customer_id in seen_customers:
                    duplicate_signal_suppressed[str(row["signal_type"])] += 1
                    continue
                seen_customers.add(customer_id)
                kept.append(row)
            sheets[sheet_name] = kept
        quick_check = con.execute("PRAGMA quick_check").fetchone()[0]
    wb = Workbook()
    wb.remove(wb.active)
    cursor_text = "; ".join(
        f"{row['source_system']}={row.get('cursor_at') or 'нет курсора'}" for row in freshness[:8]
    )
    event_text = "; ".join(f"{row['source_system']}={row['max_event_at']}" for row in freshness[:8])
    import_text = "; ".join(
        f"{row['source_system']}={row.get('imported_at') or 'нет успешного импорта'}" for row in freshness[:8]
    )
    header_text = (
        f"Данные: cursor_at по источникам: {cursor_text}; "
        f"imported_at отдельно: {import_text}; max event_at отдельно: {event_text}; "
        f"собрано {generated_at}; {reconcile_header(reconcile)}"
    )
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = sheet_name
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = header_text
        ws["A2"].fill = PatternFill("solid", fgColor="FFF2CC")
        columns = [
            "tenant_id",
            "brand",
            "customer_id",
            "display_name",
            "primary_phone",
            "primary_email",
            "signal_type",
            "severity",
            "fact_total_in",
            "fact_total_out",
            "plan_total_in",
            "fact_deals_cnt",
            "plan_deals_cnt",
            "latest_event_at",
            "signal_created_at",
            "expires_at",
            "action_today",
            "action_due_at",
            "priority_rank",
            "priority_reason",
            "identity_status",
            "signal_is_current",
            "offer_product",
            "offer_verified",
            "recommended_action",
            "next_action",
            "evidence_text",
            "evidence_event_at",
            "evidence_ref_masked",
            "row_reconcile_status",
            "reconciled_at",
            "open_amo_lead_ids",
            "open_amo_statuses",
        ]
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=4, column=col_idx, value=column)
            cell.font = Font(bold=True)
        for row_idx, row in enumerate(rows, start=5):
            for col_idx, column in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(column))
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(columns))}{max(4, len(rows) + 4)}"
        for col_idx, column in enumerate(columns, start=1):
            width = 18
            if column in {"display_name", "primary_email", "recommended_action", "evidence_text"}:
                width = 34
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(out)
    summary = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": generated_at,
        "timeline_db": str(db),
        "source_open_mode": "sqlite_mode_ro",
        "quick_check": quick_check,
        "out_xlsx": str(out),
        "rows_by_sheet": {name: len(rows) for name, rows in sheets.items()},
        "selected_customer_ids": list(
            dict.fromkeys(str(row["customer_id"]) for rows in sheets.values() for row in rows)
        ),
        "signal_counts": dict(Counter(row["signal_type"] for rows in sheets.values() for row in rows)),
        "eligibility_exclusion_counts": dict(eligibility_excluded),
        "duplicate_signal_suppressed": dict(duplicate_signal_suppressed),
        "freshness_top": freshness[:12],
        "freshness_gate": freshness_gate,
        "reconcile_status": reconcile.get("status") if reconcile else "missing",
        "reconcile_summary": {
            "customers_checked": reconcile.get("customers_checked"),
            "customers_changed": reconcile.get("customers_changed"),
            "snapshot_stale": reconcile.get("snapshot_stale"),
            "reason": reconcile.get("reason"),
        },
        "excluded_deal_rows": deal_excluded,
        "excluded_returned_rows": returned_excluded,
        "excluded_seasonal_rows": seasonal_excluded,
        "safety": {
            "prod_db_write": False,
            "crm_write": False,
            "send_messages": False,
            "pii_scope": "local_codex_local_only",
        },
    }
    summary_path = out.with_suffix(".summary.json")
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return summary


def _guard_local_wave0_output_path(path: Path, allowed_root: Path) -> Path:
    resolved = guard_customer_timeline_output_path(path, allowed_root)
    root = Path(allowed_root).expanduser().resolve(strict=False)
    try:
        relative = resolved.relative_to(root)
    except ValueError as exc:  # pragma: no cover - base guard normally catches this.
        raise ValueError("wave0 manager list output must stay under allowed root") from exc
    if not relative.parts or relative.parts[0] != ".codex_local":
        raise ValueError("wave0 manager list contains PII and must stay under .codex_local")
    return resolved


def _read_json(path: Path | None) -> Mapping[str, Any]:
    if not path:
        return {}
    try:
        value = json.loads(path.expanduser().read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    return value if isinstance(value, Mapping) else {}


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build local Wave-0 manager Excel lists from staging timeline.")
    parser.add_argument("--timeline-db", type=Path, default=DEFAULT_STAGING_DB)
    parser.add_argument("--allowed-root", type=Path, default=Path.cwd())
    parser.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT_XLSX)
    parser.add_argument("--reconcile-json", type=Path, default=DEFAULT_RECONCILE_JSON)
    parser.add_argument("--limit-per-sheet", type=int, default=500)
    parser.add_argument("--tenant-id", default="foton")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    summary = build_workbook(
        timeline_db=args.timeline_db,
        out_xlsx=args.out_xlsx,
        allowed_root=args.allowed_root,
        reconcile_json=args.reconcile_json,
        limit_per_sheet=args.limit_per_sheet,
        tenant_id=args.tenant_id,
    )
    print(json.dumps({"out_xlsx": summary["out_xlsx"], "rows_by_sheet": summary["rows_by_sheet"]}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
