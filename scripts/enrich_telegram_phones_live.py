#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib import error as url_error
from urllib import parse as url_parse
from urllib import request as url_request

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key:
            continue
        if key not in os.environ:
            os.environ[key] = value.strip()


import os  # noqa: E402

load_env_file(PROJECT_ROOT / "stable_runtime" / "amocrm_runtime" / ".env.private")

from mango_mvp.amocrm_runtime.amo_integration import (  # noqa: E402
    fetch_lead,
    fetch_pipelines_with_statuses,
    fetch_users,
    search_contacts_by_phone,
)
from mango_mvp.amocrm_runtime.db import SessionLocal  # noqa: E402
from mango_mvp.amocrm_runtime.tallanto_api import (  # noqa: E402
    TallantoApiClient,
    TallantoApiError,
    _extract_record_list,
    build_tallanto_api_config,
)
from mango_mvp.utils.phone import normalize_phone  # noqa: E402
from sqlalchemy.exc import OperationalError  # noqa: E402


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def norm_phone(value: Any) -> str:
    normalized = normalize_phone("" if value is None else str(value))
    return normalized or ""


def field_values_map(custom_fields_values: Any) -> dict[str, list[str]]:
    result: dict[str, list[str]] = defaultdict(list)
    for field in custom_fields_values or []:
        if not isinstance(field, dict):
            continue
        field_name = str(field.get("field_name") or field.get("field_code") or "").strip()
        if not field_name:
            continue
        for item in field.get("values") or []:
            if not isinstance(item, dict):
                continue
            raw_value = item.get("value")
            if raw_value is None:
                continue
            if isinstance(raw_value, dict):
                value = ", ".join(str(v).strip() for v in raw_value.values() if str(v).strip())
            else:
                value = str(raw_value).strip()
            if value:
                result[field_name].append(value)
    return {key: values for key, values in result.items() if values}


def first_value(mapping: dict[str, list[str]], *names: str) -> str:
    lowered = {key.casefold(): values for key, values in mapping.items()}
    for name in names:
        values = lowered.get(name.casefold())
        if values:
            return values[0]
    return ""


def join_unique(values: Iterable[Any], sep: str = " | ") -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = str(value or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return sep.join(result)


def parse_ts(value: Any) -> pd.Timestamp:
    if value in (None, "", 0):
        return pd.NaT
    try:
        return pd.to_datetime(int(value), unit="s", utc=True)
    except Exception:
        return pd.to_datetime(value, utc=True, errors="coerce")


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def truncate(text: str, limit: int = 500) -> str:
    value = as_text(text)
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def canonical_export_dir() -> Path:
    pointer = PROJECT_ROOT / "stable_runtime" / "CANONICAL_EXPORT.txt"
    if pointer.exists():
        path = Path(pointer.read_text(encoding="utf-8").strip())
        if path.exists():
            return path
    candidates = sorted(
        (PROJECT_ROOT / "stable_runtime").glob("sales_master_export_*_review_accepted"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError("No canonical sales_master_export_*_review_accepted directory found.")
    return candidates[0]


def telegram_phone_variants(phone: str) -> list[str]:
    normalized = norm_phone(phone)
    digits = "".join(ch for ch in normalized if ch.isdigit())
    if len(digits) != 11 or not digits.startswith("7"):
        return [phone]
    local = digits[1:]
    pretty7 = f"7 {local[0:3]} {local[3:6]}-{local[6:8]}-{local[8:10]}"
    pretty8 = f"8 {local[0:3]} {local[3:6]}-{local[6:8]}-{local[8:10]}"
    candidates = [
        normalized,
        digits,
        local,
        pretty7,
        pretty8,
        f"+7 {local[0:3]} {local[3:6]}-{local[6:8]}-{local[8:10]}",
        f"8{local}",
        f"7{local}",
    ]
    return [item for item in dict.fromkeys(candidate.strip() for candidate in candidates if candidate and candidate.strip())]


def build_status_maps(pipelines: list[dict[str, Any]]) -> tuple[dict[int, str], dict[int, dict[str, Any]]]:
    pipeline_name_by_id: dict[int, str] = {}
    status_meta_by_id: dict[int, dict[str, Any]] = {}
    for pipeline in pipelines:
        pipeline_id = int(pipeline.get("id") or 0)
        if not pipeline_id:
            continue
        pipeline_name = str(pipeline.get("name") or "").strip()
        pipeline_name_by_id[pipeline_id] = pipeline_name
        for status in (pipeline.get("_embedded") or {}).get("statuses") or []:
            if not isinstance(status, dict):
                continue
            status_id = int(status.get("id") or 0)
            if not status_id:
                continue
            status_meta_by_id[status_id] = {
                "pipeline_id": pipeline_id,
                "pipeline_name": pipeline_name,
                "status_name": str(status.get("name") or "").strip(),
                "status_type": status.get("type"),
            }
    return pipeline_name_by_id, status_meta_by_id


def lead_sort_key(lead: dict[str, Any]) -> tuple[int, int]:
    for key in ("updated_at", "closed_at", "created_at"):
        try:
            value = int(lead.get(key) or 0)
        except Exception:
            value = 0
        if value:
            return value, int(lead.get("id") or 0)
    return 0, int(lead.get("id") or 0)


def amo_contact_phones(contact: dict[str, Any]) -> str:
    values = field_values_map(contact.get("custom_fields_values"))
    phones = values.get("Телефон") or values.get("PHONE") or []
    return join_unique(phones)


@dataclass
class AmoContext:
    pipeline_name_by_id: dict[int, str]
    status_meta_by_id: dict[int, dict[str, Any]]
    user_name_by_id: dict[int, str]
    lead_cache: dict[int, dict[str, Any]]


def fetch_lead_cached(session: Any, amo_ctx: AmoContext, lead_id: int) -> dict[str, Any]:
    if lead_id not in amo_ctx.lead_cache:
        amo_ctx.lead_cache[lead_id] = fetch_lead(session, lead_id=lead_id, with_fields="contacts")
    return amo_ctx.lead_cache[lead_id]


def summarize_amo_leads(session: Any, amo_ctx: AmoContext, leads: list[dict[str, Any]], max_full_fetch: int = 5) -> dict[str, Any]:
    sorted_leads = sorted(leads, key=lead_sort_key, reverse=True)
    open_count = 0
    won_count = 0
    lost_count = 0
    lead_lines: list[str] = []
    utm_lines: list[str] = []
    latest_reason = ""
    for index, lead in enumerate(sorted_leads):
        status_id = int(lead.get("status_id") or 0)
        meta = amo_ctx.status_meta_by_id.get(status_id, {})
        pipeline_name = meta.get("pipeline_name") or amo_ctx.pipeline_name_by_id.get(int(lead.get("pipeline_id") or 0), "")
        status_name = meta.get("status_name") or str(status_id)
        if status_id == 142:
            won_count += 1
        elif status_id == 143:
            lost_count += 1
        else:
            open_count += 1
        closed_at = parse_ts(lead.get("closed_at"))
        updated_at = parse_ts(lead.get("updated_at"))
        when = closed_at if pd.notna(closed_at) else updated_at
        when_text = when.strftime("%Y-%m-%d") if pd.notna(when) else ""
        responsible_name = amo_ctx.user_name_by_id.get(int(lead.get("responsible_user_id") or 0), "")
        lead_lines.append(
            join_unique(
                [
                    when_text,
                    as_text(lead.get("name")),
                    pipeline_name,
                    status_name,
                    f"Цена: {lead.get('price')}" if lead.get("price") not in (None, "", 0) else "",
                    responsible_name,
                ],
                sep=" | ",
            )
        )
        if index < max_full_fetch:
            full = fetch_lead_cached(session, amo_ctx, int(lead.get("id")))
            field_map = field_values_map(full.get("custom_fields_values"))
            utm_parts = []
            for utm_name in ("utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term", "utm_referrer"):
                utm_value = first_value(field_map, utm_name)
                if utm_value:
                    utm_parts.append(f"{utm_name}={utm_value}")
            if utm_parts:
                utm_lines.append(f"{lead.get('id')}: " + ", ".join(utm_parts))
            if not latest_reason:
                latest_reason = first_value(field_map, "Причина отказа (лид)", "Причина отказа", "Причина")
    return {
        "open_count": open_count,
        "won_count": won_count,
        "lost_count": lost_count,
        "recent_leads_text": "\n".join(lead_lines[:8]),
        "utm_text": "\n".join(utm_lines[:5]),
        "latest_reason": latest_reason,
    }


def amo_lookup_phone(session: Any, amo_ctx: AmoContext, phone: str) -> dict[str, Any]:
    contacts = search_contacts_by_phone(session, phone=phone, limit=20)
    tallanto_ids: list[str] = []
    contact_summaries: list[dict[str, Any]] = []
    aggregate_open = 0
    aggregate_won = 0
    aggregate_lost = 0
    recent_leads_chunks: list[str] = []
    utm_chunks: list[str] = []
    latest_reason = ""
    for contact in contacts:
        fields = field_values_map(contact.get("custom_fields_values"))
        tallanto_id = first_value(fields, "Id Tallanto")
        if tallanto_id:
            tallanto_ids.append(tallanto_id)
        lead_ids: list[int] = []
        for item in (contact.get("_embedded") or {}).get("leads") or []:
            if not isinstance(item, dict):
                continue
            try:
                lead_id = int(item.get("id") or 0)
            except Exception:
                lead_id = 0
            if lead_id:
                lead_ids.append(lead_id)
        leads = [fetch_lead_cached(session, amo_ctx, lead_id) for lead_id in lead_ids[:3]]
        lead_summary = summarize_amo_leads(session, amo_ctx, leads)
        aggregate_open += lead_summary["open_count"]
        aggregate_won += lead_summary["won_count"]
        aggregate_lost += lead_summary["lost_count"]
        if lead_summary["recent_leads_text"]:
            recent_leads_chunks.append(f"Контакт {contact.get('id')}:\n{lead_summary['recent_leads_text']}")
        if lead_summary["utm_text"]:
            utm_chunks.append(lead_summary["utm_text"])
        if not latest_reason and lead_summary["latest_reason"]:
            latest_reason = lead_summary["latest_reason"]
        contact_summaries.append(
            {
                "id": int(contact.get("id") or 0),
                "name": as_text(contact.get("name")),
                "responsible": amo_ctx.user_name_by_id.get(int(contact.get("responsible_user_id") or 0), ""),
                "phones": amo_contact_phones(contact),
                "emails": join_unique(fields.get("Email") or fields.get("EMAIL") or []),
                "tallanto_id": tallanto_id,
                "tallanto_branch": first_value(fields, "Филиал Tallanto"),
                "parent_name": first_value(fields, "ФИО Родителя"),
                "telegram_id": first_value(fields, "Telegram ID"),
                "telegram_username": first_value(fields, "Telegram username"),
                "ai_priority": first_value(fields, "AI-приоритет"),
                "ai_next_step": first_value(fields, "AI-рекомендованный следующий шаг"),
                "ai_summary": first_value(fields, "Последняя AI-сводка"),
                "auto_history": first_value(fields, "Авто история общения", "История общения"),
                "lead_count": len(lead_ids),
                "open_lead_count_recent": lead_summary["open_count"],
                "won_lead_count_recent": lead_summary["won_count"],
                "lost_lead_count_recent": lead_summary["lost_count"],
                "recent_leads_text": lead_summary["recent_leads_text"],
                "utm_text": lead_summary["utm_text"],
                "latest_reason": lead_summary["latest_reason"],
            }
        )
    return {
        "contact_count": len(contact_summaries),
        "contacts": contact_summaries,
        "tallanto_ids": list(dict.fromkeys(item for item in tallanto_ids if item)),
        "open_lead_count_recent": aggregate_open,
        "won_lead_count_recent": aggregate_won,
        "lost_lead_count_recent": aggregate_lost,
        "recent_leads_text": "\n\n".join(recent_leads_chunks[:6]),
        "utm_text": "\n".join(dict.fromkeys(chunk for chunk in utm_chunks if chunk)),
        "latest_reason": latest_reason,
    }


def fast_tallanto_search(client: TallantoApiClient, phone: str) -> tuple[list[dict[str, Any]], str, str]:
    fields = ("phone_mobile", "phone_work", "phone_home", "phone_other")
    errors: list[str] = []
    records: list[dict[str, Any]] = []
    matched_via = ""
    for candidate in telegram_phone_variants(phone):
        for field_name in fields:
            try:
                payload = tallanto_get_entry_by_field_fast(client, field_name=field_name, field_value=candidate)
            except TallantoApiError as exc:
                message = str(exc)
                lowered = message.casefold()
                if "entry does not exist" in lowered or "not find by id" in lowered:
                    continue
                if "ошибка базы данных" in lowered:
                    errors.append(f"{field_name}:{candidate}:db_error")
                    continue
                errors.append(f"{field_name}:{candidate}:{message}")
                continue
            extracted = _extract_record_list(payload)
            if extracted:
                records.extend(extracted)
                matched_via = f"{field_name}:{candidate}"
        if records:
            break
    deduped: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for record in records:
        record_id = as_text(record.get("id"))
        if record_id and record_id in seen_ids:
            continue
        if record_id:
            seen_ids.add(record_id)
        deduped.append(record)
    return deduped, matched_via, "; ".join(errors[:10])


def tallanto_get_entry_by_field_fast(
    client: TallantoApiClient,
    *,
    field_name: str,
    field_value: str,
    timeout_seconds: int = 6,
) -> dict[str, Any]:
    query_items = [
        ("method", "get_entry_by_fields"),
        ("module", "Contact"),
        (f"fields_values[{field_name}]", field_value),
    ]
    url = f"{client.endpoint_url}?{url_parse.urlencode(query_items, doseq=True)}"
    request = url_request.Request(
        url,
        headers={
            "Accept": "application/json",
            "Content-Type": "application/x-www-form-urlencoded",
            "X-Auth-Token": client.config.api_token,
        },
        method="GET",
    )
    try:
        with url_request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw.strip() else {}
    except url_error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise TallantoApiError(
            f"HTTP {exc.code} from Tallanto: {details or exc.reason}",
            status_code=502,
        ) from exc
    except Exception as exc:
        raise TallantoApiError(
            f"Failed to reach Tallanto: {exc}",
            status_code=502,
        ) from exc


def compact_tallanto_contact(contact: dict[str, Any]) -> dict[str, Any]:
    filial = contact.get("filial")
    if isinstance(filial, dict):
        filial_text = join_unique(filial.values())
    else:
        filial_text = as_text(filial)
    return {
        "id": as_text(contact.get("id")),
        "name": join_unique([contact.get("last_name"), contact.get("first_name"), contact.get("middle_name")], sep=" "),
        "first_name": as_text(contact.get("first_name")),
        "last_name": as_text(contact.get("last_name")),
        "phone_mobile": as_text(contact.get("phone_mobile")),
        "email": as_text(contact.get("email1")),
        "branch": filial_text,
        "assigned_user_name": as_text(contact.get("assigned_user_name")),
        "amo_id": as_text(contact.get("amo_id")),
        "type_client": first_value({"type_client": [as_text(contact.get("type_client_c_translated")) or as_text(contact.get("type_client"))]}, "type_client"),
        "description": as_text(contact.get("description")),
        "contact_card": as_text(contact.get("contact_card")),
        "source": as_text(contact.get("source_translated") or contact.get("source")),
        "spend_money": as_text(contact.get("spend_money")),
        "recharge_money": as_text(contact.get("recharge_money")),
        "balance_money": as_text(contact.get("balance_money")),
        "parent_name": as_text(contact.get("marital_status_c")),
    }


def empty_tallanto_context(contact: dict[str, Any]) -> dict[str, Any]:
    return {
        "contact": contact,
        "opportunities": [],
        "requests": [],
        "finances": [],
        "course_relations": [],
        "class_relations": [],
    }


def summarize_tallanto_context(context: dict[str, Any]) -> dict[str, Any]:
    contact = compact_tallanto_contact(context.get("contact") or {})
    opportunities = [item for item in (context.get("opportunities") or []) if isinstance(item, dict)]
    requests = [item for item in (context.get("requests") or []) if isinstance(item, dict)]
    finances = [item for item in (context.get("finances") or []) if isinstance(item, dict)]
    course_relations = [item for item in (context.get("course_relations") or []) if isinstance(item, dict)]
    class_relations = [item for item in (context.get("class_relations") or []) if isinstance(item, dict)]
    opportunity_text = "\n".join(
        join_unique(
            [
                as_text(item.get("id")),
                as_text(item.get("name")),
                as_text(item.get("sales_stage")),
                as_text(item.get("next_step")),
                as_text(item.get("assigned_user_name")),
                as_text(item.get("date_closed") or item.get("system_date_closed")),
            ]
        )
        for item in opportunities[:8]
    )
    request_text = "\n".join(
        join_unique(
            [
                as_text(item.get("id")),
                as_text(item.get("name")),
                as_text(item.get("status")),
                as_text(item.get("date_next_contact")),
                as_text(item.get("assigned_user_name")),
            ]
        )
        for item in requests[:8]
    )
    finance_text = "\n".join(
        join_unique(
            [
                as_text(item.get("id")),
                as_text(item.get("name")),
                as_text(item.get("payment_summa")),
                as_text(item.get("payment_status")),
                as_text(item.get("date_entered")),
            ]
        )
        for item in finances[:8]
    )
    compact_context = join_unique(
        [
            truncate(contact.get("description", ""), 400),
            truncate(contact.get("contact_card", ""), 400),
            f"Сделок Tallanto: {len(opportunities)}" if opportunities else "",
            f"Заявок Tallanto: {len(requests)}" if requests else "",
            f"Оплат Tallanto: {len(finances)}" if finances else "",
            f"Групп: {len(course_relations)}" if course_relations else "",
            f"Занятий: {len(class_relations)}" if class_relations else "",
        ],
        sep="\n",
    )
    return {
        "contact": contact,
        "opportunity_count": len(opportunities),
        "request_count": len(requests),
        "finance_count": len(finances),
        "course_relation_count": len(course_relations),
        "class_relation_count": len(class_relations),
        "opportunity_text": opportunity_text,
        "request_text": request_text,
        "finance_text": finance_text,
        "compact_context": compact_context,
    }


def tallanto_lookup_phone(
    client: TallantoApiClient,
    *,
    phone: str,
    known_tallanto_ids: list[str],
    max_related_records: int,
    tallanto_context_cache: dict[str, dict[str, Any]],
    phone_search_mode: str,
) -> dict[str, Any]:
    contexts: list[dict[str, Any]] = []
    matched_via = ""
    search_error = ""

    unique_ids = [item for item in dict.fromkeys(known_tallanto_ids) if item]
    if unique_ids:
        matched_via = "known_tallanto_id"
        for tallanto_id in unique_ids[:5]:
            if tallanto_id not in tallanto_context_cache:
                try:
                    tallanto_context_cache[tallanto_id] = client.build_contact_context_by_contact_id(
                        tallanto_id,
                        max_related_records=max_related_records,
                    )
                except TallantoApiError as exc:
                    message = str(exc)
                    if "429" in message or "too many requests" in message.casefold():
                        try:
                            contact = client.contact_by_id(tallanto_id)
                            tallanto_context_cache[tallanto_id] = {
                                "contacts_found": 1 if contact else 0,
                                "contexts": [empty_tallanto_context(contact)] if contact else [],
                            }
                            search_error = join_unique([search_error, f"429_partial:{tallanto_id}"], sep="; ")
                            time.sleep(1.0)
                        except TallantoApiError as nested_exc:
                            search_error = join_unique([search_error, f"{tallanto_id}:{nested_exc}"], sep="; ")
                            continue
                    else:
                        search_error = join_unique([search_error, f"{tallanto_id}:{message}"], sep="; ")
                        continue
                time.sleep(0.1)
            payload = tallanto_context_cache[tallanto_id]
            contexts.extend(payload.get("contexts") or [])
    elif phone_search_mode == "fast":
        found_contacts, matched_via, search_error = fast_tallanto_search(client, phone)
        for contact in found_contacts[:5]:
            tallanto_id = as_text(contact.get("id"))
            if not tallanto_id:
                continue
            if tallanto_id not in tallanto_context_cache:
                try:
                    tallanto_context_cache[tallanto_id] = client.build_contact_context_by_contact_id(
                        tallanto_id,
                        max_related_records=max_related_records,
                    )
                except TallantoApiError as exc:
                    message = str(exc)
                    if "429" in message or "too many requests" in message.casefold():
                        tallanto_context_cache[tallanto_id] = {
                            "contacts_found": 1,
                            "contexts": [empty_tallanto_context(contact)],
                        }
                        search_error = join_unique([search_error, f"429_partial:{tallanto_id}"], sep="; ")
                        time.sleep(1.0)
                    else:
                        search_error = join_unique([search_error, f"{tallanto_id}:{message}"], sep="; ")
                        continue
                time.sleep(0.1)
            payload = tallanto_context_cache[tallanto_id]
            contexts.extend(payload.get("contexts") or [])
    else:
        matched_via = "skipped_no_known_id"

    deduped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for context in contexts:
        contact_id = as_text((context.get("contact") or {}).get("id"))
        if contact_id and contact_id in seen:
            continue
        if contact_id:
            seen.add(contact_id)
        deduped.append(context)

    summarized = [summarize_tallanto_context(context) for context in deduped]
    return {
        "count": len(summarized),
        "matched_via": matched_via,
        "search_error": search_error,
        "contexts": summarized,
        "contact_ids": [item["contact"]["id"] for item in summarized if item["contact"]["id"]],
        "contacts_text": join_unique(item["contact"]["name"] for item in summarized),
        "responsibles_text": join_unique(item["contact"]["assigned_user_name"] for item in summarized),
        "branches_text": join_unique(item["contact"]["branch"] for item in summarized),
        "type_clients_text": join_unique(item["contact"]["type_client"] for item in summarized),
        "opportunity_count": sum(item["opportunity_count"] for item in summarized),
        "request_count": sum(item["request_count"] for item in summarized),
        "finance_count": sum(item["finance_count"] for item in summarized),
        "course_relation_count": sum(item["course_relation_count"] for item in summarized),
        "class_relation_count": sum(item["class_relation_count"] for item in summarized),
        "opportunity_text": "\n\n".join(item["opportunity_text"] for item in summarized if item["opportunity_text"]),
        "request_text": "\n\n".join(item["request_text"] for item in summarized if item["request_text"]),
        "finance_text": "\n\n".join(item["finance_text"] for item in summarized if item["finance_text"]),
        "compact_context": "\n\n".join(item["compact_context"] for item in summarized if item["compact_context"]),
    }


def load_outreach_signals(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        df = pd.read_excel(path, sheet_name="Все кандидаты")
    except Exception:
        return pd.DataFrame()
    if "Телефон" not in df.columns:
        return pd.DataFrame()
    df["phone_norm"] = df["Телефон"].map(norm_phone)
    return df


def phone_offer_hypothesis(row: dict[str, Any]) -> tuple[str, str, str]:
    basis_parts: list[str] = []
    offer = ""
    possible = False

    if as_text(row.get("Рекомендуемый продукт")):
        offer = as_text(row["Рекомендуемый продукт"])
        basis_parts.append("рабочий contact-layer")
        possible = True
    elif as_text(row.get("Что предложить")):
        offer = as_text(row["Что предложить"])
        basis_parts.append("telegram outreach signals")
        possible = True
    elif row.get("Tallanto оплат", 0) or row.get("Tallanto групп", 0) or row.get("Tallanto занятий", 0):
        offer = "Продление, второй предмет или ближайшая программа"
        basis_parts.append("Tallanto: текущие оплаты/группы/занятия")
        possible = True
    elif row.get("AMO открытых сделок", 0):
        offer = "Продолжение текущего диалога по активной сделке"
        basis_parts.append("есть открытая сделка в AMO")
        possible = True
    elif as_text(row.get("Продукты интереса")):
        offer = as_text(row["Продукты интереса"])
        basis_parts.append("продукты интереса из рабочего слоя")
        possible = True
    elif as_text(row.get("Последний входящий Telegram")) or as_text(row.get("Последнее сообщение Telegram")):
        offer = "Мягкий follow-up и уточнение актуальной потребности"
        basis_parts.append("есть живой Telegram-контекст")
        possible = True

    return (
        "Да" if possible else "Нет",
        join_unique(basis_parts, sep=", "),
        offer,
    )


def load_telegram_phone_rows(export_root: Path) -> pd.DataFrame:
    max_dir = export_root / "local_vm_2024-04-01_max"
    with_dir = export_root / "local_vm_2024-04-01_with_contacts"
    dialogs = pd.DataFrame(load_jsonl(max_dir / "dialogs.jsonl"))
    dialogs = dialogs.rename(
        columns={
            "name": "dialog_name",
            "phone": "dialog_phone",
            "username": "dialog_username",
            "about": "tg_about",
            "first_name": "tg_first_name",
            "last_name": "tg_last_name",
        }
    )
    dialogs["phone_norm"] = dialogs["dialog_phone"].map(norm_phone)
    dialogs = dialogs[dialogs["phone_norm"].ne("")].copy()

    messages = pd.DataFrame(load_jsonl(with_dir / "messages.jsonl"))
    messages["date"] = pd.to_datetime(messages["date"], utc=True, errors="coerce")
    messages["has_media"] = messages.get("has_media", False)
    messages["has_media"] = messages["has_media"].astype("boolean").fillna(False).astype(bool)
    messages["media_path"] = messages.get("media_path", "")
    messages["media_path"] = messages["media_path"].fillna("").astype(str)
    msg_rows: list[dict[str, Any]] = []
    for dialog_id, group in messages.groupby("dialog_id"):
        group = group.sort_values("date")
        inbound = group[~group["out"].astype(bool)]
        outbound = group[group["out"].astype(bool)]
        last = group.iloc[-1]
        msg_rows.append(
            {
                "dialog_id": dialog_id,
                "tg_message_count": len(group),
                "tg_inbound_count": len(inbound),
                "tg_outbound_count": len(outbound),
                "tg_last_message_date": last["date"],
                "tg_last_message_text": as_text(last.get("text")),
                "tg_last_message_out": bool(last["out"]),
                "tg_last_inbound_text": as_text(inbound.iloc[-1].get("text")) if len(inbound) else "",
                "tg_last_outbound_text": as_text(outbound.iloc[-1].get("text")) if len(outbound) else "",
                "tg_last_inbound_date": inbound["date"].max() if len(inbound) else pd.NaT,
                "tg_last_outbound_date": outbound["date"].max() if len(outbound) else pd.NaT,
                "tg_unanswered_inbound": not bool(last["out"]),
                "tg_media_message_count": int(group["has_media"].sum()),
                "tg_media_with_path_count": int(group["media_path"].ne("").sum()),
            }
        )
    msg_df = pd.DataFrame(msg_rows)
    merged = dialogs.merge(msg_df, on="dialog_id", how="left")
    if "tg_unanswered_inbound" in merged.columns:
        merged["tg_unanswered_inbound"] = merged["tg_unanswered_inbound"].astype("boolean").fillna(False).astype(bool)

    aggregated_rows: list[dict[str, Any]] = []
    for phone_norm, group in merged.groupby("phone_norm"):
        group = group.sort_values("tg_last_message_date", ascending=False, na_position="last")
        latest = group.iloc[0]
        aggregated_rows.append(
            {
                "Телефон": phone_norm,
                "Диалогов Telegram": int(group["dialog_id"].nunique()),
                "ID диалогов Telegram": join_unique(group["dialog_id"].tolist()),
                "Диалоги Telegram": join_unique(group["dialog_name"].tolist()),
                "Username Telegram": join_unique(group["dialog_username"].tolist()),
                "Описание профиля Telegram": join_unique(group["tg_about"].tolist()),
                "Сообщений Telegram": int(group["tg_message_count"].fillna(0).sum()),
                "Входящих Telegram": int(group["tg_inbound_count"].fillna(0).sum()),
                "Исходящих Telegram": int(group["tg_outbound_count"].fillna(0).sum()),
                "Медиа Telegram": int(group["tg_media_message_count"].fillna(0).sum()),
                "Медиа с путем Telegram": int(group["tg_media_with_path_count"].fillna(0).sum()),
                "Есть неотвеченный входящий": "Да" if bool(group["tg_unanswered_inbound"].fillna(False).any()) else "Нет",
                "Дата последнего сообщения Telegram": latest.get("tg_last_message_date"),
                "Последнее сообщение Telegram": as_text(latest.get("tg_last_message_text")),
                "Последний входящий Telegram": join_unique(group.sort_values("tg_last_inbound_date", ascending=False)["tg_last_inbound_text"].tolist(), sep="\n"),
                "Последний исходящий Telegram": join_unique(group.sort_values("tg_last_outbound_date", ascending=False)["tg_last_outbound_text"].tolist(), sep="\n"),
            }
        )
    return pd.DataFrame(aggregated_rows).sort_values("Телефон").reset_index(drop=True)


def autosize_sheet(ws: Any) -> None:
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def add_dataframe_sheet(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title=title)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, pd.Timestamp):
                if pd.isna(value):
                    value = ""
                else:
                    if value.tzinfo is not None:
                        value = value.tz_convert("UTC").tz_localize(None)
                    value = value.to_pydatetime()
            elif isinstance(value, datetime) and value.tzinfo is not None:
                value = value.astimezone().replace(tzinfo=None)
            ws.cell(row=row_index, column=column_index, value=value)
    autosize_sheet(ws)


def json_default(value: Any) -> Any:
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    return str(value)


def build_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    ws = workbook.active
    ws.title = "Сводка"
    ws["A1"] = "Показатель"
    ws["B1"] = "Значение"
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    ws["B1"].font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    row = 2
    for key, value in summary.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    autosize_sheet(ws)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--export-root", type=Path, default=PROJECT_ROOT / "telegram_exports (2)")
    parser.add_argument("--canonical-dir", type=Path, default=None)
    parser.add_argument("--out-dir", type=Path, default=PROJECT_ROOT / "stable_runtime" / "audits" / "telegram_live_enrichment_20260416")
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--max-related-records", type=int, default=10)
    parser.add_argument("--tallanto-phone-search-mode", choices=["disabled", "fast"], default="disabled")
    args = parser.parse_args()

    canonical_dir = args.canonical_dir or canonical_export_dir()
    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    telegram_df = load_telegram_phone_rows(args.export_root)
    if args.offset:
        telegram_df = telegram_df.iloc[max(0, args.offset) :].copy()
    if args.limit is not None:
        telegram_df = telegram_df.head(max(0, args.limit)).copy()

    contacts_df = pd.read_csv(canonical_dir / "master_contacts_ru.csv")
    contacts_df.columns = [str(column).lstrip("\ufeff") for column in contacts_df.columns]
    contacts_df["phone_norm"] = contacts_df["Телефон клиента"].map(norm_phone)
    contacts_df = contacts_df[contacts_df["phone_norm"].ne("")].drop_duplicates("phone_norm")

    outreach_df = load_outreach_signals(PROJECT_ROOT / "stable_runtime" / "audits" / "telegram_outreach_20260416" / "telegram_openclaw_outreach_pack.xlsx")
    if not outreach_df.empty:
        outreach_df = outreach_df.drop_duplicates("phone_norm")

    client = TallantoApiClient(build_tallanto_api_config())
    tallanto_context_cache: dict[str, dict[str, Any]] = {}

    rows_for_jsonl: list[dict[str, Any]] = []
    flat_rows: list[dict[str, Any]] = []

    with SessionLocal() as bootstrap_session:
        pipelines = fetch_pipelines_with_statuses(bootstrap_session)
        users = fetch_users(bootstrap_session)
    pipeline_name_by_id, status_meta_by_id = build_status_maps(pipelines)
    user_name_by_id = {int(item.get("id") or 0): as_text(item.get("name")) for item in users if int(item.get("id") or 0)}

    total_rows = len(telegram_df)
    for index, phone_row in enumerate(telegram_df.to_dict(orient="records"), start=1):
        phone = as_text(phone_row["Телефон"])
        print(f"[{index}/{total_rows}] {phone}", flush=True)
        canonical_match_row = contacts_df[contacts_df["phone_norm"] == phone]
        canonical = canonical_match_row.iloc[0].to_dict() if not canonical_match_row.empty else {}
        amo_result = {
            "contact_count": 0,
            "contacts": [],
            "tallanto_ids": [],
            "open_lead_count_recent": 0,
            "won_lead_count_recent": 0,
            "lost_lead_count_recent": 0,
            "recent_leads_text": "",
            "utm_text": "",
            "latest_reason": "",
            "error": "",
        }
        for attempt in range(2):
            try:
                with SessionLocal() as session:
                    amo_ctx = AmoContext(
                        pipeline_name_by_id=pipeline_name_by_id,
                        status_meta_by_id=status_meta_by_id,
                        user_name_by_id=user_name_by_id,
                        lead_cache={},
                    )
                    amo_result = amo_lookup_phone(session, amo_ctx, phone)
                break
            except OperationalError as exc:
                amo_result["error"] = str(exc)
                if attempt == 0:
                    time.sleep(2.0)
                    continue
                break

        known_tallanto_ids = []
        if canonical:
            known_tallanto_ids.extend(
                item.strip()
                for item in as_text(canonical.get("ID Tallanto")).split("|")
                if item.strip()
            )
        known_tallanto_ids.extend(amo_result["tallanto_ids"])

        tallanto_result = tallanto_lookup_phone(
            client,
            phone=phone,
            known_tallanto_ids=[item for item in known_tallanto_ids if item],
            max_related_records=args.max_related_records,
            tallanto_context_cache=tallanto_context_cache,
            phone_search_mode=args.tallanto_phone_search_mode,
        )

        outreach_row = {}
        if not outreach_df.empty:
            matched = outreach_df[outreach_df["phone_norm"] == phone]
            if not matched.empty:
                outreach_row = matched.iloc[0].to_dict()

        flat_row = {
            **phone_row,
            "Есть exact match в рабочем слое": "Да" if canonical else "Нет",
            "ФИО родителя": as_text(canonical.get("ФИО родителя")),
            "ФИО ребенка": as_text(canonical.get("ФИО ребенка")),
            "Email": as_text(canonical.get("Email")),
            "Краткая история общения": as_text(canonical.get("Краткая история общения")),
            "Хронология общения": as_text(canonical.get("Хронология общения (последние 5 касаний)")),
            "Продукты интереса": as_text(canonical.get("Продукты интереса")),
            "Рекомендуемый продукт": as_text(canonical.get("Рекомендуемый продукт")),
            "Возражения": as_text(canonical.get("Возражения")),
            "Следующий шаг": as_text(canonical.get("Следующий шаг")),
            "Рекомендуемая дата следующего контакта": as_text(canonical.get("Рекомендуемая дата следующего контакта")),
            "Приоритет лида": as_text(canonical.get("Приоритет лида")),
            "Вероятность продажи, %": as_text(canonical.get("Вероятность продажи, %")),
            "Статус матчинга Tallanto": as_text(canonical.get("Статус матчинга Tallanto")),
            "ID Tallanto (рабочий слой)": as_text(canonical.get("ID Tallanto")),
            "Филиал Tallanto (рабочий слой)": as_text(canonical.get("Филиал Tallanto")),
            "AMO контактов": amo_result["contact_count"],
            "AMO ID контактов": join_unique(item["id"] for item in amo_result["contacts"]),
            "AMO контакты": join_unique(item["name"] for item in amo_result["contacts"]),
            "AMO ответственные": join_unique(item["responsible"] for item in amo_result["contacts"]),
            "AMO email": join_unique(item["emails"] for item in amo_result["contacts"]),
            "AMO Telegram ID": join_unique(item["telegram_id"] for item in amo_result["contacts"]),
            "AMO Telegram username": join_unique(item["telegram_username"] for item in amo_result["contacts"]),
            "AMO ID Tallanto": join_unique(item["tallanto_id"] for item in amo_result["contacts"]),
            "AMO Филиал Tallanto": join_unique(item["tallanto_branch"] for item in amo_result["contacts"]),
            "AMO сделок в связке": join_unique(item["lead_count"] for item in amo_result["contacts"]),
            "AMO открытых среди последних": amo_result["open_lead_count_recent"],
            "AMO выигранных среди последних": amo_result["won_lead_count_recent"],
            "AMO проигранных среди последних": amo_result["lost_lead_count_recent"],
            "AMO последние сделки": amo_result["recent_leads_text"],
            "AMO последние UTM": amo_result["utm_text"],
            "AMO причина/стоп-фактор": amo_result["latest_reason"],
            "AMO ошибка": amo_result.get("error", ""),
            "Tallanto контактов": tallanto_result["count"],
            "Tallanto способ матчинга": tallanto_result["matched_via"],
            "Tallanto ID": join_unique(tallanto_result["contact_ids"]),
            "Tallanto контакты": tallanto_result["contacts_text"],
            "Tallanto ответственные": tallanto_result["responsibles_text"],
            "Tallanto филиалы": tallanto_result["branches_text"],
            "Tallanto тип клиента": tallanto_result["type_clients_text"],
            "Tallanto сделок": tallanto_result["opportunity_count"],
            "Tallanto заявок": tallanto_result["request_count"],
            "Tallanto оплат": tallanto_result["finance_count"],
            "Tallanto групп": tallanto_result["course_relation_count"],
            "Tallanto занятий": tallanto_result["class_relation_count"],
            "Tallanto opportunities": tallanto_result["opportunity_text"],
            "Tallanto requests": tallanto_result["request_text"],
            "Tallanto finances": tallanto_result["finance_text"],
            "Tallanto краткий контекст": tallanto_result["compact_context"],
            "Tallanto ошибка поиска": tallanto_result["search_error"],
            "Сегмент outreach": as_text(outreach_row.get("Сегмент")),
            "Что предложить": as_text(outreach_row.get("Что предложить")),
            "Цель сообщения": as_text(outreach_row.get("Цель сообщения")),
            "Контекст для персонализации": as_text(outreach_row.get("Контекст для персонализации")),
            "Стратегия касания": as_text(outreach_row.get("Стратегия касания")),
        }
        possible, basis, offer = phone_offer_hypothesis(flat_row)
        flat_row["Персональное предложение возможно"] = possible
        flat_row["Основа для предложения"] = basis
        flat_row["Гипотеза предложения"] = offer
        score = 0
        score += 20 if canonical else 0
        score += 18 if amo_result["open_lead_count_recent"] else 0
        score += 12 if tallanto_result["opportunity_count"] or tallanto_result["request_count"] else 0
        score += 8 if flat_row["Есть неотвеченный входящий"] == "Да" else 0
        score += 10 if as_text(canonical.get("Приоритет лида")) == "hot" else 5 if as_text(canonical.get("Приоритет лида")) == "warm" else 0
        try:
            score += min(20, int(float(as_text(canonical.get("Вероятность продажи, %")) or 0) / 5))
        except Exception:
            pass
        flat_row["Скор полезности"] = score

        rows_for_jsonl.append(
            {
                "phone": phone,
                "telegram": phone_row,
                "canonical_contact": canonical,
                "amo": amo_result,
                "tallanto": tallanto_result,
                "outreach": outreach_row,
                "offer_possible": possible,
                "offer_basis": basis,
                "offer_hypothesis": offer,
                "utility_score": score,
            }
        )
        flat_rows.append(flat_row)

    all_df = pd.DataFrame(flat_rows).sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).reset_index(drop=True)
    matches_df = all_df[(all_df["AMO контактов"] > 0) | (all_df["Tallanto контактов"] > 0) | (all_df["Есть exact match в рабочем слое"] == "Да")].copy()
    prospects_df = all_df[(all_df["Персональное предложение возможно"] == "Да") & (all_df["Скор полезности"] >= 20)].copy()

    summary = {
        "canonical_dir": str(canonical_dir),
        "phones_total": int(len(all_df)),
        "matched_to_working_layer": int((all_df["Есть exact match в рабочем слое"] == "Да").sum()),
        "matched_to_amo": int((all_df["AMO контактов"] > 0).sum()),
        "matched_to_tallanto": int((all_df["Tallanto контактов"] > 0).sum()),
        "offer_possible": int((all_df["Персональное предложение возможно"] == "Да").sum()),
        "high_utility_phones": int((all_df["Скор полезности"] >= 30).sum()),
        "amo_contacts_total": int(pd.to_numeric(all_df["AMO контактов"], errors="coerce").fillna(0).sum()),
        "tallanto_contacts_total": int(pd.to_numeric(all_df["Tallanto контактов"], errors="coerce").fillna(0).sum()),
    }

    jsonl_path = out_dir / "telegram_phone_live_enrichment.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for row in rows_for_jsonl:
            handle.write(json.dumps(row, ensure_ascii=False, default=json_default) + "\n")

    csv_path = out_dir / "telegram_phone_live_enrichment.csv"
    all_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    workbook = Workbook()
    build_summary_sheet(workbook, summary)
    add_dataframe_sheet(workbook, "Все телефоны", all_df)
    add_dataframe_sheet(workbook, "Только матчи", matches_df)
    add_dataframe_sheet(workbook, "Перспективные", prospects_df)
    xlsx_path = out_dir / "telegram_phone_live_enrichment.xlsx"
    workbook.save(xlsx_path)

    summary_path = out_dir / "telegram_phone_live_enrichment_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"status": "ok", **summary, "xlsx": str(xlsx_path), "csv": str(csv_path), "jsonl": str(jsonl_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
