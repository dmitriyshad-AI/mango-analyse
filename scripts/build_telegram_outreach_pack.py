#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def norm_phone(x: Any) -> str:
    s = "" if pd.isna(x) else str(x)
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) == 11 and digits.startswith("7"):
        return "+" + digits
    return ""


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open() as f:
        for line in f:
            rows.append(json.loads(line))
    return rows


def first_nonempty(*values: Any) -> str:
    for v in values:
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() != "nan":
            return s
    return ""


def yes_no(val: Any) -> str:
    return "Да" if bool(val) else "Нет"


def as_excel_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return ""
        if value.tzinfo is not None:
            value = value.tz_convert("UTC").tz_localize(None)
        return value.to_pydatetime()
    return value


def build_pack(export_root: Path, canonical_dir: Path, out_dir: Path) -> dict[str, Any]:
    max_dir = export_root / "local_vm_2024-04-01_max"
    with_dir = export_root / "local_vm_2024-04-01_with_contacts"

    crm = pd.read_csv(max_dir / "crm_contacts.csv")
    crm.columns = [c.lstrip("\ufeff") for c in crm.columns]
    crm = crm.rename(
        columns={
            "name": "dialog_name",
            "phone": "crm_phone",
            "is_contact": "crm_is_contact",
            "is_mutual_contact": "crm_is_mutual_contact",
            "has_media": "crm_has_media",
            "media_count": "crm_media_count",
            "telegram_username": "crm_telegram_username",
        }
    )

    dialogs = pd.DataFrame(load_jsonl(max_dir / "dialogs.jsonl"))
    dialogs = dialogs.rename(
        columns={
            "name": "dialog_name",
            "phone": "dialog_phone",
            "username": "dialog_username",
            "is_contact": "tg_is_contact",
            "is_mutual_contact": "tg_is_mutual_contact",
            "about": "tg_about",
            "first_name": "tg_first_name",
            "last_name": "tg_last_name",
        }
    )

    with_msgs = pd.DataFrame(load_jsonl(with_dir / "messages.jsonl"))
    with_msgs["date"] = pd.to_datetime(with_msgs["date"], utc=True, errors="coerce")

    msg_agg: list[dict[str, Any]] = []
    for dialog_id, group in with_msgs.groupby("dialog_id"):
        group = group.sort_values("date")
        inbound = group[~group["out"].astype(bool)]
        outbound = group[group["out"].astype(bool)]
        last = group.iloc[-1]
        msg_agg.append(
            {
                "dialog_id": dialog_id,
                "tg_message_count": len(group),
                "tg_inbound_count": len(inbound),
                "tg_outbound_count": len(outbound),
                "tg_last_message_date": last["date"],
                "tg_last_message_out": bool(last["out"]),
                "tg_last_message_text": (last.get("text") or "")[:500],
                "tg_last_inbound_date": inbound["date"].max() if len(inbound) else pd.NaT,
                "tg_last_outbound_date": outbound["date"].max() if len(outbound) else pd.NaT,
                "tg_last_inbound_text": (inbound.iloc[-1].get("text") or "")[:500] if len(inbound) else "",
                "tg_last_outbound_text": (outbound.iloc[-1].get("text") or "")[:500] if len(outbound) else "",
                "tg_unanswered_inbound": not bool(last["out"]),
                "tg_media_message_count": int(group["has_media"].fillna(False).astype(bool).sum()),
                "tg_media_with_path_count": int(group["media_path"].fillna("").ne("").sum()),
            }
        )
    msg_agg_df = pd.DataFrame(msg_agg)

    combined = crm.merge(
        dialogs[
            [
                "dialog_id",
                "dialog_name",
                "dialog_phone",
                "dialog_username",
                "tg_about",
                "tg_is_contact",
                "tg_is_mutual_contact",
                "tg_first_name",
                "tg_last_name",
            ]
        ],
        on=["dialog_id", "dialog_name"],
        how="left",
    ).merge(msg_agg_df, on="dialog_id", how="left")

    combined["phone_norm"] = combined["dialog_phone"].map(norm_phone)

    contacts = pd.read_csv(canonical_dir / "master_contacts_ru.csv")
    contacts.columns = [c.lstrip("\ufeff") for c in contacts.columns]
    contacts["phone_norm"] = contacts["Телефон клиента"].map(norm_phone)
    contact_cols = [
        "phone_norm",
        "Краткая история общения",
        "Хронология общения (последние 5 касаний)",
        "Продукты интереса",
        "Рекомендуемый продукт",
        "Возражения",
        "Следующий шаг",
        "Рекомендуемая дата следующего контакта",
        "Приоритет лида",
        "Вероятность продажи, %",
        "ФИО родителя",
        "ФИО ребенка",
        "Email",
    ]
    combined = combined.merge(contacts[contact_cols], on="phone_norm", how="left")
    combined["matched_contact"] = combined["phone_norm"].ne("") & combined["Краткая история общения"].notna()

    interest_re = re.compile(r"математ|физик|информат|курс|олимпиад|егэ|огэ|онлайн|очно|распис|занят|летн|лагер|школ|стоим|цен|скид|договор|оплат|групп|абонем", re.I)
    warm_re = re.compile(r"пока не решили|подума|позже|позднее|напиш|перезвон|после экзамен|летом|в август|на следующий год|жд[её]м|когда можно будет|интересует|хотим|хочу", re.I)
    dead_re = re.compile(r"неактуал|не интерес|не оставлял|ошиблись|не надо|не нужно|отказ|не будем|не планир", re.I)
    existing_re = re.compile(r"чат группы|преподават|домаш|квитанц|абонемент|группа|договор|оплат|счет|сч[её]т|документ|справк|возврат", re.I)
    followup_re = re.compile(r"скид|стоим|когда могу оплатить|приступить|лагер|летн|следующ|продолж", re.I)

    text_blob = (
        combined["last_inbound_text"].fillna("")
        + " "
        + combined["last_outbound_text"].fillna("")
        + " "
        + combined["tg_last_message_text"].fillna("")
        + " "
        + combined["Краткая история общения"].fillna("")
        + " "
        + combined["Продукты интереса"].fillna("")
        + " "
        + combined["Следующий шаг"].fillna("")
    )
    combined["signal_interest"] = text_blob.str.contains(interest_re)
    combined["signal_warm"] = text_blob.str.contains(warm_re)
    combined["signal_dead"] = text_blob.str.contains(dead_re)
    combined["signal_existing"] = text_blob.str.contains(existing_re)
    combined["signal_followup"] = text_blob.str.contains(followup_re)
    combined["prob_num"] = pd.to_numeric(combined["Вероятность продажи, %"], errors="coerce").fillna(0)
    combined["priority_weight"] = combined["Приоритет лида"].map({"hot": 3, "warm": 2, "cold": 1}).fillna(0)
    now_utc = pd.Timestamp.now(tz="UTC")
    combined["days_since_last_message"] = (now_utc - pd.to_datetime(combined["tg_last_message_date"], utc=True, errors="coerce")).dt.days

    combined["score"] = 0.0
    combined.loc[combined["matched_contact"], "score"] += 25
    combined.loc[combined["tg_unanswered_inbound"].fillna(False), "score"] += 18
    combined.loc[combined["signal_interest"], "score"] += 15
    combined.loc[combined["signal_warm"], "score"] += 15
    combined.loc[combined["signal_existing"], "score"] += 8
    combined.loc[combined["signal_followup"], "score"] += 5
    combined["score"] += combined["priority_weight"] * 10
    combined["score"] += (combined["prob_num"] / 5).clip(upper=20)
    combined.loc[combined["tg_media_with_path_count"].fillna(0) > 0, "score"] += 2
    combined.loc[combined["tg_is_contact"].fillna(False), "score"] += 4
    combined.loc[(combined["days_since_last_message"] <= 30).fillna(False), "score"] += 10
    combined.loc[(combined["days_since_last_message"] > 180).fillna(False), "score"] -= 8
    combined.loc[combined["signal_dead"], "score"] -= 30

    combined["outreach_segment"] = "Низкий приоритет"
    combined.loc[combined["signal_dead"], "outreach_segment"] = "Не писать автоматически"
    combined.loc[(~combined["signal_dead"]) & combined["signal_existing"] & (combined["signal_interest"] | combined["signal_followup"]), "outreach_segment"] = "Допродажа / продление"
    combined.loc[(~combined["signal_dead"]) & combined["signal_warm"] & ~combined["signal_existing"], "outreach_segment"] = "Реактивация теплого диалога"
    combined.loc[(~combined["signal_dead"]) & combined["tg_unanswered_inbound"].fillna(False) & combined["signal_interest"], "outreach_segment"] = "Нужен быстрый ответ"
    combined.loc[(~combined["signal_dead"]) & combined["matched_contact"] & (combined["prob_num"] >= 60), "outreach_segment"] = "Высокий потенциал по CRM"

    combined["sales_fit"] = ~(combined["signal_existing"] & ~combined["signal_followup"] & ~combined["matched_contact"] & ~combined["signal_warm"])

    combined["recommended_offer"] = ""
    combined.loc[combined["Рекомендуемый продукт"].fillna("").ne(""), "recommended_offer"] = combined["Рекомендуемый продукт"].fillna("")
    combined.loc[combined["recommended_offer"].eq("") & text_blob.str.contains("математ", case=False, na=False), "recommended_offer"] = "Курс по математике / усиление текущего обучения"
    combined.loc[combined["recommended_offer"].eq("") & text_blob.str.contains("информат", case=False, na=False), "recommended_offer"] = "Курс по информатике / программированию"
    combined.loc[combined["recommended_offer"].eq("") & text_blob.str.contains("физик", case=False, na=False), "recommended_offer"] = "Курс по физике"
    combined.loc[combined["recommended_offer"].eq("") & text_blob.str.contains("олимпиад", case=False, na=False), "recommended_offer"] = "Олимпиадный трек"
    combined.loc[combined["recommended_offer"].eq("") & text_blob.str.contains("летн|лагер|школ", case=False, na=False), "recommended_offer"] = "Летняя школа / интенсив"
    combined.loc[combined["recommended_offer"].eq("") & combined["signal_existing"], "recommended_offer"] = "Продление, второй предмет или летняя программа"
    combined.loc[combined["recommended_offer"].eq(""), "recommended_offer"] = "Уточнение потребности и мягкий follow-up"

    combined["message_goal"] = "Аккуратно уточнить актуальность интереса"
    combined.loc[combined["outreach_segment"] == "Нужен быстрый ответ", "message_goal"] = "Ответить на последний вопрос и закрыть зависший запрос"
    combined.loc[combined["outreach_segment"] == "Реактивация теплого диалога", "message_goal"] = "Вернуть клиента в диалог мягким follow-up"
    combined.loc[combined["outreach_segment"] == "Допродажа / продление", "message_goal"] = "Предложить продление, второй предмет или ближайшую программу"
    combined.loc[combined["outreach_segment"] == "Высокий потенциал по CRM", "message_goal"] = "Перевести интерес в созвон, оплату или запись"

    def build_personalization(row: pd.Series) -> str:
        parts = []
        if first_nonempty(row.get("last_inbound_text")):
            parts.append(f"Последний входящий тезис: {first_nonempty(row.get('last_inbound_text'))}")
        elif first_nonempty(row.get("tg_last_inbound_text")):
            parts.append(f"Последний входящий тезис: {first_nonempty(row.get('tg_last_inbound_text'))}")
        if first_nonempty(row.get("Следующий шаг")):
            parts.append(f"Следующий шаг из CRM: {first_nonempty(row.get('Следующий шаг'))}")
        if first_nonempty(row.get("Возражения")):
            parts.append(f"Возражения: {first_nonempty(row.get('Возражения'))}")
        if first_nonempty(row.get("Продукты интереса")):
            parts.append(f"Интерес: {first_nonempty(row.get('Продукты интереса'))}")
        return " | ".join(parts[:4])

    def build_strategy(row: pd.Series) -> str:
        seg = row.get("outreach_segment")
        offer = row.get("recommended_offer")
        if seg == "Нужен быстрый ответ":
            return f"Ответить по существу и мягко довести до следующего действия. Предложение: {offer}."
        if seg == "Допродажа / продление":
            return f"Использовать текущий контекст обучения и предложить {offer} без давления."
        if seg == "Реактивация теплого диалога":
            return f"Мягко вернуться к диалогу и проверить, актуален ли {offer}."
        if seg == "Высокий потенциал по CRM":
            return f"Напомнить контекст и перевести диалог к записи или созвону по направлению: {offer}."
        return f"Коротко уточнить актуальность интереса и предложить {offer}."

    def build_prompt(row: pd.Series) -> str:
        name = first_nonempty(row.get("dialog_name"), "клиент")
        tone = "Спокойный, профессиональный, короткий, без давления."
        constraints = "Не упоминать анализ. Не писать длиннее 500 символов. Без навязчивой продажи."
        cta = row.get("message_goal")
        context = build_personalization(row)
        offer = row.get("recommended_offer")
        return (
            f"Сделай персональный черновик первого сообщения в Telegram для диалога '{name}'. "
            f"Цель: {cta}. Что предлагаем: {offer}. Контекст: {context}. "
            f"Тон: {tone} {constraints} Сообщение должно выглядеть как естественное продолжение существующего диалога."
        )

    combined["personalization_context"] = combined.apply(build_personalization, axis=1)
    combined["outreach_strategy"] = combined.apply(build_strategy, axis=1)
    combined["openclaw_prompt"] = combined.apply(build_prompt, axis=1)

    candidates = combined[
        (combined["outreach_segment"] != "Не писать автоматически")
        & combined["sales_fit"]
        & (
            (combined["score"] >= 35)
            | combined["tg_unanswered_inbound"].fillna(False)
            | (combined["matched_contact"] & (combined["prob_num"] >= 55))
        )
    ].copy()
    candidates = candidates.sort_values(["score", "tg_last_message_date"], ascending=[False, False])
    openclaw = candidates[
        (candidates["score"] >= 55)
        & (candidates["outreach_segment"] != "Низкий приоритет")
    ].copy().head(120)

    phone_present = combined["dialog_phone"].fillna("").map(norm_phone).ne("")
    username_present = combined["dialog_username"].fillna("").ne("")

    summary = {
        "source_dirs": {
            "max": str(max_dir),
            "with_contacts": str(with_dir),
        },
        "dialogs_total": int(len(crm)),
        "dialogs_with_phone": int(phone_present.sum()),
        "dialogs_with_username": int(username_present.sum()),
        "dialogs_phone_or_username": int((phone_present | username_present).sum()),
        "crm_matched_dialogs": int(combined["matched_contact"].sum()),
        "candidate_dialogs": int(len(candidates)),
        "openclaw_dialogs": int(len(openclaw)),
        "segment_counts": {k: int(v) for k, v in candidates["outreach_segment"].value_counts().to_dict().items()},
        "attachments": {
            "max_media_files": int(sum(1 for p in (max_dir / "media").rglob("*") if p.is_file())),
            "with_contacts_media_files": int(sum(1 for p in (with_dir / "media").rglob("*") if p.is_file())),
            "with_contacts_media_messages": int(with_msgs["has_media"].fillna(False).astype(bool).sum()),
            "with_contacts_media_messages_with_path": int(with_msgs["media_path"].fillna("").ne("").sum()),
        },
        "notes": [
            "local_vm_2024-04-01_max is the best source for phone and dialog metadata.",
            "local_vm_2024-04-01_with_contacts is the best source for message-level media_path linkage.",
            "Only dialogs with normalized phone can be reliably joined to the current CRM/contact layer.",
        ],
    }

    out_dir.mkdir(parents=True, exist_ok=True)

    export_cols = [
        "dialog_id",
        "dialog_name",
        "phone_norm",
        "dialog_username",
        "outreach_segment",
        "score",
        "recommended_offer",
        "message_goal",
        "tg_last_message_date",
        "tg_unanswered_inbound",
        "matched_contact",
        "Приоритет лида",
        "Вероятность продажи, %",
        "Продукты интереса",
        "Возражения",
        "Следующий шаг",
        "last_inbound_text",
        "last_outbound_text",
        "tg_last_message_text",
        "personalization_context",
        "outreach_strategy",
        "openclaw_prompt",
    ]

    def localize(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df = df.rename(
            columns={
                "dialog_id": "ID диалога Telegram",
                "dialog_name": "Диалог",
                "phone_norm": "Телефон",
                "dialog_username": "Username",
                "outreach_segment": "Сегмент",
                "score": "Скор",
                "recommended_offer": "Что предложить",
                "message_goal": "Цель сообщения",
                "tg_last_message_date": "Дата последнего сообщения",
                "tg_unanswered_inbound": "Последнее сообщение от клиента",
                "matched_contact": "Сопоставлен с CRM",
                "Приоритет лида": "Приоритет лида в CRM",
                "Вероятность продажи, %": "Вероятность продажи, %",
                "Продукты интереса": "Продукты интереса",
                "Возражения": "Возражения",
                "Следующий шаг": "Следующий шаг из CRM",
                "last_inbound_text": "Последний входящий из summary",
                "last_outbound_text": "Последний исходящий из summary",
                "tg_last_message_text": "Последнее сообщение по логу",
                "personalization_context": "Контекст для персонализации",
                "outreach_strategy": "Стратегия касания",
                "openclaw_prompt": "Промпт для openClaw",
            }
        )
        df["Последнее сообщение от клиента"] = df["Последнее сообщение от клиента"].map(yes_no)
        df["Сопоставлен с CRM"] = df["Сопоставлен с CRM"].map(yes_no)
        return df

    openclaw_df = localize(openclaw[export_cols])
    candidates_df = localize(candidates[export_cols])

    xlsx_path = out_dir / "telegram_openclaw_outreach_pack.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Показатель", "Значение"])
    for row in [
        ["Всего диалогов", summary["dialogs_total"]],
        ["Диалогов с телефоном", summary["dialogs_with_phone"]],
        ["Диалогов с username", summary["dialogs_with_username"]],
        ["Диалогов с телефоном или username", summary["dialogs_phone_or_username"]],
        ["Диалогов, сопоставленных с CRM", summary["crm_matched_dialogs"]],
        ["Кандидатов на outreach", summary["candidate_dialogs"]],
        ["Диалогов в openClaw shortlist", summary["openclaw_dialogs"]],
        ["Файлов вложений в local_vm_2024-04-01_max", summary["attachments"]["max_media_files"]],
        ["Файлов вложений в local_vm_2024-04-01_with_contacts", summary["attachments"]["with_contacts_media_files"]],
        ["Сообщений с media в with_contacts", summary["attachments"]["with_contacts_media_messages"]],
        ["Сообщений с media_path в with_contacts", summary["attachments"]["with_contacts_media_messages_with_path"]],
    ]:
        ws.append(row)
    ws.append([])
    ws.append(["Сегмент", "Количество"])
    for k, v in summary["segment_counts"].items():
        ws.append([k, v])

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"

    def add_sheet(name: str, df: pd.DataFrame) -> None:
        ws = wb.create_sheet(title=name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append([as_excel_value(v) for v in row])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for idx, col in enumerate(df.columns, start=1):
            samples = [str(col)] + [str(v) for v in df.iloc[:200, idx - 1].fillna("").astype(str).tolist()]
            ws.column_dimensions[get_column_letter(idx)].width = min(max(max(len(s) for s in samples) + 2, 12), 60)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    add_sheet("OpenClaw shortlist", openclaw_df)
    add_sheet("Все кандидаты", candidates_df)
    wb.save(xlsx_path)

    csv_path = out_dir / "telegram_openclaw_shortlist.csv"
    json_path = out_dir / "telegram_outreach_summary.json"
    md_path = out_dir / "telegram_outreach_summary.md"
    openclaw_df.to_csv(csv_path, index=False)
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2))
    md_path.write_text(
        "\n".join(
            [
                "# Telegram Outreach Analysis",
                "",
                f"- Всего диалогов: {summary['dialogs_total']}",
                f"- Диалогов с телефоном: {summary['dialogs_with_phone']}",
                f"- Диалогов с username: {summary['dialogs_with_username']}",
                f"- Сопоставлено с CRM: {summary['crm_matched_dialogs']}",
                f"- Кандидатов на outreach: {summary['candidate_dialogs']}",
                f"- В openClaw shortlist: {summary['openclaw_dialogs']}",
                "",
                "## Segment counts",
                *[f"- {k}: {v}" for k, v in summary["segment_counts"].items()],
                "",
                "## Notes",
                *[f"- {note}" for note in summary["notes"]],
            ]
        )
    )

    return {
        "xlsx": str(xlsx_path),
        "csv": str(csv_path),
        "json": str(json_path),
        "md": str(md_path),
        "summary": summary,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--export-root", required=True)
    parser.add_argument("--canonical-dir", required=True)
    parser.add_argument("--out-dir", required=True)
    args = parser.parse_args()
    result = build_pack(Path(args.export_root), Path(args.canonical_dir), Path(args.out_dir))
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
