from __future__ import annotations

import hashlib
import json
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from scripts import export_daily_mango_calls_resolve as exporter
from scripts import publish_daily_mango_calls_google as publisher


class Response:
    def __init__(self, payload: object = None, *, content: bytes = b"", status: int = 200):
        self._payload, self.content, self.status_code = payload, content, status

    def json(self) -> object:
        return self._payload


class BrokenJsonResponse(Response):
    def json(self) -> object:
        raise ValueError("broken")


class Session:
    def __init__(self, plan: dict[str, object], *, found: list[dict[str, str]] | None = None,
                 mismatch: bool = False, private: bool = True, post_error: bool = False):
        self.plan, self.xlsx, self.found = plan, plan["upload_bytes"], found or []
        self.mismatch, self.private, self.post_error = mismatch, private, post_error
        self.posts, self.patches, self.deletes, self.gets = [], [], 0, []
        first_props = found[0].get("appProperties", {}) if found else {}
        self.state = str(first_props.get("mango_report_state") or ("published" if found else "checking"))

    def get(self, url: str, **kwargs: object) -> Response:
        self.gets.append((url, kwargs))
        if url.endswith("/files"):
            return Response({"files": self.found})
        if url.endswith("/export"):
            return Response(content=b"bad" if self.mismatch else self.xlsx)
        if url.endswith("/folder_123456789"):
            return Response({
                "id": "folder_123456789", "mimeType": publisher.GOOGLE_FOLDER_MIME,
                "capabilities": {"canAddChildren": True},
                "permissions": [{"type": "user" if self.private else "anyone", "role": "writer"}],
            })
        return Response({
            "id": "sheet-123", "name": self.plan["name"] if self.state == "published" else self.plan["temporary_name"],
            "mimeType": publisher.GOOGLE_SHEET_MIME, "parents": ["folder_123456789"],
            "appProperties": publisher.expected_properties(self.plan, self.state),
            "permissions": [{"type": "user", "role": "writer"}],
        })

    def post(self, url: str, **kwargs: object) -> Response:
        self.posts.append((url, kwargs))
        if self.post_error:
            raise TimeoutError("response lost")
        return Response({"id": "sheet-123"})

    def patch(self, url: str, **kwargs: object) -> Response:
        self.patches.append((url, kwargs))
        self.state = str(kwargs.get("json", {}).get("appProperties", {}).get("mango_report_state") or self.state)
        return Response({"id": "sheet-123"})

    def delete(self, *_: object, **__: object) -> Response:
        self.deletes += 1
        return Response({})


def fixture_report(tmp_path: Path) -> tuple[Path, dict[str, object], bytes]:
    root = tmp_path / "reports"
    root.mkdir()
    xlsx = root / "report.xlsx"
    workbook = Workbook()
    workbook.active.title = "Сводка"
    workbook.active.append(["Показатель", "Значение"])
    workbook.active.append(["День", "2026-07-29"])
    calls = workbook.create_sheet("Звонки")
    calls.append(["Расшифровка", "Следующий шаг"])
    calls.append(["полный текст " + "а" * 30_000, "Отправить договор"])
    workbook.create_sheet("Проблемы данных").append(["Причина"])
    workbook.create_sheet("Описание полей").append(["Правило", "Описание"])
    workbook.save(xlsx)
    raw = xlsx.read_bytes()
    manifest = {
        "schema_version": publisher.SCHEMA,
        "day": "2026-07-29",
        "generated_at": "2026-07-30T01:15:00+00:00",
        "content_sha256": "a" * 64,
        "xlsx": xlsx.name,
        "xlsx_sha256": hashlib.sha256(raw).hexdigest(),
    }
    (root / "Отчёт РОП по звонкам 2026-07-29.manifest.json").write_text(json.dumps(manifest), encoding="utf-8")
    return root, manifest, raw


def test_export_and_publisher_share_schema_contract(tmp_path: Path) -> None:
    assert publisher.SCHEMA == exporter.EXPORT_SCHEMA_VERSION
    root, _, _ = fixture_report(tmp_path)
    assert "— v4-" in publisher.load_plan(root, date(2026, 7, 29))["name"]


def test_dry_run_validates_without_google_import_or_network(tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys: pytest.CaptureFixture[str]) -> None:
    root, _, _ = fixture_report(tmp_path)
    monkeypatch.setattr(publisher, "authorized_session", lambda *_: pytest.fail("Google auth must not run"))

    assert publisher.main(["--report-root", str(root), "--day", "2026-07-29"]) == 0

    assert json.loads(capsys.readouterr().out)["status"] == "dry_run"


def test_manifest_path_escape_and_corrupt_xlsx_fail_closed(tmp_path: Path) -> None:
    root, manifest, _ = fixture_report(tmp_path)
    manifest_path = root / "Отчёт РОП по звонкам 2026-07-29.manifest.json"
    manifest["xlsx"] = "../outside.xlsx"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    with pytest.raises(RuntimeError, match="escapes"):
        publisher.load_plan(root, date(2026, 7, 29))
    manifest["xlsx"] = "report.xlsx"
    manifest["xlsx_sha256"] = "0" * 64
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    with pytest.raises(RuntimeError, match="does not match"):
        publisher.load_plan(root, date(2026, 7, 29))


def test_manifest_wrong_day_and_oversized_xlsx_fail_closed(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    root, manifest, _ = fixture_report(tmp_path)
    manifest_path = root / "Отчёт РОП по звонкам 2026-07-29.manifest.json"
    manifest["day"] = "2026-07-28"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    with pytest.raises(RuntimeError, match="day is invalid"):
        publisher.load_plan(root, date(2026, 7, 29))
    manifest["day"] = "2026-07-29"
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    monkeypatch.setattr(publisher, "MULTIPART_LIMIT", 1)
    with pytest.raises(RuntimeError, match="does not match"):
        publisher.load_plan(root, date(2026, 7, 29))


def test_exact_generation_is_reused_after_content_readback(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan, found=[{"id": "sheet-123", "appProperties": publisher.expected_properties(plan, "published")}])

    result = publisher.publish(session, "folder_123456789", plan)

    assert result["status"] == "reused" and not session.posts


def test_ambiguous_generation_fails_closed(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan, found=[{"id": "one"}, {"id": "two"}])
    with pytest.raises(RuntimeError, match="ambiguous"):
        publisher.publish(session, "folder_123456789", plan)


def test_invalid_lookup_and_json_fail_closed(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    with pytest.raises(RuntimeError, match="lookup is invalid"):
        publisher.publish(Session(plan, found=[{}]), "folder_123456789", plan)
    with pytest.raises(RuntimeError, match="invalid JSON"):
        publisher.response_json(BrokenJsonResponse())


def test_create_converts_xlsx_with_folder_and_properties(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan)

    result = publisher.publish(session, "folder_123456789", plan)

    assert result["status"] == "created"
    _, request = session.posts[0]
    body = request["data"]
    assert publisher.GOOGLE_SHEET_MIME.encode() in body
    assert b"folder_123456789" in body and b"mango_report_content_sha256" in body
    assert plan["temporary_name"].encode() in body and plan["name"].encode() not in body
    assert request["params"]["supportsAllDrives"] == "true"
    assert session.state == "published" and session.patches[-1][1]["json"]["name"] == plan["name"]


def test_private_writable_folder_is_required_before_upload(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan, private=False)
    with pytest.raises(RuntimeError, match="not private and writable"):
        publisher.publish(session, "folder_123456789", plan)
    assert not session.posts


def test_generation_lookup_includes_day_for_identical_empty_content(tmp_path: Path) -> None:
    root, manifest, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan, found=[{"id": "sheet-123", "appProperties": publisher.expected_properties(plan, "published")}])
    publisher.publish(session, "folder_123456789", plan)
    query = next(kwargs["params"]["q"] for url, kwargs in session.gets if url.endswith("/files"))
    assert "mango_report_day" in query and "2026-07-29" in query and publisher.SCHEMA in query
    assert "mango_report_state" not in query
    assert manifest["content_sha256"] in query


def test_upload_copy_replaces_local_hyperlinks_without_changing_source(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    xlsx = root / "report.xlsx"
    workbook = load_workbook(xlsx)
    cell = workbook["Звонки"]["A2"]
    cell.hyperlink = "Расшифровки/call_123.txt"
    workbook.save(xlsx)
    workbook.close()
    source_with_link = xlsx.read_bytes()

    uploaded = publisher.google_upload_bytes(xlsx)
    converted = load_workbook(__import__("io").BytesIO(uploaded))
    assert converted["Звонки"]["A2"].value == "Яндекс Диск: Расшифровки/call_123.txt"
    assert converted["Звонки"]["A2"].hyperlink is None
    converted.close()
    assert xlsx.read_bytes() == source_with_link


def test_workbook_digest_catches_cell_type_and_hyperlink_changes(tmp_path: Path) -> None:
    path = tmp_path / "typed.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "1"
    workbook.active["A1"].hyperlink = "https://example.test/one"
    workbook.save(path)
    first = publisher.workbook_digest(path)
    workbook.active["A1"] = 1
    workbook.active["A1"].hyperlink = None
    workbook.save(path)
    workbook.close()
    assert publisher.workbook_digest(path) != first


def test_remote_content_mismatch_blocks_success(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan, mismatch=True)
    with pytest.raises(RuntimeError, match="content readback mismatch"):
        publisher.publish(session, "folder_123456789", plan)
    assert session.deletes == 0
    assert session.patches[-1][1]["json"]["name"] == plan["temporary_name"]


def test_interrupted_checking_generation_is_resumed_without_duplicate(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    session = Session(plan, found=[{
        "id": "sheet-123", "appProperties": publisher.expected_properties(plan, "checking"),
    }])

    result = publisher.publish(session, "folder_123456789", plan)

    assert result["status"] == "resumed" and not session.posts
    assert session.state == "published"


def test_lost_create_response_blocks_duplicate_until_generation_is_found(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    journal = tmp_path / "create.json"
    failed = Session(plan, post_error=True)
    with pytest.raises(TimeoutError, match="response lost"):
        publisher.publish(failed, "folder_123456789", plan, journal=journal)
    assert journal.is_file() and len(failed.posts) == 1

    retry = Session(plan)
    with pytest.raises(RuntimeError, match="uncertain"):
        publisher.publish(retry, "folder_123456789", plan, journal=journal)
    assert not retry.posts

    found = Session(plan, found=[{
        "id": "sheet-123", "appProperties": publisher.expected_properties(plan, "checking"),
    }])
    assert publisher.publish(found, "folder_123456789", plan, journal=journal)["status"] == "resumed"
    assert not journal.exists() and not found.posts


def test_google_create_journal_cannot_cross_generations(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    journal = tmp_path / "create.json"
    journal.write_text(json.dumps({"day": "2026-07-28", "content_sha256": "0" * 64,
                                   "folder_id": "folder_123456789"}), encoding="utf-8")

    with pytest.raises(RuntimeError, match="another generation"):
        publisher.publish(Session(plan), "folder_123456789", plan, journal=journal)


def test_google_create_journal_cannot_cross_target_folders(tmp_path: Path) -> None:
    root, _, _ = fixture_report(tmp_path)
    plan = publisher.load_plan(root, date(2026, 7, 29))
    journal = tmp_path / "create.json"
    journal.write_text(json.dumps({"day": plan["day"], "content_sha256": plan["content_sha256"],
                                   "folder_id": "folder_original"}), encoding="utf-8")
    with pytest.raises(RuntimeError, match="another generation"):
        publisher.publish(Session(plan), "folder_123456789", plan, journal=journal)


def test_dry_run_does_not_create_publication_lock(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    root, _, _ = fixture_report(tmp_path)
    home = tmp_path / "home"
    home.mkdir()
    monkeypatch.setattr(publisher.Path, "home", classmethod(lambda cls: home))

    publisher.main(["--report-root", str(root), "--day", "2026-07-29"])

    assert not (home / ".mango_local").exists()


def test_credentials_must_be_external_owner_only(tmp_path: Path) -> None:
    repo = tmp_path / "repo"
    repo.mkdir()
    inside = repo / "credentials.json"
    inside.write_text("{}", encoding="utf-8")
    inside.chmod(0o600)
    with pytest.raises(RuntimeError, match="repository"):
        publisher.validate_credentials(inside, repo)
    external = tmp_path / "external.json"
    external.write_text("{}", encoding="utf-8")
    external.chmod(0o644)
    with pytest.raises(RuntimeError, match="0600"):
        publisher.validate_credentials(external, repo)
    yandex = tmp_path / "yandex.disk.localized" / "credentials.json"
    yandex.parent.mkdir()
    yandex.write_text("{}", encoding="utf-8")
    yandex.chmod(0o600)
    with pytest.raises(RuntimeError, match="Yandex Disk"):
        publisher.validate_credentials(yandex, repo)


def test_dry_run_output_does_not_leak_report_pii(tmp_path: Path, capsys: pytest.CaptureFixture[str]) -> None:
    root, _, _ = fixture_report(tmp_path)
    xlsx = root / "report.xlsx"
    workbook = load_workbook(xlsx)
    workbook["Звонки"]["A2"] = "Иванов Иван +79991234567 ivan@example.test"
    workbook.save(xlsx)
    workbook.close()
    manifest_path = root / "Отчёт РОП по звонкам 2026-07-29.manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["xlsx_sha256"] = hashlib.sha256(xlsx.read_bytes()).hexdigest()
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    publisher.main(["--report-root", str(root), "--day", "2026-07-29"])

    output = capsys.readouterr()
    assert "Иванов" not in output.out + output.err
    assert "+79991234567" not in output.out + output.err
    assert "ivan@example.test" not in output.out + output.err
