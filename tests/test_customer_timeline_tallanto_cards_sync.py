from __future__ import annotations

import json
import sqlite3
import re
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Mapping, Optional, Sequence

import pytest
from openpyxl import Workbook

import mango_mvp.customer_timeline.tallanto_cards_sync as cards_sync
from mango_mvp.amocrm_runtime.tallanto_api import TallantoApiError
from mango_mvp.customer_timeline.contracts import CustomerIdentity, IdentityLink, IdentityStatus
from mango_mvp.customer_timeline.store import CustomerTimelineSQLiteStore
from mango_mvp.customer_timeline.family_graph import FamilyGraphConfig, build_family_graph
from mango_mvp.customer_timeline.stage5_money_ingest import refresh_customer_purchases_v1
from mango_mvp.customer_timeline.tallanto_attendance_import import (
    TallantoAttendanceImportConfig,
    run_tallanto_attendance_import,
)
from mango_mvp.customer_timeline.tallanto_cards_sync import (
    TALLANTO_CARDS_SOURCE_SYSTEM,
    TallantoCardsSyncConfig,
    load_tallanto_cards_checkpoint,
    map_raw_contact_to_snapshot_payload,
    run_tallanto_cards_sync,
    universe_fingerprint,
)
from scripts.import_tallanto_payments_to_timeline import (
    TallantoPaymentsImportConfig,
    run_tallanto_payments_import,
)


def test_tallanto_cards_checkpoint_with_truncated_utf8_is_ignored(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    out_root.mkdir()
    (out_root / "tallanto_cards_sync_checkpoint.json").write_bytes(b'{"schema_version":"\xff')

    assert load_tallanto_cards_checkpoint(out_root) == {}


def test_tallanto_cards_checkpoint_with_unknown_schema_is_ignored(tmp_path: Path) -> None:
    out_root = tmp_path / "out"
    out_root.mkdir()
    (out_root / "tallanto_cards_sync_checkpoint.json").write_text(
        json.dumps({"schema_version": "future", "next_offset": 999}),
        encoding="utf-8",
    )

    assert load_tallanto_cards_checkpoint(out_root) == {}


class FakeContactClient:
    """In-memory fake of TallantoApiClient.get_entry_list(module="Contact",
    ...) -- offset/result_count/total_count/next_offset pagination, exactly
    the shape the real Tallanto SugarCRM-style REST API returns. No real API
    call is made anywhere in this test file.
    """

    def __init__(self, contacts: Sequence[Mapping[str, Any]], *, page_size: int = 50) -> None:
        self.contacts = list(contacts)
        self.page_size = page_size
        self.offsets_called: list[int] = []

    def get_entry_list(
        self,
        *,
        module: str,
        select_fields: Optional[Sequence[str]] = None,
        field_values: Optional[Mapping[str, Any]] = None,
        query: Optional[str] = None,
        order_by: Optional[str] = None,
        offset: int = 0,
    ) -> Mapping[str, Any]:
        assert module == "Contact"
        offset = int(offset)
        self.offsets_called.append(offset)
        page = self.contacts[offset : offset + self.page_size]
        total = len(self.contacts)
        payload: dict[str, Any] = {
            "entry_list": page,
            "result_count": len(page),
            "total_count": total,
        }
        next_offset = offset + len(page)
        if next_offset < total:
            payload["next_offset"] = next_offset
        return payload


class FailingContactClient:
    def get_entry_list(self, **_kwargs: Any) -> Mapping[str, Any]:
        raise TallantoApiError("Tallanto Contact request failed.", status_code=502, category="server_error")


class QueryAwareContactClient(FakeContactClient):
    def get_entry_list(self, **kwargs: Any) -> Mapping[str, Any]:
        query = str(kwargs.get("query") or "")
        match = re.search(r"date_modified >= '([^']+)'", query)
        contacts = self.contacts
        if match:
            cutoff = datetime.fromisoformat(match.group(1)).replace(tzinfo=timezone(timedelta(hours=3)))
            contacts = [
                row for row in contacts
                if datetime.fromisoformat(str(row["date_modified"])).replace(tzinfo=cutoff.tzinfo) >= cutoff
            ]
        original, self.contacts = self.contacts, contacts
        try:
            return super().get_entry_list(**kwargs)
        finally:
            self.contacts = original


def _contact(*, contact_id: str, phone: Optional[str] = None, email: Optional[str] = None, amo_id: Optional[str] = None, parent_fio: Optional[str] = None, first_name: str = "Имя", last_name: str = "Фамилия", date_modified: str = "2026-07-20 10:00:00") -> dict[str, Any]:
    row: dict[str, Any] = {"id": contact_id, "first_name": first_name, "last_name": last_name, "date_modified": date_modified}
    if phone is not None:
        row["phone_mobile"] = phone
    if email is not None:
        row["email1"] = email
    if amo_id is not None:
        row["amo_id"] = amo_id
    if parent_fio is not None:
        row["marital_status_c"] = parent_fio
    return row


def _config(
    tmp_path: Path,
    client: Any,
    *,
    max_pages: int = 5,
    timeline_db: Optional[Path] = None,
) -> TallantoCardsSyncConfig:
    return TallantoCardsSyncConfig(
        timeline_db=timeline_db or tmp_path / "staging.sqlite",
        out_root=tmp_path / "out",
        allowed_root=tmp_path,
        client=client,
        max_pages=max_pages,
    )


def _event_count(db_path: Path) -> int:
    with sqlite3.connect(db_path) as con:
        return con.execute(
            "SELECT COUNT(*) FROM timeline_events WHERE source_system='tallanto_snapshot'"
        ).fetchone()[0]


def test_tallanto_cards_sync_completes_in_one_run_and_writes_staging_only(tmp_path: Path) -> None:
    client = FakeContactClient([_contact(contact_id="1", phone="+7 916 111-22-33", email="Parent@Example.com")])

    report = run_tallanto_cards_sync(_config(tmp_path, client))

    assert report["complete"] is True
    assert report["validation_ok"] is True
    assert report["checked"] == 1
    assert report["checked_with_id"] == 1
    assert report["skipped_missing_id"] == 0
    assert report["updated"] == 1
    assert report["unmatched"] == 0
    assert report["conflict"] == 0
    assert report["cursor_time"] is not None
    assert report["safety"]["tallanto_write"] is False
    assert report["safety"]["amo_write"] is False
    assert report["safety"]["staging_db_write"] is True
    assert _event_count(tmp_path / "staging.sqlite") == 1
    with sqlite3.connect(tmp_path / "staging.sqlite") as con:
        link = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='1'"
        ).fetchone()
        assert link is not None
        phone_link = con.execute(
            "SELECT match_class FROM identity_links WHERE link_type='phone' AND link_value='+79161112233'"
        ).fetchone()
        assert phone_link[0] == "strong_unique"
        assert con.execute("SELECT event_at FROM timeline_events").fetchone()[0].endswith("+03:00")


def test_tallanto_cards_sync_reuses_db_and_fetches_only_recent_changes(tmp_path: Path) -> None:
    old = (datetime.now() - timedelta(days=3)).replace(microsecond=0).isoformat(sep=" ")
    stale_overlap = (datetime.now() - timedelta(hours=12)).replace(microsecond=0).isoformat(sep=" ")
    recent = datetime.now().replace(microsecond=0).isoformat(sep=" ")
    first = QueryAwareContactClient([
        _contact(contact_id="old", date_modified=old),
        _contact(contact_id="stale-overlap", date_modified=stale_overlap),
    ])
    assert run_tallanto_cards_sync(_config(tmp_path, first))["performance"]["mode"] == "full"

    second = run_tallanto_cards_sync(
        _config(tmp_path, QueryAwareContactClient([
            _contact(contact_id="old", date_modified=old),
            _contact(contact_id="stale-overlap", date_modified=stale_overlap),
            _contact(contact_id="new", date_modified=recent),
        ]))
    )

    assert second["performance"]["mode"] == "incremental"
    assert second["performance"]["rows"]["fetched"] == 1
    assert _event_count(tmp_path / "staging.sqlite") == 3


def test_tallanto_cards_sync_repeated_import_does_not_increase_raw_events(tmp_path: Path) -> None:
    contacts = [_contact(contact_id="10", phone="+7 916 222-33-44", email="parent10@example.com")]
    first = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient(contacts)))
    assert first["complete"] is True
    count_after_first = _event_count(tmp_path / "staging.sqlite")
    assert count_after_first == 1

    second = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient(contacts)))

    assert second["complete"] is True
    assert second["updated"] == 0
    assert second["unchanged"] == 1
    assert _event_count(tmp_path / "staging.sqlite") == count_after_first


def test_tallanto_cards_sync_two_children_of_one_family_do_not_collapse(tmp_path: Path) -> None:
    shared_phone = "+7 916 333-44-55"
    db = tmp_path / ".codex_local/staging/customer_timeline.sqlite"
    db.parent.mkdir(parents=True)
    contacts = [
        _contact(contact_id="child-1", phone=shared_phone, parent_fio="Иванова Мария", first_name="Аня"),
        _contact(contact_id="child-2", phone=shared_phone, parent_fio="Иванова Мария", first_name="Оля"),
    ]

    report = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient(contacts), timeline_db=db))
    family = build_family_graph(
        FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True)
    )

    assert report["complete"] is True
    assert report["checked"] == 2
    with sqlite3.connect(db) as con:
        links = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value IN ('child-1','child-2')"
        ).fetchall()
        customer_ids = {row[0] for row in links}
        assert len(customer_ids) == 2, "two siblings must never collapse into one customer"
        phone_links = con.execute(
            "SELECT DISTINCT customer_id, match_class FROM identity_links WHERE link_type='phone' AND link_value=?",
            ("+79163334455",),
        ).fetchall()
        assert {row[0] for row in phone_links} == customer_ids
        assert {row[1] for row in phone_links} == {"ambiguous"}
        family_ids = {
            row[0]
            for row in con.execute(
                "SELECT family_id FROM family_members_v1 WHERE customer_id IN (?,?)",
                tuple(sorted(customer_ids)),
            ).fetchall()
        }
        child_keys = {
            row[0]
            for row in con.execute(
                "SELECT child_key FROM family_links_v1 WHERE customer_id IN (?,?)",
                tuple(sorted(customer_ids)),
            ).fetchall()
        }
        assert len(family_ids) == 1
        assert len(child_keys) == 2
        assert family["family_links_total"] == 2

    payment_payload = {
        "most_finances": [{
            "id": "payment-child-2",
            "contact_id": "child-2",
            "cost": "12000",
            "direction": "in",
            "date_payment": "2026-07-01 09:00:00",
        }],
        "most_abonements": [],
    }
    run_tallanto_payments_import(
        TallantoPaymentsImportConfig(
            source=None,
            timeline_db=db,
            allowed_root=tmp_path,
            tenant_id="foton",
            apply=True,
        ),
        stdin_text=json.dumps(payment_payload),
    )
    contacts_path = tmp_path / "contacts.xlsx"
    contacts_book = Workbook()
    contacts_book.active.append(("ID", "Текстовое значение штрихкода"))
    contacts_book.active.append(("child-2", "barcode-child-2"))
    contacts_book.save(contacts_path)
    attendance_path = tmp_path / "attendance.xlsx"
    attendance_book = Workbook()
    attendance_book.active.append((
        "Фамилия", "Имя", "Штрихкод", "Абонемент", "Сумма списания", "Дата списания",
        "Тип списания", "Занятие", "Филиал занятия", "Дата занятия", "День рождения",
    ))
    attendance_book.active.append((
        "Фамилия", "Оля", "barcode-child-2", "Абонемент", "1000", "01.07.2026 12:00",
        "Безналичный расчёт", "Физика 8 класс", "МФТИ", "01.07.2026 10:00", "",
    ))
    attendance_book.save(attendance_path)
    run_tallanto_attendance_import(
        TallantoAttendanceImportConfig(db, tmp_path, contacts_path, attendance_path, apply=True)
    )
    build_family_graph(FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True))

    with sqlite3.connect(db) as con:
        child_2 = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='child-2'"
        ).fetchone()[0]
        child_2_key = con.execute(
            "SELECT child_key FROM family_links_v1 WHERE customer_id=?", (child_2,)
        ).fetchone()[0]
        attributions = con.execute(
            "SELECT e.event_type,a.customer_id,a.child_key,a.status,a.reason "
            "FROM timeline_events e JOIN event_child_attribution_v1 a ON a.event_id=e.event_id "
            "WHERE e.event_type IN ('tallanto_payment','tallanto_attendance')"
        ).fetchall()
    assert attributions == [
        (event_type, child_2, child_2_key, "matched", "exact_tallanto_identity")
        for event_type in ("tallanto_payment", "tallanto_attendance")
    ]


def test_tallanto_cards_sync_shared_parent_amo_does_not_collapse_two_children(tmp_path: Path) -> None:
    contacts = [
        _contact(contact_id="child-amo-1", phone="+7 916 100-00-01", amo_id="12345", first_name="Аня"),
        _contact(contact_id="child-amo-2", phone="+7 916 100-00-02", amo_id="12345", first_name="Оля"),
    ]

    report = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient(contacts)))

    assert report["complete"] is True
    with sqlite3.connect(tmp_path / "staging.sqlite") as con:
        rows = con.execute(
            "SELECT DISTINCT customer_id FROM identity_links "
            "WHERE link_type='tallanto_student_id' AND link_value IN ('child-amo-1','child-amo-2')"
        ).fetchall()
    assert len({row[0] for row in rows}) == 2


def test_family_graph_adds_second_tallanto_child_after_first_graph_build(tmp_path: Path) -> None:
    db = tmp_path / ".codex_local/staging/customer_timeline.sqlite"
    db.parent.mkdir(parents=True)
    phone = "+7 916 777-66-55"
    first = _contact(contact_id="late-child-1", phone=phone, parent_fio="Петрова Мария", first_name="Аня")
    second = _contact(contact_id="late-child-2", phone=phone, parent_fio="Петрова Мария", first_name="Оля")
    run_tallanto_cards_sync(_config(tmp_path, FakeContactClient([first]), timeline_db=db))
    build_family_graph(FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True))

    run_tallanto_cards_sync(_config(tmp_path, FakeContactClient([first, second]), timeline_db=db))
    build_family_graph(FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True))

    with sqlite3.connect(db) as con:
        children = con.execute(
            "SELECT customer_id,family_id,child_key,canonical_name FROM family_links_v1 ORDER BY canonical_name"
        ).fetchall()
        second_attribution = con.execute(
            "SELECT a.child_key FROM event_child_attribution_v1 a "
            "JOIN timeline_events e ON e.event_id=a.event_id "
            "WHERE e.source_system='tallanto_snapshot' AND e.source_id='late-child-2'"
        ).fetchone()
    assert len(children) == 2
    assert len({row[0] for row in children}) == 2
    assert len({row[1] for row in children}) == 1
    assert len({row[2] for row in children}) == 2
    assert second_attribution == (next(row[2] for row in children if row[3] == "Оля Фамилия"),)


def test_family_graph_tallanto_name_correction_does_not_create_second_child(tmp_path: Path) -> None:
    db = tmp_path / ".codex_local/staging/customer_timeline.sqlite"
    db.parent.mkdir(parents=True)
    phone = "+7 916 777-66-44"
    first = _contact(contact_id="renamed-child", phone=phone, parent_fio="Петрова Мария", first_name="Аня")
    corrected = _contact(
        contact_id="renamed-child",
        phone=phone,
        parent_fio="Петрова Мария",
        first_name="Мария",
        date_modified="2026-07-21 10:00:00",
    )
    run_tallanto_cards_sync(_config(tmp_path, FakeContactClient([first]), timeline_db=db))
    build_family_graph(FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True))
    with sqlite3.connect(db) as con:
        child_key_before = con.execute("SELECT child_key FROM family_links_v1").fetchone()[0]

    run_tallanto_cards_sync(_config(tmp_path, FakeContactClient([corrected]), timeline_db=db))
    build_family_graph(FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True))

    with sqlite3.connect(db) as con:
        children = con.execute(
            "SELECT customer_id,child_key,canonical_name FROM family_links_v1"
        ).fetchall()
    assert len(children) == 1
    assert children[0][1] == child_key_before
    assert children[0][2] == "Мария Фамилия"


def test_tallanto_cards_payments_attendance_build_one_customer_and_family(tmp_path: Path) -> None:
    db = tmp_path / ".codex_local/staging/customer_timeline.sqlite"
    db.parent.mkdir(parents=True)
    cards = _config(
        tmp_path,
        FakeContactClient([_contact(contact_id="101", phone="+7 916 555-44-33", first_name="Аня")]),
        timeline_db=db,
    )
    assert run_tallanto_cards_sync(cards)["validation_ok"] is True
    payment_payload = {
        "most_finances": [{
            "id": "payment-e2e",
            "contact_id": "101",
            "most_abonements_id": "abonement-e2e",
            "cost": "12163",
            "direction": "in",
            "date_payment": "2026-07-01 09:00:00",
        }],
        "most_abonements": [{
            "id": "abonement-e2e",
            "contact_id": "101",
            "date_modified": "2026-07-01 09:00:00",
        }],
    }
    payment_config = TallantoPaymentsImportConfig(
        source=None,
        timeline_db=db,
        allowed_root=tmp_path,
        tenant_id="foton",
        apply=True,
    )
    first_money = run_tallanto_payments_import(payment_config, stdin_text=json.dumps(payment_payload))
    second_money = run_tallanto_payments_import(payment_config, stdin_text=json.dumps(payment_payload))

    contacts_path = tmp_path / "contacts.xlsx"
    contacts_book = Workbook()
    contacts_book.active.append(("ID", "Текстовое значение штрихкода"))
    contacts_book.active.append(("101", "barcode-101"))
    contacts_book.save(contacts_path)
    attendance_path = tmp_path / "attendance.xlsx"
    attendance_book = Workbook()
    attendance_book.active.append((
        "Фамилия", "Имя", "Штрихкод", "Абонемент", "Сумма списания", "Дата списания",
        "Тип списания", "Занятие", "Филиал занятия", "Дата занятия", "День рождения",
    ))
    attendance_book.active.append((
        "Иванова", "Анна", "barcode-101", "Абонемент", "1000", "01.07.2026 12:00",
        "Безналичный расчёт", "Физика 8 класс", "МФТИ", "01.07.2026 10:00", "",
    ))
    attendance_book.save(attendance_path)
    attendance = run_tallanto_attendance_import(
        TallantoAttendanceImportConfig(db, tmp_path, contacts_path, attendance_path, apply=True)
    )
    purchases = refresh_customer_purchases_v1(db, allowed_root=tmp_path, tenant_id="foton")
    family = build_family_graph(FamilyGraphConfig(timeline_db=db, allowed_root=tmp_path, apply=True))

    with sqlite3.connect(db) as con:
        child_id = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='101'"
        ).fetchone()[0]
        event_owners = {
            row[0]
            for row in con.execute(
                "SELECT customer_id FROM timeline_events WHERE event_type IN ('tallanto_payment','tallanto_attendance')"
            ).fetchall()
        }
        fact = con.execute(
            "SELECT total_in,deals_cnt FROM customer_purchases_v1 WHERE customer_id=? AND money_kind='fact'",
            (child_id,),
        ).fetchone()
        family_member = con.execute(
            "SELECT family_id FROM family_members_v1 WHERE customer_id=?",
            (child_id,),
        ).fetchone()
        child = con.execute(
            "SELECT family_id,child_key,canonical_name FROM family_links_v1 WHERE customer_id=?",
            (child_id,),
        ).fetchone()
        attributed_child_keys = {
            row[0]
            for row in con.execute(
                "SELECT child_key FROM event_child_attribution_v1 WHERE customer_id=? AND status='matched'",
                (child_id,),
            ).fetchall()
        }
    assert first_money["validation_ok"] is True
    assert second_money["import_report"]["write_status_counts"].get("created", 0) == 0
    assert attendance["counts"]["created"] == 1
    assert purchases["money_kind"]["fact"] == 1
    assert event_owners == {child_id}
    assert fact == (12163.0, 1)
    assert family_member is not None
    assert family["family_members_total"] == 1
    assert family["family_links_total"] == 1
    assert child[0] == family_member[0]
    assert child[1] in attributed_child_keys
    assert child[2] == "Аня Фамилия"


def test_tallanto_cards_sync_phone_change_under_same_student_id_updates_same_child(tmp_path: Path) -> None:
    old_phone = "+7 916 444-55-66"
    new_phone = "+7 916 444-55-77"
    config = _config(tmp_path, FakeContactClient([_contact(contact_id="20", phone=old_phone)]))
    first = run_tallanto_cards_sync(config)
    assert first["complete"] is True
    with sqlite3.connect(tmp_path / "staging.sqlite") as con:
        first_customer_id = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='20'"
        ).fetchone()[0]

    second_config = _config(tmp_path, FakeContactClient([_contact(contact_id="20", phone=new_phone, date_modified="2026-07-21 10:00:00")]))
    second = run_tallanto_cards_sync(second_config)

    assert second["complete"] is True
    assert second["conflict"] == 0
    with sqlite3.connect(tmp_path / "staging.sqlite") as con:
        second_customer_id = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='20'"
        ).fetchone()[0]
        assert second_customer_id == first_customer_id, "same tallanto_student_id must update the same child, not create a new one"
        new_phone_owner = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='phone' AND link_value='+79164445577'"
        ).fetchone()[0]
        assert new_phone_owner == first_customer_id
        customer_count = con.execute("SELECT COUNT(*) FROM customer_identities").fetchone()[0]
        assert customer_count == 1


def test_tallanto_cards_sync_phone_shared_by_two_different_families_is_conflict_not_merge(tmp_path: Path) -> None:
    db_path = tmp_path / "staging.sqlite"
    other_phone = "+79169998877"
    with CustomerTimelineSQLiteStore(db_path, allowed_root=tmp_path) as store:
        other = CustomerIdentity(
            tenant_id="foton",
            identity_status=IdentityStatus.STRONG,
            display_name="Другая семья",
            primary_phone=other_phone,
        )
        store.upsert_customer(other)
        store.upsert_identity_link(
            IdentityLink(
                tenant_id="foton",
                customer_id=other.customer_id,
                link_type="phone",
                link_value=other_phone,
                source_system="amocrm_snapshot",
                source_ref="amo:other",
                confidence=0.95,
            )
        )
        other_customer_id = other.customer_id

    first_config = _config(tmp_path, FakeContactClient([_contact(contact_id="conflict-1", phone="+7 916 111-00-00")]))
    first = run_tallanto_cards_sync(first_config)
    assert first["complete"] is True
    with sqlite3.connect(db_path) as con:
        tallanto_customer_id = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='conflict-1'"
        ).fetchone()[0]
    assert tallanto_customer_id != other_customer_id

    second_config = _config(
        tmp_path,
        FakeContactClient([_contact(contact_id="conflict-1", phone="+7 916 999-88-77", date_modified="2026-07-21 10:00:00")]),
    )
    second = run_tallanto_cards_sync(second_config)

    assert second["complete"] is True
    assert second["conflict"] >= 1
    with sqlite3.connect(db_path) as con:
        # Never a merge: the tallanto_student_id link must stay with its
        # original customer, not jump to the other family that already
        # holds that phone.
        holder = con.execute(
            "SELECT customer_id FROM identity_links WHERE link_type='tallanto_student_id' AND link_value='conflict-1'"
        ).fetchone()[0]
        assert holder == tallanto_customer_id
        customer_count = con.execute("SELECT COUNT(*) FROM customer_identities").fetchone()[0]
        assert customer_count == 2, "conflict must never silently merge the two families into one customer"
        conflict_types = {
            row[0] for row in con.execute("SELECT conflict_type FROM timeline_conflicts").fetchall()
        }
        assert conflict_types, "a conflict row must be recorded, not silently dropped"


def test_tallanto_cards_sync_api_unavailable_fails_source_and_publishes_nothing(tmp_path: Path) -> None:
    report = run_tallanto_cards_sync(_config(tmp_path, FailingContactClient()))

    assert report["validation_ok"] is False
    assert report["source_failed"] is True
    assert report["checked"] == 0
    assert report["cursor_time"] is None
    assert report["safety"]["staging_db_write"] is False
    assert not (tmp_path / "staging.sqlite").exists() or _event_count(tmp_path / "staging.sqlite") == 0
    assert load_tallanto_cards_checkpoint(tmp_path / "out") == {}


def test_tallanto_cards_sync_empty_first_run_is_not_fresh_success(tmp_path: Path) -> None:
    report = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient([])))

    assert report["validation_ok"] is False
    assert report["complete"] is False
    assert report["blocked_reason"] == "no_tallanto_cards"
    assert report["checked"] == 0
    assert not (tmp_path / "staging.sqlite").exists()


def test_tallanto_cards_sync_bounded_checkpoint_completes_large_contact_list_across_runs(tmp_path: Path) -> None:
    contacts = [_contact(contact_id=str(index), phone=f"+7916{index:07d}") for index in range(1, 251)]
    client = FakeContactClient(contacts, page_size=50)
    config = _config(tmp_path, client, max_pages=2)  # 250/50=5 pages, 2 pages/run -> 3 runs

    reports = []
    for _ in range(6):
        report = run_tallanto_cards_sync(config)
        reports.append(report)
        if report.get("complete"):
            break
        assert _event_count(tmp_path / "staging.sqlite") == 0 if (tmp_path / "staging.sqlite").exists() else True
        assert report["validation_ok"] is False
        assert report["apply_blocked"] is True
    else:
        pytest.fail("checkpoint cycle did not complete within the safety bound of 6 runs")

    assert len(reports) == 3
    final = reports[-1]
    assert final["checked"] == 250
    assert final["total_pages_read"] == 5
    assert final["checkpoint"]["cleared"] is True
    assert load_tallanto_cards_checkpoint(tmp_path / "out") == {}
    assert client.offsets_called == [0, 50, 50, 100, 150, 150, 200]


def test_tallanto_cards_sync_restarts_when_resume_boundary_changes(tmp_path: Path) -> None:
    contacts = [_contact(contact_id=str(index), phone=f"+7916{index:07d}") for index in range(1, 151)]
    client = FakeContactClient(contacts, page_size=50)
    config = _config(tmp_path, client, max_pages=1)

    first = run_tallanto_cards_sync(config)
    assert first["blocked_reason"] == "page_cap_hit"
    assert first["checked"] == 50

    client.contacts.insert(0, _contact(contact_id="0", phone="+79160000000"))
    second = run_tallanto_cards_sync(config)

    assert second["blocked_reason"] == "page_cap_hit"
    assert second["checked"] == 50
    assert second["checkpoint"]["checkpoint_reset_reason"] == "pagination_universe_changed"
    assert client.offsets_called == [0, 0, 0]
    assert not (tmp_path / "staging.sqlite").exists()


def test_tallanto_cards_sync_restarts_when_total_changes_after_saved_page(tmp_path: Path) -> None:
    contacts = [_contact(contact_id=str(index), phone=f"+7916{index:07d}") for index in range(1, 151)]
    client = FakeContactClient(contacts, page_size=50)
    config = _config(tmp_path, client, max_pages=1)

    first = run_tallanto_cards_sync(config)
    assert first["blocked_reason"] == "page_cap_hit"

    client.contacts.append(_contact(contact_id="151", phone="+79160000151"))
    second = run_tallanto_cards_sync(config)

    assert second["blocked_reason"] == "page_cap_hit"
    assert second["checkpoint"]["checkpoint_reset_reason"] == "pagination_universe_changed"
    assert client.offsets_called == [0, 0, 0]


def test_tallanto_cards_sync_blocks_duplicate_contact_ids(tmp_path: Path) -> None:
    contacts = [_contact(contact_id="same", phone="+79160000001"), _contact(contact_id="same", phone="+79160000002")]

    report = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient(contacts)))

    assert report["validation_ok"] is False
    assert report["complete"] is False
    assert report["blocked_reason"] == "pagination_universe_changed"
    assert not (tmp_path / "staging.sqlite").exists()


def test_universe_fingerprint_changes_when_select_fields_change() -> None:
    a = universe_fingerprint(select_fields=("id", "first_name"), order_by="id ASC")
    b = universe_fingerprint(select_fields=("id", "last_name"), order_by="id ASC")
    c = universe_fingerprint(select_fields=("id", "first_name"), order_by="id ASC")
    assert a != b
    assert a == c


def test_map_raw_contact_without_id_returns_none_never_fuzzy_matches() -> None:
    assert map_raw_contact_to_snapshot_payload({"first_name": "Аноним", "phone_mobile": "+79160000000"}, snapshot_at="2026-07-26T00:00:00+00:00") is None


def test_map_raw_contact_keeps_confirmed_business_fields_only() -> None:
    mapped = map_raw_contact_to_snapshot_payload(
        {
            "id": "student-1",
            "first_name": "Иван",
            "last_name": "Иванов",
            "phone_mobile": "+7 916 000-00-01",
            "phone_work": "+7 916 000-00-02",
            "phone_home": "+7 916 000-00-03",
            "email1": "parent@example.com",
            "email2": "student@example.com",
            "email": "unsupported@example.com",
            "marital_status_c": "Иванова Анна",
            "type_client_c": "10 класс",
            "filial": "Долгопрудный",
            "subject1_name": "Физика",
            "subject2_name": "Математика",
            "interests_c": "Олимпиада",
            "source": "Рекомендация",
            "amo_id": "123",
            "barcode": "ABC",
            "assigned_user_name": "Менеджер",
            "date_entered": "2026-01-01 10:00:00",
            "date_modified": "2026-07-25 10:00:00",
        },
        snapshot_at="2026-07-26T00:00:00+00:00",
    )

    assert mapped is not None
    assert mapped["primary_phone"] == "+7 916 000-00-01"
    assert mapped["phone_extra"] == "+7 916 000-00-02"
    assert "+7 916 000-00-03" not in str(mapped)
    assert mapped["primary_email"] == "parent@example.com"
    assert mapped["email_extra"] == "student@example.com"
    assert "unsupported@example.com" not in str(mapped)
    assert mapped["parent_fio"] == "Иванова Анна"
    assert mapped["student_type"] == "10 класс"
    assert mapped["branch"] == "Долгопрудный"
    assert mapped["subjects"] == "Физика, Математика"
    assert mapped["amo_contact_id"] == "123"


def test_tallanto_cards_sync_uses_only_explicit_env_file(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    env_path = tmp_path / "tallanto.env"
    client = FakeContactClient([_contact(contact_id="1")])
    seen: list[Path] = []

    def fake_build(path: Path) -> FakeContactClient:
        seen.append(path)
        return client

    monkeypatch.setattr(cards_sync, "_build_tallanto_client", fake_build)
    config = TallantoCardsSyncConfig(
        timeline_db=tmp_path / "staging.sqlite",
        out_root=tmp_path / "out",
        allowed_root=tmp_path,
        tallanto_env_file=env_path,
    )

    report = run_tallanto_cards_sync(config)

    assert report["validation_ok"] is True
    assert seen == [env_path]


def test_tallanto_cards_sync_private_files_and_checkpoint_survive_import_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    config = _config(tmp_path, FakeContactClient([_contact(contact_id="1")]))

    def fail_import(_config: Any) -> Mapping[str, Any]:
        raise RuntimeError("synthetic import failure")

    monkeypatch.setattr(cards_sync, "run_nightly_incremental", fail_import)
    with pytest.raises(RuntimeError, match="synthetic import failure"):
        run_tallanto_cards_sync(config)

    checkpoint = tmp_path / "out/tallanto_cards_sync_checkpoint.json"
    source = tmp_path / "out/tallanto_cards_sources/tallanto_contacts_daily.jsonl"
    assert checkpoint.exists()
    assert source.exists()
    assert checkpoint.stat().st_mode & 0o777 == 0o600
    assert source.stat().st_mode & 0o777 == 0o600


def test_tallanto_cards_sync_failed_import_report_keeps_checkpoint(tmp_path: Path, monkeypatch) -> None:
    config = _config(tmp_path, FakeContactClient([_contact(contact_id="1")]))
    monkeypatch.setattr(cards_sync, "run_nightly_incremental", lambda _config: {"gate_passed": False})

    report = run_tallanto_cards_sync(config)

    assert report["validation_ok"] is False
    assert report["complete"] is False
    assert report["checkpoint"]["cleared"] is False
    assert load_tallanto_cards_checkpoint(tmp_path / "out")


def test_tallanto_cards_sync_skips_id_less_contact_and_counts_unmatched(tmp_path: Path) -> None:
    contacts = [
        _contact(contact_id="30", phone="+7 916 555-66-77"),
        {"first_name": "Без ID", "phone_mobile": "+7 916 000-11-22"},
    ]

    report = run_tallanto_cards_sync(_config(tmp_path, FakeContactClient(contacts)))

    assert report["checked"] == 2
    assert report["checked_with_id"] == 1
    assert report["skipped_missing_id"] == 1
    assert report["unmatched"] == 1
    assert report["validation_ok"] is False
    assert report["apply_blocked"] is True
    assert report["blocked_reason"] == "contacts_missing_stable_id"
    assert not (tmp_path / "out/tallanto_cards_sources/tallanto_contacts_daily.jsonl").exists()
    assert load_tallanto_cards_checkpoint(tmp_path / "out")
