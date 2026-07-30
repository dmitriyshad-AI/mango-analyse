#!/usr/bin/env python3
r"""Deterministic, stratified sampler for the Block D semantic review.

Built for `2026-07-26_TZ_CLAUDE_dovesti_do_AI_sotrudnika_bez_pauz.md` (Block D: "досье и
локальные черновики"). The audit sandbox that wrote this script has no mount for
`~/.mango_local/...`, so the 30 real family dossiers and the 50 real Wappi drafts required by
that TZ could not be produced there. This script is the owner-gated harness that reproduces
them on the machine where staging *is* mounted, with a fixed seed so the selection is
reproducible and auditable.

Design choice (ponytail / minimum code): this script does NOT reimplement dossier building,
Owner50 classification, or draft generation. It only (a) enumerates the eligible population with
the attributes the TZ asks to stratify on, (b) samples deterministically, and (c) calls the
existing, already-tested production code:
  - `mango_mvp.customer_timeline.manager_dossier.build_manager_dossier_workbook` for dossiers
    (this is the same function customer-facing manager tooling uses; it already enforces the
    36h freshness gate via `manager_freshness_gate` and already guards PII output paths to
    `.codex_local/`).
  - `mango_mvp.customer_timeline.manager_dossier.build_owner50_family_workbook` for the owner50
    mode (Block F, 26.07) -- same function the real Owner50 queue uses: read-only DB connection
    (`_connect_ro`, `PRAGMA query_only = ON`), same freshness gate, same `.codex_local/` output
    guard. It never writes AMO notes and never sends anything to a client -- it only classifies
    already-ingested rows and writes one local XLSX.
  - `scripts/run_amo_wappi_draft_loop.py` (existing, unmodified, dry-run by default) for drafts;
    this script only *selects* 50 rows out of that run's journal.jsonl, it does not talk to
    Wappi/AMO/an LLM itself and makes zero HTTP calls.

Usage (owner, on the Mac, staging mounted)
-------------------------------------------
    cd /Users/dmitrijfabarisov/Projects/Mango\ analyse

    # 0) owner50: ALL READY families + deterministically 15 stratified CANDIDATE + 15 stratified
    #    EXCLUDED (stratified by reason code -- "Код причины" -- so the sample covers as many
    #    distinct CANDIDATE/EXCLUDED reasons as possible, not just the most common one). Same
    #    freshness-gate-blocked behaviour as dossiers below (exit 2,
    #    semantic_review_blocked_by_freshness, no synthetic substitute ever written as if real).
    PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=src python3 \
        scripts/audit_owner_gate_semantic_sample.py owner50 \
        --db ~/.mango_local/customer_timeline_nightly/.codex_local/staging/customer_timeline_staging.sqlite \
        --out-root .codex_local/semantic_sample_$(date +%Y%m%d) \
        --seed 20260726 --candidate-count 15 --excluded-count 15

    # 1) 30 stratified family dossiers (read-only; exits 2 and writes a
    #    semantic_review_blocked_by_freshness.json instead of dossiers if the staging
    #    snapshot fails the 36h freshness gate -- this is expected/correct behaviour, not a bug)
    PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=src python3 \
        scripts/audit_owner_gate_semantic_sample.py dossiers \
        --db ~/.mango_local/customer_timeline_nightly/.codex_local/staging/customer_timeline_staging.sqlite \
        --out-root .codex_local/semantic_sample_$(date +%Y%m%d) \
        --seed 20260726 --count 30

    # 2) Full personal-chat corpus dry run with the existing, unmodified script (default is
    #    dry-run / no --live-write, so this makes zero AMO POST/PATCH and zero client sends).
    #    This also satisfies the D2 "every personal inbound gets a terminal outcome" check.
    PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=src python3 \
        scripts/run_amo_wappi_draft_loop.py --once --chat-limit 0

    # 3) 50 stratified blind drafts sampled from that run's journal (no new LLM calls -- reuses
    #    the dry-run drafts step 2 already produced).
    PYTHONDONTWRITEBYTECODE=1 PYTHONPATH=src python3 \
        scripts/audit_owner_gate_semantic_sample.py drafts \
        --journal ~/.mango_local/draft_loop/journal.jsonl \
        --out-root .codex_local/semantic_sample_$(date +%Y%m%d) \
        --seed 20260726 --count 50

Every real (PII-bearing) artifact is written under `<out-root>/.codex_local/...`, which
`build_manager_dossier_workbook`/`build_owner50_family_workbook` refuse to bypass, and which must
never be copied into git or an audit pack. A scrubbed, PII-free selection manifest (ids + strata +
verdict, no names, no phones, no draft text) is written next to it for inclusion in the audit
pack. No subcommand in this script ever writes an AMO note or sends anything to a client -- all
three only read already-ingested data and write local files.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import random
import sqlite3
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "src"))

from mango_mvp.customer_timeline.manager_dossier import (  # noqa: E402
    OWNER50_CONTROL_COLUMNS,
    _family_scope_customer_ids,
    _guard_local_dossier_output_path,
    _owner50_family_rows,
    _owner50_control_row_from_ready,
    _source_freshness,
    build_manager_dossier_workbook,
    build_owner50_family_workbook,
    build_customer_dossier,
    manager_freshness_gate,
    _connect_ro,
)
from mango_mvp.customer_timeline.store import customer_entity_ref_values  # noqa: E402


def _short_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:12]


def _attendance_is_not_future(value: Any, *, now: datetime | None = None) -> bool:
    try:
        parsed = datetime.fromisoformat(str(value or "").replace("Z", "+00:00"))
    except ValueError:
        return False
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed <= (now or datetime.now(timezone.utc))


def _write_private_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
    path.chmod(0o600)


# ---------------------------------------------------------------------------
# Stratified deterministic sampling (shared helper)
# ---------------------------------------------------------------------------


def stratified_sample(
    population: list[Mapping[str, Any]],
    *,
    count: int,
    seed: int,
    strata_key: Any,
) -> list[Mapping[str, Any]]:
    """Pick `count` rows from `population`, covering as many distinct strata as possible first
    (round-robin across strata in a seed-stable shuffled order), then filling the remainder
    deterministically. Same input + same seed -> same output, always.
    """
    rng = random.Random(seed)
    buckets: dict[Any, list[Mapping[str, Any]]] = defaultdict(list)
    for row in population:
        buckets[strata_key(row)].append(row)
    for rows in buckets.values():
        rows.sort(key=lambda r: str(r.get("id", "")))
        rng.shuffle(rows)
    bucket_keys = sorted(buckets.keys(), key=lambda k: str(k))
    rng.shuffle(bucket_keys)

    picked: list[Mapping[str, Any]] = []
    picked_ids: set[str] = set()
    # Round 1: one from every stratum, to guarantee coverage.
    for key in bucket_keys:
        if len(picked) >= count:
            break
        for row in buckets[key]:
            if row.get("id") not in picked_ids:
                picked.append(row)
                picked_ids.add(row.get("id"))
                break
    # Round 2+: fill remainder round-robin across strata (still seed-stable).
    idx = 0
    offsets = {key: 1 for key in bucket_keys}
    while len(picked) < count:
        progressed = False
        for key in bucket_keys:
            if len(picked) >= count:
                break
            rows = buckets[key]
            off = offsets[key]
            if off >= len(rows):
                continue
            row = rows[off]
            offsets[key] += 1
            if row.get("id") not in picked_ids:
                picked.append(row)
                picked_ids.add(row.get("id"))
                progressed = True
        if not progressed:
            break
        idx += 1
    return picked[:count]


# ---------------------------------------------------------------------------
# Subcommand: owner50 (Block F, 26.07)
# ---------------------------------------------------------------------------


def _owner50_sheet_rows(out_xlsx: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Reads one owner50 XLSX sheet back as dicts keyed by its own header row (READY_50 uses
    OWNER50_REQUIRED_COLUMNS, CANDIDATES/EXCLUDED use OWNER50_CONTROL_COLUMNS -- both carry a
    literal "family_id" column, used here as the sampling id). Collapses to ONE row per
    family_id (a family can have several CANDIDATE/EXCLUDED rows, one per simultaneous reason
    code -- keeping only the first is enough for sampling *families*, not *reasons*; the reason
    breakdown below still reports which reason each SAMPLED family's first row carries)."""
    from openpyxl import load_workbook

    wb = load_workbook(out_xlsx, read_only=True)
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    seen: set[str] = set()
    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        record = dict(zip(header, values))
        family_id = str(record.get("family_id") or "")
        if not family_id or family_id in seen:
            continue
        seen.add(family_id)
        record["id"] = family_id
        rows.append(record)
    return rows


def cmd_owner50(args: argparse.Namespace) -> int:
    db = Path(args.db).expanduser()
    out_root = Path(args.out_root).expanduser()
    local_dir = out_root / ".codex_local"
    local_dir.mkdir(parents=True, exist_ok=True)
    local_dir.chmod(0o700)
    run_token = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_xlsx = local_dir / f"owner50_sample_{run_token}.xlsx"

    scrubbed_manifest: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "db": str(db),
        "tenant_id": args.tenant_id,
        "seed": args.seed,
        "candidate_sample_size_requested": args.candidate_count,
        "excluded_sample_size_requested": args.excluded_count,
        "current_artifact": None,
        "legacy_artifact_present": (local_dir / "owner50_sample.xlsx").exists(),
    }

    try:
        summary = build_owner50_family_workbook(
            timeline_db=db,
            allowed_root=out_root,
            out_xlsx=out_xlsx,
            tenant_id=args.tenant_id,
            limit=50,
            enforce_freshness=True,
        )
    except RuntimeError as exc:
        # требование F (26.07): свежая staging недоступна -> harness всё равно завершается
        # (не падает, не зависает) и честно помечает semantic_review_blocked_by_freshness --
        # синтетика НЕ подменяет реальную приёмку. Тот же паттерн, что и у cmd_dossiers ниже.
        scrubbed_manifest["status"] = "semantic_review_blocked_by_freshness"
        scrubbed_manifest["freshness_gate_error"] = str(exc)
        (out_root / "owner50_selection_manifest.json").write_text(
            json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        print("semantic_review_blocked_by_freshness")
        print(str(exc))
        print(f"Scrubbed selection manifest written (no PII): {out_root / 'owner50_selection_manifest.json'}")
        return 2

    ready_rows = _owner50_sheet_rows(out_xlsx, "READY_50")
    candidate_rows = _owner50_sheet_rows(out_xlsx, "CANDIDATES")
    excluded_rows = _owner50_sheet_rows(out_xlsx, "EXCLUDED")

    def reason_key(row: Mapping[str, Any]) -> tuple:
        return (str(row.get("Код причины") or ""),)

    if not summary.get("ready_audit_population_complete"):
        scrubbed_manifest.update(
            {
                "status": "semantic_review_blocked_by_ready_limit",
                "ready_population": int(summary.get("ready_total") or 0),
                "ready_sampled": len(ready_rows),
                "writes_amo_notes": False,
                "sends_to_client": False,
            }
        )
        manifest_path = out_root / "owner50_selection_manifest.json"
        manifest_path.write_text(
            json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        print(
            "semantic_review_blocked_by_ready_limit: "
            f"workbook contains {len(ready_rows)} of {summary.get('ready_total')} READY families"
        )
        return 3

    # requirement F: all READY plus deterministic CANDIDATE/EXCLUDED samples.
    candidate_sample = stratified_sample(
        candidate_rows, count=args.candidate_count, seed=args.seed, strata_key=reason_key,
    )
    excluded_sample = stratified_sample(
        excluded_rows, count=args.excluded_count, seed=args.seed, strata_key=reason_key,
    )

    scrubbed_manifest["status"] = "sample_built"
    scrubbed_manifest["build_summary"] = {k: v for k, v in summary.items() if k != "freshness_rows"}
    scrubbed_manifest["ready_population"] = len(ready_rows)
    scrubbed_manifest["ready_sampled"] = len(ready_rows)  # все READY, не подвыборка
    scrubbed_manifest["candidate_population"] = len(candidate_rows)
    scrubbed_manifest["candidate_sampled"] = len(candidate_sample)
    scrubbed_manifest["excluded_population"] = len(excluded_rows)
    scrubbed_manifest["excluded_sampled"] = len(excluded_sample)
    scrubbed_manifest["candidate_reason_breakdown"] = dict(
        sorted(Counter(str(row.get("Код причины") or "") for row in candidate_sample).items())
    )
    scrubbed_manifest["excluded_reason_breakdown"] = dict(
        sorted(Counter(str(row.get("Код причины") or "") for row in excluded_sample).items())
    )
    scrubbed_manifest["ready_family_id_hashes"] = [_short_hash(row["id"]) for row in ready_rows]
    scrubbed_manifest["candidate_family_id_hashes"] = [_short_hash(row["id"]) for row in candidate_sample]
    scrubbed_manifest["excluded_family_id_hashes"] = [_short_hash(row["id"]) for row in excluded_sample]
    # ⛔ требование E/F: harness никогда не пишет AMO notes и не отправляет клиенту --
    # build_owner50_family_workbook сам read-only (_connect_ro, PRAGMA query_only=ON) и пишет
    # только один локальный XLSX; здесь ниже -- только чтение того же XLSX и JSON-манифесты.
    scrubbed_manifest["writes_amo_notes"] = False
    scrubbed_manifest["sends_to_client"] = False
    scrubbed_manifest["current_artifact"] = f".codex_local/{out_xlsx.name}"
    (out_root / "owner50_selection_manifest.json").write_text(
        json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    # PII-bearing family_id lists for the owner/manager's own local semantic review only (raw
    # names/phones/offers stay in out_xlsx itself, already under .codex_local/; this file adds
    # only which family_ids were sampled, still PII by itself -- also local only, never git).
    _write_private_text(
        local_dir / "owner50_sample_family_ids.json",
        json.dumps(
            {
                "ready_family_ids": [row["id"] for row in ready_rows],
                "candidate_family_ids": [row["id"] for row in candidate_sample],
                "excluded_family_ids": [row["id"] for row in excluded_sample],
            },
            ensure_ascii=False, indent=2, sort_keys=True,
        ) + "\n",
    )

    print(
        f"OK: {len(ready_rows)} READY (all) + {len(candidate_sample)}/{len(candidate_rows)} CANDIDATE + "
        f"{len(excluded_sample)}/{len(excluded_rows)} EXCLUDED written to {out_xlsx} "
        "(PII, local only, not for git)."
    )
    print(f"Scrubbed selection manifest (no PII): {out_root / 'owner50_selection_manifest.json'}")
    return 0


# ---------------------------------------------------------------------------
# Subcommand: dossiers
# ---------------------------------------------------------------------------


CURRENT_TALLANTO_STUDENT_TYPES = frozenset(
    {"listener", "слушатель"}
    | {f"{grade}_klass" for grade in range(1, 11)}
    | {f"{grade} класс" for grade in range(1, 11)}
)


def _current_tallanto_business_family_ids(con: sqlite3.Connection, *, tenant_id: str) -> set[str]:
    placeholders = ",".join("?" for _ in CURRENT_TALLANTO_STUDENT_TYPES)
    params = (tenant_id, *sorted(CURRENT_TALLANTO_STUDENT_TYPES))
    rows = con.execute(
        f"""
        WITH current_students AS (
          SELECT DISTINCT customer_id
          FROM timeline_events
          WHERE tenant_id=? AND source_system='tallanto_snapshot'
            AND event_type='tallanto_student_snapshot' AND superseded_by IS NULL
            AND customer_id IS NOT NULL
            AND lower(trim(json_extract(record_json,'$.record.payload.student_type'))) IN ({placeholders})
        )
        SELECT member.family_id
        FROM family_members_v1 AS member JOIN current_students USING (customer_id)
        WHERE member.tenant_id=?
        UNION
        SELECT child.family_id
        FROM family_links_v1 AS child JOIN current_students USING (customer_id)
        WHERE child.tenant_id=?
        """,
        (*params, tenant_id, tenant_id),
    ).fetchall()
    return {str(row[0]) for row in rows if row[0]}


def _dossier_population(con: sqlite3.Connection, *, tenant_id: str) -> list[dict[str, Any]]:
    con.row_factory = sqlite3.Row
    current_family_ids = _current_tallanto_business_family_ids(con, tenant_id=tenant_id)
    if not current_family_ids:
        return []
    member_rows = con.execute(
        """
        SELECT family_id, customer_id, brand, child_key
        FROM family_links_v1
        WHERE tenant_id = ? AND status != 'excluded'
        """,
        (tenant_id,),
    ).fetchall()
    link_members: dict[str, set[str]] = defaultdict(set)
    family_brands: dict[str, set[str]] = defaultdict(set)
    family_children: dict[str, set[str]] = defaultdict(set)
    for row in member_rows:
        family_id = str(row["family_id"])
        link_members[family_id].add(str(row["customer_id"]))
        if str(row["brand"] or "").strip():
            family_brands[family_id].add(str(row["brand"]).strip().casefold())
        family_children[family_id].add(str(row["child_key"]))
    family_members: dict[str, set[str]] = {}
    if con.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name='family_members_v1'"
    ).fetchone():
        for row in con.execute(
            "SELECT family_id,customer_id FROM family_members_v1 WHERE tenant_id=? "
            "AND membership_status IN ('confident','singleton')",
            (tenant_id,),
        ):
            family_members.setdefault(str(row["family_id"]), set()).add(str(row["customer_id"]))
    for family_id, members in link_members.items():
        family_members.setdefault(family_id, set()).update(members)
    if not family_members:
        return []
    ids = sorted({customer_id for members in family_members.values() for customer_id in members})
    placeholders = ",".join("?" for _ in ids)

    paying = {
        row["customer_id"]
        for row in con.execute(
            f"""
            SELECT DISTINCT customer_id FROM customer_purchases_v1
            WHERE tenant_id = ? AND customer_id IN ({placeholders})
              AND (total_in > 0 OR deals_cnt > 0)
            """,
            (tenant_id, *ids),
        ).fetchall()
    }
    signalled = {
        row["customer_id"]
        for row in con.execute(
            f"""
            SELECT DISTINCT customer_id FROM derived_signals
            WHERE tenant_id = ? AND customer_id IN ({placeholders})
            """,
            (tenant_id, *ids),
        ).fetchall()
    }
    channel_rows = con.execute(
        f"""
        SELECT DISTINCT customer_id, source_system FROM timeline_events
        WHERE tenant_id = ? AND customer_id IN ({placeholders})
          AND source_system IN ('wappi_telegram', 'wappi_max')
        """,
        (tenant_id, *ids),
    ).fetchall()
    channel_by_customer: dict[str, set[str]] = defaultdict(set)
    for row in channel_rows:
        channel_by_customer[row["customer_id"]].add(row["source_system"])

    event_layers: dict[str, set[str]] = defaultdict(set)
    as_of = datetime.now(timezone.utc).isoformat()
    for row in con.execute(
        f"SELECT DISTINCT customer_id, event_type FROM timeline_events WHERE tenant_id=? "
        f"AND customer_id IN ({placeholders}) AND event_type IN ('email_message','mango_call','tallanto_attendance') "
        "AND (event_type != 'tallanto_attendance' OR (event_at <= ? AND ("
        "source_system != 'tallanto_attendance_api' "
        "OR json_extract(record_json, '$.record.attendance_confirmed') = 1)))",
        (tenant_id, *ids, as_of),
    ).fetchall():
        event_layers[row["customer_id"]].add(row["event_type"])

    conflicted: set[str] = set()
    conflicted_families: set[str] = set()
    customer_by_ref = {
        ref: customer_id
        for customer_id in ids
        for ref in customer_entity_ref_values(customer_id)
    }
    for row in con.execute(
        "SELECT record_json FROM timeline_conflicts WHERE tenant_id = ? AND status != 'resolved'",
        (tenant_id,),
    ).fetchall():
        # entity_refs -- тот же точный признак, что уже используется в _next_step_for_dossier
        # (manager_dossier.py:3070-3083) для увязки конфликта с customer_id. НЕ substring: id
        # одного клиента может быть текстовым префиксом другого ("customer:1" внутри
        # "customer:10") -- substring даёт ложные совпадения, entity_refs -- нет.
        refs = {str(ref) for ref in (json.loads(row["record_json"] or "{}").get("entity_refs") or ())}
        conflicted.update(customer_by_ref[ref] for ref in refs if ref in customer_by_ref)
        for family_id in family_members:
            if refs & {family_id, f"family:{family_id}"}:
                conflicted_families.add(family_id)

    population = []
    for family_id, members in sorted(family_members.items()):
        if family_id not in current_family_ids:
            continue
        cid = min(members)
        channels = {channel for member in members for channel in channel_by_customer.get(member, set())}
        layers = {layer for member in members for layer in event_layers.get(member, set())}
        brands = family_brands.get(family_id, set())
        population.append(
            {
                "id": cid,
                "family_id": family_id,
                "brand": next(iter(brands)) if len(brands) == 1 else "unknown",
                "child_bucket": (
                    "0" if not family_children[family_id]
                    else ("1" if len(family_children[family_id]) == 1 else "2+")
                ),
                "channel": (
                    "both" if len(channels) > 1 else (next(iter(channels)) if channels else "none")
                ),
                "has_payment": bool(members & paying),
                "has_conflict": bool(members & conflicted) or family_id in conflicted_families,
                "has_signal": bool(members & signalled),
                "has_mail": "email_message" in layers,
                "has_call": "mango_call" in layers,
                "has_attendance": "tallanto_attendance" in layers,
            }
        )
    return population


def cmd_dossiers(args: argparse.Namespace) -> int:
    db = Path(args.db).expanduser()
    out_root = Path(args.out_root).expanduser()
    local_dir = out_root / ".codex_local"
    local_dir.mkdir(parents=True, exist_ok=True)
    local_dir.chmod(0o700)
    run_token = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_xlsx = local_dir / f"dossiers_sample_{run_token}.xlsx"

    with _connect_ro(db) as con:
        population = _dossier_population(con, tenant_id=args.tenant_id)

    if not population:
        print(f"BLOCKED: no eligible families found in {db} (empty family_links_v1?).")
        return 3

    def strata_key(row: Mapping[str, Any]) -> tuple:
        return (row["brand"], row["channel"], row["child_bucket"], row["has_payment"], row["has_conflict"],
                row["has_signal"], row["has_mail"], row["has_call"], row["has_attendance"])

    sample = stratified_sample(population, count=args.count, seed=args.seed, strata_key=strata_key)
    strata_counter = Counter(strata_key(row) for row in sample)

    scrubbed_manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "db": str(db),
        "tenant_id": args.tenant_id,
        "seed": args.seed,
        "requested_count": args.count,
        "population_size": len(population),
        "sample_size": len(sample),
        "current_artifact": None,
        "legacy_artifact_present": (local_dir / "dossiers_sample.xlsx").exists(),
        "distinct_strata_covered": len(strata_counter),
        "strata_breakdown": [
            {
                "brand": key[0], "channel": key[1], "child_bucket": key[2],
                "has_payment": key[3], "has_conflict": key[4], "has_signal": key[5],
                "has_mail": key[6], "has_call": key[7], "has_attendance": key[8],
                "count": n,
            }
            for key, n in sorted(strata_counter.items(), key=lambda kv: str(kv[0]))
        ],
        "sample_customer_id_hashes": [_short_hash(row["id"]) for row in sample],
    }

    try:
        summary = build_manager_dossier_workbook(
            timeline_db=db,
            allowed_root=out_root,
            out_xlsx=out_xlsx,
            tenant_id=args.tenant_id,
            customer_ids=[row["id"] for row in sample],
            limit=len(sample),
            enforce_freshness=True,
        )
    except RuntimeError as exc:
        scrubbed_manifest["status"] = "semantic_review_blocked_by_freshness"
        scrubbed_manifest["freshness_gate_error"] = str(exc)
        (out_root / "dossiers_selection_manifest.json").write_text(
            json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
            encoding="utf-8",
        )
        print("semantic_review_blocked_by_freshness")
        print(str(exc))
        print(f"Scrubbed selection manifest written (no PII): {out_root / 'dossiers_selection_manifest.json'}")
        return 2

    scrubbed_manifest["status"] = "sample_built"
    scrubbed_manifest["current_artifact"] = f".codex_local/{out_xlsx.name}"
    scrubbed_manifest["dossier_summary"] = {
        k: v for k, v in summary.items() if k not in ("actuality_header", "source_freshness_top")
    }
    (out_root / "dossiers_selection_manifest.json").write_text(
        json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(f"OK: {len(sample)} dossiers written to {out_xlsx} (PII, local only, not for git).")
    print(f"Scrubbed selection manifest (no PII): {out_root / 'dossiers_selection_manifest.json'}")
    return 0


_ACCEPTANCE_SHEETS = ("Семьи 30", "Хронология", "Доказательства", "Конфликты", "Owner50")


def _write_acceptance_workbook(path: Path, sheets: Mapping[str, tuple[Sequence[str], Sequence[Sequence[Any]]]]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)
    for name in _ACCEPTANCE_SHEETS:
        headers, rows = sheets.get(name, ((), ()))
        ws = wb.create_sheet(name)
        ws.append(tuple(headers))
        for row in rows:
            ws.append(tuple(row))
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column in ws.columns:
            ws.column_dimensions[column[0].column_letter].width = min(80, max(12, *(len(str(c.value or "")) for c in column)))
    wb.save(path)
    path.chmod(0o600)


_ACCEPTANCE_EVIDENCE_KIND_BY_EVENT_TYPE = {
    "email_message": "Почта",
    "mango_call": "Звонок",
    "wappi_telegram": "Telegram",
    "wappi_max": "Max",
    "tallanto_attendance": "Посещение",
    "tallanto_absence": "Подтверждённый пропуск",
    "tallanto_scheduled_lesson": "Запланированное занятие",
    "amo_deal_stage": "Сделка",
}

_ACCEPTANCE_COMMUNICATION_EVENT_TYPES = {
    "email_message", "mango_call", "call_transcript", "telegram_message", "telegram_dialog", "max_message",
}
_ACCEPTANCE_COMMUNICATION_SOURCES = {"mail", "mail_archive_stage2", "wappi_telegram", "wappi_max", "telegram_history"}


def _acceptance_evidence_kind(event_type: Any) -> str:
    return _ACCEPTANCE_EVIDENCE_KIND_BY_EVENT_TYPE.get(str(event_type or ""), "Событие")


def _acceptance_event_type(row: Mapping[str, Any]) -> str:
    event_type = str(row["event_type"] or "")
    if event_type != "tallanto_attendance" or row["source_system"] != "tallanto_attendance_api":
        return event_type
    record = json.loads(row["record_json"] or "{}").get("record") or {}
    if record.get("attendance_confirmed"):
        return event_type
    if record.get("physical_absence_confirmed"):
        return "tallanto_absence"
    return "tallanto_scheduled_lesson"


def _acceptance_family_data(con: sqlite3.Connection, *, tenant_id: str, sample: Sequence[Mapping[str, Any]]) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]], list[list[Any]]]:
    families: list[list[Any]] = []; chronology: list[list[Any]] = []
    evidence: list[list[Any]] = []; conflicts: list[list[Any]] = []
    conflict_rows = con.execute(
        "SELECT conflict_id,conflict_type,severity,status,created_at,record_json FROM timeline_conflicts "
        "WHERE tenant_id=? AND status!='resolved' ORDER BY created_at,conflict_id", (tenant_id,)
    ).fetchall()
    # гэп №1 (EXCEL_priemka_30_semey_spec.md §3/§9.1): режим приёмки не должен подставлять
    # match_status там, где его физически нет (минимальная синтетика тестов харнесса не несёт
    # этой колонки) -- тот же защитный паттерн, что _table_exists уже применяет к опциональным
    # таблицам в manager_dossier.py, только на уровне колонки.
    event_columns = {str(row[1]) for row in con.execute("PRAGMA table_info(timeline_events)").fetchall()}
    match_status_clause = (
        " AND (event_type != 'mango_call' OR match_status = 'strong_unique')"
        if "match_status" in event_columns else ""
    )
    for number, selected in enumerate(sample, start=1):
        customer_id = str(selected["id"])
        family_id = str(selected.get("family_id") or customer_id)
        members = _family_scope_customer_ids(con, tenant_id=tenant_id, customer_id=customer_id)
        placeholders = ",".join("?" for _ in members)
        dossier = build_customer_dossier(con, tenant_id=tenant_id, customer_id=customer_id)
        children = con.execute(
            f"SELECT child_key,canonical_name,grades_json,subjects_json FROM family_links_v1 WHERE tenant_id=? "
            f"AND customer_id IN ({placeholders}) AND status!='excluded' ORDER BY child_key", (tenant_id, *members)
        ).fetchall()
        child_text = "; ".join(f"{row['canonical_name']} — {row['grades_json']} — {row['subjects_json']}" for row in children)
        opportunity = con.execute(
            f"SELECT opportunity_id,title,status FROM customer_opportunities WHERE tenant_id=? AND customer_id IN ({placeholders}) "
            "ORDER BY COALESCE(opened_at,'') DESC,opportunity_id LIMIT 1", (tenant_id, *members)
        ).fetchone()
        payment = con.execute(
            f"SELECT SUM(COALESCE(total_in,0)) total_in,MAX(last_purchase_at) last_at FROM customer_purchases_v1 "
            f"WHERE tenant_id=? AND customer_id IN ({placeholders}) AND period='all_time' AND money_kind='fact'", (tenant_id, *members)
        ).fetchone()
        event_rows = con.execute(
            f"SELECT event_id,customer_id,event_at,event_type,source_system,direction,subject,summary,text_preview,source_ref,record_json "
            f"FROM timeline_events WHERE tenant_id=? AND customer_id IN ({placeholders}) AND COALESCE(superseded_by,'')=''"
            f"{match_status_clause} "
            "ORDER BY event_at,event_id", (tenant_id, *members)
        ).fetchall()
        attendance = next(
            (
                row for row in reversed(event_rows)
                if _acceptance_event_type(row) == "tallanto_attendance"
                and _attendance_is_not_future(row["event_at"])
            ),
            None,
        )
        latest = next(
            (
                row for row in reversed(event_rows)
                if row["event_type"] in _ACCEPTANCE_COMMUNICATION_EVENT_TYPES
                or row["source_system"] in _ACCEPTANCE_COMMUNICATION_SOURCES
            ),
            None,
        )
        # entity_refs -- тот же точный признак, что уже используется в _next_step_for_dossier
        # (manager_dossier.py:3070-3083); НЕ substring -- один customer_id может быть текстовым
        # префиксом другого ("customer:1" внутри "customer:10"), substring даёт ложные конфликты.
        member_refs = {ref for member in members for ref in customer_entity_ref_values(member)}
        member_refs.update((family_id, f"family:{family_id}"))
        matched_conflicts = [
            row for row in conflict_rows
            if member_refs & {str(ref) for ref in (json.loads(row["record_json"] or "{}").get("entity_refs") or ())}
        ]
        families.append([
            number, family_id, customer_id, dossier.display_name, dossier.phone, dossier.email, dossier.brand,
            child_text, len(children), opportunity["title"] if opportunity else "", opportunity["status"] if opportunity else "",
            payment["total_in"] if payment else 0, payment["last_at"] if payment else "",
            attendance["event_at"] if attendance else "", attendance["subject"] if attendance else "",
            latest["event_at"] if latest else "", latest["source_system"] if latest else "",
            dossier.next_step, dossier.next_step_source,
            "; ".join(sorted({row["conflict_type"] for row in matched_conflicts})), "", "",
        ])
        evidence.extend([
            [family_id, "Личность", dossier.display_name, "customer_identities", "", "customer_identities", customer_id, "да"],
            *([family_id, "Ребёнок", row["canonical_name"], "family_links_v1.canonical_name", "", "family_links_v1", row["child_key"], "да"] for row in children),
            [family_id, "Оплаты", payment["total_in"] if payment else 0, "customer_purchases_v1.total_in", payment["last_at"] if payment else "", "customer_purchases_v1", customer_id, "да" if (payment and payment["last_at"]) else "нет"],
            [family_id, "Следующий шаг", dossier.next_step, "next_step_resolver.display_text", "", dossier.next_step_source or "", "", "нет"],
        ])
        # гэп №2 (спека §4/§9.1): event_id теперь есть в SQL -- каждое событие хронологии
        # одновременно становится проверяемой строкой доказательства с настоящим event_id
        # (не заглушкой вроде customer_id).
        for row in event_rows:
            display_event_type = _acceptance_event_type(row)
            full_text = row["text_preview"] or row["summary"] or ""
            evidence.append([
                family_id, _acceptance_evidence_kind(display_event_type),
                full_text or row["subject"] or "(пусто)",
                "timeline_events.text_preview" if row["text_preview"] else "timeline_events.summary",
                row["event_at"], row["source_system"], row["event_id"], "да",
            ])
        # гэп №1: "Краткое содержание" (summary) и "Полный текст" (text_preview) -- РЯДОМ, оба
        # исходные значения как есть, БЕЗ заглушки "Полный текст в базе" и без обрезки/лимита.
        for row in event_rows:
            display_event_type = _acceptance_event_type(row)
            chronology.append([
                family_id, row["customer_id"], row["event_id"], row["event_at"], display_event_type,
                row["source_system"], row["direction"], row["subject"],
                row["summary"], row["text_preview"], row["source_ref"],
            ])
        conflicts.extend([[family_id, row["conflict_id"], row["conflict_type"], row["severity"], row["status"],
                           row["created_at"], row["record_json"], "", ""] for row in matched_conflicts])
        # спека §5 "противоречия бренда": 0 или 2+ бренда у корневого клиента -- строка идёт в
        # «Конфликты», не остаётся молча пустой в «Семьи 30» (dossier.brand == "" в обоих случаях,
        # см. build_customer_dossier: brand=brands[0] if len(brands)==1 else "").
        if not dossier.brand:
            conflicts.append([
                family_id, "", "brand_ambiguous", "medium", "open", "",
                f"customer_identities:{customer_id}",
                "0 или несколько брендов у корневого клиента семьи -- не определён однозначно", "",
            ])
    return families, chronology, evidence, conflicts


def _acceptance_owner50_rows(
    candidates: Sequence[Mapping[str, Any]],
    control: Sequence[tuple[str, ...]],
    selected_family_ids: set[str],
) -> list[list[str]]:
    """One readable classification row per sampled family, including non-READY families."""
    grouped: dict[str, list[tuple[str, ...]]] = defaultdict(list)
    for row in candidates:
        family_id = str(row.get("family_id") or "")
        if family_id in selected_family_ids:
            grouped[family_id].append(
                _owner50_control_row_from_ready(row, status="READY", code="ready")
            )
    for row in control:
        family_id = str(row[0])
        if family_id in selected_family_ids:
            grouped[family_id].append(row)

    missing = sorted(selected_family_ids - set(grouped))
    if missing:
        raise RuntimeError(f"owner50 missed sampled families: {', '.join(missing)}")

    result: list[list[str]] = []
    for family_id in sorted(selected_family_ids):
        rows = grouped[family_id]
        first = list(rows[0])
        statuses = {str(row[1]).lower() for row in rows}
        first[1] = "EXCLUDED" if "excluded" in statuses else "CANDIDATE" if "candidate" in statuses else "READY"
        first[2] = "; ".join(dict.fromkeys(str(row[2]) for row in rows if row[2]))
        first[3] = "; ".join(dict.fromkeys(str(row[3]) for row in rows if row[3]))
        result.append(first)
    return result


def cmd_acceptance(args: argparse.Namespace) -> int:
    db = Path(args.db).expanduser()
    out_root = Path(args.out_root).expanduser()
    local_dir = out_root / ".codex_local"
    local_dir.mkdir(parents=True, exist_ok=True)
    local_dir.chmod(0o700)
    run_token = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    out_xlsx = _guard_local_dossier_output_path(local_dir / f"acceptance_30_families_{run_token}.xlsx", out_root)

    scrubbed_manifest: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "db": str(db),
        "tenant_id": args.tenant_id,
        "seed": args.seed,
        "requested_count": args.count,
        "sheets": list(_ACCEPTANCE_SHEETS),
        "writes_amo_notes": False,
        "sends_to_client": False,
        "current_artifact": None,
        "legacy_artifact_present": (local_dir / "acceptance_30_families.xlsx").exists(),
    }

    def strata_key(row: Mapping[str, Any]) -> tuple:
        return (
            row["brand"], row["channel"], row["child_bucket"], row["has_payment"], row["has_conflict"],
            row["has_signal"], row.get("has_mail", False), row.get("has_call", False), row.get("has_attendance", False),
        )

    with _connect_ro(db) as con:
        freshness_gate = manager_freshness_gate(_source_freshness(con, tenant_id=args.tenant_id))
        if not freshness_gate["passed"]:
            # требование задачи: недоступная свежесть -> harness завершается ЧЕСТНО (exit 2,
            # semantic_review_blocked_by_freshness), а не падает необработанным traceback --
            # тот же паттерн, что уже есть у cmd_dossiers/cmd_owner50 выше.
            reasons = ", ".join(f"{item['source_system']}:{item['reason']}" for item in freshness_gate["blockers"])
            scrubbed_manifest["status"] = "semantic_review_blocked_by_freshness"
            scrubbed_manifest["freshness_gate_error"] = f"acceptance freshness gate failed: {reasons}"
            (out_root / "acceptance_selection_manifest.json").write_text(
                json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
                encoding="utf-8",
            )
            print("semantic_review_blocked_by_freshness")
            print(scrubbed_manifest["freshness_gate_error"])
            print(f"Scrubbed selection manifest written (no PII): {out_root / 'acceptance_selection_manifest.json'}")
            return 2

        population = _dossier_population(con, tenant_id=args.tenant_id)
        if not population:
            print(f"BLOCKED: no eligible families found in {db} (empty family_links_v1?).")
            return 3

        sample = stratified_sample(population, count=args.count, seed=args.seed, strata_key=strata_key)
        if len(sample) != args.count:
            print(f"BLOCKED: acceptance requires {args.count} families, population only yields {len(sample)}.")
            return 3
        missing_layers = [
            field for field in ("has_payment", "has_attendance")
            if any(row.get(field) for row in population) and not any(row.get(field) for row in sample)
        ]
        if missing_layers:
            scrubbed_manifest.update(status="semantic_review_blocked_by_sample_coverage", missing_layers=missing_layers)
            (out_root / "acceptance_selection_manifest.json").write_text(
                json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
            )
            print(f"semantic_review_blocked_by_sample_coverage: {', '.join(missing_layers)}")
            return 4
        strata_counter = Counter(strata_key(row) for row in sample)

        families, chronology, evidence, conflicts = _acceptance_family_data(con, tenant_id=args.tenant_id, sample=sample)

        # Owner50-лист: аудитор рекомендует прогонять classify_family() на ЭТИХ ЖЕ 30 семьях,
        # а не тащить отдельный несвязанный пул 50 (build_owner50_family_workbook режет топ-50
        # ПО ВСЕМУ тенанту -- READY-семья из наших 30 может не попасть в топ-50 чужого
        # ранжирования и молча пропасть). _owner50_family_rows -- та же чистая, уже
        # протестированная функция, но без этого среза; на лист идут все три статуса именно
        # этих семей, а несколько причин одной семьи объединяются в одну читаемую строку.
        try:
            candidates, control = _owner50_family_rows(
                con,
                tenant_id=args.tenant_id,
                as_of=datetime.now(timezone.utc),
                family_ids=tuple(str(row["family_id"]) for row in sample),
            )
        except RuntimeError as exc:
            scrubbed_manifest.update(status="semantic_review_blocked_by_owner50", owner50_classification_error=str(exc))
            (out_root / "acceptance_selection_manifest.json").write_text(
                json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
            )
            print("semantic_review_blocked_by_owner50")
            return 4

    selected_family_ids = {str(row[1]) for row in families}
    try:
        owner_rows = _acceptance_owner50_rows(candidates, control, selected_family_ids)
    except RuntimeError as exc:
        scrubbed_manifest.update(status="semantic_review_blocked_by_owner50", owner50_classification_error=str(exc))
        (out_root / "acceptance_selection_manifest.json").write_text(
            json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
        )
        return 4
    owner_headers = list(OWNER50_CONTROL_COLUMNS)

    sheets = {
        "Семьи 30": (("№", "family_id", "customer_id", "Основной контакт (роль не подтверждена)", "Телефон", "Email", "Бренд", "Дети", "Число детей", "Сделка", "Статус сделки", "Оплаты", "Последняя оплата", "Последнее посещение", "Предмет посещения", "Последнее общение", "Канал", "Следующий шаг", "Источник шага", "Конфликты", "F1. Статус", "F2. Комментарий"), families),
        "Хронология": (("family_id", "customer_id", "event_id", "Дата/время", "Тип события", "Источник", "Направление", "Тема", "Краткое содержание", "Полный текст", "source_ref"), chronology),
        "Доказательства": (("family_id", "Тип", "Доказательство", "Точное поле", "Дата", "source_system", "event_id/record_id", "Проверяемо"), evidence),
        "Конфликты": (("family_id", "conflict_id", "Тип", "Критичность", "Статус", "Дата", "Исходная запись", "F1. Статус", "F2. Комментарий"), conflicts),
        "Owner50": (owner_headers, owner_rows),
    }
    _write_acceptance_workbook(out_xlsx, sheets)

    scrubbed_manifest["status"] = "sample_built"
    scrubbed_manifest["current_artifact"] = f".codex_local/{out_xlsx.name}"
    scrubbed_manifest["population_size"] = len(population)
    scrubbed_manifest["sample_size"] = len(sample)
    scrubbed_manifest["distinct_strata_covered"] = len(strata_counter)
    scrubbed_manifest["strata_breakdown"] = [
        {
            "brand": key[0], "channel": key[1], "child_bucket": key[2], "has_payment": key[3],
            "has_conflict": key[4], "has_signal": key[5], "has_mail": key[6], "has_calls": key[7],
            "has_attendance": key[8],
            "count": n,
        }
        for key, n in sorted(strata_counter.items(), key=lambda kv: str(kv[0]))
    ]
    scrubbed_manifest["sample_customer_id_hashes"] = [_short_hash(row["id"]) for row in sample]
    scrubbed_manifest["families_rows"] = len(families)
    scrubbed_manifest["chronology_rows"] = len(chronology)
    scrubbed_manifest["evidence_rows"] = len(evidence)
    scrubbed_manifest["conflicts_rows"] = len(conflicts)
    status_counts = Counter(str(row[1]) for row in owner_rows)
    scrubbed_manifest["owner50_ready_rows"] = status_counts["READY"]
    scrubbed_manifest["owner50_candidate_rows"] = status_counts["CANDIDATE"]
    scrubbed_manifest["owner50_excluded_rows"] = status_counts["EXCLUDED"]
    scrubbed_manifest["chronology_row_limit_applied"] = False
    (out_root / "acceptance_selection_manifest.json").write_text(
        json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    print(f"OK: {len(families)} families -> {out_xlsx} (5 sheets: {', '.join(_ACCEPTANCE_SHEETS)}; PII, local only, not for git).")
    print(f"Scrubbed selection manifest (no PII): {out_root / 'acceptance_selection_manifest.json'}")
    return 0


# ---------------------------------------------------------------------------
# Subcommand: drafts
# ---------------------------------------------------------------------------

_UNMATCHED_STATUSES = {"pair_missing"}
_CONFLICT_STATUSES = {"identity_conflict"}
_NON_PERSONAL_EVENTS = {"chat_skipped"}


def _iter_journal(path: Path) -> Iterable[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError:
                continue


def _draft_population(journal_path: Path) -> tuple[list[dict[str, Any]], Counter]:
    per_chat_inbound_count: Counter = Counter()
    rows_by_message: dict[str, dict[str, Any]] = {}
    zero_check = Counter()
    for rec in _iter_journal(journal_path):
        event = str(rec.get("event") or "")
        status = str(rec.get("status") or "")
        if event in _NON_PERSONAL_EVENTS:
            continue
        # Machine-zero cross-check: any event/status implying a live AMO write or client send
        # outside of dry_run must not exist in a default (dry-run) run.
        if status == "note_written" or event == "note_written":
            zero_check["note_written_like"] += 1
        if event in {"client_send", "message_sent"} or status in {"client_send", "message_sent"}:
            zero_check["client_send_like"] += 1
        chat_id = str(rec.get("chat_id") or "")
        message_id = str(rec.get("message_id") or rec.get("last_message_id") or "")
        if not message_id:
            continue
        if chat_id:
            per_chat_inbound_count[chat_id] += 1
        rows_by_message[message_id] = rec

    population = []
    for message_id, rec in rows_by_message.items():
        chat_id = str(rec.get("chat_id") or "")
        status = str(rec.get("status") or rec.get("event") or "unknown")
        route = str(rec.get("route") or "")
        safety_flags = rec.get("safety_flags") or ()
        is_p0 = any("p0" in str(flag).casefold() for flag in safety_flags) or "manager_only" in route.casefold()
        population.append(
            {
                "id": message_id,
                "chat_id": chat_id,
                "brand": str(rec.get("brand") or "unknown").strip().casefold() or "unknown",
                "channel": "max" if ":max:" in f":{rec.get('profile_id', '')}:".casefold() else "telegram",
                "is_p0": bool(is_p0),
                "is_unmatched": status in _UNMATCHED_STATUSES,
                "is_identity_conflict": status in _CONFLICT_STATUSES,
                "is_multi_turn": per_chat_inbound_count.get(chat_id, 0) > 1,
                "status": status,
                "route": route,
            }
        )
    return population, zero_check


def cmd_drafts(args: argparse.Namespace) -> int:
    journal_path = Path(args.journal).expanduser()
    out_root = Path(args.out_root).expanduser()
    local_dir = out_root / ".codex_local"
    local_dir.mkdir(parents=True, exist_ok=True)
    local_dir.chmod(0o700)

    if not journal_path.exists():
        print(f"BLOCKED: journal not found at {journal_path}. Run step 2 (run_amo_wappi_draft_loop.py --once) first.")
        return 3

    population, zero_check = _draft_population(journal_path)
    if not population:
        print(f"BLOCKED: journal at {journal_path} has zero eligible personal-inbound rows.")
        return 3

    def strata_key(row: Mapping[str, Any]) -> tuple:
        return (
            row["brand"], row["channel"], row["is_p0"], row["is_unmatched"],
            row["is_identity_conflict"], row["is_multi_turn"],
        )

    sample = stratified_sample(population, count=args.count, seed=args.seed, strata_key=strata_key)
    strata_counter = Counter(strata_key(row) for row in sample)

    terminal_outcomes = sum(1 for row in population if row["status"] not in {"", "unknown"})
    scrubbed_manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "journal": str(journal_path),
        "seed": args.seed,
        "requested_count": args.count,
        "population_size": len(population),
        "sample_size": len(sample),
        "full_corpus_terminal_outcome_count": terminal_outcomes,
        "full_corpus_total": len(population),
        "full_corpus_terminal_outcome_pct": round(100.0 * terminal_outcomes / len(population), 2),
        "machine_zero_check": {
            "note_written_like_events_in_this_run": zero_check.get("note_written_like", 0),
            "client_send_like_events_in_this_run": zero_check.get("client_send_like", 0),
            "expected": "both must be 0 unless the loop was run with --live-write (must not happen in this TZ)",
        },
        "distinct_strata_covered": len(strata_counter),
        "strata_breakdown": [
            {
                "brand": key[0], "channel": key[1], "is_p0": key[2], "is_unmatched": key[3],
                "is_identity_conflict": key[4], "is_multi_turn": key[5], "count": n,
            }
            for key, n in sorted(strata_counter.items(), key=lambda kv: str(kv[0]))
        ],
        "sample_message_id_hashes": [_short_hash(row["id"]) for row in sample],
    }
    (out_root / "drafts_selection_manifest.json").write_text(
        json.dumps(scrubbed_manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    # PII-bearing extract for the owner/manager's own local semantic review only.
    full_rows = []
    with journal_path.open("r", encoding="utf-8") as fh:
        by_message = {}
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            mid = str(rec.get("message_id") or rec.get("last_message_id") or "")
            if mid:
                by_message[mid] = rec
    for row in sample:
        full_rows.append(by_message.get(row["id"], {}))
    _write_private_text(
        local_dir / "drafts_sample_full.jsonl",
        "\n".join(json.dumps(r, ensure_ascii=False, sort_keys=True) for r in full_rows) + "\n",
    )

    print(f"OK: {len(sample)} drafts selected from {len(population)} journal rows.")
    print(f"Full-corpus terminal-outcome coverage: {terminal_outcomes}/{len(population)} ({scrubbed_manifest['full_corpus_terminal_outcome_pct']}%).")
    print(f"Machine-zero cross-check (this run): note_written_like={zero_check.get('note_written_like', 0)}, client_send_like={zero_check.get('client_send_like', 0)}.")
    print(f"Scrubbed selection manifest (no PII): {out_root / 'drafts_selection_manifest.json'}")
    print(f"Full PII extract for manual review, LOCAL ONLY, do not commit: {local_dir / 'drafts_sample_full.jsonl'}")
    return 0


# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_owner50 = sub.add_parser(
        "owner50", help="All READY + deterministically 15 stratified CANDIDATE + 15 stratified EXCLUDED.",
    )
    p_owner50.add_argument("--db", required=True, help="Path to the staging customer_timeline.sqlite.")
    p_owner50.add_argument("--out-root", required=True)
    p_owner50.add_argument("--tenant-id", default="foton")
    p_owner50.add_argument("--seed", type=int, default=20260726)
    p_owner50.add_argument("--candidate-count", type=int, default=15)
    p_owner50.add_argument("--excluded-count", type=int, default=15)
    p_owner50.set_defaults(func=cmd_owner50)

    p_dossiers = sub.add_parser("dossiers", help="Sample 30 stratified family dossiers.")
    p_dossiers.add_argument("--db", required=True, help="Path to the staging customer_timeline.sqlite.")
    p_dossiers.add_argument("--out-root", required=True)
    p_dossiers.add_argument("--tenant-id", default="foton")
    p_dossiers.add_argument("--seed", type=int, default=20260726)
    p_dossiers.add_argument("--count", type=int, default=30)
    p_dossiers.set_defaults(func=cmd_dossiers)

    p_acceptance = sub.add_parser("acceptance", help="One five-sheet XLSX for manual review of 30 families.")
    p_acceptance.add_argument("--db", required=True, help="Path to the staging customer_timeline.sqlite.")
    p_acceptance.add_argument("--out-root", required=True)
    p_acceptance.add_argument("--tenant-id", default="foton")
    p_acceptance.add_argument("--seed", type=int, default=20260726)
    p_acceptance.add_argument("--count", type=int, default=30)
    p_acceptance.set_defaults(func=cmd_acceptance)

    p_drafts = sub.add_parser("drafts", help="Sample 50 stratified blind Wappi drafts from a dry-run journal.")
    p_drafts.add_argument("--journal", required=True, help="Path to draft_loop journal.jsonl produced by a dry run.")
    p_drafts.add_argument("--out-root", required=True)
    p_drafts.add_argument("--seed", type=int, default=20260726)
    p_drafts.add_argument("--count", type=int, default=50)
    p_drafts.set_defaults(func=cmd_drafts)

    args = parser.parse_args()
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
