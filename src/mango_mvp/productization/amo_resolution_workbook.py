from __future__ import annotations

import csv
import html
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Optional

from mango_mvp.productization.amo_manual_resolution import RESOLUTION_COLUMNS


WORKBOOK_SCHEMA_VERSION = "amo_resolution_review_workbook_v1"
REVIEW_ASSETS_SCHEMA_VERSION = "amo_resolution_review_assets_v1"
DEFAULT_AMO_BASE_URL = "https://educent.amocrm.ru"

STATUS_VALUES = [
    "needs_human",
    "accepted_by_manager",
    "accepted_by_operator",
    "accepted_auto_policy",
    "blocked",
    "needs_manager",
    "needs_text_review",
    "already_written_review",
]
ALLOW_VALUES = ["no", "yes"]
REASON_VALUES = [
    "source_contact_confirmed_in_amo",
    "source_contact_confirmed_in_amo_post_merge_recheck_approved",
    "duplicate_merge_completed_post_merge_recheck_approved",
    "outside_source_approved_correct_amo_card",
    "text_quality_approved",
    "text_quality_approved_refresh_approved",
    "contact_id_mismatch_requires_operator_check",
    "multiple_exact_amo_contacts_require_manager_choice",
    "crm_text_requires_review_before_live_write",
    "do_not_rewrite_until_text_review_is_resolved",
]


REVIEW_HEADERS = [
    "N",
    "resolution_id",
    "phone",
    "queue_bucket",
    "recommended_action_ru",
    "resolution_status",
    "resolved_contact_id",
    "allow_contact_id_outside_source",
    "resolution_reason",
    "resolved_by",
    "resolution_notes",
    "validation_hint",
    "suggested_resolved_contact_id",
    "source_amo_contact_ids",
    "dry_run_contact_ids",
    "suggested_contact_link",
    "dry_run_contact_links",
    "fio_parent",
    "fio_child",
    "latest_call_date",
    "latest_call_type",
    "priority",
    "sale_probability_percent",
    "next_step",
    "amo_lead_ids",
    "lead_links",
    "written_status",
    "written_contact_id",
    "suggested_resolution_status",
    "suggested_reason",
    "queue_reason",
    "source_row_index",
    "crm_last_call_summary",
    "crm_short_history",
    "crm_timeline",
    "crm_objections",
    "crm_recommended_next_contact_date",
    "crm_ready_reason",
]


COMPARISON_HEADERS = [
    "resolution_id",
    "phone",
    "candidate_contact_id",
    "candidate_source",
    "in_source_amo_ids",
    "is_suggested",
    "is_written_contact",
    "contact_link",
    "operator_hint_ru",
    "risk_note_ru",
    "source_amo_contact_ids",
    "dry_run_contact_ids",
    "amo_lead_ids",
    "lead_links",
    "fio_parent",
    "fio_child",
    "latest_call_date",
    "priority",
    "sale_probability_percent",
    "next_step",
]


def build_amo_resolution_review_workbook(
    *,
    pack_root: Path,
    out_xlsx: Path,
    amo_base_url: str = DEFAULT_AMO_BASE_URL,
    generated_at: Optional[datetime] = None,
) -> Mapping[str, Any]:
    """Build an operator-friendly XLSX for AMO manual-resolution decisions."""

    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("openpyxl is required to build AMO resolution workbooks") from exc

    pack_root = pack_root.expanduser().resolve(strict=False)
    out_xlsx = out_xlsx.expanduser().resolve(strict=False)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    now = generated_at or datetime.now(timezone.utc)
    rows = _load_pack_rows(pack_root)
    summary = _read_json(pack_root / "summary.json")
    amo_base_url = amo_base_url.rstrip("/")

    wb = Workbook()
    ws = wb.active
    ws.title = "Решения"
    comparison = wb.create_sheet("AMO-кандидаты")
    instruction = wb.create_sheet("Инструкция")
    statuses = wb.create_sheet("Справочник")
    audit = wb.create_sheet("Audit")

    _build_instruction_sheet(instruction, pack_root, out_xlsx, now, summary)
    _build_status_sheet(statuses)
    _build_review_sheet(
        ws,
        rows=rows,
        amo_base_url=amo_base_url,
        Table=Table,
        TableStyleInfo=TableStyleInfo,
        DataValidation=DataValidation,
        FormulaRule=FormulaRule,
        Comment=Comment,
        Font=Font,
        PatternFill=PatternFill,
        Alignment=Alignment,
        Border=Border,
        Side=Side,
    )
    _build_comparison_sheet(
        comparison,
        rows=rows,
        amo_base_url=amo_base_url,
        Table=Table,
        TableStyleInfo=TableStyleInfo,
        Font=Font,
        PatternFill=PatternFill,
        Alignment=Alignment,
        Border=Border,
        Side=Side,
    )
    _build_audit_sheet(audit, rows, summary, now)

    for sheet in wb.worksheets:
        sheet.sheet_view.showGridLines = False

    wb.save(out_xlsx)
    return {
        "schema_version": WORKBOOK_SCHEMA_VERSION,
        "out_xlsx": str(out_xlsx),
        "pack_root": str(pack_root),
        "rows": len(rows),
        "generated_at": now.isoformat(timespec="seconds"),
        "summary": {
            "needs_human_rows": sum(1 for row in rows if row.get("workbook_group") == "needs_human"),
            "already_written_review_rows": sum(1 for row in rows if row.get("workbook_group") == "already_written_review"),
        },
    }


def build_amo_resolution_review_html(
    *,
    pack_root: Path,
    out_html: Path,
    amo_base_url: str = DEFAULT_AMO_BASE_URL,
    generated_at: Optional[datetime] = None,
) -> Mapping[str, Any]:
    """Build a read-only HTML review page for the same AMO manual-resolution rows."""

    pack_root = pack_root.expanduser().resolve(strict=False)
    out_html = out_html.expanduser().resolve(strict=False)
    out_html.parent.mkdir(parents=True, exist_ok=True)
    now = generated_at or datetime.now(timezone.utc)
    rows = _load_pack_rows(pack_root)
    comparison_rows = _candidate_rows(rows, amo_base_url.rstrip("/"))
    summary = {
        "schema_version": REVIEW_ASSETS_SCHEMA_VERSION,
        "generated_at": now.isoformat(timespec="seconds"),
        "pack_root": str(pack_root),
        "out_html": str(out_html),
        "review_rows": len(rows),
        "candidate_rows": len(comparison_rows),
        "safety": {
            "read_only": True,
            "live_write": False,
            "write_crm": False,
        },
    }
    out_html.write_text(_render_review_html(rows, comparison_rows, now), encoding="utf-8")
    return summary


def export_decisions_from_amo_resolution_workbook(
    *,
    workbook_path: Path,
    out_csv: Path,
) -> Mapping[str, Any]:
    """Export the Decisions sheet back to the exact manual-resolution decisions CSV schema."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment guard
        raise RuntimeError("openpyxl is required to read AMO resolution workbooks") from exc

    workbook_path = workbook_path.expanduser().resolve(strict=False)
    out_csv = out_csv.expanduser().resolve(strict=False)
    wb = load_workbook(workbook_path, data_only=False)
    if "Решения" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'Решения' sheet")
    ws = wb["Решения"]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    missing = [column for column in RESOLUTION_COLUMNS if column not in header_index]
    if missing:
        raise ValueError(f"Workbook is missing required decision columns: {', '.join(missing)}")

    rows: list[dict[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in row):
            continue
        decision = {column: _safe_text(row[header_index[column]]) for column in RESOLUTION_COLUMNS}
        if decision.get("resolution_id"):
            rows.append(decision)

    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=RESOLUTION_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    return {
        "schema_version": WORKBOOK_SCHEMA_VERSION,
        "workbook_path": str(workbook_path),
        "out_csv": str(out_csv),
        "rows": len(rows),
    }


def _load_pack_rows(pack_root: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    source_by_index, source_by_phone = _load_source_context(pack_root)
    for filename, group in (("needs_human.csv", "needs_human"), ("already_written_review.csv", "already_written_review")):
        for row in _read_csv(pack_root / filename):
            row = dict(row)
            row["workbook_group"] = group
            source = source_by_index.get(_safe_text(row.get("source_row_index"))) or source_by_phone.get(
                _safe_text(row.get("phone"))
            )
            if source:
                _merge_source_context(row, source)
            rows.append(row)
    return rows


def _build_instruction_sheet(ws: Any, pack_root: Path, out_xlsx: Path, now: datetime, summary: Mapping[str, Any]) -> None:
    ws["A1"] = "AMO manual-resolution: инструкция"
    ws["A1"].font = _font(size=18, bold=True, color="1F2937")
    ws["A2"] = "Заполняйте только желтые поля на листе 'Решения'. Остальные поля нужны для проверки и контекста."
    ws["A4"] = "Что делать"
    ws["A4"].font = _font(bold=True)
    instructions = [
        "1. Откройте AMO-ссылку по suggested_contact_link и сравните с альтернативными dry_run_contact_ids.",
        "2. Если suggested контакт верный: resolution_status=accepted_by_manager, resolved_contact_id=suggested_resolved_contact_id.",
        "3. Для accepted обязательно заполните resolution_reason и resolved_by.",
        "4. Для multi-contact дублей accepted разрешен только после AMO-склейки и post-merge recheck; reason должен содержать post_merge_recheck_approved.",
        "5. Если выбран contact_id не из source_amo_contact_ids: allow_contact_id_outside_source=yes и reason должен содержать outside_source_approved.",
        "6. Для text-quality строк reason должен содержать text_quality_approved.",
        "7. Для already-written refresh reason должен содержать refresh_approved; без явной причины лучше не переписывать.",
        "8. Если сомневаетесь, оставьте needs_human и добавьте комментарий в resolution_notes.",
    ]
    for idx, text in enumerate(instructions, start=5):
        ws[f"A{idx}"] = text
    ws["A14"] = "Исходный pack"
    ws["B14"] = str(pack_root)
    ws["A15"] = "Файл"
    ws["B15"] = str(out_xlsx)
    ws["A16"] = "Сгенерировано"
    ws["B16"] = now.isoformat(timespec="seconds")
    ws["A18"] = "Текущий summary"
    summary_obj = summary.get("summary") if isinstance(summary.get("summary"), Mapping) else summary
    for offset, key in enumerate(("review_rows", "needs_human_rows", "already_written_review_rows", "accepted_rows", "resolved_live_candidate_rows"), start=19):
        ws[f"A{offset}"] = key
        ws[f"B{offset}"] = summary_obj.get(key) if isinstance(summary_obj, Mapping) else ""
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows(min_row=1, max_row=24, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = _alignment(wrap_text=True, vertical="top")


def _build_status_sheet(ws: Any) -> None:
    ws["A1"] = "Допустимые значения"
    ws["A1"].font = _font(size=16, bold=True)
    ws.append(["resolution_status", "Когда использовать"])
    descriptions = {
        "needs_human": "Нет уверенного решения; строка не попадет в live.",
        "accepted_by_manager": "Менеджер/оператор проверил AMO и подтвердил contact_id.",
        "accepted_by_operator": "Оператор подтвердил contact_id по технической сверке.",
        "accepted_auto_policy": "Автоматическая политика применима; сейчас лучше не использовать без отдельного правила.",
        "blocked": "Строка должна остаться заблокированной.",
        "needs_manager": "Нужно решение менеджера.",
        "needs_text_review": "Нужно проверить текст CRM перед записью.",
        "already_written_review": "Строка уже записана; не переписывать без refresh-approved решения.",
    }
    for status in STATUS_VALUES:
        ws.append([status, descriptions.get(status, "")])
    start = len(STATUS_VALUES) + 4
    ws[f"A{start}"] = "resolution_reason"
    ws[f"B{start}"] = "Когда использовать"
    reason_descriptions = {
        "source_contact_confirmed_in_amo": "Выбран contact_id из source_amo_contact_ids, карточка проверена.",
        "source_contact_confirmed_in_amo_post_merge_recheck_approved": "Source contact подтвержден после склейки AMO-дублей и post-merge recheck.",
        "duplicate_merge_completed_post_merge_recheck_approved": "AMO-дубли склеены вручную, surviving contact_id подтвержден post-merge recheck.",
        "outside_source_approved_correct_amo_card": "Выбран contact_id вне source_amo_contact_ids, но карточка точно правильная.",
        "text_quality_approved": "Текст CRM проверен и не требует исправлений.",
        "text_quality_approved_refresh_approved": "Текст проверен, ранее записанную строку можно обновить.",
        "contact_id_mismatch_requires_operator_check": "Contact-id mismatch остается блокером.",
        "multiple_exact_amo_contacts_require_manager_choice": "Есть несколько точных контактов, нужен выбор.",
        "crm_text_requires_review_before_live_write": "Текст требует проверки до live.",
        "do_not_rewrite_until_text_review_is_resolved": "Не переписывать ранее записанную строку.",
    }
    for idx, reason in enumerate(REASON_VALUES, start=start + 1):
        ws[f"A{idx}"] = reason
        ws[f"B{idx}"] = reason_descriptions.get(reason, "")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 90


def _build_review_sheet(
    ws: Any,
    *,
    rows: list[Mapping[str, str]],
    amo_base_url: str,
    Table: Any,
    TableStyleInfo: Any,
    DataValidation: Any,
    FormulaRule: Any,
    Comment: Any,
    Font: Any,
    PatternFill: Any,
    Alignment: Any,
    Border: Any,
    Side: Any,
) -> None:
    ws.append(REVIEW_HEADERS)
    editable = {
        "resolution_status",
        "resolved_contact_id",
        "allow_contact_id_outside_source",
        "resolution_reason",
        "resolved_by",
        "resolution_notes",
    }
    yellow = PatternFill("solid", fgColor="FFF2CC")
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_no, source in enumerate(rows, start=2):
        review_row = _review_row(row_no - 1, source, amo_base_url)
        ws.append([review_row.get(header, "") for header in REVIEW_HEADERS])
        for col_no, header in enumerate(REVIEW_HEADERS, start=1):
            cell = ws.cell(row=row_no, column=col_no)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if header in editable:
                cell.fill = yellow
            if header.endswith("_link") or header.endswith("_links"):
                _apply_hyperlink(cell, str(cell.value or ""))
        ws.cell(row=row_no, column=12).value = _validation_formula(row_no)
        ws.cell(row=row_no, column=12).comment = Comment("OK означает, что поля заполнены достаточно для следующей автоматической валидации. Финальная проверка все равно будет в скрипте.", "Codex")

    if rows:
        table = Table(displayName="AmoManualResolution", ref=f"A1:{_excel_col(len(REVIEW_HEADERS))}{len(rows)+1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    status_dv = DataValidation(type="list", formula1=f'"{",".join(STATUS_VALUES)}"', allow_blank=False)
    allow_dv = DataValidation(type="list", formula1='"no,yes"', allow_blank=False)
    reason_dv = DataValidation(type="list", formula1=f'"{",".join(REASON_VALUES)}"', allow_blank=True)
    ws.add_data_validation(status_dv)
    ws.add_data_validation(allow_dv)
    ws.add_data_validation(reason_dv)
    max_row = len(rows) + 1
    status_dv.add(f"F2:F{max_row}")
    allow_dv.add(f"H2:H{max_row}")
    reason_dv.add(f"I2:I{max_row}")

    ws.conditional_formatting.add(f"L2:L{max_row}", FormulaRule(formula=['LEFT($L2,5)="ERROR"'], fill=PatternFill("solid", fgColor="FEE4E2")))
    ws.conditional_formatting.add(f"L2:L{max_row}", FormulaRule(formula=['LEFT($L2,2)="OK"'], fill=PatternFill("solid", fgColor="D1FADF")))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_excel_col(len(REVIEW_HEADERS))}{max_row}"
    widths = {
        "A": 6,
        "B": 28,
        "C": 17,
        "D": 30,
        "E": 54,
        "F": 24,
        "G": 18,
        "H": 16,
        "I": 42,
        "J": 16,
        "K": 52,
        "L": 42,
        "M": 18,
        "N": 24,
        "O": 30,
        "P": 24,
        "Q": 42,
        "R": 28,
        "S": 28,
        "T": 20,
        "U": 14,
        "V": 12,
        "W": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 58


def _build_comparison_sheet(
    ws: Any,
    *,
    rows: list[Mapping[str, str]],
    amo_base_url: str,
    Table: Any,
    TableStyleInfo: Any,
    Font: Any,
    PatternFill: Any,
    Alignment: Any,
    Border: Any,
    Side: Any,
) -> None:
    comparison_rows = _candidate_rows(rows, amo_base_url)
    ws.append(COMPARISON_HEADERS)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D0D5DD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row_no, row in enumerate(comparison_rows, start=2):
        ws.append([row.get(header, "") for header in COMPARISON_HEADERS])
        for col_no, header in enumerate(COMPARISON_HEADERS, start=1):
            cell = ws.cell(row=row_no, column=col_no)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border
            if header == "contact_link" or header == "lead_links":
                _apply_hyperlink(cell, str(cell.value or ""))
            if row.get("risk_note_ru") and header in {"candidate_contact_id", "risk_note_ru"}:
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
            if row.get("is_suggested") == "yes" and header in {"candidate_contact_id", "is_suggested"}:
                cell.fill = PatternFill("solid", fgColor="D1FADF")
    if comparison_rows:
        table = Table(displayName="AmoCandidateComparison", ref=f"A1:{_excel_col(len(COMPARISON_HEADERS))}{len(comparison_rows)+1}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{_excel_col(len(COMPARISON_HEADERS))}{len(comparison_rows)+1}"
    widths = {
        "A": 28,
        "B": 17,
        "C": 18,
        "D": 28,
        "E": 14,
        "F": 12,
        "G": 14,
        "H": 34,
        "I": 48,
        "J": 48,
        "K": 24,
        "L": 28,
        "M": 32,
        "N": 42,
        "O": 28,
        "P": 28,
        "Q": 20,
        "R": 12,
        "S": 12,
        "T": 48,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row_no in range(2, len(comparison_rows) + 2):
        ws.row_dimensions[row_no].height = 50


def _build_audit_sheet(ws: Any, rows: list[Mapping[str, str]], summary: Mapping[str, Any], now: datetime) -> None:
    ws["A1"] = "Audit"
    ws["A1"].font = _font(size=16, bold=True)
    checks = [
        ("schema_version", WORKBOOK_SCHEMA_VERSION),
        ("generated_at", now.isoformat(timespec="seconds")),
        ("review_rows_in_workbook", len(rows)),
        ("accepted_rows_initial", sum(1 for row in rows if str(row.get("resolution_status") or "").startswith("accepted"))),
        ("resolved_contact_ids_prefilled_initial", sum(1 for row in rows if row.get("resolved_contact_id"))),
        ("source_summary_validation_ok", _dig(summary, "summary", "validation_ok")),
        ("source_summary_review_rows", _dig(summary, "summary", "review_rows")),
    ]
    for idx, (key, value) in enumerate(checks, start=3):
        ws[f"A{idx}"] = key
        ws[f"B{idx}"] = value
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 80


def _review_row(index: int, source: Mapping[str, str], amo_base_url: str) -> dict[str, Any]:
    suggested = _safe_text(source.get("suggested_resolved_contact_id"))
    lead_ids = _split_ids(source.get("amo_lead_ids"))
    dry_ids = _split_ids(source.get("dry_run_contact_ids"))
    return {
        "N": index,
        "resolution_id": source.get("resolution_id", ""),
        "phone": source.get("phone", ""),
        "queue_bucket": source.get("queue_bucket", ""),
        "recommended_action_ru": _recommended_action(source),
        "resolution_status": source.get("resolution_status", "needs_human") or "needs_human",
        "resolved_contact_id": source.get("resolved_contact_id", ""),
        "allow_contact_id_outside_source": source.get("allow_contact_id_outside_source", "no") or "no",
        "resolution_reason": source.get("resolution_reason", ""),
        "resolved_by": source.get("resolved_by", ""),
        "resolution_notes": source.get("resolution_notes", ""),
        "validation_hint": "",
        "suggested_resolved_contact_id": suggested,
        "source_amo_contact_ids": source.get("source_amo_contact_ids", ""),
        "dry_run_contact_ids": source.get("dry_run_contact_ids", ""),
        "suggested_contact_link": _contact_url(amo_base_url, suggested) if suggested else "",
        "dry_run_contact_links": "\n".join(_contact_url(amo_base_url, contact_id) for contact_id in dry_ids),
        "fio_parent": source.get("fio_parent", ""),
        "fio_child": source.get("fio_child", ""),
        "latest_call_date": source.get("latest_call_date", ""),
        "latest_call_type": source.get("latest_call_type", ""),
        "priority": source.get("priority", ""),
        "sale_probability_percent": source.get("sale_probability_percent", ""),
        "next_step": source.get("next_step", ""),
        "amo_lead_ids": source.get("amo_lead_ids", ""),
        "lead_links": "\n".join(_lead_url(amo_base_url, lead_id) for lead_id in lead_ids),
        "written_status": source.get("written_status", ""),
        "written_contact_id": source.get("written_contact_id", ""),
        "suggested_resolution_status": source.get("suggested_resolution_status", ""),
        "suggested_reason": source.get("suggested_reason", ""),
        "queue_reason": source.get("queue_reason", ""),
        "source_row_index": source.get("source_row_index", ""),
        "crm_last_call_summary": source.get("crm_last_call_summary", ""),
        "crm_short_history": source.get("crm_short_history", ""),
        "crm_timeline": source.get("crm_timeline", ""),
        "crm_objections": source.get("crm_objections", ""),
        "crm_recommended_next_contact_date": source.get("crm_recommended_next_contact_date", ""),
        "crm_ready_reason": source.get("crm_ready_reason", ""),
    }


def _recommended_action(row: Mapping[str, str]) -> str:
    bucket = _safe_text(row.get("queue_bucket"))
    suggested = _safe_text(row.get("suggested_resolved_contact_id"))
    source_ids = _safe_text(row.get("source_amo_contact_ids"))
    dry_ids = _safe_text(row.get("dry_run_contact_ids"))
    written = _safe_text(row.get("written_status")).casefold()
    if bucket == "needs_manager_review_multi_contact":
        return f"Проверьте AMO. Если contact_id {suggested} верный, примите его; альтернативы dry-run: {dry_ids}."
    if bucket == "blocked_contact_id_mismatch":
        return f"Не принимать автоматически: source {source_ids}, live dry-run {dry_ids}. Нужно открыть обе карточки и выбрать правильную."
    if bucket == "needs_text_quality_review" and written == "written":
        return "Уже записано ранее. Не переписывать без явного refresh_approved."
    if bucket == "needs_text_quality_review":
        return "Проверить текст CRM. Если корректен, нужен reason text_quality_approved."
    return "Оставить needs_human, если нет уверенного решения."


def _candidate_rows(rows: list[Mapping[str, str]], amo_base_url: str) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for row in rows:
        source_ids = _split_ids(row.get("source_amo_contact_ids"))
        dry_ids = _split_ids(row.get("dry_run_contact_ids"))
        written_id = _safe_text(row.get("written_contact_id"))
        suggested_id = _safe_text(row.get("suggested_resolved_contact_id"))
        ordered_ids: list[str] = []
        for candidate_id in [suggested_id, *source_ids, *dry_ids, written_id]:
            if candidate_id and candidate_id not in ordered_ids:
                ordered_ids.append(candidate_id)
        if not ordered_ids:
            ordered_ids.append("")
        for candidate_id in ordered_ids:
            labels: list[str] = []
            if candidate_id and candidate_id == suggested_id:
                labels.append("suggested")
            if candidate_id and candidate_id in source_ids:
                labels.append("source")
            if candidate_id and candidate_id in dry_ids:
                labels.append("dry_run")
            if candidate_id and candidate_id == written_id:
                labels.append("already_written")
            result.append(
                {
                    "resolution_id": _safe_text(row.get("resolution_id")),
                    "phone": _safe_text(row.get("phone")),
                    "candidate_contact_id": candidate_id,
                    "candidate_source": " | ".join(labels) if labels else "no_contact_candidate",
                    "in_source_amo_ids": "yes" if candidate_id and candidate_id in source_ids else "no",
                    "is_suggested": "yes" if candidate_id and candidate_id == suggested_id else "no",
                    "is_written_contact": "yes" if candidate_id and candidate_id == written_id else "no",
                    "contact_link": _contact_url(amo_base_url, candidate_id) if candidate_id else "",
                    "operator_hint_ru": _candidate_operator_hint(row, candidate_id, source_ids, dry_ids),
                    "risk_note_ru": _candidate_risk_note(row, candidate_id, source_ids, dry_ids),
                    "source_amo_contact_ids": _safe_text(row.get("source_amo_contact_ids")),
                    "dry_run_contact_ids": _safe_text(row.get("dry_run_contact_ids")),
                    "amo_lead_ids": _safe_text(row.get("amo_lead_ids")),
                    "lead_links": "\n".join(_lead_url(amo_base_url, lead_id) for lead_id in _split_ids(row.get("amo_lead_ids"))),
                    "fio_parent": _safe_text(row.get("fio_parent")),
                    "fio_child": _safe_text(row.get("fio_child")),
                    "latest_call_date": _safe_text(row.get("latest_call_date")),
                    "priority": _safe_text(row.get("priority")),
                    "sale_probability_percent": _safe_text(row.get("sale_probability_percent")),
                    "next_step": _safe_text(row.get("next_step")),
                }
            )
    return result


def _candidate_operator_hint(
    row: Mapping[str, str],
    candidate_id: str,
    source_ids: list[str],
    dry_ids: list[str],
) -> str:
    bucket = _safe_text(row.get("queue_bucket"))
    suggested = _safe_text(row.get("suggested_resolved_contact_id"))
    if bucket == "needs_manager_review_multi_contact" and candidate_id == suggested and candidate_id in dry_ids:
        return "Вероятный выбор: source contact совпал с live exact-match; откройте карточку и подтвердите."
    if bucket == "blocked_contact_id_mismatch":
        if candidate_id in source_ids:
            return "Source-карточка из AMO-ready. Сравнить с dry-run карточкой; авто-принятие заблокировано."
        if candidate_id in dry_ids:
            return "Live dry-run нашел другую карточку. При выборе требуется outside_source approval."
    if bucket == "needs_text_quality_review":
        return "Проверить текст истории/следующего шага; принимать только с text_quality_approved."
    return "Открыть AMO и подтвердить, что это правильная карточка клиента."


def _candidate_risk_note(
    row: Mapping[str, str],
    candidate_id: str,
    source_ids: list[str],
    dry_ids: list[str],
) -> str:
    risks: list[str] = []
    bucket = _safe_text(row.get("queue_bucket"))
    if bucket == "blocked_contact_id_mismatch":
        risks.append("contact_id mismatch: live dry-run и source расходятся")
    if candidate_id and candidate_id not in source_ids:
        risks.append("contact_id вне source_amo_contact_ids; нужен allow_contact_id_outside_source=yes")
    if len(dry_ids) > 1:
        risks.append("несколько exact-match контактов по телефону")
    if _safe_text(row.get("written_status")).casefold() == "written":
        risks.append("уже записано ранее; refresh только при явном refresh_approved")
    if bucket == "needs_text_quality_review":
        risks.append("текст CRM требует проверки до записи")
    return " | ".join(risks)


def _validation_formula(row: int) -> str:
    return (
        f'=IF(OR(F{row}="accepted_by_manager",F{row}="accepted_by_operator",F{row}="accepted_auto_policy"),'
        f'IF(G{row}="","ERROR: resolved_contact_id required",'
        f'IF(I{row}="","ERROR: resolution_reason required",'
        f'IF(J{row}="","ERROR: resolved_by required",'
        f'IF(AND(H{row}<>"yes",ISERROR(SEARCH(G{row},N{row}))),"ERROR: contact_id not in source; set outside_source approval",'
        f'IF(AND(D{row}="needs_manager_review_multi_contact",ISERROR(SEARCH("post_merge_recheck_approved",I{row}))),"ERROR: post_merge_recheck_approved required for duplicate merge row",'
        f'IF(AND(D{row}="needs_text_quality_review",ISERROR(SEARCH("text_quality_approved",I{row}))),"ERROR: text_quality_approved required",'
        f'IF(AND(AA{row}="written",ISERROR(SEARCH("refresh_approved",I{row}))),"ERROR: refresh_approved required for written row","OK: accepted fields filled"))))))),'
        f'"OK: not accepted / blocked")'
    )


def _apply_hyperlink(cell: Any, value: str) -> None:
    text = value.strip()
    if not text.startswith("http"):
        return
    first = text.splitlines()[0].strip()
    cell.hyperlink = first
    cell.style = "Hyperlink"


def _contact_url(base: str, contact_id: str) -> str:
    return f"{base}/contacts/detail/{contact_id}"


def _lead_url(base: str, lead_id: str) -> str:
    return f"{base}/leads/detail/{lead_id}"


def _split_ids(value: Any) -> list[str]:
    text = _safe_text(value)
    if not text:
        return []
    return [part.strip() for part in text.replace("|", ",").replace(";", ",").split(",") if part.strip()]


def _load_source_context(pack_root: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    summary = _read_json(pack_root / "summary.json")
    source_csv = _safe_text(_dig(summary, "summary", "source_csv"))
    if not source_csv:
        return {}, {}
    source_path = Path(source_csv).expanduser().resolve(strict=False)
    rows = _read_csv(source_path)
    by_index: dict[str, dict[str, str]] = {}
    by_phone: dict[str, dict[str, str]] = {}
    for index, row in enumerate(rows, start=1):
        by_index[str(index)] = row
        phone = _safe_text(row.get("Телефон клиента"))
        if phone:
            by_phone[phone] = row
    return by_index, by_phone


def _merge_source_context(target: dict[str, str], source: Mapping[str, str]) -> None:
    mapping = {
        "crm_last_call_summary": "Краткое резюме последнего свежего звонка",
        "crm_short_history": "Краткая история общения",
        "crm_timeline": "Хронология общения (последние 5 касаний)",
        "crm_objections": "Возражения",
        "crm_recommended_next_contact_date": "Рекомендуемая дата следующего контакта",
        "crm_ready_reason": "Причина статуса AMO",
    }
    for target_key, source_key in mapping.items():
        if not _safe_text(target.get(target_key)):
            target[target_key] = _safe_text(source.get(source_key))


def _render_review_html(rows: list[Mapping[str, str]], comparison_rows: list[Mapping[str, str]], now: datetime) -> str:
    candidates_by_resolution: dict[str, list[Mapping[str, str]]] = {}
    for candidate in comparison_rows:
        candidates_by_resolution.setdefault(_safe_text(candidate.get("resolution_id")), []).append(candidate)
    cards = "\n".join(_render_review_card(row, candidates_by_resolution.get(_safe_text(row.get("resolution_id")), [])) for row in rows)
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <title>AMO manual-resolution review</title>
  <style>
    :root {{
      --bg: #f8fafc;
      --card: #ffffff;
      --ink: #101828;
      --muted: #667085;
      --line: #d0d5dd;
      --accent: #0f766e;
      --warn: #92400e;
      --warn-bg: #fffbeb;
      --ok-bg: #ecfdf3;
    }}
    body {{ margin: 0; padding: 28px; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; background: var(--bg); color: var(--ink); }}
    h1 {{ margin: 0 0 8px; font-size: 28px; }}
    .meta {{ color: var(--muted); margin-bottom: 24px; }}
    .card {{ background: var(--card); border: 1px solid var(--line); border-radius: 16px; padding: 18px; margin: 0 0 18px; box-shadow: 0 8px 24px rgba(16, 24, 40, .05); }}
    .top {{ display: flex; justify-content: space-between; gap: 16px; align-items: flex-start; }}
    .phone {{ font-size: 20px; font-weight: 700; }}
    .bucket {{ color: var(--accent); font-weight: 700; }}
    .grid {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 12px; margin-top: 14px; }}
    .box {{ background: #f9fafb; border: 1px solid #eaecf0; border-radius: 12px; padding: 10px; }}
    .label {{ color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .04em; margin-bottom: 4px; }}
    .text {{ white-space: pre-wrap; line-height: 1.35; }}
    .candidates {{ width: 100%; border-collapse: collapse; margin-top: 12px; }}
    .candidates th, .candidates td {{ border: 1px solid var(--line); padding: 8px; vertical-align: top; text-align: left; }}
    .candidates th {{ background: #0f766e; color: white; }}
    .risk {{ color: var(--warn); background: var(--warn-bg); }}
    .suggested {{ background: var(--ok-bg); }}
    a {{ color: #0b5cab; }}
    @media (max-width: 900px) {{ .grid {{ grid-template-columns: 1fr; }} body {{ padding: 14px; }} }}
  </style>
</head>
<body>
  <h1>AMO manual-resolution review</h1>
  <div class="meta">Сгенерировано: {_escape(now.isoformat(timespec="seconds"))}. Только просмотр: страница ничего не пишет в AMO.</div>
  {cards}
</body>
</html>
"""


def _render_review_card(row: Mapping[str, str], candidates: list[Mapping[str, str]]) -> str:
    candidate_rows = "\n".join(_render_candidate_html(candidate) for candidate in candidates)
    return f"""<section class="card">
  <div class="top">
    <div>
      <div class="phone">{_escape(row.get("phone"))}</div>
      <div class="bucket">{_escape(row.get("queue_bucket"))}</div>
    </div>
    <div>{_escape(row.get("resolution_id"))}</div>
  </div>
  <div class="grid">
    <div class="box"><div class="label">ФИО</div><div class="text">{_escape(_join_non_empty([row.get("fio_parent"), row.get("fio_child")], " / "))}</div></div>
    <div class="box"><div class="label">Следующий шаг</div><div class="text">{_escape(row.get("next_step"))}</div></div>
    <div class="box"><div class="label">Приоритет / вероятность</div><div class="text">{_escape(row.get("priority"))} / {_escape(row.get("sale_probability_percent"))}%</div></div>
    <div class="box"><div class="label">Рекомендация</div><div class="text">{_escape(_recommended_action(row))}</div></div>
    <div class="box"><div class="label">Возражения</div><div class="text">{_escape(row.get("crm_objections"))}</div></div>
    <div class="box"><div class="label">Последний звонок</div><div class="text">{_escape(row.get("latest_call_date"))} · {_escape(row.get("latest_call_type"))}</div></div>
  </div>
  <div class="box" style="margin-top:12px"><div class="label">Краткая история</div><div class="text">{_escape(row.get("crm_short_history"))}</div></div>
  <table class="candidates">
    <thead><tr><th>contact_id</th><th>Источник</th><th>Ссылка</th><th>Подсказка</th><th>Риск</th></tr></thead>
    <tbody>{candidate_rows}</tbody>
  </table>
</section>"""


def _render_candidate_html(candidate: Mapping[str, str]) -> str:
    link = _safe_text(candidate.get("contact_link"))
    link_html = f'<a href="{_escape(link)}">{_escape(link)}</a>' if link else ""
    cls = "suggested" if candidate.get("is_suggested") == "yes" else ""
    risk_cls = "risk" if candidate.get("risk_note_ru") else ""
    return (
        f'<tr class="{cls}">'
        f'<td>{_escape(candidate.get("candidate_contact_id"))}</td>'
        f'<td>{_escape(candidate.get("candidate_source"))}</td>'
        f"<td>{link_html}</td>"
        f'<td>{_escape(candidate.get("operator_hint_ru"))}</td>'
        f'<td class="{risk_cls}">{_escape(candidate.get("risk_note_ru"))}</td>'
        "</tr>"
    )


def _join_non_empty(values: list[Any], sep: str) -> str:
    return sep.join(_safe_text(value) for value in values if _safe_text(value))


def _escape(value: Any) -> str:
    return html.escape(_safe_text(value))


def _read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return [dict(row) for row in csv.DictReader(fh)]


def _read_json(path: Path) -> Mapping[str, Any]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return data if isinstance(data, Mapping) else {}


def _dig(value: Mapping[str, Any], *keys: str) -> Any:
    current: Any = value
    for key in keys:
        if not isinstance(current, Mapping):
            return ""
        current = current.get(key)
    return current


def _safe_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _font(*, size: int = 11, bold: bool = False, color: str = "000000") -> Any:
    from openpyxl.styles import Font

    return Font(size=size, bold=bold, color=color)


def _alignment(**kwargs: Any) -> Any:
    from openpyxl.styles import Alignment

    return Alignment(**kwargs)


def _excel_col(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


__all__ = [
    "WORKBOOK_SCHEMA_VERSION",
    "REVIEW_ASSETS_SCHEMA_VERSION",
    "build_amo_resolution_review_html",
    "build_amo_resolution_review_workbook",
    "export_decisions_from_amo_resolution_workbook",
]
