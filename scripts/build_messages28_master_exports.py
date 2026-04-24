from __future__ import annotations

import argparse
import csv
import json
import sqlite3
import urllib.parse
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from typing import Any

from mango_mvp.services.export_excel import call_to_row

try:
    import pandas as pd
except ImportError:  # pragma: no cover
    pd = None

try:
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    Alignment = None
    get_column_letter = None


MASTER_CALLS_HEADERS = [
    "ID звонка",
    "Дата и время звонка",
    "Телефон клиента",
    "Менеджер",
    "Направление звонка",
    "Длительность, сек",
    "Имя исходного файла",
    "Путь к записи",
    "Свежий период",
    "Основной ASR готов",
    "Второй ASR (GigaAM) готов",
    "Статус Resolve",
    "Статус Analyze",
    "Полная цепочка выполнена",
    "Нужна ручная проверка",
    "Причины ручной проверки",
    "Краткое резюме разговора",
    "Тип звонка",
    "ФИО родителя",
    "ФИО ребенка",
    "Email",
    "Класс",
    "Школа",
    "Продукты интереса",
    "Формат обучения",
    "Предметы интереса",
    "Целевые экзамены",
    "Рекомендуемый продукт",
    "Коммерческие ограничения",
    "Возражения",
    "Следующий шаг",
    "Срок следующего шага",
    "Приоритет лида",
    "Вероятность продажи, %",
    "Причина оценки вероятности",
    "Рекомендуемая дата следующего контакта",
    "Причина рекомендуемой даты",
    "Источник лучшего статуса",
]

MASTER_CONTACTS_HEADERS = [
    "Телефон клиента",
    "Всего звонков в истории",
    "Звонков с полным анализом",
    "Незакрытых звонков в истории",
    "Полная история проанализирована",
    "Первый звонок",
    "Последний звонок",
    "Свежих звонков за период",
    "Последний свежий звонок",
    "Последний свежий звонок проанализирован",
    "Менеджер последнего свежего звонка",
    "Краткое резюме последнего свежего звонка",
    "Тип последнего свежего звонка",
    "Краткая история общения",
    "Хронология общения (последние 5 касаний)",
    "ФИО родителя",
    "ФИО ребенка",
    "Email",
    "Продукты интереса",
    "Рекомендуемый продукт",
    "Возражения",
    "Следующий шаг",
    "Рекомендуемая дата следующего контакта",
    "Приоритет лида",
    "Вероятность продажи, %",
    "Нужна ручная проверка",
    "Статус матчинга Tallanto",
    "Количество кандидатов Tallanto",
    "ID Tallanto",
    "ФИО родителя Tallanto",
    "Контакт Tallanto",
    "Ответственный Tallanto",
    "Тип ученика Tallanto",
    "Филиал Tallanto",
    "Готово к записи в AMO",
    "Причина статуса AMO",
]

AMO_EXPORT_HEADERS = [
    "Телефон клиента",
    "ID Tallanto",
    "Статус матчинга Tallanto",
    "ФИО родителя",
    "ФИО ребенка",
    "Email",
    "Ответственный Tallanto",
    "Тип ученика Tallanto",
    "Филиал Tallanto",
    "Дата последнего свежего звонка",
    "Менеджер последнего свежего звонка",
    "Краткое резюме последнего свежего звонка",
    "Тип последнего свежего звонка",
    "Краткая история общения",
    "Хронология общения (последние 5 касаний)",
    "Продукты интереса",
    "Рекомендуемый продукт",
    "Возражения",
    "Следующий шаг",
    "Рекомендуемая дата следующего контакта",
    "Приоритет лида",
    "Вероятность продажи, %",
    "История общения Tallanto",
    "Готово к записи в AMO",
    "Причина статуса AMO",
]


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build master calls / contacts / AMO export for the current sales history universe.")
    parser.add_argument("--history-db", help="Legacy single base DB path.")
    parser.add_argument("--base-db", action="append", default=[], help="Base DB path. May be repeated.")
    parser.add_argument("--tallanto-csv", required=True)
    parser.add_argument("--out-root", required=True)
    parser.add_argument("--overlay-db", action="append", default=[])
    parser.add_argument("--fresh-from", default="2026-03-19")
    parser.add_argument("--fresh-to", default="2026-04-07")
    parser.add_argument("--analysis-date", default=date.today().isoformat())
    return parser.parse_args()


def _connect_ro(path: Path) -> sqlite3.Connection:
    uri = "file:" + urllib.parse.quote(str(path)) + "?immutable=1"
    conn = sqlite3.connect(uri, uri=True)
    conn.row_factory = sqlite3.Row
    return conn


def _rank_resolve(status: str | None) -> int:
    return {"done": 3, "skipped": 3, "manual": 3, "in_progress": 2, "pending": 1}.get(status or "", 0)


def _rank_analysis(status: str | None) -> int:
    return {"done": 3, "in_progress": 2, "pending": 1}.get(status or "", 0)


def _parse_dt(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%d.%m.%Y %H:%M",
        "%d.%m.%Y",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _clean(value: Any) -> str:
    return str(value or "").strip()


def _has_variant_b(variants_json: str | None) -> bool:
    if not variants_json or not str(variants_json).strip():
        return False
    try:
        payload = json.loads(variants_json)
    except Exception:
        return False
    for section in ("manager", "client", "full"):
        part = payload.get(section)
        if isinstance(part, dict) and str(part.get("variant_b") or "").strip():
            return True
    return False


def _normalize_phone(value: Any) -> str:
    text = _clean(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10:
        digits = "7" + digits
    if len(digits) == 11 and digits.startswith("7"):
        return f"+{digits}"
    return text if text.startswith("+") else ""


def _load_rows(db_path: Path) -> dict[str, dict[str, Any]]:
    conn = _connect_ro(db_path)
    rows = {}
    for row in conn.execute(
        """
        SELECT id, source_file, source_filename, source_call_id, duration_sec, phone, manager_name,
               direction, started_at, transcription_status, resolve_status, analysis_status,
               transcript_manager, transcript_client, transcript_text, transcript_variants_json,
               resolve_json, analysis_json, last_error, updated_at
          FROM call_records
        """
    ):
        rows[str(row["source_filename"])] = dict(row)
    conn.close()
    return rows


def _merge_rows(db_paths: list[Path]) -> tuple[dict[str, dict[str, Any]], dict[str, str]]:
    merged_rows: dict[str, dict[str, Any]] = {}
    merged_sources: dict[str, str] = {}
    for db_path in db_paths:
        if not db_path.exists():
            continue
        rows = _load_rows(db_path)
        for source_filename, row in rows.items():
            if source_filename not in merged_rows or _record_rank(row) > _record_rank(merged_rows[source_filename]):
                merged_rows[source_filename] = dict(row)
                merged_sources[source_filename] = str(db_path)
    return merged_rows, merged_sources


def _record_rank(record: dict[str, Any]) -> tuple[int, int, int, int, int, datetime]:
    analysis_json = _clean(record.get("analysis_json"))
    resolve_json = _clean(record.get("resolve_json"))
    return (
        _rank_analysis(_clean(record.get("analysis_status")) or None),
        _rank_resolve(_clean(record.get("resolve_status")) or None),
        1 if analysis_json else 0,
        1 if resolve_json else 0,
        1 if _has_variant_b(record.get("transcript_variants_json")) else 0,
        _parse_dt(record.get("updated_at")) or datetime.min,
    )


def _safe_bool_text(flag: bool) -> str:
    return "Да" if flag else "Нет"


def _safe_int_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    try:
        return str(int(value))
    except Exception:
        return str(value)


def _write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in headers})


def _write_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]], sheet_name: str) -> None:
    if pd is None:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    frame = pd.DataFrame(rows, columns=headers).fillna("")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        if Alignment is not None:
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        if get_column_letter is not None:
            for idx, column in enumerate(frame.columns, start=1):
                values = [str(column)] + [str(value)[:200] for value in frame[column].astype(str)]
                width = min(max(len(value) for value in values) + 2, 80)
                worksheet.column_dimensions[get_column_letter(idx)].width = max(width, 12)


def _amo_status(contact_row: dict[str, Any]) -> tuple[str, str]:
    if contact_row["Последний свежий звонок проанализирован"] != "Да":
        return ("Нет", "последний свежий звонок еще не проанализирован")
    if contact_row["Нужна ручная проверка"] == "Да":
        return ("Нет", "нужна ручная проверка AI-анализа")
    if _clean(contact_row["Тип последнего свежего звонка"]).lower() == "non_conversation":
        return ("Нет", "последний свежий звонок не содержит содержательного диалога")
    if contact_row["Статус матчинга Tallanto"] != "exact_phone_single":
        return ("Нет", "нет однозначного exact phone match с Tallanto")
    return ("Да", "готово к первому проходу записи в AMO")


_ANALYZE_SERVICE = None


def _needs_history_summary_repair(analysis: dict[str, Any]) -> bool:
    if not isinstance(analysis, dict):
        return False
    value = _clean(analysis.get("history_summary"))
    if not value:
        return False
    return len(value) >= 300 and value.endswith(("...", "…"))


def _repair_analysis_for_export(call: Any, transcript_text: str, analysis: dict[str, Any]) -> dict[str, Any]:
    global _ANALYZE_SERVICE
    if not _needs_history_summary_repair(analysis):
        return analysis
    if not _clean(transcript_text):
        return analysis
    repair_input = dict(analysis)
    for key in ("history_summary", "history_short"):
        if _needs_history_summary_repair({key: repair_input.get(key)}):
            repair_input[key] = ""
    if _ANALYZE_SERVICE is None:
        from mango_mvp.config import get_settings
        from mango_mvp.services.analyze import AnalyzeService

        _ANALYZE_SERVICE = AnalyzeService(get_settings())
    try:
        normalized = _ANALYZE_SERVICE._normalize_analysis(call, transcript_text, repair_input)
        return normalized if isinstance(normalized, dict) else analysis
    except Exception:
        return analysis


def _truncate(text: str, limit: int = 220) -> str:
    value = _clean(text)
    if len(value) <= limit:
        return value
    return value[: max(limit - 1, 0)].rstrip() + "…"


def _excel_safe_text(text: str, *, limit: int = 32000) -> str:
    value = _clean(text)
    if len(value) <= limit:
        return value
    suffix = " [обрезано по лимиту Excel]"
    return value[: max(limit - len(suffix), 0)].rstrip() + suffix


def _call_gist(row: dict[str, Any], *, with_step: bool, truncate: bool = True) -> str:
    summary = _clean(row.get("Краткое резюме разговора", ""))
    next_step = _clean(row.get("Следующий шаг", ""))
    objections = _clean(row.get("Возражения", ""))
    if truncate:
        summary = _truncate(summary, 180)
        next_step = _truncate(next_step, 90)
        objections = _truncate(objections, 90)
    if not summary and next_step:
        summary = f"Следующий шаг: {next_step}"
    if not summary and objections:
        summary = f"Возражение: {objections}"
    if with_step and summary and next_step and next_step.lower() not in summary.lower():
        summary = f"{summary} Следующий шаг: {next_step}"
    result = summary or _clean(row.get("Тип звонка"))
    return _truncate(result, 240) if truncate else result


def _unique_parts(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for raw in values:
        for part in str(raw or "").split("|"):
            value = _clean(part)
            key = value.casefold()
            if not value or key in seen:
                continue
            seen.add(key)
            result.append(value)
    return result


def _join_limited(values: list[str], *, limit: int = 4) -> str:
    items = _unique_parts(values)
    if not items:
        return ""
    if len(items) <= limit:
        return ", ".join(items)
    remaining = len(items) - limit
    return ", ".join(items[:limit]) + f" и еще {remaining}"


def _build_contact_summary(analyzed_desc: list[dict[str, Any]]) -> str:
    if not analyzed_desc:
        return ""

    latest = analyzed_desc[0]
    earliest = analyzed_desc[-1]
    latest_dt = _parse_dt(latest.get("Дата и время звонка"))
    earliest_dt = _parse_dt(earliest.get("Дата и время звонка"))
    period_start = earliest_dt.strftime("%d.%m.%Y") if earliest_dt else _clean(earliest.get("Дата и время звонка"))[:10]
    period_end = latest_dt.strftime("%d.%m.%Y") if latest_dt else _clean(latest.get("Дата и время звонка"))[:10]

    managers = _join_limited([row.get("Менеджер", "") for row in analyzed_desc], limit=5)
    call_types = Counter(_clean(row.get("Тип звонка")).lower() for row in analyzed_desc if _clean(row.get("Тип звонка")))
    dominant_call_type = call_types.most_common(1)[0][0] if call_types else ""
    products = _join_limited(
        [row.get("Рекомендуемый продукт", "") for row in analyzed_desc] +
        [row.get("Продукты интереса", "") for row in analyzed_desc],
        limit=5,
    )
    objections = _join_limited([row.get("Возражения", "") for row in analyzed_desc], limit=4)

    next_steps = [
        _clean(row.get("Следующий шаг"))
        for row in analyzed_desc
        if _clean(row.get("Следующий шаг"))
    ]
    latest_step = next_steps[0] if next_steps else ""
    latest_summary = _call_gist(latest, with_step=False, truncate=False)

    parts = [
        f"Контакт в истории с {period_start} по {period_end}. Проанализировано звонков: {len(analyzed_desc)}.",
    ]
    if managers:
        parts.append(f"Менеджеры: {managers}.")
    if dominant_call_type:
        parts.append(f"Основной тип общения: {dominant_call_type}.")
    if products:
        parts.append(f"Ключевой интерес: {products}.")
    if objections:
        parts.append(f"Повторяющиеся ограничения/возражения: {objections}.")
    if latest_summary:
        parts.append(f"Последний содержательный контекст: {latest_summary}")
    if latest_step:
        parts.append(f"Текущий согласованный следующий шаг: {latest_step}.")
    return " ".join(part.strip() for part in parts if part and part.strip())


def _history_line(row: dict[str, Any], *, with_step: bool) -> str:
    dt = _parse_dt(row.get("Дата и время звонка"))
    date_text = dt.strftime("%d.%m.%Y") if dt else _clean(row.get("Дата и время звонка"))[:10]
    manager = _clean(row.get("Менеджер"))
    call_type = _clean(row.get("Тип звонка"))
    prefix = date_text
    if manager:
        prefix += f" — {manager}"
    if call_type:
        prefix += f" ({call_type})"
    gist = _call_gist(row, with_step=with_step, truncate=False)
    return f"{prefix}: {gist}" if gist else prefix


def _build_contact_history(calls: list[dict[str, Any]]) -> tuple[str, str]:
    analyzed_desc = [
        row
        for row in sorted(
            calls,
            key=lambda item: (_parse_dt(item.get("Дата и время звонка")) or datetime.min, item.get("Имя исходного файла", "")),
            reverse=True,
        )
        if row.get("Статус Analyze") == "done"
    ]
    if not analyzed_desc:
        return ("", "")

    chronology_rows = list(reversed(analyzed_desc[:5]))
    chronology = "\n".join(_history_line(row, with_step=False) for row in chronology_rows)
    short_history = _build_contact_summary(analyzed_desc)
    return (_excel_safe_text(short_history), _excel_safe_text(chronology))


def _parse_probability_value(value: Any) -> int | None:
    text = _clean(value).replace("%", "")
    if not text:
        return None
    try:
        return max(0, min(100, int(round(float(text)))))
    except Exception:
        return None


def _format_probability_value(value: int | None, fallback: str) -> str:
    if value is None:
        return _clean(fallback)
    return str(max(0, min(100, value)))


def _priority_to_rank(value: str) -> int:
    normalized = _clean(value).lower()
    return {"cold": 1, "warm": 2, "hot": 3}.get(normalized, 0)


def _rank_to_priority(rank: int, fallback: str) -> str:
    if rank <= 0:
        return _clean(fallback)
    return {1: "cold", 2: "warm", 3: "hot"}.get(rank, _clean(fallback))


def _infer_priority_from_probability(probability: int | None) -> str:
    if probability is None:
        return ""
    if probability >= 75:
        return "hot"
    if probability >= 45:
        return "warm"
    return "cold"


def _adjust_contact_operational_fields(
    *,
    analysis_date: date,
    last_contact_raw: Any,
    follow_up_raw: Any,
    priority_raw: Any,
    probability_raw: Any,
) -> tuple[str, str, str]:
    last_contact_dt = _parse_dt(last_contact_raw)
    if last_contact_dt is None:
        return (_clean(follow_up_raw), _clean(priority_raw), _clean(probability_raw))

    days_since_last_contact = max(0, (analysis_date - last_contact_dt.date()).days)
    probability = _parse_probability_value(probability_raw)
    if probability is None:
        inferred_priority = _clean(priority_raw).lower() or _infer_priority_from_probability(probability)
        probability = {"hot": 80, "warm": 60, "cold": 30}.get(inferred_priority, None)

    if probability is not None:
        decay = 0
        if days_since_last_contact > 7:
            decay += 5
        if days_since_last_contact > 14:
            decay += 10
        if days_since_last_contact > 30:
            decay += 15
        if days_since_last_contact > 60:
            decay += 15
        probability = max(0, probability - decay)

    priority_rank = _priority_to_rank(_clean(priority_raw))
    if priority_rank == 0:
        priority_rank = _priority_to_rank(_infer_priority_from_probability(probability))
    if days_since_last_contact > 14 and priority_rank > 1:
        priority_rank -= 1
    if days_since_last_contact > 45 and priority_rank > 1:
        priority_rank -= 1
    adjusted_priority = _rank_to_priority(priority_rank, _clean(priority_raw) or _infer_priority_from_probability(probability))

    follow_up_dt = _parse_dt(follow_up_raw)
    adjusted_follow_up = None
    if follow_up_dt is not None:
        adjusted_follow_up = max(follow_up_dt.date(), analysis_date)
    elif adjusted_priority in {"hot", "warm"}:
        adjusted_follow_up = analysis_date if days_since_last_contact >= 1 else analysis_date + timedelta(days=1)

    adjusted_follow_up_text = adjusted_follow_up.isoformat() if adjusted_follow_up else _clean(follow_up_raw)
    return (
        adjusted_follow_up_text,
        adjusted_priority,
        _format_probability_value(probability, _clean(probability_raw)),
    )


def _read_tallanto_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        rows = []
        for row in reader:
            row = dict(row)
            row["phone_parent"] = _normalize_phone(row.get("phone_parent"))
            row["phone_extra"] = _normalize_phone(row.get("phone_extra"))
            rows.append(row)
        return rows


def main() -> int:
    args = _parse_args()
    analysis_date = (_parse_dt(args.analysis_date) or datetime.strptime(args.analysis_date, "%Y-%m-%d")).date()
    base_dbs = [Path(item).expanduser().resolve() for item in args.base_db]
    if args.history_db:
        base_dbs.append(Path(args.history_db).expanduser().resolve())
    if not base_dbs:
        raise SystemExit("Provide --base-db (may be repeated) or legacy --history-db.")
    overlay_dbs = [Path(item).expanduser().resolve() for item in args.overlay_db]
    tallanto_csv = Path(args.tallanto_csv).expanduser().resolve()
    out_root = Path(args.out_root).expanduser().resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    base_rows, base_sources = _merge_rows(base_dbs)
    best_rows = {source_filename: dict(row) for source_filename, row in base_rows.items()}
    best_sources = dict(base_sources)

    for db_path in overlay_dbs:
        if not db_path.exists():
            continue
        rows = _load_rows(db_path)
        for source_filename, row in rows.items():
            if source_filename not in best_rows:
                continue
            if _record_rank(row) > _record_rank(best_rows[source_filename]):
                best_rows[source_filename] = dict(row)
                best_sources[source_filename] = str(db_path)

    master_call_rows: list[dict[str, Any]] = []
    analyzed_flat_rows: list[dict[str, Any]] = []
    by_phone_calls: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for source_filename, base in sorted(
        base_rows.items(),
        key=lambda item: (_parse_dt(item[1].get("started_at")) or datetime.min, item[0]),
    ):
        best = best_rows[source_filename]
        phone = _normalize_phone(best.get("phone") or base.get("phone"))
        started = _clean(best.get("started_at") or base.get("started_at"))
        is_fresh = args.fresh_from <= started[:10] <= args.fresh_to
        has_primary_asr = _clean(base.get("transcription_status")) == "done"
        has_gigaam = _has_variant_b(base.get("transcript_variants_json")) or _has_variant_b(best.get("transcript_variants_json"))
        resolve_status = _clean(best.get("resolve_status")) or _clean(base.get("resolve_status")) or "pending"
        analysis_status = _clean(best.get("analysis_status")) or _clean(base.get("analysis_status")) or "pending"
        full_chain = analysis_status == "done"

        analysis_json = _clean(best.get("analysis_json"))
        flat = None
        if analysis_json:
            try:
                analysis = json.loads(analysis_json)
                call_obj = SimpleNamespace(
                    id=best.get("id"),
                    started_at=_parse_dt(best.get("started_at")),
                    phone=phone,
                    manager_name=best.get("manager_name"),
                    duration_sec=best.get("duration_sec"),
                    source_filename=best.get("source_filename"),
                    source_file=best.get("source_file"),
                    direction=best.get("direction"),
                    transcript_variants_json=best.get("transcript_variants_json"),
                )
                analysis = _repair_analysis_for_export(call_obj, _clean(best.get("transcript_text")), analysis)
                flat = call_to_row(call_obj, analysis)
            except Exception:
                flat = None

        if flat:
            analyzed_flat_rows.append(flat)

        call_row = {
            "ID звонка": _safe_int_text(best.get("id")),
            "Дата и время звонка": started,
            "Телефон клиента": phone,
            "Менеджер": _clean(best.get("manager_name") or base.get("manager_name")),
            "Направление звонка": _clean(best.get("direction") or base.get("direction")),
            "Длительность, сек": _clean(best.get("duration_sec") or base.get("duration_sec")),
            "Имя исходного файла": _clean(best.get("source_filename") or base.get("source_filename")),
            "Путь к записи": _clean(best.get("source_file") or base.get("source_file")),
            "Свежий период": _safe_bool_text(is_fresh),
            "Основной ASR готов": _safe_bool_text(has_primary_asr),
            "Второй ASR (GigaAM) готов": _safe_bool_text(has_gigaam),
            "Статус Resolve": resolve_status,
            "Статус Analyze": analysis_status,
            "Полная цепочка выполнена": _safe_bool_text(full_chain),
            "Нужна ручная проверка": _safe_bool_text(bool(flat and flat.get("needs_review"))),
            "Причины ручной проверки": _clean(flat.get("review_reasons")) if flat else "",
            "Краткое резюме разговора": _clean(flat.get("history_summary")) if flat else "",
            "Тип звонка": _clean(flat.get("call_type")) if flat else "",
            "ФИО родителя": _clean(flat.get("parent_fio")) if flat else "",
            "ФИО ребенка": _clean(flat.get("child_fio")) if flat else "",
            "Email": _clean(flat.get("email")) if flat else "",
            "Класс": _clean(flat.get("grade_current")) if flat else "",
            "Школа": _clean(flat.get("school")) if flat else "",
            "Продукты интереса": _clean(flat.get("interests_products")) if flat else "",
            "Формат обучения": _clean(flat.get("interests_format")) if flat else "",
            "Предметы интереса": _clean(flat.get("interests_subjects")) if flat else "",
            "Целевые экзамены": _clean(flat.get("exam_targets")) if flat else "",
            "Рекомендуемый продукт": _clean(flat.get("recommended_product")) if flat else "",
            "Коммерческие ограничения": " | ".join(
                part for part in [
                    _clean(flat.get("price_sensitivity")) if flat else "",
                    _clean(flat.get("budget")) if flat else "",
                    _clean(flat.get("discount_interest")) if flat else "",
                ]
                if part
            ),
            "Возражения": _clean(flat.get("objections")) if flat else "",
            "Следующий шаг": _clean(flat.get("next_step_action")) if flat else "",
            "Срок следующего шага": _clean(flat.get("next_step_due_raw")) if flat else "",
            "Приоритет лида": _clean(flat.get("lead_priority")) if flat else "",
            "Вероятность продажи, %": _clean(flat.get("sale_probability_pct")) if flat else "",
            "Причина оценки вероятности": _clean(flat.get("sale_probability_reason")) if flat else "",
            "Рекомендуемая дата следующего контакта": _clean(flat.get("recommended_followup_date")) if flat else "",
            "Причина рекомендуемой даты": _clean(flat.get("recommended_followup_reason")) if flat else "",
            "Источник лучшего статуса": best_sources[source_filename],
        }
        master_call_rows.append(call_row)
        by_phone_calls[phone].append(call_row)

    tallanto_rows = _read_tallanto_rows(tallanto_csv)
    tallanto_by_phone: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in tallanto_rows:
        for phone_field in ("phone_parent", "phone_extra"):
            phone = row.get(phone_field) or ""
            if phone:
                tallanto_by_phone[phone].append(row)

    contact_rows: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    amo_rows: list[dict[str, Any]] = []

    for phone, calls in sorted(by_phone_calls.items(), key=lambda item: item[0]):
        calls_sorted = sorted(calls, key=lambda row: (_parse_dt(row["Дата и время звонка"]) or datetime.min, row["Имя исходного файла"]), reverse=True)
        fresh_calls = [row for row in calls_sorted if row["Свежий период"] == "Да"]
        latest_fresh = fresh_calls[0] if fresh_calls else None
        latest_fresh_analyzed = next((row for row in fresh_calls if row["Статус Analyze"] == "done"), None)
        analyzed_calls = [row for row in calls_sorted if row["Статус Analyze"] == "done"]
        best_ai = latest_fresh_analyzed or (analyzed_calls[0] if analyzed_calls else calls_sorted[0])
        short_history, chronology = _build_contact_history(calls_sorted)

        candidates = list(tallanto_by_phone.get(phone, []))
        candidates.sort(
            key=lambda row: (
                _parse_dt(row.get("updated_at")) or datetime.min,
                1 if _clean(row.get("history_raw")) else 0,
                1 if _clean(row.get("email")) else 0,
            ),
            reverse=True,
        )
        match_status = "no_exact_phone_match"
        matched = None
        if len(candidates) == 1:
            match_status = "exact_phone_single"
            matched = candidates[0]
        elif len(candidates) > 1:
            match_status = "exact_phone_multiple"
            matched = candidates[0]
            for rank, cand in enumerate(candidates, start=1):
                review_rows.append(
                    {
                        "Телефон клиента": phone,
                        "Кандидат №": rank,
                        "ID Tallanto": _clean(cand.get("tallanto_id")),
                        "ФИО родителя Tallanto": _clean(cand.get("parent_fio")),
                        "Контакт Tallanto": _clean(cand.get("contact_full_name")),
                        "Ответственный Tallanto": _clean(cand.get("responsible")),
                        "Тип ученика Tallanto": _clean(cand.get("student_type")),
                        "Филиал Tallanto": _clean(cand.get("branch")),
                        "Дата обновления Tallanto": _clean(cand.get("updated_at")),
                        "История общения Tallanto": _clean(cand.get("history_raw")),
                    }
                )

        unresolved_history = sum(1 for row in calls if row["Статус Analyze"] != "done")
        base_follow_up = latest_fresh_analyzed["Рекомендуемая дата следующего контакта"] if latest_fresh_analyzed else best_ai["Рекомендуемая дата следующего контакта"]
        base_priority = latest_fresh_analyzed["Приоритет лида"] if latest_fresh_analyzed else best_ai["Приоритет лида"]
        base_probability = latest_fresh_analyzed["Вероятность продажи, %"] if latest_fresh_analyzed else best_ai["Вероятность продажи, %"]
        adjusted_follow_up, adjusted_priority, adjusted_probability = _adjust_contact_operational_fields(
            analysis_date=analysis_date,
            last_contact_raw=max((row["Дата и время звонка"] for row in calls if row["Дата и время звонка"]), default=""),
            follow_up_raw=base_follow_up,
            priority_raw=base_priority,
            probability_raw=base_probability,
        )
        contact_row = {
            "Телефон клиента": phone,
            "Всего звонков в истории": len(calls),
            "Звонков с полным анализом": sum(1 for row in calls if row["Статус Analyze"] == "done"),
            "Незакрытых звонков в истории": unresolved_history,
            "Полная история проанализирована": _safe_bool_text(unresolved_history == 0),
            "Первый звонок": min((row["Дата и время звонка"] for row in calls if row["Дата и время звонка"]), default=""),
            "Последний звонок": max((row["Дата и время звонка"] for row in calls if row["Дата и время звонка"]), default=""),
            "Свежих звонков за период": len(fresh_calls),
            "Последний свежий звонок": latest_fresh["Дата и время звонка"] if latest_fresh else "",
            "Последний свежий звонок проанализирован": _safe_bool_text(bool(latest_fresh and latest_fresh["Статус Analyze"] == "done")),
            "Менеджер последнего свежего звонка": latest_fresh["Менеджер"] if latest_fresh else "",
            "Краткое резюме последнего свежего звонка": latest_fresh_analyzed["Краткое резюме разговора"] if latest_fresh_analyzed else "",
            "Тип последнего свежего звонка": latest_fresh_analyzed["Тип звонка"] if latest_fresh_analyzed else "",
            "Краткая история общения": short_history,
            "Хронология общения (последние 5 касаний)": chronology,
            "ФИО родителя": best_ai["ФИО родителя"],
            "ФИО ребенка": best_ai["ФИО ребенка"],
            "Email": best_ai["Email"],
            "Продукты интереса": best_ai["Продукты интереса"],
            "Рекомендуемый продукт": best_ai["Рекомендуемый продукт"],
            "Возражения": latest_fresh_analyzed["Возражения"] if latest_fresh_analyzed else best_ai["Возражения"],
            "Следующий шаг": latest_fresh_analyzed["Следующий шаг"] if latest_fresh_analyzed else best_ai["Следующий шаг"],
            "Рекомендуемая дата следующего контакта": adjusted_follow_up,
            "Приоритет лида": adjusted_priority,
            "Вероятность продажи, %": adjusted_probability,
            "Нужна ручная проверка": _safe_bool_text(any(row["Нужна ручная проверка"] == "Да" for row in calls)),
            "Статус матчинга Tallanto": match_status,
            "Количество кандидатов Tallanto": len(candidates),
            "ID Tallanto": _clean(matched.get("tallanto_id")) if matched else "",
            "ФИО родителя Tallanto": _clean(matched.get("parent_fio")) if matched else "",
            "Контакт Tallanto": _clean(matched.get("contact_full_name")) if matched else "",
            "Ответственный Tallanto": _clean(matched.get("responsible")) if matched else "",
            "Тип ученика Tallanto": _clean(matched.get("student_type")) if matched else "",
            "Филиал Tallanto": _clean(matched.get("branch")) if matched else "",
            "Готово к записи в AMO": "",
            "Причина статуса AMO": "",
        }
        amo_ready, amo_reason = _amo_status(contact_row)
        contact_row["Готово к записи в AMO"] = amo_ready
        contact_row["Причина статуса AMO"] = amo_reason
        contact_rows.append(contact_row)

        if amo_ready == "Да":
            amo_rows.append(
                {
                    "Телефон клиента": phone,
                    "ID Tallanto": contact_row["ID Tallanto"],
                    "Статус матчинга Tallanto": contact_row["Статус матчинга Tallanto"],
                    "ФИО родителя": contact_row["ФИО родителя"],
                    "ФИО ребенка": contact_row["ФИО ребенка"],
                    "Email": contact_row["Email"],
                    "Ответственный Tallanto": contact_row["Ответственный Tallanto"],
                    "Тип ученика Tallanto": contact_row["Тип ученика Tallanto"],
                    "Филиал Tallanto": contact_row["Филиал Tallanto"],
                    "Дата последнего свежего звонка": contact_row["Последний свежий звонок"],
                    "Менеджер последнего свежего звонка": contact_row["Менеджер последнего свежего звонка"],
                    "Краткое резюме последнего свежего звонка": contact_row["Краткое резюме последнего свежего звонка"],
                    "Тип последнего свежего звонка": contact_row["Тип последнего свежего звонка"],
                    "Краткая история общения": contact_row["Краткая история общения"],
                    "Хронология общения (последние 5 касаний)": contact_row["Хронология общения (последние 5 касаний)"],
                    "Продукты интереса": contact_row["Продукты интереса"],
                    "Рекомендуемый продукт": contact_row["Рекомендуемый продукт"],
                    "Возражения": contact_row["Возражения"],
                    "Следующий шаг": contact_row["Следующий шаг"],
                    "Рекомендуемая дата следующего контакта": contact_row["Рекомендуемая дата следующего контакта"],
                    "Приоритет лида": contact_row["Приоритет лида"],
                    "Вероятность продажи, %": contact_row["Вероятность продажи, %"],
                    "История общения Tallanto": _clean(matched.get("history_raw")) if matched else "",
                    "Готово к записи в AMO": amo_ready,
                    "Причина статуса AMO": amo_reason,
                }
            )

    master_calls_csv = out_root / "master_calls_ru.csv"
    master_contacts_csv = out_root / "master_contacts_ru.csv"
    amo_ready_csv = out_root / "amo_export_ready_ru.csv"
    tallanto_review_csv = out_root / "tallanto_review_candidates_ru.csv"
    master_calls_xlsx = out_root / "master_calls_ru.xlsx"
    master_contacts_xlsx = out_root / "master_contacts_ru.xlsx"
    amo_ready_xlsx = out_root / "amo_export_ready_ru.xlsx"
    tallanto_review_xlsx = out_root / "tallanto_review_candidates_ru.xlsx"

    _write_csv(master_calls_csv, MASTER_CALLS_HEADERS, master_call_rows)
    _write_csv(master_contacts_csv, MASTER_CONTACTS_HEADERS, contact_rows)
    _write_csv(amo_ready_csv, AMO_EXPORT_HEADERS, amo_rows)
    tallanto_review_headers = list(review_rows[0].keys()) if review_rows else [
        "Телефон клиента", "Кандидат №", "ID Tallanto", "ФИО родителя Tallanto",
        "Контакт Tallanto", "Ответственный Tallanto", "Тип ученика Tallanto",
        "Филиал Tallanto", "Дата обновления Tallanto", "История общения Tallanto"
    ]
    _write_csv(tallanto_review_csv, tallanto_review_headers, review_rows)

    if pd is not None:
        _write_xlsx(master_calls_xlsx, MASTER_CALLS_HEADERS, master_call_rows, "Звонки")
        _write_xlsx(master_contacts_xlsx, MASTER_CONTACTS_HEADERS, contact_rows, "Контакты")
        _write_xlsx(amo_ready_xlsx, AMO_EXPORT_HEADERS, amo_rows, "AMO_Export")
        _write_xlsx(tallanto_review_xlsx, tallanto_review_headers, review_rows, "Tallanto_Review")

        workbook_path = out_root / "master_export_pack_ru.xlsx"
        with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
            pd.DataFrame(master_call_rows, columns=MASTER_CALLS_HEADERS).to_excel(writer, sheet_name="Звонки", index=False)
            pd.DataFrame(contact_rows, columns=MASTER_CONTACTS_HEADERS).to_excel(writer, sheet_name="Контакты", index=False)
            pd.DataFrame(amo_rows, columns=AMO_EXPORT_HEADERS).to_excel(writer, sheet_name="AMO_Export", index=False)
            pd.DataFrame(review_rows).to_excel(writer, sheet_name="Tallanto_Review", index=False)

    fresh_contacts = sum(1 for row in contact_rows if row["Свежих звонков за период"] > 0)
    latest_fresh_analyzed = sum(1 for row in contact_rows if row["Последний свежий звонок проанализирован"] == "Да")
    summary = {
        "base_dbs": [str(path) for path in base_dbs],
        "overlay_dbs": [str(path) for path in overlay_dbs],
        "tallanto_csv": str(tallanto_csv),
        "analysis_date": analysis_date.isoformat(),
        "master_calls_rows": len(master_call_rows),
        "master_contacts_rows": len(contact_rows),
        "amo_export_ready_rows": len(amo_rows),
        "fresh_contacts": fresh_contacts,
        "fresh_contacts_latest_call_analyzed": latest_fresh_analyzed,
        "full_chain_done_calls": sum(1 for row in master_call_rows if row["Полная цепочка выполнена"] == "Да"),
        "asr_only_calls": sum(1 for row in master_call_rows if row["Статус Resolve"] == "pending" and row["Статус Analyze"] == "pending"),
        "resolve_done_analyze_pending_calls": sum(1 for row in master_call_rows if row["Статус Resolve"] in {"done", "skipped", "manual"} and row["Статус Analyze"] != "done"),
        "tallanto_match_status_counts": dict(Counter(row["Статус матчинга Tallanto"] for row in contact_rows)),
        "amo_ready_count": sum(1 for row in contact_rows if row["Готово к записи в AMO"] == "Да"),
        "output_files": {
            "master_calls_csv": str(master_calls_csv),
            "master_contacts_csv": str(master_contacts_csv),
            "amo_export_ready_csv": str(amo_ready_csv),
            "tallanto_review_csv": str(tallanto_review_csv),
        },
    }
    if pd is not None:
        summary["output_files"]["master_calls_xlsx"] = str(master_calls_xlsx)
        summary["output_files"]["master_contacts_xlsx"] = str(master_contacts_xlsx)
        summary["output_files"]["amo_export_ready_xlsx"] = str(amo_ready_xlsx)
        summary["output_files"]["tallanto_review_xlsx"] = str(tallanto_review_xlsx)
        summary["output_files"]["workbook_xlsx"] = str(out_root / "master_export_pack_ru.xlsx")

    (out_root / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
