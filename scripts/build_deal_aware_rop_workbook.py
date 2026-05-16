#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


RENAME_COLUMNS = {
    "sample_id": "ID примера",
    "sample_bucket_ru": "Класс риска",
    "review_url_hint": "Ссылка на сделку AMO",
    "selected_deal_id": "ID сделки AMO",
    "selected_deal_name": "Название сделки",
    "selected_pipeline_name": "Воронка",
    "selected_status_name": "Статус сделки",
    "selected_loss_reason": "Причина отказа",
    "phones": "Телефоны",
    "managers": "Менеджеры",
    "candidate_call_count": "Кол-во звонков",
    "candidate_phone_count": "Кол-во телефонов",
    "tallanto_context_status": "Статус Tallanto",
    "AI-фактический статус сделки": "AI: фактический статус",
    "AI-приоритет сделки": "AI: приоритет",
    "AI-рекомендованный следующий шаг": "AI: следующий шаг",
    "AI-сводка по сделке": "AI: сводка по сделке",
    "AI-история по сделке": "AI: история по сделке",
    "AI-Tallanto статус по сделке": "AI: Tallanto статус",
    "AI-предупреждение по сделке": "AI: предупреждение",
    "risk_flags_ru": "Риски простым языком",
    "risk_classes": "Технические классы риска",
    "stage6_status": "Stage6 статус",
    "stage6_reason": "Stage6 причина",
    "stage6_finding_types": "Stage6 находки",
    "rop_row_decision": "Решение РОПа",
    "summary_correctness": "Сводка верна?",
    "next_step_quality": "Следующий шаг верен?",
    "deal_status_priority_quality": "Статус/приоритет верны?",
    "tallanto_block_quality": "Tallanto-блок верен?",
    "history_relevance": "История релевантна?",
    "sales_usefulness": "Полезность для продаж",
    "issue_type": "Тип проблемы",
    "severity": "Серьёзность",
    "comment": "Комментарий РОПа",
}

VALIDATIONS = {
    "Решение РОПа": "ready_for_rop,minor_comment,needs_fix_before_rop,block",
    "Сводка верна?": "ok,minor_inaccuracy,wrong_or_misleading,cannot_judge",
    "Следующий шаг верен?": "ok_manager_action,too_passive,customer_side_only,wrong_action,missing",
    "Статус/приоритет верны?": "ok,status_wrong,priority_wrong,amo_tallanto_mismatch_not_handled,cannot_judge",
    "Tallanto-блок верен?": "ok,wrong_student,wrong_finance_or_attendance,too_raw_or_unreadable,not_relevant",
    "История релевантна?": "ok,irrelevant_calls,important_call_missing,too_verbose,cannot_judge",
    "Полезность для продаж": "useful_as_is,useful_after_small_edit,not_useful,dangerous",
    "Тип проблемы": "wrong_deal_binding,wrong_customer_context,wrong_next_step,payment_conflict,service_feedback_routing,amo_tallanto_mismatch,bad_tenant_terms,too_verbose,duplicate_fields,other",
    "Серьёзность": "P0_blocker,P1_fix_before_rop,P2_minor,P3_note",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build human-readable ROP workbook for deal-aware preview.")
    parser.add_argument("--review-root", required=True)
    parser.add_argument("--out", default="")
    parser.add_argument("--note", default="")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = Path(args.review_root).resolve()
    out = Path(args.out).resolve() if args.out else root / "deal_aware_stage100_rop_review.xlsx"

    df = pd.read_csv(root / "stratified_preview_100_for_rop.csv")
    allowed_cols = [column for column in df.columns if column.endswith("_allowed_values")]
    main = df[[column for column in df.columns if column not in allowed_cols]].copy().rename(columns=RENAME_COLUMNS)
    rubric = pd.read_csv(root / "rop_rubric.csv").rename(columns={"field": "Поле", "allowed_values": "Допустимые значения"})
    summary = pd.DataFrame(
        [
            ["Всего строк в Stage6", 709],
            ["Dry-run кандидатов", 680],
            ["Заблокировано", 29],
            ["Строк в этой проверке", len(df)],
            ["Live-запись разрешена", "Нет"],
            ["Примечание", args.note or "ROP review workbook; live-write blocked."],
        ],
        columns=["Показатель", "Значение"],
    )

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Сводка")
        main.to_excel(writer, index=False, sheet_name="Проверка РОП")
        rubric.to_excel(writer, index=False, sheet_name="Рубрика")

    style_workbook(out)
    print(out)
    return 0


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    light_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D0D7DE")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.auto_filter.ref = ws.dimensions

    ws = wb["Проверка РОП"]
    for index in range(1, ws.max_column + 1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = 22
    for letter, width in {"A": 12, "B": 42, "C": 26, "D": 13, "E": 28, "P": 34, "Q": 54, "R": 62, "S": 42, "T": 36, "U": 44, "AI": 40}.items():
        ws.column_dimensions[letter].width = width
    for row_index in range(2, ws.max_row + 1):
        ws.row_dimensions[row_index].height = 96
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    for name, values in VALIDATIONS.items():
        if name not in headers:
            continue
        validation = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{headers[name]}2:{headers[name]}{ws.max_row}")

    ws = wb["Сводка"]
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 64
    for row in ws.iter_rows(min_row=2):
        row[0].fill = light_fill
    wb.save(path)


if __name__ == "__main__":
    raise SystemExit(main())
