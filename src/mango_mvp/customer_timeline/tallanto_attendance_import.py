from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping
from zoneinfo import ZoneInfo

from dotenv import dotenv_values

from mango_mvp.customer_timeline.contracts import (
    IdentityMatchClass,
    TimelineDirection,
    TimelineEvent,
    TimelineEventType,
    IdentityLink,
    IdentityLinkType,
)
from mango_mvp.customer_timeline.ids import stable_digest
from mango_mvp.customer_timeline.safety import guard_customer_timeline_output_path, is_customer_timeline_prod_path
from mango_mvp.customer_timeline.store import (
    CustomerTimelineSQLiteStore,
    authoritative_exact_identity_rows,
    customer_timeline_readonly_uri,
)
from mango_mvp.deal_aware.stage1_snapshot import read_writeoff_xlsx


SOURCE_SYSTEM = "tallanto_attendance"
API_SOURCE_SYSTEM = "tallanto_attendance_api"
# D4: severity for each unresolved reason recorded via record_conflict.
# identity_conflict is a real contradiction (two different existing
# customers); identity_infrastructure_gap means this run's own Contact fetch
# came back incomplete; identity_unmatched_expected is normal steady-state
# noise (nothing links this contact anywhere yet) and never blocks freshness
# on its own -- see run_tallanto_attendance_api_increment.
_UNRESOLVED_SEVERITY_BY_REASON: dict[str, str] = {
    "identity_conflict": "high",
    "identity_infrastructure_gap": "medium",
    "identity_unmatched_expected": "low",
}
_STAGING_COMPONENTS = (".codex_local", "staging")
_MOSCOW = ZoneInfo("Europe/Moscow")
_RELATIONSHIP_FIELDS = (
    "id",
    "most_class_id",
    "contact_id",
    "most_class_contacts_status",
    "most_class_abonements",
    "date_modified",
    "date_entry",
    "most_class_contacts_duration",
    "most_class_contacts_cost",
    "most_class_discount",
    "most_class_discount_type",
)
_CLASS_FIELDS = ("id", "name", "date_start", "date_finish", "status", "cource_id", "filial")


@dataclass(frozen=True)
class TallantoAttendanceImportConfig:
    timeline_db: Path
    allowed_root: Path
    contacts_workbook: Path
    attendance_report: Path
    tenant_id: str = "foton"
    apply: bool = False
    actor: str = "tallanto_attendance_import"
    tallanto_env_file: Path | None = None


@dataclass(frozen=True)
class TallantoAttendanceApiIncrementConfig:
    timeline_db: Path
    allowed_root: Path
    tallanto_env_file: Path
    initial_since: datetime
    tenant_id: str = "foton"
    apply: bool = True
    actor: str = "tallanto_attendance_api_increment"


def run_tallanto_attendance_api_increment(
    config: TallantoAttendanceApiIncrementConfig,
    *,
    client: Any | None = None,
    now: datetime | None = None,
) -> Mapping[str, Any]:
    """Import reliable Tallanto attendance and hold the cursor while identities remain unresolved."""
    root = Path(config.allowed_root).expanduser().resolve(strict=False)
    db = guard_customer_timeline_output_path(Path(config.timeline_db), root)
    if config.apply and (is_customer_timeline_prod_path(db) or not _is_staging_db(db)):
        raise ValueError("Tallanto attendance API apply requires a staging Customer Timeline DB")
    if config.initial_since.tzinfo is None:
        raise ValueError("Tallanto attendance initial_since must be timezone-aware")
    api = client or _build_tallanto_client(Path(config.tallanto_env_file).expanduser())
    run_started = (now or datetime.now(timezone.utc)).astimezone(_MOSCOW)
    cursor_before, cursor_metadata = _load_api_cursor(
        db,
        tenant_id=config.tenant_id,
        initial_since=config.initial_since,
    )
    upper_bound = _fetch_relationship_upper_bound(api) or cursor_before
    if upper_bound < cursor_before:
        upper_bound = cursor_before

    changed: list[dict[str, Any]] = []
    if upper_bound >= cursor_before:
        changed = _fetch_complete_entry_list(
            api,
            module="ClassContactsRelationship",
            select_fields=_RELATIONSHIP_FIELDS,
            query=(
                f"most_class_contacts_c.date_modified >= '{_api_datetime(cursor_before)}' "
                f"AND most_class_contacts_c.date_modified <= '{_api_datetime(upper_bound)}'"
            ),
            order_by="date_modified ASC, id ASC",
            monotonic_fields=("date_modified", "id"),
        )

    overlap_start = _parse_optional_datetime(cursor_metadata.get("class_overlap_until")) or config.initial_since
    overlap_classes = _fetch_complete_entry_list(
        api,
        module="most_class",
        select_fields=_CLASS_FIELDS,
        query=(
            f"most_class.date_start >= '{_api_datetime(overlap_start)}' "
            f"AND most_class.date_start < '{_api_datetime(run_started)}'"
        ),
        order_by="date_start ASC, id ASC",
        monotonic_fields=("date_start", "id"),
    ) if overlap_start < run_started else []
    class_by_id = {_required_id(item, "most_class"): item for item in overlap_classes}

    overlap_relations: list[dict[str, Any]] = []
    for class_id in sorted(class_by_id):
        overlap_relations.extend(
            _fetch_complete_entry_list(
                api,
                module="ClassContactsRelationship",
                select_fields=_RELATIONSHIP_FIELDS,
                field_values={"most_class_id": class_id},
                query=f"most_class_contacts_c.date_modified <= '{_api_datetime(upper_bound)}'",
                order_by="date_modified ASC, id ASC",
                monotonic_fields=("date_modified", "id"),
            )
        )

    relationships: dict[str, dict[str, Any]] = {}
    for item in (*changed, *overlap_relations):
        relationship_id = _required_id(item, "ClassContactsRelationship")
        previous = relationships.get(relationship_id)
        if previous is None or _parse_tallanto_datetime(item.get("date_modified")) >= _parse_tallanto_datetime(
            previous.get("date_modified")
        ):
            relationships[relationship_id] = item

    missing_class_ids = {
        _required_text(item.get("most_class_id"), "most_class_id")
        for item in relationships.values()
        if _required_text(item.get("most_class_id"), "most_class_id") not in class_by_id
    }
    class_by_id.update(_fetch_class_rows(api, missing_class_ids))

    tallanto_customers = _load_unique_identity_customers(
        db,
        tenant_id=config.tenant_id,
        link_type="tallanto_student_id",
    )
    amo_customers = _load_unique_identity_customers(
        db,
        tenant_id=config.tenant_id,
        link_type="amo_contact_id",
    )
    customer_families = _load_confident_customer_families(db, tenant_id=config.tenant_id)
    family_amo_contacts = _load_family_amo_contacts(db, tenant_id=config.tenant_id)
    contact_rows = _fetch_contact_rows(
        api,
        {_required_text(item.get("contact_id"), "contact_id") for item in relationships.values()},
    )

    counters: Counter[str] = Counter()
    events: list[TimelineEvent] = []
    unresolved: list[dict[str, Any]] = []
    resolved_identity_conflicts: list[dict[str, Any]] = []
    open_identity_conflicts = _load_open_attendance_identity_conflicts(db, tenant_id=config.tenant_id)
    for relationship_id, relationship in sorted(relationships.items()):
        contact_id = _required_text(relationship.get("contact_id"), "contact_id")
        class_id = _required_text(relationship.get("most_class_id"), "most_class_id")
        amo_contact_id = _scalar_text(contact_rows.get(contact_id, {}).get("amo_id"))
        tallanto_customer = tallanto_customers.get(contact_id)
        amo_customer = amo_customers.get(amo_contact_id) if amo_contact_id else None
        if tallanto_customer and amo_customer and tallanto_customer != amo_customer:
            tallanto_family = customer_families.get(tallanto_customer)
            same_family = bool(tallanto_family and tallanto_family == customer_families.get(amo_customer))
            direct_family = (tallanto_customer, amo_contact_id) in family_amo_contacts
            if same_family or direct_family:
                counters["identity_same_family"] += 1
                if direct_family:
                    counters["identity_direct_family"] += 1
                resolved_identity_conflicts.append(
                    _unresolved_relationship(
                        "identity_conflict",
                        relationship,
                        amo_contact_id=amo_contact_id,
                        tallanto_customer=tallanto_customer,
                        amo_customer=amo_customer,
                    )
                )
            else:
                counters["identity_conflict"] += 1
                unresolved.append(
                    _unresolved_relationship(
                        "identity_conflict",
                        relationship,
                        amo_contact_id=amo_contact_id,
                        tallanto_customer=tallanto_customer,
                        amo_customer=amo_customer,
                    )
                )
                continue
        customer_id = tallanto_customer or amo_customer
        if not customer_id:
            # D4: split the old single "identity_unmatched" bucket in two.
            # - identity_unmatched_expected: the Contact fetch *did* return
            #   this contact_id (resolution was actually attempted with full
            #   data) and it genuinely has no tallanto/amo link yet -- normal
            #   steady-state noise (e.g. a brand-new student not yet linked
            #   anywhere else), not a defect in this run.
            # - identity_infrastructure_gap: the Contact fetch never returned
            #   this contact_id at all, so resolution could not even be
            #   attempted -- a data-completeness problem with this run, not
            #   an "expected" unmatched identity.
            reason = "identity_unmatched_expected" if contact_id in contact_rows else "identity_infrastructure_gap"
            counters[reason] += 1
            counters["identity_unmatched"] += 1  # backward-compatible total of both buckets.
            unresolved.append(
                _unresolved_relationship(
                    reason,
                    relationship,
                    amo_contact_id=amo_contact_id,
                    tallanto_customer=tallanto_customer,
                    amo_customer=amo_customer,
                )
            )
            continue
        for reason in open_identity_conflicts.get((relationship_id, contact_id), ()):
            resolved_identity_conflicts.append(
                _unresolved_relationship(
                    reason,
                    relationship,
                    amo_contact_id=amo_contact_id,
                    tallanto_customer=tallanto_customer,
                    amo_customer=amo_customer,
                )
            )
        class_row = class_by_id.get(class_id)
        if class_row is None:
            raise ValueError(f"Tallanto most_class was not resolved: {class_id}")
        if not _scalar_text(class_row.get("date_start")):
            counters["class_date_missing"] += 1
        events.append(
            _relationship_event(
                relationship,
                class_row,
                tenant_id=config.tenant_id,
                customer_id=customer_id,
                created_at=_parse_tallanto_datetime(relationship.get("date_modified")).astimezone(
                    timezone.utc
                ),
            )
        )
    counters["relationships_changed"] = len(changed)
    counters["relationships_from_class_overlap"] = len(overlap_relations)
    counters["relationships_unique"] = len(relationships)
    counters["events_resolved"] = len(events)
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        existing_event_count = int(
            con.execute(
                "SELECT count(*) FROM timeline_events WHERE tenant_id=? AND source_system=?",
                (config.tenant_id, API_SOURCE_SYSTEM),
            ).fetchone()[0]
        )
    counters["existing_events_before"] = existing_event_count
    # Without a durable retry queue, advancing past an unmatched relationship
    # would lose it permanently once the identity appears later.
    blocking_reasons = (
        "identity_conflict",
        "identity_infrastructure_gap",
        "identity_unmatched_expected",
    )
    blocking_unresolved_count = sum(counters[reason] for reason in blocking_reasons)
    validation_errors = [reason for reason in blocking_reasons if counters[reason]]
    if not events and not existing_event_count and not validation_errors:
        validation_errors.append("no_attendance_events")
    unresolved_count = len(unresolved)
    run_status = "partial" if validation_errors else "completed"
    cursor_may_advance = config.apply and not validation_errors
    class_overlap_after = run_started if cursor_may_advance else overlap_start

    if config.apply:
        with CustomerTimelineSQLiteStore(db, allowed_root=root) as store:
            run = store.start_ingestion_run(
                tenant_id=config.tenant_id,
                source_system=API_SOURCE_SYSTEM,
                source_ref=f"tallanto:attendance-api:{_api_datetime(cursor_before)}:{_api_datetime(upper_bound)}",
                run_kind="tallanto_attendance_api_increment",
                idempotency_key=stable_digest(
                    {
                        "upper_bound": upper_bound.isoformat(),
                        "class_overlap_until": class_overlap_after.isoformat(),
                        "events": [event.to_json_dict() for event in events],
                        "unresolved": unresolved,
                    }
                ),
                input_hash=stable_digest({"cursor": cursor_before.isoformat(), "upper": upper_bound.isoformat()}),
                actor=config.actor,
            )
            try:
                with store.bulk_write():
                    for event in events:
                        result = store.upsert_event(event, actor=config.actor, ingestion_run_id=run.run_id)
                        counters[result.status] += 1
                    for item in unresolved:
                        result = store.record_conflict(
                            config.tenant_id,
                            conflict_type=f"{API_SOURCE_SYSTEM}_{item['reason']}",
                            entity_refs=(
                                f"tallanto:class-contact:{item['relationship_id']}",
                                f"tallanto:contact:{item['contact_id']}",
                            ),
                            severity=_UNRESOLVED_SEVERITY_BY_REASON.get(item["reason"], "medium"),
                            summary="Tallanto attendance relationship requires identity resolution",
                            metadata=item,
                            actor=config.actor,
                            ingestion_run_id=run.run_id,
                        )
                        counters[f"unresolved_{result.status}"] += 1
                    resolved_keys: set[tuple[str, str, str]] = set()
                    for item in resolved_identity_conflicts:
                        resolved_key = (item["reason"], item["relationship_id"], item["contact_id"])
                        if resolved_key in resolved_keys:
                            continue
                        resolved_keys.add(resolved_key)
                        store.record_conflict(
                            config.tenant_id,
                            conflict_type=f"{API_SOURCE_SYSTEM}_{item['reason']}",
                            entity_refs=(
                                f"tallanto:class-contact:{item['relationship_id']}",
                                f"tallanto:contact:{item['contact_id']}",
                            ),
                            status="resolved",
                            severity=_UNRESOLVED_SEVERITY_BY_REASON.get(item["reason"], "medium"),
                            summary="Tallanto attendance identity was resolved",
                            metadata=item,
                            actor=config.actor,
                            ingestion_run_id=run.run_id,
                        )
                    # Keep the relationship cursor on unresolved identities, but do not
                    # rescan every old class: those relationships remain covered by the
                    # held relationship window on the next run.
                    store.upsert_ingestion_cursor(
                        config.tenant_id,
                        API_SOURCE_SYSTEM,
                        last_cursor_ts=upper_bound if cursor_may_advance else cursor_before,
                        metadata={
                            "class_overlap_until": class_overlap_after.isoformat(),
                            "upper_bound": upper_bound.isoformat(),
                        },
                        actor=config.actor,
                        ingestion_run_id=run.run_id,
                    )
            except Exception:
                store.finish_ingestion_run(
                    run.run_id,
                    status="failed",
                    accepted_count=0,
                    rejected_count=1,
                    error="attendance_api_write_failed",
                    metadata={"counts": dict(counters)},
                    actor=config.actor,
                )
                raise
            store.finish_ingestion_run(
                run.run_id,
                status=run_status,
                accepted_count=len(events),
                rejected_count=unresolved_count,
                metadata={"counts": dict(counters), "unresolved": unresolved},
                actor=config.actor,
            )

    return {
        "status": run_status,
        "validation_ok": not validation_errors,
        "validation_errors": validation_errors,
        "unresolved_count": unresolved_count,
        # D4: exact numbers + reasons behind the "fresh despite unmatched"
        # decision -- never silenced, always visible even when run_status
        # ends up "completed".
        "unresolved_breakdown": {
            "identity_conflict": counters["identity_conflict"],
            "identity_infrastructure_gap": counters["identity_infrastructure_gap"],
            "identity_unmatched_expected": counters["identity_unmatched_expected"],
            "blocking_count": blocking_unresolved_count,
            "expected_unmatched_blocks_freshness": True,
            "input_fully_processed": True,
        },
        "apply": config.apply,
        "cursor_before": cursor_before.isoformat(),
        "cursor_after": upper_bound.isoformat() if cursor_may_advance else cursor_before.isoformat(),
        "class_overlap_before": overlap_start.isoformat(),
        "class_overlap_after": class_overlap_after.isoformat(),
        "counts": dict(counters),
        "safety": {
            "network_read_only": True,
            "http_methods": ["GET"],
            "writes_tallanto": False,
            "writes_prod": False,
            "writes_staging": bool(config.apply),
            "writeoff_amount_inferred": False,
        },
    }


def _unresolved_relationship(
    reason: str,
    relationship: Mapping[str, Any],
    *,
    amo_contact_id: str,
    tallanto_customer: str | None,
    amo_customer: str | None,
) -> dict[str, Any]:
    return {
        "reason": reason,
        "relationship_id": _required_id(relationship, "ClassContactsRelationship"),
        "contact_id": _required_text(relationship.get("contact_id"), "contact_id"),
        "most_class_id": _required_text(relationship.get("most_class_id"), "most_class_id"),
        "amo_contact_id": amo_contact_id or None,
        "tallanto_customer_id": tallanto_customer,
        "amo_customer_id": amo_customer,
        "date_modified": _scalar_text(relationship.get("date_modified")),
    }


def run_tallanto_attendance_import(config: TallantoAttendanceImportConfig) -> Mapping[str, Any]:
    root = Path(config.allowed_root).expanduser().resolve(strict=False)
    db = guard_customer_timeline_output_path(Path(config.timeline_db), root)
    if config.apply and (is_customer_timeline_prod_path(db) or not _is_staging_db(db)):
        raise ValueError("Tallanto attendance apply requires a staging Customer Timeline DB")
    barcode_to_tallanto, duplicate_barcodes = _load_contact_barcodes(Path(config.contacts_workbook))
    if config.apply and duplicate_barcodes:
        raise ValueError("Tallanto attendance apply blocked by duplicate contact barcodes")
    tallanto_to_customer = _load_tallanto_customers(db, tenant_id=config.tenant_id)
    attendance_rows = read_writeoff_xlsx(Path(config.attendance_report))
    unresolved_barcodes = {
        barcode
        for row in attendance_rows
        if (barcode := _normalize_barcode(row.get("barcode")))
        and (
            barcode not in barcode_to_tallanto
            or barcode_to_tallanto[barcode] not in tallanto_to_customer
        )
    }
    api_links: dict[str, tuple[str, str, str]] = {}
    api_identity_conflicts: set[str] = set()
    if config.tallanto_env_file and unresolved_barcodes:
        api_links, api_identity_conflicts = _resolve_missing_barcodes_with_api(
            db,
            env_file=Path(config.tallanto_env_file),
            barcodes=unresolved_barcodes,
            tenant_id=config.tenant_id,
        )
        for barcode, (tallanto_id, customer_id, _amo_contact_id) in tuple(api_links.items()):
            workbook_tallanto_id = barcode_to_tallanto.get(barcode)
            if workbook_tallanto_id and workbook_tallanto_id != tallanto_id:
                api_identity_conflicts.add(barcode)
                api_links.pop(barcode)
                continue
            barcode_to_tallanto[barcode] = tallanto_id
            tallanto_to_customer[tallanto_id] = customer_id
    counters: Counter[str] = Counter()
    counters["api_requested_barcodes"] = len(unresolved_barcodes) if config.tallanto_env_file else 0
    counters["api_resolved_barcodes"] = len(api_links)
    counters["api_identity_conflicts"] = len(api_identity_conflicts)
    unmatched_barcodes: set[str] = set()
    events: list[TimelineEvent] = []
    for row in attendance_rows:
        barcode = _normalize_barcode(row.get("barcode"))
        if not barcode or barcode in duplicate_barcodes:
            counters["blocked_barcode"] += 1
            continue
        tallanto_id = barcode_to_tallanto.get(barcode, "")
        customer_id = tallanto_to_customer.get(tallanto_id, "")
        if not tallanto_id or not customer_id:
            counters["unmatched"] += 1
            unmatched_barcodes.add(barcode)
            continue
        try:
            event_at = datetime.strptime(str(row.get("class_at") or ""), "%d.%m.%Y %H:%M").replace(
                tzinfo=ZoneInfo("Europe/Moscow")
            )
        except ValueError:
            counters["invalid_class_at"] += 1
            continue
        logical_key = {
            "tallanto_id": tallanto_id,
            "class_at": event_at.isoformat(),
            "class_title": str(row.get("class_title") or "").strip(),
            "subscription": str(row.get("subscription") or "").strip(),
        }
        source_id = stable_digest(logical_key)
        events.append(
            TimelineEvent(
                tenant_id=config.tenant_id,
                customer_id=customer_id,
                event_type=TimelineEventType.TALLANTO_ATTENDANCE,
                event_at=event_at,
                source_system=SOURCE_SYSTEM,
                source_id=source_id,
                source_ref=f"tallanto:attendance:{source_id}",
                direction=TimelineDirection.SYSTEM,
                subject=str(row.get("class_title") or "").strip() or "Занятие Tallanto",
                summary="Списание за занятие подтверждено в Tallanto; физическое присутствие отдельно не доказано.",
                match_status=IdentityMatchClass.STRONG_UNIQUE,
                confidence=1.0,
                record={"logical_key": logical_key, "class_branch": row.get("class_branch"), "writeoff_amount": row.get("writeoff_amount"), "writeoff_type": row.get("writeoff_type")},
                metadata={"attendance_evidence": "class_writeoff", "bot_safe": False},
                created_at=event_at,
            )
        )
    counters["resolved"] = len(events)
    counters["unmatched_barcodes"] = len(unmatched_barcodes)
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        existing = {
            str(row[0])
            for row in con.execute(
                "SELECT source_id FROM timeline_events WHERE tenant_id=? AND source_system=?",
                (config.tenant_id, SOURCE_SYSTEM),
            )
        }
    counters["would_duplicate"] = sum(event.source_id in existing for event in events)
    counters["would_create"] = len(events) - counters["would_duplicate"]
    validation_errors = _validation_errors(
        attendance_rows=len(attendance_rows),
        duplicate_barcodes=duplicate_barcodes,
        counters=counters,
    )
    if config.apply and validation_errors:
        raise ValueError(
            "Tallanto attendance apply blocked by failed validation: "
            + ", ".join(validation_errors)
        )
    if config.apply:
        with CustomerTimelineSQLiteStore(db, allowed_root=root) as store:
            run = store.start_ingestion_run(
                tenant_id=config.tenant_id,
                source_system=SOURCE_SYSTEM,
                source_ref=f"report:{_file_hash(Path(config.attendance_report))}",
                run_kind="tallanto_attendance_import",
                idempotency_key=stable_digest([event.source_id for event in events]),
                input_hash=_file_hash(Path(config.attendance_report)),
                actor=config.actor,
            )
            try:
                with store.bulk_write():
                    for _barcode, (tallanto_id, customer_id, amo_contact_id) in api_links.items():
                        result = store.upsert_identity_link(
                            IdentityLink(
                                tenant_id=config.tenant_id,
                                customer_id=customer_id,
                                link_type=IdentityLinkType.TALLANTO_STUDENT_ID,
                                link_value=tallanto_id,
                                source_system="tallanto_api",
                                source_ref=f"tallanto:contact:{tallanto_id}",
                                match_class=IdentityMatchClass.STRONG_UNIQUE,
                                confidence=1.0,
                                evidence={"method": "exact_barcode_plus_unique_amo_contact", "amo_contact_id": amo_contact_id},
                            ),
                            actor=config.actor,
                            ingestion_run_id=run.run_id,
                        )
                        counters[f"identity_{result.status}"] += 1
                    for event in events:
                        result = store.upsert_event(event, actor=config.actor, ingestion_run_id=run.run_id)
                        counters[result.status] += 1
            except Exception:
                store.finish_ingestion_run(
                    run.run_id,
                    status="failed",
                    accepted_count=counters["created"] + counters["updated"] + counters["duplicate"],
                    rejected_count=1,
                    error="attendance_write_failed",
                    metadata={"counts": dict(counters)},
                    actor=config.actor,
                )
                raise
            store.finish_ingestion_run(
                run.run_id,
                status="completed",
                accepted_count=len(events),
                rejected_count=sum(counters[key] for key in ("blocked_barcode", "unmatched", "invalid_class_at")),
                metadata={"counts": dict(counters)},
                actor=config.actor,
            )
    return {
        "validation_ok": not validation_errors,
        "validation_errors": validation_errors,
        "apply": config.apply,
        "counts": dict(counters),
        "safety": {
            "network_calls": bool(config.tallanto_env_file and unresolved_barcodes),
            "network_read_only": True,
            "writes_tallanto": False,
            "writes_prod": False,
        },
    }


def _is_staging_db(path: Path) -> bool:
    parts = path.resolve(strict=False).parts
    return any(tuple(parts[index : index + 2]) == _STAGING_COMPONENTS for index in range(len(parts) - 1))


def _load_api_cursor(
    db: Path,
    *,
    tenant_id: str,
    initial_since: datetime,
) -> tuple[datetime, Mapping[str, Any]]:
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        row = con.execute(
            "SELECT last_cursor_ts, metadata_json FROM ingestion_cursors "
            "WHERE tenant_id=? AND source_system=?",
            (tenant_id, API_SOURCE_SYSTEM),
        ).fetchone()
    if row is None:
        return initial_since.astimezone(_MOSCOW), {}
    raw_metadata: Mapping[str, Any] = {}
    try:
        decoded = json.loads(str(row[1] or "{}"))
        if isinstance(decoded, Mapping):
            nested = decoded.get("metadata")
            raw_metadata = nested if isinstance(nested, Mapping) else decoded
    except json.JSONDecodeError:
        raw_metadata = {}
    return _parse_tallanto_datetime(row[0]), raw_metadata


def _load_confident_customer_families(db: Path, *, tenant_id: str) -> dict[str, str]:
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        if con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='family_members_v1'"
        ).fetchone() is None:
            return {}
        return {
            str(customer_id): str(family_id)
            for customer_id, family_id in con.execute(
                "SELECT customer_id, family_id FROM family_members_v1 "
                "WHERE tenant_id=? AND membership_status='confident' AND confidence='high'",
                (tenant_id,),
            )
        }


def _load_open_attendance_identity_conflicts(
    db: Path, *, tenant_id: str
) -> dict[tuple[str, str], tuple[str, ...]]:
    result: dict[tuple[str, str], list[str]] = defaultdict(list)
    conflict_types = tuple(f"{API_SOURCE_SYSTEM}_{reason}" for reason in _UNRESOLVED_SEVERITY_BY_REASON)
    placeholders = ",".join("?" for _ in conflict_types)
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        rows = con.execute(
            f"SELECT conflict_type, record_json FROM timeline_conflicts "
            f"WHERE tenant_id=? AND status='open' AND conflict_type IN ({placeholders})",
            (tenant_id, *conflict_types),
        ).fetchall()
    for conflict_type, raw in rows:
        try:
            metadata = json.loads(str(raw or "{}")).get("metadata") or {}
        except (json.JSONDecodeError, AttributeError):
            continue
        relationship_id = _scalar_text(metadata.get("relationship_id"))
        contact_id = _scalar_text(metadata.get("contact_id"))
        reason = str(conflict_type).removeprefix(f"{API_SOURCE_SYSTEM}_")
        if relationship_id and contact_id:
            result[(relationship_id, contact_id)].append(reason)
    return {key: tuple(dict.fromkeys(reasons)) for key, reasons in result.items()}


def _load_family_amo_contacts(db: Path, *, tenant_id: str) -> set[tuple[str, str]]:
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        return {
            (str(customer_id), str(link_value))
            for customer_id, link_value in con.execute(
                "SELECT customer_id, link_value FROM identity_links "
                "WHERE tenant_id=? AND source_system='tallanto_snapshot' "
                "AND link_type='amo_contact_id' AND match_class='ambiguous' "
                "AND json_extract(record_json, '$.evidence.relationship')='family_amo_contact'",
                (tenant_id,),
            )
        }


def _fetch_relationship_upper_bound(client: Any) -> datetime | None:
    payload = client.request(
        method_name="get_entry_list",
        module="ClassContactsRelationship",
        http_method="GET",
        query_items=[("offset", "0"), ("order_by", "date_modified DESC, id DESC")],
        select_fields=("id", "date_modified"),
    )
    rows = payload.get("entry_list")
    if not isinstance(rows, list):
        raise ValueError("Tallanto get_entry_list misses entry_list")
    if not rows:
        return None
    first = rows[0]
    if not isinstance(first, Mapping):
        raise ValueError("Tallanto get_entry_list returned a non-object row")
    return _parse_tallanto_datetime(first.get("date_modified"))


def _fetch_complete_entry_list(
    client: Any,
    *,
    module: str,
    select_fields: tuple[str, ...],
    query: str | None = None,
    field_values: Mapping[str, str] | None = None,
    order_by: str,
    monotonic_fields: tuple[str, str] | None,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    total_count: int | None = None
    seen_offsets: set[int] = set()
    while True:
        if offset in seen_offsets:
            raise ValueError(f"Tallanto pagination repeated offset for {module}: {offset}")
        seen_offsets.add(offset)
        query_items = [("offset", str(offset)), ("order_by", order_by)]
        if query:
            query_items.append(("query", query))
        for field_name, field_value in (field_values or {}).items():
            query_items.append((f"fields_values[{field_name}]", str(field_value)))
        payload = client.request(
            method_name="get_entry_list",
            module=module,
            http_method="GET",
            query_items=query_items,
            select_fields=select_fields,
        )
        page = payload.get("entry_list")
        if not isinstance(page, list) or any(not isinstance(item, dict) for item in page):
            raise ValueError(f"Tallanto {module} page misses a valid entry_list")
        try:
            page_count = int(payload["result_count"])
            page_total = int(payload["total_count"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"Tallanto {module} page misses result_count/total_count") from exc
        if page_count != len(page):
            raise ValueError(f"Tallanto {module} result_count mismatch")
        if total_count is None:
            total_count = page_total
        elif total_count != page_total:
            raise ValueError(f"Tallanto {module} total_count changed during pagination")
        rows.extend(page)
        if not page:
            break
        if len(rows) == total_count:
            break
        if len(rows) > total_count:
            raise ValueError(f"Tallanto {module} returned more rows than total_count")
        next_offset = payload.get("next_offset")
        if next_offset in (None, ""):
            if len(rows) != total_count:
                raise ValueError(f"Tallanto {module} pagination ended before total_count")
            break
        try:
            next_value = int(next_offset)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"Tallanto {module} next_offset is invalid") from exc
        if next_value <= offset:
            raise ValueError(f"Tallanto {module} next_offset did not increase")
        offset = next_value
    if total_count is None or len(rows) != total_count:
        raise ValueError(f"Tallanto {module} pagination completeness check failed")
    identifiers = [_required_id(item, module) for item in rows]
    if len(identifiers) != len(set(identifiers)):
        raise ValueError(f"Tallanto {module} returned duplicate ids")
    if monotonic_fields is not None:
        ordered = [
            (
                _parse_tallanto_datetime(item.get(monotonic_fields[0])),
                str(item.get(monotonic_fields[1]) or ""),
            )
            for item in rows
        ]
        if ordered != sorted(ordered):
            raise ValueError(f"Tallanto {module} response is not monotonic")
    return rows


def _fetch_contact_rows(client: Any, contact_ids: set[str]) -> dict[str, Mapping[str, Any]]:
    result: dict[str, Mapping[str, Any]] = {}
    ordered_ids = sorted(contact_ids)
    for index in range(0, len(ordered_ids), 100):
        chunk = ordered_ids[index : index + 100]
        query = "contacts.id IN (" + ",".join(f"'{_api_quote(value)}'" for value in chunk) + ")"
        rows = _fetch_complete_entry_list(
            client,
            module="Contact",
            select_fields=("id", "amo_id"),
            query=query,
            order_by="id ASC",
            monotonic_fields=None,
        )
        for row in rows:
            contact_id = _required_id(row, "Contact")
            if contact_id not in chunk:
                raise ValueError("Tallanto Contact batch returned an unrequested id")
            result[contact_id] = row
    return result


def _fetch_class_rows(client: Any, class_ids: set[str]) -> dict[str, Mapping[str, Any]]:
    result: dict[str, Mapping[str, Any]] = {}
    ordered_ids = sorted(class_ids)
    for index in range(0, len(ordered_ids), 100):
        chunk = ordered_ids[index : index + 100]
        query = "most_class.id IN (" + ",".join(f"'{_api_quote(value)}'" for value in chunk) + ")"
        rows = _fetch_complete_entry_list(
            client,
            module="most_class",
            select_fields=_CLASS_FIELDS,
            query=query,
            order_by="id ASC",
            monotonic_fields=None,
        )
        for row in rows:
            class_id = _required_id(row, "most_class")
            if class_id not in chunk:
                raise ValueError("Tallanto most_class batch returned an unrequested id")
            result[class_id] = row
    return result


def _relationship_event(
    relationship: Mapping[str, Any],
    class_row: Mapping[str, Any],
    *,
    tenant_id: str,
    customer_id: str,
    created_at: datetime,
) -> TimelineEvent:
    relationship_id = _required_id(relationship, "ClassContactsRelationship")
    class_id = _required_id(class_row, "most_class")
    status = _scalar_text(relationship.get("most_class_contacts_status")).casefold()
    abonement_id = _reference_id(relationship.get("most_class_abonements"))
    attendance_confirmed = status == "visit"
    physical_absence_confirmed = status == "no-show"
    writeoff_confirmed = bool(abonement_id and status in {"visit", "no-show"})
    active_fact = attendance_confirmed or physical_absence_confirmed or writeoff_confirmed
    if attendance_confirmed and writeoff_confirmed:
        summary = "Посещение и списание занятия подтверждены в Tallanto."
    elif attendance_confirmed:
        summary = "Посещение подтверждено в Tallanto; списание не подтверждено."
    elif physical_absence_confirmed and writeoff_confirmed:
        summary = "Списание занятия подтверждено; физическое посещение не было отмечено."
    elif physical_absence_confirmed:
        summary = "Неявка подтверждена в Tallanto; списание не подтверждено."
    else:
        summary = "Статус занятия не завершён; активный факт посещения или списания отсутствует."
    class_started_at = _parse_optional_datetime(class_row.get("date_start"))
    event_at_source = "most_class.date_start"
    if class_started_at is None:
        class_started_at = _parse_optional_datetime(relationship.get("date_entry"))
        event_at_source = "relationship.date_entry"
    if class_started_at is None:
        class_started_at = _parse_tallanto_datetime(relationship.get("date_modified"))
        event_at_source = "relationship.date_modified"
    if event_at_source != "most_class.date_start":
        summary += " Точная дата занятия в карточке недоступна."
    return TimelineEvent(
        tenant_id=tenant_id,
        customer_id=customer_id,
        event_type=TimelineEventType.TALLANTO_ATTENDANCE,
        event_at=class_started_at,
        source_system=API_SOURCE_SYSTEM,
        source_id=relationship_id,
        source_ref=f"tallanto:class-contact:{relationship_id}",
        direction=TimelineDirection.SYSTEM,
        subject=_scalar_text(class_row.get("name")) or "Занятие Tallanto",
        summary=summary,
        match_status=IdentityMatchClass.STRONG_UNIQUE,
        confidence=1.0,
        record={
            "relationship_id": relationship_id,
            "tallanto_student_id": _required_text(relationship.get("contact_id"), "contact_id"),
            "most_class_id": class_id,
            "relationship_status": status,
            "abonement_id": abonement_id,
            "attendance_confirmed": attendance_confirmed,
            "physical_absence_confirmed": physical_absence_confirmed,
            "writeoff_confirmed": writeoff_confirmed,
            "writeoff_amount": None,
            "fact_active": active_fact,
            "class_status": _scalar_text(class_row.get("status")),
            "course_id": _scalar_text(class_row.get("cource_id")),
            "branch": _scalar_text(class_row.get("filial")),
            "date_modified": _scalar_text(relationship.get("date_modified")),
            "event_at_source": event_at_source,
        },
        metadata={
            "attendance_evidence": "class_contacts_relationship",
            "bot_safe": False,
            "fact_active": active_fact,
        },
        created_at=created_at,
    )


def _required_id(payload: Mapping[str, Any], module: str) -> str:
    return _required_text(payload.get("id"), f"{module}.id")


def _required_text(value: object, field_name: str) -> str:
    text = _scalar_text(value)
    if not text:
        raise ValueError(f"Tallanto response misses {field_name}")
    return text


def _scalar_text(value: object) -> str:
    if isinstance(value, Mapping):
        for key in ("id", "value", "name"):
            if value.get(key) not in (None, ""):
                return str(value[key]).strip()
        return ""
    if isinstance(value, (list, tuple)):
        return _scalar_text(value[0]) if value else ""
    return str(value or "").strip()


def _reference_id(value: object) -> str | None:
    text = _scalar_text(value)
    return None if text.casefold() in {"", "0", "none", "null", "false"} else text


def _parse_optional_datetime(value: object) -> datetime | None:
    return _parse_tallanto_datetime(value) if str(value or "").strip() else None


def _parse_tallanto_datetime(value: object) -> datetime:
    text = str(value or "").strip()
    if not text:
        raise ValueError("Tallanto response misses datetime")
    normalized = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError(f"Tallanto datetime is invalid: {text}") from exc
    return parsed.replace(tzinfo=_MOSCOW) if parsed.tzinfo is None else parsed.astimezone(_MOSCOW)


def _api_datetime(value: datetime) -> str:
    if value.tzinfo is None:
        raise ValueError("Tallanto API datetime must be timezone-aware")
    return value.astimezone(_MOSCOW).strftime("%Y-%m-%d %H:%M:%S")


def _api_quote(value: str) -> str:
    return str(value).replace("'", "''")


def _validation_errors(
    *,
    attendance_rows: int,
    duplicate_barcodes: set[str],
    counters: Mapping[str, int],
) -> list[str]:
    errors: list[str] = []
    if duplicate_barcodes:
        errors.append("duplicate_contact_barcodes")
    if attendance_rows == 0:
        errors.append("empty_report")
    elif int(counters.get("resolved", 0)) == 0:
        if int(counters.get("invalid_class_at", 0)) == attendance_rows:
            errors.append("all_rows_invalid_class_at")
        elif int(counters.get("unmatched", 0)) + int(counters.get("blocked_barcode", 0)) == attendance_rows:
            errors.append("all_rows_unmatched")
        else:
            errors.append("no_resolved_events")
    return errors


def _load_contact_barcodes(path: Path) -> tuple[dict[str, str], set[str]]:
    if path.suffix.casefold() == ".xls":
        from python_calamine import CalamineWorkbook

        raw_rows = iter(CalamineWorkbook.from_path(path).get_sheet_by_index(0).to_python())
        header = {str(value or "").strip(): index for index, value in enumerate(next(raw_rows))}
        rows = raw_rows
    else:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        raw_rows = workbook.active.iter_rows(values_only=True)
        header = {str(value or "").strip(): index for index, value in enumerate(next(raw_rows))}
        rows = raw_rows
    required = {"ID", "Текстовое значение штрихкода"}
    if not required <= header.keys():
        raise ValueError("Tallanto contacts workbook misses ID/barcode columns")
    values: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        barcode = _normalize_barcode(row[header["Текстовое значение штрихкода"]])
        tallanto_id = str(row[header["ID"]] or "").strip()
        if barcode and tallanto_id:
            values[barcode].add(tallanto_id)
    duplicates = {barcode for barcode, ids in values.items() if len(ids) != 1}
    return {barcode: next(iter(ids)) for barcode, ids in values.items() if barcode not in duplicates}, duplicates


def _normalize_barcode(value: object) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value or "").strip()
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def _resolve_missing_barcodes_with_api(
    db: Path,
    *,
    env_file: Path,
    barcodes: set[str],
    tenant_id: str,
) -> tuple[dict[str, tuple[str, str, str]], set[str]]:
    client = _build_tallanto_client(env_file)
    amo_customers = _load_unique_identity_customers(db, tenant_id=tenant_id, link_type="amo_contact_id")
    tallanto_customers = _load_unique_identity_customers(db, tenant_id=tenant_id, link_type="tallanto_student_id")
    resolved: dict[str, tuple[str, str, str]] = {}
    conflicts: set[str] = set()
    for barcode in sorted(barcodes):
        payload = client.get_entry_by_fields(
            module="Contact",
            field_values={"barcode": barcode},
            select_fields=("id", "barcode", "amo_id"),
        )
        tallanto_id = str(payload.get("id") or "").strip()
        amo_contact_id = str(payload.get("amo_id") or "").strip()
        if not tallanto_id or _normalize_barcode(payload.get("barcode")) != barcode:
            continue
        tallanto_customer_id = tallanto_customers.get(tallanto_id)
        amo_customer_id = amo_customers.get(amo_contact_id)
        if tallanto_customer_id and amo_customer_id and tallanto_customer_id != amo_customer_id:
            conflicts.add(barcode)
            continue
        customer_id = tallanto_customer_id or amo_customer_id
        if customer_id:
            resolved[barcode] = (tallanto_id, customer_id, amo_contact_id)
    return resolved, conflicts


def _build_tallanto_client(env_file: Path):
    from mango_mvp.amocrm_runtime.tallanto_api import TallantoApiClient, TallantoApiConfig

    values = dotenv_values(env_file)
    base_url = str(values.get("CRM_TALLANTO_BASE_URL") or "").strip()
    api_token = str(values.get("CRM_TALLANTO_API_TOKEN") or "").strip()
    rest_path = str(values.get("CRM_TALLANTO_STUDENT_PATH") or "/service/api/rest.php").strip()
    if not base_url or not api_token:
        raise ValueError("Tallanto read-only env is incomplete")
    if "{student_id}" in rest_path:
        rest_path = "/service/api/rest.php"
    return TallantoApiClient(TallantoApiConfig(base_url=base_url, api_token=api_token, rest_path=rest_path))


def _load_unique_identity_customers(db: Path, *, tenant_id: str, link_type: str) -> dict[str, str]:
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        con.row_factory = sqlite3.Row
        return {
            str(row["link_value"]): str(row["customer_id"])
            for row in authoritative_exact_identity_rows(con, tenant_id, link_types=(link_type,))
            if int(row["owner_count"]) == 1 and not int(row["has_open_conflict"])
        }


def _load_tallanto_customers(db: Path, *, tenant_id: str) -> dict[str, str]:
    exact = _load_unique_identity_customers(db, tenant_id=tenant_id, link_type="tallanto_student_id")
    with sqlite3.connect(customer_timeline_readonly_uri(db), uri=True) as con:
        con.execute("PRAGMA query_only=ON")
        candidates = {key: {value} for key, value in exact.items()}
        linked_values = {str(row[0]) for row in con.execute(
            "SELECT DISTINCT link_value FROM identity_links WHERE tenant_id=? AND link_type='tallanto_student_id'",
            (tenant_id,),
        )}
        for customer_id, raw in con.execute(
            "SELECT customer_id, record_json FROM timeline_events "
            "WHERE tenant_id=? AND source_system=? AND customer_id IS NOT NULL "
            "AND match_status='strong_unique' AND confidence>=0.99 AND superseded_by IS NULL",
            (tenant_id, SOURCE_SYSTEM),
        ):
            try:
                tallanto_id = json.loads(str(raw or "{}"))["record"]["logical_key"]["tallanto_id"]
            except (KeyError, TypeError, ValueError, json.JSONDecodeError):
                continue
            # Historical events are only a fallback when no current identity decision exists.
            if tallanto_id and str(tallanto_id) not in linked_values:
                candidates.setdefault(str(tallanto_id), set()).add(str(customer_id))
        return {key: next(iter(values)) for key, values in candidates.items() if len(values) == 1}


def _file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()
