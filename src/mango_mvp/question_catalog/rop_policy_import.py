from __future__ import annotations

import csv
import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Sequence

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_APPROVED_ROP_WORKBOOK = Path("product_data/question_catalog/rop_bot_policy_questionnaire_APPROVED_2026-05-15.xlsx")
DEFAULT_TAXONOMY_PATH = Path("src/mango_mvp/question_catalog/themes_taxonomy.yaml")
DEFAULT_V2_CSV = Path("product_data/question_catalog/rop_bot_policy_questionnaire_v2_2026-05-15.csv")
DEFAULT_V2_XLSX = Path("product_data/question_catalog/rop_bot_policy_questionnaire_v2_2026-05-15.xlsx")

ROP_POLICY_SHEET = "Опросник РОПа"
ROP_POLICY_SCHEMA_VERSION = "question_catalog_rop_policy_v2_2026_05_15"

ROP_DECISION_BOT_SELF = "Бот отвечает сам"
ROP_DECISION_FACT_CHECK = "Бот отвечает после проверки актуального факта"
ROP_DECISION_DRAFT_FOR_MANAGER = "Бот только собирает данные и передает менеджеру"
ROP_DECISION_MANAGER_ONLY = "Только менеджер"
ROP_DECISION_SPLIT_REQUIRED = "Тему нужно раздробить на более узкие случаи"

ALLOWED_ROP_DECISIONS = (
    ROP_DECISION_BOT_SELF,
    ROP_DECISION_FACT_CHECK,
    ROP_DECISION_DRAFT_FOR_MANAGER,
    ROP_DECISION_MANAGER_ONLY,
    ROP_DECISION_SPLIT_REQUIRED,
)

ROP_DECISION_TO_PERMISSION = {
    ROP_DECISION_BOT_SELF: "bot_self",
    ROP_DECISION_FACT_CHECK: "answer_after_fact_check",
    ROP_DECISION_DRAFT_FOR_MANAGER: "draft_for_manager",
    ROP_DECISION_MANAGER_ONLY: "manager_only",
    ROP_DECISION_SPLIT_REQUIRED: "manager_only",
}

REQUIRED_ROP_COLUMNS = (
    "№",
    "Тема",
    "✏️ Решение РОПа",
    "✏️ Утверждённая фраза для бота",
    "✏️ Что бот обязан спросить/проверить",
    "✏️ Что нельзя обещать (помимо системных)",
    "✏️ Кому передавать, если бот не сам",
    "✏️ Комментарий",
)

FINAL_V2_COLUMNS = (
    "theme_id",
    "Номер темы",
    "Блок",
    "Тема",
    "Решение РОПа",
    "Разрешение бота",
    "Утвержденная фраза для бота",
    "Что бот обязан спросить/проверить",
    "Дополнительные запреты РОПа",
    "Кому передавать",
    "Комментарий РОПа",
    "Нужные факты",
    "Системные запреты",
    "Правило эскалации",
)


@dataclass(frozen=True)
class RopThemePolicy:
    theme_id: str
    theme_number: str
    theme_name: str
    rop_decision: str
    rop_approved_phrasing: str
    rop_mandatory_data: str
    rop_forbids_extra: str
    rop_handoff_target: str
    rop_comment: str = ""

    @property
    def bot_permission(self) -> str:
        try:
            return ROP_DECISION_TO_PERMISSION[self.rop_decision]
        except KeyError as exc:
            raise ValueError(f"unsupported ROP decision for {self.theme_id}: {self.rop_decision!r}") from exc


def load_approved_rop_policies(path: str | Path = DEFAULT_APPROVED_ROP_WORKBOOK) -> dict[str, RopThemePolicy]:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=True)
    if ROP_POLICY_SHEET not in workbook.sheetnames:
        raise ValueError(f"approved ROP workbook must contain sheet {ROP_POLICY_SHEET!r}")
    sheet = workbook[ROP_POLICY_SHEET]
    headers = [clean_text(cell.value) for cell in sheet[1]]
    header_index = {name: index for index, name in enumerate(headers) if name}
    missing = [name for name in REQUIRED_ROP_COLUMNS if name not in header_index]
    if missing:
        raise ValueError(f"approved ROP workbook is missing columns: {', '.join(missing)}")

    policies: dict[str, RopThemePolicy] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        theme_number = clean_text(row[header_index["№"]])
        if not theme_number:
            continue
        theme_id = theme_id_from_number(theme_number)
        decision = clean_text(row[header_index["✏️ Решение РОПа"]])
        if decision not in ALLOWED_ROP_DECISIONS:
            raise ValueError(f"{theme_id}: unsupported ROP decision {decision!r}")
        policy = RopThemePolicy(
            theme_id=theme_id,
            theme_number=theme_number,
            theme_name=clean_text(row[header_index["Тема"]]),
            rop_decision=decision,
            rop_approved_phrasing=clean_text(row[header_index["✏️ Утверждённая фраза для бота"]]),
            rop_mandatory_data=clean_text(row[header_index["✏️ Что бот обязан спросить/проверить"]]),
            rop_forbids_extra=clean_text(row[header_index["✏️ Что нельзя обещать (помимо системных)"]]),
            rop_handoff_target=clean_text(row[header_index["✏️ Кому передавать, если бот не сам"]]),
            rop_comment=clean_text(row[header_index["✏️ Комментарий"]]),
        )
        policies[theme_id] = policy
    return policies


def apply_rop_policies_to_taxonomy(
    taxonomy: Mapping[str, Any],
    policies: Mapping[str, RopThemePolicy],
    *,
    require_all_themes: bool = True,
) -> dict[str, Any]:
    payload = copy.deepcopy(dict(taxonomy))
    payload["rop_policy_schema_version"] = ROP_POLICY_SCHEMA_VERSION
    payload["allowed_rop_decisions"] = list(ALLOWED_ROP_DECISIONS)
    source_documents = dict(payload.get("source_documents") or {})
    source_documents["rop_policy"] = str(DEFAULT_APPROVED_ROP_WORKBOOK)
    payload["source_documents"] = source_documents

    themes = payload.get("themes")
    if not isinstance(themes, list):
        raise ValueError("taxonomy themes must be a list")
    missing: list[str] = []
    for theme in themes:
        theme_id = str(theme.get("theme_id") or "")
        policy = policies.get(theme_id)
        if policy is None:
            missing.append(theme_id)
            continue
        theme["rop_decision"] = policy.rop_decision
        theme["rop_approved_phrasing"] = policy.rop_approved_phrasing
        theme["rop_mandatory_data"] = policy.rop_mandatory_data
        theme["rop_forbids_extra"] = policy.rop_forbids_extra
        theme["rop_handoff_target"] = policy.rop_handoff_target
        theme["rop_comment"] = policy.rop_comment
        theme["default_bot_permission"] = policy.bot_permission
    if require_all_themes and missing:
        raise ValueError(f"approved ROP workbook has no rows for themes: {', '.join(missing)}")
    return payload


def load_taxonomy(path: str | Path = DEFAULT_TAXONOMY_PATH) -> dict[str, Any]:
    return yaml.safe_load(Path(path).read_text(encoding="utf-8"))


def write_taxonomy(path: str | Path, taxonomy: Mapping[str, Any]) -> None:
    Path(path).write_text(yaml.safe_dump(dict(taxonomy), allow_unicode=True, sort_keys=False, width=120), encoding="utf-8")


def build_final_v2_rows(taxonomy: Mapping[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for theme in taxonomy.get("themes") or []:
        rows.append(
            {
                "theme_id": clean_text(theme.get("theme_id")),
                "Номер темы": number_from_theme_id(clean_text(theme.get("theme_id"))),
                "Блок": clean_text(theme.get("business_block")),
                "Тема": clean_text(theme.get("theme_name")),
                "Решение РОПа": clean_text(theme.get("rop_decision")),
                "Разрешение бота": clean_text(theme.get("default_bot_permission")),
                "Утвержденная фраза для бота": clean_text(theme.get("rop_approved_phrasing")),
                "Что бот обязан спросить/проверить": clean_text(theme.get("rop_mandatory_data")),
                "Дополнительные запреты РОПа": clean_text(theme.get("rop_forbids_extra")),
                "Кому передавать": clean_text(theme.get("rop_handoff_target")),
                "Комментарий РОПа": clean_text(theme.get("rop_comment")),
                "Нужные факты": "; ".join(clean_text(item) for item in theme.get("required_facts") or [] if clean_text(item)),
                "Системные запреты": "; ".join(clean_text(item) for item in theme.get("forbidden_promises") or [] if clean_text(item)),
                "Правило эскалации": clean_text(theme.get("escalation_rule")),
            }
        )
    return rows


def write_final_v2_csv(path: str | Path, rows: Sequence[Mapping[str, Any]]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=list(FINAL_V2_COLUMNS))
        writer.writeheader()
        writer.writerows(rows)


def write_final_v2_xlsx(path: str | Path, rows: Sequence[Mapping[str, Any]]) -> None:
    output_path = Path(path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Question Catalog v2"
    sheet.append(list(FINAL_V2_COLUMNS))
    for row in rows:
        sheet.append([row.get(column, "") for column in FINAL_V2_COLUMNS])
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 24,
        "B": 12,
        "C": 24,
        "D": 34,
        "E": 34,
        "F": 24,
        "G": 64,
        "H": 52,
        "I": 52,
        "J": 44,
        "K": 34,
        "L": 34,
        "M": 42,
        "N": 42,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
    workbook.save(output_path)


def theme_id_from_number(number: str) -> str:
    normalized = clean_text(number).lower()
    special = {
        "19a": "theme:019a_positive_feedback",
        "19b": "theme:019b_negative_feedback",
    }
    if normalized in special:
        return special[normalized]
    try:
        numeric = int(normalized)
    except ValueError as exc:
        raise ValueError(f"unsupported theme number: {number!r}") from exc
    prefixes = {
        1: "pricing",
        2: "payment_method",
        3: "payment_status",
        4: "payment_schedule",
        5: "discounts",
        6: "installment",
        7: "matkap_payment",
        8: "tax_deduction",
        9: "refund",
        10: "change_terms",
        11: "contract",
        12: "certificates",
        13: "schedule",
        14: "format",
        15: "address",
        16: "program",
        17: "teacher_method",
        18: "materials_homework",
        20: "enrollment",
        21: "continuation",
        22: "age_level_testing",
        23: "trial_class",
        24: "account_access",
        25: "missing_links_access",
        26: "camp_general",
        27: "camp_living_conditions",
        28: "transport_logistics",
        29: "legal_question",
        30: "partnership_b2b",
        32: "student_progress_inquiry",
    }
    suffix = prefixes.get(numeric)
    if not suffix:
        raise ValueError(f"unsupported theme number: {number!r}")
    return f"theme:{numeric:03d}_{suffix}"


def number_from_theme_id(theme_id: str) -> str:
    if theme_id.startswith("theme:019a"):
        return "19a"
    if theme_id.startswith("theme:019b"):
        return "19b"
    prefix = theme_id.split("_", 1)[0].replace("theme:", "")
    try:
        return f"{int(prefix):02d}"
    except ValueError:
        return prefix


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())

