from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import load_workbook


def test_build_rop_deal_pack_creates_expected_sheets(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    all_results = [
        {
            "matched_contact_id": 1001,
            "matched_lead_id": 2001,
            "match_confidence": 0.95,
            "call_count_for_lead": 3,
            "manager_history": ["Иванов Иван"],
            "interest_summary": "летняя школа",
            "objections_summary": "сроки",
            "current_sales_temperature": "warm",
            "recommended_next_step": "Перезвонить после экзаменов",
            "follow_up_due_at": "2026-04-20",
            "premature_close_risk": "high",
            "close_verdict": "reopen_recommended",
            "close_reason_summary": "Жесткого отказа нет, интерес подтвержден.",
            "phone": "+79990000001",
            "pipeline_name": "Лиды",
            "status_name": "Закрыто и не реализовано",
            "lead_name": "Сделка 1",
            "lead_responsible_user_name": "Иванов Иван",
            "lead_created_at": "2026-04-01 10:00:00",
            "lead_closed_at": "2026-04-10 10:00:00",
            "last_call_at": "2026-04-12 10:00:00",
            "latest_call_type": "sales_call",
            "latest_call_summary": "Клиент просил вернуться позже.",
            "history_summary": "Контакт в истории с 2026-04-01.",
            "chronology": "2026-04-01 - первый контакт | 2026-04-12 - просил вернуться позже",
            "tallanto_id": "T-1",
            "tallanto_match_status": "exact_phone_single",
            "analysis_source": "llm",
            "analysis_mode": "llm_shadow",
            "confidence": 0.91,
            "writeback_allowed": False,
            "writeback_blockers": ["shadow_mode"],
            "dossier": {
                "contact": {"name": "Петров Петр"},
                "lead": {
                    "name": "Сделка 1",
                    "pipeline_name": "Лиды",
                    "status_name": "Закрыто и не реализовано",
                    "responsible_user_name": "Иванов Иван",
                    "custom_fields": {"utm_source": "yandex", "utm_campaign": "123"},
                    "closed_at": "2026-04-10 10:00:00",
                },
                "contact_rollup": {
                    "total_calls_history": 5,
                    "history_summary": "Контакт в истории с 2026-04-01.",
                    "chronology": "2026-04-01 - первый контакт",
                    "tallanto_id": "T-1",
                    "tallanto_match_status": "exact_phone_single",
                    "tallanto_parent_fio": "Петрова Анна",
                    "tallanto_contact": "Петров Петр",
                    "tallanto_owner": "Куратор",
                    "tallanto_student_type": "student",
                    "tallanto_branch": "Филиал 1",
                },
                "call_history": [
                    {
                        "started_at": "2026-04-12 10:00:00",
                        "manager_name": "Иванов Иван",
                        "summary": "Клиент просил вернуться позже.",
                    }
                ],
                "tallanto_live": {
                    "status": "ok",
                    "matched_via": "tallanto_id",
                    "contacts_found": 1,
                    "contexts": [
                        {
                            "contact": {
                                "name": "Петров Петр",
                                "branch": "Филиал 1",
                                "assigned_user_name": "Куратор",
                            },
                            "opportunity_count": 2,
                            "finance_count": 0,
                            "opportunities": [
                                {"name": "Сделка 1", "sales_stage": "ClosedLost"},
                            ],
                        }
                    ],
                },
            },
        },
        {
            "matched_contact_id": 0,
            "matched_lead_id": 2002,
            "premature_close_risk": "manual_review",
            "close_verdict": "manual_review",
            "close_reason_summary": "Нет phone match.",
            "analysis_source": "llm",
            "analysis_mode": "llm_shadow",
            "confidence": 0.5,
            "writeback_allowed": False,
            "writeback_blockers": ["shadow_mode", "manual_review_verdict"],
        },
    ]
    (run_dir / "all_results.json").write_text(json.dumps(all_results, ensure_ascii=False), encoding="utf-8")

    out_path = tmp_path / "rop.xlsx"
    subprocess.run(
        [
            "python3",
            "scripts/build_rop_deal_pack.py",
            "--run-dir",
            str(run_dir),
            "--out",
            str(out_path),
        ],
        cwd=Path(__file__).resolve().parents[1],
        check=True,
    )

    workbook = load_workbook(out_path, read_only=True)
    assert workbook.sheetnames == [
        "Top priorities",
        "Reopen",
        "Follow-up",
        "Manual review",
        "Инструкция для РОПа",
        "Инструкция для менеджера",
    ]
    assert workbook["Reopen"].max_row == 2
    assert workbook["Manual review"].max_row == 2
    headers = [cell.value for cell in next(workbook["Reopen"].iter_rows(min_row=1, max_row=1))]
    assert headers[:12] == [
        "Ссылка на сделку",
        "Ссылка на контакт",
        "Приоритетный балл",
        "AI-риск",
        "AI-вердикт",
        "Телефон",
        "ID сделки amoCRM",
        "ID контакта amoCRM",
        "Сделка",
        "Контакт",
        "Краткая история",
        "Хронология",
    ]
    assert "Уверенность LLM" in headers
    assert "Разрешена запись в AMO" in headers
    assert "Возможностей Tallanto" in headers


def test_build_rop_deal_pack_keeps_full_post_close_history_and_skips_active_client_reopen(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    long_summary = "Очень длинная история после закрытия. " * 40
    all_results = [
        {
            "matched_contact_id": 1001,
            "matched_lead_id": 2001,
            "match_confidence": 0.95,
            "call_count_for_lead": 3,
            "manager_history": ["Иванов Иван"],
            "interest_summary": "летняя школа",
            "objections_summary": "сроки",
            "current_sales_temperature": "warm",
            "recommended_next_step": "Перезвонить позже",
            "follow_up_due_at": "2026-04-20",
            "premature_close_risk": "high",
            "close_verdict": "reopen_recommended",
            "close_reason_summary": "Старый reopen verdict.",
            "phone": "+79990000001",
            "pipeline_name": "Лиды",
            "status_name": "Закрыто и не реализовано",
            "lead_name": "Сделка 1",
            "analysis_source": "llm",
            "analysis_mode": "legacy_shadow",
            "dossier": {
                "contact": {"name": "Петров Петр"},
                "lead": {"name": "Сделка 1", "closed_at": "2026-04-10 10:00:00"},
                "call_history": [
                    {
                        "started_at": "2026-04-12 10:00:00",
                        "manager_name": "Иванов Иван",
                        "summary": long_summary,
                    }
                ],
            },
        },
        {
            "matched_contact_id": 1002,
            "matched_lead_id": 2002,
            "match_confidence": 0.91,
            "call_count_for_lead": 2,
            "manager_history": ["Сидоров Петр"],
            "interest_summary": "лагерь",
            "objections_summary": "сроки",
            "current_sales_temperature": "warm",
            "recommended_next_step": "Перезвонить",
            "follow_up_due_at": "2026-04-21",
            "premature_close_risk": "high",
            "close_verdict": "follow_up_needed",
            "close_reason_summary": "Нужен follow-up.",
            "phone": "+79990000002",
            "pipeline_name": "Лиды",
            "status_name": "Закрыто и не реализовано",
            "lead_name": "Сделка 2",
            "analysis_source": "llm",
            "analysis_mode": "legacy_shadow",
            "dossier": {
                "contact": {"name": "Сидоров Петр"},
                "lead": {"name": "Сделка 2", "closed_at": "2026-04-10 10:00:00"},
                "call_history": [
                    {
                        "started_at": "2026-04-12 10:00:00",
                        "manager_name": "Сидоров Петр",
                        "summary": long_summary,
                    }
                ],
            },
        },
    ]
    (run_dir / "all_results.json").write_text(json.dumps(all_results, ensure_ascii=False), encoding="utf-8")

    lead_snapshot_path = tmp_path / "live_recent_leads.json"
    lead_snapshot_path.write_text(
        json.dumps(
            [
                {
                    "id": 2001,
                    "name": "Сделка 1",
                    "created_at": 1773051353,
                    "updated_at": 1773573823,
                    "closed_at": 1773573823,
                    "pipeline_id": 8938034,
                    "status_id": 143,
                    "responsible_user_id": 1,
                    "custom_fields_values": [
                        {
                            "field_name": "Причина отказа (лид)",
                            "values": [{"value": "Действующий клиент"}],
                        }
                    ],
                    "_embedded": {"contacts": [{"id": 1001}]},
                },
                {
                    "id": 2002,
                    "name": "Сделка 2",
                    "created_at": 1773051353,
                    "updated_at": 1773573823,
                    "closed_at": 1773573823,
                    "pipeline_id": 8938034,
                    "status_id": 143,
                    "responsible_user_id": 1,
                    "custom_fields_values": [],
                    "_embedded": {"contacts": [{"id": 1002}]},
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    out_path = tmp_path / "rop.xlsx"
    subprocess.run(
        [
            "python3",
            "scripts/build_rop_deal_pack.py",
            "--run-dir",
            str(run_dir),
            "--out",
            str(out_path),
            "--lead-snapshot-json",
            str(lead_snapshot_path),
        ],
        cwd=Path(__file__).resolve().parents[1],
        check=True,
    )

    workbook = load_workbook(out_path, read_only=True)
    reopen_rows = list(workbook["Reopen"].iter_rows(values_only=True))
    assert reopen_rows[1][0] == "Нет данных"

    follow_rows = list(workbook["Follow-up"].iter_rows(values_only=True))
    headers = list(follow_rows[0])
    data = dict(zip(headers, follow_rows[1]))
    assert data["История после закрытия"].count("Очень длинная история после закрытия.") >= 20
    assert "..." not in data["История после закрытия"]


def test_build_rop_deal_pack_adds_open_deals_from_full_snapshot(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    all_results = [
        {
            "matched_contact_id": 1001,
            "matched_lead_id": 2001,
            "match_confidence": 0.95,
            "call_count_for_lead": 3,
            "manager_history": ["Иванов Иван"],
            "interest_summary": "летняя школа",
            "objections_summary": "сроки",
            "current_sales_temperature": "warm",
            "recommended_next_step": "Перезвонить позже",
            "follow_up_due_at": "2026-04-20",
            "premature_close_risk": "high",
            "close_verdict": "reopen_recommended",
            "close_reason_summary": "Жесткого отказа нет.",
            "phone": "+79990000001",
            "pipeline_name": "Лиды",
            "status_name": "Закрыто и не реализовано",
            "lead_name": "Сделка 1",
            "analysis_source": "llm",
            "analysis_mode": "legacy_shadow",
            "dossier": {
                "contact": {"name": "Петров Петр"},
                "lead": {"name": "Сделка 1", "closed_at": "2026-04-10 10:00:00"},
            },
        }
    ]
    (run_dir / "all_results.json").write_text(json.dumps(all_results, ensure_ascii=False), encoding="utf-8")

    lead_snapshot_path = tmp_path / "live_recent_leads.json"
    lead_snapshot_path.write_text(
        json.dumps(
            [
                {
                    "id": 2001,
                    "name": "Сделка 1",
                    "created_at": 1773051353,
                    "updated_at": 1773573823,
                    "closed_at": 1773573823,
                    "pipeline_id": 8938034,
                    "status_id": 143,
                    "responsible_user_id": 1,
                    "custom_fields_values": [],
                    "_embedded": {"contacts": [{"id": 1001}]},
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    full_snapshot_path = tmp_path / "all_leads.json"
    full_snapshot_path.write_text(
        json.dumps(
            [
                {
                    "id": 2001,
                    "name": "Сделка 1",
                    "updated_at": 1773573823,
                    "pipeline_id": 8938034,
                    "status_id": 143,
                    "_embedded": {"contacts": [{"id": 1001}]},
                },
                {
                    "id": 3001,
                    "name": "Открытая сделка",
                    "updated_at": 1773574823,
                    "pipeline_id": 10408062,
                    "status_id": 82258198,
                    "_embedded": {"contacts": [{"id": 1001}]},
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    out_path = tmp_path / "rop.xlsx"
    subprocess.run(
        [
            "python3",
            "scripts/build_rop_deal_pack.py",
            "--run-dir",
            str(run_dir),
            "--out",
            str(out_path),
            "--lead-snapshot-json",
            str(lead_snapshot_path),
            "--full-lead-snapshot-json",
            str(full_snapshot_path),
        ],
        cwd=Path(__file__).resolve().parents[1],
        check=True,
    )

    workbook = load_workbook(out_path, read_only=True)
    reopen_rows = list(workbook["Reopen"].iter_rows(values_only=True))
    headers = list(reopen_rows[0])
    data = dict(zip(headers, reopen_rows[1]))
    assert data["Открытых сделок у контакта"] == 1
    assert "3001: Открытая сделка" in data["Открытые сделки контакта"]


def test_build_rop_deal_pack_prefers_fresh_full_history_over_clipped_legacy_text(tmp_path: Path) -> None:
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    all_results = [
        {
            "matched_contact_id": 1001,
            "matched_lead_id": 2001,
            "match_confidence": 0.95,
            "call_count_for_lead": 3,
            "manager_history": ["Иванов Иван"],
            "current_sales_temperature": "warm",
            "recommended_next_step": "Перезвонить позже",
            "premature_close_risk": "high",
            "close_verdict": "reopen_recommended",
            "close_reason_summary": "Жесткого отказа нет.",
            "phone": "+79990000001",
            "pipeline_name": "Лиды",
            "status_name": "Закрыто и не реализовано",
            "lead_name": "Сделка 1",
            "analysis_source": "llm",
            "analysis_mode": "legacy_shadow",
            "latest_call_summary": "Старый срезанный summary…",
            "history_summary": "Старый срезанный history…",
            "chronology": "Старый срезанный chronology…",
            "dossier": {
                "contact": {"name": "Петров Петр"},
                "lead": {"name": "Сделка 1", "closed_at": "2026-04-10 10:00:00"},
                "contact_rollup": {
                    "history_summary": "Еще один старый history…",
                    "chronology": "Еще один старый chronology…",
                },
            },
        }
    ]
    (run_dir / "all_results.json").write_text(json.dumps(all_results, ensure_ascii=False), encoding="utf-8")

    lead_snapshot_path = tmp_path / "live_recent_leads.json"
    lead_snapshot_path.write_text(
        json.dumps(
            [
                {
                    "id": 2001,
                    "name": "Сделка 1",
                    "updated_at": 1773573823,
                    "closed_at": 1773573823,
                    "pipeline_id": 8938034,
                    "status_id": 143,
                    "custom_fields_values": [],
                    "_embedded": {"contacts": [{"id": 1001}]},
                },
            ],
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    master_export_dir = tmp_path / "master_export"
    master_export_dir.mkdir()
    (master_export_dir / "master_contacts_ru.csv").write_text(
        "\n".join(
            [
                "Телефон клиента,Краткая история общения,Хронология общения (последние 5 касаний),Краткое резюме последнего свежего звонка",
                '"+79990000001","Полная история без обрезания и с полезным хвостом","2026-04-01 — Полная хронология без обрезания","Полное последнее резюме без обрезания"',
            ]
        ),
        encoding="utf-8",
    )
    (master_export_dir / "master_calls_ru.csv").write_text(
        "Телефон клиента,Дата и время звонка\n+79990000001,2026-04-12 10:00:00\n",
        encoding="utf-8",
    )

    out_path = tmp_path / "rop.xlsx"
    subprocess.run(
        [
            "python3",
            "scripts/build_rop_deal_pack.py",
            "--run-dir",
            str(run_dir),
            "--out",
            str(out_path),
            "--lead-snapshot-json",
            str(lead_snapshot_path),
            "--master-export-dir",
            str(master_export_dir),
        ],
        cwd=Path(__file__).resolve().parents[1],
        check=True,
    )

    workbook = load_workbook(out_path, read_only=True)
    rows = list(workbook["Reopen"].iter_rows(values_only=True))
    headers = list(rows[0])
    data = dict(zip(headers, rows[1]))
    assert data["Краткая история"] == "Полная история без обрезания и с полезным хвостом"
    assert data["Хронология"] == "2026-04-01 — Полная хронология без обрезания"
    assert data["Последнее резюме звонка"] == "Полное последнее резюме без обрезания"
