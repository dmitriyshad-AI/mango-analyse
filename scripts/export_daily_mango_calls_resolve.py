#!/usr/bin/env python3
"""Read-only daily export of Mango calls for the head of sales."""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import sqlite3
import tempfile
import unicodedata
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Mapping, Sequence
from urllib.parse import quote
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import dotenv_values

from mango_mvp.amocrm_runtime.tallanto_api import (
    TallantoApiClient,
    TallantoApiConfig,
    TallantoApiError,
)
from mango_mvp.productization.mango_office_client import MangoOfficeClient, MangoOfficeCredentials
from mango_mvp.services.export_excel import call_to_row
from mango_mvp.utils.filename_repair import repair_manager_name
from mango_mvp.utils.phone import normalize_phone


ROOT = Path(__file__).resolve().parents[1]
PIPELINE_ROOT = Path(os.getenv("MANGO_CALLS_PIPELINE_ROOT", str(ROOT / "product_data/mango_calls_two_processes"))).expanduser()
DEFAULT_READY_DB = PIPELINE_ROOT / "drop/mango_calls_ready.sqlite"
DEFAULT_WORKING_DB = PIPELINE_ROOT / "working/mango_calls_pipeline.sqlite"
DEFAULT_OUT = Path(os.getenv("MANGO_CALLS_DAILY_EXPORT_OUT", str(Path.home() / "Yandex.Disk.localized/Mango Calls Resolve"))).expanduser()
DEFAULT_TALLANTO_EXPORT = Path(os.getenv("MANGO_CALLS_TALLANTO_EXPORT", str(ROOT / "_external_handoffs/tallanto_contacts_export_2026-06-20/converted/Contacts 20.06.2026.csv"))).expanduser()
DEFAULT_TALLANTO_ENV, DEFAULT_MANGO_ENV = Path(os.getenv("MANGO_CALLS_TALLANTO_ENV", "~/.mango_secrets/tallanto_readonly.env")).expanduser(), Path(os.getenv("MANGO_CALLS_MANGO_ENV", "~/.mango_secrets/mango_office.env")).expanduser()
DEFAULT_MANAGER_USERS = ROOT / (
    "_local_archive_mango_api_downloads_20260507/quarantine_import/"
    "raw_payload_archive/mango_users_config_20260507.json"
)
MOSCOW = ZoneInfo("Europe/Moscow")
TRANSCRIPT_CHUNK, EXPORT_SCHEMA_VERSION = 30_000, "daily_mango_calls_resolve_export_v3"
TIMED_LINE_RE = re.compile(
    r"^\[(?P<approx>~)?(?P<mm>\d{2}):(?P<ss>\d{2}(?:\.\d)?)\]\s+"
    r"(?P<speaker>Менеджер(?:\s*\([^)]+\))?|Клиент|Спикер\s*\(не определен\)):\s*(?P<text>.*)$"
)

CALL_TYPE_RU = {
    "sales_call": "Продажа / подбор обучения",
    "service_call": "Сервисный вопрос",
    "existing_client_progress": "Текущий клиент / продолжение",
    "technical_call": "Технический вопрос",
    "non_conversation": "Разговор не состоялся",
}
STATUS_RU = {"done": "Готово", "skipped": "Пропущено по правилу", "manual": "Нужна ручная проверка", "pending": "Ожидает", "failed": "Ошибка", "in_progress": "В работе"}
PRICE_RU = {"high": "Высокая", "medium": "Средняя", "low": "Низкая"}
CHANNEL_RU = {"phone": "Телефон", "email": "Электронная почта", "telegram": "Telegram", "whatsapp": "WhatsApp"}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def readonly_uri(path: Path, *, immutable: bool = False) -> str:
    option = "&immutable=1" if immutable else ""
    return f"file:{quote(str(path.resolve()), safe='/:')}?mode=ro{option}"


def verify_ready_drop(db: Path) -> dict[str, Any]:
    manifest_path = db.with_suffix(".manifest.json")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    actual = {"sha256": sha256_file(db), "size_bytes": db.stat().st_size}
    with sqlite3.connect(readonly_uri(db, immutable=True), uri=True) as con:
        actual["quick_check"] = str(con.execute("PRAGMA quick_check").fetchone()[0])
    expected = (manifest.get("sha256"), int(manifest.get("size_bytes") or 0), manifest.get("quick_check"), manifest.get("status"))
    observed = (actual["sha256"], actual["size_bytes"], actual["quick_check"], "ready")
    if expected != observed:
        raise RuntimeError("готовая база и её контрольный файл не совпадают")
    return {**actual, "published_at": manifest.get("published_at")}


def _manager_map(users: Sequence[Mapping[str, Any]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in users:
        extension = str((item.get("telephony") or {}).get("extension") or "").strip()
        name = repair_manager_name((item.get("general") or {}).get("name"))
        if extension and name and name.casefold() != "admin":
            result[extension] = name
    return result


def load_manager_map(path: Path | None, current_users: Sequence[Mapping[str, Any]] = ()) -> dict[str, str]:
    archived: list[Mapping[str, Any]] = []
    if path is not None and path.is_file():
        archived = json.loads(path.read_text(encoding="utf-8")).get("users", [])
    return {**_manager_map(archived), **_manager_map(current_users)}


def fetch_mango_users(env_path: Path) -> list[dict[str, Any]]:
    values = dotenv_values(env_path)
    client = MangoOfficeClient(
        MangoOfficeCredentials(str(values.get("MANGO_OFFICE_API_KEY") or ""), str(values.get("MANGO_OFFICE_API_SALT") or "")),
        base_url=str(values.get("MANGO_OFFICE_BASE_URL") or "https://app.mango-office.ru"),
    )
    payload = client.post_command("/vpbx/config/users/request", {})
    users = payload.get("users") if isinstance(payload, dict) else None
    return [item for item in users or [] if isinstance(item, dict)]


def manager_name_issue(value: Any) -> str:
    name = " ".join(str(value or "").split())
    if not name:
        return "ФИО менеджера не найдено в справочнике Mango"
    return "В справочнике Mango указано неполное имя менеджера" if len(name.split()) < 2 else ""


def parse_json(value: Any) -> dict[str, Any]:
    try:
        parsed = json.loads(value or "{}")
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, dict) else {}


def day_bounds_utc(day: date) -> tuple[str, str]:
    start = datetime.combine(day, time.min, MOSCOW)
    end = start + timedelta(days=1)
    return tuple(value.astimezone(timezone.utc).replace(tzinfo=None).isoformat(sep=" ") for value in (start, end))  # type: ignore[return-value]


def read_day(db: Path, day: date, *, immutable: bool) -> list[sqlite3.Row]:
    with sqlite3.connect(readonly_uri(db, immutable=immutable), uri=True, timeout=30) as con:
        con.row_factory = sqlite3.Row
        start, end = day_bounds_utc(day)
        return con.execute(
            "SELECT * FROM call_records WHERE started_at>=? AND started_at<? ORDER BY started_at,id",
            (start, end),
        ).fetchall()


def translate_transcript(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"(?m)^\s*MANAGER\s*:", "Менеджер:", text, flags=re.IGNORECASE)
    return re.sub(r"(?m)^\s*CLIENT\s*:", "Клиент:", text, flags=re.IGNORECASE)


def ordered_dialogue(source: Path, variants: Mapping[str, Any], fallback: str, *, allow_file_fallback: bool = True) -> tuple[str, bool]:
    raw_lines = variants.get("dialogue_lines")
    if allow_file_fallback and (not isinstance(raw_lines, list) or not raw_lines):
        exported = PIPELINE_ROOT / "working/transcripts" / source.parent.name / f"{source.stem}_text.txt"
        raw_lines = exported.read_text(encoding="utf-8", errors="ignore").splitlines() if exported.is_file() else []
    if not isinstance(raw_lines, list):
        raw_lines = []
    source_lines = [str(item).strip() for item in raw_lines if str(item).strip()]
    lines, previous, previous_role, preserve_source = [], -1.0, "", False
    for raw in source_lines:
        match = TIMED_LINE_RE.fullmatch(raw)
        if match is None or match.group("approx") or match.group("speaker").startswith("Спикер"):
            lines = []
            break
        content = match.group("text").strip()
        if not "".join(char for char in content if unicodedata.category(char) != "Cf").strip():
            continue
        stamp = int(match.group("mm")) * 60 + float(match.group("ss"))
        speaker = match.group("speaker")
        role = "Менеджер" if speaker.startswith("Менеджер") else "Клиент" if speaker == "Клиент" else "Спикер (не определён)"
        if stamp < previous or (stamp == previous and previous_role and role != previous_role):
            lines, preserve_source = [], True
            break
        previous, previous_role = stamp, role
        lines.append(f"[{match.group('mm')}:{match.group('ss')}] {role}: {content}")
    if lines:
        return "\n".join(lines), True
    warning = "Порядок реплик не сохранён в исходных данных; ниже приведён полный текст по ролям без выдуманной очередности."
    preserved = "\n".join(source_lines) if preserve_source else translate_transcript(fallback).strip() or "\n".join(source_lines)
    return f"{warning}\n\n{preserved}".strip(), False


def manager_roles_confirmed(variants: Mapping[str, Any]) -> bool:
    mapping = variants.get("role_mapping")
    channels = {(variants.get(role) if isinstance(variants.get(role), Mapping) else {}).get("physical_channel") for role in ("manager", "client")}
    return bool(
        isinstance(mapping, Mapping)
        and mapping.get("confirmed") is True
        and mapping.get("manager_quality_allowed") is True
        and mapping.get("topology") == "simple_two_party"
        and variants.get("call_topology") == "simple_two_party"
        and channels == {"left", "right"}
    )


def neutralize_unconfirmed_roles(text: str) -> str:
    text = re.sub(r"(?m)(Менеджер(?:\s*\([^)]+\))?):", "Спикер A (роль не подтверждена):", text)
    return re.sub(r"(?m)Клиент:", "Спикер B (роль не подтверждена):", text)


def load_tallanto_index(path: Path) -> dict[str, dict[str, str]]:
    index: dict[str, dict[str, str]] = {}
    if not path.is_file():
        return index
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            contact_id = str(row.get("ID") or "").strip()
            card_name = " ".join(filter(None, (str(row.get("Фамилия") or "").strip(), str(row.get("Имя") or "").strip())))
            parent_name = " ".join(str(row.get("ФИО родителя") or "").split())
            for column, name in (("Тел. (родителя)", parent_name or card_name), ("Тел. (доп.)", card_name)):
                phone = normalize_phone(row.get(column))
                if phone and contact_id:
                    index.setdefault(phone, {})[contact_id] = name
    return index


def build_tallanto_client(env_path: Path) -> TallantoApiClient:
    values = dotenv_values(env_path)
    if values.get("TALLANTO_BATCH_FETCH") is not None:
        os.environ.setdefault("TALLANTO_BATCH_FETCH", str(values["TALLANTO_BATCH_FETCH"]))
    rest_path = str(values.get("CRM_TALLANTO_STUDENT_PATH") or "/service/api/rest.php")
    if "{student_id}" in rest_path:
        rest_path = "/service/api/rest.php"
    return TallantoApiClient(TallantoApiConfig(
        base_url=str(values.get("CRM_TALLANTO_BASE_URL") or ""),
        api_token=str(values.get("CRM_TALLANTO_API_TOKEN") or ""),
        rest_path=rest_path,
    ))


def _api_contact_index(records: Sequence[Mapping[str, Any]], wanted: set[str]) -> dict[str, dict[str, str]]:
    index: dict[str, dict[str, str]] = {}
    for record in records:
        phones = {
            phone for field in TallantoApiClient.CONTACT_PHONE_FIELDS
            if (phone := normalize_phone(record.get(field))) in wanted
        }
        contact_id = str(record.get("id") or "").strip()
        if not phones or not contact_id:
            continue
        name = str(record.get("name") or "").strip() or " ".join(
            str(record.get(field) or "").strip() for field in ("last_name", "first_name", "middle_name") if str(record.get(field) or "").strip()
        )
        for phone in phones:
            index.setdefault(phone, {})[contact_id] = " ".join(name.split())
    return index


def load_tallanto_api_changes(
    client: TallantoApiClient,
    wanted: set[str],
    modified_after: str,
) -> tuple[dict[str, dict[str, str]], bool]:
    index: dict[str, dict[str, str]] = {}
    offset = 0
    try:
        while True:
            payload = client.get_entry_list(
                module="Contact",
                select_fields=("first_name", "last_name", "phone_mobile", "phone_work", "date_modified"),
                order_by="date_modified DESC",
                offset=offset,
            )
            records = [record for record in payload.get("entry_list", []) if isinstance(record, dict)]
            if not records:
                return index, True
            recent = [record for record in records if not record.get("date_modified") or str(record["date_modified"]) > modified_after]
            for phone, contacts in _api_contact_index(recent, wanted).items():
                index.setdefault(phone, {}).update(contacts)
            if len(recent) < len(records):
                return index, True
            next_offset = payload.get("next_offset")
            if next_offset in (None, "", offset):
                return index, True
            offset = int(next_offset)
    except TallantoApiError:
        return index, False


def apply_tallanto_names(
    rows: Sequence[dict[str, Any]], export_path: Path, client: TallantoApiClient | None,
    *, snapshot_as_of: datetime | None = None,
) -> None:
    if not export_path.is_file():
        raise RuntimeError(f"выгрузка Tallanto не найдена: {export_path}")
    local = load_tallanto_index(export_path)
    missing = {
        phone for row in rows
        if (phone := normalize_phone(row["phone"])) and not local.get(phone)
    }
    api_index: dict[str, dict[str, str]] = {}
    api_complete = True
    if missing and client is not None:
        snapshot = snapshot_as_of or datetime.fromtimestamp(export_path.stat().st_mtime, MOSCOW)
        if snapshot.tzinfo is None:
            snapshot = snapshot.replace(tzinfo=MOSCOW)
        modified_after = snapshot.astimezone(MOSCOW).strftime("%Y-%m-%d %H:%M:%S")
        api_index, api_complete = load_tallanto_api_changes(client, missing, modified_after)
    for row in rows:
        phone = normalize_phone(row["phone"])
        contacts, source = local.get(phone or "", {}), "выгрузка Tallanto"
        if phone and not contacts and client is not None:
            contacts, source = api_index.get(phone, {}), "Tallanto API"
        if len(contacts) == 1 and next(iter(contacts.values()), ""):
            row["client_fio"], row["tallanto_source"] = next(iter(contacts.values())), source
        elif len(contacts) > 1:
            row["issues"].append("Телефон совпал с несколькими карточками Tallanto")
        elif phone in missing and not api_complete:
            row["issues"].append("Не удалось проверить телефон через Tallanto API")
        else:
            row["issues"].append("Телефон не найден в Tallanto")
        row["manager_ready"] = bool(row.get("complete") and row.get("chronology_confirmed") and not row["issues"])


def clean_summary(value: Any) -> str:
    text = " ".join(str(value or "").split())
    text = re.sub(
        r"^\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2}\s+менеджер\s+\S+\s+(?:общался с клиентом|пытался связаться с клиентом)\.\s*",
        "",
        text,
        flags=re.IGNORECASE,
    )
    text = re.sub(r"(?:^|\s)Приоритет лида:\s*[^.]+\.\s*", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"(?:^|\s)Итог:\s*Оценка на основе содержания звонка\.\s*", " ", text, flags=re.IGNORECASE)
    return " ".join(text.split())


def processing_issues(row: sqlite3.Row, analysis: Mapping[str, Any], resolve: Mapping[str, Any]) -> list[str]:
    flags = analysis.get("quality_flags") if isinstance(analysis.get("quality_flags"), dict) else {}
    issues: list[str] = []
    if row["resolve_status"] not in {"done", "skipped"}:
        issues.append("Разделение ролей не завершено автоматически")
    if row["analysis_status"] != "done":
        issues.append("Смысловой анализ не готов")
    if analysis.get("needs_review"):
        issues.append("Смысловой анализ запросил ручную проверку")
    if flags.get("transcript_quality_requires_manual_review"):
        issues.append("Качество расшифровки требует проверки")
    if resolve.get("decision") == "manual_review_required" or row["resolve_status"] == "manual":
        issues.append("Нужно вручную проверить разделение реплик")
    return list(dict.fromkeys(issues))


def normalize_row(row: sqlite3.Row, names: Mapping[str, str], *, sealed_only: bool = False) -> dict[str, Any]:
    analysis, resolve, variants = parse_json(row["analysis_json"]), parse_json(row["resolve_json"]), parse_json(row["transcript_variants_json"])
    raw = dict(row)
    started_utc = datetime.fromisoformat(str(row["started_at"])).replace(tzinfo=timezone.utc)
    raw["started_at"] = started_utc
    base = call_to_row(SimpleNamespace(**raw), analysis) if analysis else {}
    started = started_utc.astimezone(MOSCOW)
    transcript, order_confirmed = ordered_dialogue(
        Path(str(row["source_file"])), variants, str(row["transcript_text"] or ""),
        allow_file_fallback=not sealed_only,
    )
    roles_confirmed = manager_roles_confirmed(variants)
    chronology_confirmed = order_confirmed and roles_confirmed
    extension = str(row["manager_name"] or "").strip()
    manager = names.get(extension, "")
    resolve_ok = row["resolve_status"] == "done" or (row["resolve_status"] == "skipped" and resolve.get("decision") == "skip_short_call")
    complete = bool(row["transcription_status"] == "done" and resolve_ok and row["analysis_status"] == "done" and analysis)
    issues = processing_issues(row, analysis, resolve)
    if not order_confirmed:
        issues.append("Порядок реплик не подтверждён исходными данными")
    if not roles_confirmed:
        issues.append("Роли менеджера и клиента не подтверждены")
        transcript = "Роли не подтверждены; не использовать для оценки сотрудника.\n\n" + neutralize_unconfirmed_roles(transcript)
    base = base if chronology_confirmed else {}
    if manager_issue := manager_name_issue(manager):
        issues.append(manager_issue)
    chunks = [transcript[i : i + TRANSCRIPT_CHUNK] for i in range(0, len(transcript), TRANSCRIPT_CHUNK)] or [""]
    return {
        "id": int(row["id"]), "call_id": str(row["source_call_id"] or row["id"]), "started": started,
        "extension": extension, "manager": manager, "direction": "Входящий" if row["direction"] == "inbound" else "Исходящий",
        "phone": str(row["phone"] or ""), "duration": float(row["duration_sec"] or 0), "source": Path(str(row["source_file"])),
        "complete": complete, "manager_ready": complete and chronology_confirmed and not issues, "issues": issues, "chunks": chunks, "transcript": transcript, "chronology_confirmed": chronology_confirmed,
        "base": base, "client_fio": "", "tallanto_source": "",
        "resolve_status": str(row["resolve_status"] or ""), "analysis_status": str(row["analysis_status"] or ""),
    }


def merged_day_rows(ready_db: Path, working_db: Path, day: date, names: Mapping[str, str], *, sealed_only: bool = False) -> tuple[list[dict[str, Any]], int]:
    ready = read_day(ready_db, day, immutable=True)
    working = [] if sealed_only else read_day(working_db, day, immutable=False)
    ready_ids = {str(row["source_call_id"] or row["id"]) for row in ready}
    pending = [row for row in working if str(row["source_call_id"] or row["id"]) not in ready_ids]
    merged = [normalize_row(row, names, sealed_only=sealed_only) for row in [*ready, *pending]]
    merged.sort(key=lambda item: (item["started"], item["id"]))
    return merged, len(pending)


def assign_transcript_targets(rows: Sequence[dict[str, Any]], target: Path) -> None:
    filenames: set[str] = set()
    for row in rows:
        filename = f"call_{hashlib.sha256(str(row['call_id']).encode('utf-8')).hexdigest()[:20]}.txt"
        if filename in filenames:
            raise RuntimeError("duplicate call_id in daily export")
        filenames.add(filename)
        body = row["transcript"].rstrip() + "\n"
        row["transcript_file"] = target / filename
        row["transcript_sha256"] = hashlib.sha256(body.encode("utf-8")).hexdigest()


def publish_transcripts(rows: Sequence[dict[str, Any]], target: Path) -> tuple[int, int, int]:
    expected = {Path(row["transcript_file"]).name: row["transcript_sha256"] for row in rows}
    if target.is_dir():
        actual = {path.name: sha256_file(path) if path.is_file() else "" for path in target.iterdir()}
        if actual != expected:
            raise RuntimeError("existing immutable transcript generation is inconsistent")
        return 0, len(rows), 0
    target.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    staging = Path(tempfile.mkdtemp(prefix=f".{target.name}.staging-", dir=target.parent))
    try:
        for row in rows:
            body = row["transcript"].rstrip() + "\n"
            destination = staging / Path(row["transcript_file"]).name
            destination.write_text(body, encoding="utf-8")
            destination.chmod(0o600)
        os.replace(staging, target)
    except Exception:
        for path in staging.glob("*") if staging.is_dir() else ():
            path.unlink(missing_ok=True)
        staging.rmdir() if staging.is_dir() else None
        raise
    return len(rows), 0, 0


def publication_content_sha256(rows: Sequence[Mapping[str, Any]]) -> str:
    payload = [{
        "call_id": row["call_id"], "started": row["started"].isoformat(), "manager": row["manager"],
        "extension": row["extension"], "direction": row["direction"], "phone": row["phone"],
        "duration": row["duration"], "complete": row["complete"], "manager_ready": row["manager_ready"],
        "issues": row["issues"], "transcript": row["transcript"], "base": row["base"],
        "client_fio": row["client_fio"], "tallanto_source": row["tallanto_source"],
    } for row in rows]
    document = {"schema_version": EXPORT_SCHEMA_VERSION, "rows": payload}
    return hashlib.sha256(json.dumps(document, ensure_ascii=False, sort_keys=True, default=str).encode("utf-8")).hexdigest()


def reusable_export(output_root: Path, day: date, content_sha256: str, row_count: int) -> Mapping[str, Any] | None:
    manifest_path = output_root / f"Отчёт РОП по звонкам {day.isoformat()}.manifest.json"
    if not manifest_path.is_file():
        return None
    manifest = parse_json(manifest_path.read_text(encoding="utf-8"))
    xlsx = output_root / str(manifest.get("xlsx") or "")
    transcript_dir = output_root / str(manifest.get("transcript_dir") or "")
    transcripts = manifest.get("transcripts") if isinstance(manifest.get("transcripts"), list) else []
    if manifest.get("schema_version") != EXPORT_SCHEMA_VERSION or manifest.get("content_sha256") != content_sha256:
        return None
    if not xlsx.is_file() or sha256_file(xlsx) != manifest.get("xlsx_sha256"): raise RuntimeError("existing immutable XLSX generation is inconsistent")
    if len(transcripts) != row_count or any(not isinstance(item, Mapping) for item in transcripts):
        return None
    if any(not (transcript_dir / str(item.get("file") or "")).is_file() or sha256_file(transcript_dir / str(item["file"])) != item.get("sha256") for item in transcripts):
        raise RuntimeError("existing immutable transcript generation is inconsistent")
    expected = {str(item["file"]) for item in transcripts}
    if transcript_dir.is_dir() and {path.name for path in transcript_dir.iterdir()} != expected:
        raise RuntimeError("unexpected transcript files in current daily package")
    return {**manifest, "transcripts_copied": 0, "transcripts_reused": row_count, "transcripts_updated": 0, "reused": True, "xlsx": str(xlsx), "transcript_dir": str(transcript_dir), "manifest": str(manifest_path)}


def workbook_rows(rows: Sequence[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    parts = max((len(row["chunks"]) for row in rows), default=1)
    transcript_headers = [f"Расшифровка разговора, часть {i}" for i in range(1, parts + 1)]
    headers = [
        "Дата и время", "ФИО менеджера", "Добавочный номер", "Направление", "Телефон клиента", "ФИО клиента из Tallanto",
        "Источник сопоставления клиента", "Длительность, сек", "Тип звонка по смысловому анализу", "Краткое содержание разговора",
        *transcript_headers, "Файл полной расшифровки", "Продукт", "Предметы", "Формат", "Целевые экзамены", "Класс", "Возражения и ограничения",
        "Следующий шаг", "Срок следующего шага", "Предпочтительный канал", "Озвученный бюджет",
        "Чувствительность к цене", "Интерес к скидке", "Нужна проверка", "Причина проверки",
    ]
    values: list[list[Any]] = []
    for row in rows:
        base = row["base"]
        issues = "; ".join(row["issues"])
        values.append([
            row["started"].replace(tzinfo=None), row["manager"], row["extension"], row["direction"], row["phone"], row["client_fio"], row["tallanto_source"],
            round(row["duration"], 1), CALL_TYPE_RU.get(base.get("call_type", ""), base.get("call_type", "")), clean_summary(base.get("history_summary", "")),
        ] + row["chunks"] + [""] * (parts - len(row["chunks"])) + [row["transcript_file"].name,
            base.get("interests_products") or base.get("recommended_product", ""), base.get("interests_subjects", ""),
            base.get("interests_format", ""), base.get("exam_targets", ""), base.get("grade_current", ""),
            base.get("objections", ""), base.get("next_step_action", ""), base.get("next_step_due_raw", ""),
            CHANNEL_RU.get(base.get("preferred_channel", ""), base.get("preferred_channel", "")), base.get("budget", ""),
            PRICE_RU.get(base.get("price_sensitivity", ""), base.get("price_sensitivity", "")),
            "Да" if str(base.get("discount_interest", "")).casefold() in {"true", "yes", "да", "1"} else "",
            "Да" if issues else "Нет", issues,
        ])
    return headers, values


def append_safe(sheet: Any, values: Sequence[Any]) -> None:
    if any(isinstance(value, str) and len(value) > 32_767 for value in values):
        raise RuntimeError("значение Excel превышает допустимый размер; обрезка запрещена")
    sheet.append([ILLEGAL_CHARACTERS_RE.sub("", value) if isinstance(value, str) else value for value in values])
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str) and cell.value.startswith(("=", "+", "-", "@")):
            cell.data_type = "s"


def format_sheet(sheet: Any, *, table: bool = True) -> None:
    if table:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
    style_header(sheet[1])
    for column in sheet.columns:
        header = str(column[0].value or "")
        width = 65 if "Расшифровка" in header else 42 if header in {"Краткое содержание разговора", "Возражения и ограничения", "Следующий шаг", "Причина проверки"} else 18
        sheet.column_dimensions[column[0].column_letter].width = width
        for cell in column[1:]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, datetime):
                cell.number_format = "dd.mm.yyyy hh:mm:ss"


def style_header(cells: Sequence[Any]) -> None:
    for cell in cells:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_workbook(path: Path, day: date, rows: Sequence[dict[str, Any]], manager_source: str, source_meta: Mapping[str, Any]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ready, manager_ready, unfinished = [row for row in rows if row["complete"]], [row for row in rows if row["manager_ready"]], [row for row in rows if not row["complete"]]
    summary = wb.create_sheet("Сводка")
    for line in (["Ежедневный отчёт РОПа", day.isoformat()], ["Всего звонков", len(rows)], ["Полностью обработано", len(ready)],
                 ["Обработка не завершена", len(unfinished)], ["Требуют проверки", sum(bool(row["issues"]) for row in rows)],
                 ["Допущено к оценке менеджера", len(manager_ready)],
                 ["Порядок реплик подтверждён", sum(row["chronology_confirmed"] for row in rows)],
                 ["ФИО найдено в Tallanto", sum(bool(row["client_fio"]) for row in rows)],
                 ["ФИО менеджера неполное или не найдено", sum(bool(manager_name_issue(row["manager"])) for row in rows)],
                 ["Источник ФИО менеджеров", manager_source or "не задан"], ["Готовый снимок опубликован", source_meta.get("published_at") or ""]):
        append_safe(summary, line)
    summary.append([])
    append_safe(summary, ["Менеджер", "Допущенных к оценке звонков", "Часов"])
    manager_header_row = summary.max_row
    counts = Counter((row["manager"] or f"Добавочный {row['extension']}") for row in manager_ready)
    for manager, count in counts.most_common():
        hours = sum(row["duration"] for row in manager_ready if (row["manager"] or f"Добавочный {row['extension']}") == manager) / 3600
        append_safe(summary, [manager, count, round(hours, 2)])
    for title, subset in (("Звонки", manager_ready), ("Проблемы данных", [row for row in rows if row["issues"]])):
        sheet = wb.create_sheet(title)
        headers, values = workbook_rows(subset)
        append_safe(sheet, headers)
        for value_row, source_row in zip(values, subset):
            append_safe(sheet, value_row)
            transcript_cell = sheet.cell(sheet.max_row, headers.index("Файл полной расшифровки") + 1)
            transcript_cell.hyperlink = source_row["transcript_file"].relative_to(path.parent).as_posix()
            transcript_cell.style = "Hyperlink"
        format_sheet(sheet)
    description = wb.create_sheet("Описание полей")
    for line in (["Правило", "Описание"], ["Период", "Полные календарные сутки по Москве."],
                 ["ФИО менеджера", "Из архивного справочника Mango; пустое значение означает, что подтверждённого соответствия нет."],
                 ["ФИО клиента", "Однозначное совпадение нормализованного телефона: сначала локальная выгрузка Tallanto, затем read-only Tallanto API."],
                 ["Расшифровка", "Полный последовательный диалог только при сохранённом порядке реплик. Длинный текст без обрезки разбит на соседние столбцы и продублирован в TXT."],
                 ["Следующий шаг и возражения", "Подсказка смыслового анализа, а не автоматическое поручение; сверить с записью и CRM."],
                 ["Незавершённые", "На листе «Проблемы данных»; не смешиваются с полностью обработанными звонками."],
                 ["Использование", "Внутренний отчёт. Не применять для санкций или KPI без прослушивания и контекста CRM."]):
        append_safe(description, line)
    format_sheet(summary, table=False)
    style_header(summary[manager_header_row])
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 36
    summary.column_dimensions["C"].width = 12
    format_sheet(description)
    temporary = path.with_suffix(".tmp.xlsx")
    try:
        wb.save(temporary)
        checked = load_workbook(temporary, read_only=True, data_only=False)
        try:
            if checked.sheetnames != ["Сводка", "Звонки", "Проблемы данных", "Описание полей"] or checked["Звонки"].max_row != len(manager_ready) + 1:
                raise RuntimeError("проверка созданной таблицы не пройдена")
        finally:
            checked.close()
        os.replace(temporary, path)
    finally:
        temporary.unlink(missing_ok=True)
    path.chmod(0o600)


def export_day(
    ready_db: Path,
    working_db: Path,
    output_root: Path,
    day: date,
    manager_users: Path | None,
    *,
    tallanto_export: Path = DEFAULT_TALLANTO_EXPORT,
    tallanto_env: Path = DEFAULT_TALLANTO_ENV,
    tallanto_snapshot_as_of: datetime | None = None,
    tallanto_client: TallantoApiClient | None = None,
    current_manager_users: Sequence[Mapping[str, Any]] = (),
    sealed_only: bool = False,
) -> Mapping[str, Any]:
    if day >= datetime.now(MOSCOW).date():
        raise ValueError("можно выгружать только завершённые сутки по Москве")
    source_before = verify_ready_drop(ready_db)
    rows, working_only = merged_day_rows(
        ready_db, working_db, day, load_manager_map(manager_users, current_manager_users), sealed_only=sealed_only,
    )
    client = tallanto_client or build_tallanto_client(tallanto_env)
    apply_tallanto_names(rows, tallanto_export, client, snapshot_as_of=tallanto_snapshot_as_of)
    output_root.mkdir(parents=True, exist_ok=True, mode=0o700)
    output_root.chmod(0o700)
    content_sha256 = publication_content_sha256(rows)
    source_after = verify_ready_drop(ready_db)
    if source_before["sha256"] != source_after["sha256"]:
        raise RuntimeError("готовая база изменилась во время выгрузки; повторите запуск")
    if reused := reusable_export(output_root, day, content_sha256, len(rows)):
        return reused
    generation, transcript_dir = content_sha256[:12], output_root / f"Расшифровки разговоров {day.isoformat()} v3-{content_sha256[:12]}"
    assign_transcript_targets(rows, transcript_dir)
    xlsx, manager_source = output_root / f"Отчёт РОП по звонкам {day.isoformat()} v3-{generation}.xlsx", "Mango API" if current_manager_users else manager_users.name if manager_users else ""
    if xlsx.exists(): raise RuntimeError("unreferenced immutable XLSX generation already exists")
    with tempfile.NamedTemporaryFile(prefix=f".Отчёт РОП {day.isoformat()}-", suffix=".staging.xlsx", dir=output_root, delete=False) as handle:
        staged_xlsx = Path(handle.name)
    try:
        write_workbook(staged_xlsx, day, rows, manager_source, source_before)
        copied, reused, updated = publish_transcripts(rows, transcript_dir)
        os.replace(staged_xlsx, xlsx)
    finally:
        staged_xlsx.unlink(missing_ok=True)
    manifest = {
        "schema_version": EXPORT_SCHEMA_VERSION, "generated_at": datetime.now(timezone.utc).isoformat(),
        "day": day.isoformat(), "rows": len(rows),
        "ready_rows": sum(row["complete"] for row in rows), "unfinished_rows": sum(not row["complete"] for row in rows),
        "manager_ready_rows": sum(row["manager_ready"] for row in rows), "content_sha256": content_sha256,
        "working_only_rows": working_only, "transcripts_copied": copied, "transcripts_reused": reused, "transcripts_updated": updated,
        "chronology_confirmed_rows": sum(row["chronology_confirmed"] for row in rows),
        "tallanto_names_found": sum(bool(row["client_fio"]) for row in rows),
        "tallanto_match_sources": dict(Counter(row["tallanto_source"] or "не найдено" for row in rows)),
        "current_mango_users": len(current_manager_users),
        "source_ready_db_sha256": source_before["sha256"], "tallanto_export_sha256": sha256_file(tallanto_export),
        "tallanto_snapshot_as_of": tallanto_snapshot_as_of.isoformat() if tallanto_snapshot_as_of else None,
        "xlsx": xlsx.name, "xlsx_sha256": sha256_file(xlsx), "transcript_dir": transcript_dir.name,
        "transcripts": [{"file": row["transcript_file"].name, "sha256": row["transcript_sha256"]} for row in rows],
        "reused": False,
    }
    manifest_path = output_root / f"Отчёт РОП по звонкам {day.isoformat()}.manifest.json"
    temporary = manifest_path.with_suffix(".tmp")
    temporary.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    os.replace(temporary, manifest_path)
    manifest_path.chmod(0o600)
    return {**manifest, "xlsx": str(xlsx), "transcript_dir": str(transcript_dir), "manifest": str(manifest_path)}


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Выгрузить суточный пакет звонков для РОПа.")
    parser.add_argument("--ready-db", type=Path, default=DEFAULT_READY_DB)
    parser.add_argument("--working-db", type=Path, default=DEFAULT_WORKING_DB)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    parser.add_argument("--day", type=date.fromisoformat, default=datetime.now(MOSCOW).date() - timedelta(days=1))
    parser.add_argument("--manager-users", type=Path, default=DEFAULT_MANAGER_USERS)
    parser.add_argument("--tallanto-export", type=Path, default=DEFAULT_TALLANTO_EXPORT)
    parser.add_argument("--tallanto-env", type=Path, default=DEFAULT_TALLANTO_ENV)
    parser.add_argument("--tallanto-snapshot-as-of", type=datetime.fromisoformat)
    parser.add_argument("--mango-env", type=Path, default=DEFAULT_MANGO_ENV)
    parser.add_argument("--sealed-only", action="store_true", help="Не читать рабочую DB и внешние transcript-файлы.")
    args = parser.parse_args(argv)
    result = export_day(
        args.ready_db, args.working_db, args.out, args.day, args.manager_users,
        tallanto_export=args.tallanto_export, tallanto_env=args.tallanto_env,
        tallanto_snapshot_as_of=args.tallanto_snapshot_as_of,
        current_manager_users=fetch_mango_users(args.mango_env),
        sealed_only=args.sealed_only,
    )
    print(json.dumps({key: value for key, value in result.items() if key != "transcripts"}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
