#!/usr/bin/env python3
"""Build the TZ-19 call analysis review workbook.

The workbook contains masked, columnar Analyse v7 summaries for Dmitry's
manual review. The script is read-only against canonical DB and writes the
PII-bearing workbook outside git by default.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence
from urllib.parse import quote


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = PROJECT_ROOT / "stable_runtime/canonical_master_20260523_audio_working_store_v1/canonical_calls_master.db"
DEFAULT_BLACKLIST = Path(
    "/Users/dmitrijfabarisov/Yandex.Disk.localized/OpenClaw/Actual Mango Tests/"
    "analyze_rerun_20260611/blacklist_77.txt"
)
DEFAULT_TAIL_MANIFEST = Path(
    "/Users/dmitrijfabarisov/Yandex.Disk.localized/OpenClaw/Actual Mango Tests/"
    "analyze_tail_20260612/data/manifest.json"
)
DEFAULT_OUT_DIR = Path("~/Claude Projects/Foton").expanduser()
DEFAULT_PROMPT_VERSION = "v7"
DEFAULT_LONG_THRESHOLD = 10_000

EMAIL_RE = re.compile(r"(?i)\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?7|8|9)(?:[\s().-]*\d){9,14}(?!\d)")

HEADERS = [
    "ID звонка",
    "Дата/время",
    "Длительность, сек",
    "Транскрипт, символов",
    "Длинный >10к",
    "Resolve статус",
    "Resolve класс",
    "Resolve кандидатов",
    "Клиент, телефон маска",
    "Клиент, phone_hash12",
    "AMO contact_id",
    "AMO lead_id",
    "Бренд/филиал",
    "Тип звонка",
    "Класс",
    "Продукт",
    "Предметы",
    "Формат",
    "Интерес",
    "Тема/кратко",
    "Следующий шаг",
    "Возражения",
    "Боли/сомнения",
    "Приоритет лида",
    "needs_review",
    "review_reasons",
    "blacklist-77",
    "Теги",
]


@dataclass(frozen=True)
class BuildConfig:
    db: Path
    out: Path
    blacklist: Path
    tail_manifest: Path | None
    scope: str = "baseline_22679"
    prompt_version: str = DEFAULT_PROMPT_VERSION
    long_threshold: int = DEFAULT_LONG_THRESHOLD
    summary_out: Path | None = None


def build_calls_review_table(config: BuildConfig) -> Mapping[str, Any]:
    cfg = normalize_config(config)
    rows, summary = load_review_rows(cfg)
    write_workbook(cfg.out, rows, cfg.long_threshold)
    summary = {
        **summary,
        "xlsx": str(cfg.out),
        "summary_json": str(cfg.summary_out) if cfg.summary_out else "",
        "safety": {
            "read_only_db": True,
            "run_asr": False,
            "run_analyze": False,
            "write_amo": False,
            "write_crm": False,
            "write_tallanto": False,
            "pii_workbook_in_git": False,
        },
    }
    if cfg.summary_out:
        cfg.summary_out.parent.mkdir(parents=True, exist_ok=True)
        cfg.summary_out.write_text(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return summary


def normalize_config(config: BuildConfig) -> BuildConfig:
    out = config.out.expanduser().resolve(strict=False)
    summary_out = config.summary_out.expanduser().resolve(strict=False) if config.summary_out else None
    return BuildConfig(
        db=config.db.expanduser().resolve(strict=False),
        out=out,
        blacklist=config.blacklist.expanduser().resolve(strict=False),
        tail_manifest=config.tail_manifest.expanduser().resolve(strict=False) if config.tail_manifest else None,
        scope=config.scope,
        prompt_version=config.prompt_version,
        long_threshold=config.long_threshold,
        summary_out=summary_out,
    )


def load_review_rows(config: BuildConfig) -> tuple[list[dict[str, Any]], Mapping[str, Any]]:
    blacklist_ids = read_int_set(config.blacklist)
    tail_ids = read_tail_manifest_ids(config.tail_manifest) if config.tail_manifest else set()
    rows: list[dict[str, Any]] = []
    counters: Counter[str] = Counter()
    con = connect_read_only(config.db)
    con.row_factory = sqlite3.Row
    try:
        current_v7_total = scalar_int(
            con,
            """
            SELECT COUNT(*) FROM canonical_calls
            WHERE analysis_json IS NOT NULL
              AND json_valid(analysis_json)
              AND json_extract(analysis_json, '$.analysis_meta.analysis_prompt_version') = ?
            """,
            (config.prompt_version,),
        )
        for row in con.execute(
            """
            SELECT canonical_call_id, started_at, duration_sec, transcript_chars, phone,
                   resolve_status, candidate_count, resolve_json, amocrm_contact_id,
                   amocrm_lead_id, analysis_json
            FROM canonical_calls
            WHERE analysis_json IS NOT NULL
              AND json_valid(analysis_json)
              AND json_extract(analysis_json, '$.analysis_meta.analysis_prompt_version') = ?
            ORDER BY started_at, canonical_call_id
            """,
            (config.prompt_version,),
        ):
            call_id = int(row["canonical_call_id"])
            if config.scope == "baseline_22679" and call_id in tail_ids:
                continue
            review_row = build_review_row(row, blacklist_ids=blacklist_ids, long_threshold=config.long_threshold)
            rows.append(review_row)
            counters["all"] += 1
            counters[f"resolve_{review_row['Resolve класс']}"] += 1
            if review_row["blacklist-77"]:
                counters["blacklist_77"] += 1
            if review_row["Длинный >10к"]:
                counters["long"] += 1
            counters[f"brand_{review_row['Бренд/филиал']}"] += 1
    finally:
        con.close()

    return rows, {
        "schema_version": "tz19_calls_review_table_v1",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "db": str(config.db),
        "scope": config.scope,
        "prompt_version": config.prompt_version,
        "current_v7_total_in_db": current_v7_total,
        "tail_manifest_ids_loaded": len(tail_ids),
        "blacklist_ids_loaded": len(blacklist_ids),
        "rows": len(rows),
        "long_threshold": config.long_threshold,
        "resolve_breakdown": {
            key.removeprefix("resolve_"): value for key, value in sorted(counters.items()) if key.startswith("resolve_")
        },
        "blacklist_77_rows": counters["blacklist_77"],
        "long_rows": counters["long"],
        "brand_breakdown": {
            key.removeprefix("brand_"): value for key, value in sorted(counters.items()) if key.startswith("brand_")
        },
    }


def build_review_row(row: sqlite3.Row, *, blacklist_ids: set[int], long_threshold: int) -> dict[str, Any]:
    call_id = int(row["canonical_call_id"])
    analysis = parse_json(row["analysis_json"])
    structured = as_mapping(analysis.get("structured_fields"))
    interests = as_mapping(structured.get("interests"))
    student = as_mapping(structured.get("student"))
    commercial = as_mapping(structured.get("commercial"))
    quality = as_mapping(analysis.get("quality_flags"))
    sensitive_values = sensitive_values_from(row, analysis)
    topic = sanitize_text(text_value(analysis.get("summary") or analysis.get("history_short")), sensitive_values)
    interest_values = normalize_list(analysis.get("interests")) or normalize_list(interests.get("products"))
    subjects = normalize_list(interests.get("subjects"))
    formats = normalize_list(interests.get("format"))
    objections = normalize_list(analysis.get("objections")) or normalize_list(structured.get("objections"))
    pain_points = normalize_list(analysis.get("pain_points"))
    tags = normalize_list(analysis.get("tags"))
    return {
        "ID звонка": call_id,
        "Дата/время": row["started_at"] or "",
        "Длительность, сек": round(float(row["duration_sec"] or 0), 1),
        "Транскрипт, символов": int(row["transcript_chars"] or 0),
        "Длинный >10к": int(row["transcript_chars"] or 0) > long_threshold,
        "Resolve статус": row["resolve_status"] or "",
        "Resolve класс": resolve_class(row["resolve_status"], int(row["candidate_count"] or 0)),
        "Resolve кандидатов": int(row["candidate_count"] or 0),
        "Клиент, телефон маска": mask_phone(row["phone"]),
        "Клиент, phone_hash12": phone_hash(row["phone"]),
        "AMO contact_id": row["amocrm_contact_id"] or "",
        "AMO lead_id": row["amocrm_lead_id"] or "",
        "Бренд/филиал": extract_brand(analysis),
        "Тип звонка": quality.get("call_type") or "",
        "Класс": text_value(analysis.get("student_grade") or student.get("grade_current")),
        "Продукт": text_value(analysis.get("target_product")),
        "Предметы": join_values(subjects),
        "Формат": join_values(formats),
        "Интерес": join_values(interest_values),
        "Тема/кратко": topic,
        "Следующий шаг": sanitize_text(text_value(analysis.get("next_step") or as_mapping(structured.get("next_step")).get("action")), sensitive_values),
        "Возражения": join_values(objections),
        "Боли/сомнения": join_values(pain_points),
        "Приоритет лида": text_value(as_mapping(structured.get("lead_priority")).get("value") if isinstance(structured.get("lead_priority"), Mapping) else structured.get("lead_priority") or analysis.get("lead_priority")),
        "needs_review": bool(analysis.get("needs_review") or quality.get("needs_review")),
        "review_reasons": join_values(normalize_list(analysis.get("review_reasons")) or normalize_list(quality.get("review_reasons"))),
        "blacklist-77": call_id in blacklist_ids,
        "Теги": join_values(tags),
    }


def write_workbook(path: Path, rows: Sequence[Mapping[str, Any]], long_threshold: int) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "Все", rows)
    add_sheet(workbook, "blacklist-77", [row for row in rows if row["blacklist-77"]])
    add_sheet(workbook, "длинные", [row for row in rows if int(row["Транскрипт, символов"] or 0) > long_threshold])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row_cells in sheet.iter_rows(min_row=2):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        widths = column_widths(sheet.max_column)
        for idx, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(idx)].width = width
    workbook.save(path)


def add_sheet(workbook: Any, name: str, rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])


def column_widths(column_count: int) -> list[int]:
    defaults = [12, 19, 12, 14, 12, 14, 17, 12, 18, 18, 14, 14, 14, 16, 10, 18, 22, 22, 28, 70, 32, 26, 26, 16, 14, 34, 13, 34]
    return defaults[:column_count] + [18] * max(0, column_count - len(defaults))


def read_int_set(path: Path) -> set[int]:
    if not path.exists():
        return set()
    return {int(item) for item in path.read_text(encoding="utf-8").split() if item.isdigit()}


def read_tail_manifest_ids(path: Path | None) -> set[int]:
    if path is None or not path.exists():
        return set()
    data = json.loads(path.read_text(encoding="utf-8"))
    calls = data.get("calls") if isinstance(data, Mapping) else None
    if not isinstance(calls, list):
        return set()
    return {int(item["canonical_call_id"]) for item in calls if isinstance(item, Mapping) and item.get("canonical_call_id")}


def connect_read_only(path: Path) -> sqlite3.Connection:
    resolved = path.expanduser().resolve(strict=False)
    uri = f"file:{quote(str(resolved), safe='/:')}?mode=ro&immutable=1"
    return sqlite3.connect(uri, uri=True)


def scalar_int(con: sqlite3.Connection, query: str, params: Sequence[Any] = ()) -> int:
    return int(con.execute(query, tuple(params)).fetchone()[0] or 0)


def parse_json(value: Any) -> Mapping[str, Any]:
    if isinstance(value, Mapping):
        return value
    try:
        parsed = json.loads(value or "{}")
    except (TypeError, json.JSONDecodeError):
        return {}
    return parsed if isinstance(parsed, Mapping) else {}


def as_mapping(value: Any) -> Mapping[str, Any]:
    return value if isinstance(value, Mapping) else {}


def normalize_list(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [text_value(item) for item in value if text_value(item)]
    if isinstance(value, tuple):
        return [text_value(item) for item in value if text_value(item)]
    text = text_value(value)
    return [text] if text else []


def text_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value).strip()


def join_values(values: Sequence[str]) -> str:
    return "; ".join(str(item).strip() for item in values if str(item).strip())


def normalize_phone(value: Any) -> str:
    digits = re.sub(r"\D+", "", str(value or ""))
    if len(digits) == 11 and digits.startswith("8"):
        digits = "7" + digits[1:]
    if len(digits) == 10 and digits.startswith("9"):
        digits = "7" + digits
    return digits


def mask_phone(value: Any) -> str:
    digits = normalize_phone(value)
    return f"***{digits[-4:]}" if len(digits) >= 4 else ""


def phone_hash(value: Any) -> str:
    digits = normalize_phone(value)
    if not digits:
        return ""
    digest = hashlib.sha256(digits.encode("utf-8")).hexdigest()[:12]
    return "ph_" + "_".join(digest[offset : offset + 4] for offset in range(0, len(digest), 4))


def resolve_class(status: Any, candidate_count: int) -> str:
    status_text = str(status or "").strip()
    if status_text == "done" and candidate_count == 1:
        return "нашёл одного"
    if status_text == "done" and candidate_count > 1:
        return "неоднозначно"
    if not status_text or status_text in {"skipped", "manual"} or candidate_count <= 0:
        return "не нашёл"
    return status_text


def sensitive_values_from(row: sqlite3.Row, analysis: Mapping[str, Any]) -> set[str]:
    result = {str(row["phone"] or "")}
    for section_name in ("structured_fields", "crm_blocks"):
        section = as_mapping(analysis.get(section_name))
        people = as_mapping(section.get("people"))
        contacts = as_mapping(section.get("contacts"))
        for value in list(people.values()) + list(contacts.values()):
            text = text_value(value)
            if len(text) >= 4:
                result.add(text)
    return {item for item in result if item}


def sanitize_text(value: str, sensitive_values: set[str]) -> str:
    text = str(value or "")
    text = EMAIL_RE.sub("[email]", text)
    text = PHONE_RE.sub("[phone]", text)
    for sensitive in sorted(sensitive_values, key=len, reverse=True):
        if len(sensitive) < 4:
            continue
        text = text.replace(sensitive, "[masked]")
    return text.strip()


def extract_brand(analysis: Mapping[str, Any]) -> str:
    probes = " ".join(
        normalize_list(analysis.get("tags"))
        + normalize_list(analysis.get("target_product"))
        + normalize_list(analysis.get("summary"))
        + normalize_list(as_mapping(analysis.get("quality_flags")).get("call_type"))
    ).lower()
    has_unpk = any(token in probes for token in ("унпк", "мфти", "unpk"))
    has_foton = any(token in probes for token in ("фотон", "foton"))
    if has_unpk and has_foton:
        return "mixed"
    if has_unpk:
        return "unpk"
    if has_foton:
        return "foton"
    return "unknown"


def build_parser() -> argparse.ArgumentParser:
    default_name = f"tz19_calls_review_table_{date.today().isoformat()}.xlsx"
    parser = argparse.ArgumentParser(description="Build TZ-19 Analyse v7 calls review table.")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--blacklist", type=Path, default=DEFAULT_BLACKLIST)
    parser.add_argument("--tail-manifest", type=Path, default=DEFAULT_TAIL_MANIFEST)
    parser.add_argument("--scope", choices=("baseline_22679", "current_v7"), default="baseline_22679")
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT_DIR / default_name)
    parser.add_argument("--summary-out", type=Path, default=None)
    parser.add_argument("--long-threshold", type=int, default=DEFAULT_LONG_THRESHOLD)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    summary_out = args.summary_out
    if summary_out is None:
        summary_out = args.out.with_suffix(".summary.json")
    summary = build_calls_review_table(
        BuildConfig(
            db=args.db,
            out=args.out,
            blacklist=args.blacklist,
            tail_manifest=args.tail_manifest,
            scope=args.scope,
            long_threshold=args.long_threshold,
            summary_out=summary_out,
        )
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
