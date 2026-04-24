#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib import error as url_error
from urllib import parse as url_parse
from urllib import request as url_request

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = PROJECT_ROOT / "src"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        if not key:
            continue
        if key not in os.environ:
            os.environ[key] = value.strip()


load_env_file(PROJECT_ROOT / "stable_runtime" / "amocrm_runtime" / ".env.private")

from mango_mvp.amocrm_runtime.tallanto_api import (  # noqa: E402
    TallantoApiClient,
    TallantoApiError,
    _extract_record_list,
    build_tallanto_api_config,
)
from mango_mvp.utils.phone import normalize_phone  # noqa: E402


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def norm_phone(value: Any) -> str:
    normalized = normalize_phone("" if value is None else str(value))
    return normalized or ""


def join_unique(values: Iterable[Any], sep: str = " | ") -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = as_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return sep.join(result)


def truncate(text: str, limit: int = 600) -> str:
    value = as_text(text)
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def telegram_phone_variants(phone: str) -> list[str]:
    normalized = norm_phone(phone)
    digits = "".join(ch for ch in normalized if ch.isdigit())
    if len(digits) != 11 or not digits.startswith("7"):
        return [phone]
    local = digits[1:]
    pretty7 = f"7 {local[0:3]} {local[3:6]}-{local[6:8]}-{local[8:10]}"
    pretty8 = f"8 {local[0:3]} {local[3:6]}-{local[6:8]}-{local[8:10]}"
    candidates = [
        normalized,
        digits,
        local,
        pretty7,
        pretty8,
        f"+7 {local[0:3]} {local[3:6]}-{local[6:8]}-{local[8:10]}",
        f"8{local}",
        f"7{local}",
    ]
    return [item for item in dict.fromkeys(candidate.strip() for candidate in candidates if candidate and candidate.strip())]


def is_valid_ru_phone(phone: Any) -> bool:
    normalized = norm_phone(phone)
    digits = "".join(ch for ch in normalized if ch.isdigit())
    return len(digits) == 11 and digits.startswith("7")


def tallanto_get_entry_by_field_fast(
    client: TallantoApiClient,
    *,
    field_name: str,
    field_value: str,
    timeout_seconds: int = 6,
) -> dict[str, Any]:
    query_items = [
        ("method", "get_entry_by_fields"),
        ("module", "Contact"),
        (f"fields_values[{field_name}]", field_value),
    ]
    url = f"{client.endpoint_url}?{url_parse.urlencode(query_items, doseq=True)}"
    request = url_request.Request(
        url,
        headers={
            "Accept": "application/json",
            "Content-Type": "application/x-www-form-urlencoded",
            "X-Auth-Token": client.config.api_token,
        },
        method="GET",
    )
    try:
        with url_request.urlopen(request, timeout=timeout_seconds) as response:
            raw = response.read().decode("utf-8")
            return json.loads(raw) if raw.strip() else {}
    except url_error.HTTPError as exc:
        details = exc.read().decode("utf-8", errors="replace")
        raise TallantoApiError(
            f"HTTP {exc.code} from Tallanto: {details or exc.reason}",
            status_code=502,
        ) from exc
    except Exception as exc:
        raise TallantoApiError(
            f"Failed to reach Tallanto: {exc}",
            status_code=502,
        ) from exc


def fast_tallanto_search(client: TallantoApiClient, phone: str) -> tuple[list[dict[str, Any]], str, str]:
    fields = ("phone_mobile", "phone_work", "phone_home", "phone_other")
    errors: list[str] = []
    records: list[dict[str, Any]] = []
    matched_via = ""
    for candidate in telegram_phone_variants(phone):
        for field_name in fields:
            try:
                payload = tallanto_get_entry_by_field_fast(client, field_name=field_name, field_value=candidate)
            except TallantoApiError as exc:
                message = str(exc)
                lowered = message.casefold()
                if "entry does not exist" in lowered or "not find by id" in lowered:
                    continue
                if "ошибка базы данных" in lowered:
                    errors.append(f"{field_name}:{candidate}:db_error")
                    continue
                errors.append(f"{field_name}:{candidate}:{message}")
                continue
            extracted = _extract_record_list(payload)
            if extracted:
                records.extend(extracted)
                matched_via = f"{field_name}:{candidate}"
        if records:
            break
    deduped: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for record in records:
        record_id = as_text(record.get("id"))
        if record_id and record_id in seen_ids:
            continue
        if record_id:
            seen_ids.add(record_id)
        deduped.append(record)
    return deduped, matched_via, "; ".join(errors[:10])


def compact_tallanto_contact(contact: dict[str, Any]) -> dict[str, Any]:
    filial = contact.get("filial")
    filial_text = ""
    if isinstance(filial, dict):
        filial_text = join_unique(filial.values())
    else:
        filial_text = as_text(filial)
    return {
        "id": as_text(contact.get("id")),
        "name": join_unique([contact.get("last_name"), contact.get("first_name"), contact.get("middle_name")], sep=" "),
        "branch": filial_text,
        "assigned_user_name": as_text(contact.get("assigned_user_name")),
        "type_client": as_text(contact.get("type_client_c_translated") or contact.get("type_client")),
        "description": as_text(contact.get("description")),
        "contact_card": as_text(contact.get("contact_card")),
    }


def empty_tallanto_context(contact: dict[str, Any]) -> dict[str, Any]:
    return {
        "contact": contact,
        "opportunities": [],
        "requests": [],
        "finances": [],
        "course_relations": [],
        "class_relations": [],
    }


def summarize_tallanto_context(context: dict[str, Any]) -> dict[str, Any]:
    contact = compact_tallanto_contact(context.get("contact") or {})
    opportunities = [item for item in (context.get("opportunities") or []) if isinstance(item, dict)]
    requests = [item for item in (context.get("requests") or []) if isinstance(item, dict)]
    finances = [item for item in (context.get("finances") or []) if isinstance(item, dict)]
    course_relations = [item for item in (context.get("course_relations") or []) if isinstance(item, dict)]
    class_relations = [item for item in (context.get("class_relations") or []) if isinstance(item, dict)]
    opportunity_text = "\n".join(
        join_unique(
            [
                as_text(item.get("id")),
                as_text(item.get("name")),
                as_text(item.get("sales_stage")),
                as_text(item.get("next_step")),
                as_text(item.get("assigned_user_name")),
                as_text(item.get("date_closed") or item.get("system_date_closed")),
            ]
        )
        for item in opportunities[:8]
    )
    request_text = "\n".join(
        join_unique(
            [
                as_text(item.get("id")),
                as_text(item.get("name")),
                as_text(item.get("status")),
                as_text(item.get("date_next_contact")),
                as_text(item.get("assigned_user_name")),
            ]
        )
        for item in requests[:8]
    )
    finance_text = "\n".join(
        join_unique(
            [
                as_text(item.get("id")),
                as_text(item.get("name")),
                as_text(item.get("payment_summa")),
                as_text(item.get("payment_status")),
                as_text(item.get("date_entered")),
            ]
        )
        for item in finances[:8]
    )
    compact_context = join_unique(
        [
            truncate(contact.get("description", ""), 350),
            truncate(contact.get("contact_card", ""), 350),
            f"Сделок Tallanto: {len(opportunities)}" if opportunities else "",
            f"Заявок Tallanto: {len(requests)}" if requests else "",
            f"Оплат Tallanto: {len(finances)}" if finances else "",
            f"Групп: {len(course_relations)}" if course_relations else "",
            f"Занятий: {len(class_relations)}" if class_relations else "",
        ],
        sep="\n",
    )
    return {
        "contact": contact,
        "opportunity_count": len(opportunities),
        "request_count": len(requests),
        "finance_count": len(finances),
        "course_relation_count": len(course_relations),
        "class_relation_count": len(class_relations),
        "opportunity_text": opportunity_text,
        "request_text": request_text,
        "finance_text": finance_text,
        "compact_context": compact_context,
    }


def lookup_tallanto_phone_second_pass(
    client: TallantoApiClient,
    phone: str,
    max_related_records: int,
) -> dict[str, Any]:
    found_contacts, matched_via, search_error = fast_tallanto_search(client, phone)
    contexts: list[dict[str, Any]] = []
    for contact in found_contacts[:5]:
        tallanto_id = as_text(contact.get("id"))
        if not tallanto_id:
            continue
        try:
            payload = client.build_contact_context_by_contact_id(
                tallanto_id,
                max_related_records=max_related_records,
            )
        except TallantoApiError as exc:
            message = str(exc)
            if "429" in message or "too many requests" in message.casefold():
                payload = {
                    "contacts_found": 1,
                    "contexts": [empty_tallanto_context(contact)],
                }
                search_error = join_unique([search_error, f"429_partial:{tallanto_id}"], sep="; ")
                time.sleep(1.0)
            else:
                search_error = join_unique([search_error, f"{tallanto_id}:{message}"], sep="; ")
                continue
        contexts.extend(payload.get("contexts") or [])
        time.sleep(0.15)

    deduped: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for context in contexts:
        contact_id = as_text((context.get("contact") or {}).get("id"))
        if contact_id and contact_id in seen_ids:
            continue
        if contact_id:
            seen_ids.add(contact_id)
        deduped.append(context)
    summarized = [summarize_tallanto_context(context) for context in deduped]
    return {
        "count": len(summarized),
        "matched_via": matched_via or ("second_pass_no_match" if not summarized else "second_pass"),
        "search_error": search_error,
        "contact_ids": [item["contact"]["id"] for item in summarized if item["contact"]["id"]],
        "contacts_text": join_unique(item["contact"]["name"] for item in summarized),
        "responsibles_text": join_unique(item["contact"]["assigned_user_name"] for item in summarized),
        "branches_text": join_unique(item["contact"]["branch"] for item in summarized),
        "type_clients_text": join_unique(item["contact"]["type_client"] for item in summarized),
        "opportunity_count": sum(item["opportunity_count"] for item in summarized),
        "request_count": sum(item["request_count"] for item in summarized),
        "finance_count": sum(item["finance_count"] for item in summarized),
        "course_relation_count": sum(item["course_relation_count"] for item in summarized),
        "class_relation_count": sum(item["class_relation_count"] for item in summarized),
        "opportunity_text": "\n\n".join(item["opportunity_text"] for item in summarized if item["opportunity_text"]),
        "request_text": "\n\n".join(item["request_text"] for item in summarized if item["request_text"]),
        "finance_text": "\n\n".join(item["finance_text"] for item in summarized if item["finance_text"]),
        "compact_context": "\n\n".join(item["compact_context"] for item in summarized if item["compact_context"]),
    }


def maybe_int(value: Any) -> int:
    try:
        return int(float(as_text(value) or 0))
    except Exception:
        return 0


def infer_message_goal(row: pd.Series) -> str:
    if maybe_int(row.get("Tallanto оплат")) or maybe_int(row.get("Tallanto групп")) or maybe_int(row.get("Tallanto занятий")):
        return "Предложить продление, второй предмет или ближайшую программу"
    if maybe_int(row.get("AMO открытых среди последних")):
        return "Продолжить уже начатый диалог и довести до следующего шага"
    if as_text(row.get("Есть неотвеченный входящий")) == "Да":
        return "Закрыть незавершенный вопрос и вернуть контакт в диалог"
    if as_text(row.get("Последний входящий Telegram")):
        return "Мягко ответить на последний запрос и актуализировать потребность"
    return "Мягко реактивировать диалог и уточнить текущую потребность"


def infer_touch_strategy(row: pd.Series) -> str:
    if maybe_int(row.get("Tallanto оплат")) or maybe_int(row.get("Tallanto групп")) or maybe_int(row.get("Tallanto занятий")):
        return "Опора на действующее обучение и логичное продолжение, без агрессивной продажи"
    if maybe_int(row.get("AMO открытых среди последних")):
        return "Продолжение уже знакомого контекста с одним конкретным CTA"
    if as_text(row.get("Возражения")):
        return "Аккуратно снять основное возражение и предложить следующий удобный формат"
    return "Короткий теплый follow-up с одним вопросом и без давления"


def build_personalization_context(row: pd.Series) -> str:
    parts = [
        as_text(row.get("ФИО родителя")),
        as_text(row.get("ФИО ребенка")),
        truncate(as_text(row.get("Продукты интереса")), 200),
        truncate(as_text(row.get("Краткая история общения")), 350),
        truncate(as_text(row.get("Возражения")), 180),
        truncate(as_text(row.get("AMO причина/стоп-фактор")), 180),
        truncate(as_text(row.get("Tallanto краткий контекст")), 280),
        truncate(as_text(row.get("Последний входящий Telegram")), 180),
    ]
    return join_unique([part for part in parts if part], sep="\n")


def openclaw_priority(row: pd.Series) -> str:
    score = maybe_int(row.get("Скор полезности"))
    if score >= 50:
        return "P1"
    if score >= 35:
        return "P2"
    if score >= 20:
        return "P3"
    return "P4"


def build_openclaw_prompt(row: pd.Series) -> str:
    offer = as_text(row.get("Что предложить")) or as_text(row.get("Гипотеза предложения")) or "актуальный следующий шаг"
    goal = as_text(row.get("Цель сообщения")) or infer_message_goal(row)
    strategy = as_text(row.get("Стратегия касания")) or infer_touch_strategy(row)
    context = as_text(row.get("Контекст для персонализации")) or build_personalization_context(row)
    dialog_name = as_text(row.get("Диалоги Telegram"))
    phone = as_text(row.get("Телефон"))
    return (
        "Составь одно короткое персональное сообщение в Telegram от имени менеджера школы.\n"
        f"Диалог: {dialog_name}\n"
        f"Телефон: {phone}\n"
        f"Цель: {goal}\n"
        f"Что предложить: {offer}\n"
        f"Стратегия: {strategy}\n"
        f"Контекст для персонализации:\n{context}\n\n"
        "Требования:\n"
        "- на русском\n"
        "- теплый человеческий тон\n"
        "- без давления и без выдумывания фактов\n"
        "- не упоминать CRM, AMO, Tallanto, score, внутренние заметки\n"
        "- максимум 500 знаков\n"
        "- один понятный CTA\n"
        "- если уместно, предложить созвон, подбор программы или отправку информации\n"
    )


def add_dataframe_sheet(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title=title)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, pd.Timestamp):
                if pd.isna(value):
                    value = ""
                else:
                    if value.tzinfo is not None:
                        value = value.tz_convert("UTC").tz_localize(None)
                    value = value.to_pydatetime()
            elif isinstance(value, datetime) and value.tzinfo is not None:
                value = value.astimezone().replace(tzinfo=None)
            ws.cell(row=row_index, column=column_index, value=value)
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def build_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    ws = workbook.active
    ws.title = "Сводка"
    ws["A1"] = "Показатель"
    ws["B1"] = "Значение"
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    ws["B1"].font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    row_index = 2
    for key, value in summary.items():
        ws.cell(row=row_index, column=1, value=key)
        ws.cell(row=row_index, column=2, value=value)
        row_index += 1
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-csv",
        type=Path,
        default=PROJECT_ROOT / "stable_runtime" / "audits" / "telegram_live_enrichment_20260416" / "telegram_phone_live_enrichment.csv",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=PROJECT_ROOT / "stable_runtime" / "audits" / "telegram_openclaw_final_20260416",
    )
    parser.add_argument("--target-limit", type=int, default=25)
    parser.add_argument("--target-min-score", type=int, default=5)
    parser.add_argument("--max-related-records", type=int, default=5)
    parser.add_argument("--top-n", type=int, default=30)
    args = parser.parse_args()

    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(args.input_csv)
    df["Телефон"] = df["Телефон"].map(norm_phone)
    if "Дата последнего сообщения Telegram" in df.columns:
        df["Дата последнего сообщения Telegram"] = pd.to_datetime(df["Дата последнего сообщения Telegram"], errors="coerce", utc=True)
    for column in [
        "Цель сообщения",
        "Стратегия касания",
        "Контекст для персонализации",
        "Что предложить",
        "Промпт для OpenClaw",
        "Черновик сообщения",
    ]:
        if column not in df.columns:
            df[column] = ""
        df[column] = df[column].fillna("").astype(str)

    target_mask = (
        (df["Персональное предложение возможно"] == "Да")
        & (df["Скор полезности"].fillna(0) >= args.target_min_score)
        & ((df["Tallanto контактов"].fillna(0) == 0) | df["Tallanto контактов"].isna())
        & df["Телефон"].map(is_valid_ru_phone)
    )
    target_df = df[target_mask].sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).head(args.target_limit).copy()

    client = TallantoApiClient(build_tallanto_api_config())
    second_pass_rows: list[dict[str, Any]] = []

    for index, row in enumerate(target_df.to_dict(orient="records"), start=1):
        phone = as_text(row.get("Телефон"))
        print(f"[{index}/{len(target_df)}] second pass {phone}", flush=True)
        result = lookup_tallanto_phone_second_pass(client, phone, max_related_records=args.max_related_records)
        second_pass_row = {
            "Телефон": phone,
            "Диалоги Telegram": as_text(row.get("Диалоги Telegram")),
            "Скор полезности": maybe_int(row.get("Скор полезности")),
            "Было Tallanto контактов": maybe_int(row.get("Tallanto контактов")),
            "Стало Tallanto контактов": result["count"],
            "Tallanto способ матчинга": result["matched_via"],
            "Tallanto ID": join_unique(result["contact_ids"]),
            "Tallanto контакты": result["contacts_text"],
            "Tallanto филиалы": result["branches_text"],
            "Tallanto тип клиента": result["type_clients_text"],
            "Tallanto сделок": result["opportunity_count"],
            "Tallanto заявок": result["request_count"],
            "Tallanto оплат": result["finance_count"],
            "Tallanto групп": result["course_relation_count"],
            "Tallanto занятий": result["class_relation_count"],
            "Tallanto краткий контекст": result["compact_context"],
            "Tallanto ошибка поиска": result["search_error"],
            "Успешно добрали Tallanto": "Да" if result["count"] else "Нет",
        }
        second_pass_rows.append(second_pass_row)
        if result["count"]:
            idx = df.index[df["Телефон"] == phone]
            if len(idx):
                i = idx[0]
                df.at[i, "Tallanto контактов"] = result["count"]
                df.at[i, "Tallanto способ матчинга"] = result["matched_via"]
                df.at[i, "Tallanto ID"] = join_unique(result["contact_ids"])
                df.at[i, "Tallanto контакты"] = result["contacts_text"]
                df.at[i, "Tallanto ответственные"] = result["responsibles_text"]
                df.at[i, "Tallanto филиалы"] = result["branches_text"]
                df.at[i, "Tallanto тип клиента"] = result["type_clients_text"]
                df.at[i, "Tallanto сделок"] = result["opportunity_count"]
                df.at[i, "Tallanto заявок"] = result["request_count"]
                df.at[i, "Tallanto оплат"] = result["finance_count"]
                df.at[i, "Tallanto групп"] = result["course_relation_count"]
                df.at[i, "Tallanto занятий"] = result["class_relation_count"]
                df.at[i, "Tallanto opportunities"] = result["opportunity_text"]
                df.at[i, "Tallanto requests"] = result["request_text"]
                df.at[i, "Tallanto finances"] = result["finance_text"]
                df.at[i, "Tallanto краткий контекст"] = result["compact_context"]
                df.at[i, "Tallanto ошибка поиска"] = result["search_error"]
        time.sleep(0.7)

    for i, row in df.iterrows():
        if not as_text(row.get("Цель сообщения")):
            df.at[i, "Цель сообщения"] = infer_message_goal(row)
        if not as_text(row.get("Стратегия касания")):
            df.at[i, "Стратегия касания"] = infer_touch_strategy(row)
        if not as_text(row.get("Контекст для персонализации")):
            df.at[i, "Контекст для персонализации"] = build_personalization_context(row)
        if not as_text(row.get("Что предложить")):
            df.at[i, "Что предложить"] = as_text(row.get("Гипотеза предложения"))
        df.at[i, "OpenClaw приоритет"] = openclaw_priority(row)
        df.at[i, "Промпт для OpenClaw"] = build_openclaw_prompt(df.loc[i])
        df.at[i, "Черновик сообщения"] = ""

    df = df.sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).reset_index(drop=True)
    top_df = df[df["Персональное предложение возможно"] == "Да"].head(args.top_n).copy()
    high_utility_df = df[(df["Персональное предложение возможно"] == "Да") & (df["Скор полезности"] >= 30)].copy()
    prospects_df = df[(df["Персональное предложение возможно"] == "Да") & (df["Скор полезности"] >= 20)].copy()
    second_pass_df = pd.DataFrame(second_pass_rows).sort_values(["Успешно добрали Tallanto", "Скор полезности"], ascending=[False, False])

    openclaw_cols = [
        "OpenClaw приоритет",
        "Телефон",
        "Диалоги Telegram",
        "Username Telegram",
        "Дата последнего сообщения Telegram",
        "Последний входящий Telegram",
        "ФИО родителя",
        "ФИО ребенка",
        "Краткая история общения",
        "Хронология общения",
        "Продукты интереса",
        "Возражения",
        "AMO контакты",
        "AMO ответственные",
        "AMO последние сделки",
        "AMO последние UTM",
        "AMO причина/стоп-фактор",
        "Tallanto контакты",
        "Tallanto филиалы",
        "Tallanto тип клиента",
        "Tallanto оплат",
        "Tallanto групп",
        "Tallanto занятий",
        "Tallanto краткий контекст",
        "Что предложить",
        "Цель сообщения",
        "Контекст для персонализации",
        "Стратегия касания",
        "Основа для предложения",
        "Гипотеза предложения",
        "Скор полезности",
        "Промпт для OpenClaw",
        "Черновик сообщения",
    ]
    top_df = top_df[openclaw_cols]
    high_utility_df = high_utility_df[openclaw_cols]
    prospects_df = prospects_df[openclaw_cols]

    summary = {
        "phones_total": int(len(df)),
        "offer_possible": int((df["Персональное предложение возможно"] == "Да").sum()),
        "high_utility_score_ge_30": int((df["Скор полезности"] >= 30).sum()),
        "prospects_score_ge_20": int(((df["Персональное предложение возможно"] == "Да") & (df["Скор полезности"] >= 20)).sum()),
        "second_pass_targeted": int(len(second_pass_df)),
        "second_pass_got_tallanto": int((second_pass_df.get("Успешно добрали Tallanto") == "Да").sum()) if len(second_pass_df) else 0,
        "top_sheet_rows": int(len(top_df)),
        "xlsx": str(out_dir / "telegram_openclaw_final.xlsx"),
        "csv_top": str(out_dir / "telegram_openclaw_top.csv"),
        "csv_all": str(out_dir / "telegram_live_enrichment_refined.csv"),
    }

    workbook = Workbook()
    build_summary_sheet(workbook, summary)
    add_dataframe_sheet(workbook, "OpenClaw Top", top_df)
    add_dataframe_sheet(workbook, "High Utility", high_utility_df)
    add_dataframe_sheet(workbook, "Prospects", prospects_df)
    add_dataframe_sheet(workbook, "Tallanto 2nd pass", second_pass_df)
    workbook.save(out_dir / "telegram_openclaw_final.xlsx")

    df.to_csv(out_dir / "telegram_live_enrichment_refined.csv", index=False)
    top_df.to_csv(out_dir / "telegram_openclaw_top.csv", index=False)
    second_pass_df.to_csv(out_dir / "telegram_tallanto_second_pass.csv", index=False)
    (out_dir / "telegram_openclaw_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
