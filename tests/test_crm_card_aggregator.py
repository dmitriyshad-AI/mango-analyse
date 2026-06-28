from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook

from mango_mvp.crm_card_aggregator import (
    apply_deal_card_payload,
    build_crm_card_projection,
    compact_objection_explicit,
)
from mango_mvp.crm_card_history_summary import CrmHistorySummaryConfig, CrmHistorySummarizer
from mango_mvp.crm_card_workbook import CrmCardWorkbookConfig, build_crm_card_workbook
from mango_mvp.customer_timeline import CustomerTimelineReadApi, CustomerTimelineReadApiConfig
from mango_mvp.quality.crm_text_quality_detector import detect_crm_text_quality_risks
from tests.test_customer_timeline_read_api import seed_timeline_db


def _profile(tmp_path: Path) -> dict:
    db_path, customer_id = seed_timeline_db(tmp_path)
    with CustomerTimelineReadApi.open(CustomerTimelineReadApiConfig(timeline_db=db_path, allowed_root=tmp_path)) as api:
        return dict(api.customer_profile("foton", customer_id, event_limit=10, bot_context_limit=10))


def test_crm_card_aggregator_builds_two_projections_from_one_profile(tmp_path: Path) -> None:
    profile = _profile(tmp_path)
    profile["manager_projection"] = {
        "amo_contact_ids": ["123"],
        "amo_lead_ids": ["456"],
        "identity_links": [
            {"link_type": "amo_contact_id", "link_value": "123", "match_class": "strong_unique"},
            {"link_type": "amo_lead_id", "link_value": "456", "match_class": "strong_unique"},
        ],
        "opportunities": [
            {
                "opportunity_id": "opp-manager",
                "opportunity_type": "amo_deal",
                "source_system": "amocrm_snapshot",
                "source_id": "456",
                "title": "ЕГЭ математика",
                "status": "open",
            }
        ],
    }

    card = build_crm_card_projection(
        profile,
        manager_facts={
            "Следующий шаг": "Согласовать дату звонка",
            "Возражения": "цена",
            "Вероятность продажи, %": "70",
            "AI-бюджет диапазон": "50k_100k",
            "AMO contact IDs": "123",
            "selected_deal_id": "456",
        },
    )

    assert card["snapshot_as_of"] == profile["snapshot_as_of"]
    assert card["deal_card"]["fields"]["Следующий шаг"] == "Уточнить у менеджера: открыт конфликт идентичности"
    assert "AI-бюджет диапазон" not in card["deal_card"]["fields"]
    assert "AI-дата обновления сделки" not in card["deal_card"]["fields"]
    assert set(card["contact_card"]["fields"]).issubset({"Запрос", "Последняя сводка", "История общения"})
    assert set(card["deal_card"]["fields"]).issubset({"Статус сделки", "Возражения", "Следующий шаг", "Tallanto", "Предупреждения"})
    assert card["bot_safety"]["money_fields_manager_only"] is True


def test_crm_card_uses_full_call_analysis_and_filters_non_conversation() -> None:
    live_summary = "Полный разбор живого звонка. " + ("Клиент обсуждал курс и оплату. " * 20)
    older_summary = "Полный разбор предыдущего звонка: семья выбирала формат и предмет."
    profile = {
        "found": True,
        "customer_id": "customer:call-analysis",
        "snapshot_as_of": "2026-06-18T10:00:00+00:00",
        "last_event_at": "2026-06-18T10:00:00+00:00",
        "customer": {"customer_id": "customer:call-analysis", "identity_status": "strong", "display_name": "Татьяна Николаевна", "summary": {}},
        "customer_id_mappings": [],
        "identity_links": [{"match_class": "strong_unique"}],
        "manager_projection": {
            "amo_contact_ids": ["123"],
            "amo_lead_ids": ["456"],
            "opportunities": [{"opportunity_id": "opp1", "opportunity_type": "amo_deal", "source_system": "amocrm_snapshot", "source_id": "456"}],
        },
        "opportunities": [],
        "timeline": {
            "items": [
                {
                    "event_type": "amo_contact_snapshot",
                    "event_at": "2026-06-19T10:00:00+00:00",
                    "source_system": "amocrm_snapshot",
                    "summary": "Read-only AMO contact snapshot",
                    "text_preview": "Read-only AMO contact snapshot",
                },
                {
                    "event_type": "tallanto_student_snapshot",
                    "event_at": "2026-06-19T09:00:00+00:00",
                    "source_system": "tallanto_snapshot",
                    "summary": "exact_phone_single",
                    "text_preview": "Статус: exact_phone_single | Тип ученика: 7 класс | Филиал: Фотон",
                },
                {
                    "event_type": "amo_deal_stage",
                    "event_at": "2026-06-19T08:00:00+00:00",
                    "source_system": "amocrm_snapshot",
                    "summary": "Закрыто и не реализовано",
                    "stage_after": "Закрыто и не реализовано",
                },
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-18T10:00:00+00:00",
                    "source_system": "mango",
                    "summary": "Недозвон не должен попасть в историю",
                    "call_type": "non_conversation",
                    "call_history_eligible": False,
                    "call_analysis": {
                        "history_summary": "Недозвон не должен попасть в историю",
                        "call_type": "non_conversation",
                        "call_history_eligible": False,
                    },
                },
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-17T10:00:00+00:00",
                    "source_system": "mango",
                    "summary": live_summary,
                    "call_type": "sales_call",
                    "call_history_eligible": True,
                    "call_analysis": {
                        "history_summary": live_summary,
                        "call_type": "sales_call",
                        "call_history_eligible": True,
                        "objections": ["цена"],
                        "next_step": "Перезвонить завтра",
                        "pain_points": ["нет времени"],
                        "interests": ["математика", "математика"],
                        "target_product": "годовой курс",
                        "structured_fields": {
                            "people": {"child_fio": "Анна"},
                            "student": {"grade_current": "7"},
                            "interests": {
                                "products": ["годовой курс", "годовой курс"],
                                "subjects": ["математика"],
                            },
                        },
                    },
                },
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-16T10:00:00+00:00",
                    "source_system": "mango",
                    "summary": older_summary,
                    "call_type": "sales_call",
                    "call_history_eligible": True,
                    "call_analysis": {
                        "history_summary": older_summary,
                        "call_type": "sales_call",
                        "call_history_eligible": True,
                        "next_step": "Дождаться выбора формата",
                        "interests": ["физика"],
                    },
                },
            ]
        },
        "signals": [],
        "bot_context": {"items": []},
        "conflicts": {"items": [], "summary": {"open_conflicts": 0}},
        "readiness": {"open_conflicts": 0},
    }

    card = build_crm_card_projection(profile)
    history = card["contact_card"]["fields"]["История общения"]
    preview = card["workbook"]["what_goes_to_amo"]

    assert card["contact_card"]["fields"]["Запрос"] == "Анна, 7 класс; годовой курс; математика"
    assert "Татьяна Николаевна" not in card["contact_card"]["fields"]["Запрос"]
    assert card["contact_card"]["fields"]["Последняя сводка"] == "Сводка:\n" + live_summary.strip()
    assert "Полный разбор живого звонка" not in history
    assert preview.count("Полный разбор живого звонка") == 1
    assert older_summary in history
    assert preview.count(older_summary) == 1
    assert "Возражения: цена" not in history
    assert "Интересы: математика" not in history
    assert "Шаг: Перезвонить на следующий день (18.06.2026)" in history
    assert "Недозвон не должен попасть" not in history
    assert "Read-only AMO contact snapshot" not in history
    assert "exact_phone_single" not in history
    assert "Закрыто и не реализовано" not in history
    assert "Закрыто и не реализовано" in card["deal_card"]["fields"]["Статус сделки"]
    assert card["deal_card"]["fields"]["Следующий шаг"] == "Перезвонить на следующий день (18.06.2026)"
    assert card["deal_card"]["fields"]["Возражения"] == "цена"
    assert "[сжато]" not in history
    assert "Read-only AMO contact snapshot" not in preview
    assert "exact_phone_single" not in preview
    assert "Tallanto: найден один ученик по телефону." not in history
    assert card["deal_card"]["fields"]["Tallanto"] == "Один ученик по телефону\nТип ученика: 7 класс\nФилиал: Фотон"
    assert card["workbook"]["ready"] == "да"


def test_crm_card_history_uses_cached_summary_and_excludes_email_handoff(tmp_path: Path) -> None:
    calls: list[str] = []

    def fake_runner(prompt: str) -> str:
        calls.append(prompt)
        source_part = prompt.split("История для сжатия:", 1)[-1]
        assert "Email handoff" not in source_part
        assert "Связанных писем" not in source_part
        assert "End of History" not in source_part
        return json.dumps(
            {
                "history": (
                    "Сводка:\n"
                    "Клиент выбирал курс по математике.\n\n"
                    "Хронология:\n"
                    "- 17.06.2026: обсудили оплату и следующий шаг."
                )
            },
            ensure_ascii=False,
        )

    summarizer = CrmHistorySummarizer(
        CrmHistorySummaryConfig(provider="codex_cli", cache_dir=tmp_path / "cache", model="fake"),
        runner=fake_runner,
    )
    profile = {
        "found": True,
        "customer_id": "customer:history-summary",
        "snapshot_as_of": "2026-06-18T10:00:00+00:00",
        "customer": {"customer_id": "customer:history-summary", "identity_status": "strong", "summary": {}},
        "identity_links": [{"match_class": "strong_unique"}],
        "manager_projection": {"amo_contact_ids": ["123"], "amo_lead_ids": ["456"]},
        "timeline": {
            "items": [
                {
                    "event_type": "email_message",
                    "source_system": "mail_archive",
                    "event_at": "2026-06-18T10:00:00+00:00",
                    "summary": "Email handoff: 8 сообщений; кандидатов: 1",
                    "text_preview": "Связанных писем: 8",
                },
                {
                    "event_type": "whatsapp_message",
                    "source_system": "channel_snapshot",
                    "event_at": "2026-06-18T09:00:00+00:00",
                    "summary": "Клиент спросил про курс.\nEnd of History\nApache 2.0 License",
                },
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-17T10:00:00+00:00",
                    "source_system": "mango",
                    "summary": "Клиент выбирал курс по математике. Итог: Есть согласованный следующий шаг. Контакты: канал: email.",
                    "call_history_eligible": True,
                    "call_analysis": {
                        "history_summary": "Клиент выбирал курс по математике. Итог: Есть согласованный следующий шаг. Контакты: канал: email.",
                        "next_step": "Перезвонить",
                    },
                },
            ]
        },
        "signals": [],
        "bot_context": {"items": []},
        "conflicts": {"items": [], "summary": {"open_conflicts": 0}},
        "readiness": {"open_conflicts": 0},
    }

    first = build_crm_card_projection(profile, history_summarizer=summarizer)
    second = build_crm_card_projection(profile, history_summarizer=summarizer)
    history = first["contact_card"]["fields"]["История общения"]

    assert history.startswith("Сводка:\nКлиент выбирал курс")
    assert "Email handoff" not in history
    assert "End of History" not in history
    assert first["contact_card"]["fields"]["История общения"] == second["contact_card"]["fields"]["История общения"]
    assert len(calls) == 1
    assert summarizer.summary()["cache_hits"] == 1


def test_crm_card_accepts_prod_mango_summary_without_legacy_eligibility_flag(tmp_path: Path) -> None:
    profile = {
        "found": True,
        "customer_id": "customer:prod-summary",
        "snapshot_as_of": "2026-06-21T00:00:00+00:00",
        "customer": {"customer_id": "customer:prod-summary", "identity_status": "strong", "summary": {}},
        "identity_links": [{"match_class": "strong_unique"}],
        "manager_projection": {
            "amo_contact_ids": ["123"],
            "amo_lead_ids": ["456"],
            "opportunities": [{"opportunity_id": "opp1", "opportunity_type": "amo_deal", "source_system": "amocrm_snapshot", "source_id": "456"}],
        },
        "timeline": {
            "items": [
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-20T10:00:00+00:00",
                    "source_system": "mango_processed_summary",
                    "subject": "sales_call",
                    "summary": "20.06.2026 менеджер обсудил с клиентом летнюю школу. Договорились: перезвонить завтра.",
                },
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-21T10:00:00+00:00",
                    "source_system": "mango_processed_summary",
                    "subject": "non_conversation",
                    "summary": "Недозвон не должен стать последней сводкой.",
                    "call_history_eligible": False,
                },
            ]
        },
        "signals": [],
        "bot_context": {"items": []},
        "conflicts": {"items": [], "summary": {"open_conflicts": 0}},
        "readiness": {"open_conflicts": 0},
    }

    summarizer = CrmHistorySummarizer(CrmHistorySummaryConfig(provider="rule", cache_dir=tmp_path / "cache"))
    card = build_crm_card_projection(profile, history_summarizer=summarizer)
    quality_payload = {
        "Последняя AI-сводка": card["contact_card"]["fields"]["Последняя сводка"],
        "Авто история общения": card["contact_card"]["fields"]["История общения"],
    }

    assert card["contact_card"]["fields"]["Последняя сводка"].startswith("Сводка:\n20.06.2026 менеджер")
    assert "завтра" not in card["contact_card"]["fields"]["Последняя сводка"].casefold()
    assert "на следующий день (21.06.2026)" in card["contact_card"]["fields"]["Последняя сводка"]
    assert "Недозвон" not in card["contact_card"]["fields"]["История общения"]
    assert "20.06.2026 менеджер обсудил с клиентом летнюю школу" not in card["contact_card"]["fields"]["История общения"]
    assert not [
        item
        for item in detect_crm_text_quality_risks(quality_payload, min_severity="P2")
        if item.risk_type == "cross_field_duplicate_information"
    ]
    assert card["workbook"]["ready"] == "да"


def test_crm_card_history_summarizer_accepts_multiline_history_json(tmp_path: Path) -> None:
    summarizer = CrmHistorySummarizer(
        CrmHistorySummaryConfig(provider="codex_cli", cache_dir=tmp_path / "cache", model="fake"),
        runner=lambda _prompt: '{"history":"Сводка:\nКлиент выбрал курс.\n\nШаг:\nПерезвонить"}',
    )

    history = summarizer("2026-06-17 mango_call: Клиент выбрал курс. Следующий шаг: перезвонить.")

    assert "Сводка:\nКлиент выбрал курс." in history
    assert summarizer.summary()["llm_calls"] == 1
    assert summarizer.summary()["rule_fallbacks"] == 0


def test_crm_card_internal_text_fields_are_not_cut_to_old_caps() -> None:
    long_summary = "Полная история важного разговора. " * 120
    long_next_step = "Согласовать программу и расписание с родителем. " * 60
    profile = {
        "found": True,
        "customer_id": "customer:no-caps",
        "snapshot_as_of": "2026-06-18T10:00:00+00:00",
        "customer": {"customer_id": "customer:no-caps", "identity_status": "strong", "summary": {}},
        "identity_links": [{"match_class": "strong_unique"}],
        "manager_projection": {"amo_contact_ids": ["123"], "amo_lead_ids": ["456"]},
        "timeline": {
            "items": [
                {
                    "event_type": "mango_call",
                    "event_at": "2026-06-17T10:00:00+00:00",
                    "source_system": "mango",
                    "summary": long_summary,
                    "call_history_eligible": True,
                    "call_analysis": {
                        "history_summary": long_summary,
                        "next_step": long_next_step,
                    },
                }
            ]
        },
        "signals": [],
        "bot_context": {"items": []},
        "conflicts": {"items": [], "summary": {"open_conflicts": 0}},
        "readiness": {"open_conflicts": 0},
    }

    card = build_crm_card_projection(profile)

    assert "[сжато]" not in card["contact_card"]["fields"]["Последняя сводка"]
    assert "…" not in card["contact_card"]["fields"]["Последняя сводка"]
    assert len(card["contact_card"]["fields"]["Последняя сводка"]) > 1600
    assert len(card["deal_card"]["fields"]["Следующий шаг"]) > 800


def test_crm_card_empty_timeline_falls_back_to_analyze_and_is_idempotent() -> None:
    profile = {
        "found": True,
        "customer_id": "customer:empty",
        "snapshot_as_of": "2026-06-18T10:00:00+00:00",
        "last_event_at": None,
        "customer": {"customer_id": "customer:empty", "identity_status": "strong", "summary": {}},
        "customer_id_mappings": [{"old_customer_id": "customer:legacy", "new_customer_id": "customer:empty"}],
        "identity_links": [{"match_class": "strong_unique"}],
        "opportunities": [],
        "timeline": {"items": []},
        "signals": [],
        "bot_context": {"items": []},
        "conflicts": {"items": [], "summary": {"open_conflicts": 0}},
        "readiness": {"open_conflicts": 0},
    }
    facts = {
        "Краткая история общения": "Клиент интересовался курсом.",
        "Следующий шаг": "Перезвонить и уточнить предмет.",
        "AMO contact IDs": "123",
    }

    first = build_crm_card_projection(profile, manager_facts=facts)
    second = build_crm_card_projection(profile, manager_facts=facts)

    assert first == second
    assert first["contact_card"]["fields"]["Последняя сводка"] == "Сводка:\nКлиент интересовался курсом."
    assert first["deal_card"]["fields"]["Следующий шаг"] == "Перезвонить и уточнить предмет."


def test_crm_card_ambiguous_identity_blocks_ready_and_keeps_family_summary() -> None:
    profile = {
        "found": True,
        "customer_id": "customer:family",
        "snapshot_as_of": "2026-06-18T10:00:00+00:00",
        "customer": {"customer_id": "customer:family", "identity_status": "ambiguous", "summary": {}},
        "customer_id_mappings": [{"old_customer_id": "old", "new_customer_id": "customer:family"}],
        "identity_links": [{"match_class": "ambiguous", "link_type": "phone"}],
        "opportunities": [{"opportunity_id": "opp1", "opportunity_type": "amo_deal", "title": "ЛВШ", "status": "open"}],
        "timeline": {"items": [{"event_at": "2026-06-18T10:00:00+00:00", "summary": "Семейный контакт: несколько учеников."}]},
        "signals": [],
        "bot_context": {"items": []},
        "conflicts": {"items": [{"status": "open", "conflict_type": "ambiguous_identity"}], "summary": {"open_conflicts": 1}},
        "readiness": {"open_conflicts": 1},
    }

    card = build_crm_card_projection(profile, manager_facts={"AMO contact IDs": "123", "selected_deal_id": "456"})

    assert card["contact_card"]["fields"]["Последняя сводка"] == "Сводка:\nСемейный контакт: несколько учеников."
    assert card["workbook"]["ready"] == "нет"
    assert "На телефоне несколько человек — проверьте, к кому относится" in card["workbook"]["blockers"]
    assert card["deal_card"]["ready_for_amo"] is False


def test_deal_card_payload_is_replaced_only_when_flag_enabled(monkeypatch) -> None:
    base = {"Статус сделки": "old"}
    card = {"Статус сделки": "new"}
    monkeypatch.delenv("CRM_CARD_AGGREGATOR_ENABLED", raising=False)

    assert apply_deal_card_payload(base, card)["Статус сделки"] == "old"

    monkeypatch.setenv("CRM_CARD_AGGREGATOR_ENABLED", "1")
    assert apply_deal_card_payload(base, card)["Статус сделки"] == "new"


def test_explicit_objection_compaction_marker_keeps_text_visible() -> None:
    text = " ".join(["клиент сомневается из-за стоимости программы"] * 5)

    compacted = compact_objection_explicit(text, limit=90)

    assert len(compacted) <= 90
    assert compacted.endswith("[сжато]")
    assert "…" not in compacted


def test_crm_card_workbook_is_transient_read_only_xlsx(tmp_path: Path) -> None:
    db_path, _customer_id = seed_timeline_db(tmp_path)
    out_xlsx = tmp_path / "crm_cards.xlsx"

    summary = build_crm_card_workbook(
        CrmCardWorkbookConfig(
            timeline_db=db_path,
            allowed_root=tmp_path,
            out_xlsx=out_xlsx,
            sample_size=2,
        )
    )

    assert summary["generated_at"] == "2026-05-12T12:01:00+00:00"
    assert summary["safety"]["write_amo"] is False
    assert summary["safety"]["new_storage_table"] is False
    assert out_xlsx.exists()
    assert out_xlsx.with_suffix(".csv").exists()
    wb = load_workbook(out_xlsx)
    assert wb.sheetnames == ["Сводка", "Клиенты"]
    assert [cell.value for cell in wb["Клиенты"][1]][:17] == [
        "Имя",
        "Телефон",
        "Бренд",
        "Открыть в AMO",
        "Запрос",
        "Статус сделки",
        "Возражения",
        "Следующий шаг",
        "Последняя сводка",
        "Tallanto",
        "Предупреждения",
        "История общения",
        "Готово",
        "Блокеры",
        "Вердикт",
        "Комментарий",
        "customer_id",
    ]
    phone_cell = wb["Клиенты"]["B2"].value
    assert phone_cell and "***" not in str(phone_cell)
    csv_rows = out_xlsx.with_suffix(".csv").read_text(encoding="utf-8-sig")
    assert "crm_card_contact_payload_json" in csv_rows
    saved_summary = json.loads(out_xlsx.with_suffix(".summary.json").read_text(encoding="utf-8"))
    assert saved_summary["generated_at"] == summary["generated_at"]
