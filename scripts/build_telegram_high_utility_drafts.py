#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def maybe_int(value: Any) -> int:
    try:
        return int(float(as_text(value) or 0))
    except Exception:
        return 0


def truncate(text: str, limit: int = 500) -> str:
    value = as_text(text)
    if len(value) <= limit:
        return value
    return value[: limit - 3].rstrip() + "..."


def extract_name_candidate(value: str) -> str:
    text = as_text(value)
    if not text:
        return ""
    text = re.sub(r"[_|/]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    if " - " in text:
        text = text.split(" - ", 1)[0].strip()
    words = re.findall(r"[A-Za-zА-Яа-яЁё]+", text)
    if not words:
        return ""
    if len(words) >= 2 and len(words[1]) > 2:
        return f"{words[0]} {words[1]}"
    return words[0]


def recipient_name(row: pd.Series) -> str:
    for key in ("ФИО родителя", "Диалоги Telegram", "AMO контакты"):
        value = as_text(row.get(key))
        if not value:
            continue
        if key == "AMO контакты" and "|" in value:
            value = value.split("|", 1)[0].strip()
        name = extract_name_candidate(value)
        if name:
            return name
    return ""


def student_phrase(row: pd.Series) -> str:
    if as_text(row.get("ФИО ребенка")):
        return "по поводу дальнейшего обучения"
    if as_text(row.get("Что предложить")) or as_text(row.get("Гипотеза предложения")):
        return "по поводу продолжения обучения"
    return "по поводу обучения"


def offer_phrase(row: pd.Series) -> str:
    offer = as_text(row.get("Что предложить")) or as_text(row.get("Гипотеза предложения")) or as_text(row.get("Рекомендуемый продукт"))
    products = as_text(row.get("Продукты интереса"))
    offer_lower = offer.lower()
    products_lower = products.lower()
    if "годов" in offer_lower:
        if "мат" in products_lower and "физ" in products_lower:
            return "варианты по годовым курсам по математике и физике"
        if "мат" in products_lower:
            return "варианты по годовым курсам по математике"
        if "физ" in products_lower:
            return "варианты по годовым курсам по физике"
        if "информ" in products_lower:
            return "варианты по годовым курсам по информатике"
        return "варианты по обучению на следующий учебный год"
    if "лагер" in offer_lower or "лвш" in products_lower:
        return "подходящий вариант летней программы"
    if "интенсив" in offer_lower:
        return "подходящий интенсив"
    if "второй предмет" in offer_lower.lower():
        return "вариант продолжения или второго предмета"
    return offer_lower or "подходящий вариант обучения"


def constraints_phrase(row: pd.Series) -> str:
    objections = as_text(row.get("Возражения")).lower()
    parts: list[str] = []
    if any(token in objections for token in ("время", "распис", "нагруз")):
        parts.append("с учетом расписания")
    if any(token in objections for token in ("цен", "стоим", "бюдж")):
        parts.append("с вариантами по нагрузке и стоимости")
    if any(token in objections for token in ("онлайн", "интернет", "формат")):
        parts.append("с удобным форматом занятий")
    if any(token in objections for token in ("преподав", "групп", "довер")):
        parts.append("с понятной группой и преподавателем")
    if not parts:
        return ""
    return ", ".join(parts)


def continuity_phrase(row: pd.Series) -> str:
    if maybe_int(row.get("Tallanto оплат")) or maybe_int(row.get("Tallanto групп")) or maybe_int(row.get("Tallanto занятий")):
        return "Вижу, что у вас уже был опыт обучения у нас, поэтому можем подобрать логичное продолжение."
    if maybe_int(row.get("AMO открытых среди последних")):
        return "Возвращаюсь к нашему текущему обсуждению, чтобы не потерять удобный момент."
    if as_text(row.get("Есть неотвеченный входящий")) == "Да":
        return "Возвращаюсь к вашему сообщению, чтобы разговор не потерялся."
    if as_text(row.get("Последний входящий Telegram")):
        return "Возвращаюсь к нашему диалогу и хочу аккуратно продолжить обсуждение."
    return ""


def next_step_phrase(row: pd.Series) -> str:
    next_step = as_text(row.get("Следующий шаг"))
    if not next_step:
        return ""
    lowered = next_step.lower()
    if "май" in lowered:
        return "Если вам удобно, могу в ближайшее время коротко написать по вариантам и вернуться к вопросу в мае."
    if "август" in lowered:
        return "Если удобнее, могу уже сейчас кратко прислать варианты, а к августу вернуться к финальному выбору."
    if "материал" in lowered:
        return "Если актуально, я могу сразу прислать короткую информацию и несколько подходящих вариантов."
    if "оплат" in lowered or "зафиксировать" in lowered:
        return "Если удобно, я могу коротко расписать варианты и помочь спокойно дойти до решения."
    return "Если актуально, я могу коротко прислать варианты и помочь выбрать самый удобный."


def build_draft(row: pd.Series) -> str:
    greeting_name = recipient_name(row)
    greeting = f"Здравствуйте, {greeting_name}!" if greeting_name else "Здравствуйте!"
    student = student_phrase(row)
    offer = offer_phrase(row)
    continuity = continuity_phrase(row)
    constraints = constraints_phrase(row)
    next_step = next_step_phrase(row)

    sentence_two = f"Пишу {student}: можем подобрать {offer}"
    if constraints:
        sentence_two += f" {constraints}"
    sentence_two += "."

    cta = next_step or "Если удобно, я могу коротко прислать варианты и помочь выбрать наиболее подходящий формат."
    draft = " ".join(part for part in [greeting, continuity, sentence_two, cta] if part)
    return truncate(re.sub(r"\s+", " ", draft).strip(), 500)


def add_dataframe_sheet(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title=title)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, pd.Timestamp):
                if pd.isna(value):
                    value = ""
                else:
                    if value.tzinfo is not None:
                        value = value.tz_convert("UTC").tz_localize(None)
                    value = value.to_pydatetime()
            elif isinstance(value, datetime) and value.tzinfo is not None:
                value = value.astimezone().replace(tzinfo=None)
            ws.cell(row=row_index, column=column_index, value=value)
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def build_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    ws = workbook.active
    ws.title = "Сводка"
    ws["A1"] = "Показатель"
    ws["B1"] = "Значение"
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    ws["B1"].font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    row_index = 2
    for key, value in summary.items():
        ws.cell(row=row_index, column=1, value=key)
        ws.cell(row=row_index, column=2, value=value)
        row_index += 1
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-csv",
        type=Path,
        default=PROJECT_ROOT / "stable_runtime" / "audits" / "telegram_openclaw_final_20260416" / "telegram_live_enrichment_refined.csv",
    )
    parser.add_argument(
        "--out-dir",
        type=Path,
        default=PROJECT_ROOT / "stable_runtime" / "audits" / "telegram_high_utility_drafts_20260416",
    )
    args = parser.parse_args()

    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(args.input_csv)
    if "Дата последнего сообщения Telegram" in df.columns:
        df["Дата последнего сообщения Telegram"] = pd.to_datetime(df["Дата последнего сообщения Telegram"], errors="coerce", utc=True)

    high_utility = df[(df["Персональное предложение возможно"] == "Да") & (df["Скор полезности"] >= 30)].copy()
    high_utility["Черновик сообщения"] = high_utility.apply(build_draft, axis=1)

    high_utility_short_cols = [
        "OpenClaw приоритет",
        "Телефон",
        "Диалоги Telegram",
        "Username Telegram",
        "Дата последнего сообщения Telegram",
        "ФИО родителя",
        "ФИО ребенка",
        "Краткая история общения",
        "Хронология общения",
        "Продукты интереса",
        "Возражения",
        "Следующий шаг",
        "Приоритет лида",
        "Вероятность продажи, %",
        "AMO контакты",
        "AMO ответственные",
        "AMO последние сделки",
        "AMO последние UTM",
        "AMO причина/стоп-фактор",
        "Tallanto контакты",
        "Tallanto филиалы",
        "Tallanto тип клиента",
        "Tallanto оплат",
        "Tallanto групп",
        "Tallanto занятий",
        "Tallanto краткий контекст",
        "Что предложить",
        "Цель сообщения",
        "Контекст для персонализации",
        "Стратегия касания",
        "Основа для предложения",
        "Гипотеза предложения",
        "Скор полезности",
        "Промпт для OpenClaw",
        "Черновик сообщения",
    ]

    top30 = high_utility.sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).head(30).copy()
    high_utility_short = high_utility[high_utility_short_cols].sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).copy()
    high_utility_full = high_utility.sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).copy()

    workbook = Workbook()
    summary = {
        "high_utility_rows": int(len(high_utility_short)),
        "top30_rows": int(len(top30)),
        "source_csv": str(args.input_csv),
        "xlsx": str(out_dir / "telegram_high_utility_drafts_96.xlsx"),
        "csv": str(out_dir / "telegram_high_utility_drafts_96.csv"),
    }
    build_summary_sheet(workbook, summary)
    add_dataframe_sheet(workbook, "96 High Utility Drafts", high_utility_short)
    add_dataframe_sheet(workbook, "96 High Utility Full", high_utility_full)
    add_dataframe_sheet(workbook, "Top 30", top30[high_utility_short_cols])
    workbook.save(out_dir / "telegram_high_utility_drafts_96.xlsx")

    high_utility_short.to_csv(out_dir / "telegram_high_utility_drafts_96.csv", index=False)
    (out_dir / "telegram_high_utility_drafts_96_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
