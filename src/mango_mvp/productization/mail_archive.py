from __future__ import annotations

import csv
import email
import hashlib
import io
import json
import logging
import os
import platform
import re
import signal
import sqlite3
import subprocess
import sys
import threading
import time
import zipfile
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from email import policy
from email.message import Message
from email.utils import getaddresses, parsedate_to_datetime
from pathlib import Path
from typing import Any, Mapping, Optional, Sequence
from urllib.parse import quote
from xml.etree import ElementTree as ET

from mango_mvp.services.ingest import SUPPORTED_EXTENSIONS, parse_filename_metadata
from mango_mvp.utils.phone import normalize_phone as canonical_normalize_phone
from mango_mvp.productization.mail_imap_snapshot import (
    HEADER_FETCH_QUERY,
    ImapClient,
    ImapLibClient,
    MailImapCredentials,
    clean_header,
    first_fetch_payload,
    parse_header_payload,
    parse_search_ids,
)


MAIL_ARCHIVE_SCHEMA_VERSION = "mail_archive_v1"
TALLANTO_IDENTITY_MAP_SCHEMA_VERSION = "tallanto_email_identity_map_v1"
MAIL_MATCHING_REPORT_SCHEMA_VERSION = "mail_matching_report_v1"
MAIL_MANGO_BRIDGE_PREVIEW_SCHEMA_VERSION = "mail_mango_bridge_preview_v1"
MAIL_PHONE_LIFT_PREVIEW_SCHEMA_VERSION = "mail_phone_lift_preview_v1"
MAIL_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION = "mail_customer_relink_preview_v1"
MAIL_STAGE2_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION = "mail_stage2_customer_relink_preview_v1"
MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION = "mail_attachment_parse_plan_v1"
MAIL_ATTACHMENT_TEXT_EXTRACT_SCHEMA_VERSION = "mail_attachment_text_extract_v1"
MAIL_ATTACHMENT_PDF_EXTRACT_SCHEMA_VERSION = "mail_attachment_pdf_extract_v1"
MAIL_ATTACHMENT_IMAGE_OCR_PLAN_SCHEMA_VERSION = "mail_attachment_image_ocr_plan_v1"
MAIL_ATTACHMENT_TEXT_INDEX_SCHEMA_VERSION = "mail_attachment_text_index_v1"
MAIL_ATTACHMENT_STAGE6_PLAN_SCHEMA_VERSION = "mail_attachment_stage6_plan_v1"
MAIL_ATTACHMENT_OCR_PREFLIGHT_SCHEMA_VERSION = "mail_attachment_ocr_preflight_v1"
MAIL_ATTACHMENT_OCR_PILOT_SCHEMA_VERSION = "mail_attachment_ocr_pilot_v1"
MANGO_PHONE_INDEX_PREVIEW_SCHEMA_VERSION = "mango_phone_index_preview_v1"
FULL_MESSAGE_FETCH_QUERY = "(BODY.PEEK[])"

DEFAULT_TALLANTO_EMAIL_COLUMNS = ("E-mail", "Другой E-mail")
DEFAULT_TALLANTO_PHONE_COLUMNS = (
    "Тел. (родителя)",
    "Тел. (доп.)",
    "Тел. (дом.)",
    "Другой тел.",
    "Помощник - тел.",
    "Тел. цифровой (моб.)",
    "Тел. цифровой (доп.)",
)
DEFAULT_TALLANTO_CANDIDATE_COLUMNS = (
    "ID",
    "ID Tallanto",
    "amoCRM ID",
    "Имя",
    "Фамилия",
    "ФИО родителя",
    "Тип ученика",
    "Филиал",
    "Ответственный(ая)",
    "Ответственный(ая) (ID)",
    "Группа(ID)",
)
TALLANTO_ID_COLUMNS = ("ID", "ID Tallanto", "Tallanto ID", "ID ученика")
DEFAULT_INTERNAL_DOMAINS = ("kmipt.ru",)
SERVICE_LOCAL_PARTS = (
    "no-reply",
    "noreply",
    "mailer-daemon",
    "postmaster",
    "notification",
    "notifications",
    "robot",
)
PILOT_SINCE_DAYS_RANGE = (1, 7)
PILOT_MAX_MESSAGES_RANGE = (1, 5)
LARGE_BATCH_SINCE_DAYS_RANGE = (1, 31)
LARGE_BATCH_MAX_MESSAGES_RANGE = (1, 500)
DEFAULT_ATTACHMENT_PARSE_ALLOW_EXTENSIONS = (
    ".pdf",
    ".docx",
    ".xlsx",
    ".csv",
    ".txt",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
)
DEFAULT_ATTACHMENT_PARSE_REVIEW_EXTENSIONS = (
    ".doc",
    ".xls",
    ".ppt",
    ".gif",
    ".heic",
    ".ics",
    ".html",
    ".htm",
)
DEFAULT_ATTACHMENT_TEXT_EXTRACT_EXTENSIONS = (".txt", ".csv", ".docx", ".xlsx")
DEFAULT_ATTACHMENT_PDF_EXTRACT_EXTENSIONS = (".pdf",)
DEFAULT_ATTACHMENT_IMAGE_OCR_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp")
ATTACHMENT_PARSE_BLOCK_EXTENSIONS = {
    ".exe",
    ".dll",
    ".msi",
    ".bat",
    ".cmd",
    ".ps1",
    ".vbs",
    ".js",
    ".jse",
    ".wsf",
    ".hta",
    ".scr",
    ".com",
    ".jar",
    ".sh",
    ".py",
    ".pl",
    ".rb",
    ".docm",
    ".xlsm",
    ".pptm",
    ".dotm",
    ".xltm",
    ".ppam",
    ".zip",
    ".rar",
    ".7z",
    ".iso",
    ".img",
    ".cab",
    ".ace",
    ".gz",
    ".tar",
}
ATTACHMENT_PARSE_ARCHIVE_EXTENSIONS = {
    ".zip",
    ".rar",
    ".7z",
    ".iso",
    ".img",
    ".cab",
    ".ace",
    ".gz",
    ".tar",
}
ATTACHMENT_PARSE_EXECUTABLE_EXTENSIONS = {
    ".exe",
    ".dll",
    ".msi",
    ".bat",
    ".cmd",
    ".ps1",
    ".vbs",
    ".js",
    ".jse",
    ".wsf",
    ".hta",
    ".scr",
    ".com",
    ".jar",
    ".sh",
    ".py",
    ".pl",
    ".rb",
}
ATTACHMENT_PARSE_MACRO_EXTENSIONS = {".docm", ".xlsm", ".pptm", ".dotm", ".xltm", ".ppam"}
ATTACHMENT_PARSE_BLOCK_CONTENT_TYPES = {
    "application/zip",
    "application/x-zip-compressed",
    "application/x-rar",
    "application/vnd.rar",
    "application/x-7z-compressed",
    "application/x-msdownload",
    "application/vnd.microsoft.portable-executable",
    "application/x-msdos-program",
    "application/java-archive",
    "application/x-sh",
}
ATTACHMENT_TEXT_EXTRACT_SUPPORTED_EXTENSIONS = set(DEFAULT_ATTACHMENT_TEXT_EXTRACT_EXTENSIONS)
ATTACHMENT_PDF_EXTRACT_SUPPORTED_EXTENSIONS = set(DEFAULT_ATTACHMENT_PDF_EXTRACT_EXTENSIONS)
ATTACHMENT_IMAGE_OCR_SUPPORTED_EXTENSIONS = set(DEFAULT_ATTACHMENT_IMAGE_OCR_EXTENSIONS)
PDF_BLOCK_TOKENS = {
    "/JavaScript",
    "/JS",
    "/OpenAction",
    "/AA",
    "/EmbeddedFile",
    "/EmbeddedFiles",
    "/Filespec",
    "/Launch",
    "/RichMedia",
    "/XFA",
    "/AcroForm",
    "/SubmitForm",
    "/ImportData",
    "/GoToR",
    "/GoToE",
    "/URI",
}
PDF_BLOCK_TOKEN_REASONS = {
    "/JavaScript": "pdf_javascript_declared",
    "/JS": "pdf_javascript_declared",
    "/OpenAction": "pdf_open_action_declared",
    "/AA": "pdf_additional_action_declared",
    "/EmbeddedFile": "pdf_embedded_file_declared",
    "/EmbeddedFiles": "pdf_embedded_file_declared",
    "/Filespec": "pdf_file_spec_declared",
    "/Launch": "pdf_launch_action_declared",
    "/RichMedia": "pdf_rich_media_declared",
    "/XFA": "pdf_xfa_form_declared",
    "/AcroForm": "pdf_acroform_declared",
    "/SubmitForm": "pdf_submit_form_declared",
    "/ImportData": "pdf_import_data_declared",
    "/GoToR": "pdf_remote_goto_declared",
    "/GoToE": "pdf_embedded_goto_declared",
    "/URI": "pdf_external_link_declared",
}


@dataclass(frozen=True)
class TallantoIdentityMapConfig:
    tallanto_csv_path: Path
    out_dir: Path
    encoding: str = "cp1251"
    delimiter: str = "\t"
    email_columns: Sequence[str] = DEFAULT_TALLANTO_EMAIL_COLUMNS
    phone_columns: Sequence[str] = DEFAULT_TALLANTO_PHONE_COLUMNS
    candidate_columns: Sequence[str] = DEFAULT_TALLANTO_CANDIDATE_COLUMNS


@dataclass(frozen=True)
class MailArchiveIngestConfig:
    out_dir: Path
    mailbox: str = "INBOX"
    mailbox_label: str = "INBOX"
    since_days: int = 30
    before_days: Optional[int] = None
    max_messages: int = 25
    account_label: str = "regru_edu"
    internal_domains: Sequence[str] = DEFAULT_INTERNAL_DOMAINS
    extracted_text_max_chars: int = 250_000
    exclude_message_sha256s: Sequence[str] = ()


@dataclass(frozen=True)
class MailArchivePreflightConfig:
    out_dir: Path
    mailbox: str = "INBOX"
    since_days: int = 3
    before_days: Optional[int] = None
    max_messages: int = 1
    account_label: str = "regru_edu"
    host: str = "mail.hosting.reg.ru"
    port: int = 993
    email_address: str = ""
    password_env_name: str = "MAIL_IMAP_PASSWORD"
    password_env_present: bool = False
    identity_db_path: Optional[Path] = None
    allow_large_batch: bool = False


@dataclass(frozen=True)
class MailMatchingReportConfig:
    archive_db_path: Path
    identity_db_path: Path
    out_dir: Path
    mailbox_email: str = ""
    internal_domains: Sequence[str] = DEFAULT_INTERNAL_DOMAINS


@dataclass(frozen=True)
class MailArchiveVerificationConfig:
    archive_dir: Path
    expected_max_messages: int = 5


@dataclass(frozen=True)
class MailCustomerHistoryHandoffConfig:
    archive_db_paths: Sequence[Path]
    identity_db_path: Path
    out_dir: Path
    mailbox_email: str = ""
    internal_domains: Sequence[str] = DEFAULT_INTERNAL_DOMAINS


@dataclass(frozen=True)
class MailMangoBridgePreviewConfig:
    mail_handoff_db_path: Path
    identity_db_path: Path
    product_db_path: Path
    out_dir: Path
    max_call_refs_per_candidate: int = 50
    mango_phone_index_db_path: Optional[Path] = None


@dataclass(frozen=True)
class MailPhoneLiftPreviewConfig:
    archive_db_paths: Sequence[Path]
    identity_db_path: Path
    out_dir: Path
    max_text_chars_per_message: int = 250_000


@dataclass(frozen=True)
class MailCustomerRelinkPreviewConfig:
    mail_handoff_db_path: Path
    identity_db_path: Path
    out_dir: Path
    classification_index_path: Optional[Path] = None
    live_phone_lookup: Optional[Any] = None
    require_real_correspondence: bool = True
    max_text_chars_per_message: int = 250_000


@dataclass(frozen=True)
class MailStage2CustomerRelinkPreviewConfig:
    event_jsonl_paths: Sequence[Path]
    identity_db_path: Path
    out_dir: Path
    internal_domains: Sequence[str] = DEFAULT_INTERNAL_DOMAINS
    max_text_chars_per_message: int = 250_000


@dataclass(frozen=True)
class MailAttachmentParsePlanConfig:
    archive_db_paths: Sequence[Path]
    out_dir: Path
    max_size_bytes: int = 20_000_000
    allow_extensions: Sequence[str] = DEFAULT_ATTACHMENT_PARSE_ALLOW_EXTENSIONS
    review_extensions: Sequence[str] = DEFAULT_ATTACHMENT_PARSE_REVIEW_EXTENSIONS


@dataclass(frozen=True)
class MailAttachmentTextExtractConfig:
    archive_db_paths: Sequence[Path]
    parse_plan_db_path: Path
    out_dir: Path
    stage_extensions: Sequence[str] = DEFAULT_ATTACHMENT_TEXT_EXTRACT_EXTENSIONS
    max_attachment_bytes: int = 10_000_000
    max_text_chars_per_attachment: int = 100_000
    max_csv_rows: int = 200
    max_csv_columns: int = 50
    max_xlsx_sheets: int = 5
    max_xlsx_rows_per_sheet: int = 100
    max_xlsx_columns_per_sheet: int = 50
    max_zip_members: int = 2_000
    max_zip_uncompressed_bytes: int = 50_000_000
    max_attachments: int = 0


@dataclass(frozen=True)
class MailAttachmentPdfExtractConfig:
    archive_db_paths: Sequence[Path]
    parse_plan_db_path: Path
    out_dir: Path
    stage_extensions: Sequence[str] = DEFAULT_ATTACHMENT_PDF_EXTRACT_EXTENSIONS
    max_attachment_bytes: int = 20_000_000
    max_pdf_pages: int = 10
    max_text_chars_per_attachment: int = 100_000
    max_page_text_chars: int = 25_000
    max_pdf_objects_to_scan: int = 10_000
    pdf_timeout_seconds: int = 10
    max_attachments: int = 0


@dataclass(frozen=True)
class MailAttachmentImageOcrPlanConfig:
    archive_db_paths: Sequence[Path]
    parse_plan_db_path: Path
    out_dir: Path
    stage_extensions: Sequence[str] = DEFAULT_ATTACHMENT_IMAGE_OCR_EXTENSIONS
    max_attachment_bytes: int = 20_000_000
    max_image_dimension: int = 20_000
    max_image_pixels: int = 50_000_000
    inspect_headers: bool = False
    max_attachments: int = 0


@dataclass(frozen=True)
class MailAttachmentTextIndexConfig:
    out_dir: Path
    text_extract_db_paths: Sequence[Path] = ()
    pdf_extract_db_paths: Sequence[Path] = ()
    image_ocr_plan_db_paths: Sequence[Path] = ()
    parse_plan_db_path: Optional[Path] = None
    max_rows: int = 0


@dataclass(frozen=True)
class MailAttachmentStage6PlanConfig:
    parse_plan_db_path: Path
    text_index_db_path: Path
    out_dir: Path
    ocr_pilot_limit: int = 15
    min_pilot_attachment_bytes: int = 0
    max_pilot_attachment_bytes: int = 5_000_000
    pilot_extensions: Sequence[str] = DEFAULT_ATTACHMENT_IMAGE_OCR_EXTENSIONS


@dataclass(frozen=True)
class MailAttachmentOcrPreflightConfig:
    archive_db_paths: Sequence[Path]
    stage6_plan_db_path: Path
    out_dir: Path
    max_candidates: int = 15
    max_attachment_bytes: int = 5_000_000
    verify_sha256: bool = True


@dataclass(frozen=True)
class MailAttachmentOcrPilotConfig:
    archive_db_paths: Sequence[Path]
    ocr_preflight_db_path: Path
    out_dir: Path
    max_candidates: int = 15
    max_attachment_bytes: int = 5_000_000
    languages: str = "rus+eng"
    page_segmentation_mode: int = 6
    tesseract_timeout_seconds: int = 30
    max_text_chars_per_attachment: int = 100_000
    workers: int = 1
    tesseract_thread_limit: int = 1
    reuse_existing_ocr_text: bool = False


@dataclass(frozen=True)
class MangoPhoneIndexPreviewConfig:
    product_db_path: Path
    recording_roots: Sequence[Path]
    out_dir: Path
    include_product_db: bool = True
    include_recording_filenames: bool = True


def normalize_email(value: object) -> str:
    text = clean_text(value).strip().strip(";,")
    if not text:
        return ""
    text = re.sub(r"^mailto:", "", text, flags=re.IGNORECASE).strip()
    if "<" in text and ">" in text:
        parsed = getaddresses([text])
        if parsed:
            text = parsed[0][1]
    text = text.strip().strip("<>;,").casefold()
    if not re.fullmatch(r"[a-z0-9.!#$%&'*+/=?^_`{|}~-]+@[a-z0-9.-]+\.[a-z0-9-]+", text):
        return ""
    local, domain = text.rsplit("@", 1)
    domain = domain.strip(".")
    if not local or not domain:
        return ""
    return f"{local}@{domain}"


def extract_email_addresses(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    found: list[str] = []
    for _name, address in getaddresses([text.replace(";", ",").replace("\n", ",")]):
        normalized = normalize_email(address)
        if normalized:
            found.append(normalized)
    for match in re.findall(
        r"[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Za-z0-9.-]+\.[A-Za-z0-9-]+",
        text,
    ):
        normalized = normalize_email(match)
        if normalized:
            found.append(normalized)
    return unique_preserving_order(found)


def normalize_phone(value: object) -> str:
    return canonical_normalize_phone(clean_text(value)) or ""


def extract_phone_numbers(value: object) -> list[str]:
    text = clean_text(value)
    if not text:
        return []
    candidates: list[str] = []
    for part in re.split(r"[,;\n\r/|]+", text):
        normalized = normalize_phone(part)
        if normalized:
            candidates.append(normalized)
    for match in re.findall(r"\+?\d[\d\s().-]{8,}\d", text):
        normalized = normalize_phone(match)
        if normalized:
            candidates.append(normalized)
    return unique_preserving_order(candidates)


def build_tallanto_identity_map(config: TallantoIdentityMapConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.out_dir, "identity map output directory")
    out_dir = config.out_dir.resolve(strict=False)
    out_dir.mkdir(parents=True, exist_ok=True)
    db_path = out_dir / "tallanto_email_identity_map.sqlite"
    report_path = out_dir / "tallanto_email_identity_map_report.json"

    rows = read_tallanto_rows(
        config.tallanto_csv_path,
        encoding=config.encoding,
        delimiter=config.delimiter,
    )
    tallanto_id_counts = Counter(tallanto_id_for_row(row) for row in rows if tallanto_id_for_row(row))
    candidate_rows: list[dict[str, Any]] = []
    identity_links: dict[tuple[str, str], dict[str, Any]] = {}
    row_identity_classes: dict[str, dict[str, str]] = {}

    for row_number, row in enumerate(rows, start=2):
        candidate_key = candidate_key_for_row(row_number, row, tallanto_id_counts=tallanto_id_counts)
        candidate_payload = {
            column: clean_text(row.get(column))
            for column in config.candidate_columns
            if column in row
        }
        tallanto_id = tallanto_id_for_row(row)
        if tallanto_id:
            candidate_payload.setdefault("ID", tallanto_id)
        candidate_rows.append(
            {
                "candidate_key": candidate_key,
                "row_number": row_number,
                "tallanto_id": tallanto_id,
                "amocrm_id": candidate_payload.get("amoCRM ID", ""),
                "first_name": candidate_payload.get("Имя", ""),
                "last_name": candidate_payload.get("Фамилия", ""),
                "parent_name": candidate_payload.get("ФИО родителя", ""),
                "student_type": candidate_payload.get("Тип ученика", ""),
                "manager": candidate_payload.get("Ответственный(ая)", ""),
                "manager_id": candidate_payload.get("Ответственный(ая) (ID)", ""),
                "group_id": candidate_payload.get("Группа(ID)", ""),
                "candidate_json": json.dumps(
                    candidate_payload,
                    ensure_ascii=False,
                    sort_keys=True,
                ),
            }
        )

        email_values = collect_row_identity_values(
            row,
            columns=config.email_columns,
            extractor=extract_email_addresses,
        )
        phone_values = collect_row_identity_values(
            row,
            columns=config.phone_columns,
            extractor=extract_phone_numbers,
        )
        for value, columns in email_values.items():
            append_identity_link(identity_links, "email", value, candidate_key, columns)
        for value, columns in phone_values.items():
            append_identity_link(identity_links, "phone", value, candidate_key, columns)
        row_identity_classes[candidate_key] = {
            "email": "missing" if not email_values else "pending",
            "phone": "missing" if not phone_values else "pending",
        }

    for item in identity_links.values():
        candidate_count = len(item["candidate_keys"])
        item["match_class"] = "strong_unique" if candidate_count == 1 else "duplicate"
        for candidate_key in item["candidate_keys"]:
            kind = item["kind"]
            current = row_identity_classes[candidate_key][kind]
            if current == "missing":
                row_identity_classes[candidate_key][kind] = item["match_class"]
            elif item["match_class"] == "duplicate":
                row_identity_classes[candidate_key][kind] = "duplicate"
            elif current == "pending":
                row_identity_classes[candidate_key][kind] = "strong_unique"

    for classes in row_identity_classes.values():
        for kind in ("email", "phone"):
            if classes[kind] == "pending":
                classes[kind] = "strong_unique"

    write_identity_map_db(
        db_path,
        source_path=config.tallanto_csv_path,
        candidate_rows=candidate_rows,
        identity_links=identity_links,
        row_identity_classes=row_identity_classes,
    )
    remove_sqlite_sidecars(db_path)

    report = build_identity_map_report(
        db_path=db_path,
        report_path=report_path,
        source_path=config.tallanto_csv_path,
        candidate_rows=candidate_rows,
        identity_links=identity_links,
        row_identity_classes=row_identity_classes,
        config=config,
        duplicate_tallanto_id_values=sum(1 for count in tallanto_id_counts.values() if count > 1),
    )
    write_json(report_path, report)
    return report


def build_mail_archive_preflight(config: MailArchivePreflightConfig) -> Mapping[str, Any]:
    out_dir = config.out_dir.resolve(strict=False)
    blocking_risks: list[str] = []
    warnings: list[str] = []

    try:
        guard_not_stable_runtime(out_dir, "mail archive preflight output directory")
        out_dir_not_stable_runtime = True
    except ValueError:
        out_dir_not_stable_runtime = False
        blocking_risks.append("out_dir_under_stable_runtime")

    out_dir_under_external_handoffs = "_external_handoffs" in out_dir.parts
    if not out_dir_under_external_handoffs:
        warnings.append("out_dir_not_under_external_handoffs")

    out_dir_git_ignored = git_check_ignored(out_dir)
    if not out_dir_git_ignored:
        blocking_risks.append("out_dir_not_git_ignored")

    email_normalized = normalize_email(config.email_address)
    if not email_normalized:
        blocking_risks.append("missing_or_invalid_mailbox_email")
    password_env_name = (
        config.password_env_name
        if valid_env_var_name(config.password_env_name)
        else "<invalid_env_var_name>"
    )
    if not valid_env_var_name(config.password_env_name):
        blocking_risks.append("invalid_password_env_name")
    elif not config.password_env_present:
        blocking_risks.append("missing_password_env")
    if not clean_text(config.mailbox):
        blocking_risks.append("missing_mailbox")
    if not clean_text(config.host):
        blocking_risks.append("missing_imap_host")
    if int(config.port or 0) <= 0:
        blocking_risks.append("invalid_imap_port")
    window_min, window_max = (
        LARGE_BATCH_SINCE_DAYS_RANGE if config.allow_large_batch else PILOT_SINCE_DAYS_RANGE
    )
    messages_min, messages_max = (
        LARGE_BATCH_MAX_MESSAGES_RANGE if config.allow_large_batch else PILOT_MAX_MESSAGES_RANGE
    )
    since_days = int(config.since_days)
    before_days = int(config.before_days) if config.before_days is not None else None
    if before_days is None:
        window_days = since_days
    elif before_days < 0:
        window_days = 0
        blocking_risks.append("before_days_must_be_non_negative")
    elif before_days >= since_days:
        window_days = 0
        blocking_risks.append("before_days_must_be_less_than_since_days")
    else:
        window_days = since_days - before_days
    if not (window_min <= window_days <= window_max):
        blocking_risks.append(
            f"{'batch' if config.allow_large_batch else 'pilot'}_window_days_must_be_"
            f"{window_min}_to_{window_max}"
        )
    if not (messages_min <= int(config.max_messages) <= messages_max):
        blocking_risks.append(
            f"{'batch' if config.allow_large_batch else 'pilot'}_max_messages_must_be_"
            f"{messages_min}_to_{messages_max}"
        )

    identity_db_exists = None
    identity_db_not_stable_runtime = None
    if config.identity_db_path is not None:
        identity_db = config.identity_db_path.resolve(strict=False)
        identity_db_exists = identity_db.exists()
        if not identity_db_exists:
            blocking_risks.append("identity_db_missing")
        try:
            guard_not_stable_runtime(identity_db, "identity database")
            identity_db_not_stable_runtime = True
        except ValueError:
            identity_db_not_stable_runtime = False
            blocking_risks.append("identity_db_under_stable_runtime")

    if out_dir.exists():
        warnings.append("out_dir_already_exists")

    report_path = out_dir / "mail_archive_preflight.json"
    report = {
        "schema_version": "mail_archive_preflight_v1",
        "created_at": utc_now(),
        "account_label": config.account_label,
        "preflight_pass": not blocking_risks,
        "blocking_risks": blocking_risks,
        "warnings": warnings,
        "batch_mode": "approved_large_batch" if config.allow_large_batch else "pilot",
        "requested_pilot": {
            "host": config.host,
            "port": config.port,
            "email_present": bool(email_normalized),
            "mailbox": config.mailbox,
            "since_days": since_days,
            "before_days": before_days,
            "window_days": window_days,
            "max_messages": int(config.max_messages),
        },
        "checks": {
            "password_env_name": password_env_name,
            "password_env_present": bool(config.password_env_present),
            "out_dir_not_stable_runtime": out_dir_not_stable_runtime,
            "out_dir_under_external_handoffs": out_dir_under_external_handoffs,
            "out_dir_git_ignored": out_dir_git_ignored,
            "identity_db_exists": identity_db_exists,
            "identity_db_not_stable_runtime": identity_db_not_stable_runtime,
        },
        "safety": {
            "network_calls": False,
            "imap_login": False,
            "send_mail": False,
            "delete_or_move_mail": False,
            "write_crm": False,
            "write_tallanto": False,
            "password_written": False,
            "run_asr": False,
            "run_ra": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
        },
        "recommended_command": (
            "python3 scripts/mango_office_mail_archive.py ingest "
            f"--email <redacted> --mailbox {shell_safe_token(config.mailbox)} "
            f"--since-days {since_days} "
            f"{f'--before-days {before_days} ' if before_days is not None else ''}"
            f"--max-messages {int(config.max_messages)} "
            f"--out-dir {shell_safe_token(str(config.out_dir))}"
            f"{' --allow-large-batch' if config.allow_large_batch else ''}"
        ),
        "paths": {
            "out_dir": str(out_dir),
            "preflight_report": str(report_path),
            "identity_db": str(config.identity_db_path) if config.identity_db_path else "",
        },
        "privacy": {
            "contains_password": False,
            "contains_raw_mail": False,
            "contains_raw_personal_values": False,
        },
    }
    if out_dir_not_stable_runtime:
        write_json(report_path, report)
    return report


def build_mail_archive_ingest(
    *,
    credentials: MailImapCredentials,
    config: MailArchiveIngestConfig,
    client: Optional[ImapClient] = None,
) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.out_dir, "mail archive output directory")
    out_dir = config.out_dir.resolve(strict=False)
    if client is None:
        guard_git_ignored_output(out_dir, "mail archive output directory")
    raw_dir = out_dir / "raw_eml"
    attachment_dir = out_dir / "attachments"
    text_dir = out_dir / "extracted_text"
    for path in (raw_dir, attachment_dir, text_dir):
        path.mkdir(parents=True, exist_ok=True)
    db_path = out_dir / "mail_archive.sqlite"
    report_path = out_dir / "mail_ingest_report.json"

    since_days = max(1, int(config.since_days))
    since_dt = (datetime.now(timezone.utc) - timedelta(days=since_days)).date()
    since_imap = since_dt.strftime("%d-%b-%Y")
    before_days = int(config.before_days) if config.before_days is not None else None
    before_imap = ""
    if before_days is not None:
        before_dt = (datetime.now(timezone.utc) - timedelta(days=max(0, before_days))).date()
        before_imap = before_dt.strftime("%d-%b-%Y")
    max_messages = max(0, int(config.max_messages))
    search_criteria = ["SINCE", since_imap]
    if before_imap:
        search_criteria.extend(["BEFORE", before_imap])

    report: dict[str, Any] = {
        "schema_version": MAIL_ARCHIVE_SCHEMA_VERSION,
        "created_at": utc_now(),
        "account_label": config.account_label,
        "host": credentials.host,
        "port": credentials.port,
        "email": credentials.email_address,
        "mailbox": config.mailbox_label,
        "mailbox_raw": config.mailbox,
        "since_days": since_days,
        "since_imap": since_imap,
        "before_days": before_days,
        "before_imap": before_imap,
        "window_days": since_days - before_days if before_days is not None else since_days,
        "search_criteria": search_criteria,
        "max_messages": max_messages,
        "safety": {
            "readonly_select": True,
            "fetch_uses_body_peek": True,
            "send_mail": False,
            "delete_or_move_mail": False,
            "write_crm": False,
            "write_tallanto": False,
            "open_attachments": False,
            "password_written": False,
            "run_asr": False,
            "run_ra": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
        },
        "messages_found_since": 0,
        "messages_attempted": 0,
        "messages_inserted_or_seen": 0,
        "messages_excluded_by_sha256": 0,
        "message_fetch_retries": 0,
        "imap_connection_retries": 0,
        "raw_eml_written": 0,
        "attachments_written": 0,
        "text_files_written": 0,
        "errors": [],
        "paths": {
            "archive_db": str(db_path),
            "raw_eml_dir": str(raw_dir),
            "attachments_dir": str(attachment_dir),
            "extracted_text_dir": str(text_dir),
            "report": str(report_path),
        },
    }

    init_mail_archive_db(db_path)
    imap: Optional[ImapClient] = client
    if imap is None:
        imap, connection_retries = open_imap_client_with_retries(credentials)
        report["imap_connection_retries"] = connection_retries
    try:
        login_status, _ = imap.login(credentials.email_address, credentials.password)
        report["login_status"] = login_status
        select_status, select_data = imap.select(config.mailbox, readonly=True)
        report["select_status"] = select_status
        report["mailbox_total_messages"] = int(select_data[0]) if select_data and select_data[0] else 0
        if select_status != "OK":
            raise RuntimeError(f"IMAP SELECT failed for {config.mailbox}: {select_status}")
        search_status, search_data = imap.search(None, *search_criteria)
        report["search_status"] = search_status
        message_ids = parse_search_ids(search_status, search_data)
        report["messages_found_since"] = len(message_ids)
        selected_ids = message_ids[-max_messages:] if max_messages > 0 else []
        report["messages_attempted"] = len(selected_ids)

        def record_fetched_message(fetched: Mapping[str, Any]) -> None:
            if fetched.get("excluded_by_sha256"):
                report["messages_excluded_by_sha256"] += 1
            else:
                report["messages_inserted_or_seen"] += 1
                report["raw_eml_written"] += int(bool(fetched["raw_eml_written"]))
                report["attachments_written"] += int(fetched["attachments_written"])
                report["text_files_written"] += int(bool(fetched["text_file_written"]))

        for msg_id in selected_ids:
            try:
                fetched = ingest_one_message(
                    imap,
                    msg_id=msg_id,
                    credentials=credentials,
                    config=config,
                    db_path=db_path,
                    raw_dir=raw_dir,
                    attachment_dir=attachment_dir,
                    text_dir=text_dir,
                )
                record_fetched_message(fetched)
            except Exception as exc:  # noqa: BLE001
                if client is None and is_transient_imap_fetch_error(exc):
                    report["message_fetch_retries"] += 1
                    try:
                        imap.close()
                    except Exception:  # noqa: BLE001
                        pass
                    try:
                        imap.logout()
                    except Exception:  # noqa: BLE001
                        pass
                    try:
                        imap = open_readonly_mailbox(credentials, config.mailbox)
                        fetched = ingest_one_message(
                            imap,
                            msg_id=msg_id,
                            credentials=credentials,
                            config=config,
                            db_path=db_path,
                            raw_dir=raw_dir,
                            attachment_dir=attachment_dir,
                            text_dir=text_dir,
                        )
                        record_fetched_message(fetched)
                        continue
                    except Exception as retry_exc:  # noqa: BLE001
                        report["errors"].append(
                            {
                                "imap_seq": msg_id.decode("ascii", "ignore"),
                                "error": f"{type(retry_exc).__name__}: {retry_exc}",
                                "first_error": f"{type(exc).__name__}: {exc}",
                            }
                        )
                        imap = None
                        break
                report["errors"].append(
                    {
                        "imap_seq": msg_id.decode("ascii", "ignore"),
                        "error": f"{type(exc).__name__}: {exc}",
                    }
                )
    finally:
        if imap is not None:
            try:
                imap.close()
            except Exception:  # noqa: BLE001
                pass
            try:
                imap.logout()
            except Exception:  # noqa: BLE001
                pass

    write_json(report_path, report)
    return report


def is_transient_imap_fetch_error(exc: Exception) -> bool:
    kind = type(exc).__name__.lower()
    text = str(exc).lower()
    transient_kinds = {"abort", "timeouterror", "timeout"}
    transient_markers = (
        "socket error",
        "eof",
        "timed out",
        "timeout",
        "handshake operation timed out",
        "connection reset",
        "nodename nor servname provided",
        "name or service not known",
        "temporary failure in name resolution",
        "temporarily unavailable",
    )
    return kind in transient_kinds or any(marker in text for marker in transient_markers)


def open_imap_client_with_retries(
    credentials: MailImapCredentials,
    *,
    attempts: int = 3,
    delay_seconds: float = 3.0,
) -> tuple[ImapClient, int]:
    retries = 0
    for attempt in range(1, max(1, attempts) + 1):
        try:
            return ImapLibClient(host=credentials.host, port=credentials.port), retries
        except Exception as exc:  # noqa: BLE001
            if attempt >= attempts or not is_transient_imap_fetch_error(exc):
                raise
            retries += 1
            if delay_seconds > 0:
                time.sleep(delay_seconds)
    raise RuntimeError("IMAP connection retry loop exited unexpectedly")


def open_readonly_mailbox(credentials: MailImapCredentials, mailbox: str) -> ImapClient:
    imap, _ = open_imap_client_with_retries(credentials)
    login_status, _ = imap.login(credentials.email_address, credentials.password)
    if login_status != "OK":
        raise RuntimeError(f"IMAP LOGIN failed during retry: {login_status}")
    select_status, _ = imap.select(mailbox, readonly=True)
    if select_status != "OK":
        raise RuntimeError(f"IMAP SELECT failed during retry for {mailbox}: {select_status}")
    return imap


def build_mail_matching_report(config: MailMatchingReportConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.out_dir, "matching report output directory")
    guard_not_stable_runtime(config.archive_db_path, "mail archive database")
    out_dir = config.out_dir.resolve(strict=False)
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "mail_matching_report.json"

    identity_index = load_identity_index(config.identity_db_path)
    internal_domains = normalize_domains(config.internal_domains, config.mailbox_email)

    with sqlite3.connect(str(config.archive_db_path)) as con:
        con.row_factory = sqlite3.Row
        init_mail_match_tables(con)
        con.execute("DELETE FROM message_matches")
        messages = list(con.execute("SELECT * FROM messages ORDER BY message_date_iso, sha256"))
        counts: dict[str, int] = {
            "strong_unique": 0,
            "ambiguous": 0,
            "missing": 0,
            "internal_or_service": 0,
        }
        kind_counts: Counter[str] = Counter()
        class_by_kind: dict[str, Counter[str]] = {}
        candidate_messages: set[tuple[str, str]] = set()
        for message in messages:
            participants = list(
                con.execute(
                    "SELECT * FROM message_participants WHERE message_sha256 = ?",
                    (message["sha256"],),
                )
            )
            external_emails = [
                row["email_normalized"]
                for row in participants
                if row["email_normalized"]
                and not email_domain_is_internal(row["email_normalized"], internal_domains)
            ]
            external_emails = unique_preserving_order(external_emails)
            message_kind = clean_text(message["message_kind"])
            kind_counts[message_kind] += 1
            if message_kind in {"internal", "service"}:
                match_class = "internal_or_service"
                candidate_keys: list[str] = []
            else:
                candidate_keys = unique_preserving_order(
                    candidate_key
                    for address in external_emails
                    for candidate_key in identity_index.get(address, [])
                )
                if len(candidate_keys) == 1:
                    match_class = "strong_unique"
                elif len(candidate_keys) > 1:
                    match_class = "ambiguous"
                else:
                    match_class = "missing"
            counts[match_class] += 1
            class_by_kind.setdefault(message_kind, Counter())[match_class] += 1
            con.execute(
                """
                INSERT OR REPLACE INTO message_matches (
                  message_sha256, match_class, candidate_count, matched_email_count, matched_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (
                    message["sha256"],
                    match_class,
                    len(candidate_keys),
                    sum(1 for address in external_emails if address in identity_index),
                    utc_now(),
                ),
            )
            for candidate_key in candidate_keys:
                candidate_messages.add((candidate_key, message["sha256"]))
        con.commit()

    report = {
        "schema_version": MAIL_MATCHING_REPORT_SCHEMA_VERSION,
        "created_at": utc_now(),
        "archive_db_path": str(config.archive_db_path),
        "identity_db_path": str(config.identity_db_path),
        "message_count": len(messages),
        "matched_message_count": counts["strong_unique"] + counts["ambiguous"],
        "distinct_matched_candidates": len({candidate for candidate, _sha in candidate_messages}),
        "counts": counts,
        "message_kind_counts": {kind: int(count) for kind, count in sorted(kind_counts.items())},
        "match_class_by_message_kind": {
            kind: {match_class: int(count) for match_class, count in sorted(class_counts.items())}
            for kind, class_counts in sorted(class_by_kind.items())
        },
        "privacy": {
            "contains_raw_personal_values": False,
            "raw_emails_written": False,
            "raw_names_written": False,
            "raw_phones_written": False,
        },
        "paths": {"report": str(report_path)},
    }
    write_json(report_path, report)
    return report


def verify_mail_archive_pilot(config: MailArchiveVerificationConfig) -> Mapping[str, Any]:
    archive_dir = config.archive_dir.resolve(strict=False)
    report_path = archive_dir / "mail_ingest_report.json"
    db_path = archive_dir / "mail_archive.sqlite"
    verification_path = archive_dir / "mail_archive_verification.json"
    raw_dir = archive_dir / "raw_eml"
    text_dir = archive_dir / "extracted_text"
    attachment_dir = archive_dir / "attachments"

    blocking_risks: list[str] = []
    warnings: list[str] = []
    try:
        guard_not_stable_runtime(archive_dir, "mail archive verification directory")
        archive_dir_not_stable_runtime = True
    except ValueError:
        archive_dir_not_stable_runtime = False
        blocking_risks.append("archive_dir_under_stable_runtime")

    archive_dir_git_ignored = git_check_ignored(archive_dir)
    if not archive_dir_git_ignored:
        blocking_risks.append("archive_dir_not_git_ignored")

    ingest_report: Mapping[str, Any] = {}
    if not report_path.exists():
        blocking_risks.append("mail_ingest_report_missing")
    else:
        try:
            ingest_report = json.loads(report_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            blocking_risks.append("mail_ingest_report_invalid_json")

    if not db_path.exists():
        blocking_risks.append("mail_archive_db_missing")

    safety = ingest_report.get("safety") if isinstance(ingest_report, Mapping) else {}
    if isinstance(safety, Mapping):
        if safety.get("readonly_select") is not True:
            blocking_risks.append("readonly_select_not_confirmed")
        if safety.get("fetch_uses_body_peek") is not True:
            blocking_risks.append("body_peek_not_confirmed")
        for key in (
            "send_mail",
            "delete_or_move_mail",
            "write_crm",
            "write_tallanto",
            "password_written",
            "run_asr",
            "run_ra",
            "runtime_db_writes",
            "stable_runtime_writes",
        ):
            if safety.get(key) is not False:
                blocking_risks.append(f"{key}_not_false")
    elif ingest_report:
        blocking_risks.append("safety_block_missing")

    attempted = int(ingest_report.get("messages_attempted") or 0) if ingest_report else 0
    expected_max = int(config.expected_max_messages)
    if expected_max > 0 and attempted > expected_max:
        blocking_risks.append("messages_attempted_exceeds_expected_max")
    if ingest_report and ingest_report.get("errors"):
        warnings.append("ingest_report_contains_errors")

    db_counts = empty_archive_db_counts()
    db_schema_ok = False
    if db_path.exists():
        try:
            with sqlite3.connect(str(db_path)) as con:
                db_counts = archive_db_counts(con)
                db_schema_ok = archive_db_has_required_tables(con)
        except sqlite3.Error:
            blocking_risks.append("mail_archive_db_unreadable")
    if db_path.exists() and not db_schema_ok:
        blocking_risks.append("mail_archive_db_missing_required_tables")

    raw_file_count = count_files(raw_dir, "*.eml")
    text_file_count = count_files(text_dir, "*.txt")
    attachment_file_count = count_files(attachment_dir, "*.bin")
    if db_counts["messages"] != raw_file_count:
        blocking_risks.append("raw_eml_count_mismatch")
    if db_counts["messages"] and text_file_count > db_counts["messages"]:
        warnings.append("text_file_count_exceeds_message_count")
    if db_counts["attachments"] != attachment_file_count:
        warnings.append("attachment_file_count_differs_from_db_count")
    if db_counts["messages"] and db_counts["message_sources"] < db_counts["messages"]:
        blocking_risks.append("message_sources_less_than_messages")

    report_text = report_path.read_text(encoding="utf-8", errors="ignore") if report_path.exists() else ""
    if re.search(r"(?i)password\\s*[:=]", report_text):
        warnings.append("report_contains_password_key_text")

    verification = {
        "schema_version": "mail_archive_verification_v1",
        "created_at": utc_now(),
        "verification_pass": not blocking_risks,
        "blocking_risks": blocking_risks,
        "warnings": warnings,
        "archive_dir": str(archive_dir),
        "expected_max_messages": expected_max,
        "checks": {
            "archive_dir_not_stable_runtime": archive_dir_not_stable_runtime,
            "archive_dir_git_ignored": archive_dir_git_ignored,
            "mail_ingest_report_exists": report_path.exists(),
            "mail_archive_db_exists": db_path.exists(),
            "mail_archive_db_schema_ok": db_schema_ok,
            "safety_block_present": isinstance(safety, Mapping),
        },
        "ingest_summary": {
            "messages_found_since": ingest_report.get("messages_found_since", 0),
            "messages_attempted": attempted,
            "messages_inserted_or_seen": ingest_report.get("messages_inserted_or_seen", 0),
            "raw_eml_written": ingest_report.get("raw_eml_written", 0),
            "attachments_written": ingest_report.get("attachments_written", 0),
            "text_files_written": ingest_report.get("text_files_written", 0),
            "error_count": len(ingest_report.get("errors") or []) if ingest_report else 0,
        },
        "db_counts": db_counts,
        "file_counts": {
            "raw_eml": raw_file_count,
            "extracted_text": text_file_count,
            "attachments": attachment_file_count,
        },
        "paths": {
            "archive_db": str(db_path),
            "ingest_report": str(report_path),
            "verification_report": str(verification_path),
            "raw_eml_dir": str(raw_dir),
            "extracted_text_dir": str(text_dir),
            "attachments_dir": str(attachment_dir),
        },
        "privacy": {
            "verification_opens_raw_eml": False,
            "verification_opens_attachments": False,
            "contains_password": False,
            "contains_raw_mail": False,
            "contains_raw_personal_values": False,
        },
    }
    if archive_dir_not_stable_runtime:
        write_json(verification_path, verification)
    return verification


def build_mail_customer_history_handoff(
    config: MailCustomerHistoryHandoffConfig,
) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.out_dir, "customer history handoff output directory")
    guard_not_stable_runtime(config.identity_db_path, "identity database")
    out_dir = config.out_dir.resolve(strict=False)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_customer_history_handoff.sqlite"
    report_path = out_dir / "mail_customer_history_handoff_report.json"

    identity_index = load_identity_index(config.identity_db_path)
    internal_domains = normalize_domains(config.internal_domains, config.mailbox_email)
    source_paths = [path.resolve(strict=False) for path in config.archive_db_paths]
    for source_path in source_paths:
        guard_not_stable_runtime(source_path, "mail archive database")
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS source_archives (
              source_archive_id INTEGER PRIMARY KEY AUTOINCREMENT,
              archive_db_path TEXT NOT NULL UNIQUE,
              message_count INTEGER NOT NULL,
              imported_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS mail_customer_links (
              source_archive_id INTEGER NOT NULL,
              message_sha256 TEXT NOT NULL,
              mailbox TEXT,
              message_date_iso TEXT,
              message_kind TEXT NOT NULL,
              match_class TEXT NOT NULL,
              candidate_count INTEGER NOT NULL,
              matched_email_count INTEGER NOT NULL,
              candidate_keys_json TEXT NOT NULL,
              link_status TEXT NOT NULL,
              blocked_reason TEXT,
              PRIMARY KEY (source_archive_id, message_sha256)
            );
            DELETE FROM meta;
            DELETE FROM source_archives;
            DELETE FROM mail_customer_links;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", "mail_customer_history_handoff_v1"),
                ("created_at", utc_now()),
                ("identity_db_path", str(config.identity_db_path)),
            ],
        )

        for source_path in source_paths:
            with open_sqlite_readonly(source_path) as src:
                src.row_factory = sqlite3.Row
                messages = list(src.execute("SELECT * FROM messages ORDER BY message_date_iso, sha256"))
                out.execute(
                    """
                    INSERT INTO source_archives (archive_db_path, message_count, imported_at)
                    VALUES (?, ?, ?)
                    """,
                    (str(source_path), len(messages), utc_now()),
                )
                source_archive_id = int(out.execute("SELECT last_insert_rowid()").fetchone()[0])
                for message in messages:
                    participants = list(
                        src.execute(
                            "SELECT * FROM message_participants WHERE message_sha256 = ?",
                            (message["sha256"],),
                        )
                    )
                    match = classify_message_against_identity(
                        message_kind=clean_text(message["message_kind"]),
                        participants=participants,
                        identity_index=identity_index,
                        internal_domains=internal_domains,
                    )
                    out.execute(
                        """
                        INSERT INTO mail_customer_links (
                          source_archive_id, message_sha256, mailbox, message_date_iso,
                          message_kind, match_class, candidate_count, matched_email_count,
                          candidate_keys_json, link_status, blocked_reason
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            source_archive_id,
                            message["sha256"],
                            clean_text(message["mailbox"]),
                            clean_text(message["message_date_iso"]),
                            clean_text(message["message_kind"]),
                            match["match_class"],
                            len(match["candidate_keys"]),
                            int(match["matched_email_count"]),
                            json.dumps(match["candidate_keys"], ensure_ascii=False),
                            link_status_for_match_class(match["match_class"]),
                            blocked_reason_for_match_class(match["match_class"]),
                        ),
                    )
        out.executescript(
            """
            CREATE VIEW IF NOT EXISTS v_strong_customer_mail_links AS
            SELECT * FROM mail_customer_links WHERE match_class = 'strong_unique';
            CREATE VIEW IF NOT EXISTS v_manual_review_mail_links AS
            SELECT * FROM mail_customer_links WHERE match_class IN ('ambiguous', 'missing');
            """
        )
        counts = dict(
            out.execute(
                "SELECT match_class, COUNT(*) FROM mail_customer_links GROUP BY match_class"
            ).fetchall()
        )
        kind_counts = dict(
            out.execute(
                "SELECT message_kind, COUNT(*) FROM mail_customer_links GROUP BY message_kind"
            ).fetchall()
        )
        total_links = int(out.execute("SELECT COUNT(*) FROM mail_customer_links").fetchone()[0])
        distinct_candidates = {
            candidate
            for row in out.execute("SELECT candidate_keys_json FROM mail_customer_links")
            for candidate in json.loads(row[0] or "[]")
        }
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": "mail_customer_history_handoff_v1",
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "message_count": total_links,
        "counts": {
            "strong_unique": int(counts.get("strong_unique", 0)),
            "ambiguous": int(counts.get("ambiguous", 0)),
            "missing": int(counts.get("missing", 0)),
            "internal_or_service": int(counts.get("internal_or_service", 0)),
        },
        "message_kind_counts": {str(k): int(v) for k, v in kind_counts.items()},
        "distinct_candidate_keys": len(distinct_candidates),
        "safety": {
            "read_only_source_archives": True,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "store_raw_files_in_sqlite": False,
        },
        "privacy": {
            "contains_raw_personal_values": False,
            "raw_emails_written": False,
            "raw_names_written": False,
            "raw_phones_written": False,
            "raw_mail_written": False,
            "handoff_sqlite_contains_identity_refs": True,
        },
        "paths": {
            "handoff_db": str(out_db_path),
            "report": str(report_path),
        },
    }
    write_json(report_path, report)
    return report


def build_mango_phone_index_preview(config: MangoPhoneIndexPreviewConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.product_db_path, "Mango product database")
    for root in config.recording_roots:
        guard_not_stable_runtime(root, "Mango recording root")
    guard_not_stable_runtime(config.out_dir, "Mango phone index preview output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "Mango phone index preview output directory")
    guard_git_ignored_output(out_dir, "Mango phone index preview output directory")
    if config.include_product_db and not config.product_db_path.exists():
        raise FileNotFoundError(f"Mango product database not found: {config.product_db_path}")

    recording_roots = [Path(root).resolve(strict=False) for root in config.recording_roots]
    if config.include_recording_filenames:
        for root in recording_roots:
            if not root.exists():
                raise FileNotFoundError(f"Mango recording root not found: {root}")

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mango_phone_index_preview.sqlite"
    report_path = out_dir / "mango_phone_index_preview_report.json"

    rows_by_key: dict[tuple[str, str], dict[str, Any]] = {}
    product_source_counts: Mapping[str, int] = {}
    product_call_refs_seen = 0
    product_distinct_phones: set[str] = set()
    if config.include_product_db:
        product_by_phone, product_source_counts = load_mango_calls_by_phone(config.product_db_path)
        for phone, calls in product_by_phone.items():
            product_distinct_phones.add(phone)
            for call in calls:
                product_call_refs_seen += 1
                row = mango_phone_index_row_from_call(
                    phone=phone,
                    call=call,
                    source_kind=f"product_db:{clean_text(call.get('source_table')) or 'unknown'}",
                    source_path_sha256="",
                    source_filename_sha256="",
                    source_root_index=None,
                )
                rows_by_key.setdefault((phone, row["call_ref_key"]), row)

    recording_stats = (
        scan_recording_filename_phone_refs(recording_roots)
        if config.include_recording_filenames
        else {
            "rows": [],
            "files_seen": 0,
            "files_with_filename_phone": 0,
            "distinct_filename_phones": 0,
        }
    )
    for row in recording_stats["rows"]:
        rows_by_key.setdefault((row["normalized_phone"], row["call_ref_key"]), row)

    source_kind_counts = Counter(row["source_kind"] for row in rows_by_key.values())
    distinct_phones = {phone for phone, _call_ref in rows_by_key}

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS mango_phone_call_refs (
              normalized_phone TEXT NOT NULL,
              phone_sha256 TEXT NOT NULL,
              call_ref_key TEXT NOT NULL,
              source_kind TEXT NOT NULL,
              tenant_id TEXT NOT NULL,
              provider TEXT NOT NULL,
              event_key TEXT NOT NULL,
              provider_call_id TEXT NOT NULL,
              started_at TEXT,
              direction TEXT,
              manager_ref TEXT,
              recording_ref TEXT,
              status TEXT,
              product_call_started_at TEXT,
              duration_sec REAL,
              manager_extension TEXT,
              crm_owner_id INTEGER,
              crm_match_status TEXT,
              source_path_sha256 TEXT,
              source_filename_sha256 TEXT,
              source_root_index INTEGER,
              PRIMARY KEY (normalized_phone, call_ref_key)
            );
            DELETE FROM meta;
            DELETE FROM mango_phone_call_refs;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MANGO_PHONE_INDEX_PREVIEW_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("product_db_path", str(config.product_db_path)),
                ("recording_root_count", str(len(recording_roots))),
            ],
        )
        out.executemany(
            """
            INSERT INTO mango_phone_call_refs (
              normalized_phone, phone_sha256, call_ref_key, source_kind,
              tenant_id, provider, event_key, provider_call_id, started_at,
              direction, manager_ref, recording_ref, status,
              product_call_started_at, duration_sec, manager_extension,
              crm_owner_id, crm_match_status, source_path_sha256,
              source_filename_sha256, source_root_index
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    row["normalized_phone"],
                    row["phone_sha256"],
                    row["call_ref_key"],
                    row["source_kind"],
                    row["tenant_id"],
                    row["provider"],
                    row["event_key"],
                    row["provider_call_id"],
                    row["started_at"],
                    row["direction"],
                    row["manager_ref"],
                    row["recording_ref"],
                    row["status"],
                    row["product_call_started_at"],
                    row["duration_sec"],
                    row["manager_extension"],
                    row["crm_owner_id"],
                    row["crm_match_status"],
                    row["source_path_sha256"],
                    row["source_filename_sha256"],
                    row["source_root_index"],
                )
                for row in sorted(
                    rows_by_key.values(),
                    key=lambda item: (
                        clean_text(item.get("normalized_phone")),
                        clean_text(item.get("started_at")),
                        clean_text(item.get("call_ref_key")),
                    ),
                )
            ],
        )
        out.executescript(
            """
            CREATE VIEW IF NOT EXISTS v_mango_phone_index_source_counts AS
            SELECT source_kind, COUNT(*) AS call_refs, COUNT(DISTINCT normalized_phone) AS phones
            FROM mango_phone_call_refs
            GROUP BY source_kind;
            """
        )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MANGO_PHONE_INDEX_PREVIEW_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_file_counts": {
            "recording_root_count": len(recording_roots),
            "recording_files_seen": int(recording_stats["files_seen"]),
            "recording_files_with_filename_phone": int(recording_stats["files_with_filename_phone"]),
            "distinct_recording_filename_phones": int(recording_stats["distinct_filename_phones"]),
            "product_call_refs_seen": product_call_refs_seen,
            "distinct_product_db_phones": len(product_distinct_phones),
        },
        "product_source_counts": dict(product_source_counts),
        "phone_index_counts": {
            "rows_written": len(rows_by_key),
            "distinct_normalized_phones": len(distinct_phones),
            "duplicate_call_refs_collapsed": (
                product_call_refs_seen + len(recording_stats["rows"]) - len(rows_by_key)
            ),
            "source_kind_counts": {kind: int(count) for kind, count in sorted(source_kind_counts.items())},
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "read_only_product_db": True,
            "open_audio_files": False,
            "scan_filenames_only": True,
            "write_crm": False,
            "write_tallanto": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values": False,
            "raw_phones_written_to_json": False,
            "raw_filenames_written_to_json": False,
            "raw_paths_written_to_json": False,
            "phone_index_sqlite_contains_phone_refs": True,
            "source_paths_stored_as_sha256": True,
            "source_filenames_stored_as_sha256": True,
        },
        "paths": {
            "index_db": str(out_db_path),
            "report": str(report_path),
            "product_db": str(config.product_db_path),
            "recording_root_count": len(recording_roots),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_mango_bridge_preview(config: MailMangoBridgePreviewConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.mail_handoff_db_path, "mail handoff database")
    guard_not_stable_runtime(config.identity_db_path, "identity database")
    guard_not_stable_runtime(config.product_db_path, "Mango product database")
    if config.mango_phone_index_db_path is not None:
        guard_not_stable_runtime(config.mango_phone_index_db_path, "Mango phone index preview database")
    guard_not_stable_runtime(config.out_dir, "mail Mango bridge preview output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail Mango bridge preview output directory")
    guard_git_ignored_output(out_dir, "mail Mango bridge preview output directory")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_mango_bridge_preview.sqlite"
    report_path = out_dir / "mail_mango_bridge_preview_report.json"

    for source_path, label in (
        (config.mail_handoff_db_path, "mail handoff database"),
        (config.identity_db_path, "identity database"),
        (config.product_db_path, "Mango product database"),
        (config.mango_phone_index_db_path, "Mango phone index preview database"),
    ):
        if source_path is not None and not source_path.exists():
            raise FileNotFoundError(f"{label} not found: {source_path}")

    mail_summary = load_mail_handoff_candidate_summary(config.mail_handoff_db_path)
    candidate_refs = load_identity_candidate_refs(config.identity_db_path)
    candidate_phones = load_candidate_phone_links(config.identity_db_path)
    mango_by_phone, mango_source_counts = load_mango_calls_by_phone(config.product_db_path)
    if config.mango_phone_index_db_path is not None:
        index_by_phone, index_counts = load_mango_phone_index_calls_by_phone(
            config.mango_phone_index_db_path
        )
        mango_by_phone, index_rows_added = merge_mango_calls_by_phone(mango_by_phone, index_by_phone)
        mango_source_counts = {
            **mango_source_counts,
            "phone_index_enabled": 1,
            "phone_index_call_refs_loaded": int(index_counts["call_refs_loaded"]),
            "phone_index_distinct_phones": int(index_counts["distinct_normalized_phones"]),
            "phone_index_rows_added_after_dedupe": index_rows_added,
            "combined_distinct_normalized_phones": len(mango_by_phone),
        }
    else:
        mango_source_counts = {
            **mango_source_counts,
            "phone_index_enabled": 0,
            "combined_distinct_normalized_phones": len(mango_by_phone),
        }

    status_counts: Counter[str] = Counter()
    blocked_reason_counts: Counter[str] = Counter()
    phone_ref_rows = 0
    call_ref_rows = 0
    resolved_candidates_with_calls = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS candidate_mango_preview (
              candidate_key TEXT PRIMARY KEY,
              tallanto_id TEXT,
              amocrm_id TEXT,
              mail_message_count INTEGER NOT NULL,
              first_mail_date_iso TEXT,
              last_mail_date_iso TEXT,
              phone_value_count INTEGER NOT NULL,
              strong_phone_value_count INTEGER NOT NULL,
              duplicate_phone_value_count INTEGER NOT NULL,
              mango_call_count INTEGER NOT NULL,
              first_call_started_at TEXT,
              last_call_started_at TEXT,
              bridge_status TEXT NOT NULL,
              blocked_reason TEXT
            );
            CREATE TABLE IF NOT EXISTS candidate_phone_refs (
              candidate_key TEXT NOT NULL,
              normalized_phone TEXT NOT NULL,
              phone_sha256 TEXT NOT NULL,
              phone_match_class TEXT NOT NULL,
              phone_candidate_count INTEGER NOT NULL,
              PRIMARY KEY (candidate_key, normalized_phone)
            );
            CREATE TABLE IF NOT EXISTS mango_call_refs (
              candidate_key TEXT NOT NULL,
              normalized_phone TEXT NOT NULL,
              call_ref_key TEXT NOT NULL,
              tenant_id TEXT NOT NULL,
              provider TEXT NOT NULL,
              event_key TEXT NOT NULL,
              provider_call_id TEXT NOT NULL,
              started_at TEXT,
              direction TEXT,
              manager_ref TEXT,
              recording_ref TEXT,
              status TEXT,
              product_call_started_at TEXT,
              duration_sec REAL,
              manager_extension TEXT,
              crm_owner_id INTEGER,
              crm_match_status TEXT,
              PRIMARY KEY (candidate_key, normalized_phone, call_ref_key)
            );
            CREATE TABLE IF NOT EXISTS blocked_bridge_candidates (
              candidate_key TEXT PRIMARY KEY,
              blocked_reason TEXT NOT NULL
            );
            DELETE FROM meta;
            DELETE FROM candidate_mango_preview;
            DELETE FROM candidate_phone_refs;
            DELETE FROM mango_call_refs;
            DELETE FROM blocked_bridge_candidates;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_MANGO_BRIDGE_PREVIEW_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("mail_handoff_db_path", str(config.mail_handoff_db_path)),
                ("identity_db_path", str(config.identity_db_path)),
                ("product_db_path", str(config.product_db_path)),
                ("mango_phone_index_db_path", str(config.mango_phone_index_db_path or "")),
            ],
        )

        for candidate_key, mail in sorted(mail_summary["candidate_summaries"].items()):
            refs = candidate_refs.get(candidate_key, {})
            phone_rows = candidate_phones.get(candidate_key, [])
            strong_phone_values = [
                row["normalized_phone"]
                for row in phone_rows
                if row["match_class"] == "strong_unique"
            ]
            duplicate_phone_rows = [
                row for row in phone_rows if row["match_class"] != "strong_unique"
            ]
            calls_by_key: dict[str, Mapping[str, Any]] = {}
            for phone in strong_phone_values:
                for call in mango_by_phone.get(phone, []):
                    calls_by_key.setdefault(str(call["call_ref_key"]), call)
            calls = sorted(
                calls_by_key.values(),
                key=lambda row: (clean_text(row.get("started_at")), clean_text(row.get("call_ref_key"))),
            )

            if not phone_rows:
                bridge_status = "blocked"
                blocked_reason = "no_phone_for_candidate"
            elif not strong_phone_values:
                bridge_status = "blocked"
                blocked_reason = "phone_multiple_candidates"
            elif not calls:
                bridge_status = "blocked"
                blocked_reason = "mango_no_phone_match"
            else:
                bridge_status = "resolved"
                blocked_reason = ""
                resolved_candidates_with_calls += 1

            status_counts[bridge_status] += 1
            if blocked_reason:
                blocked_reason_counts[blocked_reason] += 1
                out.execute(
                    """
                    INSERT INTO blocked_bridge_candidates (candidate_key, blocked_reason)
                    VALUES (?, ?)
                    """,
                    (candidate_key, blocked_reason),
                )

            out.execute(
                """
                INSERT INTO candidate_mango_preview (
                  candidate_key, tallanto_id, amocrm_id, mail_message_count,
                  first_mail_date_iso, last_mail_date_iso, phone_value_count,
                  strong_phone_value_count, duplicate_phone_value_count, mango_call_count,
                  first_call_started_at, last_call_started_at, bridge_status, blocked_reason
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    candidate_key,
                    clean_text(refs.get("tallanto_id")),
                    clean_text(refs.get("amocrm_id")),
                    int(mail["mail_message_count"]),
                    clean_text(mail["first_mail_date_iso"]),
                    clean_text(mail["last_mail_date_iso"]),
                    len(phone_rows),
                    len(strong_phone_values),
                    len(duplicate_phone_rows),
                    len(calls),
                    clean_text(calls[0].get("started_at")) if calls else "",
                    clean_text(calls[-1].get("started_at")) if calls else "",
                    bridge_status,
                    blocked_reason,
                ),
            )
            for row in phone_rows:
                out.execute(
                    """
                    INSERT INTO candidate_phone_refs (
                      candidate_key, normalized_phone, phone_sha256,
                      phone_match_class, phone_candidate_count
                    ) VALUES (?, ?, ?, ?, ?)
                    """,
                    (
                        candidate_key,
                        row["normalized_phone"],
                        hashlib.sha256(row["normalized_phone"].encode("utf-8")).hexdigest(),
                        row["match_class"],
                        int(row["candidate_count"]),
                    ),
                )
                phone_ref_rows += 1
            for call in calls[: max(0, int(config.max_call_refs_per_candidate))]:
                out.execute(
                    """
                    INSERT INTO mango_call_refs (
                      candidate_key, normalized_phone, call_ref_key, tenant_id, provider,
                      event_key, provider_call_id, started_at, direction, manager_ref,
                      recording_ref, status, product_call_started_at, duration_sec,
                      manager_extension, crm_owner_id, crm_match_status
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        candidate_key,
                        call["normalized_phone"],
                        call["call_ref_key"],
                        call["tenant_id"],
                        call["provider"],
                        call["event_key"],
                        call["provider_call_id"],
                        call["started_at"],
                        call["direction"],
                        call["manager_ref"],
                        call["recording_ref"],
                        call["status"],
                        call["product_call_started_at"],
                        call["duration_sec"],
                        call["manager_extension"],
                        call["crm_owner_id"],
                        call["crm_match_status"],
                    ),
                )
                call_ref_rows += 1
        out.executescript(
            """
            CREATE VIEW IF NOT EXISTS v_resolved_mail_mango_links AS
            SELECT * FROM candidate_mango_preview WHERE bridge_status = 'resolved';
            CREATE VIEW IF NOT EXISTS v_manual_review_mail_mango_links AS
            SELECT * FROM candidate_mango_preview WHERE bridge_status <> 'resolved';
            """
        )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    candidate_count = len(mail_summary["candidate_summaries"])
    report = {
        "schema_version": MAIL_MANGO_BRIDGE_PREVIEW_SCHEMA_VERSION,
        "created_at": utc_now(),
        "candidate_count": candidate_count,
        "mail_link_reconciliation": {
            **mail_summary["link_status_counts"],
            "total": int(mail_summary["total_links"]),
            "pass": (
                int(mail_summary["total_links"])
                == sum(int(value) for value in mail_summary["link_status_counts"].values())
            ),
        },
        "counts": {
            "resolved": int(status_counts.get("resolved", 0)),
            "blocked": candidate_count - int(status_counts.get("resolved", 0)),
            "no_phone_for_candidate": int(blocked_reason_counts.get("no_phone_for_candidate", 0)),
            "phone_multiple_candidates": int(blocked_reason_counts.get("phone_multiple_candidates", 0)),
            "mango_no_phone_match": int(blocked_reason_counts.get("mango_no_phone_match", 0)),
        },
        "mango_source_counts": mango_source_counts,
        "artifact_counts": {
            "candidate_phone_refs": phone_ref_rows,
            "mango_call_refs_written": call_ref_rows,
            "resolved_candidates_with_calls": resolved_candidates_with_calls,
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "source_db_attached_to_writer": False,
            "read_only_source_archives": True,
            "read_only_identity_db": True,
            "read_only_product_db": True,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
            "store_raw_files_in_sqlite": False,
        },
        "privacy": {
            "contains_raw_personal_values": False,
            "raw_emails_written": False,
            "raw_names_written": False,
            "raw_mail_written": False,
            "raw_payloads_written": False,
            "raw_audio_written": False,
            "bridge_sqlite_contains_phone_refs": True,
            "bridge_sqlite_contains_customer_refs": True,
        },
        "paths": {
            "preview_db": str(out_db_path),
            "report": str(report_path),
            "mail_handoff_db": str(config.mail_handoff_db_path),
            "identity_db": str(config.identity_db_path),
            "product_db": str(config.product_db_path),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_phone_lift_preview(config: MailPhoneLiftPreviewConfig) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.identity_db_path, "identity database")
    guard_not_stable_runtime(config.out_dir, "mail phone lift preview output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail phone lift preview output directory")
    guard_git_ignored_output(out_dir, "mail phone lift preview output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")
    if not config.identity_db_path.exists():
        raise FileNotFoundError(f"identity database not found: {config.identity_db_path}")

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_phone_lift_preview.sqlite"
    report_path = out_dir / "mail_phone_lift_preview_report.json"

    phone_identity = load_phone_identity_index(config.identity_db_path)
    source_rows = load_mail_phone_lift_source_messages(source_paths)

    original_counts: Counter[str] = Counter()
    lift_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    distinct_phone_hashes: set[str] = set()
    text_files_read = 0
    text_files_missing = 0
    text_files_unreadable = 0
    text_paths_rejected = 0
    total_extracted_phone_values = 0
    total_identity_phone_matches = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS message_phone_lift_preview (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              message_date_iso TEXT,
              original_match_class TEXT NOT NULL,
              text_available INTEGER NOT NULL,
              text_path_status TEXT NOT NULL,
              extracted_phone_count INTEGER NOT NULL,
              identity_phone_match_count INTEGER NOT NULL,
              identity_candidate_count INTEGER NOT NULL,
              lift_class TEXT NOT NULL,
              phone_sha256_json TEXT NOT NULL,
              candidate_keys_json TEXT NOT NULL,
              evaluated_at TEXT NOT NULL,
              PRIMARY KEY (source_archive_id, message_sha256)
            );
            DELETE FROM meta;
            DELETE FROM message_phone_lift_preview;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_PHONE_LIFT_PREVIEW_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("identity_db_path", str(config.identity_db_path)),
                ("source_archive_count", str(len(source_paths))),
            ],
        )

        for row in source_rows:
            original_match_class = clean_text(row["match_class"])
            original_counts[original_match_class] += 1
            source_counts[clean_text(row["source_archive_id"])] += 1

            text, text_path_status = read_safe_extracted_text(
                row["extracted_text_path"],
                archive_db_path=Path(row["source_archive_path"]),
                max_chars=config.max_text_chars_per_message,
            )
            text_available = bool(text)
            if text_path_status == "read":
                text_files_read += 1
            elif text_path_status == "missing":
                text_files_missing += 1
            elif text_path_status == "rejected":
                text_paths_rejected += 1
            elif text_path_status == "unreadable":
                text_files_unreadable += 1

            phone_values = extract_phone_numbers(text)
            phone_hashes = [
                hashlib.sha256(phone.encode("utf-8")).hexdigest()
                for phone in phone_values
            ]
            distinct_phone_hashes.update(phone_hashes)
            matched_phone_count = sum(1 for phone in phone_values if phone in phone_identity)
            total_extracted_phone_values += len(phone_values)
            total_identity_phone_matches += matched_phone_count
            candidate_keys = unique_preserving_order(
                candidate_key
                for phone in phone_values
                for candidate_key in phone_identity.get(phone, {}).get("candidate_keys", [])
            )

            if text_path_status != "read":
                lift_class = "no_text"
            elif not phone_values:
                lift_class = "no_phone_detected"
            elif not candidate_keys:
                lift_class = "phone_no_identity_match"
            elif len(candidate_keys) == 1:
                lift_class = "phone_strong_unique"
            else:
                lift_class = "phone_ambiguous"
            lift_counts[lift_class] += 1

            out.execute(
                """
                INSERT INTO message_phone_lift_preview (
                  source_archive_id, message_sha256, message_date_iso,
                  original_match_class, text_available, text_path_status,
                  extracted_phone_count, identity_phone_match_count,
                  identity_candidate_count, lift_class, phone_sha256_json,
                  candidate_keys_json, evaluated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["source_archive_id"],
                    row["message_sha256"],
                    row["message_date_iso"],
                    original_match_class,
                    int(text_available),
                    text_path_status,
                    len(phone_values),
                    matched_phone_count,
                    len(candidate_keys),
                    lift_class,
                    json.dumps(phone_hashes, ensure_ascii=False),
                    json.dumps(candidate_keys, ensure_ascii=False, sort_keys=True),
                    utc_now(),
                ),
            )
        out.executescript(
            """
            CREATE VIEW IF NOT EXISTS v_phone_strong_unique_lift AS
            SELECT * FROM message_phone_lift_preview
            WHERE lift_class = 'phone_strong_unique';
            CREATE VIEW IF NOT EXISTS v_phone_manual_review_lift AS
            SELECT * FROM message_phone_lift_preview
            WHERE lift_class <> 'phone_strong_unique';
            """
        )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    evaluated_count = len(source_rows)
    strong_lift_count = int(lift_counts.get("phone_strong_unique", 0))
    report = {
        "schema_version": MAIL_PHONE_LIFT_PREVIEW_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "evaluated_message_count": evaluated_count,
        "by_original_match_class": {
            "ambiguous": int(original_counts.get("ambiguous", 0)),
            "missing": int(original_counts.get("missing", 0)),
        },
        "lift_class_counts": {
            "phone_strong_unique": strong_lift_count,
            "phone_ambiguous": int(lift_counts.get("phone_ambiguous", 0)),
            "phone_no_identity_match": int(lift_counts.get("phone_no_identity_match", 0)),
            "no_phone_detected": int(lift_counts.get("no_phone_detected", 0)),
            "no_text": int(lift_counts.get("no_text", 0)),
        },
        "potential_lift": {
            "strong_unique_messages": strong_lift_count,
            "manual_review_messages": evaluated_count - strong_lift_count,
            "ambiguous_to_phone_strong_unique": count_phone_lift_rows(
                out_db_path,
                original_match_class="ambiguous",
                lift_class="phone_strong_unique",
            ),
            "missing_to_phone_strong_unique": count_phone_lift_rows(
                out_db_path,
                original_match_class="missing",
                lift_class="phone_strong_unique",
            ),
        },
        "text_access_counts": {
            "text_files_read": text_files_read,
            "text_files_missing": text_files_missing,
            "text_files_unreadable": text_files_unreadable,
            "text_paths_rejected": text_paths_rejected,
        },
        "artifact_counts": {
            "preview_rows_written": evaluated_count,
            "extracted_phone_values_seen": total_extracted_phone_values,
            "identity_phone_value_matches": total_identity_phone_matches,
            "distinct_phone_hashes": len(distinct_phone_hashes),
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "source_db_attached_to_writer": False,
            "read_only_source_archives": True,
            "read_only_identity_db": True,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
            "open_attachments": False,
            "read_extracted_text_files": True,
            "store_raw_files_in_sqlite": False,
        },
        "privacy": {
            "contains_raw_personal_values": False,
            "raw_emails_written": False,
            "raw_names_written": False,
            "raw_phones_written": False,
            "raw_mail_written": False,
            "raw_text_written": False,
            "phone_hashes_written": True,
            "preview_sqlite_contains_candidate_refs": True,
        },
        "paths": {
            "preview_db": str(out_db_path),
            "report": str(report_path),
            "identity_db": str(config.identity_db_path),
            "archive_db_count": len(source_paths),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_customer_relink_preview(config: MailCustomerRelinkPreviewConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.mail_handoff_db_path, "mail handoff database")
    guard_not_stable_runtime(config.identity_db_path, "identity database")
    guard_not_stable_runtime(config.out_dir, "mail customer relink preview output directory")
    if config.classification_index_path is not None:
        guard_not_stable_runtime(config.classification_index_path, "mail classification index")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail customer relink preview output directory")
    guard_git_ignored_output(out_dir, "mail customer relink preview output directory")
    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_customer_relink_preview.sqlite"
    report_path = out_dir / "mail_customer_relink_preview_report.json"

    for source_path, label in (
        (config.mail_handoff_db_path, "mail handoff database"),
        (config.identity_db_path, "identity database"),
        (config.classification_index_path, "mail classification index"),
    ):
        if source_path is not None and not Path(source_path).exists():
            raise FileNotFoundError(f"{label} not found: {source_path}")

    address_book = load_tallanto_customer_address_book(config.identity_db_path)
    classification = load_mail_classification_index(config.classification_index_path)
    source_archives, manual_rows, baseline_counts = load_manual_mail_handoff_rows(config.mail_handoff_db_path)

    decisions: list[dict[str, Any]] = []
    learned_rows: dict[str, dict[str, Any]] = {}
    archive_cache: dict[str, sqlite3.Connection] = {}
    reason_counts: Counter[str] = Counter()
    decision_counts: Counter[str] = Counter()
    live_counts: Counter[str] = Counter()

    try:
        for row in manual_rows:
            message_sha = clean_text(row["message_sha256"])
            archive_path = Path(source_archives[int(row["source_archive_id"])])
            message, participants = read_archive_message_for_relink(
                archive_path,
                message_sha256=message_sha,
                cache=archive_cache,
            )
            class_row = classification.get(message_sha, {})
            decision = decide_mail_customer_relink(
                row=row,
                archive_path=archive_path,
                message=message,
                participants=participants,
                address_book=address_book,
                classification=class_row,
                internal_domains=DEFAULT_INTERNAL_DOMAINS,
                live_phone_lookup=config.live_phone_lookup,
                live_counts=live_counts,
                require_real_correspondence=config.require_real_correspondence,
                max_text_chars_per_message=config.max_text_chars_per_message,
            )
            decisions.append(decision)
            decision_counts[decision["decision"]] += 1
            reason_counts[decision["reason"] or "linked"] += 1
            for learned in decision.get("learned_values", []):
                learned_rows.setdefault(clean_text(learned["idempotency_key"]), learned)
    finally:
        for con in archive_cache.values():
            con.close()

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS address_book_clients (
              tallanto_id TEXT PRIMARY KEY,
              candidate_count INTEGER NOT NULL,
              duplicate_tallanto_id INTEGER NOT NULL,
              emails_json TEXT NOT NULL,
              phones_json TEXT NOT NULL,
              common_emails_json TEXT NOT NULL,
              common_phones_json TEXT NOT NULL,
              brand_scope_json TEXT NOT NULL,
              amocrm_ids_json TEXT NOT NULL,
              names_json TEXT NOT NULL,
              managers_json TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS address_book_values (
              kind TEXT NOT NULL,
              value TEXT NOT NULL,
              value_sha256 TEXT NOT NULL,
              match_class TEXT NOT NULL,
              candidate_count INTEGER NOT NULL,
              tallanto_ids_json TEXT NOT NULL,
              brand_scope_json TEXT NOT NULL,
              common_value INTEGER NOT NULL,
              PRIMARY KEY (kind, value)
            );
            CREATE TABLE IF NOT EXISTS mail_relink_decisions (
              source_archive_id INTEGER NOT NULL,
              source_archive_path TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              message_date_iso TEXT,
              original_match_class TEXT NOT NULL,
              original_candidate_keys_json TEXT NOT NULL,
              decision TEXT NOT NULL,
              reason TEXT NOT NULL,
              tallanto_id TEXT,
              candidate_keys_json TEXT NOT NULL,
              signal_kind TEXT NOT NULL,
              signal_value_sha256 TEXT NOT NULL,
              phones_found_count INTEGER NOT NULL,
              external_email_count INTEGER NOT NULL,
              brand_signal TEXT NOT NULL,
              brand_source TEXT NOT NULL,
              message_classification TEXT NOT NULL,
              learned_values_json TEXT NOT NULL,
              evaluated_at TEXT NOT NULL,
              PRIMARY KEY (source_archive_id, message_sha256)
            );
            CREATE TABLE IF NOT EXISTS learned_address_book_values (
              idempotency_key TEXT PRIMARY KEY,
              tallanto_id TEXT NOT NULL,
              kind TEXT NOT NULL,
              value TEXT NOT NULL,
              value_sha256 TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              signal_kind TEXT NOT NULL,
              source TEXT NOT NULL,
              created_at TEXT NOT NULL
            );
            DELETE FROM meta;
            DELETE FROM address_book_clients;
            DELETE FROM address_book_values;
            DELETE FROM mail_relink_decisions;
            DELETE FROM learned_address_book_values;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("mail_handoff_db_path", str(config.mail_handoff_db_path)),
                ("identity_db_path", str(config.identity_db_path)),
                ("classification_index_path", str(config.classification_index_path or "")),
                ("live_tallanto_lookup", "1" if config.live_phone_lookup is not None else "0"),
            ],
        )
        for client in address_book["clients"].values():
            out.execute(
                """
                INSERT INTO address_book_clients (
                  tallanto_id, candidate_count, duplicate_tallanto_id,
                  emails_json, phones_json, common_emails_json, common_phones_json,
                  brand_scope_json, amocrm_ids_json, names_json, managers_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    client["tallanto_id"],
                    len(client["candidate_keys"]),
                    int(client["duplicate_tallanto_id"]),
                    json.dumps(sorted(client["emails"]), ensure_ascii=False),
                    json.dumps(sorted(client["phones"]), ensure_ascii=False),
                    json.dumps(sorted(client["common_emails"]), ensure_ascii=False),
                    json.dumps(sorted(client["common_phones"]), ensure_ascii=False),
                    json.dumps(sorted(client["brand_scope"]), ensure_ascii=False),
                    json.dumps(sorted(client["amocrm_ids"]), ensure_ascii=False),
                    json.dumps(sorted(client["names"]), ensure_ascii=False),
                    json.dumps(sorted(client["managers"]), ensure_ascii=False),
                ),
            )
        for values_by_kind in address_book["values"].values():
            for item in values_by_kind.values():
                out.execute(
                    """
                    INSERT INTO address_book_values (
                      kind, value, value_sha256, match_class, candidate_count,
                      tallanto_ids_json, brand_scope_json, common_value
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        item["kind"],
                        item["value"],
                        stable_value_hash(item["value"]),
                        item["match_class"],
                        int(item["candidate_count"]),
                        json.dumps(sorted(item["tallanto_ids"]), ensure_ascii=False),
                        json.dumps(sorted(item["brand_scope"]), ensure_ascii=False),
                        int(item["common_value"]),
                    ),
                )
        for decision in decisions:
            out.execute(
                """
                INSERT INTO mail_relink_decisions (
                  source_archive_id, source_archive_path, message_sha256, message_date_iso,
                  original_match_class, original_candidate_keys_json, decision, reason,
                  tallanto_id, candidate_keys_json, signal_kind, signal_value_sha256,
                  phones_found_count, external_email_count, brand_signal, brand_source,
                  message_classification, learned_values_json, evaluated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    int(decision["source_archive_id"]),
                    decision["source_archive_path"],
                    decision["message_sha256"],
                    decision["message_date_iso"],
                    decision["original_match_class"],
                    decision["original_candidate_keys_json"],
                    decision["decision"],
                    decision["reason"],
                    decision["tallanto_id"],
                    json.dumps(decision["candidate_keys"], ensure_ascii=False),
                    decision["signal_kind"],
                    decision["signal_value_sha256"],
                    int(decision["phones_found_count"]),
                    int(decision["external_email_count"]),
                    decision["brand_signal"],
                    decision["brand_source"],
                    decision["message_classification"],
                    json.dumps(decision["learned_values"], ensure_ascii=False, sort_keys=True),
                    decision["evaluated_at"],
                ),
            )
        for learned in learned_rows.values():
            out.execute(
                """
                INSERT INTO learned_address_book_values (
                  idempotency_key, tallanto_id, kind, value, value_sha256,
                  message_sha256, signal_kind, source, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    learned["idempotency_key"],
                    learned["tallanto_id"],
                    learned["kind"],
                    learned["value"],
                    stable_value_hash(learned["value"]),
                    learned["message_sha256"],
                    learned["signal_kind"],
                    learned["source"],
                    learned["created_at"],
                ),
            )
        out.executescript(
            """
            CREATE VIEW IF NOT EXISTS v_linked_mail_relinks AS
            SELECT * FROM mail_relink_decisions WHERE decision = 'linked';
            CREATE VIEW IF NOT EXISTS v_unmatched_mail_relinks AS
            SELECT * FROM mail_relink_decisions WHERE decision <> 'linked';
            """
        )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    ready_before = int(baseline_counts.get("ready", 0))
    manual_before = int(baseline_counts.get("manual_review", 0))
    linkable_before = ready_before + manual_before
    linked_new = int(decision_counts.get("linked", 0))
    ready_after = ready_before + linked_new
    report = {
        "schema_version": MAIL_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION,
        "created_at": utc_now(),
        "paths": {
            "preview_db": str(out_db_path),
            "report": str(report_path),
            "mail_handoff_db": str(config.mail_handoff_db_path),
            "identity_db": str(config.identity_db_path),
            "classification_index": str(config.classification_index_path or ""),
        },
        "baseline": {
            "ready": ready_before,
            "manual_review": manual_before,
            "excluded": int(baseline_counts.get("excluded", 0)),
            "linkable_messages": linkable_before,
            "ready_share": ready_before / linkable_before if linkable_before else 0.0,
        },
        "after_preview": {
            "new_links": linked_new,
            "ready": ready_after,
            "manual_review_remaining": max(0, manual_before - linked_new),
            "linkable_messages": linkable_before,
            "ready_share": ready_after / linkable_before if linkable_before else 0.0,
        },
        "address_book": {
            "clients": len(address_book["clients"]),
            "values": sum(len(values) for values in address_book["values"].values()),
            "duplicate_tallanto_ids": int(address_book["duplicate_tallanto_ids"]),
            "learned_values": len(learned_rows),
        },
        "decision_counts": dict(sorted(decision_counts.items())),
        "unmatched_reasons": dict(
            sorted(
                (k, v)
                for k, v in reason_counts.items()
                if k not in {"linked", "already_linked"}
            )
        ),
        "live_tallanto": {
            "enabled": config.live_phone_lookup is not None,
            "counts": dict(sorted(live_counts.items())),
        },
        "review_examples": build_relink_review_examples(decisions, limit=20),
        "safety": {
            "live_crm_reads": config.live_phone_lookup is not None,
            "read_only_identity_db": True,
            "read_only_source_archives": True,
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "write_crm": False,
            "write_tallanto": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "store_raw_files_in_sqlite": False,
            "raw_personal_values_in_public_report": False,
        },
    }
    write_json(report_path, report)
    return report


def build_mail_stage2_customer_relink_preview(
    config: MailStage2CustomerRelinkPreviewConfig,
) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.identity_db_path, "identity database")
    guard_not_stable_runtime(config.out_dir, "stage2 relink preview output directory")
    for path in config.event_jsonl_paths:
        guard_not_stable_runtime(path, "stage2 event jsonl")
        if not Path(path).exists():
            raise FileNotFoundError(f"stage2 event jsonl not found: {path}")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "stage2 relink preview output directory")
    guard_git_ignored_output(out_dir, "stage2 relink preview output directory")
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "mail_stage2_customer_relink_preview_report.json"
    decisions_path = out_dir / "mail_stage2_customer_relink_preview_decisions.csv"

    address_book = load_tallanto_customer_address_book(config.identity_db_path)
    events = load_stage2_mail_events(config.event_jsonl_paths)
    prepared_events = [
        {
            "event": event,
            "emails": stage2_event_external_emails(
                event,
                internal_domains=config.internal_domains,
            ),
            "phones": stage2_event_phone_signals(
                event,
                max_chars=config.max_text_chars_per_message,
            ),
        }
        for event in events
    ]
    value_brand_scope = build_stage2_value_brand_scope(prepared_events)

    decisions: list[dict[str, Any]] = []
    decision_counts: Counter[str] = Counter()
    reason_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    old_match_counts: Counter[str] = Counter()
    baseline_linked = 0
    for prepared in prepared_events:
        decision = decide_stage2_event_customer_relink(
            prepared["event"],
            emails=prepared["emails"],
            phones=prepared["phones"],
            address_book=address_book,
            value_brand_scope=value_brand_scope,
        )
        decisions.append(decision)
        decision_counts[decision["decision"]] += 1
        reason_counts[decision["reason"] or "linked"] += 1
        source_counts[decision["source_file"]] += 1
        old_match_counts[decision["old_match_method"]] += 1
        baseline_linked += int(decision["baseline_linked"])

    write_stage2_relink_decisions_csv(decisions_path, decisions)
    new_links = int(decision_counts.get("linked", 0))
    event_count = len(decisions)
    report = {
        "schema_version": MAIL_STAGE2_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION,
        "created_at": utc_now(),
        "paths": {
            "report": str(report_path),
            "decisions_csv": str(decisions_path),
            "identity_db": str(config.identity_db_path),
            "event_jsonl": [str(path) for path in config.event_jsonl_paths],
        },
        "input": {
            "events": event_count,
            "source_files": dict(sorted(source_counts.items())),
            "old_match_methods": dict(sorted(old_match_counts.items())),
        },
        "baseline": {
            "linked": baseline_linked,
            "unlinked": max(0, event_count - baseline_linked),
            "linked_share": baseline_linked / event_count if event_count else 0.0,
        },
        "after_preview": {
            "new_links": new_links,
            "linked": baseline_linked + new_links,
            "unlinked_remaining": max(0, event_count - baseline_linked - new_links),
            "linked_share": (baseline_linked + new_links) / event_count if event_count else 0.0,
        },
        "address_book": {
            "clients": len(address_book["clients"]),
            "values": sum(len(values) for values in address_book["values"].values()),
            "duplicate_tallanto_ids": int(address_book["duplicate_tallanto_ids"]),
            "brand_from_tallanto_trusted": False,
        },
        "decision_counts": dict(sorted(decision_counts.items())),
        "unmatched_reasons": dict(
            sorted(
                (k, v)
                for k, v in reason_counts.items()
                if k not in {"linked", "already_linked"}
            )
        ),
        "cross_brand_signal_values": {
            kind: sum(1 for brands in values.values() if len(brands) > 1)
            for kind, values in value_brand_scope.items()
        },
        "review_examples": build_stage2_relink_review_examples(decisions, limit=20),
        "safety": {
            "read_only_identity_db": True,
            "read_only_stage2_jsonl": True,
            "source_sqlite_query_only": False,
            "write_crm": False,
            "write_tallanto": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "raw_personal_values_in_public_report": False,
            "brand_source": "mail_event_channel_only",
        },
    }
    write_json(report_path, report)
    return report


def load_stage2_mail_events(paths: Sequence[Path]) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    for path in paths:
        with Path(path).open("r", encoding="utf-8") as fh:
            for line_number, line in enumerate(fh, start=1):
                if not line.strip():
                    continue
                payload = json.loads(line)
                if not isinstance(payload, dict):
                    continue
                payload = dict(payload)
                payload["_source_file"] = Path(path).name
                payload["_line_number"] = line_number
                events.append(payload)
    return events


def stage2_event_external_emails(
    event: Mapping[str, Any],
    *,
    internal_domains: Sequence[str],
) -> list[str]:
    domains = normalize_domains(internal_domains, "")
    result: list[str] = []

    def add_value(value: object) -> None:
        for email_value in extract_email_addresses(value):
            if email_domain_is_internal(email_value, domains):
                continue
            if email_value not in result:
                result.append(email_value)

    def walk(value: object) -> None:
        if isinstance(value, Mapping):
            add_value(value.get("email") or value.get("address") or value.get("email_normalized"))
            return
        if isinstance(value, Sequence) and not isinstance(value, (str, bytes, bytearray)):
            for item in value:
                walk(item)
            return
        add_value(value)

    for key in ("from", "to", "cc", "bcc", "reply_to", "participants"):
        if key in event:
            walk(event.get(key))
    return result


def stage2_event_phone_signals(event: Mapping[str, Any], *, max_chars: int) -> list[str]:
    text, status = read_safe_stage2_event_text(
        event.get("extracted_text_path"),
        max_chars=max_chars,
    )
    if status != "read":
        return []
    return extract_phone_numbers(text)


def read_safe_stage2_event_text(value: object, *, max_chars: int) -> tuple[str, str]:
    raw_path = clean_text(value)
    if not raw_path:
        return "", "missing"
    path = Path(raw_path).resolve(strict=False)
    if "_external_handoffs" not in path.parts:
        return "", "rejected"
    if "stable_runtime" in {part.casefold() for part in path.parts}:
        return "", "rejected"
    if "extracted_text" not in path.parts:
        return "", "rejected"
    if path.suffix.casefold() != ".txt":
        return "", "rejected"
    if not path.exists():
        return "", "missing"
    try:
        return path.read_text(encoding="utf-8", errors="replace")[: max(0, int(max_chars))], "read"
    except OSError:
        return "", "unreadable"


def build_stage2_value_brand_scope(
    prepared_events: Sequence[Mapping[str, Any]],
) -> dict[str, dict[str, set[str]]]:
    value_brand_scope: dict[str, dict[str, set[str]]] = {"email": {}, "phone": {}}
    for prepared in prepared_events:
        event = prepared["event"]
        brand = normalize_brand_token(event.get("brand") or event.get("brand_signal"))
        if brand not in {"foton", "unpk"}:
            continue
        for kind in ("email", "phone"):
            for value in prepared[f"{kind}s"]:
                value_brand_scope[kind].setdefault(value, set()).add(brand)
    return value_brand_scope


def normalize_existing_customer_id(value: object) -> str:
    text = clean_text(value)
    if text.startswith("tallanto:"):
        return text.removeprefix("tallanto:")
    return text


def decide_stage2_event_customer_relink(
    event: Mapping[str, Any],
    *,
    emails: Sequence[str],
    phones: Sequence[str],
    address_book: Mapping[str, Any],
    value_brand_scope: Mapping[str, Mapping[str, set[str]]],
) -> dict[str, Any]:
    message_sha = clean_text(event.get("message_sha256") or event.get("sha"))
    baseline_customer_id = normalize_existing_customer_id(event.get("customer_id"))
    old_match_method = clean_text(event.get("match_method") or event.get("match") or "unknown")
    brand_signal = normalize_brand_token(event.get("brand") or event.get("brand_signal"))
    base = {
        "source_file": clean_text(event.get("_source_file")),
        "line_number": int(event.get("_line_number") or 0),
        "message_sha256": message_sha,
        "date_iso": clean_text(event.get("date_iso") or event.get("date_first") or event.get("date")),
        "subject_hash": stable_value_hash(event.get("subject"))[:16] if event.get("subject") else "",
        "baseline_linked": bool(baseline_customer_id),
        "old_customer_id_hash": stable_value_hash(baseline_customer_id)[:16] if baseline_customer_id else "",
        "old_match_method": old_match_method,
        "decision": "already_linked" if baseline_customer_id else "unmatched",
        "reason": "",
        "tallanto_id": "",
        "tallanto_id_hash": "",
        "signal_kind": "",
        "signal_value_sha256": "",
        "email_signal_count": len(emails),
        "phone_signal_count": len(phones),
        "brand_signal": brand_signal or "unknown",
        "brand_source": clean_text(event.get("brand_source") or event.get("brand_src") or event.get("brand_note")),
    }
    if baseline_customer_id:
        return {**base, "reason": "already_linked"}
    blocked: list[str] = []
    for kind, values in (("email", emails), ("phone", phones)):
        linked: list[dict[str, Any]] = []
        for value in values:
            brands = value_brand_scope.get(kind, {}).get(value, set())
            if len(brands) > 1:
                blocked.append("cross_brand_signal")
                continue
            resolved = resolve_address_book_value(
                address_book,
                kind=kind,
                value=value,
                message_brand=brand_signal,
            )
            if resolved["status"] == "linked":
                linked.append({**resolved, "signal_kind": kind, "signal_value": value})
                continue
            blocked.append(resolved["reason"])
        target_ids = sorted({item["tallanto_id"] for item in linked if item.get("tallanto_id")})
        if len(target_ids) == 1:
            selected = next(item for item in linked if item["tallanto_id"] == target_ids[0])
            return {
                **base,
                "decision": "linked",
                "reason": "linked",
                "tallanto_id": selected["tallanto_id"],
                "tallanto_id_hash": stable_value_hash(selected["tallanto_id"])[:16],
                "signal_kind": selected["signal_kind"],
                "signal_value_sha256": stable_value_hash(selected["signal_value"]),
            }
        if len(target_ids) > 1:
            return {**base, "reason": "multiple_identity_targets", "decision": "unmatched"}
    reason = most_common_reason(blocked) or "no_identity_signal"
    return {**base, "reason": reason, "decision": "unmatched"}


def write_stage2_relink_decisions_csv(path: Path, decisions: Sequence[Mapping[str, Any]]) -> None:
    fieldnames = [
        "source_file",
        "line_number",
        "message_sha256",
        "date_iso",
        "subject_hash",
        "baseline_linked",
        "old_customer_id_hash",
        "old_match_method",
        "decision",
        "reason",
        "tallanto_id_hash",
        "signal_kind",
        "signal_value_sha256",
        "email_signal_count",
        "phone_signal_count",
        "brand_signal",
        "brand_source",
    ]
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in decisions:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def build_stage2_relink_review_examples(
    decisions: Sequence[Mapping[str, Any]],
    *,
    limit: int,
) -> list[dict[str, Any]]:
    linked = [row for row in decisions if row.get("decision") == "linked"]
    unmatched = [row for row in decisions if row.get("decision") == "unmatched"]
    result: list[dict[str, Any]] = []
    for row in linked[: max(0, limit // 2)] + unmatched[: max(0, limit - min(len(linked), limit // 2))]:
        result.append(
            {
                "source_file": clean_text(row.get("source_file")),
                "line_number": int(row.get("line_number") or 0),
                "message_sha256": clean_text(row.get("message_sha256"))[:16],
                "date": clean_text(row.get("date_iso")),
                "decision": clean_text(row.get("decision")),
                "reason": clean_text(row.get("reason")),
                "old_match_method": clean_text(row.get("old_match_method")),
                "tallanto_id_hash": clean_text(row.get("tallanto_id_hash")),
                "signal_kind": clean_text(row.get("signal_kind")),
                "signal_value_sha256": clean_text(row.get("signal_value_sha256"))[:16],
                "brand_signal": clean_text(row.get("brand_signal")),
            }
        )
        if len(result) >= limit:
            break
    return result


def load_tallanto_customer_address_book(identity_db_path: Path) -> dict[str, Any]:
    clients: dict[str, dict[str, Any]] = {}
    candidate_to_tallanto: dict[str, str] = {}
    with open_sqlite_readonly(identity_db_path) as con:
        for row in con.execute(
            """
            SELECT candidate_key, tallanto_id, amocrm_id, first_name, last_name,
                   parent_name, manager, candidate_json
            FROM identity_candidates
            ORDER BY tallanto_id, candidate_key
            """
        ):
            tallanto_id = clean_text(row["tallanto_id"])
            if not tallanto_id:
                continue
            candidate_key = clean_text(row["candidate_key"])
            candidate_to_tallanto[candidate_key] = tallanto_id
            payload = parse_json_object(row["candidate_json"])
            # Tallanto "Филиал" is a location field and is not a reliable Foton/UNPK
            # brand source. Mail relink brand isolation is derived from the mail
            # channel/event, never from Tallanto branch-like fields.
            brand_scope: set[str] = set()
            client = clients.setdefault(
                tallanto_id,
                {
                    "tallanto_id": tallanto_id,
                    "candidate_keys": set(),
                    "emails": set(),
                    "phones": set(),
                    "common_emails": set(),
                    "common_phones": set(),
                    "brand_scope": set(),
                    "amocrm_ids": set(),
                    "names": set(),
                    "managers": set(),
                    "duplicate_tallanto_id": False,
                },
            )
            client["candidate_keys"].add(candidate_key)
            client["brand_scope"].update(brand_scope)
            for key in ("amocrm_id", "manager"):
                if clean_text(row[key]):
                    client[f"{key}s" if key == "amocrm_id" else "managers"].add(clean_text(row[key]))
            for name in (row["first_name"], row["last_name"], row["parent_name"]):
                if clean_text(name):
                    client["names"].add(clean_text(name))

        for client in clients.values():
            client["duplicate_tallanto_id"] = len(client["candidate_keys"]) > 1

        values: dict[str, dict[str, dict[str, Any]]] = {"email": {}, "phone": {}}
        for row in con.execute(
            """
            SELECT il.kind, il.value, il.candidate_key, iv.match_class, iv.candidate_count
            FROM identity_links il
            JOIN identity_values iv
              ON iv.kind = il.kind AND iv.value = il.value
            WHERE il.kind IN ('email', 'phone')
            ORDER BY il.kind, il.value, il.candidate_key
            """
        ):
            kind = clean_text(row["kind"])
            value = clean_text(row["value"])
            candidate_key = clean_text(row["candidate_key"])
            tallanto_id = candidate_to_tallanto.get(candidate_key, "")
            if not tallanto_id:
                continue
            item = values[kind].setdefault(
                value,
                {
                    "kind": kind,
                    "value": value,
                    "candidate_keys": set(),
                    "tallanto_ids": set(),
                    "brand_scope": set(),
                    "match_class": clean_text(row["match_class"]),
                    "candidate_count": int(row["candidate_count"]),
                    "common_value": clean_text(row["match_class"]) != "strong_unique",
                },
            )
            item["candidate_keys"].add(candidate_key)
            item["tallanto_ids"].add(tallanto_id)
            item["brand_scope"].update(clients[tallanto_id]["brand_scope"])
            item["candidate_count"] = max(int(item["candidate_count"]), int(row["candidate_count"]))
            if clean_text(row["match_class"]) != "strong_unique":
                item["common_value"] = True
            target = "emails" if kind == "email" else "phones"
            common_target = "common_emails" if kind == "email" else "common_phones"
            clients[tallanto_id][target].add(value)
            if item["common_value"]:
                clients[tallanto_id][common_target].add(value)
        duplicate_tallanto_ids = sum(1 for client in clients.values() if client["duplicate_tallanto_id"])
    return {
        "clients": clients,
        "values": values,
        "duplicate_tallanto_ids": duplicate_tallanto_ids,
    }


def load_mail_classification_index(path: Optional[Path]) -> dict[str, dict[str, str]]:
    if path is None or not path.exists():
        return {}
    result: dict[str, dict[str, str]] = {}
    with Path(path).open("r", encoding="utf-8", newline="", errors="replace") as fh:
        for row in csv.DictReader(fh):
            message_sha = clean_text(row.get("message_sha256"))
            if not message_sha:
                continue
            result[message_sha] = {
                "brand": clean_text(row.get("brand")),
                "brand_source": clean_text(row.get("brand_source")),
                "klass": clean_text(row.get("klass")),
            }
    return result


def load_manual_mail_handoff_rows(path: Path) -> tuple[dict[int, str], list[sqlite3.Row], dict[str, int]]:
    with open_sqlite_readonly(path) as con:
        sources = {
            int(row["source_archive_id"]): clean_text(row["archive_db_path"])
            for row in con.execute("SELECT source_archive_id, archive_db_path FROM source_archives")
        }
        rows = list(
            con.execute(
                """
                SELECT *
                FROM mail_customer_links
                WHERE match_class IN ('ambiguous', 'missing')
                ORDER BY message_date_iso, message_sha256
                """
            )
        )
        baseline_counts = {
            clean_text(row["link_status"]): int(row["count"])
            for row in con.execute(
                "SELECT link_status, COUNT(*) AS count FROM mail_customer_links GROUP BY link_status"
            )
        }
    return sources, rows, baseline_counts


def read_archive_message_for_relink(
    archive_path: Path,
    *,
    message_sha256: str,
    cache: dict[str, sqlite3.Connection],
) -> tuple[sqlite3.Row | None, list[sqlite3.Row]]:
    key = str(Path(archive_path).resolve(strict=False))
    con = cache.get(key)
    if con is None:
        con = open_sqlite_readonly(Path(archive_path))
        cache[key] = con
    message = con.execute("SELECT * FROM messages WHERE sha256 = ?", (message_sha256,)).fetchone()
    participants = list(
        con.execute(
            "SELECT * FROM message_participants WHERE message_sha256 = ?",
            (message_sha256,),
        )
    )
    return message, participants


def decide_mail_customer_relink(
    *,
    row: Mapping[str, Any],
    archive_path: Path,
    message: sqlite3.Row | None,
    participants: Sequence[Mapping[str, Any]],
    address_book: Mapping[str, Any],
    classification: Mapping[str, str],
    internal_domains: Sequence[str],
    live_phone_lookup: Optional[Any],
    live_counts: Counter[str],
    require_real_correspondence: bool,
    max_text_chars_per_message: int,
) -> dict[str, Any]:
    message_sha = clean_text(row["message_sha256"])
    brand_signal = normalize_brand_token(classification.get("brand"))
    brand_source = clean_text(classification.get("brand_source"))
    klass = clean_text(classification.get("klass"))
    base = {
        "source_archive_id": int(row["source_archive_id"]),
        "source_archive_path": str(archive_path),
        "message_sha256": message_sha,
        "message_date_iso": clean_text(row["message_date_iso"]),
        "original_match_class": clean_text(row["match_class"]),
        "original_candidate_keys_json": clean_text(row["candidate_keys_json"]) or "[]",
        "decision": "unmatched",
        "reason": "",
        "tallanto_id": "",
        "candidate_keys": [],
        "signal_kind": "",
        "signal_value_sha256": "",
        "phones_found_count": 0,
        "external_email_count": 0,
        "brand_signal": brand_signal or "unknown",
        "brand_source": brand_source,
        "message_classification": klass,
        "learned_values": [],
        "evaluated_at": utc_now(),
    }
    if require_real_correspondence and klass and klass != "real_correspondence":
        return {**base, "reason": f"classification_{klass}"}
    if message is None:
        return {**base, "reason": "message_not_found"}

    text, text_status = read_safe_extracted_text(
        message["extracted_text_path"],
        archive_db_path=archive_path,
        max_chars=max_text_chars_per_message,
    )
    if text_status != "read":
        return {**base, "reason": f"text_{text_status}"}
    phones = extract_phone_numbers(text)
    external_emails = external_participant_emails(participants, internal_domains=internal_domains)
    base["phones_found_count"] = len(phones)
    base["external_email_count"] = len(external_emails)
    if not phones:
        return {**base, "reason": "no_phone_signal"}

    linked: list[dict[str, Any]] = []
    blocked: list[str] = []
    for phone in phones:
        resolved = resolve_address_book_value(
            address_book,
            kind="phone",
            value=phone,
            message_brand=brand_signal,
        )
        if resolved["status"] == "linked":
            linked.append({**resolved, "signal_kind": "phone", "signal_value": phone})
            continue
        if resolved["reason"] == "identity_value_missing" and live_phone_lookup is not None:
            live = resolve_live_tallanto_phone(
                phone,
                live_phone_lookup=live_phone_lookup,
                message_brand=brand_signal,
                live_counts=live_counts,
            )
            if live["status"] == "linked":
                linked.append({**live, "signal_kind": "live_phone", "signal_value": phone})
                continue
            blocked.append(live["reason"])
            continue
        blocked.append(resolved["reason"])

    target_ids = sorted({item["tallanto_id"] for item in linked if item.get("tallanto_id")})
    if len(target_ids) != 1:
        reason = "multiple_identity_targets" if len(target_ids) > 1 else most_common_reason(blocked) or "phone_no_identity_match"
        return {**base, "reason": reason}

    selected = next(item for item in linked if item["tallanto_id"] == target_ids[0])
    learned = learned_values_for_message(
        address_book,
        tallanto_id=selected["tallanto_id"],
        message_sha256=message_sha,
        signal_kind=selected["signal_kind"],
        external_emails=external_emails,
        phones=phones if selected["signal_kind"] == "live_phone" else (),
    )
    return {
        **base,
        "decision": "linked",
        "reason": "linked",
        "tallanto_id": selected["tallanto_id"],
        "candidate_keys": sorted(selected.get("candidate_keys") or []),
        "signal_kind": selected["signal_kind"],
        "signal_value_sha256": stable_value_hash(selected["signal_value"]),
        "learned_values": learned,
    }


def resolve_address_book_value(
    address_book: Mapping[str, Any],
    *,
    kind: str,
    value: str,
    message_brand: str,
) -> dict[str, Any]:
    item = address_book["values"].get(kind, {}).get(value)
    if not item:
        return {"status": "blocked", "reason": "identity_value_missing"}
    if len(item["brand_scope"]) > 1:
        return {"status": "blocked", "reason": "brand_conflict"}
    if item["common_value"] or item["match_class"] != "strong_unique":
        return {"status": "blocked", "reason": "duplicate_identity_value"}
    tallanto_ids = sorted(item["tallanto_ids"])
    if len(tallanto_ids) != 1:
        return {"status": "blocked", "reason": "ambiguous_tallanto_id"}
    tallanto_id = tallanto_ids[0]
    client = address_book["clients"].get(tallanto_id)
    if not client:
        return {"status": "blocked", "reason": "missing_address_book_client"}
    if client["duplicate_tallanto_id"]:
        return {"status": "blocked", "reason": "duplicate_tallanto_id"}
    if not brand_is_compatible(client["brand_scope"], message_brand):
        return {"status": "blocked", "reason": "brand_conflict"}
    return {
        "status": "linked",
        "reason": "linked",
        "tallanto_id": tallanto_id,
        "candidate_keys": sorted(item["candidate_keys"]),
    }


def resolve_live_tallanto_phone(
    phone: str,
    *,
    live_phone_lookup: Any,
    message_brand: str,
    live_counts: Counter[str],
) -> dict[str, Any]:
    live_counts["queries"] += 1
    try:
        records = list(live_phone_lookup(phone) or [])
    except Exception as exc:  # noqa: BLE001
        live_counts["errors"] += 1
        return {"status": "blocked", "reason": f"live_lookup_error:{type(exc).__name__}"}
    if not records:
        live_counts["empty"] += 1
        return {"status": "blocked", "reason": "live_no_contact"}
    ids = {
        clean_text(record.get("tallanto_id") or record.get("id") or record.get("contact_id"))
        for record in records
        if clean_text(record.get("tallanto_id") or record.get("id") or record.get("contact_id"))
    }
    if len(ids) != 1:
        live_counts["ambiguous"] += 1
        return {"status": "blocked", "reason": "live_multiple_contacts"}
    brand_scope = set()
    for record in records:
        brand_scope.update(infer_brand_scope(record, record))
    if not brand_is_compatible(brand_scope, message_brand):
        live_counts["brand_conflict"] += 1
        return {"status": "blocked", "reason": "live_brand_conflict"}
    live_counts["linked"] += 1
    return {
        "status": "linked",
        "reason": "linked",
        "tallanto_id": sorted(ids)[0],
        "candidate_keys": [],
    }


def learned_values_for_message(
    address_book: Mapping[str, Any],
    *,
    tallanto_id: str,
    message_sha256: str,
    signal_kind: str,
    external_emails: Sequence[str],
    phones: Sequence[str],
) -> list[dict[str, str]]:
    learned: list[dict[str, str]] = []
    for kind, values in (("email", external_emails), ("phone", phones)):
        for value in values:
            current = address_book["values"].get(kind, {}).get(value)
            if current:
                current_ids = set(current["tallanto_ids"])
                if current_ids != {tallanto_id} or current["common_value"]:
                    continue
                continue
            key = hashlib.sha256(
                f"{tallanto_id}|{kind}|{value}|{message_sha256}".encode("utf-8")
            ).hexdigest()
            learned.append(
                {
                    "idempotency_key": key,
                    "tallanto_id": tallanto_id,
                    "kind": kind,
                    "value": value,
                    "message_sha256": message_sha256,
                    "signal_kind": signal_kind,
                    "source": "mail_customer_relink_preview",
                    "created_at": utc_now(),
                }
            )
    return learned


def external_participant_emails(
    participants: Sequence[Mapping[str, Any]],
    *,
    internal_domains: Sequence[str],
) -> list[str]:
    domains = normalize_domains(internal_domains, "")
    result: list[str] = []
    for row in participants:
        value = normalize_email(
            mapping_get(row, "email_normalized")
            or mapping_get(row, "email")
            or mapping_get(row, "address")
        )
        if not value or email_domain_is_internal(value, domains):
            continue
        if value not in result:
            result.append(value)
    return result


def parse_json_object(value: object) -> dict[str, Any]:
    try:
        decoded = json.loads(clean_text(value) or "{}")
    except json.JSONDecodeError:
        return {}
    return decoded if isinstance(decoded, dict) else {}


def mapping_get(source: Mapping[str, Any], key: str, default: Any = "") -> Any:
    if hasattr(source, "get"):
        return source.get(key, default)
    if isinstance(source, sqlite3.Row) and key in source.keys():
        return source[key]
    return default


def infer_brand_scope(*sources: Mapping[str, Any]) -> set[str]:
    def iter_items(source: Mapping[str, Any]) -> Sequence[tuple[str, Any]]:
        if hasattr(source, "items"):
            return tuple(source.items())
        if isinstance(source, sqlite3.Row):
            return tuple((key, source[key]) for key in source.keys())
        return ()

    text = " ".join(
        clean_text(value)
        for source in sources
        for key, value in iter_items(source)
        if key and value is not None
    ).casefold()
    brands: set[str] = set()
    if re.search(r"\b(?:foton|фотон|цдпо)\b", text):
        brands.add("foton")
    if re.search(r"\b(?:unpk|унпк|мфти|kmipt)\b", text):
        brands.add("unpk")
    return brands


def normalize_brand_token(value: object) -> str:
    text = clean_text(value).casefold()
    if text in {"foton", "фотон"}:
        return "foton"
    if text in {"unpk", "унпк"}:
        return "unpk"
    return "unknown" if not text or text == "unknown" else text


def brand_is_compatible(client_brand_scope: set[str], message_brand: str) -> bool:
    brand = normalize_brand_token(message_brand)
    if not brand or brand == "unknown":
        return True
    if not client_brand_scope:
        return True
    return brand in client_brand_scope


def stable_value_hash(value: object) -> str:
    return hashlib.sha256(clean_text(value).encode("utf-8")).hexdigest()


def most_common_reason(reasons: Sequence[str]) -> str:
    if not reasons:
        return ""
    return Counter(reason for reason in reasons if reason).most_common(1)[0][0]


def build_relink_review_examples(decisions: Sequence[Mapping[str, Any]], *, limit: int) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    linked = [row for row in decisions if row.get("decision") == "linked"]
    unmatched = [row for row in decisions if row.get("decision") != "linked"]
    for row in (linked[: max(0, limit // 2)] + unmatched[: max(0, limit - len(linked[: max(0, limit // 2)]))]):
        result.append(
            {
                "message_sha256": clean_text(row.get("message_sha256"))[:16],
                "date": clean_text(row.get("message_date_iso")),
                "decision": clean_text(row.get("decision")),
                "reason": clean_text(row.get("reason")),
                "tallanto_id_hash": stable_value_hash(row.get("tallanto_id"))[:16] if row.get("tallanto_id") else "",
                "signal_kind": clean_text(row.get("signal_kind")),
                "signal_value_sha256": clean_text(row.get("signal_value_sha256"))[:16],
                "brand_signal": clean_text(row.get("brand_signal")),
                "message_classification": clean_text(row.get("message_classification")),
                "source_archive_path": clean_text(row.get("source_archive_path")),
            }
        )
        if len(result) >= limit:
            break
    return result


def build_mail_attachment_parse_plan(config: MailAttachmentParsePlanConfig) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.out_dir, "mail attachment parse plan output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment parse plan output directory")
    guard_git_ignored_output(out_dir, "mail attachment parse plan output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_attachment_parse_plan.sqlite"
    report_path = out_dir / "mail_attachment_parse_plan_report.json"

    allow_extensions = normalize_attachment_extensions(config.allow_extensions)
    review_extensions = normalize_attachment_extensions(config.review_extensions)
    max_size_bytes = max(0, int(config.max_size_bytes))
    rows = load_attachment_parse_plan_source_rows(source_paths)

    action_counts: Counter[str] = Counter()
    risk_level_counts: Counter[str] = Counter()
    risk_reason_counts: Counter[str] = Counter()
    extension_counts: Counter[str] = Counter()
    content_type_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    total_size_bytes = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attachment_parse_plan (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              filename_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              risk_level TEXT NOT NULL,
              action TEXT NOT NULL,
              risk_reasons_json TEXT NOT NULL,
              planned_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_parse_queue AS
            SELECT * FROM attachment_parse_plan
            WHERE action = 'parse_later';
            CREATE VIEW IF NOT EXISTS v_attachment_manual_review AS
            SELECT * FROM attachment_parse_plan
            WHERE action = 'manual_review';
            CREATE VIEW IF NOT EXISTS v_attachment_blocked AS
            SELECT * FROM attachment_parse_plan
            WHERE action = 'blocked';
            DELETE FROM meta;
            DELETE FROM attachment_parse_plan;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("source_archive_count", str(len(source_paths))),
                ("max_size_bytes", str(max_size_bytes)),
            ],
        )

        for row in rows:
            classification = classify_attachment_for_parse_plan(
                filename=row["filename"],
                declared_content_type=row["content_type"],
                size_bytes=int(row["size_bytes"]),
                max_size_bytes=max_size_bytes,
                allow_extensions=allow_extensions,
                review_extensions=review_extensions,
            )
            action_counts[classification["action"]] += 1
            risk_level_counts[classification["risk_level"]] += 1
            source_counts[row["source_archive_id"]] += 1
            extension_counts[classification["extension"]] += 1
            content_type_counts[classification["declared_content_type"]] += 1
            total_size_bytes += int(row["size_bytes"])
            for reason in classification["risk_reasons"]:
                risk_reason_counts[reason] += 1

            out.execute(
                """
                INSERT OR REPLACE INTO attachment_parse_plan (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  filename_sha256, extension, declared_content_type, size_bytes,
                  risk_level, action, risk_reasons_json, planned_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["source_archive_id"],
                    row["message_sha256"],
                    int(row["part_index"]),
                    row["attachment_sha256"],
                    row["filename_sha256"],
                    classification["extension"],
                    classification["declared_content_type"],
                    int(row["size_bytes"]),
                    classification["risk_level"],
                    classification["action"],
                    json.dumps(classification["risk_reasons"], ensure_ascii=False),
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    total = len(rows)
    report = {
        "schema_version": MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "attachment_count": total,
        "parse_plan_counts": {
            "parse_later": int(action_counts.get("parse_later", 0)),
            "manual_review": int(action_counts.get("manual_review", 0)),
            "blocked": int(action_counts.get("blocked", 0)),
        },
        "risk_level_counts": {
            "allow": int(risk_level_counts.get("allow", 0)),
            "review": int(risk_level_counts.get("review", 0)),
            "block": int(risk_level_counts.get("block", 0)),
        },
        "risk_reason_counts": dict(sorted(risk_reason_counts.items())),
        "artifact_counts": {
            "plan_rows_written": total,
            "distinct_source_archives": len(source_counts),
            "total_size_bytes": total_size_bytes,
            "max_size_bytes": max_size_bytes,
        },
        "top_extensions": extension_counts.most_common(30),
        "top_declared_content_types": content_type_counts.most_common(30),
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "source_db_attached_to_writer": False,
            "read_only_source_archives": True,
            "open_attachments": False,
            "read_attachment_bytes": False,
            "read_raw_eml": False,
            "execute_files": False,
            "extract_archives": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
            "store_raw_files_in_sqlite": False,
        },
        "privacy": {
            "contains_raw_personal_values": False,
            "raw_filenames_written": False,
            "filename_hashes_written": True,
            "raw_paths_written": False,
            "raw_mail_written": False,
            "raw_text_written": False,
            "raw_attachment_content_written": False,
            "attachment_sha256_written": True,
            "message_sha256_written": True,
        },
        "paths": {
            "plan_db": str(out_db_path),
            "report": str(report_path),
            "archive_db_count": len(source_paths),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_text_extract(config: MailAttachmentTextExtractConfig) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.parse_plan_db_path, "mail attachment parse plan database")
    guard_not_stable_runtime(config.out_dir, "mail attachment text extract output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment text extract output directory")
    guard_git_ignored_output(out_dir, "mail attachment text extract output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")
    parse_plan_db_path = config.parse_plan_db_path.resolve(strict=False)
    if not parse_plan_db_path.exists():
        raise FileNotFoundError(f"mail attachment parse plan database not found: {parse_plan_db_path}")

    stage_extensions = normalize_attachment_extensions(config.stage_extensions)
    if not stage_extensions:
        raise ValueError("at least one stage attachment extension is required")
    unsupported_extensions = sorted(stage_extensions - ATTACHMENT_TEXT_EXTRACT_SUPPORTED_EXTENSIONS)
    if unsupported_extensions:
        raise ValueError(
            "attachment text extract stage supports only "
            f"{sorted(ATTACHMENT_TEXT_EXTRACT_SUPPORTED_EXTENSIONS)}; got {unsupported_extensions}"
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    derived_text_dir = out_dir / "attachment_text"
    out_db_path = out_dir / "mail_attachment_text_extract.sqlite"
    report_path = out_dir / "mail_attachment_text_extract_report.json"

    archive_rows = load_attachment_text_extract_source_rows(source_paths)
    archive_row_by_key = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        ): row
        for row in archive_rows
    }
    archive_db_path_by_source_id = {mail_archive_source_id(path): path for path in source_paths}
    plan_rows, plan_total_parse_later, plan_stage_skipped = load_attachment_text_extract_plan_rows(
        parse_plan_db_path=parse_plan_db_path,
        source_archive_ids=set(archive_db_path_by_source_id),
        stage_extensions=stage_extensions,
        max_attachments=max(0, int(config.max_attachments)),
    )

    status_counts: Counter[str] = Counter()
    status_reason_counts: Counter[str] = Counter()
    extension_counts: Counter[str] = Counter()
    parser_counts: Counter[str] = Counter()
    warning_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    extracted_text_chars = 0
    extracted_text_files_written = 0
    attachment_bytes_read = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attachment_text_extracts (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              parser TEXT NOT NULL,
              status TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              warnings_json TEXT NOT NULL,
              text_sha256 TEXT NOT NULL,
              text_chars INTEGER NOT NULL,
              text_truncated INTEGER NOT NULL,
              derived_text_path TEXT NOT NULL,
              extracted_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_text_ready AS
            SELECT * FROM attachment_text_extracts
            WHERE status = 'extracted';
            CREATE VIEW IF NOT EXISTS v_attachment_text_needs_review AS
            SELECT * FROM attachment_text_extracts
            WHERE status != 'extracted';
            DELETE FROM meta;
            DELETE FROM attachment_text_extracts;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_TEXT_EXTRACT_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("source_archive_count", str(len(source_paths))),
                ("parse_plan_db_path", str(parse_plan_db_path)),
                ("stage_extensions", json.dumps(sorted(stage_extensions), ensure_ascii=False)),
                ("max_attachment_bytes", str(max(0, int(config.max_attachment_bytes)))),
                (
                    "max_text_chars_per_attachment",
                    str(max(0, int(config.max_text_chars_per_attachment))),
                ),
            ],
        )

        for plan_row in plan_rows:
            key = (
                plan_row["source_archive_id"],
                plan_row["message_sha256"],
                int(plan_row["part_index"]),
                plan_row["attachment_sha256"],
            )
            source_counts[plan_row["source_archive_id"]] += 1
            extension_counts[plan_row["extension"]] += 1
            archive_row = archive_row_by_key.get(key)
            result = extract_one_attachment_text(
                plan_row=plan_row,
                archive_row=archive_row,
                archive_db_path_by_source_id=archive_db_path_by_source_id,
                derived_text_dir=derived_text_dir,
                config=config,
            )
            status_counts[result["status"]] += 1
            status_reason_counts[result["status_reason"]] += 1
            if result["parser"]:
                parser_counts[result["parser"]] += 1
            for warning in result["warnings"]:
                warning_counts[warning] += 1
            if result["status"] == "extracted":
                extracted_text_chars += int(result["text_chars"])
                if result["derived_text_written"]:
                    extracted_text_files_written += 1
            attachment_bytes_read += int(result["bytes_read"])

            out.execute(
                """
                INSERT OR REPLACE INTO attachment_text_extracts (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, parser, status,
                  status_reason, warnings_json, text_sha256, text_chars,
                  text_truncated, derived_text_path, extracted_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    plan_row["source_archive_id"],
                    plan_row["message_sha256"],
                    int(plan_row["part_index"]),
                    plan_row["attachment_sha256"],
                    result["extension"],
                    result["declared_content_type"],
                    int(result["size_bytes"]),
                    result["parser"],
                    result["status"],
                    result["status_reason"],
                    json.dumps(result["warnings"], ensure_ascii=False),
                    result["text_sha256"],
                    int(result["text_chars"]),
                    1 if result["text_truncated"] else 0,
                    result["derived_text_path"],
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MAIL_ATTACHMENT_TEXT_EXTRACT_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "parse_plan_queue_count": plan_total_parse_later,
        "stage_supported_queue_count": len(plan_rows),
        "stage_skipped_queue_count": plan_stage_skipped,
        "status_counts": {
            "extracted": int(status_counts.get("extracted", 0)),
            "empty_text": int(status_counts.get("empty_text", 0)),
            "skipped": int(status_counts.get("skipped", 0)),
            "blocked_safety": int(status_counts.get("blocked_safety", 0)),
            "parse_error": int(status_counts.get("parse_error", 0)),
        },
        "status_reason_counts": dict(sorted(status_reason_counts.items())),
        "parser_counts": dict(sorted(parser_counts.items())),
        "warning_counts": dict(sorted(warning_counts.items())),
        "top_stage_extensions": extension_counts.most_common(30),
        "artifact_counts": {
            "rows_written": len(plan_rows),
            "derived_text_files_written": extracted_text_files_written,
            "extracted_text_chars": extracted_text_chars,
            "attachment_bytes_read": attachment_bytes_read,
            "max_attachment_bytes": max(0, int(config.max_attachment_bytes)),
            "max_text_chars_per_attachment": max(0, int(config.max_text_chars_per_attachment)),
            "distinct_source_archives": len(source_counts),
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "parse_plan_sqlite_mode": "mode=ro",
            "parse_plan_sqlite_query_only": True,
            "source_db_attached_to_writer": False,
            "read_raw_eml": False,
            "read_extracted_mail_text": False,
            "read_attachment_bytes": True,
            "read_only_parse_later_queue": True,
            "allowed_extensions_only": sorted(stage_extensions),
            "parse_pdf": False,
            "parse_images": False,
            "run_ocr": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "derived_text_files_may_contain_personal_data": True,
            "derived_text_paths_written": True,
            "attachment_sha256_written": True,
            "message_sha256_written": True,
        },
        "paths": {
            "extract_db": str(out_db_path),
            "report": str(report_path),
            "derived_text_dir": str(derived_text_dir),
            "parse_plan_db": str(parse_plan_db_path),
            "archive_db_count": len(source_paths),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_pdf_extract(config: MailAttachmentPdfExtractConfig) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.parse_plan_db_path, "mail attachment parse plan database")
    guard_not_stable_runtime(config.out_dir, "mail attachment PDF extract output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment PDF extract output directory")
    guard_git_ignored_output(out_dir, "mail attachment PDF extract output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")
    parse_plan_db_path = config.parse_plan_db_path.resolve(strict=False)
    if not parse_plan_db_path.exists():
        raise FileNotFoundError(f"mail attachment parse plan database not found: {parse_plan_db_path}")

    stage_extensions = normalize_attachment_extensions(config.stage_extensions)
    if not stage_extensions:
        raise ValueError("at least one stage attachment extension is required")
    unsupported_extensions = sorted(stage_extensions - ATTACHMENT_PDF_EXTRACT_SUPPORTED_EXTENSIONS)
    if unsupported_extensions:
        raise ValueError(
            "attachment PDF extract stage supports only "
            f"{sorted(ATTACHMENT_PDF_EXTRACT_SUPPORTED_EXTENSIONS)}; got {unsupported_extensions}"
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    derived_text_dir = out_dir / "attachment_pdf_text"
    out_db_path = out_dir / "mail_attachment_pdf_extract.sqlite"
    report_path = out_dir / "mail_attachment_pdf_extract_report.json"

    archive_rows = load_attachment_text_extract_source_rows(source_paths)
    archive_row_by_key = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        ): row
        for row in archive_rows
    }
    archive_db_path_by_source_id = {mail_archive_source_id(path): path for path in source_paths}
    plan_rows, plan_total_parse_later, plan_stage_skipped = load_attachment_text_extract_plan_rows(
        parse_plan_db_path=parse_plan_db_path,
        source_archive_ids=set(archive_db_path_by_source_id),
        stage_extensions=stage_extensions,
        max_attachments=max(0, int(config.max_attachments)),
    )

    status_counts: Counter[str] = Counter()
    status_reason_counts: Counter[str] = Counter()
    extension_counts: Counter[str] = Counter()
    parser_counts: Counter[str] = Counter()
    warning_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    extracted_text_chars = 0
    extracted_text_files_written = 0
    attachment_bytes_read = 0
    pdf_pages_total = 0
    pdf_pages_processed = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attachment_pdf_extracts (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              parser TEXT NOT NULL,
              status TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              warnings_json TEXT NOT NULL,
              page_count INTEGER NOT NULL,
              pages_processed INTEGER NOT NULL,
              text_sha256 TEXT NOT NULL,
              text_chars INTEGER NOT NULL,
              text_truncated INTEGER NOT NULL,
              derived_text_path TEXT NOT NULL,
              extracted_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_pdf_text_ready AS
            SELECT * FROM attachment_pdf_extracts
            WHERE status = 'extracted';
            CREATE VIEW IF NOT EXISTS v_attachment_pdf_text_needs_review AS
            SELECT * FROM attachment_pdf_extracts
            WHERE status != 'extracted';
            DELETE FROM meta;
            DELETE FROM attachment_pdf_extracts;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_PDF_EXTRACT_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("source_archive_count", str(len(source_paths))),
                ("parse_plan_db_path", str(parse_plan_db_path)),
                ("stage_extensions", json.dumps(sorted(stage_extensions), ensure_ascii=False)),
                ("max_attachment_bytes", str(max(0, int(config.max_attachment_bytes)))),
                ("max_pdf_pages", str(max(0, int(config.max_pdf_pages)))),
                ("pdf_timeout_seconds", str(max(0, int(config.pdf_timeout_seconds)))),
                (
                    "max_text_chars_per_attachment",
                    str(max(0, int(config.max_text_chars_per_attachment))),
                ),
            ],
        )

        for plan_row in plan_rows:
            key = (
                plan_row["source_archive_id"],
                plan_row["message_sha256"],
                int(plan_row["part_index"]),
                plan_row["attachment_sha256"],
            )
            source_counts[plan_row["source_archive_id"]] += 1
            extension_counts[plan_row["extension"]] += 1
            archive_row = archive_row_by_key.get(key)
            result = extract_one_attachment_pdf_text(
                plan_row=plan_row,
                archive_row=archive_row,
                archive_db_path_by_source_id=archive_db_path_by_source_id,
                derived_text_dir=derived_text_dir,
                config=config,
            )
            status_counts[result["status"]] += 1
            status_reason_counts[result["status_reason"]] += 1
            if result["parser"]:
                parser_counts[result["parser"]] += 1
            for warning in result["warnings"]:
                warning_counts[warning] += 1
            if result["status"] == "extracted":
                extracted_text_chars += int(result["text_chars"])
                if result["derived_text_written"]:
                    extracted_text_files_written += 1
            attachment_bytes_read += int(result["bytes_read"])
            pdf_pages_total += int(result["page_count"])
            pdf_pages_processed += int(result["pages_processed"])

            out.execute(
                """
                INSERT OR REPLACE INTO attachment_pdf_extracts (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, parser, status,
                  status_reason, warnings_json, page_count, pages_processed,
                  text_sha256, text_chars, text_truncated, derived_text_path,
                  extracted_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    plan_row["source_archive_id"],
                    plan_row["message_sha256"],
                    int(plan_row["part_index"]),
                    plan_row["attachment_sha256"],
                    result["extension"],
                    result["declared_content_type"],
                    int(result["size_bytes"]),
                    result["parser"],
                    result["status"],
                    result["status_reason"],
                    json.dumps(result["warnings"], ensure_ascii=False),
                    int(result["page_count"]),
                    int(result["pages_processed"]),
                    result["text_sha256"],
                    int(result["text_chars"]),
                    1 if result["text_truncated"] else 0,
                    result["derived_text_path"],
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MAIL_ATTACHMENT_PDF_EXTRACT_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "parse_plan_queue_count": plan_total_parse_later,
        "stage_supported_queue_count": len(plan_rows),
        "stage_skipped_queue_count": plan_stage_skipped,
        "status_counts": {
            "extracted": int(status_counts.get("extracted", 0)),
            "empty_text": int(status_counts.get("empty_text", 0)),
            "skipped": int(status_counts.get("skipped", 0)),
            "blocked_safety": int(status_counts.get("blocked_safety", 0)),
            "parse_error": int(status_counts.get("parse_error", 0)),
        },
        "status_reason_counts": dict(sorted(status_reason_counts.items())),
        "parser_counts": dict(sorted(parser_counts.items())),
        "warning_counts": dict(sorted(warning_counts.items())),
        "top_stage_extensions": extension_counts.most_common(30),
        "artifact_counts": {
            "rows_written": len(plan_rows),
            "derived_text_files_written": extracted_text_files_written,
            "extracted_text_chars": extracted_text_chars,
            "attachment_bytes_read": attachment_bytes_read,
            "pdf_pages_total": pdf_pages_total,
            "pdf_pages_processed": pdf_pages_processed,
            "max_attachment_bytes": max(0, int(config.max_attachment_bytes)),
            "max_pdf_pages": max(0, int(config.max_pdf_pages)),
            "pdf_timeout_seconds": max(0, int(config.pdf_timeout_seconds)),
            "max_text_chars_per_attachment": max(0, int(config.max_text_chars_per_attachment)),
            "distinct_source_archives": len(source_counts),
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "parse_plan_sqlite_mode": "mode=ro",
            "parse_plan_sqlite_query_only": True,
            "source_db_attached_to_writer": False,
            "read_raw_eml": False,
            "read_extracted_mail_text": False,
            "read_attachment_bytes": True,
            "read_only_parse_later_queue": True,
            "allowed_extensions_only": sorted(stage_extensions),
            "parse_pdf": True,
            "render_pdf": False,
            "parse_images": False,
            "run_ocr": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "allow_pdf_actions": False,
            "allow_pdf_embedded_files": False,
            "allow_pdf_javascript": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "derived_text_files_may_contain_personal_data": True,
            "derived_text_paths_written": True,
            "attachment_sha256_written": True,
            "message_sha256_written": True,
        },
        "paths": {
            "extract_db": str(out_db_path),
            "report": str(report_path),
            "derived_text_dir": str(derived_text_dir),
            "parse_plan_db": str(parse_plan_db_path),
            "archive_db_count": len(source_paths),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_image_ocr_plan(
    config: MailAttachmentImageOcrPlanConfig,
) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.parse_plan_db_path, "mail attachment parse plan database")
    guard_not_stable_runtime(config.out_dir, "mail attachment image OCR plan output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment image OCR plan output directory")
    guard_git_ignored_output(out_dir, "mail attachment image OCR plan output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")
    parse_plan_db_path = config.parse_plan_db_path.resolve(strict=False)
    if not parse_plan_db_path.exists():
        raise FileNotFoundError(f"mail attachment parse plan database not found: {parse_plan_db_path}")

    stage_extensions = normalize_attachment_extensions(config.stage_extensions)
    if not stage_extensions:
        raise ValueError("at least one stage attachment extension is required")
    unsupported_extensions = sorted(stage_extensions - ATTACHMENT_IMAGE_OCR_SUPPORTED_EXTENSIONS)
    if unsupported_extensions:
        raise ValueError(
            "attachment image OCR plan stage supports only "
            f"{sorted(ATTACHMENT_IMAGE_OCR_SUPPORTED_EXTENSIONS)}; got {unsupported_extensions}"
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_attachment_image_ocr_plan.sqlite"
    report_path = out_dir / "mail_attachment_image_ocr_plan_report.json"

    archive_rows = load_attachment_text_extract_source_rows(source_paths)
    archive_row_by_key = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        ): row
        for row in archive_rows
    }
    archive_db_path_by_source_id = {mail_archive_source_id(path): path for path in source_paths}
    plan_rows, plan_total_parse_later, plan_stage_skipped = load_attachment_text_extract_plan_rows(
        parse_plan_db_path=parse_plan_db_path,
        source_archive_ids=set(archive_db_path_by_source_id),
        stage_extensions=stage_extensions,
        max_attachments=max(0, int(config.max_attachments)),
    )

    status_counts: Counter[str] = Counter()
    status_reason_counts: Counter[str] = Counter()
    extension_counts: Counter[str] = Counter()
    image_format_counts: Counter[str] = Counter()
    warning_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    attachment_bytes_read = 0
    total_image_pixels = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attachment_image_ocr_plan (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              image_format TEXT NOT NULL,
              width INTEGER NOT NULL,
              height INTEGER NOT NULL,
              pixel_count INTEGER NOT NULL,
              status TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              ocr_status TEXT NOT NULL,
              warnings_json TEXT NOT NULL,
              planned_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_image_ocr_candidates AS
            SELECT * FROM attachment_image_ocr_plan
            WHERE status = 'planned' AND ocr_status = 'disabled';
            CREATE VIEW IF NOT EXISTS v_attachment_image_ocr_needs_review AS
            SELECT * FROM attachment_image_ocr_plan
            WHERE status != 'planned';
            DELETE FROM meta;
            DELETE FROM attachment_image_ocr_plan;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_IMAGE_OCR_PLAN_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("source_archive_count", str(len(source_paths))),
                ("parse_plan_db_path", str(parse_plan_db_path)),
                ("stage_extensions", json.dumps(sorted(stage_extensions), ensure_ascii=False)),
                ("max_attachment_bytes", str(max(0, int(config.max_attachment_bytes)))),
                ("max_image_dimension", str(max(0, int(config.max_image_dimension)))),
                ("max_image_pixels", str(max(0, int(config.max_image_pixels)))),
                ("inspect_headers", "true" if config.inspect_headers else "false"),
                ("ocr_status", "disabled"),
            ],
        )

        for plan_row in plan_rows:
            key = (
                plan_row["source_archive_id"],
                plan_row["message_sha256"],
                int(plan_row["part_index"]),
                plan_row["attachment_sha256"],
            )
            source_counts[plan_row["source_archive_id"]] += 1
            extension_counts[plan_row["extension"]] += 1
            archive_row = archive_row_by_key.get(key)
            result = inspect_one_attachment_image_for_ocr_plan(
                plan_row=plan_row,
                archive_row=archive_row,
                archive_db_path_by_source_id=archive_db_path_by_source_id,
                config=config,
            )
            status_counts[result["status"]] += 1
            status_reason_counts[result["status_reason"]] += 1
            image_format_counts[result["image_format"]] += 1
            for warning in result["warnings"]:
                warning_counts[warning] += 1
            attachment_bytes_read += int(result["bytes_read"])
            total_image_pixels += int(result["pixel_count"])

            out.execute(
                """
                INSERT OR REPLACE INTO attachment_image_ocr_plan (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, image_format,
                  width, height, pixel_count, status, status_reason, ocr_status,
                  warnings_json, planned_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    plan_row["source_archive_id"],
                    plan_row["message_sha256"],
                    int(plan_row["part_index"]),
                    plan_row["attachment_sha256"],
                    result["extension"],
                    result["declared_content_type"],
                    int(result["size_bytes"]),
                    result["image_format"],
                    int(result["width"]),
                    int(result["height"]),
                    int(result["pixel_count"]),
                    result["status"],
                    result["status_reason"],
                    result["ocr_status"],
                    json.dumps(result["warnings"], ensure_ascii=False),
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MAIL_ATTACHMENT_IMAGE_OCR_PLAN_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "parse_plan_queue_count": plan_total_parse_later,
        "stage_supported_queue_count": len(plan_rows),
        "stage_skipped_queue_count": plan_stage_skipped,
        "status_counts": {
            "planned": int(status_counts.get("planned", 0)),
            "skipped": int(status_counts.get("skipped", 0)),
            "blocked_safety": int(status_counts.get("blocked_safety", 0)),
            "parse_error": int(status_counts.get("parse_error", 0)),
        },
        "status_reason_counts": dict(sorted(status_reason_counts.items())),
        "image_format_counts": dict(sorted(image_format_counts.items())),
        "warning_counts": dict(sorted(warning_counts.items())),
        "top_stage_extensions": extension_counts.most_common(30),
        "artifact_counts": {
            "rows_written": len(plan_rows),
            "attachment_bytes_read": attachment_bytes_read,
            "derived_text_files_written": 0,
            "derived_image_files_written": 0,
            "thumbnails_written": 0,
            "ocr_text_files_written": 0,
            "total_image_pixels": total_image_pixels,
            "max_attachment_bytes": max(0, int(config.max_attachment_bytes)),
            "max_image_dimension": max(0, int(config.max_image_dimension)),
            "max_image_pixels": max(0, int(config.max_image_pixels)),
            "distinct_source_archives": len(source_counts),
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "parse_plan_sqlite_mode": "mode=ro",
            "parse_plan_sqlite_query_only": True,
            "source_db_attached_to_writer": False,
            "read_raw_eml": False,
            "read_extracted_mail_text": False,
            "read_attachment_bytes": bool(config.inspect_headers),
            "read_only_parse_later_queue": True,
            "allowed_extensions_only": sorted(stage_extensions),
            "inspect_image_headers_only": bool(config.inspect_headers),
            "decode_full_images": False,
            "extract_exif": False,
            "write_thumbnails": False,
            "write_images": False,
            "parse_pdf": False,
            "render_pdf": False,
            "run_ocr": False,
            "ocr_enabled": False,
            "parse_images": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "raw_exif_written": False,
            "raw_ocr_text_written": False,
            "derived_text_files_may_contain_personal_data": False,
            "attachment_sha256_written": True,
            "message_sha256_written": True,
        },
        "paths": {
            "plan_db": str(out_db_path),
            "report": str(report_path),
            "parse_plan_db": str(parse_plan_db_path),
            "archive_db_count": len(source_paths),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_text_index(config: MailAttachmentTextIndexConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.out_dir, "mail attachment text index output directory")
    source_specs = [
        *[(Path(path), "text_extract") for path in config.text_extract_db_paths],
        *[(Path(path), "pdf_extract") for path in config.pdf_extract_db_paths],
        *[(Path(path), "image_ocr_plan") for path in config.image_ocr_plan_db_paths],
    ]
    if not source_specs:
        raise ValueError("at least one attachment extract or OCR-plan database is required")
    for source_path, source_stage in source_specs:
        guard_not_stable_runtime(source_path, f"mail attachment {source_stage} database")
    if config.parse_plan_db_path is not None:
        guard_not_stable_runtime(config.parse_plan_db_path, "mail attachment parse plan database")

    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment text index output directory")
    guard_git_ignored_output(out_dir, "mail attachment text index output directory")

    resolved_source_specs: list[tuple[Path, str]] = []
    for source_path, source_stage in source_specs:
        resolved_source_path = source_path.resolve(strict=False)
        if not resolved_source_path.exists():
            raise FileNotFoundError(f"mail attachment {source_stage} database not found: {resolved_source_path}")
        resolved_source_specs.append((resolved_source_path, source_stage))

    parse_plan_queue_count = 0
    parse_plan_db_provided = config.parse_plan_db_path is not None
    if config.parse_plan_db_path is not None:
        parse_plan_db_path = config.parse_plan_db_path.resolve(strict=False)
        if not parse_plan_db_path.exists():
            raise FileNotFoundError(f"mail attachment parse plan database not found: {parse_plan_db_path}")
        parse_plan_queue_count = load_attachment_parse_later_count(parse_plan_db_path)

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_attachment_text_index.sqlite"
    report_path = out_dir / "mail_attachment_text_index_report.json"

    source_rows: list[dict[str, Any]] = []
    source_summaries: dict[tuple[str, str], dict[str, Any]] = {}
    for source_path, source_stage in resolved_source_specs:
        rows, schema_version = load_attachment_text_index_rows(
            source_path,
            source_stage=source_stage,
        )
        source_db_id = mail_archive_source_id(source_path)
        for row in rows:
            row["source_db_id"] = source_db_id
        source_summaries[(source_db_id, source_stage)] = {
            "source_db_id": source_db_id,
            "source_stage": source_stage,
            "source_schema_version": schema_version,
            "row_count": len(rows),
        }
        source_rows.extend(rows)

    source_rows.sort(
        key=lambda row: (
            row["source_stage"],
            row["source_db_id"],
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        )
    )
    if int(config.max_rows) > 0:
        source_rows = source_rows[: int(config.max_rows)]

    source_stage_counts: Counter[str] = Counter()
    source_status_counts: Counter[str] = Counter()
    text_status_counts: Counter[str] = Counter()
    extension_counts: Counter[str] = Counter()
    parser_counts: Counter[str] = Counter()
    warning_counts: Counter[str] = Counter()
    attachment_key_counts: Counter[tuple[str, str, int, str]] = Counter()

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS stage5_sources (
              source_db_id TEXT NOT NULL,
              source_stage TEXT NOT NULL,
              source_schema_version TEXT NOT NULL,
              row_count INTEGER NOT NULL,
              indexed_at TEXT NOT NULL,
              PRIMARY KEY (source_db_id, source_stage)
            );
            CREATE TABLE IF NOT EXISTS attachment_text_index (
              source_db_id TEXT NOT NULL,
              source_stage TEXT NOT NULL,
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              parser TEXT NOT NULL,
              source_status TEXT NOT NULL,
              source_status_reason TEXT NOT NULL,
              text_status TEXT NOT NULL,
              text_sha256 TEXT NOT NULL,
              text_chars INTEGER NOT NULL,
              text_truncated INTEGER NOT NULL,
              derived_text_path_sha256 TEXT NOT NULL,
              page_count INTEGER NOT NULL,
              pages_processed INTEGER NOT NULL,
              warnings_json TEXT NOT NULL,
              indexed_at TEXT NOT NULL,
              PRIMARY KEY (
                source_db_id, source_stage, source_archive_id,
                message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_text_available AS
            SELECT * FROM attachment_text_index
            WHERE text_status = 'available';
            CREATE VIEW IF NOT EXISTS v_attachment_text_needs_review AS
            SELECT * FROM attachment_text_index
            WHERE text_status != 'available';
            CREATE VIEW IF NOT EXISTS v_attachment_ocr_pending AS
            SELECT * FROM attachment_text_index
            WHERE text_status = 'ocr_pending';
            DELETE FROM meta;
            DELETE FROM stage5_sources;
            DELETE FROM attachment_text_index;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_TEXT_INDEX_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("source_db_count", str(len(resolved_source_specs))),
                ("parse_plan_db_provided", "true" if parse_plan_db_provided else "false"),
                ("parse_plan_queue_count", str(parse_plan_queue_count)),
                ("max_rows", str(max(0, int(config.max_rows)))),
                ("read_derived_text_content", "false"),
            ],
        )
        for summary in source_summaries.values():
            out.execute(
                """
                INSERT OR REPLACE INTO stage5_sources (
                  source_db_id, source_stage, source_schema_version, row_count, indexed_at
                ) VALUES (?, ?, ?, ?, ?)
                """,
                (
                    summary["source_db_id"],
                    summary["source_stage"],
                    summary["source_schema_version"],
                    int(summary["row_count"]),
                    utc_now(),
                ),
            )

        for row in source_rows:
            source_stage = clean_text(row["source_stage"])
            source_status = clean_text(row["source_status"])
            text_status = attachment_text_index_status(source_stage, source_status, row)
            source_stage_counts[source_stage] += 1
            source_status_counts[f"{source_stage}:{source_status}"] += 1
            text_status_counts[text_status] += 1
            extension_counts[clean_text(row["extension"])] += 1
            if clean_text(row["parser"]):
                parser_counts[clean_text(row["parser"])] += 1
            warnings = parse_json_list(row.get("warnings_json", "[]"))
            for warning in warnings:
                warning_counts[warning] += 1
            attachment_key_counts[
                (
                    clean_text(row["source_archive_id"]),
                    clean_text(row["message_sha256"]),
                    int(row["part_index"]),
                    clean_text(row["attachment_sha256"]),
                )
            ] += 1
            derived_path = clean_text(row.get("derived_text_path", ""))
            derived_path_sha256 = (
                hashlib.sha256(derived_path.encode("utf-8")).hexdigest() if derived_path else ""
            )

            out.execute(
                """
                INSERT OR REPLACE INTO attachment_text_index (
                  source_db_id, source_stage, source_archive_id, message_sha256,
                  part_index, attachment_sha256, extension, declared_content_type,
                  size_bytes, parser, source_status, source_status_reason,
                  text_status, text_sha256, text_chars, text_truncated,
                  derived_text_path_sha256, page_count, pages_processed,
                  warnings_json, indexed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    clean_text(row["source_db_id"]),
                    source_stage,
                    clean_text(row["source_archive_id"]),
                    clean_text(row["message_sha256"]),
                    int(row["part_index"]),
                    clean_text(row["attachment_sha256"]),
                    clean_text(row["extension"]),
                    clean_text(row["declared_content_type"]),
                    int(row["size_bytes"]),
                    clean_text(row["parser"]),
                    source_status,
                    clean_text(row["source_status_reason"]),
                    text_status,
                    clean_text(row["text_sha256"]),
                    int(row["text_chars"]),
                    int(row["text_truncated"]),
                    derived_path_sha256,
                    int(row["page_count"]),
                    int(row["pages_processed"]),
                    json.dumps(warnings, ensure_ascii=False),
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    duplicate_attachment_key_count = sum(1 for count in attachment_key_counts.values() if count > 1)
    covered_row_count = len(source_rows)
    stage_gap_count = max(0, parse_plan_queue_count - covered_row_count) if parse_plan_queue_count else 0
    report = {
        "schema_version": MAIL_ATTACHMENT_TEXT_INDEX_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_db_count": len(resolved_source_specs),
        "source_text_extract_count": len(config.text_extract_db_paths),
        "source_pdf_extract_count": len(config.pdf_extract_db_paths),
        "source_image_ocr_plan_count": len(config.image_ocr_plan_db_paths),
        "parse_plan_queue_count": parse_plan_queue_count,
        "parse_plan_db_provided": parse_plan_db_provided,
        "source_stage_counts": dict(sorted(source_stage_counts.items())),
        "source_status_counts": dict(sorted(source_status_counts.items())),
        "text_status_counts": dict(sorted(text_status_counts.items())),
        "parser_counts": dict(sorted(parser_counts.items())),
        "warning_counts": dict(sorted(warning_counts.items())),
        "top_stage_extensions": extension_counts.most_common(30),
        "coverage_counts": {
            "covered_source_rows": covered_row_count,
            "available_text_rows": int(text_status_counts.get("available", 0)),
            "ocr_pending_rows": int(text_status_counts.get("ocr_pending", 0)),
            "needs_review_rows": covered_row_count - int(text_status_counts.get("available", 0)),
            "parse_later_rows_without_stage5_source": stage_gap_count,
            "duplicate_attachment_keys": duplicate_attachment_key_count,
        },
        "artifact_counts": {
            "rows_written": covered_row_count,
            "stage_source_rows_written": len(source_summaries),
            "derived_text_files_read": 0,
            "raw_text_chars_written": 0,
            "attachment_bytes_read": 0,
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "parse_plan_sqlite_mode": "mode=ro" if parse_plan_db_provided else "",
            "parse_plan_sqlite_query_only": bool(parse_plan_db_provided),
            "read_raw_eml": False,
            "read_attachment_bytes": False,
            "read_existing_derived_text": False,
            "read_derived_text_content": False,
            "parse_pdf": False,
            "render_pdf": False,
            "parse_images": False,
            "run_ocr": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "raw_text_written_to_sqlite": False,
            "raw_text_written_to_json": False,
            "derived_text_content_written_to_sqlite": False,
            "derived_text_content_written_to_json": False,
            "derived_text_paths_written": False,
            "derived_text_path_hashes_written": True,
            "message_sha256_written": True,
            "attachment_sha256_written": True,
        },
        "paths": {
            "index_db": str(out_db_path),
            "report": str(report_path),
            "source_db_count": len(resolved_source_specs),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_stage6_plan(config: MailAttachmentStage6PlanConfig) -> Mapping[str, Any]:
    guard_not_stable_runtime(config.parse_plan_db_path, "mail attachment parse plan database")
    guard_not_stable_runtime(config.text_index_db_path, "mail attachment text index database")
    guard_not_stable_runtime(config.out_dir, "mail attachment stage 6 output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment stage 6 output directory")
    guard_git_ignored_output(out_dir, "mail attachment stage 6 output directory")

    parse_plan_db_path = config.parse_plan_db_path.resolve(strict=False)
    text_index_db_path = config.text_index_db_path.resolve(strict=False)
    if not parse_plan_db_path.exists():
        raise FileNotFoundError(f"mail attachment parse plan database not found: {parse_plan_db_path}")
    if not text_index_db_path.exists():
        raise FileNotFoundError(f"mail attachment text index database not found: {text_index_db_path}")

    pilot_extensions = normalize_attachment_extensions(config.pilot_extensions)
    if not pilot_extensions:
        raise ValueError("at least one OCR pilot extension is required")
    unsupported_extensions = sorted(pilot_extensions - ATTACHMENT_IMAGE_OCR_SUPPORTED_EXTENSIONS)
    if unsupported_extensions:
        raise ValueError(
            "attachment stage 6 OCR pilot supports only "
            f"{sorted(ATTACHMENT_IMAGE_OCR_SUPPORTED_EXTENSIONS)}; got {unsupported_extensions}"
        )

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_attachment_stage6_plan.sqlite"
    report_path = out_dir / "mail_attachment_stage6_plan_report.json"

    parse_later_rows = load_attachment_stage6_parse_later_rows(parse_plan_db_path)
    text_index_rows = load_attachment_stage6_text_index_rows(text_index_db_path)
    text_index_keys = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        )
        for row in text_index_rows
    }
    gap_rows = [
        {
            **row,
            **classify_stage6_gap_row(row),
        }
        for row in parse_later_rows
        if (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        )
        not in text_index_keys
    ]

    ocr_rows = [
        row
        for row in text_index_rows
        if row["source_stage"] == "image_ocr_plan"
        and row["text_status"] == "ocr_pending"
        and row["extension"] in pilot_extensions
    ]
    min_pilot_attachment_bytes = max(0, int(config.min_pilot_attachment_bytes))
    max_pilot_attachment_bytes = max(0, int(config.max_pilot_attachment_bytes))
    if max_pilot_attachment_bytes and min_pilot_attachment_bytes >= max_pilot_attachment_bytes:
        raise ValueError("min_pilot_attachment_bytes must be lower than max_pilot_attachment_bytes")
    eligible_ocr_rows = [
        row
        for row in ocr_rows
        if min_pilot_attachment_bytes < int(row["size_bytes"]) <= max_pilot_attachment_bytes
    ]
    selected_ocr_rows = select_balanced_ocr_pilot_rows(
        eligible_ocr_rows,
        limit=max(0, int(config.ocr_pilot_limit)),
    )
    selected_keys = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        ): rank
        for rank, row in enumerate(selected_ocr_rows, start=1)
    }

    gap_extension_counts: Counter[str] = Counter(row["extension"] for row in gap_rows)
    gap_class_counts: Counter[str] = Counter(row["gap_class"] for row in gap_rows)
    gap_action_counts: Counter[str] = Counter(row["recommended_action"] for row in gap_rows)
    ocr_candidate_extension_counts: Counter[str] = Counter(row["extension"] for row in ocr_rows)
    selected_extension_counts: Counter[str] = Counter(row["extension"] for row in selected_ocr_rows)
    pilot_status_counts: Counter[str] = Counter()

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS stage6_gap_plan (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              gap_class TEXT NOT NULL,
              recommended_action TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              risk_reasons_json TEXT NOT NULL,
              planned_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE TABLE IF NOT EXISTS stage6_ocr_pilot_plan (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              pilot_status TEXT NOT NULL,
              pilot_rank INTEGER NOT NULL,
              pilot_batch_label TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              planned_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_stage6_gap_needs_decision AS
            SELECT * FROM stage6_gap_plan
            WHERE recommended_action != 'ignore';
            CREATE VIEW IF NOT EXISTS v_stage6_ocr_pilot_selected AS
            SELECT * FROM stage6_ocr_pilot_plan
            WHERE pilot_status = 'selected'
            ORDER BY pilot_rank;
            CREATE VIEW IF NOT EXISTS v_stage6_ocr_pilot_deferred AS
            SELECT * FROM stage6_ocr_pilot_plan
            WHERE pilot_status != 'selected';
            DELETE FROM meta;
            DELETE FROM stage6_gap_plan;
            DELETE FROM stage6_ocr_pilot_plan;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_STAGE6_PLAN_SCHEMA_VERSION),
                ("created_at", utc_now()),
                (
                    "parse_plan_db_path_sha256",
                    hashlib.sha256(str(parse_plan_db_path).encode("utf-8")).hexdigest(),
                ),
                (
                    "text_index_db_path_sha256",
                    hashlib.sha256(str(text_index_db_path).encode("utf-8")).hexdigest(),
                ),
                ("ocr_pilot_limit", str(max(0, int(config.ocr_pilot_limit)))),
                ("min_pilot_attachment_bytes", str(min_pilot_attachment_bytes)),
                ("max_pilot_attachment_bytes", str(max_pilot_attachment_bytes)),
                ("pilot_extensions", json.dumps(sorted(pilot_extensions), ensure_ascii=False)),
            ],
        )
        for row in gap_rows:
            out.execute(
                """
                INSERT OR REPLACE INTO stage6_gap_plan (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, gap_class,
                  recommended_action, status_reason, risk_reasons_json, planned_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["source_archive_id"],
                    row["message_sha256"],
                    int(row["part_index"]),
                    row["attachment_sha256"],
                    row["extension"],
                    row["declared_content_type"],
                    int(row["size_bytes"]),
                    row["gap_class"],
                    row["recommended_action"],
                    row["status_reason"],
                    row["risk_reasons_json"],
                    utc_now(),
                ),
            )
        for row in ocr_rows:
            key = (
                row["source_archive_id"],
                row["message_sha256"],
                int(row["part_index"]),
                row["attachment_sha256"],
            )
            rank = int(selected_keys.get(key, 0))
            if rank:
                pilot_status = "selected"
                status_reason = "balanced_extension_sample"
            elif int(row["size_bytes"]) <= 0:
                pilot_status = "excluded"
                status_reason = "empty_or_unknown_size"
            elif int(row["size_bytes"]) <= min_pilot_attachment_bytes:
                pilot_status = "excluded"
                status_reason = "pilot_min_size_not_reached"
            elif int(row["size_bytes"]) > max_pilot_attachment_bytes:
                pilot_status = "excluded"
                status_reason = "pilot_size_limit_exceeded"
            else:
                pilot_status = "deferred"
                status_reason = "outside_pilot_limit"
            pilot_status_counts[pilot_status] += 1
            out.execute(
                """
                INSERT OR REPLACE INTO stage6_ocr_pilot_plan (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, pilot_status,
                  pilot_rank, pilot_batch_label, status_reason, planned_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["source_archive_id"],
                    row["message_sha256"],
                    int(row["part_index"]),
                    row["attachment_sha256"],
                    row["extension"],
                    row["declared_content_type"],
                    int(row["size_bytes"]),
                    pilot_status,
                    rank,
                    "pilot_001",
                    status_reason,
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MAIL_ATTACHMENT_STAGE6_PLAN_SCHEMA_VERSION,
        "created_at": utc_now(),
        "parse_plan_queue_count": len(parse_later_rows),
        "text_index_row_count": len(text_index_rows),
        "gap_count": len(gap_rows),
        "gap_extension_counts": dict(sorted(gap_extension_counts.items())),
        "gap_class_counts": dict(sorted(gap_class_counts.items())),
        "gap_action_counts": dict(sorted(gap_action_counts.items())),
        "ocr_pilot": {
            "candidate_count": len(ocr_rows),
            "eligible_count": len(eligible_ocr_rows),
            "selected_count": len(selected_ocr_rows),
            "limit": max(0, int(config.ocr_pilot_limit)),
            "min_attachment_bytes": min_pilot_attachment_bytes,
            "max_attachment_bytes": max_pilot_attachment_bytes,
            "candidate_extension_counts": dict(sorted(ocr_candidate_extension_counts.items())),
            "selected_extension_counts": dict(sorted(selected_extension_counts.items())),
            "pilot_status_counts": dict(sorted(pilot_status_counts.items())),
        },
        "artifact_counts": {
            "gap_rows_written": len(gap_rows),
            "ocr_pilot_rows_written": len(ocr_rows),
            "ocr_selected_rows_written": len(selected_ocr_rows),
            "raw_text_chars_written": 0,
            "derived_text_files_read": 0,
            "attachment_bytes_read": 0,
            "ocr_text_files_written": 0,
            "images_written": 0,
            "thumbnails_written": 0,
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "read_raw_eml": False,
            "read_attachment_bytes": False,
            "read_existing_derived_text": False,
            "read_derived_text_content": False,
            "parse_pdf": False,
            "render_pdf": False,
            "parse_images": False,
            "decode_full_images": False,
            "extract_exif": False,
            "run_ocr": False,
            "ocr_enabled": False,
            "write_images": False,
            "write_thumbnails": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "raw_text_written_to_sqlite": False,
            "raw_text_written_to_json": False,
            "derived_text_content_written_to_sqlite": False,
            "derived_text_content_written_to_json": False,
            "derived_text_paths_written": False,
            "raw_ocr_text_written": False,
            "raw_exif_written": False,
            "message_sha256_written": True,
            "attachment_sha256_written": True,
        },
        "paths": {
            "stage6_db": str(out_db_path),
            "report": str(report_path),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_ocr_preflight(
    config: MailAttachmentOcrPreflightConfig,
) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.stage6_plan_db_path, "mail attachment stage 6 plan database")
    guard_not_stable_runtime(config.out_dir, "mail attachment OCR preflight output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment OCR preflight output directory")
    guard_git_ignored_output(out_dir, "mail attachment OCR preflight output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")
    stage6_plan_db_path = config.stage6_plan_db_path.resolve(strict=False)
    if not stage6_plan_db_path.exists():
        raise FileNotFoundError(f"mail attachment stage 6 plan database not found: {stage6_plan_db_path}")

    out_dir.mkdir(parents=True, exist_ok=True)
    out_db_path = out_dir / "mail_attachment_ocr_preflight.sqlite"
    report_path = out_dir / "mail_attachment_ocr_preflight_report.json"

    selected_rows = load_attachment_ocr_preflight_selected_rows(
        stage6_plan_db_path,
        max_candidates=max(0, int(config.max_candidates)),
    )
    archive_rows = load_attachment_text_extract_source_rows(source_paths)
    archive_row_by_key = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        ): row
        for row in archive_rows
    }
    archive_db_path_by_source_id = {mail_archive_source_id(path): path for path in source_paths}

    status_counts: Counter[str] = Counter()
    status_reason_counts: Counter[str] = Counter()
    extension_counts: Counter[str] = Counter()
    total_bytes_read = 0
    sha256_verified_count = 0

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attachment_ocr_preflight (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              pilot_rank INTEGER NOT NULL,
              preflight_status TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              sha256_verified INTEGER NOT NULL,
              bytes_read INTEGER NOT NULL,
              checked_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_ocr_preflight_verified AS
            SELECT * FROM attachment_ocr_preflight
            WHERE preflight_status = 'verified';
            CREATE VIEW IF NOT EXISTS v_attachment_ocr_preflight_needs_review AS
            SELECT * FROM attachment_ocr_preflight
            WHERE preflight_status != 'verified';
            DELETE FROM meta;
            DELETE FROM attachment_ocr_preflight;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_OCR_PREFLIGHT_SCHEMA_VERSION),
                ("created_at", utc_now()),
                (
                    "stage6_plan_db_path_sha256",
                    hashlib.sha256(str(stage6_plan_db_path).encode("utf-8")).hexdigest(),
                ),
                ("archive_db_count", str(len(source_paths))),
                ("max_candidates", str(max(0, int(config.max_candidates)))),
                ("max_attachment_bytes", str(max(0, int(config.max_attachment_bytes)))),
                ("verify_sha256", "true" if config.verify_sha256 else "false"),
            ],
        )
        for selected_row in selected_rows:
            key = (
                selected_row["source_archive_id"],
                selected_row["message_sha256"],
                int(selected_row["part_index"]),
                selected_row["attachment_sha256"],
            )
            archive_row = archive_row_by_key.get(key)
            result = preflight_one_ocr_candidate(
                selected_row=selected_row,
                archive_row=archive_row,
                archive_db_path_by_source_id=archive_db_path_by_source_id,
                max_attachment_bytes=max(0, int(config.max_attachment_bytes)),
                verify_sha256=bool(config.verify_sha256),
            )
            status_counts[result["preflight_status"]] += 1
            status_reason_counts[result["status_reason"]] += 1
            extension_counts[result["extension"]] += 1
            total_bytes_read += int(result["bytes_read"])
            if result["sha256_verified"]:
                sha256_verified_count += 1
            out.execute(
                """
                INSERT OR REPLACE INTO attachment_ocr_preflight (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, pilot_rank,
                  preflight_status, status_reason, sha256_verified, bytes_read,
                  checked_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    result["source_archive_id"],
                    result["message_sha256"],
                    int(result["part_index"]),
                    result["attachment_sha256"],
                    result["extension"],
                    result["declared_content_type"],
                    int(result["size_bytes"]),
                    int(result["pilot_rank"]),
                    result["preflight_status"],
                    result["status_reason"],
                    1 if result["sha256_verified"] else 0,
                    int(result["bytes_read"]),
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MAIL_ATTACHMENT_OCR_PREFLIGHT_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "selected_candidate_count": len(selected_rows),
        "status_counts": {
            "verified": int(status_counts.get("verified", 0)),
            "skipped": int(status_counts.get("skipped", 0)),
            "blocked_safety": int(status_counts.get("blocked_safety", 0)),
        },
        "status_reason_counts": dict(sorted(status_reason_counts.items())),
        "extension_counts": dict(sorted(extension_counts.items())),
        "artifact_counts": {
            "rows_written": len(selected_rows),
            "attachment_bytes_read": total_bytes_read,
            "sha256_verified": sha256_verified_count,
            "ocr_text_files_written": 0,
            "images_written": 0,
            "thumbnails_written": 0,
            "max_attachment_bytes": max(0, int(config.max_attachment_bytes)),
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "stage6_plan_sqlite_mode": "mode=ro",
            "stage6_plan_sqlite_query_only": True,
            "read_raw_eml": False,
            "read_attachment_bytes": bool(config.verify_sha256),
            "read_attachment_bytes_scope": "selected_ocr_pilot_candidates_only",
            "decode_full_images": False,
            "parse_images": False,
            "extract_exif": False,
            "write_images": False,
            "write_thumbnails": False,
            "run_ocr": False,
            "ocr_enabled": False,
            "read_existing_derived_text": False,
            "read_derived_text_content": False,
            "parse_pdf": False,
            "render_pdf": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "raw_exif_written": False,
            "raw_ocr_text_written": False,
            "raw_text_written_to_sqlite": False,
            "raw_text_written_to_json": False,
            "message_sha256_written": True,
            "attachment_sha256_written": True,
        },
        "paths": {
            "preflight_db": str(out_db_path),
            "report": str(report_path),
        },
    }
    write_json(report_path, report)
    return report


def build_mail_attachment_ocr_pilot(config: MailAttachmentOcrPilotConfig) -> Mapping[str, Any]:
    for archive_db_path in config.archive_db_paths:
        guard_not_stable_runtime(archive_db_path, "mail archive database")
    guard_not_stable_runtime(config.ocr_preflight_db_path, "mail attachment OCR preflight database")
    guard_not_stable_runtime(config.out_dir, "mail attachment OCR pilot output directory")
    out_dir = config.out_dir.resolve(strict=False)
    guard_external_handoffs_output(out_dir, "mail attachment OCR pilot output directory")
    guard_git_ignored_output(out_dir, "mail attachment OCR pilot output directory")

    source_paths = [Path(path).resolve(strict=False) for path in config.archive_db_paths]
    if not source_paths:
        raise ValueError("at least one mail archive database is required")
    for source_path in source_paths:
        if not source_path.exists():
            raise FileNotFoundError(f"mail archive database not found: {source_path}")
    ocr_preflight_db_path = config.ocr_preflight_db_path.resolve(strict=False)
    if not ocr_preflight_db_path.exists():
        raise FileNotFoundError(f"mail attachment OCR preflight database not found: {ocr_preflight_db_path}")

    languages = normalize_tesseract_languages(config.languages)
    if not languages:
        raise ValueError("at least one OCR language is required")
    worker_count = max(1, int(config.workers))
    tesseract_thread_limit = max(1, int(config.tesseract_thread_limit))

    out_dir.mkdir(parents=True, exist_ok=True)
    derived_text_dir = out_dir / "attachment_ocr_text"
    out_db_path = out_dir / "mail_attachment_ocr_pilot.sqlite"
    report_path = out_dir / "mail_attachment_ocr_pilot_report.json"

    verified_rows = load_attachment_ocr_pilot_verified_rows(
        ocr_preflight_db_path,
        max_candidates=max(0, int(config.max_candidates)),
    )
    archive_rows = load_attachment_text_extract_source_rows(source_paths)
    archive_row_by_key = {
        (
            row["source_archive_id"],
            row["message_sha256"],
            int(row["part_index"]),
            row["attachment_sha256"],
        ): row
        for row in archive_rows
    }
    archive_db_path_by_source_id = {mail_archive_source_id(path): path for path in source_paths}

    def _process_verified_row(verified_row: Mapping[str, Any]) -> dict[str, Any]:
        key = (
            verified_row["source_archive_id"],
            verified_row["message_sha256"],
            int(verified_row["part_index"]),
            verified_row["attachment_sha256"],
        )
        archive_row = archive_row_by_key.get(key)
        return run_one_attachment_ocr_pilot(
            verified_row=verified_row,
            archive_row=archive_row,
            archive_db_path_by_source_id=archive_db_path_by_source_id,
            derived_text_dir=derived_text_dir,
            languages=languages,
            page_segmentation_mode=int(config.page_segmentation_mode),
            timeout_seconds=max(0, int(config.tesseract_timeout_seconds)),
            max_attachment_bytes=max(0, int(config.max_attachment_bytes)),
            max_text_chars=max(0, int(config.max_text_chars_per_attachment)),
            tesseract_thread_limit=tesseract_thread_limit,
            reuse_existing_ocr_text=config.reuse_existing_ocr_text,
        )

    processing_started = time.monotonic()
    if worker_count == 1 or len(verified_rows) <= 1:
        results = [_process_verified_row(row) for row in verified_rows]
    else:
        indexed_results: list[tuple[int, dict[str, Any]]] = []
        with ThreadPoolExecutor(max_workers=worker_count, thread_name_prefix="mail-ocr") as executor:
            futures = {
                executor.submit(_process_verified_row, row): index
                for index, row in enumerate(verified_rows)
            }
            for future in as_completed(futures):
                indexed_results.append((futures[future], future.result()))
        results = [
            result
            for _index, result in sorted(
                indexed_results,
                key=lambda item: (int(item[1]["pilot_rank"]), item[0]),
            )
        ]
    processing_wall_seconds = time.monotonic() - processing_started

    status_counts: Counter[str] = Counter(result["status"] for result in results)
    status_reason_counts: Counter[str] = Counter(result["status_reason"] for result in results)
    extension_counts: Counter[str] = Counter(result["extension"] for result in results)
    warning_counts: Counter[str] = Counter(
        warning for result in results for warning in result["warnings"]
    )
    total_text_chars = sum(int(result["text_chars"]) for result in results)
    ocr_text_files_written = sum(1 for result in results if result["derived_text_written"])
    ocr_text_files_reused = sum(1 for result in results if result["derived_text_reused"])
    attachment_bytes_submitted = sum(
        int(result["size_bytes"]) for result in results if result["ocr_attempted"]
    )

    with sqlite3.connect(str(out_db_path)) as out:
        out.row_factory = sqlite3.Row
        out.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS attachment_ocr_pilot (
              source_archive_id TEXT NOT NULL,
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              attachment_sha256 TEXT NOT NULL,
              extension TEXT NOT NULL,
              declared_content_type TEXT NOT NULL,
              size_bytes INTEGER NOT NULL,
              pilot_rank INTEGER NOT NULL,
              ocr_engine TEXT NOT NULL,
              ocr_languages TEXT NOT NULL,
              page_segmentation_mode INTEGER NOT NULL,
              status TEXT NOT NULL,
              status_reason TEXT NOT NULL,
              warnings_json TEXT NOT NULL,
              text_sha256 TEXT NOT NULL,
              text_chars INTEGER NOT NULL,
              text_truncated INTEGER NOT NULL,
              derived_text_path_sha256 TEXT NOT NULL,
              ocr_seconds REAL NOT NULL,
              processed_at TEXT NOT NULL,
              PRIMARY KEY (
                source_archive_id, message_sha256, part_index, attachment_sha256
              )
            );
            CREATE VIEW IF NOT EXISTS v_attachment_ocr_pilot_extracted AS
            SELECT * FROM attachment_ocr_pilot
            WHERE status = 'extracted';
            CREATE VIEW IF NOT EXISTS v_attachment_ocr_pilot_needs_review AS
            SELECT * FROM attachment_ocr_pilot
            WHERE status != 'extracted';
            DELETE FROM meta;
            DELETE FROM attachment_ocr_pilot;
            """
        )
        out.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ATTACHMENT_OCR_PILOT_SCHEMA_VERSION),
                ("created_at", utc_now()),
                (
                    "ocr_preflight_db_path_sha256",
                    hashlib.sha256(str(ocr_preflight_db_path).encode("utf-8")).hexdigest(),
                ),
                ("archive_db_count", str(len(source_paths))),
                ("max_candidates", str(max(0, int(config.max_candidates)))),
                ("max_attachment_bytes", str(max(0, int(config.max_attachment_bytes)))),
                ("languages", languages),
                ("page_segmentation_mode", str(int(config.page_segmentation_mode))),
                ("tesseract_timeout_seconds", str(max(0, int(config.tesseract_timeout_seconds)))),
                ("workers", str(worker_count)),
                ("tesseract_thread_limit", str(tesseract_thread_limit)),
                ("reuse_existing_ocr_text", "true" if config.reuse_existing_ocr_text else "false"),
                (
                    "max_text_chars_per_attachment",
                    str(max(0, int(config.max_text_chars_per_attachment))),
                ),
            ],
        )
        for result in results:
            out.execute(
                """
                INSERT OR REPLACE INTO attachment_ocr_pilot (
                  source_archive_id, message_sha256, part_index, attachment_sha256,
                  extension, declared_content_type, size_bytes, pilot_rank,
                  ocr_engine, ocr_languages, page_segmentation_mode, status,
                  status_reason, warnings_json, text_sha256, text_chars,
                  text_truncated, derived_text_path_sha256, ocr_seconds,
                  processed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    result["source_archive_id"],
                    result["message_sha256"],
                    int(result["part_index"]),
                    result["attachment_sha256"],
                    result["extension"],
                    result["declared_content_type"],
                    int(result["size_bytes"]),
                    int(result["pilot_rank"]),
                    "tesseract_cli",
                    languages,
                    int(config.page_segmentation_mode),
                    result["status"],
                    result["status_reason"],
                    json.dumps(result["warnings"], ensure_ascii=False),
                    result["text_sha256"],
                    int(result["text_chars"]),
                    1 if result["text_truncated"] else 0,
                    result["derived_text_path_sha256"],
                    float(result["ocr_seconds"]),
                    utc_now(),
                ),
            )
        out.commit()

    remove_sqlite_sidecars(out_db_path)
    report = {
        "schema_version": MAIL_ATTACHMENT_OCR_PILOT_SCHEMA_VERSION,
        "created_at": utc_now(),
        "source_archive_count": len(source_paths),
        "verified_candidate_count": len(verified_rows),
        "status_counts": {
            "extracted": int(status_counts.get("extracted", 0)),
            "empty_text": int(status_counts.get("empty_text", 0)),
            "skipped": int(status_counts.get("skipped", 0)),
            "blocked_safety": int(status_counts.get("blocked_safety", 0)),
            "ocr_error": int(status_counts.get("ocr_error", 0)),
        },
        "status_reason_counts": dict(sorted(status_reason_counts.items())),
        "extension_counts": dict(sorted(extension_counts.items())),
        "warning_counts": dict(sorted(warning_counts.items())),
        "artifact_counts": {
            "rows_written": len(verified_rows),
            "attachment_bytes_submitted_to_ocr": attachment_bytes_submitted,
            "ocr_text_files_written": ocr_text_files_written,
            "ocr_text_files_reused": ocr_text_files_reused,
            "ocr_text_chars": total_text_chars,
            "images_written": 0,
            "thumbnails_written": 0,
            "max_attachment_bytes": max(0, int(config.max_attachment_bytes)),
        },
        "ocr": {
            "engine": "tesseract_cli",
            "languages": languages,
            "page_segmentation_mode": int(config.page_segmentation_mode),
            "timeout_seconds": max(0, int(config.tesseract_timeout_seconds)),
            "workers": worker_count,
            "tesseract_thread_limit": tesseract_thread_limit,
            "reuse_existing_ocr_text": bool(config.reuse_existing_ocr_text),
            "processing_wall_seconds": round(processing_wall_seconds, 3),
        },
        "safety": {
            "source_sqlite_mode": "mode=ro",
            "source_sqlite_query_only": True,
            "ocr_preflight_sqlite_mode": "mode=ro",
            "ocr_preflight_sqlite_query_only": True,
            "read_raw_eml": False,
            "read_attachment_bytes": True,
            "read_attachment_bytes_scope": "verified_ocr_pilot_candidates_only",
            "decode_full_images": True,
            "decode_full_images_scope": "tesseract_cli_only",
            "parse_images": True,
            "extract_exif": False,
            "write_images": False,
            "write_thumbnails": False,
            "run_ocr": True,
            "ocr_enabled": True,
            "read_existing_derived_text": bool(config.reuse_existing_ocr_text),
            "read_derived_text_content": bool(config.reuse_existing_ocr_text),
            "parse_pdf": False,
            "render_pdf": False,
            "execute_files": False,
            "extract_archives_to_disk": False,
            "follow_external_links": False,
            "write_crm": False,
            "write_tallanto": False,
            "live_crm_reads": False,
            "runtime_db_writes": False,
            "stable_runtime_writes": False,
            "run_asr": False,
            "run_ra": False,
        },
        "privacy": {
            "contains_raw_personal_values_in_json": False,
            "raw_filenames_written": False,
            "raw_source_attachment_paths_written": False,
            "raw_attachment_content_written_to_sqlite": False,
            "raw_attachment_content_written_to_json": False,
            "raw_exif_written": False,
            "raw_ocr_text_written_to_json": False,
            "raw_ocr_text_written_to_sqlite": False,
            "ocr_text_files_may_contain_personal_data": True,
            "derived_text_paths_written": False,
            "derived_text_path_hashes_written": True,
            "message_sha256_written": True,
            "attachment_sha256_written": True,
        },
        "paths": {
            "ocr_pilot_db": str(out_db_path),
            "report": str(report_path),
            "ocr_text_dir": str(derived_text_dir),
        },
    }
    write_json(report_path, report)
    return report


def ingest_one_message(
    imap: ImapClient,
    *,
    msg_id: bytes,
    credentials: MailImapCredentials,
    config: MailArchiveIngestConfig,
    db_path: Path,
    raw_dir: Path,
    attachment_dir: Path,
    text_dir: Path,
) -> Mapping[str, Any]:
    fetch_status, fetch_data = imap.fetch(msg_id, FULL_MESSAGE_FETCH_QUERY)
    if fetch_status != "OK":
        raise RuntimeError(f"IMAP FETCH failed: {fetch_status}")
    raw = first_fetch_payload(fetch_data)
    if not raw:
        raise RuntimeError("IMAP FETCH returned empty payload")
    raw_sha256 = hashlib.sha256(raw).hexdigest()
    if raw_sha256 in set(config.exclude_message_sha256s):
        return {
            "sha256": raw_sha256,
            "excluded_by_sha256": True,
            "raw_eml_written": False,
            "text_file_written": False,
            "attachments_written": 0,
        }
    raw_path = raw_message_path(raw_dir, raw_sha256)
    raw_written = write_bytes_once(raw_path, raw)

    msg = email.message_from_bytes(raw, policy=policy.default)
    metadata = message_metadata(msg)
    source_key = hashlib.sha256(
        "|".join(
            [
                config.account_label,
                config.mailbox,
                msg_id.decode("ascii", "ignore"),
                metadata["message_id"],
                raw_sha256,
            ]
        ).encode("utf-8")
    ).hexdigest()
    message_kind = classify_message_kind(
        mailbox_email=credentials.email_address,
        participants=message_participants(msg),
        configured_internal_domains=config.internal_domains,
    )
    extracted_text = extract_message_text(msg, max_chars=config.extracted_text_max_chars)
    text_path = text_dir / f"{raw_sha256}.txt"
    text_written = write_text_once(text_path, extracted_text) if extracted_text else False
    attachments_written = save_attachments(msg, raw_sha256=raw_sha256, attachment_dir=attachment_dir)

    with sqlite3.connect(str(db_path)) as con:
        con.row_factory = sqlite3.Row
        upsert_message(
            con,
            sha256=raw_sha256,
            source_key=source_key,
            msg_id=msg_id,
            config=config,
            metadata=metadata,
            message_kind=message_kind,
            raw_path=raw_path,
            text_path=text_path if extracted_text else None,
            raw_size_bytes=len(raw),
            extracted_text_chars=len(extracted_text),
        )
        replace_participants(con, raw_sha256, message_participants(msg))
        replace_attachments(con, raw_sha256, msg, attachment_dir / raw_sha256)
        con.commit()

    return {
        "sha256": raw_sha256,
        "excluded_by_sha256": False,
        "raw_eml_written": raw_written,
        "text_file_written": text_written,
        "attachments_written": attachments_written,
    }


def read_tallanto_rows(path: Path, *, encoding: str, delimiter: str) -> list[dict[str, str]]:
    with path.open("r", encoding=encoding, newline="", errors="replace") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        return [dict(row) for row in reader]


def tallanto_id_for_row(row: Mapping[str, Any]) -> str:
    for column in TALLANTO_ID_COLUMNS:
        value = clean_text(row.get(column))
        if value:
            return value
    return ""


def collect_row_identity_values(
    row: Mapping[str, Any],
    *,
    columns: Sequence[str],
    extractor: Any,
) -> dict[str, list[str]]:
    values: dict[str, list[str]] = {}
    for column in columns:
        if column not in row:
            continue
        for value in extractor(row.get(column)):
            values.setdefault(value, [])
            if column not in values[value]:
                values[value].append(column)
    return values


def append_identity_link(
    identity_links: dict[tuple[str, str], dict[str, Any]],
    kind: str,
    value: str,
    candidate_key: str,
    columns: Sequence[str],
) -> None:
    item = identity_links.setdefault(
        (kind, value),
        {
            "kind": kind,
            "value": value,
            "candidate_keys": [],
            "columns_by_candidate": {},
            "match_class": "",
        },
    )
    if candidate_key not in item["candidate_keys"]:
        item["candidate_keys"].append(candidate_key)
    item["columns_by_candidate"].setdefault(candidate_key, [])
    for column in columns:
        if column not in item["columns_by_candidate"][candidate_key]:
            item["columns_by_candidate"][candidate_key].append(column)


def write_identity_map_db(
    db_path: Path,
    *,
    source_path: Path,
    candidate_rows: Sequence[Mapping[str, Any]],
    identity_links: Mapping[tuple[str, str], Mapping[str, Any]],
    row_identity_classes: Mapping[str, Mapping[str, str]],
) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(db_path)) as con:
        con.executescript(
            """
            PRAGMA journal_mode=DELETE;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS identity_candidates (
              candidate_key TEXT PRIMARY KEY,
              row_number INTEGER NOT NULL,
              tallanto_id TEXT,
              amocrm_id TEXT,
              first_name TEXT,
              last_name TEXT,
              parent_name TEXT,
              student_type TEXT,
              manager TEXT,
              manager_id TEXT,
              group_id TEXT,
              email_class TEXT NOT NULL,
              phone_class TEXT NOT NULL,
              candidate_json TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS identity_values (
              kind TEXT NOT NULL,
              value TEXT NOT NULL,
              match_class TEXT NOT NULL,
              candidate_count INTEGER NOT NULL,
              PRIMARY KEY (kind, value)
            );
            CREATE TABLE IF NOT EXISTS identity_links (
              kind TEXT NOT NULL,
              value TEXT NOT NULL,
              candidate_key TEXT NOT NULL,
              source_columns_json TEXT NOT NULL,
              PRIMARY KEY (kind, value, candidate_key)
            );
            DELETE FROM meta;
            DELETE FROM identity_candidates;
            DELETE FROM identity_values;
            DELETE FROM identity_links;
            """
        )
        con.executemany(
            "INSERT INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", TALLANTO_IDENTITY_MAP_SCHEMA_VERSION),
                ("created_at", utc_now()),
                ("source_path", str(source_path)),
            ],
        )
        for row in candidate_rows:
            classes = row_identity_classes[row["candidate_key"]]
            con.execute(
                """
                INSERT INTO identity_candidates (
                  candidate_key, row_number, tallanto_id, amocrm_id, first_name, last_name,
                  parent_name, student_type, manager, manager_id, group_id,
                  email_class, phone_class, candidate_json
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    row["candidate_key"],
                    row["row_number"],
                    row.get("tallanto_id", ""),
                    row.get("amocrm_id", ""),
                    row.get("first_name", ""),
                    row.get("last_name", ""),
                    row.get("parent_name", ""),
                    row.get("student_type", ""),
                    row.get("manager", ""),
                    row.get("manager_id", ""),
                    row.get("group_id", ""),
                    classes["email"],
                    classes["phone"],
                    row["candidate_json"],
                ),
            )
        for item in identity_links.values():
            con.execute(
                """
                INSERT INTO identity_values (kind, value, match_class, candidate_count)
                VALUES (?, ?, ?, ?)
                """,
                (
                    item["kind"],
                    item["value"],
                    item["match_class"],
                    len(item["candidate_keys"]),
                ),
            )
            for candidate_key, columns in item["columns_by_candidate"].items():
                con.execute(
                    """
                    INSERT INTO identity_links (
                      kind, value, candidate_key, source_columns_json
                    ) VALUES (?, ?, ?, ?)
                    """,
                    (
                        item["kind"],
                        item["value"],
                        candidate_key,
                        json.dumps(columns, ensure_ascii=False),
                    ),
                )
        con.commit()


def build_identity_map_report(
    *,
    db_path: Path,
    report_path: Path,
    source_path: Path,
    candidate_rows: Sequence[Mapping[str, Any]],
    identity_links: Mapping[tuple[str, str], Mapping[str, Any]],
    row_identity_classes: Mapping[str, Mapping[str, str]],
    config: TallantoIdentityMapConfig,
    duplicate_tallanto_id_values: int = 0,
) -> Mapping[str, Any]:
    values_by_kind: dict[str, dict[str, int]] = {
        "email": {"strong_unique": 0, "duplicate": 0},
        "phone": {"strong_unique": 0, "duplicate": 0},
    }
    for item in identity_links.values():
        values_by_kind[item["kind"]][item["match_class"]] += 1

    row_counts: dict[str, dict[str, int]] = {
        "email": {"strong_unique": 0, "duplicate": 0, "missing": 0},
        "phone": {"strong_unique": 0, "duplicate": 0, "missing": 0},
    }
    rows_with_any_strong = 0
    rows_with_both_strong = 0
    for classes in row_identity_classes.values():
        row_counts["email"][classes["email"]] += 1
        row_counts["phone"][classes["phone"]] += 1
        email_strong = classes["email"] == "strong_unique"
        phone_strong = classes["phone"] == "strong_unique"
        rows_with_any_strong += int(email_strong or phone_strong)
        rows_with_both_strong += int(email_strong and phone_strong)

    source_size = source_path.stat().st_size if source_path.exists() else 0
    identity_db_size = db_path.stat().st_size if db_path.exists() else 0
    sanity_checks = {
        "email_row_classes_equal_row_count": sum(row_counts["email"].values()) == len(candidate_rows),
        "phone_row_classes_equal_row_count": sum(row_counts["phone"].values()) == len(candidate_rows),
        "identity_values_equal_links_by_value": len(identity_links) == sum(values_by_kind[kind][klass] for kind in values_by_kind for klass in values_by_kind[kind]),
        "source_file_exists": source_path.exists(),
        "identity_db_exists": db_path.exists(),
    }
    warnings = build_identity_report_warnings(
        duplicate_tallanto_id_values=duplicate_tallanto_id_values,
        values_by_kind=values_by_kind,
        row_counts=row_counts,
    )
    blocking_risks = [name for name, passed in sanity_checks.items() if not passed]

    return {
        "schema_version": TALLANTO_IDENTITY_MAP_SCHEMA_VERSION,
        "created_at": utc_now(),
        "provenance": build_local_provenance(),
        "source_path": str(source_path),
        "source_file": {
            "size_bytes": source_size,
            "sha256": sha256_file(source_path) if source_path.exists() else "",
        },
        "row_count": len(candidate_rows),
        "duplicate_tallanto_id_values": duplicate_tallanto_id_values,
        "source_id_quality": {
            "duplicate_tallanto_id_values": duplicate_tallanto_id_values,
        },
        "identity_values": values_by_kind,
        "row_identity_classes": row_counts,
        "row_coverage": {
            "rows_with_contact_email": len(candidate_rows) - row_counts["email"]["missing"],
            "rows_without_contact_email": row_counts["email"]["missing"],
            "rows_with_contact_phone": len(candidate_rows) - row_counts["phone"]["missing"],
            "rows_without_contact_phone": row_counts["phone"]["missing"],
            "rows_with_strong_unique_email": row_counts["email"]["strong_unique"],
            "rows_with_strong_unique_phone": row_counts["phone"]["strong_unique"],
            "rows_with_any_strong_email_or_phone": rows_with_any_strong,
            "rows_with_both_strong_email_and_phone": rows_with_both_strong,
        },
        "identity_link_counts": {
            "candidate_rows": len(candidate_rows),
            "identity_values": len(identity_links),
            "identity_links": sum(len(item["candidate_keys"]) for item in identity_links.values()),
        },
        "sanity_checks": sanity_checks,
        "columns_used": {
            "emails": list(config.email_columns),
            "phones": list(config.phone_columns),
            "candidate": list(config.candidate_columns),
        },
        "artifact_integrity": {
            "identity_db_size_bytes": identity_db_size,
            "identity_db_sha256": sha256_file(db_path) if db_path.exists() else "",
        },
        "audit_readiness": {
            "pass": not blocking_risks,
            "blocking_risks": blocking_risks,
            "warnings": warnings,
            "recommended_next_action": (
                "run_small_imap_pilot"
                if not blocking_risks
                else "fix_identity_map_sanity_checks_before_imap_pilot"
            ),
        },
        "pii_artifacts": {
            "source_csv": {
                "path": str(source_path),
                "contains_pii": True,
                "commit_safe": False,
            },
            "identity_db": {
                "path": str(db_path),
                "contains_pii": True,
                "commit_safe": False,
            },
            "identity_report": {
                "path": str(report_path),
                "contains_pii": False,
                "commit_safe": False,
                "reason": "operational handoff report under ignored raw artifact root",
            },
        },
        "privacy": {
            "report_contains_raw_personal_values": False,
            "sqlite_contains_selected_identity_values": True,
            "raw_history_columns_written": False,
            "raw_export_copied": False,
        },
        "paths": {
            "identity_db": str(db_path),
            "report": str(report_path),
        },
    }


def build_identity_report_warnings(
    *,
    duplicate_tallanto_id_values: int,
    values_by_kind: Mapping[str, Mapping[str, int]],
    row_counts: Mapping[str, Mapping[str, int]],
) -> list[str]:
    warnings: list[str] = []
    if duplicate_tallanto_id_values:
        warnings.append("duplicate_tallanto_id_values_present")
    for kind in ("email", "phone"):
        if int(values_by_kind[kind].get("duplicate") or 0):
            warnings.append(f"duplicate_{kind}_identity_values_present")
        if int(row_counts[kind].get("missing") or 0):
            warnings.append(f"rows_missing_{kind}_identity_present")
    warnings.append("identity_sqlite_contains_pii_keep_out_of_git")
    return warnings


def build_local_provenance() -> Mapping[str, Any]:
    return {
        "python_version": sys.version.split()[0],
        "platform": platform.platform(),
        "git": best_effort_git_provenance(),
    }


def best_effort_git_provenance() -> Mapping[str, Any]:
    def run_git(args: Sequence[str]) -> str:
        try:
            completed = subprocess.run(
                ["git", *args],
                check=False,
                capture_output=True,
                text=True,
                timeout=5,
            )
        except Exception:  # noqa: BLE001
            return ""
        if completed.returncode != 0:
            return ""
        return completed.stdout.strip()

    commit = run_git(["rev-parse", "HEAD"])
    status = run_git(["status", "--short"])
    return {
        "commit": commit,
        "dirty": bool(status),
        "status_short_line_count": len([line for line in status.splitlines() if line.strip()]),
    }


def git_check_ignored(path: Path) -> bool:
    try:
        completed = subprocess.run(
            ["git", "check-ignore", "-q", str(path)],
            check=False,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=5,
        )
    except Exception:  # noqa: BLE001
        return False
    return completed.returncode == 0


def guard_git_ignored_output(path: Path, label: str) -> None:
    if not git_check_ignored(path):
        raise ValueError(f"{label} must be git-ignored before writing raw mail artifacts")


def guard_external_handoffs_output(path: Path, label: str) -> None:
    if "_external_handoffs" not in path.resolve(strict=False).parts:
        raise ValueError(f"{label} must be under _external_handoffs")


def open_sqlite_readonly(path: Path) -> sqlite3.Connection:
    uri = f"file:{quote(str(path.resolve(strict=False)), safe='/:')}?mode=ro"
    con = sqlite3.connect(uri, uri=True, timeout=15)
    con.row_factory = sqlite3.Row
    con.execute("PRAGMA query_only = ON")
    return con


def init_mail_archive_db(db_path: Path) -> None:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(str(db_path)) as con:
        con.executescript(
            """
            PRAGMA journal_mode=WAL;
            CREATE TABLE IF NOT EXISTS meta (
              key TEXT PRIMARY KEY,
              value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS messages (
              sha256 TEXT PRIMARY KEY,
              message_id TEXT,
              message_date_header TEXT,
              message_date_iso TEXT,
              subject TEXT,
              from_header TEXT,
              to_header TEXT,
              cc_header TEXT,
              mailbox TEXT NOT NULL,
              mailbox_raw TEXT NOT NULL,
              message_kind TEXT NOT NULL,
              raw_eml_path TEXT NOT NULL,
              extracted_text_path TEXT,
              raw_size_bytes INTEGER NOT NULL,
              extracted_text_chars INTEGER NOT NULL,
              first_ingested_at TEXT NOT NULL,
              updated_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS message_sources (
              source_key TEXT PRIMARY KEY,
              message_sha256 TEXT NOT NULL,
              account_label TEXT NOT NULL,
              mailbox TEXT NOT NULL,
              mailbox_raw TEXT NOT NULL,
              imap_seq TEXT NOT NULL,
              source_message_id TEXT,
              ingested_at TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS message_participants (
              message_sha256 TEXT NOT NULL,
              header_name TEXT NOT NULL,
              display_name TEXT,
              email_raw TEXT,
              email_normalized TEXT,
              domain TEXT,
              PRIMARY KEY (message_sha256, header_name, email_normalized, email_raw)
            );
            CREATE TABLE IF NOT EXISTS attachments (
              message_sha256 TEXT NOT NULL,
              part_index INTEGER NOT NULL,
              filename TEXT,
              content_type TEXT,
              content_disposition TEXT,
              size_bytes INTEGER NOT NULL,
              sha256 TEXT NOT NULL,
              path TEXT NOT NULL,
              PRIMARY KEY (message_sha256, part_index)
            );
            """
        )
        con.executemany(
            "INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
            [
                ("schema_version", MAIL_ARCHIVE_SCHEMA_VERSION),
                ("updated_at", utc_now()),
            ],
        )
        init_mail_match_tables(con)
        con.commit()


def init_mail_match_tables(con: sqlite3.Connection) -> None:
    con.executescript(
        """
        CREATE TABLE IF NOT EXISTS message_matches (
          message_sha256 TEXT PRIMARY KEY,
          match_class TEXT NOT NULL,
          candidate_count INTEGER NOT NULL,
          matched_email_count INTEGER NOT NULL,
          matched_at TEXT NOT NULL
        );
        """
    )


def empty_archive_db_counts() -> dict[str, int]:
    return {
        "messages": 0,
        "message_sources": 0,
        "message_participants": 0,
        "attachments": 0,
        "message_matches": 0,
    }


def archive_db_counts(con: sqlite3.Connection) -> dict[str, int]:
    counts = empty_archive_db_counts()
    for table in counts:
        try:
            counts[table] = int(con.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])
        except sqlite3.Error:
            counts[table] = 0
    return counts


def archive_db_has_required_tables(con: sqlite3.Connection) -> bool:
    required = {
        "messages",
        "message_sources",
        "message_participants",
        "attachments",
        "message_matches",
    }
    rows = con.execute("SELECT name FROM sqlite_master WHERE type = 'table'").fetchall()
    present = {str(row[0]) for row in rows}
    return required.issubset(present)


def upsert_message(
    con: sqlite3.Connection,
    *,
    sha256: str,
    source_key: str,
    msg_id: bytes,
    config: MailArchiveIngestConfig,
    metadata: Mapping[str, str],
    message_kind: str,
    raw_path: Path,
    text_path: Optional[Path],
    raw_size_bytes: int,
    extracted_text_chars: int,
) -> None:
    now = utc_now()
    existing = con.execute("SELECT first_ingested_at FROM messages WHERE sha256 = ?", (sha256,)).fetchone()
    first_ingested_at = existing["first_ingested_at"] if existing else now
    con.execute(
        """
        INSERT OR REPLACE INTO messages (
          sha256, message_id, message_date_header, message_date_iso, subject,
          from_header, to_header, cc_header, mailbox, mailbox_raw, message_kind,
          raw_eml_path, extracted_text_path, raw_size_bytes, extracted_text_chars,
          first_ingested_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            sha256,
            metadata["message_id"],
            metadata["date"],
            metadata["date_iso"],
            metadata["subject"],
            metadata["from"],
            metadata["to"],
            metadata["cc"],
            config.mailbox_label,
            config.mailbox,
            message_kind,
            str(raw_path),
            str(text_path) if text_path else "",
            raw_size_bytes,
            extracted_text_chars,
            first_ingested_at,
            now,
        ),
    )
    con.execute(
        """
        INSERT OR REPLACE INTO message_sources (
          source_key, message_sha256, account_label, mailbox, mailbox_raw,
          imap_seq, source_message_id, ingested_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            source_key,
            sha256,
            config.account_label,
            config.mailbox_label,
            config.mailbox,
            msg_id.decode("ascii", "ignore"),
            metadata["message_id"],
            now,
        ),
    )


def replace_participants(
    con: sqlite3.Connection,
    message_sha256: str,
    participants: Sequence[Mapping[str, str]],
) -> None:
    con.execute("DELETE FROM message_participants WHERE message_sha256 = ?", (message_sha256,))
    for row in participants:
        con.execute(
            """
            INSERT OR REPLACE INTO message_participants (
              message_sha256, header_name, display_name, email_raw, email_normalized, domain
            ) VALUES (?, ?, ?, ?, ?, ?)
            """,
            (
                message_sha256,
                row["header_name"],
                row["display_name"],
                row["email_raw"],
                row["email_normalized"],
                domain_from_email(row["email_normalized"]),
            ),
        )


def replace_attachments(
    con: sqlite3.Connection,
    message_sha256: str,
    msg: Message,
    attachment_message_dir: Path,
) -> None:
    con.execute("DELETE FROM attachments WHERE message_sha256 = ?", (message_sha256,))
    for part_index, part in enumerate(iter_attachment_parts(msg), start=1):
        payload = part.get_payload(decode=True) or b""
        attachment_sha = hashlib.sha256(payload).hexdigest()
        filename = safe_part_filename(part)
        path = attachment_message_dir / f"part_{part_index:03d}_{attachment_sha[:16]}.bin"
        con.execute(
            """
            INSERT OR REPLACE INTO attachments (
              message_sha256, part_index, filename, content_type, content_disposition,
              size_bytes, sha256, path
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                message_sha256,
                part_index,
                filename,
                safe_part_content_type(part),
                safe_part_header(part, "Content-Disposition"),
                len(payload),
                attachment_sha,
                str(path),
            ),
        )


def save_attachments(msg: Message, *, raw_sha256: str, attachment_dir: Path) -> int:
    written = 0
    message_dir = attachment_dir / raw_sha256
    for part_index, part in enumerate(iter_attachment_parts(msg), start=1):
        payload = part.get_payload(decode=True) or b""
        attachment_sha = hashlib.sha256(payload).hexdigest()
        path = message_dir / f"part_{part_index:03d}_{attachment_sha[:16]}.bin"
        if write_bytes_once(path, payload):
            written += 1
    return written


def iter_attachment_parts(msg: Message) -> list[Message]:
    parts: list[Message] = []
    if not msg.is_multipart():
        return parts
    for part in msg.walk():
        if part.is_multipart():
            continue
        disposition = safe_part_content_disposition(part)
        filename = safe_part_filename(part)
        if disposition == "attachment" or filename:
            parts.append(part)
    return parts


def safe_part_content_disposition(part: Message) -> str:
    try:
        return clean_text(part.get_content_disposition()).lower()
    except Exception:  # noqa: BLE001
        raw = safe_part_header(part, "Content-Disposition").split(";", 1)[0]
        return clean_text(raw).lower()


def safe_part_filename(part: Message) -> str:
    try:
        return clean_header(part.get_filename()) or ""
    except Exception:  # noqa: BLE001
        return ""


def safe_part_content_type(part: Message) -> str:
    try:
        return clean_text(part.get_content_type())
    except Exception:  # noqa: BLE001
        return "application/octet-stream"


def safe_part_header(part: Message, name: str) -> str:
    try:
        return clean_text(part.get(name))
    except Exception:  # noqa: BLE001
        return ""


def message_metadata(msg: Message) -> dict[str, str]:
    date_header = safe_message_header(msg, "Date")
    return {
        "message_id": safe_message_header(msg, "Message-ID"),
        "date": date_header,
        "date_iso": parse_email_date_iso(date_header),
        "from": safe_message_header(msg, "From"),
        "to": safe_message_header(msg, "To"),
        "cc": safe_message_header(msg, "Cc"),
        "subject": safe_message_header(msg, "Subject"),
    }


def message_participants(msg: Message) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for header_name in ("From", "To", "Cc", "Bcc", "Reply-To"):
        values = safe_message_header_values(msg, header_name)
        for display_name, address in getaddresses(values):
            normalized = normalize_email(address)
            rows.append(
                {
                    "header_name": header_name,
                    "display_name": clean_header(display_name),
                    "email_raw": clean_text(address),
                    "email_normalized": normalized,
                }
            )
    return rows


def safe_message_header(msg: Message, name: str) -> str:
    values = safe_message_header_values(msg, name)
    return clean_header(values[0]) if values else ""


def safe_message_header_values(msg: Message, name: str) -> list[str]:
    try:
        return [clean_text(value) for value in msg.get_all(name, [])]
    except Exception:  # noqa: BLE001
        target = name.lower()
        try:
            return [
                clean_text(value)
                for key, value in msg.raw_items()
                if clean_text(key).lower() == target
            ]
        except Exception:  # noqa: BLE001
            return []


def classify_message_kind(
    *,
    mailbox_email: str,
    participants: Sequence[Mapping[str, str]],
    configured_internal_domains: Sequence[str],
) -> str:
    internal_domains = normalize_domains(configured_internal_domains, mailbox_email)
    normalized = [row["email_normalized"] for row in participants if row.get("email_normalized")]
    if any(is_service_email(address) for address in normalized):
        return "service"
    external = [address for address in normalized if not email_domain_is_internal(address, internal_domains)]
    if normalized and not external:
        return "internal"
    return "external"


def extract_message_text(msg: Message, *, max_chars: int) -> str:
    chunks: list[str] = []
    for part in msg.walk() if msg.is_multipart() else [msg]:
        if part.is_multipart():
            continue
        if safe_part_content_disposition(part) == "attachment":
            continue
        content_type = safe_part_content_type(part).lower()
        if content_type not in {"text/plain", "text/html"}:
            continue
        try:
            text = part.get_content()
        except Exception:  # noqa: BLE001
            payload = part.get_payload(decode=True) or b""
            charset = part.get_content_charset() or "utf-8"
            text = payload.decode(charset, errors="replace")
        text = clean_html_text(text) if content_type == "text/html" else str(text)
        text = text.strip()
        if text:
            chunks.append(text)
        if sum(len(chunk) for chunk in chunks) >= max_chars:
            break
    joined = "\n\n--- part ---\n\n".join(chunks)
    return joined[:max(0, max_chars)]


def clean_html_text(text: str) -> str:
    without_scripts = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", text)
    without_tags = re.sub(r"(?s)<[^>]+>", " ", without_scripts)
    return re.sub(r"\s+", " ", without_tags).strip()


def load_identity_index(identity_db_path: Path) -> dict[str, list[str]]:
    index: dict[str, list[str]] = {}
    with open_sqlite_readonly(identity_db_path) as con:
        for row in con.execute(
            """
            SELECT value, candidate_key
            FROM identity_links
            WHERE kind = 'email'
            ORDER BY value, candidate_key
            """
        ):
            index.setdefault(row["value"], []).append(row["candidate_key"])
    return index


def load_phone_identity_index(identity_db_path: Path) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    with open_sqlite_readonly(identity_db_path) as con:
        for row in con.execute(
            """
            SELECT il.value, il.candidate_key, iv.match_class, iv.candidate_count
            FROM identity_links il
            JOIN identity_values iv
              ON iv.kind = il.kind AND iv.value = il.value
            WHERE il.kind = 'phone'
            ORDER BY il.value, il.candidate_key
            """
        ):
            phone = clean_text(row["value"])
            item = index.setdefault(
                phone,
                {
                    "candidate_keys": [],
                    "match_class": clean_text(row["match_class"]),
                    "candidate_count": int(row["candidate_count"]),
                },
            )
            item["candidate_keys"].append(clean_text(row["candidate_key"]))
    return index


def load_mail_phone_lift_source_messages(archive_db_paths: Sequence[Path]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for archive_db_path in archive_db_paths:
        source_path = Path(archive_db_path).resolve(strict=False)
        source_archive_id = hashlib.sha256(str(source_path).encode("utf-8")).hexdigest()[:16]
        with open_sqlite_readonly(source_path) as con:
            required_tables = {"messages", "message_matches"}
            present = {
                clean_text(row["name"])
                for row in con.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
            }
            missing = sorted(required_tables - present)
            if missing:
                raise ValueError(
                    f"mail archive database is missing required tables for phone lift: {', '.join(missing)}"
                )
            for row in con.execute(
                """
                SELECT
                  m.sha256 AS message_sha256,
                  m.message_date_iso,
                  m.extracted_text_path,
                  mm.match_class
                FROM message_matches mm
                JOIN messages m ON m.sha256 = mm.message_sha256
                WHERE mm.match_class IN ('ambiguous', 'missing')
                ORDER BY m.message_date_iso, m.sha256
                """
            ):
                rows.append(
                    {
                        "source_archive_id": source_archive_id,
                        "source_archive_path": str(source_path),
                        "message_sha256": clean_text(row["message_sha256"]),
                        "message_date_iso": clean_text(row["message_date_iso"]),
                        "extracted_text_path": clean_text(row["extracted_text_path"]),
                        "match_class": clean_text(row["match_class"]),
                    }
                )
    return rows


def read_safe_extracted_text(
    value: object,
    *,
    archive_db_path: Path,
    max_chars: int,
) -> tuple[str, str]:
    raw_path = clean_text(value)
    if not raw_path:
        return "", "missing"
    path = Path(raw_path).resolve(strict=False)
    archive_dir = archive_db_path.resolve(strict=False).parent
    try:
        path.relative_to(archive_dir)
    except ValueError:
        return "", "rejected"
    if "_external_handoffs" not in path.parts:
        return "", "rejected"
    if "stable_runtime" in {part.casefold() for part in path.parts}:
        return "", "rejected"
    if "extracted_text" not in path.parts:
        return "", "rejected"
    if path.suffix.casefold() != ".txt":
        return "", "rejected"
    if not path.exists():
        return "", "missing"
    try:
        return path.read_text(encoding="utf-8", errors="replace")[: max(0, int(max_chars))], "read"
    except OSError:
        return "", "unreadable"


def count_phone_lift_rows(
    preview_db_path: Path,
    *,
    original_match_class: str,
    lift_class: str,
) -> int:
    with open_sqlite_readonly(preview_db_path) as con:
        return int(
            con.execute(
                """
                SELECT COUNT(*)
                FROM message_phone_lift_preview
                WHERE original_match_class = ? AND lift_class = ?
                """,
                (original_match_class, lift_class),
            ).fetchone()[0]
        )


def mail_archive_source_id(path: Path) -> str:
    return hashlib.sha256(str(Path(path).resolve(strict=False)).encode("utf-8")).hexdigest()[:16]


def load_attachment_parse_later_count(parse_plan_db_path: Path) -> int:
    with open_sqlite_readonly(parse_plan_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION:
            raise ValueError(
                "mail attachment parse plan schema mismatch: "
                f"expected {MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION}, got {schema_version}"
            )
        return int(
            con.execute(
                "SELECT COUNT(*) FROM attachment_parse_plan WHERE action = 'parse_later'"
            ).fetchone()[0]
        )


def load_attachment_stage6_parse_later_rows(parse_plan_db_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with open_sqlite_readonly(parse_plan_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION:
            raise ValueError(
                "mail attachment parse plan schema mismatch: "
                f"expected {MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION}, got {schema_version}"
            )
        for row in con.execute(
            """
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes, risk_reasons_json
            FROM attachment_parse_plan
            WHERE action = 'parse_later'
            ORDER BY source_archive_id, message_sha256, part_index, attachment_sha256
            """
        ):
            rows.append(
                {
                    "source_archive_id": clean_text(row["source_archive_id"]),
                    "message_sha256": clean_text(row["message_sha256"]),
                    "part_index": int(row["part_index"]),
                    "attachment_sha256": clean_text(row["attachment_sha256"]),
                    "extension": clean_text(row["extension"]),
                    "declared_content_type": clean_text(row["declared_content_type"]),
                    "size_bytes": int(row["size_bytes"] or 0),
                    "risk_reasons_json": clean_text(row["risk_reasons_json"] or "[]"),
                }
            )
    return rows


def load_attachment_stage6_text_index_rows(text_index_db_path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with open_sqlite_readonly(text_index_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != MAIL_ATTACHMENT_TEXT_INDEX_SCHEMA_VERSION:
            raise ValueError(
                "mail attachment text index schema mismatch: "
                f"expected {MAIL_ATTACHMENT_TEXT_INDEX_SCHEMA_VERSION}, got {schema_version}"
            )
        table_row = con.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'attachment_text_index'"
        ).fetchone()
        if table_row is None:
            raise ValueError("mail attachment text index database missing table: attachment_text_index")
        for row in con.execute(
            """
            SELECT
              source_stage, source_archive_id, message_sha256, part_index,
              attachment_sha256, extension, declared_content_type, size_bytes,
              text_status
            FROM attachment_text_index
            ORDER BY source_stage, source_archive_id, message_sha256, part_index, attachment_sha256
            """
        ):
            rows.append(
                {
                    "source_stage": clean_text(row["source_stage"]),
                    "source_archive_id": clean_text(row["source_archive_id"]),
                    "message_sha256": clean_text(row["message_sha256"]),
                    "part_index": int(row["part_index"]),
                    "attachment_sha256": clean_text(row["attachment_sha256"]),
                    "extension": clean_text(row["extension"]),
                    "declared_content_type": clean_text(row["declared_content_type"]),
                    "size_bytes": int(row["size_bytes"] or 0),
                    "text_status": clean_text(row["text_status"]),
                }
            )
    return rows


def classify_stage6_gap_row(row: Mapping[str, Any]) -> dict[str, str]:
    extension = clean_text(row.get("extension")).casefold()
    if extension in {".gif", ".heic"}:
        gap_class = "unsupported_image_format"
        recommended_action = "manual_review_future_image_parser"
        status_reason = "image_format_not_in_safe_ocr_pilot"
    elif extension in {".html", ".htm"}:
        gap_class = "unsafe_markup_format"
        recommended_action = "manual_review_html_parser"
        status_reason = "html_parser_not_enabled"
    elif extension == ".ics":
        gap_class = "calendar_invite_format"
        recommended_action = "manual_review_calendar_parser"
        status_reason = "calendar_parser_not_enabled"
    else:
        gap_class = "uncovered_parse_later_format"
        recommended_action = "manual_review"
        status_reason = "no_downstream_stage_for_extension"
    return {
        "gap_class": gap_class,
        "recommended_action": recommended_action,
        "status_reason": status_reason,
    }


def select_balanced_ocr_pilot_rows(
    rows: Sequence[Mapping[str, Any]],
    *,
    limit: int,
) -> list[Mapping[str, Any]]:
    if limit <= 0:
        return []
    grouped: dict[str, list[Mapping[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(clean_text(row["extension"]), []).append(row)
    for extension_rows in grouped.values():
        extension_rows.sort(
            key=lambda row: (
                int(row["size_bytes"]),
                clean_text(row["source_archive_id"]),
                clean_text(row["message_sha256"]),
                int(row["part_index"]),
                clean_text(row["attachment_sha256"]),
            )
        )
    selected: list[Mapping[str, Any]] = []
    offsets = {extension: 0 for extension in grouped}
    extensions = sorted(grouped)
    while len(selected) < limit:
        progressed = False
        for extension in extensions:
            offset = offsets[extension]
            extension_rows = grouped[extension]
            if offset >= len(extension_rows):
                continue
            selected.append(extension_rows[offset])
            offsets[extension] += 1
            progressed = True
            if len(selected) >= limit:
                break
        if not progressed:
            break
    return selected


def load_attachment_ocr_preflight_selected_rows(
    stage6_plan_db_path: Path,
    *,
    max_candidates: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with open_sqlite_readonly(stage6_plan_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != MAIL_ATTACHMENT_STAGE6_PLAN_SCHEMA_VERSION:
            raise ValueError(
                "mail attachment stage 6 plan schema mismatch: "
                f"expected {MAIL_ATTACHMENT_STAGE6_PLAN_SCHEMA_VERSION}, got {schema_version}"
            )
        table_row = con.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'stage6_ocr_pilot_plan'"
        ).fetchone()
        if table_row is None:
            raise ValueError("mail attachment stage 6 plan database missing table: stage6_ocr_pilot_plan")
        query = """
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes, pilot_rank
            FROM stage6_ocr_pilot_plan
            WHERE pilot_status = 'selected'
            ORDER BY pilot_rank, source_archive_id, message_sha256, part_index, attachment_sha256
        """
        if max_candidates > 0:
            query += " LIMIT ?"
            params: Sequence[Any] = (max_candidates,)
        else:
            params = ()
        for row in con.execute(query, params):
            rows.append(
                {
                    "source_archive_id": clean_text(row["source_archive_id"]),
                    "message_sha256": clean_text(row["message_sha256"]),
                    "part_index": int(row["part_index"]),
                    "attachment_sha256": clean_text(row["attachment_sha256"]),
                    "extension": clean_text(row["extension"]),
                    "declared_content_type": clean_text(row["declared_content_type"]),
                    "size_bytes": int(row["size_bytes"] or 0),
                    "pilot_rank": int(row["pilot_rank"]),
                }
            )
    return rows


def preflight_one_ocr_candidate(
    *,
    selected_row: Mapping[str, Any],
    archive_row: Optional[Mapping[str, Any]],
    archive_db_path_by_source_id: Mapping[str, Path],
    max_attachment_bytes: int,
    verify_sha256: bool,
) -> dict[str, Any]:
    result = {
        "source_archive_id": clean_text(selected_row["source_archive_id"]),
        "message_sha256": clean_text(selected_row["message_sha256"]),
        "part_index": int(selected_row["part_index"]),
        "attachment_sha256": clean_text(selected_row["attachment_sha256"]),
        "extension": clean_text(selected_row["extension"]),
        "declared_content_type": clean_text(selected_row["declared_content_type"]),
        "size_bytes": int(selected_row["size_bytes"] or 0),
        "pilot_rank": int(selected_row["pilot_rank"]),
        "preflight_status": "skipped",
        "status_reason": "",
        "sha256_verified": False,
        "bytes_read": 0,
    }
    if archive_row is None:
        return {**result, "status_reason": "missing_source_attachment_row"}
    source_size = int(archive_row["size_bytes"] or 0)
    if source_size != int(selected_row["size_bytes"] or 0):
        return {
            **result,
            "preflight_status": "blocked_safety",
            "status_reason": "metadata_size_mismatch",
            "size_bytes": source_size,
        }
    if source_size <= 0:
        return {
            **result,
            "preflight_status": "blocked_safety",
            "status_reason": "empty_attachment",
            "size_bytes": source_size,
        }
    if source_size > max(0, int(max_attachment_bytes)):
        return {
            **result,
            "preflight_status": "skipped",
            "status_reason": "preflight_size_limit_exceeded",
            "size_bytes": source_size,
        }
    archive_db_path = archive_db_path_by_source_id.get(clean_text(selected_row["source_archive_id"]))
    if archive_db_path is None:
        return {**result, "status_reason": "missing_source_archive", "size_bytes": source_size}
    attachment_path, path_status = resolve_safe_attachment_path(
        archive_row["path"],
        archive_db_path=archive_db_path,
        message_sha256=selected_row["message_sha256"],
        part_index=int(selected_row["part_index"]),
        attachment_sha256=selected_row["attachment_sha256"],
        expected_size=source_size,
    )
    if attachment_path is None:
        status = "blocked_safety" if path_status.startswith("unsafe_") else "skipped"
        return {
            **result,
            "preflight_status": status,
            "status_reason": path_status,
            "size_bytes": source_size,
        }
    if not verify_sha256:
        return {
            **result,
            "preflight_status": "verified",
            "status_reason": "metadata_path_verified",
            "size_bytes": source_size,
        }
    try:
        payload = attachment_path.read_bytes()
    except OSError:
        return {
            **result,
            "preflight_status": "skipped",
            "status_reason": "attachment_unreadable",
            "size_bytes": source_size,
        }
    actual_sha256 = hashlib.sha256(payload).hexdigest()
    if actual_sha256 != clean_text(selected_row["attachment_sha256"]):
        return {
            **result,
            "preflight_status": "blocked_safety",
            "status_reason": "attachment_sha256_mismatch",
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }
    return {
        **result,
        "preflight_status": "verified",
        "status_reason": "sha256_verified",
        "sha256_verified": True,
        "size_bytes": len(payload),
        "bytes_read": len(payload),
    }


def load_attachment_ocr_pilot_verified_rows(
    ocr_preflight_db_path: Path,
    *,
    max_candidates: int,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with open_sqlite_readonly(ocr_preflight_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != MAIL_ATTACHMENT_OCR_PREFLIGHT_SCHEMA_VERSION:
            raise ValueError(
                "mail attachment OCR preflight schema mismatch: "
                f"expected {MAIL_ATTACHMENT_OCR_PREFLIGHT_SCHEMA_VERSION}, got {schema_version}"
            )
        table_row = con.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'attachment_ocr_preflight'"
        ).fetchone()
        if table_row is None:
            raise ValueError("mail attachment OCR preflight database missing table: attachment_ocr_preflight")
        query = """
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes, pilot_rank
            FROM attachment_ocr_preflight
            WHERE preflight_status = 'verified' AND sha256_verified = 1
            ORDER BY pilot_rank, source_archive_id, message_sha256, part_index, attachment_sha256
        """
        if max_candidates > 0:
            query += " LIMIT ?"
            params: Sequence[Any] = (max_candidates,)
        else:
            params = ()
        for row in con.execute(query, params):
            rows.append(
                {
                    "source_archive_id": clean_text(row["source_archive_id"]),
                    "message_sha256": clean_text(row["message_sha256"]),
                    "part_index": int(row["part_index"]),
                    "attachment_sha256": clean_text(row["attachment_sha256"]),
                    "extension": clean_text(row["extension"]),
                    "declared_content_type": clean_text(row["declared_content_type"]),
                    "size_bytes": int(row["size_bytes"] or 0),
                    "pilot_rank": int(row["pilot_rank"]),
                }
            )
    return rows


def run_one_attachment_ocr_pilot(
    *,
    verified_row: Mapping[str, Any],
    archive_row: Optional[Mapping[str, Any]],
    archive_db_path_by_source_id: Mapping[str, Path],
    derived_text_dir: Path,
    languages: str,
    page_segmentation_mode: int,
    timeout_seconds: int,
    max_attachment_bytes: int,
    max_text_chars: int,
    tesseract_thread_limit: int,
    reuse_existing_ocr_text: bool,
) -> dict[str, Any]:
    result = {
        "source_archive_id": clean_text(verified_row["source_archive_id"]),
        "message_sha256": clean_text(verified_row["message_sha256"]),
        "part_index": int(verified_row["part_index"]),
        "attachment_sha256": clean_text(verified_row["attachment_sha256"]),
        "extension": clean_text(verified_row["extension"]),
        "declared_content_type": clean_text(verified_row["declared_content_type"]),
        "size_bytes": int(verified_row["size_bytes"] or 0),
        "pilot_rank": int(verified_row["pilot_rank"]),
        "status": "skipped",
        "status_reason": "",
        "warnings": [],
        "text_sha256": "",
        "text_chars": 0,
        "text_truncated": False,
        "derived_text_path_sha256": "",
        "derived_text_written": False,
        "derived_text_reused": False,
        "ocr_seconds": 0.0,
        "ocr_attempted": False,
    }
    if archive_row is None:
        return {**result, "status_reason": "missing_source_attachment_row"}
    source_size = int(archive_row["size_bytes"] or 0)
    if source_size != int(verified_row["size_bytes"] or 0):
        return {
            **result,
            "status": "blocked_safety",
            "status_reason": "metadata_size_mismatch",
            "size_bytes": source_size,
        }
    if source_size <= 0:
        return {
            **result,
            "status": "blocked_safety",
            "status_reason": "empty_attachment",
            "size_bytes": source_size,
        }
    if source_size > max(0, int(max_attachment_bytes)):
        return {
            **result,
            "status_reason": "ocr_size_limit_exceeded",
            "size_bytes": source_size,
        }
    archive_db_path = archive_db_path_by_source_id.get(clean_text(verified_row["source_archive_id"]))
    if archive_db_path is None:
        return {**result, "status_reason": "missing_source_archive", "size_bytes": source_size}
    attachment_path, path_status = resolve_safe_attachment_path(
        archive_row["path"],
        archive_db_path=archive_db_path,
        message_sha256=verified_row["message_sha256"],
        part_index=int(verified_row["part_index"]),
        attachment_sha256=verified_row["attachment_sha256"],
        expected_size=source_size,
    )
    if attachment_path is None:
        status = "blocked_safety" if path_status.startswith("unsafe_") else "skipped"
        return {
            **result,
            "status": status,
            "status_reason": path_status,
            "size_bytes": source_size,
        }

    try:
        actual_sha256 = sha256_file(attachment_path)
    except OSError:
        return {
            **result,
            "status": "skipped",
            "status_reason": "attachment_unreadable",
            "size_bytes": source_size,
        }
    if actual_sha256 != clean_text(verified_row["attachment_sha256"]):
        return {
            **result,
            "status": "blocked_safety",
            "status_reason": "attachment_sha256_mismatch",
            "size_bytes": source_size,
        }

    derived_path = derived_attachment_text_path(
        derived_text_dir=derived_text_dir,
        source_archive_id=verified_row["source_archive_id"],
        message_sha256=verified_row["message_sha256"],
        part_index=int(verified_row["part_index"]),
        attachment_sha256=verified_row["attachment_sha256"],
    )
    if reuse_existing_ocr_text and derived_path.exists():
        try:
            text = derived_path.read_text(encoding="utf-8")
        except OSError:
            return {
                **result,
                "status": "skipped",
                "status_reason": "existing_ocr_text_unreadable",
                "size_bytes": source_size,
                "derived_text_path_sha256": hashlib.sha256(str(derived_path).encode("utf-8")).hexdigest(),
            }
        text = text[:max(0, int(max_text_chars))]
        text_sha256 = hashlib.sha256(text.encode("utf-8")).hexdigest()
        return {
            **result,
            "status": "extracted" if text.strip() else "empty_text",
            "status_reason": "existing_ocr_text_reused" if text.strip() else "existing_ocr_text_empty",
            "text_sha256": text_sha256 if text.strip() else "",
            "text_chars": len(text) if text.strip() else 0,
            "text_truncated": False,
            "derived_text_path_sha256": hashlib.sha256(str(derived_path).encode("utf-8")).hexdigest(),
            "derived_text_reused": bool(text.strip()),
            "size_bytes": source_size,
        }

    started = time.monotonic()
    try:
        ocr = run_tesseract_ocr(
            attachment_path,
            languages=languages,
            page_segmentation_mode=page_segmentation_mode,
            timeout_seconds=timeout_seconds,
            tesseract_thread_limit=tesseract_thread_limit,
        )
    except TimeoutError:
        return {
            **result,
            "status": "ocr_error",
            "status_reason": "tesseract_timeout",
            "size_bytes": source_size,
            "ocr_seconds": time.monotonic() - started,
            "ocr_attempted": True,
        }
    except FileNotFoundError:
        return {
            **result,
            "status": "ocr_error",
            "status_reason": "tesseract_not_found",
            "size_bytes": source_size,
            "ocr_seconds": time.monotonic() - started,
            "ocr_attempted": True,
        }
    except subprocess.SubprocessError:
        return {
            **result,
            "status": "ocr_error",
            "status_reason": "tesseract_subprocess_error",
            "size_bytes": source_size,
            "ocr_seconds": time.monotonic() - started,
            "ocr_attempted": True,
        }
    ocr_seconds = time.monotonic() - started
    if int(ocr["returncode"]) != 0:
        return {
            **result,
            "status": "ocr_error",
            "status_reason": "tesseract_failed",
            "warnings": ["tesseract_stderr_present"] if ocr["stderr_present"] else [],
            "size_bytes": source_size,
            "ocr_seconds": ocr_seconds,
            "ocr_attempted": True,
        }

    text = normalize_extracted_attachment_text(ocr["text"])
    text_truncated = len(text) > max_text_chars
    if text_truncated:
        text = text[:max_text_chars]
    warnings = sorted(
        set(
            [
                *(["tesseract_stderr_present"] if ocr["stderr_present"] else []),
                *(["text_truncated"] if text_truncated else []),
            ]
        )
    )
    if not text.strip():
        return {
            **result,
            "status": "empty_text",
            "status_reason": "no_text_extracted",
            "warnings": warnings,
            "size_bytes": source_size,
            "ocr_seconds": ocr_seconds,
            "ocr_attempted": True,
        }

    text_sha256 = hashlib.sha256(text.encode("utf-8")).hexdigest()
    derived_text_written = write_text_once(derived_path, text)
    return {
        **result,
        "status": "extracted",
        "status_reason": "ok",
        "warnings": warnings,
        "text_sha256": text_sha256,
        "text_chars": len(text),
        "text_truncated": text_truncated,
        "derived_text_path_sha256": hashlib.sha256(str(derived_path).encode("utf-8")).hexdigest(),
        "derived_text_written": derived_text_written,
        "size_bytes": source_size,
        "ocr_seconds": ocr_seconds,
        "ocr_attempted": True,
    }


def run_tesseract_ocr(
    image_path: Path,
    *,
    languages: str,
    page_segmentation_mode: int,
    timeout_seconds: int,
    tesseract_thread_limit: int,
) -> dict[str, Any]:
    env = os.environ.copy()
    safe_thread_limit = str(max(1, int(tesseract_thread_limit)))
    env["OMP_THREAD_LIMIT"] = safe_thread_limit
    env["OMP_NUM_THREADS"] = safe_thread_limit
    try:
        completed = subprocess.run(
            [
                "tesseract",
                str(image_path),
                "stdout",
                "-l",
                normalize_tesseract_languages(languages),
                "--psm",
                str(int(page_segmentation_mode)),
            ],
            check=False,
            capture_output=True,
            text=True,
            timeout=max(1, int(timeout_seconds)),
            env=env,
        )
    except subprocess.TimeoutExpired:
        raise TimeoutError("tesseract_timeout") from None
    return {
        "returncode": completed.returncode,
        "text": completed.stdout or "",
        "stderr_present": bool(clean_text(completed.stderr)),
    }


def normalize_tesseract_languages(value: object) -> str:
    text = clean_text(value)
    parts = [part for part in re.split(r"[+,;\s]+", text) if part]
    safe_parts = [
        part
        for part in parts
        if re.fullmatch(r"[A-Za-z0-9_]+", part)
    ]
    return "+".join(unique_preserving_order(safe_parts))


def load_attachment_text_index_rows(
    source_db_path: Path,
    *,
    source_stage: str,
) -> tuple[list[dict[str, Any]], str]:
    source_stage = clean_text(source_stage)
    if source_stage == "text_extract":
        expected_schema = MAIL_ATTACHMENT_TEXT_EXTRACT_SCHEMA_VERSION
        required_table = "attachment_text_extracts"
        query = """
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes, parser,
              status AS source_status, status_reason AS source_status_reason,
              warnings_json, text_sha256, text_chars, text_truncated,
              derived_text_path, 0 AS page_count, 0 AS pages_processed
            FROM attachment_text_extracts
            ORDER BY source_archive_id, message_sha256, part_index, attachment_sha256
        """
    elif source_stage == "pdf_extract":
        expected_schema = MAIL_ATTACHMENT_PDF_EXTRACT_SCHEMA_VERSION
        required_table = "attachment_pdf_extracts"
        query = """
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes, parser,
              status AS source_status, status_reason AS source_status_reason,
              warnings_json, text_sha256, text_chars, text_truncated,
              derived_text_path, page_count, pages_processed
            FROM attachment_pdf_extracts
            ORDER BY source_archive_id, message_sha256, part_index, attachment_sha256
        """
    elif source_stage == "image_ocr_plan":
        expected_schema = MAIL_ATTACHMENT_IMAGE_OCR_PLAN_SCHEMA_VERSION
        required_table = "attachment_image_ocr_plan"
        query = """
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes, '' AS parser,
              status AS source_status, status_reason AS source_status_reason,
              warnings_json, '' AS text_sha256, 0 AS text_chars, 0 AS text_truncated,
              '' AS derived_text_path, 0 AS page_count, 0 AS pages_processed,
              ocr_status
            FROM attachment_image_ocr_plan
            ORDER BY source_archive_id, message_sha256, part_index, attachment_sha256
        """
    else:
        raise ValueError(f"unsupported attachment text index source stage: {source_stage}")

    rows: list[dict[str, Any]] = []
    with open_sqlite_readonly(source_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != expected_schema:
            raise ValueError(
                f"mail attachment {source_stage} schema mismatch: "
                f"expected {expected_schema}, got {schema_version}"
            )
        table_row = con.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = ?",
            (required_table,),
        ).fetchone()
        if table_row is None:
            raise ValueError(f"mail attachment {source_stage} database missing table: {required_table}")
        for row in con.execute(query):
            rows.append(
                {
                    "source_stage": source_stage,
                    "source_archive_id": clean_text(row["source_archive_id"]),
                    "message_sha256": clean_text(row["message_sha256"]),
                    "part_index": int(row["part_index"]),
                    "attachment_sha256": clean_text(row["attachment_sha256"]),
                    "extension": clean_text(row["extension"]),
                    "declared_content_type": clean_text(row["declared_content_type"]),
                    "size_bytes": int(row["size_bytes"] or 0),
                    "parser": clean_text(row["parser"]),
                    "source_status": clean_text(row["source_status"]),
                    "source_status_reason": clean_text(row["source_status_reason"]),
                    "warnings_json": clean_text(row["warnings_json"]),
                    "text_sha256": clean_text(row["text_sha256"]),
                    "text_chars": int(row["text_chars"] or 0),
                    "text_truncated": int(row["text_truncated"] or 0),
                    "derived_text_path": clean_text(row["derived_text_path"]),
                    "page_count": int(row["page_count"] or 0),
                    "pages_processed": int(row["pages_processed"] or 0),
                    "ocr_status": clean_text(row["ocr_status"]) if "ocr_status" in row.keys() else "",
                }
            )
    return rows, schema_version


def attachment_text_index_status(
    source_stage: str,
    source_status: str,
    row: Mapping[str, Any],
) -> str:
    source_stage = clean_text(source_stage)
    source_status = clean_text(source_status)
    if source_stage == "image_ocr_plan":
        if source_status == "planned" and clean_text(row.get("ocr_status")) == "disabled":
            return "ocr_pending"
        if source_status == "blocked_safety":
            return "blocked"
        return source_status or "unknown"

    if source_status == "extracted":
        if clean_text(row.get("text_sha256")) and int(row.get("text_chars") or 0) > 0:
            return "available"
        return "invalid_extracted_text_metadata"
    if source_status == "empty_text":
        return "empty_text"
    if source_status == "blocked_safety":
        return "blocked"
    return source_status or "unknown"


def parse_json_list(value: object) -> list[str]:
    try:
        parsed = json.loads(clean_text(value) or "[]")
    except Exception:  # noqa: BLE001
        return ["invalid_json_list"]
    if not isinstance(parsed, list):
        return ["invalid_json_list"]
    return [clean_text(item) for item in parsed if clean_text(item)]


def load_attachment_text_extract_plan_rows(
    *,
    parse_plan_db_path: Path,
    source_archive_ids: set[str],
    stage_extensions: set[str],
    max_attachments: int,
) -> tuple[list[dict[str, Any]], int, int]:
    rows: list[dict[str, Any]] = []
    with open_sqlite_readonly(parse_plan_db_path) as con:
        schema_row = con.execute("SELECT value FROM meta WHERE key = 'schema_version'").fetchone()
        schema_version = clean_text(schema_row["value"] if schema_row else "")
        if schema_version != MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION:
            raise ValueError(
                "mail attachment parse plan schema mismatch: "
                f"expected {MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION}, got {schema_version}"
            )
        total_parse_later = int(
            con.execute(
                "SELECT COUNT(*) FROM attachment_parse_plan WHERE action = 'parse_later'"
            ).fetchone()[0]
        )
        placeholders = ", ".join("?" for _ in stage_extensions)
        query = f"""
            SELECT
              source_archive_id, message_sha256, part_index, attachment_sha256,
              extension, declared_content_type, size_bytes
            FROM attachment_parse_plan
            WHERE action = 'parse_later'
              AND extension IN ({placeholders})
            ORDER BY source_archive_id, message_sha256, part_index, attachment_sha256
        """
        params: list[Any] = sorted(stage_extensions)
        if max_attachments > 0:
            query += " LIMIT ?"
            params.append(max_attachments)
        for row in con.execute(query, params):
            if clean_text(row["source_archive_id"]) not in source_archive_ids:
                continue
            rows.append(
                {
                    "source_archive_id": clean_text(row["source_archive_id"]),
                    "message_sha256": clean_text(row["message_sha256"]),
                    "part_index": int(row["part_index"]),
                    "attachment_sha256": clean_text(row["attachment_sha256"]),
                    "extension": clean_text(row["extension"]),
                    "declared_content_type": clean_text(row["declared_content_type"]),
                    "size_bytes": int(row["size_bytes"] or 0),
                }
            )
    return rows, total_parse_later, max(0, total_parse_later - len(rows))


def load_attachment_text_extract_source_rows(source_paths: Sequence[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source_path in source_paths:
        resolved_source_path = Path(source_path).resolve(strict=False)
        source_archive_id = mail_archive_source_id(resolved_source_path)
        with open_sqlite_readonly(resolved_source_path) as con:
            for row in con.execute(
                """
                SELECT
                  message_sha256,
                  part_index,
                  content_type,
                  size_bytes,
                  sha256 AS attachment_sha256,
                  path
                FROM attachments
                ORDER BY message_sha256, part_index, sha256
                """
            ):
                rows.append(
                    {
                        "source_archive_id": source_archive_id,
                        "source_archive_path": str(resolved_source_path),
                        "message_sha256": clean_text(row["message_sha256"]),
                        "part_index": int(row["part_index"]),
                        "content_type": clean_text(row["content_type"]),
                        "size_bytes": int(row["size_bytes"] or 0),
                        "attachment_sha256": clean_text(row["attachment_sha256"]),
                        "path": clean_text(row["path"]),
                    }
                )
    return rows


def extract_one_attachment_text(
    *,
    plan_row: Mapping[str, Any],
    archive_row: Optional[Mapping[str, Any]],
    archive_db_path_by_source_id: Mapping[str, Path],
    derived_text_dir: Path,
    config: MailAttachmentTextExtractConfig,
) -> dict[str, Any]:
    base_result = {
        "extension": clean_text(plan_row["extension"]),
        "declared_content_type": clean_text(plan_row["declared_content_type"]),
        "size_bytes": int(plan_row["size_bytes"] or 0),
        "parser": "",
        "status": "skipped",
        "status_reason": "",
        "warnings": [],
        "text_sha256": "",
        "text_chars": 0,
        "text_truncated": False,
        "derived_text_path": "",
        "derived_text_written": False,
        "bytes_read": 0,
    }
    if archive_row is None:
        return {**base_result, "status_reason": "missing_source_attachment_row"}

    archive_db_path = archive_db_path_by_source_id.get(clean_text(plan_row["source_archive_id"]))
    if archive_db_path is None:
        return {**base_result, "status_reason": "missing_source_archive"}

    source_size = int(archive_row["size_bytes"] or 0)
    if source_size != int(plan_row["size_bytes"] or 0):
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": "metadata_size_mismatch",
            "size_bytes": source_size,
        }
    if source_size <= 0:
        return {**base_result, "status": "blocked_safety", "status_reason": "empty_attachment"}
    if source_size > max(0, int(config.max_attachment_bytes)):
        return {**base_result, "status_reason": "stage_size_limit_exceeded", "size_bytes": source_size}

    attachment_path, path_status = resolve_safe_attachment_path(
        archive_row["path"],
        archive_db_path=archive_db_path,
        message_sha256=plan_row["message_sha256"],
        part_index=int(plan_row["part_index"]),
        attachment_sha256=plan_row["attachment_sha256"],
        expected_size=source_size,
    )
    if attachment_path is None:
        status = "blocked_safety" if path_status.startswith("unsafe_") else "skipped"
        return {**base_result, "status": status, "status_reason": path_status, "size_bytes": source_size}

    try:
        payload = attachment_path.read_bytes()
    except OSError:
        return {**base_result, "status_reason": "attachment_unreadable", "size_bytes": source_size}
    actual_sha256 = hashlib.sha256(payload).hexdigest()
    if actual_sha256 != clean_text(plan_row["attachment_sha256"]):
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": "attachment_sha256_mismatch",
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }

    try:
        parsed = parse_attachment_text_payload(
            payload=payload,
            extension=clean_text(plan_row["extension"]),
            config=config,
        )
    except AttachmentTextBlockedError as exc:
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": exc.reason,
            "warnings": sorted(set(exc.warnings)),
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }
    except AttachmentTextParseError as exc:
        return {
            **base_result,
            "status": "parse_error",
            "status_reason": exc.reason,
            "warnings": sorted(set(exc.warnings)),
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }

    text = normalize_extracted_attachment_text(parsed["text"])
    max_text_chars = max(0, int(config.max_text_chars_per_attachment))
    text_truncated = len(text) > max_text_chars
    if text_truncated:
        text = text[:max_text_chars]
    warnings = sorted(set([*parsed["warnings"], *(["text_truncated"] if text_truncated else [])]))
    if not text.strip():
        return {
            **base_result,
            "parser": parsed["parser"],
            "status": "empty_text",
            "status_reason": "no_text_extracted",
            "warnings": warnings,
            "size_bytes": len(payload),
            "bytes_read": len(payload),
            "text_truncated": text_truncated,
        }

    text_sha256 = hashlib.sha256(text.encode("utf-8")).hexdigest()
    derived_path = derived_attachment_text_path(
        derived_text_dir=derived_text_dir,
        source_archive_id=plan_row["source_archive_id"],
        message_sha256=plan_row["message_sha256"],
        part_index=int(plan_row["part_index"]),
        attachment_sha256=plan_row["attachment_sha256"],
    )
    derived_payload = "\n".join(
        [
            f"source_archive_id={clean_text(plan_row['source_archive_id'])}",
            f"message_sha256={clean_text(plan_row['message_sha256'])}",
            f"part_index={int(plan_row['part_index'])}",
            f"attachment_sha256={clean_text(plan_row['attachment_sha256'])}",
            f"extension={clean_text(plan_row['extension'])}",
            "",
            text,
        ]
    )
    derived_path.parent.mkdir(parents=True, exist_ok=True)
    previous = derived_path.read_text(encoding="utf-8", errors="replace") if derived_path.exists() else None
    derived_text_written = previous != derived_payload
    if derived_text_written:
        derived_path.write_text(derived_payload, encoding="utf-8")

    return {
        **base_result,
        "parser": parsed["parser"],
        "status": "extracted",
        "status_reason": "ok",
        "warnings": warnings,
        "text_sha256": text_sha256,
        "text_chars": len(text),
        "text_truncated": text_truncated,
        "derived_text_path": str(derived_path),
        "derived_text_written": derived_text_written,
        "size_bytes": len(payload),
        "bytes_read": len(payload),
    }


def extract_one_attachment_pdf_text(
    *,
    plan_row: Mapping[str, Any],
    archive_row: Optional[Mapping[str, Any]],
    archive_db_path_by_source_id: Mapping[str, Path],
    derived_text_dir: Path,
    config: MailAttachmentPdfExtractConfig,
) -> dict[str, Any]:
    base_result = {
        "extension": clean_text(plan_row["extension"]),
        "declared_content_type": clean_text(plan_row["declared_content_type"]),
        "size_bytes": int(plan_row["size_bytes"] or 0),
        "parser": "",
        "status": "skipped",
        "status_reason": "",
        "warnings": [],
        "page_count": 0,
        "pages_processed": 0,
        "text_sha256": "",
        "text_chars": 0,
        "text_truncated": False,
        "derived_text_path": "",
        "derived_text_written": False,
        "bytes_read": 0,
    }
    if archive_row is None:
        return {**base_result, "status_reason": "missing_source_attachment_row"}

    archive_db_path = archive_db_path_by_source_id.get(clean_text(plan_row["source_archive_id"]))
    if archive_db_path is None:
        return {**base_result, "status_reason": "missing_source_archive"}

    source_size = int(archive_row["size_bytes"] or 0)
    if source_size != int(plan_row["size_bytes"] or 0):
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": "metadata_size_mismatch",
            "size_bytes": source_size,
        }
    if source_size <= 0:
        return {**base_result, "status": "blocked_safety", "status_reason": "empty_attachment"}
    if source_size > max(0, int(config.max_attachment_bytes)):
        return {**base_result, "status_reason": "stage_size_limit_exceeded", "size_bytes": source_size}

    attachment_path, path_status = resolve_safe_attachment_path(
        archive_row["path"],
        archive_db_path=archive_db_path,
        message_sha256=plan_row["message_sha256"],
        part_index=int(plan_row["part_index"]),
        attachment_sha256=plan_row["attachment_sha256"],
        expected_size=source_size,
    )
    if attachment_path is None:
        status = "blocked_safety" if path_status.startswith("unsafe_") else "skipped"
        return {**base_result, "status": status, "status_reason": path_status, "size_bytes": source_size}

    try:
        payload = attachment_path.read_bytes()
    except OSError:
        return {**base_result, "status_reason": "attachment_unreadable", "size_bytes": source_size}
    actual_sha256 = hashlib.sha256(payload).hexdigest()
    if actual_sha256 != clean_text(plan_row["attachment_sha256"]):
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": "attachment_sha256_mismatch",
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }

    try:
        with pdf_parse_timeout(max(0, int(config.pdf_timeout_seconds))):
            parsed = parse_pdf_attachment(payload=payload, config=config)
    except TimeoutError:
        return {
            **base_result,
            "status": "parse_error",
            "status_reason": "pdf_timeout",
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }
    except AttachmentTextBlockedError as exc:
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": exc.reason,
            "warnings": sorted(set(exc.warnings)),
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }
    except AttachmentTextParseError as exc:
        return {
            **base_result,
            "status": "parse_error",
            "status_reason": exc.reason,
            "warnings": sorted(set(exc.warnings)),
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }

    text = normalize_extracted_attachment_text(parsed["text"])
    max_text_chars = max(0, int(config.max_text_chars_per_attachment))
    text_truncated = len(text) > max_text_chars
    if text_truncated:
        text = text[:max_text_chars]
    warnings = sorted(set([*parsed["warnings"], *(["text_truncated"] if text_truncated else [])]))
    if not text.strip():
        return {
            **base_result,
            "parser": parsed["parser"],
            "status": "empty_text",
            "status_reason": "no_text_extracted",
            "warnings": warnings,
            "page_count": int(parsed["page_count"]),
            "pages_processed": int(parsed["pages_processed"]),
            "size_bytes": len(payload),
            "bytes_read": len(payload),
            "text_truncated": text_truncated,
        }

    text_sha256 = hashlib.sha256(text.encode("utf-8")).hexdigest()
    derived_path = derived_attachment_text_path(
        derived_text_dir=derived_text_dir,
        source_archive_id=plan_row["source_archive_id"],
        message_sha256=plan_row["message_sha256"],
        part_index=int(plan_row["part_index"]),
        attachment_sha256=plan_row["attachment_sha256"],
    )
    derived_payload = "\n".join(
        [
            f"source_archive_id={clean_text(plan_row['source_archive_id'])}",
            f"message_sha256={clean_text(plan_row['message_sha256'])}",
            f"part_index={int(plan_row['part_index'])}",
            f"attachment_sha256={clean_text(plan_row['attachment_sha256'])}",
            f"extension={clean_text(plan_row['extension'])}",
            f"page_count={int(parsed['page_count'])}",
            f"pages_processed={int(parsed['pages_processed'])}",
            "",
            text,
        ]
    )
    derived_path.parent.mkdir(parents=True, exist_ok=True)
    previous = derived_path.read_text(encoding="utf-8", errors="replace") if derived_path.exists() else None
    derived_text_written = previous != derived_payload
    if derived_text_written:
        derived_path.write_text(derived_payload, encoding="utf-8")

    return {
        **base_result,
        "parser": parsed["parser"],
        "status": "extracted",
        "status_reason": "ok",
        "warnings": warnings,
        "page_count": int(parsed["page_count"]),
        "pages_processed": int(parsed["pages_processed"]),
        "text_sha256": text_sha256,
        "text_chars": len(text),
        "text_truncated": text_truncated,
        "derived_text_path": str(derived_path),
        "derived_text_written": derived_text_written,
        "size_bytes": len(payload),
        "bytes_read": len(payload),
    }


def inspect_one_attachment_image_for_ocr_plan(
    *,
    plan_row: Mapping[str, Any],
    archive_row: Optional[Mapping[str, Any]],
    archive_db_path_by_source_id: Mapping[str, Path],
    config: MailAttachmentImageOcrPlanConfig,
) -> dict[str, Any]:
    base_result = {
        "extension": clean_text(plan_row["extension"]),
        "declared_content_type": clean_text(plan_row["declared_content_type"]),
        "size_bytes": int(plan_row["size_bytes"] or 0),
        "image_format": "",
        "width": 0,
        "height": 0,
        "pixel_count": 0,
        "status": "skipped",
        "status_reason": "",
        "ocr_status": "disabled",
        "warnings": [],
        "bytes_read": 0,
    }
    if archive_row is None:
        return {**base_result, "status_reason": "missing_source_attachment_row"}

    archive_db_path = archive_db_path_by_source_id.get(clean_text(plan_row["source_archive_id"]))
    if archive_db_path is None:
        return {**base_result, "status_reason": "missing_source_archive"}

    source_size = int(archive_row["size_bytes"] or 0)
    if source_size != int(plan_row["size_bytes"] or 0):
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": "metadata_size_mismatch",
            "size_bytes": source_size,
        }
    if source_size <= 0:
        return {**base_result, "status": "blocked_safety", "status_reason": "empty_attachment"}
    if source_size > max(0, int(config.max_attachment_bytes)):
        return {**base_result, "status_reason": "stage_size_limit_exceeded", "size_bytes": source_size}
    if not config.inspect_headers:
        return {
            **base_result,
            "image_format": "not_inspected",
            "status": "planned",
            "status_reason": "ocr_disabled",
            "size_bytes": source_size,
        }

    attachment_path, path_status = resolve_safe_attachment_path(
        archive_row["path"],
        archive_db_path=archive_db_path,
        message_sha256=plan_row["message_sha256"],
        part_index=int(plan_row["part_index"]),
        attachment_sha256=plan_row["attachment_sha256"],
        expected_size=source_size,
    )
    if attachment_path is None:
        status = "blocked_safety" if path_status.startswith("unsafe_") else "skipped"
        return {**base_result, "status": status, "status_reason": path_status, "size_bytes": source_size}

    try:
        payload = attachment_path.read_bytes()
    except OSError:
        return {**base_result, "status_reason": "attachment_unreadable", "size_bytes": source_size}
    actual_sha256 = hashlib.sha256(payload).hexdigest()
    if actual_sha256 != clean_text(plan_row["attachment_sha256"]):
        return {
            **base_result,
            "status": "blocked_safety",
            "status_reason": "attachment_sha256_mismatch",
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }

    try:
        image_info = inspect_image_header(payload)
    except AttachmentTextParseError as exc:
        return {
            **base_result,
            "status": "parse_error",
            "status_reason": exc.reason,
            "warnings": sorted(set(exc.warnings)),
            "size_bytes": len(payload),
            "bytes_read": len(payload),
        }
    width = int(image_info["width"])
    height = int(image_info["height"])
    pixel_count = width * height
    warnings = sorted(set(image_info["warnings"]))
    if width <= 0 or height <= 0:
        status = "parse_error"
        reason = "invalid_image_dimensions"
    elif width > max(0, int(config.max_image_dimension)) or height > max(0, int(config.max_image_dimension)):
        status = "blocked_safety"
        reason = "image_dimension_limit_exceeded"
    elif pixel_count > max(0, int(config.max_image_pixels)):
        status = "blocked_safety"
        reason = "image_pixel_limit_exceeded"
    else:
        status = "planned"
        reason = "ocr_disabled"

    return {
        **base_result,
        "image_format": clean_text(image_info["image_format"]),
        "width": width,
        "height": height,
        "pixel_count": pixel_count,
        "status": status,
        "status_reason": reason,
        "warnings": warnings,
        "size_bytes": len(payload),
        "bytes_read": len(payload),
    }


def inspect_image_header(payload: bytes) -> dict[str, Any]:
    if payload.startswith(b"\x89PNG\r\n\x1a\n"):
        return inspect_png_header(payload)
    if payload.startswith(b"\xff\xd8"):
        return inspect_jpeg_header(payload)
    if payload.startswith(b"RIFF") and payload[8:12] == b"WEBP":
        return inspect_webp_header(payload)
    raise AttachmentTextParseError("invalid_image_header")


def inspect_png_header(payload: bytes) -> dict[str, Any]:
    if len(payload) < 33:
        raise AttachmentTextParseError("invalid_png_header")
    if payload[12:16] != b"IHDR":
        raise AttachmentTextParseError("invalid_png_ihdr")
    width = int.from_bytes(payload[16:20], "big")
    height = int.from_bytes(payload[20:24], "big")
    bit_depth = payload[24]
    color_type = payload[25]
    warnings: list[str] = []
    if b"eXIf" in payload[:4096]:
        warnings.append("image_exif_declared_not_read")
    if bit_depth not in {1, 2, 4, 8, 16}:
        warnings.append("png_unusual_bit_depth")
    return {
        "image_format": "png",
        "width": width,
        "height": height,
        "warnings": [*warnings, f"png_color_type_{color_type}"],
    }


def inspect_jpeg_header(payload: bytes) -> dict[str, Any]:
    warnings: list[str] = []
    offset = 2
    while offset + 4 <= len(payload):
        if payload[offset] != 0xFF:
            offset += 1
            continue
        marker = payload[offset + 1]
        offset += 2
        if marker in {0xD8, 0xD9}:
            continue
        if marker == 0xDA:
            break
        if offset + 2 > len(payload):
            break
        segment_length = int.from_bytes(payload[offset : offset + 2], "big")
        if segment_length < 2:
            raise AttachmentTextParseError("invalid_jpeg_segment")
        segment_start = offset + 2
        segment_end = offset + segment_length
        if segment_end > len(payload):
            raise AttachmentTextParseError("truncated_jpeg_segment")
        if marker == 0xE1:
            warnings.append("image_exif_declared_not_read")
        if marker in {
            0xC0,
            0xC1,
            0xC2,
            0xC3,
            0xC5,
            0xC6,
            0xC7,
            0xC9,
            0xCA,
            0xCB,
            0xCD,
            0xCE,
            0xCF,
        }:
            if segment_start + 5 >= segment_end:
                raise AttachmentTextParseError("invalid_jpeg_sof")
            height = int.from_bytes(payload[segment_start + 1 : segment_start + 3], "big")
            width = int.from_bytes(payload[segment_start + 3 : segment_start + 5], "big")
            return {
                "image_format": "jpeg",
                "width": width,
                "height": height,
                "warnings": warnings,
            }
        offset = segment_end
    raise AttachmentTextParseError("jpeg_dimensions_not_found", warnings)


def inspect_webp_header(payload: bytes) -> dict[str, Any]:
    if len(payload) < 30:
        raise AttachmentTextParseError("invalid_webp_header")
    chunk_type = payload[12:16]
    if chunk_type == b"VP8X":
        if len(payload) < 30:
            raise AttachmentTextParseError("invalid_webp_vp8x")
        width = int.from_bytes(payload[24:27], "little") + 1
        height = int.from_bytes(payload[27:30], "little") + 1
        warnings = []
        flags = payload[20]
        if flags & 0b00001000:
            warnings.append("image_exif_declared_not_read")
        return {"image_format": "webp", "width": width, "height": height, "warnings": warnings}
    if chunk_type == b"VP8 ":
        if len(payload) < 30:
            raise AttachmentTextParseError("invalid_webp_vp8")
        width = int.from_bytes(payload[26:28], "little") & 0x3FFF
        height = int.from_bytes(payload[28:30], "little") & 0x3FFF
        return {"image_format": "webp", "width": width, "height": height, "warnings": []}
    if chunk_type == b"VP8L":
        if len(payload) < 25:
            raise AttachmentTextParseError("invalid_webp_vp8l")
        bits = int.from_bytes(payload[21:25], "little")
        width = (bits & 0x3FFF) + 1
        height = ((bits >> 14) & 0x3FFF) + 1
        return {"image_format": "webp", "width": width, "height": height, "warnings": []}
    raise AttachmentTextParseError("unsupported_webp_chunk")


class AttachmentTextParseError(Exception):
    def __init__(self, reason: str, warnings: Optional[Sequence[str]] = None) -> None:
        super().__init__(reason)
        self.reason = reason
        self.warnings = tuple(warnings or ())


class AttachmentTextBlockedError(AttachmentTextParseError):
    pass


@contextmanager
def pdf_parse_timeout(seconds: int) -> Any:
    if seconds <= 0 or threading.current_thread() is not threading.main_thread():
        yield
        return
    previous_handler = signal.getsignal(signal.SIGALRM)

    def _raise_timeout(_signum: int, _frame: object) -> None:
        raise TimeoutError("pdf_timeout")

    try:
        signal.signal(signal.SIGALRM, _raise_timeout)
        signal.setitimer(signal.ITIMER_REAL, float(seconds))
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous_handler)


def resolve_safe_attachment_path(
    value: object,
    *,
    archive_db_path: Path,
    message_sha256: object,
    part_index: int,
    attachment_sha256: object,
    expected_size: int,
) -> tuple[Optional[Path], str]:
    raw_path = clean_text(value)
    if not raw_path:
        return None, "missing_attachment_path"
    path = Path(raw_path).resolve(strict=False)
    if "_external_handoffs" not in path.parts:
        return None, "unsafe_path_outside_external_handoffs"
    if "stable_runtime" in {part.casefold() for part in path.parts}:
        return None, "unsafe_path_under_stable_runtime"
    archive_root = archive_db_path.resolve(strict=False).parent
    try:
        relative = path.relative_to(archive_root)
    except ValueError:
        return None, "unsafe_path_outside_archive_root"
    if "attachments" not in relative.parts:
        return None, "unsafe_path_not_under_attachments"
    expected_name = f"part_{int(part_index):03d}_{clean_text(attachment_sha256)[:16]}.bin"
    if path.name != expected_name:
        return None, "unsafe_attachment_filename_mismatch"
    expected_parent = archive_root / "attachments" / clean_text(message_sha256)
    try:
        path.parent.relative_to(expected_parent)
    except ValueError:
        return None, "unsafe_attachment_message_dir_mismatch"
    if path.suffix.casefold() != ".bin":
        return None, "unsafe_attachment_suffix"
    if not path.exists():
        return None, "attachment_file_missing"
    if not path.is_file():
        return None, "attachment_path_not_file"
    try:
        if path.stat().st_size != int(expected_size):
            return None, "attachment_file_size_mismatch"
    except OSError:
        return None, "attachment_stat_failed"
    return path, "ok"


def derived_attachment_text_path(
    *,
    derived_text_dir: Path,
    source_archive_id: object,
    message_sha256: object,
    part_index: int,
    attachment_sha256: object,
) -> Path:
    source = clean_text(source_archive_id)
    message_hash = clean_text(message_sha256)
    attachment_hash = clean_text(attachment_sha256)
    return (
        derived_text_dir
        / source
        / f"{message_hash[:16]}_part_{int(part_index):03d}_{attachment_hash[:16]}.txt"
    )


def parse_attachment_text_payload(
    *,
    payload: bytes,
    extension: str,
    config: MailAttachmentTextExtractConfig,
) -> dict[str, Any]:
    normalized_extension = clean_text(extension).casefold()
    if normalized_extension == ".txt":
        text, warnings = decode_text_attachment(payload)
        return {"parser": "plain_text", "text": text, "warnings": warnings}
    if normalized_extension == ".csv":
        text, warnings = parse_csv_attachment(payload, config=config)
        return {"parser": "csv", "text": text, "warnings": warnings}
    if normalized_extension == ".docx":
        text, warnings = parse_docx_attachment(payload, config=config)
        return {"parser": "docx_xml", "text": text, "warnings": warnings}
    if normalized_extension == ".xlsx":
        text, warnings = parse_xlsx_attachment(payload, config=config)
        return {"parser": "xlsx_openpyxl", "text": text, "warnings": warnings}
    raise AttachmentTextBlockedError("unsupported_stage_extension")


def parse_pdf_attachment(
    *,
    payload: bytes,
    config: MailAttachmentPdfExtractConfig,
) -> dict[str, Any]:
    if not payload.lstrip().startswith(b"%PDF"):
        raise AttachmentTextParseError("invalid_pdf_header")
    if b"/Encrypt" in payload[: min(len(payload), 2_000_000)]:
        raise AttachmentTextBlockedError("pdf_encrypted")

    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise AttachmentTextParseError("pypdf_unavailable") from exc

    with quiet_pypdf_warnings():
        try:
            reader = PdfReader(io.BytesIO(payload), strict=False)
        except Exception as exc:  # noqa: BLE001
            raise AttachmentTextParseError("invalid_pdf") from exc

        if getattr(reader, "is_encrypted", False):
            raise AttachmentTextBlockedError("pdf_encrypted")

        warnings = inspect_pdf_for_disallowed_features(
            reader=reader,
            payload=payload,
            max_objects=max(0, int(config.max_pdf_objects_to_scan)),
        )
        try:
            page_count = len(reader.pages)
        except Exception as exc:  # noqa: BLE001
            raise AttachmentTextParseError("pdf_page_count_failed", warnings) from exc
        if page_count <= 0:
            raise AttachmentTextBlockedError("pdf_no_pages", warnings)
        if page_count > max(0, int(config.max_pdf_pages)):
            raise AttachmentTextBlockedError("pdf_page_limit_exceeded", warnings)

        chunks: list[str] = []
        pages_processed = 0
        max_page_text_chars = max(0, int(config.max_page_text_chars))
        for page_index, page in enumerate(reader.pages, start=1):
            try:
                page_text = page.extract_text() or ""
            except Exception:  # noqa: BLE001
                warnings.append("pdf_page_extract_error")
                continue
            pages_processed += 1
            page_text = normalize_extracted_attachment_text(page_text)
            if len(page_text) > max_page_text_chars:
                warnings.append("pdf_page_text_truncated")
                page_text = page_text[:max_page_text_chars]
            if page_text:
                chunks.append(f"[page {page_index}]\n{page_text}")

    return {
        "parser": "pypdf_text",
        "text": "\n\n".join(chunks),
        "warnings": sorted(set(warnings)),
        "page_count": page_count,
        "pages_processed": pages_processed,
    }


@contextmanager
def quiet_pypdf_warnings() -> Any:
    logger_names = ("pypdf", "pypdf._reader", "pypdf._page", "pypdf._cmap", "pypdf.generic")
    previous_levels = {name: logging.getLogger(name).level for name in logger_names}
    try:
        for name in logger_names:
            logging.getLogger(name).setLevel(logging.CRITICAL + 1)
        yield
    finally:
        for name, level in previous_levels.items():
            logging.getLogger(name).setLevel(level)


def inspect_pdf_for_disallowed_features(
    *,
    reader: object,
    payload: bytes,
    max_objects: int,
) -> list[str]:
    warnings: set[str] = set()
    blocked: set[str] = set()
    for token, reason in PDF_BLOCK_TOKEN_REASONS.items():
        if token.encode("ascii") in payload:
            blocked.add(reason)

    visited: set[tuple[int, int]] = set()
    seen_objects = 0

    def visit(obj: object, depth: int) -> None:
        nonlocal seen_objects
        if depth > 8 or seen_objects >= max_objects:
            if seen_objects >= max_objects:
                warnings.add("pdf_object_scan_limit_reached")
            return
        seen_objects += 1
        if hasattr(obj, "idnum") and hasattr(obj, "generation") and hasattr(obj, "get_object"):
            key = (int(getattr(obj, "idnum")), int(getattr(obj, "generation")))
            if key in visited:
                return
            visited.add(key)
            try:
                obj = obj.get_object()
            except Exception:  # noqa: BLE001
                warnings.add("pdf_indirect_object_read_error")
                return
        if isinstance(obj, Mapping):
            for key, value in obj.items():
                token = clean_text(key)
                if token in PDF_BLOCK_TOKEN_REASONS:
                    blocked.add(PDF_BLOCK_TOKEN_REASONS[token])
                visit(value, depth + 1)
            return
        if isinstance(obj, (list, tuple)):
            for value in obj:
                visit(value, depth + 1)
            return
        token = clean_text(obj)
        if token in PDF_BLOCK_TOKEN_REASONS:
            blocked.add(PDF_BLOCK_TOKEN_REASONS[token])

    try:
        visit(getattr(reader, "trailer", {}), 0)
    except Exception:  # noqa: BLE001
        warnings.add("pdf_structure_scan_error")

    if blocked:
        raise AttachmentTextBlockedError(sorted(blocked)[0], sorted(blocked | warnings))
    return sorted(warnings)


def decode_text_attachment(payload: bytes) -> tuple[str, list[str]]:
    warnings: list[str] = []
    candidates: list[tuple[int, str, str]] = []
    for encoding in ("utf-8-sig", "cp1251", "latin-1"):
        text = payload.decode(encoding, errors="replace")
        candidates.append((text.count("\ufffd"), encoding, text))
    replacement_count, encoding, text = min(candidates, key=lambda item: item[0])
    if replacement_count:
        warnings.append("decode_replacement_chars")
    if encoding != "utf-8-sig":
        warnings.append(f"decoded_as_{encoding}")
    if b"\x00" in payload:
        warnings.append("nul_bytes_removed")
    return normalize_extracted_attachment_text(text), warnings


def parse_csv_attachment(
    payload: bytes,
    *,
    config: MailAttachmentTextExtractConfig,
) -> tuple[str, list[str]]:
    text, warnings = decode_text_attachment(payload)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
        warnings.append("csv_dialect_fallback")
    rows: list[str] = []
    reader = csv.reader(io.StringIO(text), dialect)
    max_rows = max(0, int(config.max_csv_rows))
    max_columns = max(0, int(config.max_csv_columns))
    for row_index, row in enumerate(reader, start=1):
        if row_index > max_rows:
            warnings.append("csv_rows_truncated")
            break
        cells = [normalize_extracted_attachment_text(cell)[:500] for cell in row[:max_columns]]
        if len(row) > max_columns:
            warnings.append("csv_columns_truncated")
        if any(cell.strip() for cell in cells):
            rows.append("\t".join(cells))
    return "\n".join(rows), sorted(set(warnings))


def parse_docx_attachment(
    payload: bytes,
    *,
    config: MailAttachmentTextExtractConfig,
) -> tuple[str, list[str]]:
    warnings = inspect_office_zip_payload(payload, config=config)
    texts: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            names = sorted(
                name
                for name in zf.namelist()
                if name == "word/document.xml"
                or re.fullmatch(r"word/(?:header|footer|footnotes|endnotes|comments)\d*\.xml", name)
            )
            for name in names:
                try:
                    texts.append(extract_wordprocessing_xml_text(zf.read(name)))
                except (ET.ParseError, KeyError, RuntimeError, zipfile.BadZipFile):
                    warnings.append("docx_xml_parse_error")
    except zipfile.BadZipFile as exc:
        raise AttachmentTextParseError("invalid_docx_zip", warnings) from exc
    return "\n\n".join(text for text in texts if text.strip()), sorted(set(warnings))


def parse_xlsx_attachment(
    payload: bytes,
    *,
    config: MailAttachmentTextExtractConfig,
) -> tuple[str, list[str]]:
    warnings = inspect_office_zip_payload(payload, config=config)
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise AttachmentTextParseError("openpyxl_unavailable", warnings) from exc

    try:
        workbook = load_workbook(
            io.BytesIO(payload),
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:  # noqa: BLE001
        raise AttachmentTextParseError("invalid_xlsx_workbook", warnings) from exc

    lines: list[str] = []
    max_sheets = max(0, int(config.max_xlsx_sheets))
    max_rows = max(0, int(config.max_xlsx_rows_per_sheet))
    max_columns = max(0, int(config.max_xlsx_columns_per_sheet))
    try:
        sheets = workbook.worksheets
        if len(sheets) > max_sheets:
            warnings.append("xlsx_sheets_truncated")
        for sheet_index, worksheet in enumerate(sheets[:max_sheets], start=1):
            lines.append(f"[sheet {sheet_index}]")
            for row in worksheet.iter_rows(
                max_row=max_rows,
                max_col=max_columns,
                values_only=True,
            ):
                cells = [safe_spreadsheet_cell_text(value) for value in row]
                if any(cell.strip() for cell in cells):
                    lines.append("\t".join(cells).rstrip())
            if getattr(worksheet, "max_row", 0) and int(worksheet.max_row or 0) > max_rows:
                warnings.append("xlsx_rows_truncated")
            if getattr(worksheet, "max_column", 0) and int(worksheet.max_column or 0) > max_columns:
                warnings.append("xlsx_columns_truncated")
    finally:
        workbook.close()
    return "\n".join(lines), sorted(set(warnings))


def inspect_office_zip_payload(
    payload: bytes,
    *,
    config: MailAttachmentTextExtractConfig,
) -> list[str]:
    warnings: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as zf:
            infos = zf.infolist()
            if len(infos) > max(0, int(config.max_zip_members)):
                raise AttachmentTextBlockedError("office_zip_member_limit_exceeded", warnings)
            total_uncompressed = 0
            rels_to_scan: list[str] = []
            worksheets_to_scan: list[str] = []
            for info in infos:
                name = clean_text(info.filename)
                parts = Path(name).parts
                if name.startswith("/") or "\\" in name or any(part == ".." for part in parts):
                    raise AttachmentTextBlockedError("office_zip_unsafe_member_path", warnings)
                lowered = name.casefold()
                if "vbaproject.bin" in lowered or "/vba" in lowered:
                    raise AttachmentTextBlockedError("office_macro_payload_detected", warnings)
                if "externallinks/" in lowered:
                    raise AttachmentTextBlockedError("office_external_links_declared", warnings)
                if lowered in {"xl/connections.xml", "word/afchunk.mht"} or "altchunk" in lowered:
                    raise AttachmentTextBlockedError("office_external_connection_declared", warnings)
                if "/activex/" in lowered or "/embeddings/" in lowered or "oleobject" in lowered:
                    raise AttachmentTextBlockedError("office_embedded_object_declared", warnings)
                if lowered.endswith(".rels"):
                    rels_to_scan.append(info.filename)
                if lowered.startswith("xl/worksheets/") and lowered.endswith(".xml"):
                    worksheets_to_scan.append(info.filename)
                total_uncompressed += int(info.file_size)
                if total_uncompressed > max(0, int(config.max_zip_uncompressed_bytes)):
                    raise AttachmentTextBlockedError("office_zip_uncompressed_limit_exceeded", warnings)
            for rels_name in rels_to_scan:
                rels = zf.read(rels_name)
                if b'TargetMode="External"' in rels or b"TargetMode='External'" in rels:
                    raise AttachmentTextBlockedError("office_external_relationship_declared", warnings)
            for worksheet_name in worksheets_to_scan:
                worksheet_xml = zf.read(worksheet_name)
                if b"<f" in worksheet_xml:
                    warnings.append("xlsx_formulas_present_data_only")
                if b"<hyperlink" in worksheet_xml:
                    warnings.append("xlsx_hyperlinks_present_not_followed")
    except zipfile.BadZipFile as exc:
        raise AttachmentTextParseError("invalid_office_zip", warnings) from exc
    return sorted(set(warnings))


def extract_wordprocessing_xml_text(payload: bytes) -> str:
    root = ET.fromstring(payload)
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if xml_local_name(paragraph.tag) != "p":
            continue
        chunks: list[str] = []
        for node in paragraph.iter():
            local_name = xml_local_name(node.tag)
            if local_name == "t" and node.text:
                chunks.append(node.text)
            elif local_name == "tab":
                chunks.append("\t")
            elif local_name in {"br", "cr"}:
                chunks.append("\n")
        text = normalize_extracted_attachment_text("".join(chunks))
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def xml_local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def safe_spreadsheet_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return normalize_extracted_attachment_text(str(value))[:500]


def normalize_extracted_attachment_text(value: object) -> str:
    text = clean_text(value).encode("utf-8", errors="replace").decode("utf-8").replace("\x00", "")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{4,}", "\n\n\n", text)
    return text.strip()


def load_attachment_parse_plan_source_rows(source_paths: Sequence[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source_path in source_paths:
        source_archive_id = mail_archive_source_id(source_path)
        with open_sqlite_readonly(source_path) as con:
            for row in con.execute(
                """
                SELECT
                  message_sha256,
                  part_index,
                  filename,
                  content_type,
                  size_bytes,
                  sha256 AS attachment_sha256
                FROM attachments
                ORDER BY message_sha256, part_index, sha256
                """
            ):
                filename = clean_text(row["filename"])
                rows.append(
                    {
                        "source_archive_id": source_archive_id,
                        "message_sha256": clean_text(row["message_sha256"]),
                        "part_index": int(row["part_index"]),
                        "filename": filename,
                        "filename_sha256": (
                            hashlib.sha256(filename.encode("utf-8")).hexdigest()
                            if filename
                            else ""
                        ),
                        "content_type": clean_text(row["content_type"]),
                        "size_bytes": int(row["size_bytes"] or 0),
                        "attachment_sha256": clean_text(row["attachment_sha256"]),
                    }
                )
    return rows


def classify_attachment_for_parse_plan(
    *,
    filename: str,
    declared_content_type: str,
    size_bytes: int,
    max_size_bytes: int,
    allow_extensions: set[str],
    review_extensions: set[str],
) -> dict[str, Any]:
    extension = safe_attachment_extension(filename)
    suffixes = safe_attachment_suffixes(filename)
    content_type = safe_declared_content_type(declared_content_type)
    reasons: list[str] = []

    if int(size_bytes) <= 0:
        reasons.append("empty_attachment")
    if int(size_bytes) > int(max_size_bytes):
        reasons.append("size_limit_exceeded")
    if has_path_traversal_filename(filename):
        reasons.append("path_traversal_filename")
    if has_dangerous_double_extension(suffixes):
        reasons.append("double_extension")
    if extension in ATTACHMENT_PARSE_ARCHIVE_EXTENSIONS:
        reasons.append("archive_extension")
    if extension in ATTACHMENT_PARSE_EXECUTABLE_EXTENSIONS:
        reasons.append("executable_or_script_extension")
    if extension in ATTACHMENT_PARSE_MACRO_EXTENSIONS:
        reasons.append("macro_enabled_office_extension")
    if extension in {".doc", ".xls", ".ppt"}:
        reasons.append("legacy_office_manual_review")
    if extension in {"[missing]", "[none]"}:
        reasons.append("missing_extension")
    if extension == "[other_or_non_ascii]":
        reasons.append("nonstandard_extension")
    if content_type in ATTACHMENT_PARSE_BLOCK_CONTENT_TYPES:
        reasons.append("blocked_declared_content_type")
    if (
        extension not in allow_extensions
        and extension not in review_extensions
        and extension not in ATTACHMENT_PARSE_BLOCK_EXTENSIONS
        and extension not in {"[missing]", "[none]", "[other_or_non_ascii]"}
    ):
        reasons.append("not_allowlisted")

    block_reasons = {
        "empty_attachment",
        "size_limit_exceeded",
        "path_traversal_filename",
        "double_extension",
        "archive_extension",
        "executable_or_script_extension",
        "macro_enabled_office_extension",
        "blocked_declared_content_type",
    }
    if any(reason in block_reasons for reason in reasons):
        risk_level = "block"
        action = "blocked"
    elif reasons:
        risk_level = "review"
        action = "manual_review"
    else:
        risk_level = "allow"
        action = "parse_later"

    return {
        "extension": extension,
        "declared_content_type": content_type,
        "risk_level": risk_level,
        "action": action,
        "risk_reasons": sorted(set(reasons)),
    }


def normalize_attachment_extensions(values: Sequence[str]) -> set[str]:
    result: set[str] = set()
    for value in values:
        text = clean_text(value).casefold()
        if not text:
            continue
        result.add(text if text.startswith(".") else f".{text}")
    return result


def safe_attachment_extension(filename: str) -> str:
    name = clean_text(filename)
    if not name:
        return "[missing]"
    suffix = Path(name).suffix.casefold()
    if not suffix:
        return "[none]"
    if not re.fullmatch(r"\.[a-z0-9]{1,10}", suffix):
        return "[other_or_non_ascii]"
    return suffix


def safe_attachment_suffixes(filename: str) -> list[str]:
    suffixes: list[str] = []
    for suffix in Path(clean_text(filename)).suffixes:
        lowered = suffix.casefold()
        if re.fullmatch(r"\.[a-z0-9]{1,10}", lowered):
            suffixes.append(lowered)
        elif lowered:
            suffixes.append("[other_or_non_ascii]")
    return suffixes


def has_dangerous_double_extension(suffixes: Sequence[str]) -> bool:
    if len(suffixes) < 2:
        return False
    return any(suffix in ATTACHMENT_PARSE_BLOCK_EXTENSIONS for suffix in suffixes)


def has_path_traversal_filename(filename: str) -> bool:
    text = clean_text(filename)
    if "/" in text or "\\" in text:
        return True
    return any(part == ".." for part in Path(text).parts)


def safe_declared_content_type(value: str) -> str:
    text = clean_text(value).casefold()
    if not text:
        return "unknown"
    if re.fullmatch(r"[a-z0-9.+-]+/[a-z0-9.+-]+", text):
        return text
    return "invalid"


def load_mail_handoff_candidate_summary(mail_handoff_db_path: Path) -> dict[str, Any]:
    with open_sqlite_readonly(mail_handoff_db_path) as con:
        link_status_counts = {
            "ready": 0,
            "manual_review": 0,
            "excluded": 0,
        }
        link_status_counts.update(
            {
                str(row["link_status"]): int(row["count"])
                for row in con.execute(
                    """
                    SELECT link_status, COUNT(*) AS count
                    FROM mail_customer_links
                    GROUP BY link_status
                    """
                )
            }
        )
        total_links = int(con.execute("SELECT COUNT(*) FROM mail_customer_links").fetchone()[0])
        candidate_summaries: dict[str, dict[str, Any]] = {}
        anomalies = 0
        for row in con.execute(
            """
            SELECT message_sha256, message_date_iso, candidate_keys_json
            FROM mail_customer_links
            WHERE match_class = 'strong_unique'
            ORDER BY message_date_iso, message_sha256
            """
        ):
            try:
                candidate_keys = json.loads(row["candidate_keys_json"] or "[]")
            except json.JSONDecodeError:
                candidate_keys = []
            if len(candidate_keys) != 1:
                anomalies += 1
                continue
            candidate_key = clean_text(candidate_keys[0])
            if not candidate_key:
                anomalies += 1
                continue
            summary = candidate_summaries.setdefault(
                candidate_key,
                {
                    "mail_message_count": 0,
                    "first_mail_date_iso": "",
                    "last_mail_date_iso": "",
                    "message_sha256_refs": [],
                },
            )
            message_date = clean_text(row["message_date_iso"])
            summary["mail_message_count"] += 1
            if message_date and (
                not summary["first_mail_date_iso"] or message_date < summary["first_mail_date_iso"]
            ):
                summary["first_mail_date_iso"] = message_date
            if message_date and (
                not summary["last_mail_date_iso"] or message_date > summary["last_mail_date_iso"]
            ):
                summary["last_mail_date_iso"] = message_date
            summary["message_sha256_refs"].append(clean_text(row["message_sha256"]))
    return {
        "candidate_summaries": candidate_summaries,
        "link_status_counts": link_status_counts,
        "total_links": total_links,
        "strong_candidate_key_anomalies": anomalies,
    }


def load_identity_candidate_refs(identity_db_path: Path) -> dict[str, dict[str, str]]:
    refs: dict[str, dict[str, str]] = {}
    with open_sqlite_readonly(identity_db_path) as con:
        for row in con.execute(
            """
            SELECT candidate_key, tallanto_id, amocrm_id
            FROM identity_candidates
            ORDER BY candidate_key
            """
        ):
            refs[row["candidate_key"]] = {
                "tallanto_id": clean_text(row["tallanto_id"]),
                "amocrm_id": clean_text(row["amocrm_id"]),
            }
    return refs


def load_candidate_phone_links(identity_db_path: Path) -> dict[str, list[dict[str, Any]]]:
    by_candidate: dict[str, list[dict[str, Any]]] = {}
    with open_sqlite_readonly(identity_db_path) as con:
        for row in con.execute(
            """
            SELECT il.candidate_key, il.value, iv.match_class, iv.candidate_count
            FROM identity_links il
            JOIN identity_values iv
              ON iv.kind = il.kind AND iv.value = il.value
            WHERE il.kind = 'phone'
            ORDER BY il.candidate_key, il.value
            """
        ):
            by_candidate.setdefault(row["candidate_key"], []).append(
                {
                    "normalized_phone": clean_text(row["value"]),
                    "match_class": clean_text(row["match_class"]),
                    "candidate_count": int(row["candidate_count"]),
                }
            )
    return by_candidate


def load_mango_calls_by_phone(product_db_path: Path) -> tuple[dict[str, list[dict[str, Any]]], dict[str, int]]:
    by_phone: dict[str, dict[str, dict[str, Any]]] = {}
    counts = {
        "capture_rows_seen": 0,
        "capture_rows_with_phone": 0,
        "capture_rows_with_normalized_phone": 0,
        "distinct_normalized_phones": 0,
        "product_calls_seen": 0,
        "product_calls_with_filename_phone": 0,
        "distinct_product_filename_phones": 0,
        "capture_rows_joined_to_product_calls": 0,
    }
    with open_sqlite_readonly(product_db_path) as con:
        counts["product_calls_seen"] = int(con.execute("SELECT COUNT(*) FROM product_calls").fetchone()[0])
        rows = list(
            con.execute(
                """
                SELECT
                  c.tenant_id,
                  c.provider,
                  c.event_key,
                  c.provider_call_id,
                  c.started_at,
                  c.direction,
                  c.client_phone,
                  c.manager_ref,
                  c.recording_ref,
                  c.status,
                  pc.started_at AS product_call_started_at,
                  pc.duration_sec,
                  pc.manager_extension,
                  pc.crm_owner_id,
                  pc.crm_match_status
                FROM capture_inbox_items c
                LEFT JOIN product_calls pc
                  ON pc.tenant_id = c.tenant_id
                 AND pc.telephony_provider = c.provider
                 AND (
                   pc.event_key = c.event_key
                   OR pc.provider_call_id = c.provider_call_id
                 )
                ORDER BY c.started_at, c.event_key, c.provider_call_id
                """
            )
        )
        product_rows = list(
            con.execute(
                """
                SELECT
                  tenant_id,
                  telephony_provider AS provider,
                  event_key,
                  provider_call_id,
                  recording_id,
                  source_filename,
                  started_at,
                  duration_sec,
                  manager_extension,
                  crm_owner_id,
                  crm_match_status
                FROM product_calls
                ORDER BY started_at, event_key, provider_call_id
                """
            )
        )
    counts["capture_rows_seen"] = len(rows)
    seen_capture_keys: set[tuple[str, str, str]] = set()
    normalized_capture_keys: set[tuple[str, str, str]] = set()
    joined_capture_keys: set[tuple[str, str, str]] = set()
    for row in rows:
        capture_key = (
            clean_text(row["tenant_id"]),
            clean_text(row["provider"]),
            clean_text(row["event_key"]),
        )
        if capture_key not in seen_capture_keys:
            seen_capture_keys.add(capture_key)
            if clean_text(row["client_phone"]):
                counts["capture_rows_with_phone"] += 1
        normalized_phone = normalize_phone(row["client_phone"])
        if not normalized_phone:
            continue
        if capture_key not in normalized_capture_keys:
            normalized_capture_keys.add(capture_key)
            counts["capture_rows_with_normalized_phone"] += 1
        call_ref_key = "|".join(
            [
                clean_text(row["tenant_id"]),
                clean_text(row["provider"]),
                clean_text(row["event_key"]),
                clean_text(row["provider_call_id"]),
            ]
        )
        if clean_text(row["product_call_started_at"]):
            joined_capture_keys.add(capture_key)
        by_phone.setdefault(normalized_phone, {})
        by_phone[normalized_phone].setdefault(
            call_ref_key,
            {
                "source_table": "capture_inbox_items",
                "normalized_phone": normalized_phone,
                "call_ref_key": call_ref_key,
                "tenant_id": clean_text(row["tenant_id"]),
                "provider": clean_text(row["provider"]),
                "event_key": clean_text(row["event_key"]),
                "provider_call_id": clean_text(row["provider_call_id"]),
                "started_at": clean_text(row["started_at"]),
                "direction": clean_text(row["direction"]),
                "manager_ref": clean_text(row["manager_ref"]),
                "recording_ref": clean_text(row["recording_ref"]),
                "status": clean_text(row["status"]),
                "product_call_started_at": clean_text(row["product_call_started_at"]),
                "duration_sec": row["duration_sec"],
                "manager_extension": clean_text(row["manager_extension"]),
                "crm_owner_id": row["crm_owner_id"],
                "crm_match_status": clean_text(row["crm_match_status"]),
            },
        )
    product_filename_phones: set[str] = set()
    for row in product_rows:
        filename_meta = parse_filename_metadata(clean_text(row["source_filename"]))
        normalized_phone = normalize_phone(filename_meta.get("phone"))
        if not normalized_phone:
            continue
        counts["product_calls_with_filename_phone"] += 1
        product_filename_phones.add(normalized_phone)
        call_ref_key = "|".join(
            [
                clean_text(row["tenant_id"]),
                clean_text(row["provider"]),
                clean_text(row["event_key"]),
                clean_text(row["provider_call_id"]),
            ]
        )
        by_phone.setdefault(normalized_phone, {})
        by_phone[normalized_phone].setdefault(
            call_ref_key,
            {
                "source_table": "product_calls_filename",
                "normalized_phone": normalized_phone,
                "call_ref_key": call_ref_key,
                "tenant_id": clean_text(row["tenant_id"]),
                "provider": clean_text(row["provider"]),
                "event_key": clean_text(row["event_key"]),
                "provider_call_id": clean_text(row["provider_call_id"]),
                "started_at": clean_text(row["started_at"]),
                "direction": "",
                "manager_ref": clean_text(row["manager_extension"]),
                "recording_ref": clean_text(row["recording_id"]),
                "status": "",
                "product_call_started_at": clean_text(row["started_at"]),
                "duration_sec": row["duration_sec"],
                "manager_extension": clean_text(row["manager_extension"]),
                "crm_owner_id": row["crm_owner_id"],
                "crm_match_status": clean_text(row["crm_match_status"]),
            },
        )
    counts["distinct_normalized_phones"] = len(by_phone)
    counts["distinct_product_filename_phones"] = len(product_filename_phones)
    counts["capture_rows_joined_to_product_calls"] = len(joined_capture_keys)
    return (
        {
            phone: sorted(
                calls.values(),
                key=lambda item: (clean_text(item.get("started_at")), clean_text(item.get("call_ref_key"))),
            )
            for phone, calls in by_phone.items()
        },
        counts,
    )


def mango_phone_index_row_from_call(
    *,
    phone: str,
    call: Mapping[str, Any],
    source_kind: str,
    source_path_sha256: str,
    source_filename_sha256: str,
    source_root_index: Optional[int],
) -> dict[str, Any]:
    normalized_phone = normalize_phone(phone)
    call_ref_key = clean_text(call.get("call_ref_key"))
    if not call_ref_key:
        call_ref_key = hashlib.sha256(
            json.dumps(call, ensure_ascii=False, sort_keys=True).encode("utf-8")
        ).hexdigest()
    return {
        "normalized_phone": normalized_phone,
        "phone_sha256": hashlib.sha256(normalized_phone.encode("utf-8")).hexdigest(),
        "call_ref_key": call_ref_key,
        "source_kind": clean_text(source_kind),
        "tenant_id": clean_text(call.get("tenant_id")) or "local_mango_archive",
        "provider": clean_text(call.get("provider")) or "mango",
        "event_key": clean_text(call.get("event_key")) or call_ref_key,
        "provider_call_id": clean_text(call.get("provider_call_id")) or call_ref_key,
        "started_at": clean_text(call.get("started_at")),
        "direction": clean_text(call.get("direction")),
        "manager_ref": clean_text(call.get("manager_ref")),
        "recording_ref": clean_text(call.get("recording_ref")),
        "status": clean_text(call.get("status")),
        "product_call_started_at": clean_text(call.get("product_call_started_at")),
        "duration_sec": call.get("duration_sec"),
        "manager_extension": clean_text(call.get("manager_extension")),
        "crm_owner_id": call.get("crm_owner_id"),
        "crm_match_status": clean_text(call.get("crm_match_status")),
        "source_path_sha256": clean_text(source_path_sha256),
        "source_filename_sha256": clean_text(source_filename_sha256),
        "source_root_index": source_root_index,
    }


def scan_recording_filename_phone_refs(recording_roots: Sequence[Path]) -> dict[str, Any]:
    rows: list[dict[str, Any]] = []
    files_seen = 0
    files_with_filename_phone = 0
    distinct_filename_phones: set[str] = set()
    for root_index, root in enumerate(recording_roots):
        for path in sorted(Path(root).rglob("*")):
            if not path.is_file() or path.suffix.casefold() not in SUPPORTED_EXTENSIONS:
                continue
            files_seen += 1
            metadata = parse_filename_metadata(path.name)
            normalized_phone = normalize_phone(metadata.get("phone"))
            if not normalized_phone:
                continue
            files_with_filename_phone += 1
            distinct_filename_phones.add(normalized_phone)
            started_at = metadata.get("started_at")
            if isinstance(started_at, datetime):
                started_at_text = started_at.isoformat()
            else:
                started_at_text = clean_text(started_at)
            filename_sha256 = hashlib.sha256(path.name.encode("utf-8")).hexdigest()
            path_sha256 = hashlib.sha256(str(path.resolve(strict=False)).encode("utf-8")).hexdigest()
            rows.append(
                mango_phone_index_row_from_call(
                    phone=normalized_phone,
                    call={
                        "call_ref_key": f"recording_filename|{filename_sha256}",
                        "tenant_id": "local_mango_archive",
                        "provider": "mango",
                        "event_key": filename_sha256,
                        "provider_call_id": clean_text(metadata.get("source_call_id")) or filename_sha256,
                        "started_at": started_at_text,
                        "direction": "",
                        "manager_ref": clean_text(metadata.get("manager_name")),
                        "recording_ref": f"filename_sha256:{filename_sha256}",
                        "status": "filename_indexed",
                        "product_call_started_at": started_at_text,
                        "duration_sec": None,
                        "manager_extension": "",
                        "crm_owner_id": None,
                        "crm_match_status": "",
                    },
                    source_kind="recording_filename",
                    source_path_sha256=path_sha256,
                    source_filename_sha256=filename_sha256,
                    source_root_index=root_index,
                )
            )
    return {
        "rows": rows,
        "files_seen": files_seen,
        "files_with_filename_phone": files_with_filename_phone,
        "distinct_filename_phones": len(distinct_filename_phones),
    }


def load_mango_phone_index_calls_by_phone(
    index_db_path: Path,
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, int]]:
    by_phone: dict[str, dict[str, dict[str, Any]]] = {}
    with open_sqlite_readonly(index_db_path) as con:
        rows = list(
            con.execute(
                """
                SELECT
                  normalized_phone,
                  call_ref_key,
                  source_kind,
                  tenant_id,
                  provider,
                  event_key,
                  provider_call_id,
                  started_at,
                  direction,
                  manager_ref,
                  recording_ref,
                  status,
                  product_call_started_at,
                  duration_sec,
                  manager_extension,
                  crm_owner_id,
                  crm_match_status
                FROM mango_phone_call_refs
                ORDER BY started_at, call_ref_key
                """
            )
        )
    for row in rows:
        normalized_phone = clean_text(row["normalized_phone"])
        if not normalized_phone:
            continue
        call_ref_key = clean_text(row["call_ref_key"])
        by_phone.setdefault(normalized_phone, {})
        by_phone[normalized_phone].setdefault(
            call_ref_key,
            {
                "source_table": clean_text(row["source_kind"]),
                "normalized_phone": normalized_phone,
                "call_ref_key": call_ref_key,
                "tenant_id": clean_text(row["tenant_id"]),
                "provider": clean_text(row["provider"]),
                "event_key": clean_text(row["event_key"]),
                "provider_call_id": clean_text(row["provider_call_id"]),
                "started_at": clean_text(row["started_at"]),
                "direction": clean_text(row["direction"]),
                "manager_ref": clean_text(row["manager_ref"]),
                "recording_ref": clean_text(row["recording_ref"]),
                "status": clean_text(row["status"]),
                "product_call_started_at": clean_text(row["product_call_started_at"]),
                "duration_sec": row["duration_sec"],
                "manager_extension": clean_text(row["manager_extension"]),
                "crm_owner_id": row["crm_owner_id"],
                "crm_match_status": clean_text(row["crm_match_status"]),
            },
        )
    return (
        {
            phone: sorted(
                calls.values(),
                key=lambda item: (clean_text(item.get("started_at")), clean_text(item.get("call_ref_key"))),
            )
            for phone, calls in by_phone.items()
        },
        {
            "call_refs_loaded": len(rows),
            "distinct_normalized_phones": len(by_phone),
        },
    )


def merge_mango_calls_by_phone(
    base: Mapping[str, Sequence[Mapping[str, Any]]],
    extra: Mapping[str, Sequence[Mapping[str, Any]]],
) -> tuple[dict[str, list[dict[str, Any]]], int]:
    merged: dict[str, dict[str, dict[str, Any]]] = {}
    for source in (base, extra):
        for phone, calls in source.items():
            for call in calls:
                normalized_phone = normalize_phone(phone) or clean_text(phone)
                call_ref_key = clean_text(call.get("call_ref_key"))
                if not normalized_phone or not call_ref_key:
                    continue
                merged.setdefault(normalized_phone, {})
                merged[normalized_phone].setdefault(call_ref_key, dict(call))
    base_keys = {
        (normalize_phone(phone) or clean_text(phone), clean_text(call.get("call_ref_key")))
        for phone, calls in base.items()
        for call in calls
        if clean_text(call.get("call_ref_key"))
    }
    merged_keys = {
        (phone, call_ref_key)
        for phone, calls in merged.items()
        for call_ref_key in calls
    }
    return (
        {
            phone: sorted(
                calls.values(),
                key=lambda item: (clean_text(item.get("started_at")), clean_text(item.get("call_ref_key"))),
            )
            for phone, calls in merged.items()
        },
        len(merged_keys - base_keys),
    )


def classify_message_against_identity(
    *,
    message_kind: str,
    participants: Sequence[Mapping[str, Any]],
    identity_index: Mapping[str, Sequence[str]],
    internal_domains: set[str],
) -> dict[str, Any]:
    external_emails = [
        clean_text(row["email_normalized"])
        for row in participants
        if row["email_normalized"]
        and not email_domain_is_internal(clean_text(row["email_normalized"]), internal_domains)
    ]
    external_emails = unique_preserving_order(external_emails)
    if message_kind in {"internal", "service"}:
        return {
            "match_class": "internal_or_service",
            "candidate_keys": [],
            "matched_email_count": 0,
        }
    candidate_keys = unique_preserving_order(
        candidate_key
        for address in external_emails
        for candidate_key in identity_index.get(address, [])
    )
    if len(candidate_keys) == 1:
        match_class = "strong_unique"
    elif len(candidate_keys) > 1:
        match_class = "ambiguous"
    else:
        match_class = "missing"
    return {
        "match_class": match_class,
        "candidate_keys": candidate_keys,
        "matched_email_count": sum(1 for address in external_emails if address in identity_index),
    }


def link_status_for_match_class(match_class: str) -> str:
    if match_class == "strong_unique":
        return "ready"
    if match_class in {"ambiguous", "missing"}:
        return "manual_review"
    return "excluded"


def blocked_reason_for_match_class(match_class: str) -> str:
    return {
        "ambiguous": "ambiguous_identity_match",
        "missing": "missing_identity_match",
        "internal_or_service": "internal_or_service_message",
    }.get(match_class, "")


def raw_message_path(raw_dir: Path, raw_sha256: str) -> Path:
    return raw_dir / raw_sha256[:2] / f"{raw_sha256}.eml"


def write_bytes_once(path: Path, payload: bytes) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return False
    path.write_bytes(payload)
    return True


def write_text_once(path: Path, text: str) -> bool:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return False
    path.write_text(text, encoding="utf-8")
    return True


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def count_files(root: Path, pattern: str) -> int:
    if not root.exists():
        return 0
    return sum(1 for path in root.rglob(pattern) if path.is_file())


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def remove_sqlite_sidecars(db_path: Path) -> None:
    for suffix in ("-wal", "-shm"):
        sidecar = Path(f"{db_path}{suffix}")
        try:
            if sidecar.exists():
                sidecar.unlink()
        except OSError:
            pass


def parse_email_date_iso(date_header: str) -> str:
    if not date_header:
        return ""
    try:
        parsed = parsedate_to_datetime(date_header)
    except Exception:  # noqa: BLE001
        return ""
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc).isoformat()


def candidate_key_for_row(
    row_number: int,
    row: Mapping[str, Any],
    *,
    tallanto_id_counts: Mapping[str, int],
) -> str:
    tallanto_id = tallanto_id_for_row(row)
    if tallanto_id:
        if int(tallanto_id_counts.get(tallanto_id) or 0) > 1:
            return f"tallanto:{tallanto_id}:row:{row_number}"
        return f"tallanto:{tallanto_id}"
    digest = hashlib.sha256(
        json.dumps(row, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()[:16]
    return f"row:{row_number}:{digest}"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def shell_safe_token(value: str) -> str:
    text = clean_text(value)
    if not text:
        return "''"
    if re.fullmatch(r"[A-Za-z0-9_./:@=+-]+", text):
        return text
    return "'" + text.replace("'", "'\"'\"'") + "'"


def valid_env_var_name(value: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z_][A-Za-z0-9_]*", clean_text(value)))


def unique_preserving_order(values: Sequence[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def normalize_domains(domains: Sequence[str], mailbox_email: str = "") -> set[str]:
    result = {
        domain.casefold().strip().lstrip("@")
        for domain in domains
        if clean_text(domain)
    }
    mailbox_domain = domain_from_email(normalize_email(mailbox_email))
    if mailbox_domain:
        result.add(mailbox_domain)
    return result


def domain_from_email(address: str) -> str:
    normalized = normalize_email(address)
    if "@" not in normalized:
        return ""
    return normalized.rsplit("@", 1)[1]


def email_domain_is_internal(address: str, internal_domains: set[str]) -> bool:
    domain = domain_from_email(address)
    return bool(domain and domain in internal_domains)


def is_service_email(address: str) -> bool:
    normalized = normalize_email(address)
    if "@" not in normalized:
        return False
    local = normalized.split("@", 1)[0]
    return any(token in local for token in SERVICE_LOCAL_PARTS)


def guard_not_stable_runtime(path: Path, label: str) -> None:
    parts = {part.casefold() for part in path.resolve(strict=False).parts}
    if "stable_runtime" in parts:
        raise ValueError(f"{label} must not be under stable_runtime")


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


__all__ = [
    "FULL_MESSAGE_FETCH_QUERY",
    "DEFAULT_ATTACHMENT_IMAGE_OCR_EXTENSIONS",
    "DEFAULT_ATTACHMENT_PARSE_ALLOW_EXTENSIONS",
    "DEFAULT_ATTACHMENT_PDF_EXTRACT_EXTENSIONS",
    "DEFAULT_ATTACHMENT_PARSE_REVIEW_EXTENSIONS",
    "DEFAULT_ATTACHMENT_TEXT_EXTRACT_EXTENSIONS",
    "MAIL_ARCHIVE_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_IMAGE_OCR_PLAN_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_OCR_PREFLIGHT_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_OCR_PILOT_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_PDF_EXTRACT_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_PARSE_PLAN_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_STAGE6_PLAN_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_TEXT_EXTRACT_SCHEMA_VERSION",
    "MAIL_ATTACHMENT_TEXT_INDEX_SCHEMA_VERSION",
    "MAIL_MANGO_BRIDGE_PREVIEW_SCHEMA_VERSION",
    "MAIL_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION",
    "MAIL_STAGE2_CUSTOMER_RELINK_PREVIEW_SCHEMA_VERSION",
    "MAIL_MATCHING_REPORT_SCHEMA_VERSION",
    "MAIL_PHONE_LIFT_PREVIEW_SCHEMA_VERSION",
    "MANGO_PHONE_INDEX_PREVIEW_SCHEMA_VERSION",
    "TALLANTO_IDENTITY_MAP_SCHEMA_VERSION",
    "MailArchiveIngestConfig",
    "MailArchivePreflightConfig",
    "MailArchiveVerificationConfig",
    "MailAttachmentImageOcrPlanConfig",
    "MailAttachmentOcrPreflightConfig",
    "MailAttachmentOcrPilotConfig",
    "MailAttachmentPdfExtractConfig",
    "MailAttachmentParsePlanConfig",
    "MailAttachmentStage6PlanConfig",
    "MailAttachmentTextExtractConfig",
    "MailAttachmentTextIndexConfig",
    "MailCustomerHistoryHandoffConfig",
    "MailCustomerRelinkPreviewConfig",
    "MailStage2CustomerRelinkPreviewConfig",
    "MailMangoBridgePreviewConfig",
    "MailMatchingReportConfig",
    "MailPhoneLiftPreviewConfig",
    "MangoPhoneIndexPreviewConfig",
    "TallantoIdentityMapConfig",
    "build_mail_archive_ingest",
    "build_mail_archive_preflight",
    "build_mail_attachment_image_ocr_plan",
    "build_mail_attachment_ocr_preflight",
    "build_mail_attachment_ocr_pilot",
    "build_mail_attachment_pdf_extract",
    "build_mail_attachment_parse_plan",
    "build_mail_attachment_stage6_plan",
    "build_mail_attachment_text_extract",
    "build_mail_attachment_text_index",
    "build_mail_customer_history_handoff",
    "build_mail_customer_relink_preview",
    "build_mail_stage2_customer_relink_preview",
    "build_mail_mango_bridge_preview",
    "build_mail_matching_report",
    "build_mail_phone_lift_preview",
    "build_mango_phone_index_preview",
    "build_tallanto_identity_map",
    "extract_email_addresses",
    "extract_phone_numbers",
    "normalize_email",
    "normalize_phone",
    "valid_env_var_name",
    "verify_mail_archive_pilot",
]
