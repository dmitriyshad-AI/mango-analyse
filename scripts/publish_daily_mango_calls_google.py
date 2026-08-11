#!/usr/bin/env python3
"""Publish a verified daily XLSX as an immutable native Google spreadsheet."""
from __future__ import annotations

import argparse
import fcntl
import hashlib
import io
import json
import os
import re
import secrets
import stat
import tempfile
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Mapping, Sequence
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"
GOOGLE_FOLDER_MIME = "application/vnd.google-apps.folder"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SCHEMA = "daily_mango_calls_resolve_export_v5"
CONFIRMATION = "UPLOAD_MANGO_DAILY_REPORT"
MULTIPART_LIMIT = 5 * 1024 * 1024
MOSCOW = ZoneInfo("Europe/Moscow")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def workbook_digest(source: Path | bytes) -> str:
    workbook = load_workbook(source if isinstance(source, Path) else io.BytesIO(source), read_only=False, data_only=False)
    payload = []
    for sheet in workbook:
        cells = [
            [cell.coordinate, cell.data_type, cell.value, cell.hyperlink.target if cell.hyperlink else None]
            for row in sheet.iter_rows() for cell in row if cell.value is not None or cell.hyperlink
        ]
        payload.append([sheet.title, sheet.max_row, sheet.max_column, sorted(map(str, sheet.merged_cells.ranges)), cells])
    workbook.close()
    raw = json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str).encode()
    return hashlib.sha256(raw).hexdigest()


def google_upload_bytes(path: Path) -> bytes:
    workbook = load_workbook(path, read_only=False, data_only=False)
    for sheet in workbook:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    cell.value = f"Яндекс Диск: {cell.hyperlink.target}"
                    cell.hyperlink = None
    target = io.BytesIO()
    workbook.save(target)
    workbook.close()
    return target.getvalue()


def load_plan(root: Path, day: date) -> dict[str, Any]:
    root = root.resolve()
    manifest_path = root / f"Отчёт РОП по звонкам {day.isoformat()}.manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    if not isinstance(manifest, Mapping) or manifest.get("day") != day.isoformat():
        raise RuntimeError("daily report manifest day is invalid")
    content_hash = str(manifest.get("content_sha256") or "")
    if manifest.get("schema_version") != SCHEMA or not re.fullmatch(r"[0-9a-f]{64}", content_hash):
        raise RuntimeError("daily report manifest schema or content hash is invalid")
    xlsx = (root / str(manifest.get("xlsx") or "")).resolve()
    try:
        xlsx.relative_to(root)
    except ValueError as exc:
        raise RuntimeError("daily report XLSX escapes report root") from exc
    if not xlsx.is_file() or xlsx.stat().st_size > MULTIPART_LIMIT or sha256_file(xlsx) != manifest.get("xlsx_sha256"):
        raise RuntimeError("daily report XLSX does not match manifest")
    try:
        generated = datetime.fromisoformat(str(manifest["generated_at"]).replace("Z", "+00:00"))
        if generated.tzinfo is None:
            raise ValueError
    except (KeyError, ValueError) as exc:
        raise RuntimeError("daily report generation time is invalid") from exc
    uploaded = google_upload_bytes(xlsx)
    stamp = generated.astimezone(MOSCOW).strftime("%d.%m.%Y %H-%M МСК")
    return {
        "day": day.isoformat(), "content_sha256": content_hash, "xlsx_sha256": manifest["xlsx_sha256"],
        "xlsx": xlsx, "upload_bytes": uploaded, "workbook_digest": workbook_digest(uploaded),
        "name": f"Отчёт РОП по звонкам {day.isoformat()} — опубликован {stamp} — v5-{content_hash[:12]}",
        "temporary_name": f"ПРОВЕРКА — НЕ ИСПОЛЬЗОВАТЬ — {day.isoformat()} — {content_hash[:12]}",
    }


def validate_credentials(path: Path, repo: Path) -> Path:
    path = path.expanduser().resolve()
    if not path.is_file() or stat.S_IMODE(path.stat().st_mode) != 0o600:
        raise RuntimeError("Google credentials must be an existing owner-only 0600 file")
    if repo.resolve() in path.parents or any("yandex.disk" in part.casefold() for part in path.parts):
        raise RuntimeError("Google credentials must not live in the repository or Yandex Disk")
    return path


def response_json(response: Any) -> Mapping[str, Any]:
    if not 200 <= int(response.status_code) < 300:
        raise RuntimeError(f"Google Drive HTTP {response.status_code}")
    try:
        payload = response.json()
    except Exception as exc:
        raise RuntimeError("Google Drive returned invalid JSON") from exc
    if not isinstance(payload, Mapping):
        raise RuntimeError("Google Drive returned invalid JSON")
    return payload


def permissions_are_private(metadata: Mapping[str, Any]) -> bool:
    permissions = metadata.get("permissions")
    return bool(permissions) and isinstance(permissions, list) and all(
        isinstance(item, Mapping) and item.get("type") not in {"anyone", "domain"} for item in permissions
    )


def validate_target_folder(session: Any, folder_id: str) -> None:
    fields = "id,mimeType,capabilities(canAddChildren),permissions(type,role,allowFileDiscovery)"
    meta = response_json(session.get(
        f"https://www.googleapis.com/drive/v3/files/{folder_id}",
        params={"fields": fields, "supportsAllDrives": "true"}, timeout=60,
    ))
    capabilities = meta.get("capabilities") if isinstance(meta.get("capabilities"), Mapping) else {}
    if meta.get("mimeType") != GOOGLE_FOLDER_MIME or capabilities.get("canAddChildren") is not True or not permissions_are_private(meta):
        raise RuntimeError("Google Drive target folder is not private and writable")


def expected_properties(plan: Mapping[str, Any], state: str) -> dict[str, str]:
    return {
        "mango_report_day": str(plan["day"]), "mango_report_content_sha256": str(plan["content_sha256"]),
        "mango_report_schema": SCHEMA, "mango_report_state": state,
    }


def readback(session: Any, file_id: str, folder_id: str, plan: Mapping[str, Any], *, state: str, check_content: bool) -> None:
    fields = "id,name,mimeType,parents,appProperties,permissions(type,role,allowFileDiscovery)"
    meta = response_json(session.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={"fields": fields, "supportsAllDrives": "true"}, timeout=60,
    ))
    props = meta.get("appProperties") if isinstance(meta.get("appProperties"), Mapping) else {}
    expected_name = plan["name"] if state == "published" else plan["temporary_name"]
    if (meta.get("name") != expected_name or meta.get("mimeType") != GOOGLE_SHEET_MIME
            or folder_id not in (meta.get("parents") or []) or not permissions_are_private(meta)
            or any(props.get(key) != value for key, value in expected_properties(plan, state).items())):
        raise RuntimeError("Google spreadsheet metadata readback mismatch")
    if not check_content:
        return
    exported = session.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}/export",
        params={"mimeType": XLSX_MIME}, timeout=120,
    )
    try:
        if not 200 <= int(exported.status_code) < 300 or workbook_digest(exported.content) != plan["workbook_digest"]:
            raise RuntimeError
    except Exception as exc:
        raise RuntimeError("Google spreadsheet content readback mismatch") from exc


def best_effort_quarantine(session: Any, file_id: str, plan: Mapping[str, Any]) -> None:
    try:
        response_json(session.patch(
            f"https://www.googleapis.com/drive/v3/files/{file_id}",
            params={"supportsAllDrives": "true"},
            json={"name": plan["temporary_name"], "appProperties": expected_properties(plan, "checking")}, timeout=60,
        ))
        readback(session, file_id, str(plan["folder_id"]), plan, state="checking", check_content=False)
    except Exception:
        pass


def atomic_state(path: Path, payload: Mapping[str, Any]) -> None:
    descriptor, name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    temporary = Path(name)
    try:
        os.fchmod(descriptor, 0o600)
        with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, sort_keys=True)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        temporary.replace(path)
        directory = os.open(path.parent, os.O_RDONLY)
        try:
            os.fsync(directory)
        finally:
            os.close(directory)
    finally:
        if temporary.exists():
            temporary.unlink()


def remove_state(path: Path | None) -> None:
    if path is None or not path.exists():
        return
    path.unlink()
    directory = os.open(path.parent, os.O_RDONLY)
    try:
        os.fsync(directory)
    finally:
        os.close(directory)


@contextmanager
def publication_lock(root: Path, day: str, content_sha256: str, folder_id: str):
    root_key = hashlib.sha256(str(root.expanduser().resolve()).encode()).hexdigest()
    generation = hashlib.sha256(f"{folder_id}:{day}:{content_sha256}".encode()).hexdigest()
    directory = Path.home() / ".mango_local" / "google_publish_locks"
    directory.mkdir(parents=True, mode=0o700, exist_ok=True)
    directory.chmod(0o700)
    path = directory / f"{root_key}.lock"
    with path.open("a+", encoding="utf-8") as handle:
        path.chmod(0o600)
        try:
            fcntl.flock(handle, fcntl.LOCK_EX | fcntl.LOCK_NB)
        except BlockingIOError as exc:
            raise RuntimeError("Google publication is already running") from exc
        yield directory / f"{root_key}.{generation}.create.json"


def validate_journal(path: Path, plan: Mapping[str, Any]) -> None:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RuntimeError("Google create journal is invalid") from exc
    if (not isinstance(payload, Mapping) or payload.get("day") != plan["day"]
            or payload.get("content_sha256") != plan["content_sha256"]
            or payload.get("folder_id") != plan["folder_id"]):
        raise RuntimeError("Google create journal belongs to another generation")


def finalize_file(session: Any, file_id: str, folder_id: str, plan: Mapping[str, Any]) -> None:
    response_json(session.patch(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        params={"fields": "id", "supportsAllDrives": "true"},
        json={"name": plan["name"], "appProperties": expected_properties(plan, "published")}, timeout=60,
    ))
    try:
        readback(session, file_id, folder_id, plan, state="published", check_content=False)
    except Exception:
        best_effort_quarantine(session, file_id, plan)
        raise


def publish(session: Any, folder_id: str, plan: Mapping[str, Any], *, journal: Path | None = None) -> Mapping[str, Any]:
    if not re.fullmatch(r"[A-Za-z0-9_-]{10,}", folder_id):
        raise RuntimeError("Google Drive folder id is invalid")
    validate_target_folder(session, folder_id)
    plan = {**plan, "folder_id": folder_id}
    if journal and journal.exists():
        validate_journal(journal, plan)
    query = (
        f"'{folder_id}' in parents and appProperties has {{ key='mango_report_day' and value='{plan['day']}' }} "
        f"and appProperties has {{ key='mango_report_content_sha256' and value='{plan['content_sha256']}' }} "
        f"and appProperties has {{ key='mango_report_schema' and value='{SCHEMA}' }} "
        "and trashed=false"
    )
    lookup = response_json(session.get(
        "https://www.googleapis.com/drive/v3/files",
        params={"q": query, "fields": "files(id,appProperties),nextPageToken", "spaces": "drive", "pageSize": 2,
                "supportsAllDrives": "true", "includeItemsFromAllDrives": "true"}, timeout=60,
    ))
    found = lookup.get("files")
    if not isinstance(found, list) or len(found) > 1 or lookup.get("nextPageToken"):
        raise RuntimeError("Google Drive exact generation lookup is ambiguous")
    if found:
        if not isinstance(found[0], Mapping) or not (file_id := str(found[0].get("id") or "")):
            raise RuntimeError("Google Drive exact generation lookup is invalid")
        props = found[0].get("appProperties") if isinstance(found[0].get("appProperties"), Mapping) else {}
        state = str(props.get("mango_report_state") or "")
        if state not in {"checking", "published"}:
            raise RuntimeError("Google Drive exact generation state is invalid")
        readback(session, file_id, folder_id, plan, state=state, check_content=True)
        if state == "checking":
            finalize_file(session, file_id, folder_id, plan)
            remove_state(journal)
            return {"status": "resumed", "day": plan["day"], "spreadsheet_id": file_id}
        remove_state(journal)
        return {"status": "reused", "day": plan["day"], "spreadsheet_id": file_id}
    if journal and journal.exists():
        raise RuntimeError("Google create result is uncertain; reconcile the quarantined generation before retry")
    boundary = f"mango_{secrets.token_hex(12)}"
    metadata = {
        "name": plan["temporary_name"], "mimeType": GOOGLE_SHEET_MIME, "parents": [folder_id],
        "appProperties": expected_properties(plan, "checking"),
    }
    body = (
        f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{json.dumps(metadata, ensure_ascii=False)}"
        f"\r\n--{boundary}\r\nContent-Type: {XLSX_MIME}\r\n\r\n"
    ).encode() + plan["upload_bytes"] + f"\r\n--{boundary}--\r\n".encode()
    if journal:
        atomic_state(journal, {"day": plan["day"], "content_sha256": plan["content_sha256"],
                               "folder_id": folder_id, "state": "create_pending"})
    created = response_json(session.post(
        "https://www.googleapis.com/upload/drive/v3/files",
        params={"uploadType": "multipart", "fields": "id", "supportsAllDrives": "true"},
        headers={"Content-Type": f"multipart/related; boundary={boundary}"}, data=body, timeout=180,
    ))
    file_id = str(created.get("id") or "")
    if not file_id:
        raise RuntimeError("Google Drive create response has no file id")
    if journal:
        atomic_state(journal, {"day": plan["day"], "content_sha256": plan["content_sha256"],
                               "folder_id": folder_id, "state": "checking", "spreadsheet_id": file_id})
    try:
        readback(session, file_id, folder_id, plan, state="checking", check_content=True)
    except Exception:
        best_effort_quarantine(session, file_id, plan)
        raise
    finalize_file(session, file_id, folder_id, plan)
    remove_state(journal)
    return {"status": "created", "day": plan["day"], "spreadsheet_id": file_id}


def authorized_session(credentials: Path) -> Any:
    from google.auth.transport.requests import AuthorizedSession
    from google.oauth2.service_account import Credentials
    return AuthorizedSession(Credentials.from_service_account_file(
        str(credentials), scopes=["https://www.googleapis.com/auth/drive.file"],
    ))


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Publish verified daily Mango XLSX as Google Sheets.")
    parser.add_argument("--report-root", type=Path, required=True)
    parser.add_argument("--day", type=date.fromisoformat, default=datetime.now(MOSCOW).date() - timedelta(days=1))
    parser.add_argument("--folder-id", default="")
    parser.add_argument("--credentials", type=Path)
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--confirmation", default="")
    args = parser.parse_args(argv)
    plan = load_plan(args.report_root, args.day)
    if not args.execute:
        result = {
            "status": "dry_run",
            "day": plan["day"],
            "name": plan["name"],
            "content_sha256": plan["content_sha256"],
            "external_write_allowed": False,
            "reason": "full_daily_transcript_google_publication_disabled",
        }
    else:
        raise RuntimeError(
            "full daily transcript Google publication is disabled; use the safe current-day projection"
        )
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
