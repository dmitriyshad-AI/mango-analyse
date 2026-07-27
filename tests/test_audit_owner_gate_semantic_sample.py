from __future__ import annotations

import importlib.util
import json
import sqlite3
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook


SCRIPT = Path(__file__).resolve().parents[1] / "scripts" / "audit_owner_gate_semantic_sample.py"
SPEC = importlib.util.spec_from_file_location("audit_owner_gate_semantic_sample", SCRIPT)
assert SPEC and SPEC.loader
MODULE = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MODULE)


def _review_db(path: Path) -> sqlite3.Connection:
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    con.executescript(
        """
        CREATE TABLE family_links_v1 (
          tenant_id TEXT, family_id TEXT, customer_id TEXT, child_key TEXT,
          canonical_name TEXT, grades_json TEXT, subjects_json TEXT, status TEXT, brand TEXT
        );
        CREATE TABLE customer_opportunities (
          tenant_id TEXT, customer_id TEXT, opportunity_id TEXT, title TEXT,
          status TEXT, opened_at TEXT
        );
        CREATE TABLE customer_purchases_v1 (
          tenant_id TEXT, customer_id TEXT, period TEXT, money_kind TEXT,
          total_in REAL, last_purchase_at TEXT, deals_cnt INTEGER
        );
        CREATE TABLE derived_signals (tenant_id TEXT, customer_id TEXT);
        CREATE TABLE timeline_events (
          tenant_id TEXT, customer_id TEXT, event_id TEXT, event_at TEXT,
          event_type TEXT, source_system TEXT, direction TEXT, subject TEXT,
          summary TEXT, text_preview TEXT, source_ref TEXT, superseded_by TEXT
        );
        CREATE TABLE timeline_conflicts (
          tenant_id TEXT, conflict_id TEXT, conflict_type TEXT, severity TEXT,
          status TEXT, created_at TEXT, record_json TEXT
        );
        """
    )
    con.execute(
        "INSERT INTO family_links_v1 VALUES (?,?,?,?,?,?,?,?,?)",
        ("foton", "family:1", "customer:1", "child:1", "Ученик", '["8"]', '["математика"]', "confident", "foton"),
    )
    con.execute(
        "INSERT INTO customer_opportunities VALUES (?,?,?,?,?,?)",
        ("foton", "customer:1", "lead:1", "Курс", "open", "2026-07-01"),
    )
    con.execute(
            "INSERT INTO customer_purchases_v1 VALUES (?,?,?,?,?,?,?)",
        ("foton", "customer:1", "all_time", "fact", 50000, "2026-07-02", 1),
    )
    con.executemany(
        "INSERT INTO timeline_events VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            ("foton", "customer:1", "event:mail", "2026-07-03", "email_message", "mail", "inbound", "Вопрос", "Коротко", "Полный исходный текст письма", "mail:1", None),
            ("foton", "customer:1", "event:visit", "2026-07-04", "tallanto_attendance", "tallanto", "system", "математика", "Посещение", "", "tallanto:1", None),
            ("foton", "customer:other", "event:other", "2026-07-05", "email_message", "mail", "inbound", "Чужое", "Не должно попасть", "Чужой полный текст", "mail:other", None),
        ],
    )
    con.execute(
        "INSERT INTO timeline_conflicts VALUES (?,?,?,?,?,?,?)",
        ("foton", "conflict:1", "shared_family_phone", "high", "open", "2026-07-05", '{"entity_refs":["customer:1"]}'),
    )
    con.commit()
    return con


def test_acceptance_workbook_has_five_raw_review_sheets(tmp_path: Path, monkeypatch) -> None:
    con = _review_db(tmp_path / "review.sqlite")
    monkeypatch.setattr(MODULE, "_family_scope_customer_ids", lambda *_args, **_kwargs: ("customer:1",))
    monkeypatch.setattr(
        MODULE,
        "build_customer_dossier",
        lambda *_args, **_kwargs: SimpleNamespace(
            display_name="Родитель", phone="+70000000000", email="parent@example.com",
            brand="foton", next_step="Ответить", next_step_source="derived_signals:signal:1",
        ),
    )
    families, chronology, evidence, conflicts = MODULE._acceptance_family_data(
        con, tenant_id="foton", sample=[{"id": "customer:1", "family_id": "family:1"}],
    )
    out = tmp_path / "review.xlsx"
    MODULE._write_acceptance_workbook(
        out,
        {
            "Семьи 30": (("family_id", "Родитель"), [[families[0][1], families[0][3]]]),
            "Хронология": (("family_id", "event_id", "Полный текст"), [[row[0], row[2], row[9]] for row in chronology]),
            "Доказательства": (("family_id", "Тип", "source_system", "event_id"), [[row[0], row[1], row[5], row[6]] for row in evidence]),
            "Конфликты": (("family_id", "conflict_id"), [[row[0], row[1]] for row in conflicts]),
            "Owner50": (("family_id", "Статус"), [["family:1", "READY"]]),
        },
    )
    wb = load_workbook(out, read_only=True)
    assert wb.sheetnames == list(MODULE._ACCEPTANCE_SHEETS)
    chronology_values = [value for row in wb["Хронология"].iter_rows(values_only=True) for value in row]
    assert "event:mail" in chronology_values
    assert "Полный исходный текст письма" in chronology_values
    assert "event:other" not in chronology_values
    assert "customer_identities" in [value for row in wb["Доказательства"].iter_rows(values_only=True) for value in row]
    assert "conflict:1" in [value for row in wb["Конфликты"].iter_rows(values_only=True) for value in row]
    assert out.stat().st_mode & 0o777 == 0o600


def test_acceptance_blocks_when_owner50_classification_fails(tmp_path: Path, monkeypatch) -> None:
    con = _review_db(tmp_path / "review.sqlite")
    monkeypatch.setattr(MODULE, "_connect_ro", lambda _db: con)
    monkeypatch.setattr(MODULE, "_source_freshness", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(MODULE, "manager_freshness_gate", lambda _rows: {"passed": True, "blockers": []})
    monkeypatch.setattr(MODULE, "_dossier_population", lambda *_args, **_kwargs: [{
        "id": "customer:1", "family_id": "family:1", "brand": "foton", "channel": "email",
        "child_bucket": "1", "has_payment": True, "has_conflict": False, "has_signal": True,
        "has_mail": True, "has_call": False, "has_attendance": True,
    }])
    monkeypatch.setattr(MODULE, "_acceptance_family_data", lambda *_args, **_kwargs: ([], [], [], []))
    monkeypatch.setattr(MODULE, "_owner50_family_rows", lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("broken")))
    args = SimpleNamespace(db=tmp_path / "review.sqlite", out_root=tmp_path / "out", tenant_id="foton", seed=1, count=1)
    legacy = args.out_root / ".codex_local" / "acceptance_30_families.xlsx"
    legacy.parent.mkdir(parents=True)
    legacy.write_text("old", encoding="utf-8")

    rc = MODULE.cmd_acceptance(args)

    manifest = json.loads((args.out_root / "acceptance_selection_manifest.json").read_text(encoding="utf-8"))
    assert rc == 4
    assert manifest["status"] == "semantic_review_blocked_by_owner50"
    assert manifest["current_artifact"] is None
    assert manifest["legacy_artifact_present"] is True
    assert legacy.read_text(encoding="utf-8") == "old"
    assert not list(legacy.parent.glob("acceptance_30_families_*.xlsx"))


def test_cli_rejects_freshness_bypass(monkeypatch) -> None:
    monkeypatch.setattr("sys.argv", [str(SCRIPT), "acceptance", "--db", "x", "--out-root", "y", "--skip-freshness-gate"])

    with pytest.raises(SystemExit) as exc:
        MODULE.main()

    assert exc.value.code == 2


def test_acceptance_blocks_when_sample_loses_attendance_layer(tmp_path: Path, monkeypatch) -> None:
    con = _review_db(tmp_path / "review.sqlite")
    base = {
        "brand": "foton", "channel": "email", "child_bucket": "1", "has_payment": True,
        "has_conflict": False, "has_signal": True, "has_mail": True, "has_call": False,
    }
    population = [
        {**base, "id": "customer:1", "family_id": "family:1", "has_attendance": False},
        {**base, "id": "customer:2", "family_id": "family:2", "has_attendance": True},
    ]
    monkeypatch.setattr(MODULE, "_connect_ro", lambda _db: con)
    monkeypatch.setattr(MODULE, "_source_freshness", lambda *_args, **_kwargs: [])
    monkeypatch.setattr(MODULE, "manager_freshness_gate", lambda _rows: {"passed": True, "blockers": []})
    monkeypatch.setattr(MODULE, "_dossier_population", lambda *_args, **_kwargs: population)
    monkeypatch.setattr(MODULE, "stratified_sample", lambda *_args, **_kwargs: population[:1])
    args = SimpleNamespace(db=tmp_path / "review.sqlite", out_root=tmp_path / "out", tenant_id="foton", seed=1, count=1)

    rc = MODULE.cmd_acceptance(args)

    manifest = json.loads((args.out_root / "acceptance_selection_manifest.json").read_text(encoding="utf-8"))
    assert rc == 4
    assert manifest["status"] == "semantic_review_blocked_by_sample_coverage"
    assert manifest["missing_layers"] == ["has_attendance"]
    assert not (args.out_root / ".codex_local" / "acceptance_30_families.xlsx").exists()


def test_population_is_unique_by_family_and_conflicts_match_exact_refs(tmp_path: Path) -> None:
    con = _review_db(tmp_path / "population.sqlite")
    con.execute(
        "INSERT INTO family_links_v1 VALUES (?,?,?,?,?,?,?,?,?)",
        ("foton", "family:1", "customer:1b", "child:2", "Второй ребёнок", '["6"]', '["физика"]', "confident", "foton"),
    )
    con.execute("DELETE FROM timeline_conflicts")
    con.execute(
        "INSERT INTO timeline_conflicts VALUES (?,?,?,?,?,?,?)",
        ("foton", "conflict:10", "shared_family_phone", "high", "open", "2026-07-05", '{"entity_refs":["customer:10"]}'),
    )
    con.commit()
    population = MODULE._dossier_population(con, tenant_id="foton")
    assert len(population) == 1
    assert population[0]["family_id"] == "family:1"
    assert population[0]["has_conflict"] is False
