#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def autosize_sheet(ws: Any) -> None:
    for column_cells in ws.columns:
        max_length = 0
        letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 60)


def add_dataframe_sheet(workbook: Workbook, title: str, dataframe: pd.DataFrame) -> None:
    ws = workbook.create_sheet(title=title)
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(dataframe.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, pd.Timestamp):
                if pd.isna(value):
                    value = ""
                else:
                    if value.tzinfo is not None:
                        value = value.tz_convert("UTC").tz_localize(None)
                    value = value.to_pydatetime()
            elif isinstance(value, datetime) and value.tzinfo is not None:
                value = value.astimezone().replace(tzinfo=None)
            ws.cell(row=row_index, column=column_index, value=value)
    autosize_sheet(ws)


def build_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    ws = workbook.active
    ws.title = "Сводка"
    ws["A1"] = "Показатель"
    ws["B1"] = "Значение"
    ws["A1"].font = Font(color="FFFFFF", bold=True)
    ws["B1"].font = Font(color="FFFFFF", bold=True)
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    row = 2
    for key, value in summary.items():
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        row += 1
    autosize_sheet(ws)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, required=True)
    parser.add_argument("--pattern", default="telegram_live_enrichment_20260416_chunk*")
    parser.add_argument("--out-dir", type=Path, required=True)
    args = parser.parse_args()

    chunk_dirs = sorted(path for path in args.root.glob(args.pattern) if path.is_dir())
    if not chunk_dirs:
        raise SystemExit("No chunk directories found.")

    csv_frames: list[pd.DataFrame] = []
    jsonl_rows: list[dict[str, Any]] = []
    chunk_summaries: list[dict[str, Any]] = []
    for chunk_dir in chunk_dirs:
        csv_path = chunk_dir / "telegram_phone_live_enrichment.csv"
        jsonl_path = chunk_dir / "telegram_phone_live_enrichment.jsonl"
        summary_path = chunk_dir / "telegram_phone_live_enrichment_summary.json"
        if not csv_path.exists() or not jsonl_path.exists() or not summary_path.exists():
            raise SystemExit(f"Chunk is incomplete: {chunk_dir}")
        csv_frames.append(pd.read_csv(csv_path))
        with jsonl_path.open(encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                jsonl_rows.append(json.loads(line))
        chunk_summaries.append(json.loads(summary_path.read_text(encoding="utf-8")))

    all_df = pd.concat(csv_frames, ignore_index=True)
    if "Скор полезности" in all_df.columns:
        all_df = all_df.sort_values(["Скор полезности", "Телефон"], ascending=[False, True]).reset_index(drop=True)
    matches_df = all_df[
        (pd.to_numeric(all_df.get("AMO контактов", 0), errors="coerce").fillna(0) > 0)
        | (pd.to_numeric(all_df.get("Tallanto контактов", 0), errors="coerce").fillna(0) > 0)
        | (all_df.get("Есть exact match в рабочем слое", "") == "Да")
    ].copy()
    prospects_df = all_df[
        (all_df.get("Персональное предложение возможно", "") == "Да")
        & (pd.to_numeric(all_df.get("Скор полезности", 0), errors="coerce").fillna(0) >= 20)
    ].copy()

    summary = {
        "chunks": len(chunk_dirs),
        "phones_total": int(len(all_df)),
        "matched_to_working_layer": int((all_df.get("Есть exact match в рабочем слое", "") == "Да").sum()),
        "matched_to_amo": int((pd.to_numeric(all_df.get("AMO контактов", 0), errors="coerce").fillna(0) > 0).sum()),
        "matched_to_tallanto": int((pd.to_numeric(all_df.get("Tallanto контактов", 0), errors="coerce").fillna(0) > 0).sum()),
        "offer_possible": int((all_df.get("Персональное предложение возможно", "") == "Да").sum()),
        "high_utility_phones": int((pd.to_numeric(all_df.get("Скор полезности", 0), errors="coerce").fillna(0) >= 30).sum()),
        "amo_contacts_total": int(pd.to_numeric(all_df.get("AMO контактов", 0), errors="coerce").fillna(0).sum()),
        "tallanto_contacts_total": int(pd.to_numeric(all_df.get("Tallanto контактов", 0), errors="coerce").fillna(0).sum()),
    }

    args.out_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = args.out_dir / "telegram_phone_live_enrichment.jsonl"
    with jsonl_path.open("w", encoding="utf-8") as handle:
        for row in jsonl_rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")

    csv_path = args.out_dir / "telegram_phone_live_enrichment.csv"
    all_df.to_csv(csv_path, index=False, encoding="utf-8-sig")

    workbook = Workbook()
    build_summary_sheet(workbook, summary)
    add_dataframe_sheet(workbook, "Все телефоны", all_df)
    add_dataframe_sheet(workbook, "Только матчи", matches_df)
    add_dataframe_sheet(workbook, "Перспективные", prospects_df)
    xlsx_path = args.out_dir / "telegram_phone_live_enrichment.xlsx"
    workbook.save(xlsx_path)

    summary_path = args.out_dir / "telegram_phone_live_enrichment_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"status": "ok", **summary, "xlsx": str(xlsx_path), "csv": str(csv_path), "jsonl": str(jsonl_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
