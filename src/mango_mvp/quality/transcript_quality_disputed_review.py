from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sqlite3
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import unquote


BACKFILL_VERSION = "safe_non_contentful_v1"
CONTENTFUL_CALL_TYPES = {"sales_call", "service_call", "technical_call", "existing_client_progress"}


@dataclass(frozen=True)
class DisputedReviewConfig:
    database_url: str
    guardrails_root: Path
    out_root: Path
    transcript_jsonl_max_chars: int = 30000
    xlsx_text_max_chars: int = 8000
    human_sample_per_bucket: int = 80


def build_transcript_quality_disputed_review(config: DisputedReviewConfig) -> dict[str, Any]:
    guardrails_root = config.guardrails_root.resolve()
    out_root = config.out_root.resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    db_path = _sqlite_path_from_url(config.database_url)
    auto_fix_path = guardrails_root / "auto_fix_candidates.csv"
    manual_path = guardrails_root / "manual_review_candidates.csv"
    if not auto_fix_path.exists():
        raise FileNotFoundError(f"Missing auto-fix candidates CSV: {auto_fix_path}")
    if not manual_path.exists():
        raise FileNotFoundError(f"Missing manual review candidates CSV: {manual_path}")

    raw_candidates = _merge_candidate_sources(
        _read_csv(auto_fix_path, source="auto_fix_candidates"),
        _read_csv(manual_path, source="manual_review_candidates"),
    )

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    try:
        enriched_rows: list[dict[str, Any]] = []
        llm_tasks: list[dict[str, Any]] = []
        already_backfilled_rows: list[dict[str, Any]] = []
        missing_rows: list[dict[str, Any]] = []

        for row in raw_candidates:
            call_id = _safe_int(row.get("id"))
            if call_id is None:
                missing_rows.append({**row, "review_bucket": "missing_invalid_id"})
                continue
            db_row = conn.execute(
                """
                SELECT id, source_filename, started_at, phone, manager_name, duration_sec,
                       transcription_status, resolve_status, analysis_status,
                       transcript_text, transcript_manager, transcript_client,
                       transcript_variants_json, analysis_json
                  FROM call_records
                 WHERE id = ?
                """,
                (call_id,),
            ).fetchone()
            if db_row is None:
                missing_rows.append({**row, "review_bucket": "missing_db_row"})
                continue

            analysis = _json_object(db_row["analysis_json"])
            backfill_meta = _nested(_nested(analysis, "quality_flags"), "transcript_quality_backfill")
            already_backfilled = backfill_meta.get("version") == BACKFILL_VERSION
            review_bucket = _review_bucket(row, already_backfilled=already_backfilled)
            review_priority = _review_priority(row, review_bucket)
            variants = _variant_texts(db_row["transcript_variants_json"])
            transcript_text = _clean(db_row["transcript_text"])

            enriched = {
                "review_id": _stable_hash(_clean(row.get("id")), _clean(row.get("source_filename")), _clean(row.get("guardrail_label")), _clean(row.get("candidate_sources"))),
                "review_bucket": review_bucket,
                "review_priority": review_priority,
                "recommended_review_owner": _review_owner(review_bucket),
                "candidate_sources": row.get("candidate_sources", ""),
                "already_backfilled": already_backfilled,
                "id": call_id,
                "source_filename": _clean(db_row["source_filename"]) or _clean(row.get("source_filename")),
                "started_at": _clean(db_row["started_at"]) or _clean(row.get("started_at")),
                "month": _clean(row.get("month")) or _month(_clean(db_row["started_at"])),
                "phone": _clean(db_row["phone"]) or _clean(row.get("phone")),
                "manager_name": _clean(db_row["manager_name"]) or _clean(row.get("manager_name")),
                "duration_sec": db_row["duration_sec"],
                "transcription_status": _clean(db_row["transcription_status"]) or _clean(row.get("transcription_status")),
                "resolve_status": _clean(db_row["resolve_status"]) or _clean(row.get("resolve_status")),
                "analysis_status": _clean(db_row["analysis_status"]) or _clean(row.get("analysis_status")),
                "current_call_type": _clean(row.get("current_call_type")) or "unknown",
                "current_contentful": _is_true(row.get("current_contentful")),
                "guardrail_label": _clean(row.get("guardrail_label")),
                "guardrail_score": row.get("guardrail_score", ""),
                "guardrail_reason_codes": _clean(row.get("guardrail_reason_codes")),
                "should_force_non_conversation": _is_true(row.get("should_force_non_conversation")),
                "requires_manual_review": _is_true(row.get("requires_manual_review")),
                "protected_live_dialogue": _is_true(row.get("protected_live_dialogue")),
                "recommended_call_type": _clean(row.get("recommended_call_type")),
                "recommended_contentful": _is_true(row.get("recommended_contentful")),
                "next_step": _clean(row.get("next_step")),
                "products": _clean(row.get("products")),
                "subjects": _clean(row.get("subjects")),
                "objections": _clean(row.get("objections")),
                "history_summary_excerpt": _clean(row.get("history_summary_excerpt")),
                "analysis_history_summary": _analysis_summary(analysis),
                "analysis_next_step": _analysis_next_step(analysis),
                "analysis_products": "|".join(_analysis_products(analysis)),
                "analysis_objections": "|".join(_analysis_objections(analysis)),
                "transcript_chars": len(transcript_text),
                "transcript_excerpt_for_xlsx": _limit_text(transcript_text, config.xlsx_text_max_chars),
                "variant_a_excerpt_for_xlsx": _limit_text(variants.get("variant_a", ""), min(4000, config.xlsx_text_max_chars)),
                "variant_b_excerpt_for_xlsx": _limit_text(variants.get("variant_b", ""), min(4000, config.xlsx_text_max_chars)),
                "review_decision": "",
                "review_comment": "",
            }

            if already_backfilled:
                already_backfilled_rows.append(enriched)
                continue

            enriched_rows.append(enriched)
            llm_tasks.append(_llm_task(enriched, transcript_text, variants, analysis, config.transcript_jsonl_max_chars))
    finally:
        conn.close()

    bucket_counts = Counter(row["review_bucket"] for row in enriched_rows)
    type_counts = Counter(row["current_call_type"] for row in enriched_rows)
    month_counts = Counter(row["month"] or "unknown" for row in enriched_rows)
    source_counts = Counter(row["candidate_sources"] for row in enriched_rows)
    human_priority_rows = _human_priority_sample(enriched_rows, per_bucket=config.human_sample_per_bucket)

    outputs = {
        "summary_json": out_root / "summary.json",
        "report_markdown": out_root / "DISPUTED_REVIEW_REPORT.md",
        "remaining_disputed_candidates_csv": out_root / "remaining_disputed_candidates.csv",
        "llm_review_queue_csv": out_root / "llm_review_queue.csv",
        "human_review_priority_csv": out_root / "human_review_priority.csv",
        "already_backfilled_excluded_csv": out_root / "already_backfilled_excluded.csv",
        "missing_rows_csv": out_root / "missing_rows.csv",
        "bucket_summary_csv": out_root / "bucket_summary.csv",
        "monthly_summary_csv": out_root / "monthly_summary.csv",
        "llm_review_tasks_jsonl": out_root / "llm_review_tasks.jsonl",
        "llm_review_prompt_md": out_root / "LLM_REVIEW_PROMPT.md",
    }
    _write_csv(outputs["remaining_disputed_candidates_csv"], enriched_rows)
    _write_csv(outputs["llm_review_queue_csv"], [row for row in enriched_rows if row["recommended_review_owner"] == "llm_or_human"])
    _write_csv(outputs["human_review_priority_csv"], human_priority_rows)
    _write_csv(outputs["already_backfilled_excluded_csv"], already_backfilled_rows)
    _write_csv(outputs["missing_rows_csv"], missing_rows)
    _write_csv(outputs["bucket_summary_csv"], _counter_rows(bucket_counts, "review_bucket"))
    _write_csv(outputs["monthly_summary_csv"], _counter_rows(month_counts, "month"))
    _write_jsonl(outputs["llm_review_tasks_jsonl"], llm_tasks)
    outputs["llm_review_prompt_md"].write_text(_llm_prompt(), encoding="utf-8")

    xlsx_path, xlsx_error = _write_xlsx_if_available(
        out_root / "transcript_quality_disputed_review.xlsx",
        summary_rows=[
            {"metric": "input_candidate_rows", "value": len(raw_candidates)},
            {"metric": "remaining_disputed_candidates", "value": len(enriched_rows)},
            {"metric": "llm_review_tasks", "value": len(llm_tasks)},
            {"metric": "human_priority_rows", "value": len(human_priority_rows)},
            {"metric": "already_backfilled_excluded", "value": len(already_backfilled_rows)},
            {"metric": "missing_rows", "value": len(missing_rows)},
        ],
        bucket_rows=_counter_rows(bucket_counts, "review_bucket"),
        type_rows=_counter_rows(type_counts, "current_call_type"),
        month_rows=_counter_rows(month_counts, "month"),
        source_rows=_counter_rows(source_counts, "candidate_sources"),
        remaining_rows=enriched_rows,
        human_rows=human_priority_rows,
    )

    summary: dict[str, Any] = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "database_url": _redact_database_url(config.database_url),
        "database_path": str(db_path),
        "guardrails_root": str(guardrails_root),
        "input_candidate_rows": len(raw_candidates),
        "remaining_disputed_candidates": len(enriched_rows),
        "llm_review_tasks": len(llm_tasks),
        "human_priority_rows": len(human_priority_rows),
        "already_backfilled_excluded": len(already_backfilled_rows),
        "missing_rows": len(missing_rows),
        "bucket_counts": dict(bucket_counts.most_common()),
        "current_call_type_counts": dict(type_counts.most_common()),
        "month_counts_top": month_counts.most_common(20),
        "candidate_source_counts": dict(source_counts.most_common()),
        "outputs": {key: str(path) for key, path in outputs.items()},
    }
    if xlsx_path is not None:
        summary["outputs"]["xlsx"] = str(xlsx_path)
    if xlsx_error:
        summary["xlsx_error"] = xlsx_error

    outputs["summary_json"].write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    outputs["report_markdown"].write_text(_markdown_report(summary), encoding="utf-8")
    return summary


def _merge_candidate_sources(*row_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for group in row_groups:
        for row in group:
            key = _clean(row.get("id")) or _stable_hash(_clean(row.get("source_filename")))
            existing = merged.get(key)
            if existing is None:
                merged[key] = dict(row)
                continue
            sources = set(_split_sources(existing.get("candidate_sources"))) | set(_split_sources(row.get("candidate_sources")))
            existing["candidate_sources"] = "|".join(sorted(sources))
            # Prefer manual-review metadata where it adds stricter handling.
            if _is_true(row.get("requires_manual_review")):
                existing.update(row)
                existing["candidate_sources"] = "|".join(sorted(sources))
    return list(merged.values())


def _read_csv(path: Path, *, source: str) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            item = dict(row)
            item["candidate_sources"] = source
            rows.append(item)
    return rows


def _review_bucket(row: dict[str, Any], *, already_backfilled: bool) -> str:
    if already_backfilled:
        return "already_backfilled_safe_non_contentful"
    label = _clean(row.get("guardrail_label"))
    call_type = _clean(row.get("current_call_type")) or "unknown"
    contentful = _is_true(row.get("current_contentful")) or call_type in CONTENTFUL_CALL_TYPES
    force = _is_true(row.get("should_force_non_conversation"))
    manual = _is_true(row.get("requires_manual_review"))

    if label == "manual_review_borderline_live_context":
        return "human_review_borderline_live_context"
    if force and call_type == "sales_call":
        return "human_review_sales_call_conflict"
    if force and contentful:
        return "llm_review_contentful_auto_fix_conflict"
    if manual and contentful:
        return "llm_review_contentful_probable_no_live"
    if manual and not contentful:
        return "llm_review_non_contentful_probable_no_live"
    if force and not contentful:
        return "llm_review_non_contentful_force_candidate"
    return "human_review_uncategorized"


def _review_priority(row: dict[str, Any], review_bucket: str) -> int:
    if review_bucket in {"human_review_sales_call_conflict", "human_review_borderline_live_context"}:
        return 100
    if _clean(row.get("current_call_type")) == "sales_call":
        return 95
    if review_bucket == "llm_review_contentful_auto_fix_conflict":
        return 90
    if review_bucket == "llm_review_contentful_probable_no_live":
        return 80
    if review_bucket == "llm_review_non_contentful_probable_no_live":
        return 60
    if review_bucket == "llm_review_non_contentful_force_candidate":
        return 50
    return 40


def _review_owner(review_bucket: str) -> str:
    if review_bucket.startswith("llm_review_"):
        return "llm_or_human"
    if review_bucket.startswith("human_review_"):
        return "human_required"
    return "skip"


def _human_priority_sample(rows: list[dict[str, Any]], *, per_bucket: int) -> list[dict[str, Any]]:
    selected: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row["recommended_review_owner"] == "human_required":
            selected[row["review_id"]] = row
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        grouped[row["review_bucket"]].append(row)
    for group_rows in grouped.values():
        ordered = sorted(group_rows, key=lambda item: (-int(item["review_priority"]), item["review_id"]))
        for row in ordered[: max(0, per_bucket)]:
            selected[row["review_id"]] = row
    return sorted(selected.values(), key=lambda item: (-int(item["review_priority"]), item["review_bucket"], item["started_at"], item["id"]))


def _llm_task(enriched: dict[str, Any], transcript_text: str, variants: dict[str, str], analysis: dict[str, Any], max_chars: int) -> dict[str, Any]:
    return {
        "task_id": enriched["review_id"],
        "schema": "transcript_quality_disputed_review_v1",
        "required_output_json_schema": {
            "decision": "keep_current_analysis | force_non_conversation | reanalyze_required | human_review_required",
            "confidence": "0.0-1.0",
            "reason": "short Russian explanation",
            "evidence": ["short quotes or paraphrases from transcript/analysis"],
            "safe_to_auto_apply": "boolean",
            "recommended_call_type": "non_conversation | sales_call | service_call | technical_call | existing_client_progress | unknown",
        },
        "call": {
            "id": enriched["id"],
            "source_filename": enriched["source_filename"],
            "started_at": enriched["started_at"],
            "duration_sec": enriched["duration_sec"],
            "manager_name": enriched["manager_name"],
            "phone": enriched["phone"],
        },
        "guardrail": {
            "review_bucket": enriched["review_bucket"],
            "current_call_type": enriched["current_call_type"],
            "current_contentful": enriched["current_contentful"],
            "label": enriched["guardrail_label"],
            "score": enriched["guardrail_score"],
            "reason_codes": enriched["guardrail_reason_codes"],
            "should_force_non_conversation": enriched["should_force_non_conversation"],
            "requires_manual_review": enriched["requires_manual_review"],
            "protected_live_dialogue": enriched["protected_live_dialogue"],
        },
        "current_analysis": {
            "history_summary": _limit_text(_analysis_summary(analysis), max_chars),
            "next_step": _analysis_next_step(analysis),
            "products": _analysis_products(analysis),
            "objections": _analysis_objections(analysis),
            "quality_flags": _nested(analysis, "quality_flags"),
        },
        "transcript_text": _limit_text(transcript_text, max_chars),
        "asr_variants": {key: _limit_text(value, max_chars) for key, value in variants.items() if value},
    }


def _variant_texts(raw: str | None) -> dict[str, str]:
    payload = _json_object(raw)
    if not payload:
        return {}
    out: dict[str, str] = {}
    full = _nested(payload, "full")
    for key in ("variant_a", "variant_b"):
        value = _clean(full.get(key))
        if value:
            out[key] = value
    manager = _nested(payload, "manager")
    client = _nested(payload, "client")
    for key in ("variant_a", "variant_b"):
        parts = []
        m = _clean(manager.get(key))
        c = _clean(client.get(key))
        if m:
            parts.append(f"Менеджер: {m}")
        if c:
            parts.append(f"Клиент: {c}")
        if parts:
            out.setdefault(key, "\n\n".join(parts))
    return out


def _analysis_summary(analysis: dict[str, Any]) -> str:
    return (
        _clean(analysis.get("history_summary"))
        or _clean(analysis.get("history_short"))
        or _clean(analysis.get("summary"))
    )


def _analysis_next_step(analysis: dict[str, Any]) -> str:
    blocks = _nested(analysis, "structured_fields") or _nested(analysis, "crm_blocks")
    next_step = _nested(blocks, "next_step")
    return _clean(next_step.get("action")) or _clean(analysis.get("next_step"))


def _analysis_products(analysis: dict[str, Any]) -> list[str]:
    blocks = _nested(analysis, "structured_fields") or _nested(analysis, "crm_blocks")
    interests = _nested(blocks, "interests")
    values = _clean_list(interests.get("products")) + _clean_list(analysis.get("interests"))
    target = _clean(analysis.get("target_product"))
    if target:
        values.append(target)
    return _unique(values)


def _analysis_objections(analysis: dict[str, Any]) -> list[str]:
    blocks = _nested(analysis, "structured_fields") or _nested(analysis, "crm_blocks")
    return _unique(_clean_list(blocks.get("objections")) + _clean_list(analysis.get("objections")))


def _write_xlsx_if_available(
    path: Path,
    *,
    summary_rows: list[dict[str, Any]],
    bucket_rows: list[dict[str, Any]],
    type_rows: list[dict[str, Any]],
    month_rows: list[dict[str, Any]],
    source_rows: list[dict[str, Any]],
    remaining_rows: list[dict[str, Any]],
    human_rows: list[dict[str, Any]],
) -> tuple[Path | None, str | None]:
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - optional local runtime
        return None, f"pandas/openpyxl unavailable: {exc}"
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="Summary")
        pd.DataFrame(bucket_rows).to_excel(writer, index=False, sheet_name="Buckets")
        pd.DataFrame(type_rows).to_excel(writer, index=False, sheet_name="Call types")
        pd.DataFrame(month_rows).to_excel(writer, index=False, sheet_name="Months")
        pd.DataFrame(source_rows).to_excel(writer, index=False, sheet_name="Sources")
        pd.DataFrame(human_rows).to_excel(writer, index=False, sheet_name="Human priority")
        pd.DataFrame(remaining_rows).to_excel(writer, index=False, sheet_name="All remaining")
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return path, None
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            letter = col[0].column_letter
            header = str(col[0].value or "")
            if "transcript" in header or "summary" in header or "comment" in header:
                ws.column_dimensions[letter].width = 55
            elif header in {"source_filename", "guardrail_reason_codes", "review_bucket"}:
                ws.column_dimensions[letter].width = 34
            else:
                ws.column_dimensions[letter].width = min(24, max(10, len(header) + 2))
    wb.save(path)
    return path, None


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fieldnames: list[str] = []
    for row in rows:
        for key in row:
            if key not in fieldnames:
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def _counter_rows(counter: Counter[str], key_name: str) -> list[dict[str, Any]]:
    return [{key_name: key, "count": count} for key, count in counter.most_common()]


def _llm_prompt() -> str:
    return """# Transcript Quality Disputed Review Prompt

You receive JSONL tasks. For each task decide whether the current analysis is reliable or whether this call should be treated as `non_conversation`.

Primary goal: prevent ASR artifacts, voicemail, robot/system messages, and no-live-dialogue calls from contaminating CRM history, ROP reports, and bot knowledge base.

Decision rules:

- `force_non_conversation`: use only when the transcript is clearly voicemail/system/no live client-manager dialogue or ASR artifact, and existing contentful analysis appears hallucinated or derived from non-dialogue text.
- Treat outbound voicemail as `force_non_conversation`: even if the manager left a meaningful pitch, if the client side is voicemail/virtual secretary/IVR and the client did not participate, it is not a contentful client dialogue.
- No-live markers include voice mail, unavailable/busy/out of coverage, call forwarded, free SMS prompt, press 1, virtual secretary, voice assistant/helper, assistant Mia, company IVR greetings, and call-recording prompts.
- `keep_current_analysis`: use when transcript contains real client-manager dialogue with meaningful sales/service/technical content, even if there are some artifact phrases.
- `reanalyze_required`: use when transcript is real but current analysis is clearly wrong or incomplete; this should go through Analyze again, not be converted to non-conversation.
- `human_review_required`: use when there is a real sales risk, a borderline live dialogue, contradictory ASR variants, or confidence is below 0.75.

Return one JSON object per input task with exactly these fields:

```json
{
  "task_id": "...",
  "decision": "keep_current_analysis | force_non_conversation | reanalyze_required | human_review_required",
  "confidence": 0.0,
  "reason": "Короткое объяснение на русском",
  "evidence": ["краткие признаки из transcript/current_analysis"],
  "safe_to_auto_apply": false,
  "recommended_call_type": "non_conversation | sales_call | service_call | technical_call | existing_client_progress | unknown"
}
```

Set `safe_to_auto_apply=true` only for high-confidence `force_non_conversation` or `keep_current_analysis` decisions. Never auto-apply sales-call conflicts unless confidence is extremely high and transcript is clearly not a conversation.
"""


def _markdown_report(summary: dict[str, Any]) -> str:
    lines = [
        "# Transcript Quality Disputed Review",
        "",
        f"- Generated at: `{summary['generated_at']}`",
        f"- Database: `{summary['database_url']}`",
        f"- Guardrails root: `{summary['guardrails_root']}`",
        f"- Input candidate rows: `{summary['input_candidate_rows']}`",
        f"- Remaining disputed candidates: `{summary['remaining_disputed_candidates']}`",
        f"- LLM review tasks: `{summary['llm_review_tasks']}`",
        f"- Human priority rows: `{summary['human_priority_rows']}`",
        f"- Already backfilled excluded: `{summary['already_backfilled_excluded']}`",
        f"- Missing rows: `{summary['missing_rows']}`",
        "",
        "## Review Buckets",
        "",
    ]
    for label, count in summary["bucket_counts"].items():
        lines.append(f"- `{label}`: `{count}`")
    lines.extend(["", "## Current Call Types", ""])
    for label, count in summary["current_call_type_counts"].items():
        lines.append(f"- `{label}`: `{count}`")
    lines.extend(["", "## Outputs", ""])
    for key, value in summary["outputs"].items():
        lines.append(f"- `{key}`: `{value}`")
    if summary.get("xlsx_error"):
        lines.extend(["", f"XLSX skipped: `{summary['xlsx_error']}`"])
    lines.append("")
    return "\n".join(lines)


def _sqlite_path_from_url(database_url: str) -> Path:
    if database_url.startswith("sqlite:///"):
        return Path(unquote(database_url[len("sqlite:///") :])).expanduser().resolve()
    if database_url.startswith("sqlite://"):
        raw = database_url[len("sqlite://") :]
        return Path(unquote(raw)).expanduser().resolve()
    path = Path(database_url).expanduser().resolve()
    if path.suffix in {".db", ".sqlite", ".sqlite3"}:
        return path
    raise ValueError("Only file-backed SQLite databases are supported for disputed review")


def _redact_database_url(database_url: str) -> str:
    return database_url if database_url.startswith("sqlite") else database_url.split("@")[-1]


def _json_object(raw: Any) -> dict[str, Any]:
    text = _clean(raw)
    if not text:
        return {}
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return {}
    return payload if isinstance(payload, dict) else {}


def _nested(payload: dict[str, Any], key: str) -> dict[str, Any]:
    value = payload.get(key)
    return value if isinstance(value, dict) else {}


def _clean_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return [_clean(item) for item in value if _clean(item)]


def _unique(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        key = value.lower()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def _safe_int(value: Any) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _is_true(value: Any) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "да"}


def _clean(value: Any) -> str:
    return str(value or "").replace("\x00", "").strip()


def _limit_text(value: str, limit: int) -> str:
    text = _clean(value)
    if limit <= 0 or len(text) <= limit:
        return text
    # Keep both beginning and ending; endings often contain the ASR artifact/no-dialogue marker.
    head = max(1, int(limit * 0.65))
    tail = max(1, limit - head - 80)
    return f"{text[:head]}\n\n[... middle truncated, original_chars={len(text)} ...]\n\n{text[-tail:]}"


def _split_sources(value: Any) -> list[str]:
    return [part for part in _clean(value).split("|") if part]


def _stable_hash(*parts: str) -> str:
    digest = hashlib.sha256()
    for part in parts:
        digest.update(part.encode("utf-8", errors="ignore"))
        digest.update(b"\0")
    return digest.hexdigest()[:16]


def _month(started_at: str) -> str:
    return started_at[:7] if len(started_at) >= 7 else "unknown"


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build review queues for disputed transcript-quality candidates.")
    parser.add_argument("--database-url", default=None, help="sqlite:///... URL. Defaults to --db if provided.")
    parser.add_argument("--db", default=None, help="SQLite DB path shortcut.")
    parser.add_argument("--guardrails-root", type=Path, required=True)
    parser.add_argument("--out-root", type=Path, required=True)
    parser.add_argument("--transcript-jsonl-max-chars", type=int, default=30000)
    parser.add_argument("--xlsx-text-max-chars", type=int, default=8000)
    parser.add_argument("--human-sample-per-bucket", type=int, default=80)
    return parser.parse_args(argv)


def config_from_args(args: argparse.Namespace) -> DisputedReviewConfig:
    if args.database_url:
        database_url = args.database_url
    elif args.db:
        database_url = f"sqlite:///{Path(args.db).expanduser().resolve()}"
    else:
        raise SystemExit("Either --database-url or --db is required")
    return DisputedReviewConfig(
        database_url=database_url,
        guardrails_root=args.guardrails_root,
        out_root=args.out_root,
        transcript_jsonl_max_chars=args.transcript_jsonl_max_chars,
        xlsx_text_max_chars=args.xlsx_text_max_chars,
        human_sample_per_bucket=args.human_sample_per_bucket,
    )
