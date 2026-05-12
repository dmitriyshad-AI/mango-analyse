from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from mango_mvp.productization.amo_manual_resolution import build_amo_manual_resolution_pack
from mango_mvp.productization.amo_resolution_pipeline import (
    AmoResolutionAfterXlsxConfig,
    run_amo_resolution_after_xlsx_pipeline,
)
from mango_mvp.productization.amo_resolution_workbook import build_amo_resolution_review_workbook
from tests.test_productization_amo_manual_resolution import _fixture_queue


def test_amo_resolution_after_xlsx_pipeline_stays_fail_closed_without_decisions(tmp_path: Path) -> None:
    queue_root, source = _fixture_queue(tmp_path)
    pack_root = tmp_path / "manual_resolution"
    build_amo_manual_resolution_pack(queue_root=queue_root, source_csv=source, out_root=pack_root)
    workbook = pack_root / "resolution_decisions_manual_template.xlsx"
    build_amo_resolution_review_workbook(pack_root=pack_root, out_xlsx=workbook)
    out_root = tmp_path / "after_xlsx"
    audit_root = tmp_path / "audit_pack"

    result = run_amo_resolution_after_xlsx_pipeline(
        AmoResolutionAfterXlsxConfig(
            project_root=Path.cwd(),
            pack_root=pack_root,
            workbook_path=workbook,
            out_root=out_root,
            audit_pack_root=audit_root,
            update_operator_status=False,
        )
    )

    assert result["manual_resolution_summary"]["resolved_live_candidate_rows"] == 0
    assert result["quality_gate"]["status"] == "skipped_no_resolved_candidates"
    assert result["decision_qa"]["ready_for_quality_gate"] is False
    assert (out_root / "next_real_tunnel_dry_run_command.sh").exists()
    assert "No resolved live candidates" in (out_root / "next_real_tunnel_dry_run_command.sh").read_text(encoding="utf-8")
    assert (audit_root / "README.md").exists()
    assert (audit_root / "pipeline_summary.json").exists()


def test_amo_resolution_after_xlsx_pipeline_builds_candidates_from_accepted_workbook(tmp_path: Path) -> None:
    queue_root, source = _fixture_queue(tmp_path)
    pack_root = tmp_path / "manual_resolution"
    build_amo_manual_resolution_pack(queue_root=queue_root, source_csv=source, out_root=pack_root)
    workbook = pack_root / "resolution_decisions_manual_template.xlsx"
    build_amo_resolution_review_workbook(pack_root=pack_root, out_xlsx=workbook)
    wb = load_workbook(workbook)
    ws = wb["Решения"]
    ws["F2"] = "accepted_by_manager"
    ws["G2"] = "111"
    ws["H2"] = "no"
    ws["I2"] = "source_contact_confirmed_in_amo_post_merge_recheck_approved"
    ws["J2"] = "test_manager"
    wb.save(workbook)
    out_root = tmp_path / "after_xlsx"

    result = run_amo_resolution_after_xlsx_pipeline(
        AmoResolutionAfterXlsxConfig(
            project_root=Path.cwd(),
            pack_root=pack_root,
            workbook_path=workbook,
            out_root=out_root,
            audit_pack_root=tmp_path / "audit_pack",
            run_quality_gate=False,
            update_operator_status=False,
        )
    )

    assert result["manual_resolution_summary"]["accepted_rows"] == 1
    assert result["manual_resolution_summary"]["resolved_live_candidate_rows"] == 1
    assert result["decision_qa"]["ready_for_quality_gate"] is True
    assert result["quality_gate"]["status"] == "skipped_by_config"
    assert "CRM quality gate has not passed" in (out_root / "next_real_tunnel_dry_run_command.sh").read_text(encoding="utf-8")
