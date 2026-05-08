from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


VALIDATION_COLUMNS = ["Решение РОПа", "Комментарий РОПа", "Исправленный ответ / формулировка"]
REVIEW_COLUMNS = [
    "Категория проверки",
    "Приоритет",
    "Сигнал клиента",
    "Стадия",
    "Паттерн ответа",
    "Итог сделки",
    "Коммерческая полезность",
    "Статус для бота",
    "Оценка",
    "Уверенность",
    "Менеджер",
    "Телефон",
    "Дата звонка",
    "Вопрос клиента",
    "Ответ менеджера",
    "Идеальный ответ",
    "Что хорошо",
    "Что упущено",
    "Риски",
    "Что сделать РОПу",
    "Когда не использовать",
    "Ограничение данных",
    "ID момента",
    "Файл звонка",
]


@dataclass(frozen=True)
class ROPValidationPackConfig:
    project_root: Path
    kb_root: Path
    out_root: Path
    top_answers_limit: int = 150
    bot_seeds_limit: int = 250


def build_rop_validation_pack(config: ROPValidationPackConfig) -> dict[str, Any]:
    out_root = config.out_root.resolve()
    out_root.mkdir(parents=True, exist_ok=True)
    enriched = read_csv(config.kb_root / "enriched_reviews.csv")

    top_answers = select_diverse_rows(
        [row for row in enriched if clean(row.get("commercial_usefulness")) == "playbook_candidate"],
        limit=config.top_answers_limit,
        caps={"signal_ru": 12, "answer_pattern_ru": 22, "manager_name": 12},
    )
    revenue_risks = sorted_revenue_risks([row for row in enriched if clean(row.get("commercial_usefulness")) == "revenue_leakage_risk"])
    process_problems = sorted_rows([row for row in enriched if clean(row.get("commercial_usefulness")) == "process_fix_needed"], ascending=True)
    bot_seeds = select_diverse_rows(
        [
            row
            for row in enriched
            if clean(row.get("bot_seed_status")) in {"ready_for_bot_draft", "needs_rop_validation"}
            and clean(row.get("answer_pattern")) != "no_live_contact_or_voicemail"
            and clean(row.get("ideal_answer_example"))
        ],
        limit=config.bot_seeds_limit,
        caps={"signal_ru": 14, "answer_pattern_ru": 28, "manager_name": 18},
    )

    sheets = {
        "Инструкция": instruction_rows(len(enriched), top_answers, revenue_risks, process_problems, bot_seeds),
        "Проверка РОПа": build_combined_review_rows(
            [
                ("ТОП ответ для скрипта", top_answers),
                ("Риск потери выручки", revenue_risks),
                ("Процессная проблема", process_problems),
                ("Черновик для бота", bot_seeds),
            ]
        ),
        "ТОП ответы для скриптов": [validation_row(row, "ТОП ответ для скрипта") for row in top_answers],
        "Риски потери выручки": [validation_row(row, "Риск потери выручки") for row in revenue_risks],
        "Процессные проблемы": [validation_row(row, "Процессная проблема") for row in process_problems],
        "Черновики для бота": [validation_row(row, "Черновик для бота") for row in bot_seeds],
        "Сводка": summary_rows(enriched, top_answers, revenue_risks, process_problems, bot_seeds),
    }

    outputs: dict[str, Path] = {}
    for name, rows in sheets.items():
        if name == "Инструкция":
            continue
        csv_path = out_root / f"{safe_name(name)}.csv"
        write_csv(csv_path, rows)
        outputs[f"{safe_name(name)}_csv"] = csv_path

    xlsx_path = out_root / "ROP_validation_pack_v1.xlsx"
    write_xlsx(xlsx_path, sheets)
    outputs["xlsx"] = xlsx_path

    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "kb_root": str(config.kb_root.resolve()),
        "totals": {
            "source_reviews": len(enriched),
            "top_answers_for_validation": len(top_answers),
            "revenue_risks_for_validation": len(revenue_risks),
            "process_problems_for_validation": len(process_problems),
            "bot_seeds_for_validation": len(bot_seeds),
            "combined_unique_moments": len({row.get("ID момента") for row in sheets["Проверка РОПа"] if row.get("ID момента")}),
        },
        "outputs": {key: str(path) for key, path in outputs.items()},
        "notes": [
            "Файл предназначен для ручной проверки РОПом/методистом, а не для автоматической загрузки в CRM.",
            "Выводы сделаны только по звонкам: WhatsApp/Telegram/email в этом слое не учтены.",
            "После заполнения колонок Решение РОПа / Комментарий / Исправленный ответ можно собрать golden dataset для скриптов и Telegram-бота.",
        ],
    }
    (out_root / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def instruction_rows(
    source_count: int,
    top_answers: list[dict[str, Any]],
    revenue_risks: list[dict[str, Any]],
    process_problems: list[dict[str, Any]],
    bot_seeds: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    return [
        {
            "Раздел": "Назначение",
            "Что смотреть": "Это файл для ручной проверки РОПом перед превращением аналитики в скрипты, обучение и базу Telegram-бота.",
            "Как принять решение": "Заполнять первые три колонки: Решение РОПа, Комментарий РОПа, Исправленный ответ / формулировка.",
        },
        {
            "Раздел": "Ограничение",
            "Что смотреть": "В анализе учтены только звонки.",
            "Как принять решение": "Если менеджер мог закрыть вопрос в WhatsApp/Telegram/email, считать замечание гипотезой для проверки, а не обвинением.",
        },
        {
            "Раздел": "Объем",
            "Что смотреть": f"Источник: {source_count} LLM-review моментов. В pack: {len(top_answers)} топ-ответов, {len(revenue_risks)} рисков выручки, {len(process_problems)} процессных проблем, {len(bot_seeds)} bot-seed строк.",
            "Как принять решение": "Начинать с листа Проверка РОПа; остальные листы дают детализацию по категориям.",
        },
        {
            "Раздел": "Решение РОПа",
            "Что смотреть": "Рекомендуемые значения: одобрить / отклонить / переписать / проверить контекст.",
            "Как принять решение": "Одобрять только то, что можно показывать менеджерам или использовать как основу ответа бота.",
        },
        {
            "Раздел": "Следующий шаг",
            "Что смотреть": "После заполнения файла собрать golden dataset.",
            "Как принять решение": "Golden dataset = утвержденные вопросы клиента, идеальные ответы, сигналы, стадии, правила эскалации на менеджера.",
        },
    ]


def build_combined_review_rows(groups: list[tuple[str, list[dict[str, Any]]]]) -> list[dict[str, Any]]:
    by_id: dict[str, dict[str, Any]] = {}
    categories: dict[str, list[str]] = defaultdict(list)
    for category, rows in groups:
        for row in rows:
            moment_id = clean(row.get("moment_id"))
            if not moment_id:
                continue
            if moment_id not in by_id:
                by_id[moment_id] = row
            if category not in categories[moment_id]:
                categories[moment_id].append(category)
    ranked = sorted(by_id.values(), key=lambda row: (priority_rank(row, categories[clean(row.get("moment_id"))]), -score(row), clean(row.get("moment_id"))))
    return [validation_row(row, " / ".join(categories[clean(row.get("moment_id"))])) for row in ranked]


def validation_row(row: dict[str, Any], category: str) -> dict[str, Any]:
    payload = {
        "Решение РОПа": "",
        "Комментарий РОПа": "",
        "Исправленный ответ / формулировка": "",
        "Категория проверки": category,
        "Приоритет": priority_label(row, category),
        "Сигнал клиента": row.get("signal_ru", ""),
        "Стадия": row.get("stage_ru", ""),
        "Паттерн ответа": row.get("answer_pattern_ru", ""),
        "Итог сделки": row.get("final_outcome_ru", ""),
        "Коммерческая полезность": row.get("commercial_usefulness_ru", ""),
        "Статус для бота": row.get("bot_seed_status_ru", ""),
        "Оценка": row.get("overall_quality_score", ""),
        "Уверенность": row.get("extraction_confidence", ""),
        "Менеджер": row.get("manager_name", ""),
        "Телефон": row.get("phone", ""),
        "Дата звонка": row.get("started_at", ""),
        "Вопрос клиента": row.get("customer_question", ""),
        "Ответ менеджера": row.get("manager_answer", ""),
        "Идеальный ответ": row.get("ideal_answer_example", ""),
        "Что хорошо": row.get("what_manager_did_well", ""),
        "Что упущено": row.get("what_manager_missed", ""),
        "Риски": row.get("risk_flags", ""),
        "Что сделать РОПу": row.get("rop_action", ""),
        "Когда не использовать": row.get("avoid_using_when", ""),
        "Ограничение данных": row.get("data_scope_note", ""),
        "ID момента": row.get("moment_id", ""),
        "Файл звонка": row.get("source_filename", ""),
    }
    return {key: payload.get(key, "") for key in [*VALIDATION_COLUMNS, *REVIEW_COLUMNS]}


def summary_rows(
    enriched: list[dict[str, Any]],
    top_answers: list[dict[str, Any]],
    revenue_risks: list[dict[str, Any]],
    process_problems: list[dict[str, Any]],
    bot_seeds: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows = [
        {"Раздел": "Объем", "Метрика": "Всего review в KB v2", "Значение": len(enriched), "Комментарий": ""},
        {"Раздел": "Объем", "Метрика": "ТОП ответы для проверки", "Значение": len(top_answers), "Комментарий": "Диверсифицированная выборка из playbook_candidate."},
        {"Раздел": "Объем", "Метрика": "Риски потери выручки", "Значение": len(revenue_risks), "Комментарий": "Все строки commercial_usefulness = revenue_leakage_risk."},
        {"Раздел": "Объем", "Метрика": "Процессные проблемы", "Значение": len(process_problems), "Комментарий": "Все строки commercial_usefulness = process_fix_needed."},
        {"Раздел": "Объем", "Метрика": "Черновики для бота", "Значение": len(bot_seeds), "Комментарий": "Диверсифицированная выборка ready/needs validation."},
    ]
    for label, counter in (
        ("Сигналы ТОП ответов", Counter(clean(row.get("signal_ru")) for row in top_answers if clean(row.get("signal_ru")))),
        ("Сигналы рисков выручки", Counter(clean(row.get("signal_ru")) for row in revenue_risks if clean(row.get("signal_ru")))),
        ("Паттерны процессных проблем", Counter(clean(row.get("answer_pattern_ru")) for row in process_problems if clean(row.get("answer_pattern_ru")))),
        ("Сигналы bot-seed", Counter(clean(row.get("signal_ru")) for row in bot_seeds if clean(row.get("signal_ru")))),
    ):
        rows.append({"Раздел": "Распределение", "Метрика": label, "Значение": top_counter(counter), "Комментарий": ""})
    return rows


def select_diverse_rows(rows: list[dict[str, Any]], *, limit: int, caps: dict[str, int]) -> list[dict[str, Any]]:
    ordered = sorted_rows(rows, ascending=False)
    selected: list[dict[str, Any]] = []
    counts: dict[str, Counter[str]] = {field: Counter() for field in caps}
    seen: set[str] = set()
    for row in ordered:
        moment_id = clean(row.get("moment_id"))
        if moment_id in seen:
            continue
        if all(counts[field][clean(row.get(field))] < cap for field, cap in caps.items()):
            selected.append(row)
            seen.add(moment_id)
            for field in caps:
                counts[field][clean(row.get(field))] += 1
        if len(selected) >= limit:
            return selected
    for row in ordered:
        moment_id = clean(row.get("moment_id"))
        if moment_id and moment_id not in seen:
            selected.append(row)
            seen.add(moment_id)
        if len(selected) >= limit:
            break
    return selected


def sorted_rows(rows: Iterable[dict[str, Any]], *, ascending: bool) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            score(row) if ascending else -score(row),
            -confidence(row),
            clean(row.get("signal_ru")),
            clean(row.get("answer_pattern_ru")),
            clean(row.get("moment_id")),
        ),
    )


def sorted_revenue_risks(rows: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda row: (
            clean(row.get("answer_pattern")) == "no_live_contact_or_voicemail",
            score(row),
            -confidence(row),
            clean(row.get("signal_ru")),
            clean(row.get("moment_id")),
        ),
    )


def priority_label(row: dict[str, Any], category: str) -> str:
    if "Риск потери выручки" in category:
        if clean(row.get("answer_pattern")) == "no_live_contact_or_voicemail":
            return "P2 недозвон / контакт"
        return "P0 риск выручки" if score(row) < 45 else "P1 риск выручки"
    if "Процессная проблема" in category:
        return "P1 процесс"
    if "ТОП ответ" in category:
        return "P1 скрипт"
    if "Черновик для бота" in category:
        return "P1 бот" if clean(row.get("bot_seed_status")) == "ready_for_bot_draft" else "P2 бот"
    return "P2 проверка"


def priority_rank(row: dict[str, Any], categories: list[str]) -> tuple[int, int]:
    joined = " / ".join(categories)
    if "Риск потери выручки" in joined:
        if clean(row.get("answer_pattern")) == "no_live_contact_or_voicemail":
            return (2, score(row))
        return (0, score(row))
    if "Процессная проблема" in joined:
        return (1, score(row))
    if "ТОП ответ" in joined:
        return (2, -score(row))
    if "Черновик для бота" in joined:
        return (3, -score(row))
    return (4, 0)


def write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    import pandas as pd

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, rows in sheets.items():
            pd.DataFrame(rows).to_excel(writer, sheet_name=sheet_name, index=False)
        style_workbook(writer.book)


def style_workbook(workbook: Any) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    validation_fill = PatternFill("solid", fgColor="FFF2CC")
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = validation_fill if cell.value in VALIDATION_COLUMNS else header_fill
            cell.font = header_font if cell.value not in VALIDATION_COLUMNS else Font(color="000000", bold=True)
        if sheet.max_row > 1 and sheet.max_column > 0 and sheet.title != "Инструкция":
            validation = DataValidation(type="list", formula1='"одобрить,отклонить,переписать,проверить контекст"', allow_blank=True)
            sheet.add_data_validation(validation)
            validation.add(f"A2:A{sheet.max_row}")
        for column_cells in sheet.columns:
            max_len = 0
            column = column_cells[0].column_letter
            for cell in column_cells[:250]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if len(value) > 80:
                    alignment = copy(cell.alignment)
                    alignment.wrap_text = True
                    alignment.vertical = "top"
                    cell.alignment = alignment
            sheet.column_dimensions[column].width = min(max(max_len + 2, 10), 58)


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def top_counter(counter: Counter[str], limit: int = 8) -> str:
    return " | ".join(f"{key}: {value}" for key, value in counter.most_common(limit))


def score(row: dict[str, Any]) -> int:
    return clamp_int(row.get("overall_quality_score"), 0, 100, 0)


def confidence(row: dict[str, Any]) -> float:
    return clamp_float(row.get("extraction_confidence"), 0.0, 1.0, 0.0)


def clamp_int(value: Any, low: int, high: int, default: int) -> int:
    try:
        parsed = int(round(float(str(value).strip())))
    except (TypeError, ValueError):
        return default
    return max(low, min(high, parsed))


def clamp_float(value: Any, low: float, high: float, default: float) -> float:
    try:
        parsed = float(str(value).strip())
    except (TypeError, ValueError):
        return default
    return max(low, min(high, parsed))


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return " ".join(text.split())


def safe_name(value: str) -> str:
    return {
        "Проверка РОПа": "rop_validation",
        "ТОП ответы для скриптов": "top_script_answers",
        "Риски потери выручки": "revenue_leakage_risks",
        "Процессные проблемы": "process_problems",
        "Черновики для бота": "bot_knowledge_drafts",
        "Сводка": "summary",
    }.get(value, "sheet")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build ROP validation workbook from sales insight KB.")
    parser.add_argument("--project-root", default=".")
    parser.add_argument("--kb-root", default="stable_runtime/sales_insight_knowledge_base_20260507_v2")
    parser.add_argument("--out-root", default="stable_runtime/rop_validation_pack_20260507_v1")
    parser.add_argument("--top-answers-limit", type=int, default=150)
    parser.add_argument("--bot-seeds-limit", type=int, default=250)
    return parser.parse_args(argv)


def config_from_args(args: argparse.Namespace) -> ROPValidationPackConfig:
    project_root = Path(args.project_root).expanduser().resolve()
    return ROPValidationPackConfig(
        project_root=project_root,
        kb_root=(project_root / args.kb_root).resolve(),
        out_root=(project_root / args.out_root).resolve(),
        top_answers_limit=int(args.top_answers_limit),
        bot_seeds_limit=int(args.bot_seeds_limit),
    )


__all__ = [
    "ROPValidationPackConfig",
    "build_rop_validation_pack",
    "config_from_args",
    "parse_args",
    "select_diverse_rows",
]
