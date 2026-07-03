#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from mango_mvp.customer_timeline.a2_mail_ingest import (
    A2V3MailIngestConfig,
    apply_a2v3_mail_ingest,
    create_test_db_backup,
    ensure_not_prod_apply_path,
    validate_a2v3_mail_ingest,
)
from mango_mvp.customer_timeline.canonical_readonly_import import infer_offline_brand
from mango_mvp.customer_timeline.mail_stage2_visibility import assert_mail_stage2_visibility_gate
from mango_mvp.customer_timeline.store import json_loads
try:
    from email_pipeline.quality import evaluate_quality, quality_to_dict, sanitize_summary_payload_for_quality
    from email_pipeline.summary import SummaryItem, split_thread_context, summarize_items
except ModuleNotFoundError:
    from scripts.email_pipeline.quality import evaluate_quality, quality_to_dict, sanitize_summary_payload_for_quality
    from scripts.email_pipeline.summary import SummaryItem, split_thread_context, summarize_items


PROMPT_VERSION = "email_summary_v1_20260703"
SHORT_TEXT_LIMIT = 600
SOURCE_SYSTEM = "mail_archive_stage2"
PROTECTED_TERM_RE = re.compile(
    r"(?<!\d)(?:\d{1,3}(?:[\s\u00a0]\d{3})+|\d+)\s*(?:руб\.?|₽|р\.)|"
    r"\b(?:20\d{2}|19\d{2})[/-]\d{1,2}[/-]\d{1,2}\b|"
    r"\b\d{1,2}[./-]\d{1,2}(?:[./-]\d{2,4})?\b|"
    r"\b\d{1,2}[:.]\d{2}\b|"
    r"\b(?:фотон|унпк|мфти|физтех|cdpofoton|kmipt)\b",
    re.I,
)
NUMERIC_TOKEN_RE = re.compile(r"(?<!\w)\d[\d\s\u00a0.,:/\\-]{0,32}\d|\b\d\b")
COURSE_OR_SUBJECT_RE = re.compile(
    r"\b(?:математик\w*|физик\w*|информатик\w*|программировани\w*|"
    r"русск\w+\s+язык\w*|английск\w+\s+язык\w*|егэ|огэ|олимпиад\w*|"
    r"профильн\w*|базов\w*|курс\w*|групп\w*)\b",
    re.I,
)
PAYMENT_FACT_RE = re.compile(r"\b(?:оплат\w*|предоплат\w*|плат[её]ж\w*|внес\w*|поступил\w*|чек\w*|квитанц\w*)\b", re.I)
REFUND_FACT_RE = re.compile(r"\b(?:возврат\w*|верн\w+\s+деньг\w*|компенсац\w*)\b", re.I)
FORBIDDEN_SUMMARY_RE = re.compile(
    r"\b(?:паспорт\w*|снилс|инн|кпп|бик|р/?с|расч[её]тн\w+\s+сч[её]т|"
    r"корр[её]спондентск\w+\s+сч[её]т|банк\w+\s+реквизит\w*)\b",
    re.I,
)
MODEL_FACT_FIELDS = (
    "student_name",
    "payer_name",
    "contact_name",
    "grade",
    "subject_area",
    "deadline_date",
    "contract_no",
    "document_no",
)
NEXT_STEP_GENERIC_TOKENS = {
    "клиент",
    "клиенту",
    "менеджер",
    "менеджеру",
    "связаться",
    "позвонить",
    "ответить",
    "уточнить",
    "отправить",
    "подготовить",
    "написать",
    "проверить",
    "обсудить",
    "согласовать",
    "напомнить",
    "вернуться",
}


@dataclass(frozen=True)
class EnrichConfig:
    timeline_db: Path
    prod_timeline_db: Path
    allowed_root: Path
    out_dir: Path
    crm_export_dir: Path
    review_workbook: Path | None
    tenant_id: str
    provider: str
    model: str
    reasoning: str
    batch_size: int
    max_llm_calls: int
    codex_home: Path | None
    summarize: bool
    apply: bool


def main(argv: Sequence[str] | None = None) -> int:
    args = _parse_args(argv)
    config = EnrichConfig(
        timeline_db=args.timeline_db.expanduser(),
        prod_timeline_db=args.prod_timeline_db.expanduser(),
        allowed_root=args.allowed_root.expanduser(),
        out_dir=args.out_dir.expanduser(),
        crm_export_dir=args.crm_export_dir.expanduser(),
        review_workbook=args.review_workbook.expanduser() if args.review_workbook else None,
        tenant_id=args.tenant_id,
        provider=args.provider,
        model=args.model,
        reasoning=args.reasoning,
        batch_size=args.batch_size,
        max_llm_calls=args.max_llm_calls,
        codex_home=args.codex_home.expanduser() if args.codex_home else None,
        summarize=bool(args.summarize or args.apply),
        apply=bool(args.apply),
    )
    summary = run(config)
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


def run(config: EnrichConfig) -> Mapping[str, Any]:
    ensure_not_prod_apply_path(config.timeline_db, allowed_root=config.allowed_root)
    if "staging" not in config.timeline_db.resolve(strict=False).parts:
        raise ValueError("mail summary enrich must target .codex_local/staging")
    _ensure_local_staging_out_dir(config.out_dir, allowed_root=config.allowed_root)
    config.out_dir.mkdir(parents=True, exist_ok=True)

    crm_customer_ids = _load_crm_customer_ids(config.crm_export_dir)
    review_customer_ids = _load_review_customer_ids(config.review_workbook) if config.review_workbook else set()
    target_customer_ids = sorted(crm_customer_ids | review_customer_ids)
    target_manifest = {
        "schema_version": "marathon2_block1_mail_summary_targets_v1",
        "crm_customer_ids_count": len(crm_customer_ids),
        "review_customer_ids_count": len(review_customer_ids),
        "target_customer_ids_count": len(target_customer_ids),
        "crm_review_overlap": len(crm_customer_ids & review_customer_ids),
        "customer_ids": target_customer_ids,
    }
    target_manifest_path = config.out_dir / "target_manifest.json"
    target_manifest_path.write_text(json.dumps(target_manifest, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    with sqlite3.connect(config.timeline_db) as con:
        con.row_factory = sqlite3.Row
        rows = _load_target_mail_rows(con, tenant_id=config.tenant_id, customer_ids=target_customer_ids)
        _ensure_cache_table(con)
        cache_before = _cache_counts(con)
        prepared_rows, cache_stats = _prepare_rows_with_summaries(con, rows, config=config)
        cache_after = _cache_counts(con)
        con.commit()

    input_jsonl = config.out_dir / "mail_summary_input.jsonl"
    plan_report = {
        "schema_version": "marathon2_mail_summary_enrich_v1",
        "mode": "apply" if config.apply else ("summarize" if config.summarize else "plan"),
        "target_customers": len(target_customer_ids),
        "crm_customers": len(crm_customer_ids),
        "review_customers": len(review_customer_ids),
        "crm_review_overlap": len(crm_customer_ids & review_customer_ids),
        "target_mail_events": len(rows),
        "input_jsonl": str(input_jsonl),
        "target_manifest": str(target_manifest_path),
        "cache_before": cache_before,
        "cache_after": cache_after,
        "summary_cache": cache_stats,
        "safety": {
            "prod_write": False,
            "crm_write": False,
            "client_sends": False,
            "raw_pii_scope": ".codex_local/staging",
        },
    }
    if cache_stats.get("missing_long_requires_summary") and not config.summarize:
        report_path = config.out_dir / "mail_summary_enrich_report.json"
        report_path.write_text(json.dumps(plan_report, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
        return {**plan_report, "report_path": str(report_path)}
    _write_jsonl(input_jsonl, prepared_rows)
    a2_config = A2V3MailIngestConfig(
        input_jsonl=input_jsonl,
        prod_timeline_db=config.prod_timeline_db,
        timeline_db_path=config.timeline_db,
        allowed_root=config.allowed_root,
        out_dir=config.out_dir / "a2v3_apply",
        tenant_id=config.tenant_id,
        source_ref="marathon2_block1_mail_summaries_20260702",
        enrich_existing=True,
        chunk_rich_text=True,
        refresh_purchases=False,
    )
    validation = validate_a2v3_mail_ingest(a2_config)
    plan_report["validation"] = _compact_validation(validation)
    if config.apply:
        backup = create_test_db_backup(a2_config, label="marathon2_block1_mail_summary")
        apply_report = apply_a2v3_mail_ingest(a2_config, backup_manifest_path=Path(backup["manifest_path"]))
        plan_report["backup"] = {
            "manifest_path": backup["manifest_path"],
            "backup_sha256": backup["backup_sha256"],
        }
        plan_report["apply_report"] = _compact_apply(apply_report)
        with sqlite3.connect(config.timeline_db) as con:
            con.row_factory = sqlite3.Row
            plan_report["post_apply_checks"] = _post_apply_checks(con, allowed_root=config.allowed_root)
    report_path = config.out_dir / "mail_summary_enrich_report.json"
    report_path.write_text(json.dumps(plan_report, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    return {**plan_report, "report_path": str(report_path)}


def _parse_args(argv: Sequence[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Marathon2 Block1 targeted mail summary enrich on staging DB")
    parser.add_argument("--timeline-db", type=Path, required=True)
    parser.add_argument("--prod-timeline-db", type=Path, required=True)
    parser.add_argument("--allowed-root", type=Path, required=True)
    parser.add_argument("--out-dir", type=Path, required=True)
    parser.add_argument("--crm-export-dir", type=Path, required=True)
    parser.add_argument("--review-workbook", type=Path)
    parser.add_argument("--tenant-id", default="foton")
    parser.add_argument("--provider", default="codex_cli")
    parser.add_argument("--model", default="gpt-5.5")
    parser.add_argument("--reasoning", default="high")
    parser.add_argument("--batch-size", type=int, default=10)
    parser.add_argument("--max-llm-calls", type=int, default=100)
    parser.add_argument("--codex-home", type=Path)
    parser.add_argument("--summarize", action="store_true")
    parser.add_argument("--apply", action="store_true")
    return parser.parse_args(argv)


def _load_crm_customer_ids(path: Path) -> set[str]:
    ready = _load_customer_ids_from_jsonl(path / "batch_ready_crm_card_candidates.jsonl")
    if ready:
        return ready
    return _load_customer_ids_from_jsonl(path / "pilot_20_crm_card_candidates.jsonl")


def _load_customer_ids_from_jsonl(path: Path) -> set[str]:
    result: set[str] = set()
    if not path.exists():
        return result
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        customer_id = str(row.get("customer_id") or "").strip()
        if customer_id:
            result.add(customer_id)
    return result


def _load_review_customer_ids(path: Path | None) -> set[str]:
    if path is None or not path.exists():
        return set()
    try:
        from openpyxl import load_workbook
    except Exception:
        return set()
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Обзор клиентов" not in workbook.sheetnames:
        return set()
    sheet = workbook["Обзор клиентов"]
    result: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        value = row[1] if len(row) > 1 else None
        customer_id = str(value or "").strip()
        if customer_id:
            result.add(customer_id)
    return result


def _load_target_mail_rows(
    con: sqlite3.Connection,
    *,
    tenant_id: str,
    customer_ids: Sequence[str],
) -> list[dict[str, Any]]:
    if not customer_ids:
        return []
    placeholders = ",".join("?" for _ in customer_ids)
    rows = con.execute(
        f"""
        WITH phones AS (
          SELECT customer_id, MIN(link_value) AS contact_phone
          FROM identity_links
          WHERE tenant_id = ?
            AND link_type = 'phone'
            AND match_class = 'strong_unique'
          GROUP BY customer_id
        )
        SELECT
          e.event_id, e.tenant_id, e.customer_id, e.event_at, e.source_id, e.source_ref,
          e.direction, e.subject, e.summary, e.text_preview, e.record_json, phones.contact_phone
        FROM timeline_events e
        LEFT JOIN phones ON phones.customer_id = e.customer_id
        WHERE e.tenant_id = ?
          AND e.source_system = ?
          AND COALESCE(e.superseded_by, '') = ''
          AND e.customer_id IN ({placeholders})
        ORDER BY e.customer_id, e.event_at, e.source_id
        """,
        (tenant_id, tenant_id, SOURCE_SYSTEM, *customer_ids),
    ).fetchall()
    result = []
    for line_number, row in enumerate(rows, start=1):
        payload = json_loads(str(row["record_json"] or "{}"))
        record = payload.get("record") if isinstance(payload.get("record"), Mapping) else {}
        metadata = payload.get("metadata") if isinstance(payload.get("metadata"), Mapping) else {}
        full_text = str(record.get("full_clean_text") or "")
        clean_text, thread_context = split_thread_context(full_text)
        body_missing = not bool(clean_text.strip())
        subject = str(row["subject"] or record.get("subject") or "Email message")
        brand = _brand_for_row(row=row, record=record, metadata=metadata, text=clean_text)
        result.append(
            {
                "_line_number": line_number,
                "event_id": str(row["event_id"]),
                "customer_id": str(row["customer_id"]),
                "message_sha256": str(row["source_id"]).lower(),
                "date_iso": str(row["event_at"] or ""),
                "direction": str(row["direction"] or "unknown"),
                "brand": brand,
                "brand_source": str(record.get("brand_source") or record.get("brand_signal") or metadata.get("brand_source") or "staging_event"),
                "raw_infer_offline_brand": brand,
                "classification_reason": "marathon2_existing_stage2_event",
                "subject_full": subject,
                "subject": subject,
                "full_clean_text": clean_text,
                "body_missing": body_missing,
                "thread_context": thread_context,
                "thread_context_source": "raw_body_split_thread_context" if thread_context else "none",
                "full_clean_text_chars": len(clean_text),
                "body_chars": len(full_text),
                "body": clean_text,
                "contact_phone": row["contact_phone"],
                "contact_email": None,
                "contact_name": None,
                "contact_source": "staging_identity_link_phone" if row["contact_phone"] else "missing",
                "contact_missing": not bool(row["contact_phone"]),
                "contact_ambiguous": False,
                "contact_reason": "strong_phone_identity_link" if row["contact_phone"] else "no_strong_phone_identity_link",
                "from_email": None,
                "from_domain": None,
                "to_domains": [],
                "to_emails": [],
                "cc_emails": [],
                "external_recipient_count": 0,
                "outbound_template_freq": 0,
                "is_outbound_template": False,
                "is_mass_recipient": False,
                "has_attachment": False,
            }
        )
    return result


def _brand_for_row(*, row: sqlite3.Row, record: Mapping[str, Any], metadata: Mapping[str, Any], text: str) -> str:
    for value in (record.get("brand"), record.get("brand_signal"), metadata.get("brand")):
        normalized = str(value or "").strip().lower()
        if normalized in {"foton", "unpk"}:
            return normalized
    inferred = infer_offline_brand({"subject": row["subject"], "summary": row["summary"], "full_clean_text": text})
    return inferred if inferred in {"foton", "unpk"} else "unknown"


def _ensure_cache_table(con: sqlite3.Connection) -> None:
    con.execute(
        """
        CREATE TABLE IF NOT EXISTS email_summary_cache_v1 (
          message_sha256 TEXT PRIMARY KEY,
          text_sha256 TEXT NOT NULL,
          prompt_version TEXT NOT NULL,
          provider TEXT NOT NULL,
          model TEXT NOT NULL,
          reasoning TEXT NOT NULL,
          source_kind TEXT NOT NULL,
          summary_text TEXT NOT NULL,
          summary_payload_json TEXT NOT NULL,
          created_at TEXT NOT NULL
        )
        """
    )
    con.execute(
        """
        CREATE INDEX IF NOT EXISTS idx_email_summary_cache_v1_prompt
          ON email_summary_cache_v1(prompt_version, provider, model, reasoning)
        """
    )


def _prepare_rows_with_summaries(
    con: sqlite3.Connection,
    rows: Sequence[Mapping[str, Any]],
    *,
    config: EnrichConfig,
) -> tuple[list[dict[str, Any]], Mapping[str, Any]]:
    prepared = [dict(row) for row in rows]
    missing_long: list[dict[str, Any]] = []
    stats = {
        "rows": len(prepared),
        "short_text_rows": 0,
        "long_text_rows": 0,
        "cache_hits": 0,
        "cache_misses_short": 0,
        "cache_misses_long": 0,
        "llm_calls_total": 0,
        "missing_full_text_rows": 0,
        "fallback_rows": 0,
        "llm_ok": 0,
        "hallucination_suspect": 0,
        "sanitized_payloads": 0,
        "review_reason_counts": {},
    }
    for row in prepared:
        text_hash = _text_hash(row)
        if row.get("body_missing"):
            row["summary_payload"] = _summary_review_needed_payload(
                row,
                reasons=["missing_full_clean_text"],
                original_payload={},
            )
            _cache_put(con, row, text_hash=text_hash, config=config, source_kind="missing_full_text")
            stats["missing_full_text_rows"] += 1
            stats["fallback_rows"] += 1
            continue
        cached = _cache_get(con, row, text_hash=text_hash, config=config)
        if cached is not None:
            row["summary_payload"] = cached
            stats["cache_hits"] += 1
            continue
        if int(row.get("full_clean_text_chars") or 0) < SHORT_TEXT_LIMIT:
            row["summary_payload"] = _short_summary_payload(row)
            _cache_put(con, row, text_hash=text_hash, config=config, source_kind="short_text")
            stats["short_text_rows"] += 1
            stats["cache_misses_short"] += 1
            stats["fallback_rows"] += 1
            continue
        stats["long_text_rows"] += 1
        stats["cache_misses_long"] += 1
        missing_long.append(row)
    if missing_long and not config.summarize:
        stats["missing_long_requires_summary"] = len(missing_long)
        return prepared, stats
    if missing_long:
        stats["summary_review_needed"] = 0
        stats["llm_batches_completed"] = 0
        calls_used = 0
        for offset in range(0, len(missing_long), config.batch_size):
            batch_rows = missing_long[offset : offset + config.batch_size]
            calls_remaining = config.max_llm_calls - calls_used
            if calls_remaining <= 0:
                raise RuntimeError(f"LLM call limit exceeded before batch offset {offset}")
            items = [
                SummaryItem(
                    message_sha256=str(row["message_sha256"]),
                    direction=str(row.get("direction") or "unknown"),
                    brand=str(row.get("brand") or "unknown"),
                    brand_source=str(row.get("brand_source") or "unknown"),
                    subject=str(row.get("subject_full") or ""),
                    body=str(row.get("full_clean_text") or ""),
                )
                for row in batch_rows
            ]
            result = summarize_items(
                items,
                provider=config.provider,
                model=config.model,
                reasoning=config.reasoning,
                batch_size=len(items),
                max_llm_calls=calls_remaining,
                project_root=config.allowed_root,
                codex_home=config.codex_home,
            )
            calls_used += result.llm_calls_total
            stats["llm_calls_total"] = calls_used
            stats["llm_batches_completed"] += 1
            for row in batch_rows:
                payload = result.summaries.get(str(row["message_sha256"]))
                if not payload:
                    raise RuntimeError(f"summary provider did not return message {row['message_sha256']}")
                review_reasons = _anti_hallucination_reasons(row, payload)
                if review_reasons:
                    stats["summary_review_needed"] += 1
                    stats["hallucination_suspect"] += 1
                    _increment_review_reason_counts(stats, review_reasons)
                    payload = _summary_review_needed_payload(row, reasons=review_reasons, original_payload=payload)
                else:
                    stats["llm_ok"] += 1
                row["summary_payload"] = payload
                _cache_put(
                    con,
                    row,
                    text_hash=_text_hash(row),
                    config=config,
                    source_kind="llm_review_needed" if review_reasons else "llm",
                )
            con.commit()
    for row in prepared:
        quality = evaluate_quality(row)
        sanitized = sanitize_summary_payload_for_quality(dict(row.get("summary_payload") or {}), quality)
        if sanitized != row.get("summary_payload"):
            row["summary_payload"] = sanitized
            _cache_put(con, row, text_hash=_text_hash(row), config=config, source_kind="sanitized")
            stats["sanitized_payloads"] += 1
            quality = evaluate_quality(row)
        if _is_summary_review_needed_payload(row.get("summary_payload")):
            quality_dict = quality_to_dict(quality)
            flags = list(quality_dict.get("quality_flags") or [])
            flags.append("summary_review_needed")
            quality_dict["quality_flags"] = list(dict.fromkeys(flags))
            quality_dict["memory_status"] = "summary_review_needed"
            row["quality"] = quality_dict
            continue
        row["quality"] = quality_to_dict(quality)
    return prepared, stats


def _cache_get(
    con: sqlite3.Connection,
    row: Mapping[str, Any],
    *,
    text_hash: str,
    config: EnrichConfig,
) -> dict[str, Any] | None:
    cached = con.execute(
        """
        SELECT summary_payload_json
        FROM email_summary_cache_v1
        WHERE message_sha256 = ?
          AND text_sha256 = ?
          AND prompt_version = ?
          AND provider = ?
          AND model = ?
          AND reasoning = ?
        """,
        (row["message_sha256"], text_hash, PROMPT_VERSION, config.provider, config.model, config.reasoning),
    ).fetchone()
    if cached is None:
        cached = con.execute(
            """
            SELECT text_sha256, summary_payload_json
            FROM email_summary_cache_v1
            WHERE message_sha256 = ?
              AND prompt_version = ?
              AND provider = ?
              AND model = ?
              AND reasoning = ?
            """,
            (row["message_sha256"], PROMPT_VERSION, config.provider, config.model, config.reasoning),
        ).fetchone()
        if cached is None:
            return None
        if str(cached["text_sha256"]) != _legacy_brand_text_hash(row):
            return None
        con.execute(
            """
            UPDATE email_summary_cache_v1
            SET text_sha256 = ?
            WHERE message_sha256 = ?
              AND prompt_version = ?
              AND provider = ?
              AND model = ?
              AND reasoning = ?
            """,
            (text_hash, row["message_sha256"], PROMPT_VERSION, config.provider, config.model, config.reasoning),
        )
    payload = json.loads(str(cached["summary_payload_json"]))
    return payload if isinstance(payload, dict) else None


def _cache_put(
    con: sqlite3.Connection,
    row: Mapping[str, Any],
    *,
    text_hash: str,
    config: EnrichConfig,
    source_kind: str,
) -> None:
    payload = dict(row.get("summary_payload") or {})
    con.execute(
        """
        INSERT INTO email_summary_cache_v1 (
          message_sha256, text_sha256, prompt_version, provider, model, reasoning,
          source_kind, summary_text, summary_payload_json, created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(message_sha256) DO UPDATE SET
          text_sha256 = excluded.text_sha256,
          prompt_version = excluded.prompt_version,
          provider = excluded.provider,
          model = excluded.model,
          reasoning = excluded.reasoning,
          source_kind = excluded.source_kind,
          summary_text = excluded.summary_text,
          summary_payload_json = excluded.summary_payload_json,
          created_at = excluded.created_at
        """,
        (
            row["message_sha256"],
            text_hash,
            PROMPT_VERSION,
            config.provider,
            config.model,
            config.reasoning,
            source_kind,
            str(payload.get("summary") or ""),
            json.dumps(payload, ensure_ascii=False, sort_keys=True),
            datetime.now(timezone.utc).isoformat(),
        ),
    )


def _short_summary_payload(row: Mapping[str, Any]) -> dict[str, Any]:
    text = str(row.get("full_clean_text") or "").strip()
    return {
        "message_sha256": row["message_sha256"],
        "summary": text,
        "topic": str(row.get("subject_full") or "Короткое письмо")[:120],
        "next_step": None,
        "confidence": 0.7,
        "extraction_source": "fallback",
        "event_type": "other",
        "money_direction": "none",
        "student_name": None,
        "payer_name": None,
        "contact_name": None,
        "grade": None,
        "subject_area": None,
        "amount_rub": None,
        "amount_kind": None,
        "amount_is_total": False,
        "amount_items": [],
        "amount_uncertain": False,
        "deadline_date": None,
        "contract_no": None,
        "document_no": None,
        "requisites": [],
        "has_attachment": False,
        "is_plain_acknowledgement": False,
    }


def _anti_hallucination_reasons(row: Mapping[str, Any], payload: Mapping[str, Any]) -> list[str]:
    raw_source = "\n".join([str(row.get("subject_full") or ""), str(row.get("full_clean_text") or "")])
    source_digits = re.sub(r"\D", "", raw_source)
    source = _normalized_guard_text(
        "\n".join(
            [
                str(row.get("subject_full") or ""),
                str(row.get("full_clean_text") or ""),
            ]
        )
    )
    summary_text = _normalized_guard_text(
        "\n".join(_payload_text_values(payload))
    )
    reasons: list[str] = []
    if str(payload.get("extraction_source") or "").strip() != "model":
        reasons.append("extraction_source_not_model")
    if not str(payload.get("summary") or "").strip():
        reasons.append("empty_summary")
    for term in sorted(set(PROTECTED_TERM_RE.findall(summary_text))):
        cleaned = _normalized_guard_text(term)
        if cleaned and cleaned not in source:
            reasons.append(f"new_protected_term:{cleaned[:40]}")
    for term in sorted(set(NUMERIC_TOKEN_RE.findall(summary_text))):
        digits = re.sub(r"\D", "", term)
        if digits and digits not in source_digits:
            reasons.append(f"new_numeric_token:{digits[:24]}")
    for term in sorted(set(COURSE_OR_SUBJECT_RE.findall(summary_text))):
        cleaned = _normalized_guard_text(term)
        if cleaned and cleaned not in source:
            reasons.append(f"new_course_or_subject_term:{cleaned[:40]}")
    for field in MODEL_FACT_FIELDS:
        value = str(payload.get(field) or "").strip()
        if value and _normalized_guard_text(value) not in source:
            reasons.append(f"{field}_not_in_source")
    requisites = payload.get("requisites")
    if isinstance(requisites, list) and requisites:
        reasons.append("requisites_in_summary_payload")
    if FORBIDDEN_SUMMARY_RE.search(summary_text):
        reasons.append("forbidden_requisite_or_document_term")
    event_type = str(payload.get("event_type") or "").strip()
    amount_kind = str(payload.get("amount_kind") or "").strip()
    if PAYMENT_FACT_RE.search(summary_text) and not PAYMENT_FACT_RE.search(raw_source):
        reasons.append("payment_text_not_supported_by_source")
    if REFUND_FACT_RE.search(summary_text) and not REFUND_FACT_RE.search(raw_source):
        reasons.append("refund_text_not_supported_by_source")
    if (event_type == "payment" or amount_kind == "actual_payment") and not PAYMENT_FACT_RE.search(raw_source):
        reasons.append("actual_payment_not_supported_by_source")
    if (event_type == "refund" or amount_kind == "refund") and not REFUND_FACT_RE.search(raw_source):
        reasons.append("refund_not_supported_by_source")
    if _next_step_not_supported(row, payload):
        reasons.append("next_step_not_supported_by_source")
    amount = payload.get("amount_rub")
    if amount not in (None, "") and str(amount) not in source_digits:
        reasons.append("amount_rub_not_in_source")
    for item in payload.get("amount_items") or []:
        if not isinstance(item, Mapping):
            continue
        item_amount = item.get("amount_rub")
        if item_amount not in (None, "") and str(item_amount) not in source_digits:
            reasons.append("amount_item_not_in_source")
    return list(dict.fromkeys(reasons))


def _summary_review_needed_payload(
    row: Mapping[str, Any],
    *,
    reasons: Sequence[str],
    original_payload: Mapping[str, Any],
) -> dict[str, Any]:
    payload = _short_summary_payload(row)
    payload.update(
        {
            "summary": "Требуется ручная проверка модельной выжимки; полный текст письма сохранён в timeline.",
            "topic": str(row.get("subject_full") or "Письмо требует проверки")[:120],
            "next_step": None,
            "confidence": 0.0,
            "extraction_source": "fallback",
            "event_type": "other",
            "money_direction": "none",
            "amount_rub": None,
            "amount_kind": None,
            "amount_uncertain": True,
            "summary_review_needed": True,
            "summary_review_reasons": list(reasons),
            "rejected_summary_payload_sha256": sha256(
                json.dumps(dict(original_payload), ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest(),
        }
    )
    return payload


def _is_summary_review_needed_payload(payload: object) -> bool:
    return isinstance(payload, Mapping) and bool(payload.get("summary_review_needed"))


def _normalized_guard_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "").casefold()).replace("\u00a0", " ").replace("ё", "е").strip()


def _payload_text_values(payload: Mapping[str, Any]) -> list[str]:
    values: list[str] = []
    for key in ("summary", "topic", "next_step", *MODEL_FACT_FIELDS):
        value = payload.get(key)
        if value not in (None, ""):
            values.append(str(value))
    for key in ("requisites", "amount_items"):
        value = payload.get(key)
        if isinstance(value, list):
            for item in value:
                if isinstance(item, Mapping):
                    values.extend(str(v) for v in item.values() if v not in (None, ""))
                elif item not in (None, ""):
                    values.append(str(item))
    return values


def _next_step_not_supported(row: Mapping[str, Any], payload: Mapping[str, Any]) -> bool:
    next_step = str(payload.get("next_step") or "").strip()
    if not next_step:
        return False
    source = _normalized_guard_text("\n".join([str(row.get("subject_full") or ""), str(row.get("full_clean_text") or "")]))
    if not source:
        return True
    meaningful = [
        token
        for token in re.findall(r"[a-zа-яё0-9]{5,}", _normalized_guard_text(next_step), flags=re.I)
        if token not in NEXT_STEP_GENERIC_TOKENS
    ]
    if not meaningful:
        return True
    return not any(token in source for token in meaningful)


def _increment_review_reason_counts(stats: dict[str, Any], reasons: Sequence[str]) -> None:
    raw_counts = stats.setdefault("review_reason_counts", {})
    if not isinstance(raw_counts, dict):
        raw_counts = {}
        stats["review_reason_counts"] = raw_counts
    for reason in reasons:
        key = str(reason).split(":", 1)[0]
        raw_counts[key] = int(raw_counts.get(key, 0)) + 1


def _text_hash(row: Mapping[str, Any]) -> str:
    material = "\n".join(
        [
            str(row.get("subject_full") or ""),
            str(row.get("direction") or ""),
            str(row.get("full_clean_text") or ""),
        ]
    )
    return sha256(material.encode("utf-8", errors="ignore")).hexdigest()


def _legacy_brand_text_hash(row: Mapping[str, Any]) -> str:
    material = "\n".join(
        [
            str(row.get("subject_full") or ""),
            str(row.get("direction") or ""),
            str(row.get("brand") or ""),
            str(row.get("full_clean_text") or ""),
        ]
    )
    return sha256(material.encode("utf-8", errors="ignore")).hexdigest()


def _ensure_local_staging_out_dir(path: Path, *, allowed_root: Path) -> None:
    resolved = path.resolve(strict=False)
    staging_root = (allowed_root / ".codex_local" / "staging").resolve(strict=False)
    try:
        resolved.relative_to(staging_root)
    except ValueError as exc:
        raise ValueError("mail summary enrich out_dir must stay under .codex_local/staging") from exc


def _cache_counts(con: sqlite3.Connection) -> Mapping[str, Any]:
    if (
        con.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='email_summary_cache_v1'"
        ).fetchone()
        is None
    ):
        return {"rows": 0}
    return {
        "rows": int(con.execute("SELECT count(*) FROM email_summary_cache_v1").fetchone()[0]),
        "source_kind_counts": {
            str(row["source_kind"]): int(row["c"])
            for row in con.execute("SELECT source_kind, count(*) AS c FROM email_summary_cache_v1 GROUP BY source_kind")
        },
    }


def _write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")


def _compact_validation(report: Mapping[str, Any]) -> Mapping[str, Any]:
    counts = report.get("counts") if isinstance(report.get("counts"), Mapping) else {}
    return {
        "would_create_events": report.get("would_create_events"),
        "would_skip_existing_test_events": report.get("would_skip_existing_test_events"),
        "would_skip_existing_content_events": report.get("would_skip_existing_content_events"),
        "would_skip_prod_duplicate_events": report.get("would_skip_prod_duplicate_events"),
        "counts": {
            key: counts.get(key)
            for key in (
                "input_rows",
                "linked",
                "blocked",
                "unmatched",
                "bot_eligible_candidate",
                "chunks_planned",
                "customer_brand.foton",
                "customer_brand.unpk",
                "customer_brand.unknown",
            )
            if key in counts
        },
    }


def _compact_apply(report: Mapping[str, Any]) -> Mapping[str, Any]:
    counts = report.get("counts") if isinstance(report.get("counts"), Mapping) else {}
    return {
        "selected_events": report.get("selected_events"),
        "counts": {
            key: counts.get(key, 0)
            for key in (
                "enrich_existing_events",
                "created_events",
                "created_chunks",
                "upserted_a2v3_event_facts",
                "upserted_a2v3_customer_brand_profiles",
                "upserted_customer_purchases_v1",
                "reconciled_a2v3_event_facts",
                "hash_mismatch_events",
                "skipped_identical_events",
            )
        },
    }


def _post_apply_checks(con: sqlite3.Connection, *, allowed_root: Path) -> Mapping[str, Any]:
    db_path = Path(con.execute("PRAGMA database_list").fetchone()[2])
    try:
        assert_mail_stage2_visibility_gate(db_path, allowed_root=allowed_root)
        visibility_assertion = "passed"
    except Exception as exc:
        visibility_assertion = f"failed:{type(exc).__name__}:{exc}"
    return {
        "quick_check": con.execute("PRAGMA quick_check").fetchone()[0],
        "foreign_key_check_rows": len(con.execute("PRAGMA foreign_key_check").fetchall()),
        "a2v3_fact_rows": int(con.execute("SELECT count(*) FROM a2v3_mail_event_facts").fetchone()[0]),
        "a2v3_fact_bot_visible": int(
            con.execute("SELECT count(*) FROM a2v3_mail_event_facts WHERE bot_visible = 1").fetchone()[0]
        ),
        "mail_stage2_unsafe_chunks": int(
            con.execute(
                """
                SELECT count(*)
                FROM bot_context_chunks
                WHERE source_system = ?
                  AND COALESCE(superseded_by, '') = ''
                  AND (allowed_for_bot != 0 OR requires_manager_review != 1)
                """,
                (SOURCE_SYSTEM,),
            ).fetchone()[0]
        ),
        "mail_stage2_visibility_assertion": visibility_assertion,
    }


if __name__ == "__main__":
    sys.exit(main())
