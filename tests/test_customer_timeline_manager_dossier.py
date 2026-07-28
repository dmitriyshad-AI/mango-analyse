from __future__ import annotations

import json
import sqlite3
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from mango_mvp.customer_timeline import manager_dossier as manager_dossier_module
from mango_mvp.customer_timeline.contracts import (
    CustomerIdentity,
    CustomerOpportunity,
    DerivedSignal,
    IdentityStatus,
    OpportunityType,
    SignalSeverity,
    TimelineDirection,
    TimelineEvent,
    TimelineEventType,
)
from mango_mvp.customer_timeline.manager_dossier import (
    OWNER50_REQUIRED_COLUMNS,
    OWNER50_CONTROL_COLUMNS,
    _owner50_event_p0_is_stale_and_resolved,
    _owner50_family_rows,
    _product_interest_values,
    _season_purchase_matches,
    build_customer_dossier,
    build_evidence_record,
    build_manager_dossier_workbook,
    build_owner50_family_workbook,
    classify_family,
    dedupe_family_payment_rows,
    load_canonical_call_client_texts,
    manager_outreach_eligibility,
    owner50_action_text,
    resolve_evidence_source,
)
from mango_mvp.customer_timeline.freshness import (
    MANAGER_REQUIRED_SOURCE_SYSTEMS,
    manager_freshness_gate,
    source_freshness_rows,
)
from mango_mvp.customer_timeline.store import CustomerTimelineSQLiteStore


NOW = datetime(2026, 7, 3, 12, 0, tzinfo=timezone.utc)


def test_product_interest_reads_existing_amo_product_context_fields() -> None:
    values = _product_interest_values(
        None,
        ({"record_json": json.dumps({"product_context": {
            "course": "ОГЭ по физике", "group": "Группа 8", "filial": "МФТИ",
            "subject_name": "Физика",
        }}, ensure_ascii=False)},),
    )

    assert values == ("ОГЭ по физике", "Группа 8", "МФТИ", "Физика")


def test_manager_dossier_extracts_interests_and_pains_from_client_text_only(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    canonical_db = _canonical_calls_db(
        tmp_path,
        {
            "call-client": "Нас интересует летняя школа, телефон +7 916 123-45-67, но переживаем, что не успеваем по математике.",
            "call-manager-only": "",
        },
    )

    with sqlite3.connect(db) as con:
        canonical_calls = load_canonical_call_client_texts(canonical_db)
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=canonical_calls,
        )

    interest_text = "\n".join(item.text for item in dossier.interests)
    pain_text = "\n".join(item.text for item in dossier.pains)
    assert dossier.brand == "foton"
    assert "Из данных: Летняя школа по математике" in interest_text
    assert "Служебная акция из title" not in interest_text
    assert "Нас интересует летняя школа" in interest_text
    assert "916" not in interest_text
    assert "123-45-67" not in interest_text
    assert "[contact]" in interest_text
    assert "переживаем" in pain_text.casefold()
    assert "не успеваем" in pain_text.casefold()
    assert "сложно оплатить" not in pain_text


def test_manager_dossier_names_tallanto_attendance_without_overclaiming_presence(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    with CustomerTimelineSQLiteStore(db, allowed_root=tmp_path) as store:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.TALLANTO_ATTENDANCE,
                event_at=NOW + timedelta(hours=1),
                source_system="tallanto_attendance",
                source_id="attendance-1",
                direction=TimelineDirection.SYSTEM,
                subject="Физика 8 класс",
                summary="Занятие с подтверждённым списанием в Tallanto.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(con, tenant_id="foton", customer_id="customer:1")

    text = "\n".join(row.text for row in dossier.chronology)
    assert "Списание за занятие: Физика 8 класс" in text
    assert "посетил" not in text.casefold()


def test_manager_dossier_excludes_ambiguous_calls(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW + timedelta(minutes=1),
                source_system="mango_processed_summary",
                source_id="ambiguous-call",
                direction=TimelineDirection.INBOUND,
                summary="Неоднозначный звонок не должен попасть в досье.",
                record={"next_step": "Позвонить по неоднозначному звонку"},
                match_status="ambiguous",
                created_at=NOW,
            )
        )
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton",
                customer_id="customer:ambiguous-only",
                identity_status=IdentityStatus.STRONG,
                display_name="Ambiguous only",
                created_at=NOW,
                updated_at=NOW,
            )
        )
        for event_type, source_id, match_status in (
            (TimelineEventType.MANGO_CALL, "ambiguous-only-call", "ambiguous"),
            (TimelineEventType.EMAIL_MESSAGE, "ambiguous-only-mail", "strong_unique"),
        ):
            store.upsert_event(
                TimelineEvent(
                    tenant_id="foton",
                    customer_id="customer:ambiguous-only",
                    event_type=event_type,
                    event_at=NOW,
                    source_system="mango_processed_summary" if event_type == TimelineEventType.MANGO_CALL else "mail_archive_stage2",
                    source_id=source_id,
                    direction=TimelineDirection.INBOUND,
                    summary="Тестовая запись.",
                    match_status=match_status,
                    created_at=NOW,
                )
            )
    finally:
        store.close()
    canonical_db = _canonical_calls_db(tmp_path, {"ambiguous-call": "Нас интересует ошибочный курс."})

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=load_canonical_call_client_texts(canonical_db),
        )

    assert all("ошибочный курс" not in item.text for item in dossier.interests)
    assert all("Неоднозначный звонок" not in item.text for item in dossier.chronology)
    assert "неоднозначному" not in dossier.next_step.casefold()
    summary = build_manager_dossier_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=tmp_path / ".codex_local" / "ambiguous_segment.xlsx",
        enforce_freshness=False,
    )
    assert summary["full_dossier_segment_total"] == 1
    assert summary["customers"] == 1


def test_manager_outreach_eligibility_blocks_safety_risks(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    with sqlite3.connect(db) as con:
        con.execute(
            """
            CREATE TABLE family_links_v1 (
              tenant_id TEXT, customer_id TEXT, status TEXT, confidence TEXT
            )
            """
        )
        event_id = str(con.execute("SELECT event_id FROM timeline_events WHERE source_id='call-client'").fetchone()[0])
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_signal(
            DerivedSignal(
                tenant_id="foton",
                customer_id="customer:1",
                event_id=event_id,
                signal_type="client_returned",
                severity=SignalSeverity.MEDIUM,
                evidence_text="Клиент вернулся после паузы.",
                expires_at=NOW + timedelta(days=30),
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW + timedelta(hours=1),
                source_system="mail_archive_stage2",
                source_id="service-notification",
                direction=TimelineDirection.OUTBOUND,
                summary="Служебное уведомление.",
                record={"service_notification": True},
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.record_conflict(
            "foton",
            conflict_type="ambiguous_identity",
            entity_refs=("customer:10",),
            summary="Prefix collision must not affect customer:1.",
        )
    finally:
        store.close()

    with sqlite3.connect(db) as con:
        eligible = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )
        assert eligible["eligible"] is True
        identity_record = json.loads(
            con.execute("SELECT record_json FROM customer_identities WHERE customer_id='customer:1'").fetchone()[0]
        )
        for alias, expected_brand in (("МФТИ", "unpk"), ("ЦДПО", "foton")):
            identity_record["metadata"]["brands"] = [alias]
            con.execute(
                "UPDATE customer_identities SET record_json=? WHERE customer_id='customer:1'",
                (json.dumps(identity_record, ensure_ascii=False),),
            )
            alias_result = manager_outreach_eligibility(
                con, tenant_id="foton", customer_id="customer:1", as_of=NOW + timedelta(days=1),
            )
            assert "brand_not_exactly_one_known" not in alias_result["reasons"]
            assert alias_result["brand"] == expected_brand
        con.execute("UPDATE timeline_events SET superseded_by='replacement' WHERE event_id=?", (event_id,))
        superseded = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )
        assert "signal_evidence_superseded" in superseded["reasons"]
        con.execute("UPDATE timeline_events SET superseded_by=NULL, match_status='ambiguous' WHERE event_id=?", (event_id,))
        ambiguous = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )
        assert "signal_evidence_ambiguous_call" in ambiguous["reasons"]
        con.execute("UPDATE timeline_events SET match_status='strong_unique' WHERE event_id=?", (event_id,))
        con.execute(
            "INSERT INTO family_links_v1 VALUES (?,?,?,?)",
            ("foton", "customer:1", "ambiguous", "low"),
        )
        blocked = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )
        assert blocked["eligible"] is False
        assert "family_ambiguous" in blocked["reasons"]


def test_manager_outreach_eligibility_blocks_p0_optout_and_meaningful_outbound(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    with sqlite3.connect(db) as con:
        con.execute("CREATE TABLE family_links_v1 (tenant_id TEXT, customer_id TEXT, status TEXT, confidence TEXT)")
        event_id = str(con.execute("SELECT event_id FROM timeline_events WHERE source_id='call-client'").fetchone()[0])
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_signal(
            DerivedSignal(
                tenant_id="foton",
                customer_id="customer:1",
                event_id=event_id,
                signal_type="client_returned",
                severity=SignalSeverity.MEDIUM,
                evidence_text="Клиент вернулся после паузы.",
                expires_at=NOW + timedelta(days=30),
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW + timedelta(hours=1),
                source_system="mail_archive_stage2",
                source_id="p0-refund",
                direction=TimelineDirection.INBOUND,
                summary="Требую возврат денег.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW - timedelta(days=90),
                source_system="mail_archive_stage2",
                source_id="old-optout",
                direction=TimelineDirection.INBOUND,
                summary="Прошу больше со мной не связываться.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW + timedelta(hours=2),
                source_system="mail_archive_stage2",
                source_id="manager-followup",
                direction=TimelineDirection.OUTBOUND,
                summary="Менеджер уже отправил персональное предложение.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
    finally:
        store.close()

    with sqlite3.connect(db) as con:
        blocked = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )

    assert blocked["eligible"] is False
    assert {"durable_p0_history", "durable_opt_out", "meaningful_outbound_after_evidence"}.issubset(
        blocked["reasons"]
    )


def test_manager_outreach_blocks_structured_refund_with_neutral_text(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    with sqlite3.connect(db) as con:
        con.execute(
            "CREATE TABLE family_links_v1 ("
            "tenant_id TEXT, family_id TEXT, customer_id TEXT, status TEXT, confidence TEXT)"
        )
        con.executemany(
            "INSERT INTO family_links_v1 VALUES (?,?,?,?,?)",
            (
                ("foton", "family:1", "customer:1", "confident", "high"),
                ("foton", "family:1", "customer:2", "confident", "high"),
            ),
        )
        event_id = str(con.execute("SELECT event_id FROM timeline_events WHERE source_id='call-client'").fetchone()[0])
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton",
                customer_id="customer:2",
                identity_status=IdentityStatus.STRONG,
                display_name="Second family member",
                created_at=NOW,
                updated_at=NOW,
            )
        )
        store.upsert_signal(
            DerivedSignal(
                tenant_id="foton",
                customer_id="customer:1",
                event_id=event_id,
                signal_type="client_returned",
                severity=SignalSeverity.MEDIUM,
                evidence_text="Клиент вернулся после паузы.",
                expires_at=NOW + timedelta(days=30),
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:2",
                event_type=TimelineEventType.TALLANTO_PAYMENT,
                event_at=NOW - timedelta(days=2),
                source_system="tallanto_crm_call",
                source_id="structured-refund",
                direction=TimelineDirection.SYSTEM,
                summary="Структурная операция Tallanto.",
                record={"payment_direction": "school_out", "direction": "refund", "amount": 1000},
                match_status="strong_unique",
                created_at=NOW - timedelta(days=2),
            )
        )
    finally:
        store.close()

    with sqlite3.connect(db) as con:
        blocked = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )

    assert blocked["eligible"] is False
    assert "durable_p0_history" in blocked["reasons"]


def test_manager_outreach_eligibility_blocks_risk_signal_and_exact_conflict(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    with sqlite3.connect(db) as con:
        con.execute("CREATE TABLE family_links_v1 (tenant_id TEXT, customer_id TEXT, status TEXT, confidence TEXT)")
        event_id = str(con.execute("SELECT event_id FROM timeline_events WHERE source_id='call-client'").fetchone()[0])
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        for signal_type in ("client_returned", "paid_no_access"):
            store.upsert_signal(
                DerivedSignal(
                    tenant_id="foton",
                    customer_id="customer:1",
                    event_id=event_id,
                    signal_type=signal_type,
                    severity=SignalSeverity.HIGH if signal_type == "paid_no_access" else SignalSeverity.MEDIUM,
                    evidence_text=f"Evidence for {signal_type}.",
                    expires_at=NOW + timedelta(days=30),
                    created_at=NOW,
                )
            )
        store.record_conflict(
            "foton",
            conflict_type="ambiguous_identity",
            entity_refs=("customer:1",),
            summary="Exact customer conflict.",
        )
    finally:
        store.close()

    with sqlite3.connect(db) as con:
        blocked = manager_outreach_eligibility(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            as_of=NOW + timedelta(days=1),
        )

    assert blocked["eligible"] is False
    assert {"active_risk_signal:paid_no_access", "open_identity_conflict"}.issubset(blocked["reasons"])


def test_manager_dossier_workbook_stays_under_allowed_root_and_writes_summary(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    canonical_db = _canonical_calls_db(tmp_path, {"call-client": "Хотим рассмотреть курс, но сложно по времени."})
    out = tmp_path / ".codex_local" / "review" / "dossier.xlsx"

    summary = build_manager_dossier_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        customer_ids=("customer:1",),
        canonical_calls_db=canonical_db,
        enforce_freshness=False,
    )

    assert summary["customers"] == 1
    assert summary["interests_total"] >= 2
    assert summary["pains_total"] == 1
    assert summary["safety"]["write_crm"] is False
    assert out.exists()
    assert out.stat().st_mode & 0o777 == 0o600
    summary_path = out.with_suffix(".summary.json")
    assert json.loads(summary_path.read_text(encoding="utf-8"))["customers"] == 1
    wb = load_workbook(out, read_only=True)
    assert "Оглавление" in wb.sheetnames
    assert "Клиент 1" in wb.sheetnames
    values = [row for row in wb["Клиент 1"].iter_rows(values_only=True)]
    assert ("Интересы", "Из данных: Летняя школа по математике", "Данные клиента/сделки") in values
    assert any(row[0] == "Боли" and "сложно по времени" in str(row[1]).casefold() for row in values)

    with pytest.raises(ValueError, match=".codex_local"):
        build_manager_dossier_workbook(
            timeline_db=db,
            allowed_root=tmp_path,
            out_xlsx=tmp_path / "dossier.xlsx",
            customer_ids=("customer:1",),
        )

    with pytest.raises(ValueError, match="allowed root"):
        build_manager_dossier_workbook(
            timeline_db=db,
            allowed_root=tmp_path / "inside",
            out_xlsx=tmp_path / "outside.xlsx",
            customer_ids=("customer:1",),
        )


def test_manager_dossier_workbook_includes_full_manager_sections(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db)
    reconcile = tmp_path / ".codex_local" / "reconcile.json"
    reconcile.parent.mkdir(parents=True)
    reconcile.write_text(
        json.dumps(
            {
                "status": "checked",
                "generated_at": "2026-07-03T12:00:00+00:00",
                "customers_checked": 42,
                "customers_changed": 0,
                "snapshot_stale": False,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    out = tmp_path / ".codex_local" / "review" / "dossier_full.xlsx"

    summary = build_manager_dossier_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        customer_ids=("customer:1",),
        canonical_calls_db=tmp_path / "missing.sqlite",
        reconcile_json=reconcile,
        enforce_freshness=False,
    )

    assert summary["canonical_calls_loaded"] == 0
    assert "not found" in summary["canonical_calls_warning"]
    assert summary["family_rows_total"] == 1
    assert summary["money_rows_total"] == 2
    assert summary["signals_total"] == 1
    assert summary["objections_total"] == 1
    assert summary["next_step_rows_total"] == 1
    assert summary["missing_next_step_rows_total"] == 0
    wb = load_workbook(out, read_only=True)
    values = [row for row in wb["Клиент 1"].iter_rows(values_only=True)]
    joined = "\n".join(str(cell) for row in values for cell in row if cell)
    assert "0 расхождений из 42" in joined
    assert "Иван" in joined and "класс: 8" in joined and "предметы: математика" in joined
    assert "факт оплат" in joined and "120 000 руб." in joined
    assert "план сделок" in joined and "80 000 руб." in joined
    assert "Бренд\nfoton" in joined
    assert "списания/расход" in joined
    assert "возвраты/исход" not in joined
    assert "сделка зависла" in joined
    assert ("Следующий шаг", "Позвонить в понедельник по оплате", "Сигнал Customer Timeline") in values
    assert values[0] == ("Раздел", "Значение", "Откуда")
    assert "family_links_v1" not in joined
    assert "derived_signals" not in joined
    assert "price: Дорого, просит рассрочку." in joined
    assert "Письмо «Расписание»: полный текст в базе." in joined
    assert "Требуется ручная проверка модельной выжимки" not in joined


def test_manager_season_evidence_accepts_tallanto_balance_charge(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db)
    evidence_at = datetime(2026, 6, 1, tzinfo=timezone.utc)

    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        assert _season_purchase_matches(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            evidence_at=evidence_at,
            as_of=NOW,
        )
        con.execute(
            "UPDATE customer_purchases_v1 SET total_out=1000 WHERE customer_id='customer:1' AND money_kind='fact'"
        )
        assert _season_purchase_matches(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            evidence_at=evidence_at,
            as_of=NOW,
        )
        future = NOW + timedelta(days=10)
        con.execute(
            "UPDATE customer_purchases_v1 SET last_purchase_at=? "
            "WHERE customer_id='customer:1' AND money_kind='fact'",
            (future.isoformat(),),
        )
        assert not _season_purchase_matches(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            evidence_at=future,
            as_of=NOW,
        )


def test_manager_dossier_marks_cross_brand_family_as_historical(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db)
    with sqlite3.connect(db) as con:
        con.execute("UPDATE family_links_v1 SET brand='unpk' WHERE customer_id='customer:1'")
        dossier = build_customer_dossier(con, tenant_id="foton", customer_id="customer:1")

    assert dossier.brand == "foton"
    assert any("исторический другой бренд" in row.text for row in dossier.family)


def test_manager_dossier_skips_missing_customer_ids(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    out = tmp_path / ".codex_local" / "review" / "dossier_missing.xlsx"

    summary = build_manager_dossier_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        customer_ids=("customer:1", "customer:missing"),
        enforce_freshness=False,
    )

    assert summary["requested_customers"] == 2
    assert summary["customers"] == 1
    assert summary["missing_customer_ids_count"] == 1
    assert summary["missing_customer_ids_sample"] == ["customer:missing"]
    assert out.exists()

    empty = build_manager_dossier_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=tmp_path / ".codex_local" / "review" / "dossier_empty.xlsx",
        customer_ids=(),
        enforce_freshness=False,
    )
    assert empty["requested_customers"] == 0
    assert empty["customers"] == 0


def test_manager_dossier_omits_generic_history_next_step(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db, signal_action="Посмотреть историю и ответить с учётом прошлого запроса клиента.")
    out = tmp_path / ".codex_local" / "review" / "dossier_no_generic_next.xlsx"

    summary = build_manager_dossier_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        customer_ids=("customer:1",),
        enforce_freshness=False,
    )

    assert summary["signals_total"] == 1
    assert summary["next_step_rows_total"] == 0
    assert summary["missing_next_step_rows_total"] == 1
    wb = load_workbook(out, read_only=True)
    values = [row for row in wb["Клиент 1"].iter_rows(values_only=True)]
    assert (
        "Следующий шаг",
        "Не определён: менеджеру нужно выбрать действие после проверки истории.",
        "Требует решения менеджера",
    ) in values
    joined = "\n".join(str(cell) for row in values for cell in row if cell is not None)
    assert "Посмотреть историю" not in joined


def test_manager_dossier_prefers_resolved_active_timeline_step(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db)
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW + timedelta(minutes=1),
                source_system="mango_call",
                source_id="call-next-step",
                direction=TimelineDirection.INBOUND,
                summary="Договорились отправить материалы клиенту.",
                record={"next_step": "Отправить материалы клиенту"},
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.record_conflict(
            "foton",
            conflict_type="ambiguous_identity",
            entity_refs=("customer:10",),
            summary="Prefix collision must not affect customer:1.",
        )
    finally:
        store.close()

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(con, tenant_id="foton", customer_id="customer:1")

    assert dossier.next_step.startswith("Отправить материалы клиенту (")
    assert dossier.next_step_source == "timeline_events"


def test_manager_dossier_does_not_fall_back_to_signal_after_step_closed(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db)
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW + timedelta(minutes=1),
                source_system="mango_call",
                source_id="call-next-step",
                direction=TimelineDirection.INBOUND,
                record={"next_step": "Отправить материалы клиенту"},
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW + timedelta(minutes=2),
                source_system="mail_archive_stage2",
                source_id="mail-materials-sent",
                direction=TimelineDirection.OUTBOUND,
                summary="Материалы отправлены клиенту.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
    finally:
        store.close()

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(con, tenant_id="foton", customer_id="customer:1")

    assert dossier.next_step == ""
    assert dossier.next_step_source == ""


def test_manager_dossier_does_not_show_step_with_open_ambiguous_identity(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    _seed_full_dossier_tables(db)
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW + timedelta(minutes=1),
                source_system="mango_call",
                source_id="call-next-step",
                direction=TimelineDirection.INBOUND,
                record={"next_step": "Отправить материалы клиенту"},
                match_status="strong_unique",
                created_at=NOW,
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT INTO timeline_conflicts VALUES (?,?,?,?,?,?,?,?,?)",
            (
                "conflict:ambiguous",
                "foton",
                "ambiguous_identity",
                "high",
                "open",
                NOW.isoformat(),
                None,
                "hash-conflict",
                json.dumps({"entity_refs": ["customer:1"]}),
            ),
        )
        con.commit()
        dossier = build_customer_dossier(con, tenant_id="foton", customer_id="customer:1")

    assert dossier.next_step == ""
    assert dossier.next_step_source == ""


def test_manager_freshness_gate_blocks_missing_or_stale_sources() -> None:
    rows = [
        {
            "source_system": "amocrm_snapshot",
            "expected": True,
            "missing": False,
            "cursor_complete": False,
            "imported_at": "2026-07-20T00:00:00+00:00",
        },
        {
            "source_system": "mail_archive_stage2",
            "expected": True,
            "missing": True,
            "cursor_complete": False,
            "imported_at": None,
        },
    ]

    gate = manager_freshness_gate(rows, now=datetime(2026, 7, 22, tzinfo=timezone.utc))

    assert gate["passed"] is False
    assert {item["reason"] for item in gate["blockers"]} == {
        "cursor_incomplete",
        "successful_import_stale",
        "missing",
    }


def test_manager_freshness_gate_blocks_missing_and_future_import_times() -> None:
    rows = [
        {
            "source_system": "wappi_max",
            "expected": True,
            "missing": False,
            "cursor_complete": True,
            "imported_at": None,
        },
        {
            "source_system": "wappi_telegram",
            "expected": True,
            "missing": False,
            "cursor_complete": True,
            "imported_at": "2026-07-22T00:06:00+00:00",
        },
    ]

    gate = manager_freshness_gate(rows, now=datetime(2026, 7, 22, tzinfo=timezone.utc))

    assert gate["passed"] is False
    assert {item["reason"] for item in gate["blockers"]} == {
        "successful_import_missing",
        "imported_at_in_future",
    }


def test_manager_freshness_requires_tallanto_payments_and_attendance_data() -> None:
    assert {"tallanto_crm_call", "tallanto_attendance_api"} <= set(MANAGER_REQUIRED_SOURCE_SYSTEMS)
    rows = [
        {"source_system": source, "expected": True, "missing": False, "cursor_complete": True,
         "cursor_updated_at": "2026-07-22T00:00:00+00:00", "imported_at": "2026-07-22T00:00:00+00:00",
         "max_event_at": None}
        for source in ("tallanto_crm_call", "tallanto_attendance_api")
    ]

    gate = manager_freshness_gate(rows, now=datetime(2026, 7, 22, tzinfo=timezone.utc))

    assert gate["passed"] is False
    assert gate["blockers"] == [
        {"source_system": "tallanto_crm_call", "reason": "data_boundary_missing"},
        {"source_system": "tallanto_attendance_api", "reason": "data_boundary_missing"},
    ]


def test_source_freshness_reads_attendance_increment_and_cursor() -> None:
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.executescript(
        """
        CREATE TABLE timeline_events (tenant_id TEXT, source_system TEXT, event_at TEXT);
        CREATE TABLE ingestion_cursors (tenant_id TEXT, source_system TEXT, last_cursor_ts TEXT, updated_at TEXT);
        CREATE TABLE ingestion_runs (tenant_id TEXT, source_system TEXT, source_ref TEXT, run_kind TEXT, status TEXT, finished_at TEXT);
        INSERT INTO timeline_events VALUES ('foton','tallanto_attendance_api','2026-07-21T12:00:00+00:00');
        INSERT INTO ingestion_cursors VALUES ('foton','tallanto_attendance_api','2026-07-22T00:00:00+00:00','2026-07-22T00:01:00+00:00');
        INSERT INTO ingestion_runs VALUES ('foton','tallanto_attendance_api','tallanto:attendance-api','tallanto_attendance_api_increment','completed','2026-07-22T00:02:00+00:00');
        """
    )

    row = next(item for item in source_freshness_rows(
        con, expected_sources=("tallanto_attendance_api",)
    ) if item["source_system"] == "tallanto_attendance_api")

    assert row["imported_at"] == "2026-07-22T00:02:00+00:00"
    assert row["cursor_complete"] is True
    assert row["events"] == 1


def test_source_freshness_rejects_local_payment_snapshot_as_api_proof() -> None:
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.executescript(
        """
        CREATE TABLE timeline_events (tenant_id TEXT, source_system TEXT, event_at TEXT);
        CREATE TABLE ingestion_runs (tenant_id TEXT, source_system TEXT, source_ref TEXT, run_kind TEXT, status TEXT, finished_at TEXT);
        INSERT INTO timeline_events VALUES ('foton','tallanto_crm_call','2026-07-01T00:00:00+00:00');
        INSERT INTO ingestion_runs VALUES ('foton','tallanto_crm_call','old-local.json','timeline_import','completed','2026-07-22T00:00:00+00:00');
        """
    )

    row = next(item for item in source_freshness_rows(
        con, expected_sources=("tallanto_crm_call",)
    ) if item["source_system"] == "tallanto_crm_call")

    assert row["imported_at"] is None
    assert manager_freshness_gate([row], now=datetime(2026, 7, 22, tzinfo=timezone.utc))["passed"] is False


def test_source_freshness_uses_live_cards_cursor_not_bootstrap_import() -> None:
    con = sqlite3.connect(":memory:")
    con.row_factory = sqlite3.Row
    con.executescript(
        """
        CREATE TABLE timeline_events (tenant_id TEXT, source_system TEXT, event_at TEXT);
        CREATE TABLE ingestion_cursors (tenant_id TEXT, source_system TEXT, last_cursor_ts TEXT, updated_at TEXT);
        CREATE TABLE ingestion_runs (tenant_id TEXT, source_system TEXT, source_ref TEXT, run_kind TEXT, status TEXT, finished_at TEXT);
        INSERT INTO timeline_events VALUES ('foton','tallanto_snapshot','2026-01-01T00:00:00+00:00');
        INSERT INTO ingestion_cursors VALUES ('foton','tallanto_cards_daily','2026-07-22T00:00:00+00:00','2026-07-22T00:01:00+00:00');
        INSERT INTO ingestion_runs VALUES ('foton','tallanto_cards_daily','tallanto:contacts:daily','timeline_import','completed','2026-07-22T00:02:00+00:00');
        """
    )

    row = next(item for item in source_freshness_rows(
        con, expected_sources=("tallanto_snapshot",)
    ) if item["source_system"] == "tallanto_snapshot")

    assert row["cursor_sources"] == ["tallanto_cards_daily"]
    assert row["imported_at"] == "2026-07-22T00:02:00+00:00"
    assert manager_freshness_gate([row], now=datetime(2026, 7, 22, 1, tzinfo=timezone.utc))["passed"] is True


def test_manager_freshness_gate_does_not_accept_local_amo_reindex() -> None:
    rows = [
        {
            "source_system": "amocrm_snapshot",
            "expected": True,
            "missing": False,
            "cursor_complete": True,
            "cursor_updated_at": "2026-07-18T00:00:00+00:00",
            "imported_at": "2026-07-22T00:00:00+00:00",
            "max_event_at": "2026-07-21T12:00:00+00:00",
        }
    ]

    gate = manager_freshness_gate(rows, now=datetime(2026, 7, 22, tzinfo=timezone.utc))

    assert gate["passed"] is False
    assert gate["blockers"] == [
        {"source_system": "amocrm_snapshot", "reason": "cursor_check_stale"}
    ]


def test_manager_freshness_gate_blocks_future_but_accepts_old_contact_after_fresh_scan() -> None:
    rows = [
        {
            "source_system": "wappi_max",
            "expected": True,
            "missing": False,
            "cursor_complete": True,
            "cursor_updated_at": "2026-07-22T00:00:00+00:00",
            "imported_at": "2026-07-22T00:00:00+00:00",
            "max_event_at": "2026-07-23T00:00:00+00:00",
        },
        {
            "source_system": "tallanto_snapshot",
            "expected": True,
            "missing": False,
            "cursor_complete": True,
            "cursor_updated_at": "2026-07-22T00:00:00+00:00",
            "imported_at": "2026-07-22T00:00:00+00:00",
            "max_event_at": "2026-05-01T00:00:00+00:00",
        },
    ]

    gate = manager_freshness_gate(rows, now=datetime(2026, 7, 22, tzinfo=timezone.utc))

    assert gate["blockers"] == [{"source_system": "wappi_max", "reason": "max_event_at_in_future"}]


def test_manager_dossier_matches_canonical_call_id_and_prefixed_source_id(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW,
                source_system="mango_call",
                source_id="source-without-canonical-match",
                direction=TimelineDirection.INBOUND,
                summary="Summary не должен использоваться.",
                match_status="strong_unique",
                record={"canonical_call_id": "canonical-from-record"},
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW,
                source_system="mango_call",
                source_id="call:prefixed-id",
                direction=TimelineDirection.INBOUND,
                summary="Summary не должен использоваться.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
    finally:
        store.close()
    canonical_db = _canonical_calls_db(
        tmp_path,
        {
            "canonical-from-record": "Рассматриваем курс по программированию.",
            "prefixed-id": "Хотим интенсив, но очень сложно с расписанием.",
        },
    )

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=load_canonical_call_client_texts(canonical_db),
        )

    interest_text = "\n".join(item.text for item in dossier.interests)
    pain_text = "\n".join(item.text for item in dossier.pains)
    assert "Рассматриваем курс по программированию" in interest_text
    assert "Хотим интенсив" in interest_text
    assert "сложно с расписанием" in pain_text.casefold()


def test_manager_dossier_cleans_speech_fillers_and_repeated_words(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    canonical_db = _canonical_calls_db(
        tmp_path,
        {
            "call-client": (
                "Ну эээ вот нас нас интересует олимпиадная математика онлайн, "
                "и хотим попробовать, потому что очень переживаем переживаем из-за экзамена"
            )
        },
    )

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=load_canonical_call_client_texts(canonical_db),
        )

    interest_from_call = next(item.text for item in dossier.interests if item.source == "mango_call:call-client")
    pain_from_call = next(item.text for item in dossier.pains if item.source == "mango_call:call-client")

    assert "эээ" not in interest_from_call.casefold()
    assert "вот нас нас" not in interest_from_call.casefold()
    assert "Нас интересует олимпиадная математика онлайн" in interest_from_call
    assert interest_from_call.endswith(".")
    assert "переживаем переживаем" not in pain_from_call.casefold()
    assert pain_from_call.endswith(".")


def test_manager_dossier_interest_quote_is_sentence_not_raw_transcript(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    canonical_db = _canonical_calls_db(
        tmp_path,
        {
            "call-client": (
                "Да я как раз хотела сама вам звонить Нас интересует математика очная "
                "Можете там счет скинуть или как там сколько будет полгод"
            )
        },
    )

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=load_canonical_call_client_texts(canonical_db),
        )

    interest_from_call = next(item.text for item in dossier.interests if item.source == "mango_call:call-client")

    assert interest_from_call == "Интерес из звонка: Нас интересует математика очная."
    assert "сколько будет полгод" not in interest_from_call


def test_manager_dossier_ignores_generic_wanted_phrases_without_product_context(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    canonical_db = _canonical_calls_db(
        tmp_path,
        {
            "call-client": (
                "Хотела спросить, как дела. "
                "Мы хотим уже бы хотим, чтобы он начал учиться. "
                "Нас интересует математика онлайн."
            )
        },
    )

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=load_canonical_call_client_texts(canonical_db),
        )

    interest_text = "\n".join(item.text for item in dossier.interests if item.source == "mango_call:call-client")

    assert "Хотела спросить" not in interest_text
    assert "Мы хотим уже бы хотим" not in interest_text
    assert "Нас интересует математика онлайн" in interest_text


def test_manager_dossier_pain_quote_trims_adjacent_asr_tail(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    canonical_db = _canonical_calls_db(
        tmp_path,
        {"call-client": "Переживаю, у вас же мест мало Сейчас Катя у вас рисунок карандаши."},
    )

    with sqlite3.connect(db) as con:
        dossier = build_customer_dossier(
            con,
            tenant_id="foton",
            customer_id="customer:1",
            canonical_calls=load_canonical_call_client_texts(canonical_db),
        )

    pain_text = "\n".join(item.text for item in dossier.pains if item.source == "mango_call:call-client")

    assert "Переживаю, у вас же мест мало." in pain_text
    assert "рисунок" not in pain_text.casefold()


def test_manager_dossier_reads_family_chronology_without_merging_customer_records(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_customer_with_call_and_opportunity(db, tmp_path)
    with CustomerTimelineSQLiteStore(db, allowed_root=tmp_path) as store:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton",
                customer_id="customer:2",
                identity_status=IdentityStatus.STRONG,
                display_name="Второй ученик",
                primary_phone="+79000000001",
                metadata={"brands": ["foton"]},
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:2",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW + timedelta(hours=1),
                source_system="mail_archive_stage2",
                source_id="family-mail",
                direction=TimelineDirection.INBOUND,
                summary="Родитель уточнил расписание второго ребёнка.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
    with sqlite3.connect(db) as con:
        con.execute("DELETE FROM family_members_v1 WHERE tenant_id='foton'")
        con.executemany(
            """
            INSERT INTO family_members_v1 (
              tenant_id, family_id, customer_id, membership_status, confidence,
              reason, created_at, updated_at, record_hash, record_json
            ) VALUES (?,?,?,?,?,?,?,?,?,?)
            """,
            (
                ("foton", "family:shared", "customer:1", "confident", "high", "test", NOW.isoformat(), NOW.isoformat(), "hash-1", "{}"),
                ("foton", "family:shared", "customer:2", "confident", "high", "test", NOW.isoformat(), NOW.isoformat(), "hash-2", "{}"),
            ),
        )
        dossier = build_customer_dossier(con, tenant_id="foton", customer_id="customer:1")

    chronology = "\n".join(row.text for row in dossier.chronology)
    assert dossier.customer_id == "customer:1"
    assert "второго ребёнка" in chronology
    assert "карточка: Второй ученик" in chronology
    assert any(row.source.endswith(":customer:2") for row in dossier.chronology)


@pytest.fixture
def owner50_workbook(tmp_path: Path) -> tuple[Path, dict[str, object]]:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(db, tmp_path, family_id="family:shared", customer_id="customer:a", signal_type="deal_stalling")
    _seed_owner50_member(db, tmp_path, family_id="family:shared", customer_id="customer:b", signal_type="callback_due")
    out = tmp_path / ".codex_local" / "owner50.xlsx"
    summary = dict(
        build_owner50_family_workbook(
            timeline_db=db,
            allowed_root=tmp_path,
            out_xlsx=out,
            as_of=NOW,
            enforce_freshness=False,
        )
    )
    return out, summary


def test_owner50_deduplicates_family_and_uses_best_family_signal(owner50_workbook: tuple[Path, dict[str, object]]) -> None:
    out, summary = owner50_workbook
    assert out.stat().st_mode & 0o777 == 0o600
    rows = list(load_workbook(out, read_only=True)["READY_50"].iter_rows(values_only=True))

    assert summary["families"] == 1
    assert len(rows) == 2
    assert rows[1][1] == "family:shared"
    assert rows[1][9] == "callback_due"
    assert rows[1][5] == "+79000000002"
    assert rows[1][6] == "customer-b@example.com"
    assert rows[1][7] == "Email"
    assert rows[1][8] == (NOW - timedelta(days=10)).isoformat()
    assert rows[1][11] == "Проверить историю и написать клиенту."
    assert rows[1][16] == "tier=0; due=1; fresh_intent=0; specific_offer=1; child_fit=1; payment_history=1"


def test_owner50_has_source_evidence_for_every_family(owner50_workbook: tuple[Path, dict[str, object]]) -> None:
    out, _ = owner50_workbook
    wb = load_workbook(out, read_only=True)
    families = {row[1] for row in list(wb["READY_50"].iter_rows(values_only=True))[1:]}
    evidence = list(wb["EVIDENCE"].iter_rows(values_only=True))[1:]

    assert families
    assert families == {row[0] for row in evidence}
    assert any(str(row[4]).startswith("derived_signals:") for row in evidence)
    assert any(str(row[4]).startswith("timeline_events:") for row in evidence)
    assert any(str(row[4]).startswith("customer_opportunities:") for row in evidence)
    assert any(str(row[4]).startswith("family_links_v1:") for row in evidence)
    assert any(row[4] == "customer_purchases_v1" for row in evidence)
    payment = next(row for row in evidence if row[4] == "customer_purchases_v1")
    assert "customer:a [all_time]" in payment[3]
    assert "customer:b [all_time]" in payment[3]
    # требование аудиторов BLOCKED #1 (fail-open доказательств): "Проверяемо" (resolvable)
    # теперь честно отражает разрешение в known_records, а не просто "поля не пустые" -- у
    # этой READY-семьи signal/event/offer резолвятся в реальные строки БД.
    resolvable_kinds = {row[2] for row in evidence if row[8] is True}
    assert {"signal", "event", "offer"}.issubset(resolvable_kinds)


def test_owner50_workbook_has_exactly_five_owner_sheets(owner50_workbook: tuple[Path, dict[str, object]]) -> None:
    """Требование аудиторов BLOCKED #5 (25.07): пять отдельных рабочих листов с литеральными
    именами READY_50/CANDIDATES/EXCLUDED/EVIDENCE/CONTROL -- заменяет прежние три листа
    ("Кому писать"/"Доказательства"/"Контроль")."""
    out, summary = owner50_workbook
    wb = load_workbook(out, read_only=True)

    assert wb.sheetnames == ["READY_50", "CANDIDATES", "EXCLUDED", "EVIDENCE", "CONTROL"]
    assert summary["sheets"] == ("READY_50", "CANDIDATES", "EXCLUDED", "EVIDENCE", "CONTROL")
    assert tuple(row.value for row in wb["READY_50"][1]) == OWNER50_REQUIRED_COLUMNS
    # требование архитектора #10 (лист кандидатов неполноценный): CANDIDATES/EXCLUDED несут
    # бренд/контакт/детей/сигнал/действие, не только family_id и причину -- см.
    # OWNER50_CONTROL_COLUMNS.
    assert tuple(row.value for row in wb["CANDIDATES"][1]) == OWNER50_CONTROL_COLUMNS
    assert tuple(row.value for row in wb["EXCLUDED"][1]) == OWNER50_CONTROL_COLUMNS
    assert summary["required_business_columns"] == OWNER50_REQUIRED_COLUMNS
    # требование аудиторов BLOCKED #5 (лист CONTROL несёт свежесть/counts/версию базы):
    # CONTROL -- теперь мета-лист сборки, не построчные причины.
    control_meta = {row[0]: row[1] for row in wb["CONTROL"].iter_rows(values_only=True, min_row=2) if row[0]}
    assert control_meta["ready_50"] == summary["families"]
    assert control_meta["tenant_id"] == "foton"
    assert "freshness_gate_passed" in control_meta
    assert "price_axes_catalog_version" in control_meta


def test_owner50_excludes_family_level_safety_risks(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:unsafe",
        customer_id="customer:unsafe",
        signal_type="client_returned",
        display_name="Тестовый сотрудник",
        no_contact=True,
        grade="11",
        event_summary="Хочу возврат денег за курс.",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:active-learning",
        customer_id="customer:c",
        signal_type="season_return_candidate",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:weak",
        customer_id="customer:d",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:closed-deal",
        customer_id="customer:e",
        signal_type="deal_stalling",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:optout",
        customer_id="customer:f",
        signal_type="client_returned",
        event_summary="Прошу больше со мной не связываться.",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:ambiguous-event",
        customer_id="customer:j",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:brand-conflict",
        customer_id="customer:k",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:superseded",
        customer_id="customer:l",
        signal_type="deal_stalling",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:partial-identity",
        customer_id="customer:m",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:outflow",
        customer_id="customer:n",
        signal_type="client_returned",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:explicit-refund",
        customer_id="customer:refund",
        signal_type="client_returned",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:test-opportunity",
        customer_id="customer:s",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:graduate-word",
        customer_id="customer:v",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:cross-brand-signal",
        customer_id="customer:w",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:test-signal",
        customer_id="customer:x",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:risky-signal-text",
        customer_id="customer:signal-risk",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:cross-brand-event",
        customer_id="customer:event-brand",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:test-event",
        customer_id="customer:event-test",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:empty-event",
        customer_id="customer:empty-event",
        signal_type="callback_due",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:c",
                event_type=TimelineEventType.TALLANTO_GROUP,
                event_at=NOW - timedelta(days=1),
                source_system="tallanto_snapshot",
                source_id="active-group",
                direction=TimelineDirection.SYSTEM,
                summary="Активная учебная группа.",
                match_status="strong_unique",
                created_at=NOW - timedelta(days=1),
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:refund",
                event_type=TimelineEventType.TALLANTO_PAYMENT,
                event_at=NOW - timedelta(days=2),
                source_system="tallanto_crm_call",
                source_id="explicit-refund",
                direction=TimelineDirection.SYSTEM,
                summary="Структурная операция Tallanto.",
                record={"amount": 1000, "payment_direction": "refund"},
                match_status="strong_unique",
                created_at=NOW - timedelta(days=2),
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        con.execute(
            "UPDATE family_members_v1 SET membership_status='conflict', confidence='low' WHERE customer_id='customer:d'"
        )
        con.execute(
            "UPDATE family_links_v1 SET status='ambiguous', confidence='low' WHERE customer_id='customer:d'"
        )
        con.execute(
            "UPDATE customer_opportunities SET status='closed', closed_at=? WHERE customer_id='customer:e'",
            ((NOW - timedelta(days=1)).isoformat(),),
        )
        con.execute(
            "UPDATE timeline_events SET match_status='ambiguous' WHERE customer_id='customer:j'"
        )
        con.execute(
            "UPDATE timeline_events SET superseded_by='replacement' WHERE customer_id='customer:l'"
        )
        con.execute(
            "UPDATE customer_identities SET identity_status='partial' WHERE customer_id='customer:m'"
        )
        con.execute(
            "UPDATE customer_purchases_v1 SET total_out=1000 WHERE customer_id='customer:n'"
        )
        test_record = json.loads(
            con.execute(
                "SELECT record_json FROM customer_opportunities WHERE customer_id='customer:s'"
            ).fetchone()[0]
        )
        test_record["metadata"] = {"tags": ["test"]}
        con.execute(
            "UPDATE customer_opportunities SET record_json=? WHERE customer_id='customer:s'",
            (json.dumps(test_record, ensure_ascii=False),),
        )
        con.execute(
            "UPDATE family_links_v1 SET canonical_name='Выпускница' WHERE customer_id='customer:v'"
        )
        cross_brand_record = json.loads(
            con.execute(
                "SELECT record_json FROM derived_signals WHERE customer_id='customer:w'"
            ).fetchone()[0]
        )
        cross_brand_record["recommended_action"] = "Предложить программу УНПК."
        con.execute(
            "UPDATE derived_signals SET record_json=? WHERE customer_id='customer:w'",
            (json.dumps(cross_brand_record, ensure_ascii=False),),
        )
        test_signal_record = json.loads(
            con.execute(
                "SELECT record_json FROM derived_signals WHERE customer_id='customer:x'"
            ).fetchone()[0]
        )
        test_signal_record["metadata"] = {"tags": ["test"]}
        con.execute(
            "UPDATE derived_signals SET record_json=? WHERE customer_id='customer:x'",
            (json.dumps(test_signal_record, ensure_ascii=False),),
        )
        risky_signal_record = json.loads(
            con.execute(
                "SELECT record_json FROM derived_signals WHERE customer_id='customer:signal-risk'"
            ).fetchone()[0]
        )
        risky_signal_record["evidence_text"] = "Клиент ждёт обещанный звонок."
        risky_signal_record["recommended_action"] = "Больше не пишите клиенту."
        con.execute(
            "UPDATE derived_signals SET record_json=? WHERE customer_id='customer:signal-risk'",
            (json.dumps(risky_signal_record, ensure_ascii=False),),
        )
        con.execute(
            "UPDATE timeline_events SET summary='Обсуждали программу УНПК.' "
            "WHERE customer_id='customer:event-brand'"
        )
        con.execute(
            "UPDATE timeline_events SET summary='Тестовый клиент для проверки импорта.' "
            "WHERE customer_id='customer:event-test'"
        )
        con.execute(
            "UPDATE timeline_events SET summary='', text_preview='', subject='' "
            "WHERE customer_id='customer:empty-event'"
        )
        brand_record = json.loads(
            con.execute(
                "SELECT record_json FROM customer_identities WHERE customer_id='customer:k'"
            ).fetchone()[0]
        )
        brand_record["metadata"]["brands"] = ["unpk"]
        con.execute(
            "UPDATE customer_identities SET record_json=? WHERE customer_id='customer:k'",
            (json.dumps(brand_record, ensure_ascii=False),),
        )
        con.commit()
    out = tmp_path / ".codex_local" / "owner50_excluded.xlsx"

    summary = build_owner50_family_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        as_of=NOW,
        enforce_freshness=False,
    )
    wb = load_workbook(out, read_only=True)
    control_rows = list(wb["EXCLUDED"].iter_rows(values_only=True))[1:]
    candidate_rows = list(wb["CANDIDATES"].iter_rows(values_only=True))[1:]
    control = "\n".join(str(row[2]) for row in control_rows)

    ready_rows = list(wb["READY_50"].iter_rows(values_only=True))[1:]
    assert summary["families"] == 2
    assert {row[1] for row in ready_rows} == {"family:outflow", "family:risky-signal-text"}
    assert {
        "structured_no_contact",
        "staff_test_system",
        "grade_11_or_graduate",
        "durable_p0_history",
        "durable_opt_out",
        "active_access_or_learning",
        "family_ambiguous",
        "child_ambiguous",
        "brand_ambiguous",
        "identity_not_strong",
    }.issubset(
        set(control.replace(",", "").split())
    )
    assert ("family:graduate-word", "excluded", "grade_11_or_graduate") in {
        row[:3] for row in control_rows
    }
    assert ("family:cross-brand-signal", "excluded", "brand_ambiguous") in {
        row[:3] for row in control_rows
    }
    assert ("family:test-signal", "excluded", "staff_test_system") in {
        row[:3] for row in control_rows
    }
    assert not any(row[0] == "family:risky-signal-text" for row in control_rows), (
        "manager-authored recommended_action is not a customer P0/opt-out"
    )
    assert ("family:cross-brand-event", "excluded", "brand_ambiguous") in {
        row[:3] for row in control_rows
    }
    assert ("family:test-event", "excluded", "staff_test_system") in {
        row[:3] for row in control_rows
    }
    assert ("family:explicit-refund", "excluded", "durable_p0_history") in {
        row[:3] for row in control_rows
    }
    assert ("family:ambiguous-event", "candidate", "signal_evidence_ambiguous") in {row[:3] for row in candidate_rows}
    assert ("family:superseded", "candidate", "signal_evidence_superseded") in {row[:3] for row in candidate_rows}
    assert ("family:closed-deal", "candidate", "active_deal_missing") in {row[:3] for row in candidate_rows}
    assert ("family:empty-event", "candidate", "signal_evidence_text_missing") in {row[:3] for row in candidate_rows}


def test_owner50_bulk_selection_has_constant_query_count(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    for index in range(20):
        _seed_owner50_member(
            db,
            tmp_path,
            family_id=f"family:{index:02d}",
            customer_id=f"customer:{index:02d}",
            signal_type="callback_due",
        )
    queries: list[str] = []
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        con.set_trace_callback(
            lambda sql: queries.append(sql)
            if sql.lstrip().upper().startswith(("SELECT", "WITH"))
            else None
        )
        candidates, _ = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert len(candidates) == 20
    assert len(queries) <= 10


def test_owner50_signal_budget_ignores_unlinked_signal_noise(tmp_path: Path, monkeypatch) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:linked",
        customer_id="customer:linked",
        signal_type="callback_due",
    )
    with sqlite3.connect(db) as con:
        template = con.execute(
            "SELECT * FROM derived_signals WHERE customer_id='customer:linked'"
        ).fetchone()
        placeholders = ",".join("?" for _ in template)
        for index in range(3):
            row = list(template)
            row[0] = f"signal:unlinked:{index}"
            row[2] = f"customer:unlinked:{index}"
            row[11] = (NOW + timedelta(days=index + 1)).isoformat()
            con.execute(f"INSERT INTO derived_signals VALUES ({placeholders})", row)
        con.commit()
        monkeypatch.setattr(manager_dossier_module, "OWNER50_SIGNAL_SCAN_LIMIT", 1)
        con.row_factory = sqlite3.Row
        candidates, _ = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert [row["family_id"] for row in candidates] == ["family:linked"]


def test_owner50_accepts_only_evidence_backed_brand_child_and_channels(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:child-brand",
        customer_id="customer:g",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:offer-grade",
        customer_id="customer:h",
        signal_type="client_returned",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:max-channel",
        customer_id="customer:i",
        signal_type="deal_stalling",
        # deal_stalling НЕ входит в OWNER50_PRODUCT_OPTIONAL_SIGNALS -- продукт решает
        # READY/CANDIDATE. Требование архитектора #10 (по итогам ревью 25.07): без явных
        # маркеров формата/периода в тексте продукт больше не резолвится догадкой.
        offer_title="Курс математики 8 класс, онлайн, годовой курс",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:active-foreign-offer",
        customer_id="customer:r",
        signal_type="callback_due",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_opportunity(
            CustomerOpportunity(
                tenant_id="foton",
                customer_id="customer:h",
                opportunity_type=OpportunityType.AMO_DEAL,
                source_system="amo",
                source_id="closed-unpk",
                title="UNPK 11 класс",
                status="closed",
                opened_at=NOW - timedelta(days=100),
                closed_at=NOW - timedelta(days=90),
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        raw = json.loads(
            con.execute(
                "SELECT record_json FROM customer_identities WHERE customer_id='customer:g'"
            ).fetchone()[0]
        )
        raw["metadata"]["brands"] = ["Фотон"]
        con.execute(
            "UPDATE customer_identities SET record_json=? WHERE customer_id='customer:g'",
            (json.dumps(raw, ensure_ascii=False),),
        )
        con.execute("DELETE FROM family_links_v1 WHERE customer_id='customer:h'")
        con.execute(
            "UPDATE timeline_events SET event_type='channel_message', source_system='wappi_max' "
            "WHERE customer_id='customer:i'"
        )
        foreign_record = json.loads(
            con.execute(
                "SELECT record_json FROM customer_opportunities WHERE customer_id='customer:r'"
            ).fetchone()[0]
        )
        foreign_record["product_context"] = {"products_of_interest": ["УНПК 8 класс"]}
        con.execute(
            "UPDATE customer_opportunities SET title='УНПК 8 класс', record_json=? WHERE customer_id='customer:r'",
            (json.dumps(foreign_record, ensure_ascii=False),),
        )
        con.commit()
    out = tmp_path / ".codex_local" / "owner50_fallbacks.xlsx"

    summary = build_owner50_family_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        as_of=NOW,
        enforce_freshness=False,
    )
    rows = {
        row[1]: row
        for row in list(load_workbook(out, read_only=True)["READY_50"].iter_rows(values_only=True))[1:]
    }

    assert summary["families"] == 2
    assert rows["family:child-brand"][2] == "foton"
    assert "family:offer-grade" not in rows
    assert rows["family:max-channel"][7] == "MAX"
    evidence = list(load_workbook(out, read_only=True)["EVIDENCE"].iter_rows(values_only=True))[1:]
    assert not any(row[0] == "family:offer-grade" and row[2] == "child" for row in evidence)
    control = list(load_workbook(out, read_only=True)["EXCLUDED"].iter_rows(values_only=True))[1:]
    candidate_rows = list(load_workbook(out, read_only=True)["CANDIDATES"].iter_rows(values_only=True))[1:]
    assert any(row[:3] == ("family:active-foreign-offer", "excluded", "brand_ambiguous") for row in control)
    assert any(row[:3] == ("family:offer-grade", "candidate", "target_child_unproven") for row in candidate_rows)


def test_owner50_never_substitutes_another_family_member_contact(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:contact",
        customer_id="customer:o",
        signal_type="client_returned",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:contact",
        customer_id="customer:p",
        signal_type="callback_due",
    )
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        con.execute(
            "UPDATE customer_identities SET primary_phone='', primary_email='' WHERE customer_id='customer:o'"
        )
        con.execute("DELETE FROM derived_signals WHERE customer_id='customer:p'")
        con.commit()
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert candidates == []
    assert ("family:contact", "excluded", "contact_missing", "У уверенного члена семьи нет телефона или email") in {
        row[:4] for row in control
    }


def test_owner50_rejects_bad_signal_evidence_without_family_fallback(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:no-fallback",
        customer_id="customer:t",
        signal_type="deal_stalling",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:no-fallback",
        customer_id="customer:u",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:no-fallback",
        customer_id="customer:ambig",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:no-fallback",
        customer_id="customer:missing",
        signal_type="callback_due",
    )
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:season-foreign",
        customer_id="customer:y",
        signal_type="season_return_candidate",
    )
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        con.execute(
            "UPDATE timeline_events SET superseded_by='replacement', summary='Хочу возврат денег.' "
            "WHERE customer_id='customer:u'"
        )
        con.execute(
            "UPDATE timeline_events SET match_status='ambiguous' WHERE customer_id='customer:ambig'"
        )
        missing_record = json.loads(
            con.execute(
                "SELECT record_json FROM derived_signals WHERE customer_id='customer:missing'"
            ).fetchone()[0]
        )
        missing_record.pop("event_id", None)
        con.execute(
            "UPDATE derived_signals SET event_id='', record_json=? WHERE customer_id='customer:missing'",
            (json.dumps(missing_record, ensure_ascii=False),),
        )
        con.execute(
            "UPDATE derived_signals SET event_id='foreign-event' WHERE customer_id='customer:y'"
        )
        con.commit()
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert candidates == []
    assert any(row[:3] == ("family:no-fallback", "excluded", "durable_p0_history") for row in control)
    assert any(row[:3] == ("family:season-foreign", "candidate", "signal_evidence_not_owned") for row in control)


def test_owner50_deal_signal_requires_its_linked_active_deal(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:linked-deal",
        customer_id="customer:z",
        signal_type="deal_stalling",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_opportunity(
            CustomerOpportunity(
                tenant_id="foton",
                customer_id="customer:z",
                opportunity_type=OpportunityType.AMO_DEAL,
                source_system="amo",
                source_id="another-active-deal",
                title="Другой активный курс 8 класс",
                status="active",
                opened_at=NOW - timedelta(days=2),
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        linked_id = con.execute(
            "SELECT opportunity_id FROM derived_signals WHERE customer_id='customer:z'"
        ).fetchone()[0]
        con.execute(
            "UPDATE customer_opportunities SET status='closed', closed_at=? WHERE opportunity_id=?",
            ((NOW - timedelta(days=1)).isoformat(), linked_id),
        )
        con.commit()
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert candidates == []
    assert any(row[:3] == ("family:linked-deal", "candidate", "active_deal_missing") for row in control)


def test_owner50_fails_closed_when_scan_budget_is_exceeded(tmp_path: Path, monkeypatch) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db,
        tmp_path,
        family_id="family:event-budget",
        customer_id="customer:q",
        signal_type="callback_due",
    )
    event_limit = manager_dossier_module.OWNER50_EVENT_SCAN_LIMIT
    monkeypatch.setattr(manager_dossier_module, "OWNER50_EVENT_SCAN_LIMIT", 0)

    with sqlite3.connect(db) as con, pytest.raises(RuntimeError, match="event budget exceeded"):
        con.row_factory = sqlite3.Row
        _owner50_family_rows(con, tenant_id="foton", as_of=NOW)
    monkeypatch.setattr(manager_dossier_module, "OWNER50_EVENT_SCAN_LIMIT", event_limit)
    monkeypatch.setattr(manager_dossier_module, "OWNER50_RELATED_SCAN_LIMIT", 0)
    with sqlite3.connect(db) as con, pytest.raises(RuntimeError, match="members budget exceeded"):
        con.row_factory = sqlite3.Row
        _owner50_family_rows(con, tenant_id="foton", as_of=NOW)


def test_owner50_caps_at_50_without_padding(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    for index in range(52):
        _seed_owner50_member(
            db,
            tmp_path,
            family_id=f"family:cap:{index:02d}",
            customer_id=f"customer:cap:{index:02d}",
            signal_type="callback_due",
        )
    out = tmp_path / ".codex_local" / "owner50_cap.xlsx"

    summary = build_owner50_family_workbook(
        timeline_db=db,
        allowed_root=tmp_path,
        out_xlsx=out,
        limit=999,
        as_of=NOW,
        enforce_freshness=False,
    )
    wb = load_workbook(out, read_only=True)
    # требование аудиторов BLOCKED #5: READY, не попавшие в топ-50 ("outside_limit"), теперь
    # живут на листе CANDIDATES (они не исключены -- им просто не хватило места).
    candidates_rows = list(wb["CANDIDATES"].iter_rows(values_only=True))[1:]

    assert summary["candidate_families"] == 52
    assert summary["families"] == 50
    assert summary["effective_limit"] == 50
    assert summary["ready_total"] == 52
    assert summary["ready_audit_population_complete"] is False
    assert sum(row[1] == "outside_limit" for row in candidates_rows) == 2
    control_meta = {row[0]: row[1] for row in wb["CONTROL"].iter_rows(values_only=True, min_row=2) if row[0]}
    assert control_meta["ready_total"] == 52
    assert control_meta["ready_audit_population_complete"] is False


def test_owner50_family_rows_wires_classify_family_into_ready_candidate_excluded(tmp_path: Path) -> None:
    """Требование архитектора #1/#9: classify_family подключён в финал _owner50_family_rows.
    READY идёт в candidates (лист READY_50); CANDIDATE никогда не туда -- только в control со
    статусом "candidate" (лист CANDIDATES); EXCLUDED -- в control со статусом "excluded" (лист
    EXCLUDED). Все три категории видны раздельно на выходе одной и той же функции, а не только
    у изолированного classify_family."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:ready", customer_id="customer:ready", signal_type="callback_due",
    )
    _seed_owner50_member(
        db, tmp_path, family_id="family:stale", customer_id="customer:stale", signal_type="callback_due",
    )
    _seed_owner50_member(
        db, tmp_path, family_id="family:blocked", customer_id="customer:blocked", signal_type="callback_due",
        event_summary="Хочу возврат денег за курс.",
    )
    with sqlite3.connect(db) as con:
        con.execute(
            "UPDATE derived_signals SET created_at=? WHERE customer_id='customer:stale'",
            ((NOW - timedelta(days=45)).isoformat(),),
        )
        con.commit()
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert [row["family_id"] for row in candidates] == ["family:ready"]
    assert candidates[0]["status"] == "READY"
    assert candidates[0]["action_text"]
    control_index = {(family_id, status, code) for family_id, status, code, *_rest in control}
    assert ("family:stale", "candidate", "stale_signal") in control_index
    assert ("family:blocked", "excluded", "durable_p0_history") in control_index
    assert "family:stale" not in {row["family_id"] for row in candidates}

    out = tmp_path / ".codex_local" / "owner50_tiers.xlsx"
    summary = build_owner50_family_workbook(
        timeline_db=db, allowed_root=tmp_path, out_xlsx=out, as_of=NOW, enforce_freshness=False,
    )
    assert summary["families"] == 1
    assert summary["candidate_queue_families"] == 1
    assert summary["excluded_families"] == 1
    wb = load_workbook(out, read_only=True)
    ready_rows = list(wb["READY_50"].iter_rows(values_only=True))[1:]
    assert {row[1] for row in ready_rows} == {"family:ready"}
    # требование архитектора #8: последняя колонка -- готовая фраза "действие + кому + срок".
    assert ready_rows[0][-1]
    candidate_rows = list(wb["CANDIDATES"].iter_rows(values_only=True))[1:]
    excluded_rows = list(wb["EXCLUDED"].iter_rows(values_only=True))[1:]
    assert any(row[:3] == ("family:stale", "candidate", "stale_signal") for row in candidate_rows)
    assert any(row[:3] == ("family:blocked", "excluded", "durable_p0_history") for row in excluded_rows)
    # требование архитектора #9: доказательства несут дату+source_system+event_id отдельно.
    evidence_rows = list(wb["EVIDENCE"].iter_rows(values_only=True))[1:]
    signal_evidence = next(row for row in evidence_rows if row[2] == "signal")
    assert signal_evidence[6] == "mail_archive_stage2"  # source_system
    assert signal_evidence[7]  # event_id непустой
    assert signal_evidence[5]  # дата непустая


def test_owner50_signal_expiry_is_not_manager_due_date(tmp_path: Path) -> None:
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:no-manager-due", customer_id="customer:no-manager-due",
        signal_type="callback_due",
    )
    with sqlite3.connect(db) as con:
        signal_row = con.execute(
            "SELECT signal_id, record_json, expires_at FROM derived_signals WHERE customer_id='customer:no-manager-due'",
        ).fetchone()
        record = json.loads(signal_row[1])
        record.pop("follow_up_due_at", None)
        con.execute(
            "UPDATE derived_signals SET record_json=? WHERE signal_id=?",
            (json.dumps(record, ensure_ascii=False), signal_row[0]),
        )
        assert signal_row[2]  # срок жизни сигнала остался, но это не deadline менеджера
        con.commit()
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:no-manager-due" not in {row["family_id"] for row in candidates}
    assert ("family:no-manager-due", "candidate", "next_step_missing_or_vague") in {
        row[:3] for row in control
    }


def test_owner50_stale_freshness_gate_stops_entire_build_no_workbook_written(tmp_path: Path) -> None:
    """Требование аудиторов BLOCKED #3 (25.07): непройденный manager_freshness_gate должен
    остановить ВЕСЬ build (raise), а не молча пометить каждую семью EXCLUDED и всё равно
    записать workbook. Эта БД не содержит ingestion_cursors/ingestion_runs -- гейт заведомо не
    проходит; enforce_freshness намеренно НЕ передан False (прод-дефолт True)."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:would-be-ready", customer_id="customer:would-be-ready",
        signal_type="callback_due",
    )
    out = tmp_path / ".codex_local" / "owner50_stale.xlsx"

    with pytest.raises(RuntimeError, match="freshness gate failed"):
        build_owner50_family_workbook(
            timeline_db=db, allowed_root=tmp_path, out_xlsx=out, as_of=NOW,
        )

    assert not out.exists()


def test_owner50_family_without_any_signal_is_classified_not_silently_dropped(tmp_path: Path) -> None:
    """Требование аудиторов BLOCKED #4 (25.07): семья, у которой нет НИ ОДНОГО активного
    канонического сигнала (никогда не проходила старый SQL-предфильтр по derived_signals),
    раньше не попадала НИКУДА -- ни в candidates, ни в control, ни на один лист. Теперь она
    обязана получить статус (candidate с кодом no_active_outreach_signal, как решил бы
    classify_family для одинокого Г5) -- каждая семья тенанта классифицируется."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:has-signal", customer_id="customer:has-signal",
        signal_type="callback_due",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton", customer_id="customer:no-signal", identity_status=IdentityStatus.STRONG,
                display_name="Родитель без сигнала", primary_phone="+79000007777",
                primary_email="no-signal@example.com", source_ref="amocrm:contact:customer:no-signal",
                metadata={"brands": ["foton"]},
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        # family_links_v1 уже создана предыдущим _seed_owner50_member (CREATE TABLE IF NOT
        # EXISTS) -- добавляем ещё одну семью БЕЗ единого derived_signals, но с чистым
        # брендом/классом 1-10, чтобы дойти именно до ветки "нет сигнала", а не отвалиться
        # раньше по другой (не связанной с этим тестом) причине.
        con.execute(
            "INSERT OR REPLACE INTO family_members_v1 VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "family:no-signal", "customer:no-signal", "confident", "high", "test",
                NOW.isoformat(), NOW.isoformat(), "hash-no-signal", "{}",
            ),
        )
        con.execute(
            "INSERT OR REPLACE INTO family_links_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "family:no-signal", "customer:no-signal", "child:no-signal", "Ребёнок без сигнала", "[]",
                json.dumps(["7"]), json.dumps(["математика"], ensure_ascii=False), "foton", "confident", "high",
                "test", "[]", 1, NOW.isoformat(), "hash-child-no-signal", "{}",
            ),
        )
        con.commit()
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:has-signal" in {row["family_id"] for row in candidates}
    assert "family:no-signal" not in {row["family_id"] for row in candidates}
    # раньше family:no-signal не появлялась бы вообще нигде в control -- молча исчезала.
    assert ("family:no-signal", "candidate", "no_active_outreach_signal") in {row[:3] for row in control}


def _synthetic_price_axes_catalog(*, brand: str = "foton", grade_min: int = 5, grade_max: int = 11) -> dict[str, object]:
    return {
        "entries": [
            {
                "entry_id": f"synthetic:{brand}:{grade_min}-{grade_max}:online:year",
                "brand": brand,
                "format": "online",
                "period": "year",
                "grade_min": grade_min,
                "grade_max": grade_max,
                "grade_values": list(range(grade_min, grade_max + 1)),
                "product_code": "regular_course",
                "tariff_id": "",
                "schedule": "",
                "subjects": ["math", "physics", "informatics", "russian", "ai"],
                "amount": 47250,
                "client_safe_text": "Тест: онлайн-курс, годовой.",
            }
        ]
    }


def test_owner50_product_confirmed_via_injected_catalog_entry_id_only(tmp_path: Path) -> None:
    """Требование архитектора #3: продукт подтверждается ТОЛЬКО по точному entry_id
    действующего price_axes_catalog (передан явно через параметр -- тест не зависит от
    содержимого реального продуктового каталога на диске). deal_stalling НЕ входит в
    OWNER50_PRODUCT_OPTIONAL_SIGNALS -- продукт напрямую решает READY vs CANDIDATE, а
    неизвестный класс (нет entry в каталоге) -- CANDIDATE, никогда не выдуманный READY."""
    db = _timeline_db(tmp_path)
    # требование архитектора #10 (по итогам ревью 25.07): продукт резолвится ТОЛЬКО по
    # формату/периоду, явно упомянутым в тексте предложения -- offer_title обязан нести
    # реальные маркеры ("онлайн"/"годовой"), догадка по каноническим комбинациям убрана.
    _seed_owner50_member(
        db, tmp_path, family_id="family:in-catalog", customer_id="customer:in-catalog",
        signal_type="deal_stalling", grade="8", offer_title="Курс математики, 8 класс, онлайн, годовой курс",
    )
    _seed_owner50_member(
        db, tmp_path, family_id="family:out-of-catalog", customer_id="customer:out-of-catalog",
        signal_type="deal_stalling", grade="3", offer_title="Курс математики, 3 класс, онлайн, годовой курс",
    )
    catalog = _synthetic_price_axes_catalog(brand="foton", grade_min=5, grade_max=11)

    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW, price_axes_catalog=catalog)

    ready_ids = {row["family_id"] for row in candidates}
    assert "family:in-catalog" in ready_ids
    ready_row = next(row for row in candidates if row["family_id"] == "family:in-catalog")
    assert ready_row["product_entry_id"] == "synthetic:foton:5-11:online:year"
    assert "family:out-of-catalog" not in ready_ids
    assert ("family:out-of-catalog", "candidate", "product_not_confirmed_by_kb") in {
        row[:3] for row in control
    }


def test_owner50_does_not_retarget_an_older_child_signal_to_a_younger_sibling(tmp_path: Path) -> None:
    """Смешанная семья "11 + младший" не исключается, но сигнал старшего ребёнка
    нельзя молча переадресовать младшему только потому, что ему подходит продукт."""
    db = _timeline_db(tmp_path)
    # offer_title несёт явные маркеры формата/периода ("онлайн"/"годовой") -- требование
    # архитектора #10: продукт больше не резолвится догадкой по каноническим комбинациям.
    _seed_owner50_member(
        db, tmp_path, family_id="family:mixed", customer_id="customer:mixed-a",
        signal_type="deal_stalling", grade="11",
        offer_title="Курс математики, онлайн, годовой курс",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton", customer_id="customer:mixed-b", identity_status=IdentityStatus.STRONG,
                display_name="Родитель customer:mixed-b", primary_phone="+79000009999",
                primary_email="mixed-b@example.com", source_ref="amocrm:contact:customer:mixed-b",
                metadata={"brands": ["foton"]},
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT OR REPLACE INTO family_members_v1 VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "family:mixed", "customer:mixed-b", "confident", "high", "test",
                NOW.isoformat(), NOW.isoformat(), "hash-mixed-b", "{}",
            ),
        )
        con.execute(
            "INSERT OR REPLACE INTO family_links_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "family:mixed", "customer:mixed-b", "child:mixed-b", "Младший", "[]",
                json.dumps(["6"]), json.dumps(["математика"], ensure_ascii=False), "foton", "confident", "high",
                "test", "[]", 1, NOW.isoformat(), "hash-child-mixed-b", "{}",
            ),
        )
        con.commit()

    catalog = {
        "entries": [
            {
                "entry_id": "grade-6-only", "brand": "foton", "format": "online", "period": "year",
                "grade_min": 6, "grade_max": 6, "grade_values": [6], "product_code": "regular_course",
                "tariff_id": "", "schedule": "", "subjects": ["math"], "amount": 10000,
                "client_safe_text": "6 класс",
            },
            {
                "entry_id": "grade-11-only", "brand": "foton", "format": "online", "period": "year",
                "grade_min": 11, "grade_max": 11, "grade_values": [11], "product_code": "regular_course",
                "tariff_id": "", "schedule": "", "subjects": ["math"], "amount": 20000,
                "client_safe_text": "11 класс",
            },
        ]
    }

    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW, price_axes_catalog=catalog)

    assert not any(row[:3] == ("family:mixed", "excluded", "grade_11_or_graduate") for row in control)
    assert "family:mixed" not in {row["family_id"] for row in candidates}
    assert ("family:mixed", "candidate", "target_child_unproven") in {row[:3] for row in control}


def test_owner50_ready_offer_uses_kb_text_and_carries_target_child(tmp_path: Path) -> None:
    """E2 (26.07): для READY с подтверждённым продуктом "Предложение" -- ТОЛЬКО текст выбранного
    price_axes_catalog entry (client_safe_text), никогда старое название сделки AMO (то же
    название есть в offers, но не в "Предложение"); строка несёт конкретного адресата
    (target_child_key/name/grade). E4 (26.07): evidence "product"/"child"/"payment" резолвятся
    в реальные записи (раньше все три были структурно нерезолвируемыми)."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:kb-offer", customer_id="customer:kb-offer",
        signal_type="deal_stalling", grade="8",
        offer_title="Курс математики, 8 класс, онлайн, годовой курс",
    )
    catalog = _synthetic_price_axes_catalog(brand="foton", grade_min=5, grade_max=11)

    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW, price_axes_catalog=catalog)

    assert "family:kb-offer" in {row["family_id"] for row in candidates}
    ready_row = next(row for row in candidates if row["family_id"] == "family:kb-offer")
    assert ready_row["offer"] == "Тест: онлайн-курс, годовой."
    assert ready_row["offer"] != "Курс математики, 8 класс, онлайн, годовой курс"
    assert ready_row["target_child_key"] == "child:customer:kb-offer"
    assert ready_row["target_child_name"] == "Ребёнок customer:kb-offer"
    assert ready_row["target_child_grade"] == "8"

    product_evidence = next(item for item in ready_row["evidence"] if item["kind"] == "product")
    assert product_evidence["resolvable"] is True
    assert product_evidence["event_id"] == ready_row["product_entry_id"]
    child_evidence = next(item for item in ready_row["evidence"] if item["kind"] == "child")
    assert child_evidence["resolvable"] is True
    payment_evidence = next(item for item in ready_row["evidence"] if item["kind"] == "payment")
    assert payment_evidence["resolvable"] is True  # ровно один вклад (один customer, all_time)


def test_owner50_two_different_children_with_confirmed_grades_is_target_child_ambiguous(tmp_path: Path) -> None:
    """E2 (26.07): семья с ДВУМЯ РАЗНЫМИ верифицированными не-выпускниками, у каждого класс
    известен однозначно (по одному значению в истории у каждого) -- но продукт всё равно нельзя
    адресовать никому из них без доказательства, кому именно. deal_stalling НЕ входит в
    OWNER50_PRODUCT_OPTIONAL_SIGNALS -- продукт обязателен -> CANDIDATE "target_child_ambiguous",
    а не молчаливое "берём младшего" (это и есть отменённая эвристика "минимум из истории",
    только на уровне выбора МЕЖДУ детьми, а не внутри истории одного ребёнка)."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:two-kids", customer_id="customer:two-kids",
        signal_type="deal_stalling", grade="5",
        offer_title="Курс математики, онлайн, годовой курс",
    )
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT OR REPLACE INTO family_links_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "family:two-kids", "customer:two-kids", "child:two-kids-b", "Второй ребёнок", "[]",
                json.dumps(["9"]), json.dumps(["математика"], ensure_ascii=False), "foton", "confident", "high",
                "test", "[]", 1, NOW.isoformat(), "hash-child-two-kids-b", "{}",
            ),
        )
        con.commit()
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(
            con, tenant_id="foton", as_of=NOW,
            price_axes_catalog=_synthetic_price_axes_catalog(brand="foton", grade_min=1, grade_max=11),
        )

    assert "family:two-kids" not in {row["family_id"] for row in candidates}
    assert ("family:two-kids", "candidate", "target_child_ambiguous") in {row[:3] for row in control}


def test_owner50_single_child_conflicting_grade_history_falls_back_to_offer_not_minimum(tmp_path: Path) -> None:
    """E2 (26.07): один ребёнок упомянут в истории то с классом 5, то с классом 8 (grades_json --
    множество БЕЗ дат, family_graph.py только добавляет в set) -- запрещённая догадка "минимум
    из истории" выбрала бы grade-5. Правильно: класс этого ребёнка неизвестен (два разных
    значения), продукт резолвится по классу, явно названному в тексте предложения (8-й), адресат
    (target_child) остаётся пустым (не выдумываем, КАКОЙ это ребёнок)."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:grade-history", customer_id="customer:grade-history",
        signal_type="deal_stalling", grade="5",
        offer_title="Курс математики, 8 класс, онлайн, годовой курс",
    )
    with sqlite3.connect(db) as con:
        con.execute(
            "UPDATE family_links_v1 SET grades_json=? WHERE customer_id='customer:grade-history'",
            (json.dumps(["5", "8"]),),
        )
        con.commit()
    catalog = {
        "entries": [
            {
                "entry_id": "grade-5-only", "brand": "foton", "format": "online", "period": "year",
                "grade_min": 5, "grade_max": 5, "grade_values": [5], "product_code": "regular_course",
                "tariff_id": "", "schedule": "", "subjects": ["math"], "amount": 10000,
                "client_safe_text": "5 класс",
            },
            {
                "entry_id": "grade-8-only", "brand": "foton", "format": "online", "period": "year",
                "grade_min": 8, "grade_max": 8, "grade_values": [8], "product_code": "regular_course",
                "tariff_id": "", "schedule": "", "subjects": ["math"], "amount": 15000,
                "client_safe_text": "8 класс",
            },
        ]
    }

    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW, price_axes_catalog=catalog)

    assert "family:grade-history" not in {row["family_id"] for row in candidates}
    assert ("family:grade-history", "candidate", "target_child_unproven") in {row[:3] for row in control}


def test_owner50_brand_unproven_reaches_candidate_not_excluded_via_full_workbook(tmp_path: Path) -> None:
    """E1 (26.07): семья без единого распознанного бренда обязана
    получить CANDIDATE "brand_unproven" (Г2: 0 брендов -> CANDIDATE, не EXCLUDED). Раньше рабочий
    SQL-путь (_owner50_family_rows) сам жёстко исключал такую семью ДО того, как classify_family
    вообще её видел. Этот тест идёт через ПОЛНЫЙ build_owner50_family_workbook (не только
    classify_family/_owner50_family_rows по отдельности), включая запись в XLSX-листы."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:brand-unproven", customer_id="customer:brand-unproven",
        signal_type="client_returned",
    )
    with sqlite3.connect(db) as con:
        raw = json.loads(
            con.execute(
                "SELECT record_json FROM customer_identities WHERE customer_id='customer:brand-unproven'"
            ).fetchone()[0]
        )
        raw["metadata"]["brands"] = ["неизвестный бренд"]
        con.execute(
            "UPDATE customer_identities SET record_json=? WHERE customer_id='customer:brand-unproven'",
            (json.dumps(raw, ensure_ascii=False),),
        )
        con.execute(
            "UPDATE family_links_v1 SET brand='неизвестный бренд' WHERE customer_id='customer:brand-unproven'"
        )
        con.commit()
    out = tmp_path / ".codex_local" / "owner50_brand_unproven.xlsx"

    summary = build_owner50_family_workbook(
        timeline_db=db, allowed_root=tmp_path, out_xlsx=out, as_of=NOW, enforce_freshness=False,
    )

    wb = load_workbook(out, read_only=True)
    ready_ids = {row[1] for row in list(wb["READY_50"].iter_rows(values_only=True))[1:]}
    candidate_rows = list(wb["CANDIDATES"].iter_rows(values_only=True))[1:]
    excluded_rows = list(wb["EXCLUDED"].iter_rows(values_only=True))[1:]

    assert "family:brand-unproven" not in ready_ids
    assert any(row[0] == "family:brand-unproven" and row[2] == "brand_unproven" for row in candidate_rows)
    assert not any(row[0] == "family:brand-unproven" for row in excluded_rows)
    assert summary["candidate_queue_reason_counts"].get("brand_unproven", 0) >= 1


def test_owner50_fake_client_name_ochno_dva_predmeta_never_ready(tmp_path: Path) -> None:
    """E5 (26.07), синтетическая регрессия владельца #1: "очно два предмета" читается как
    название сделки/шага, не как имя человека. READY требует структурного person-contact origin
    (стабильный AMO-контакт или Tallanto person ID) -- проверка идёт по структурному
    record.source_ref, а не по смыслу текста display_name (никакого нового regex понимания)."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:fake-name-1", customer_id="customer:fake-name-1",
        signal_type="client_returned", display_name="очно два предмета", source_ref="",
    )
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:fake-name-1" not in {row["family_id"] for row in candidates}
    assert ("family:fake-name-1", "excluded", "person_origin_unproven") in {row[:3] for row in control}


def test_owner50_fake_client_name_lvsh_2_chast_never_ready(tmp_path: Path) -> None:
    """E5 (26.07), синтетическая регрессия владельца #2: "лвш 2 часть"."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:fake-name-2", customer_id="customer:fake-name-2",
        signal_type="client_returned", display_name="лвш 2 часть", source_ref="",
    )
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:fake-name-2" not in {row["family_id"] for row in candidates}
    assert ("family:fake-name-2", "excluded", "person_origin_unproven") in {row[:3] for row in control}


def test_owner50_fake_client_name_os_ot_roditelya_never_ready(tmp_path: Path) -> None:
    """E5 (26.07), синтетическая регрессия владельца #3: "ОС от родителя + предложить 26/27
    уч.г." -- следующий шаг менеджера, не имя клиента."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:fake-name-3", customer_id="customer:fake-name-3",
        signal_type="client_returned", display_name="ОС от родителя + предложить 26/27 уч.г.", source_ref="",
    )
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:fake-name-3" not in {row["family_id"] for row in candidates}
    assert ("family:fake-name-3", "excluded", "person_origin_unproven") in {row[:3] for row in control}


def test_owner50_refund_mention_excludes_regardless_of_age_without_structural_resolution(tmp_path: Path) -> None:
    """Требование аудиторов BLOCKED #2 (отменяет прежнее требование архитектора #6): без
    ДОКАЗАННОГО структурного статуса резолюции refund-упоминание исключает семью НАВСЕГДА,
    независимо от возраста -- эвристика "после жалобы было ещё какое-то событие = закрыто"
    удалена целиком, потому что она ничего не доказывает о РЕЗУЛЬТАТЕ жалобы."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:old-refund", customer_id="customer:old-refund",
        signal_type="callback_due", event_summary="Обычная переписка про расписание.",
    )
    _seed_owner50_member(
        db, tmp_path, family_id="family:recent-refund", customer_id="customer:recent-refund",
        signal_type="callback_due", event_summary="Хочу возврат денег за курс.",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton", customer_id="customer:old-refund",
                event_type=TimelineEventType.EMAIL_MESSAGE, event_at=NOW - timedelta(days=800),
                source_system="mail_archive_stage2", source_id="old-refund-event",
                direction=TimelineDirection.INBOUND, summary="Хочу возврат денег за курс.",
                match_status="strong_unique", created_at=NOW - timedelta(days=800),
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    ready_ids = {row["family_id"] for row in candidates}
    # СТАРЫЙ (800 дней) refund БЕЗ структурного статуса резолюции -- ВСЁ РАВНО исключает.
    assert "family:old-refund" not in ready_ids
    assert ("family:old-refund", "excluded", "durable_p0_history") in {row[:3] for row in control}
    assert "family:recent-refund" not in ready_ids
    assert ("family:recent-refund", "excluded", "durable_p0_history") in {row[:3] for row in control}


def test_refund_forgiveness_requires_explicit_structural_resolution_status() -> None:
    """Требование аудиторов BLOCKED #2: единственное, что может простить refund-упоминание --
    ЯВНЫЙ структурный статус резолюции на самой записи (record/metadata.resolution_status).
    Более позднее событие (даже содержательное) -- НЕ доказательство результата жалобы и
    больше не участвует в решении вообще (эвристика удалена, а не ослаблена)."""
    old_refund_unresolved = {"event_at": (OWNER50_CLASSIFY_NOW - timedelta(days=800)).isoformat()}
    meaningful_later_event = {
        "event_at": (OWNER50_CLASSIFY_NOW - timedelta(days=5)).isoformat(),
        "event_type": "email_message",
        "summary": "Клиент уточнил расписание на новый сезон.",
    }

    # Без структурного статуса -- НЕ прощается, даже с содержательным более поздним событием.
    still_unresolved = _owner50_event_p0_is_stale_and_resolved(
        old_refund_unresolved, codes=["refund"],
        all_events=[old_refund_unresolved, meaningful_later_event], as_of=OWNER50_CLASSIFY_NOW,
    )
    assert still_unresolved is False

    # Явный структурный статус в record (формат _owner50_event) -- прощает.
    old_refund_resolved = {
        "event_at": (OWNER50_CLASSIFY_NOW - timedelta(days=800)).isoformat(),
        "record": {"resolution_status": "resolved"},
    }
    resolved_via_record = _owner50_event_p0_is_stale_and_resolved(
        old_refund_resolved, codes=["refund"], all_events=[old_refund_resolved], as_of=OWNER50_CLASSIFY_NOW,
    )
    assert resolved_via_record is True

    # Тот же статус в metadata (альтернативное вложение _owner50_event) -- тоже прощает.
    old_refund_resolved_metadata = {
        "event_at": (OWNER50_CLASSIFY_NOW - timedelta(days=800)).isoformat(),
        "metadata": {"resolution_status": "closed"},
    }
    resolved_via_metadata = _owner50_event_p0_is_stale_and_resolved(
        old_refund_resolved_metadata, codes=["refund"],
        all_events=[old_refund_resolved_metadata], as_of=OWNER50_CLASSIFY_NOW,
    )
    assert resolved_via_metadata is True

    # Серьёзные коды (не чистый refund) НЕ прощаются, даже со структурным статусом.
    legal_with_status = {
        "event_at": (OWNER50_CLASSIFY_NOW - timedelta(days=800)).isoformat(),
        "record": {"resolution_status": "resolved"},
    }
    legal_never_forgiven = _owner50_event_p0_is_stale_and_resolved(
        legal_with_status, codes=["refund", "legal"], all_events=[legal_with_status], as_of=OWNER50_CLASSIFY_NOW,
    )
    assert legal_never_forgiven is False


def test_owner50_objection_downgrades_family_to_candidate(tmp_path: Path) -> None:
    """Требование архитектора #7: свежее возражение (после сигнала-основания, без более
    свежего клиентского позитива) откатывает семью в CANDIDATE -- customer_objections_v1
    теперь читается в _owner50_snapshot и доходит до classify_family."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:objecting", customer_id="customer:objecting",
        signal_type="callback_due",
    )
    with sqlite3.connect(db) as con:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_objections_v1 (
              tenant_id TEXT NOT NULL, customer_id TEXT NOT NULL, source_event_id TEXT NOT NULL,
              source_channel TEXT NOT NULL, objection_type TEXT NOT NULL, quote_preview TEXT NOT NULL,
              budget_hint_rub INTEGER, price_sensitivity TEXT NOT NULL, extracted_at TEXT NOT NULL,
              extractor_version TEXT NOT NULL, speaker TEXT NOT NULL DEFAULT 'unknown',
              direction TEXT NOT NULL DEFAULT 'unknown', confidence TEXT NOT NULL DEFAULT 'low',
              PRIMARY KEY (tenant_id, customer_id, source_event_id, objection_type)
            )
            """
        )
        con.execute(
            "INSERT INTO customer_objections_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "customer:objecting", "event:objection-1", "email", "price",
                "Дорого, надо подумать.", 30000, "high",
                (NOW - timedelta(days=1)).isoformat(),
                "test", "client", "inbound", "high",
            ),
        )
        con.commit()
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:objecting" not in {row["family_id"] for row in candidates}
    assert ("family:objecting", "candidate", "late_objection_no_fresh_positive") in {
        row[:3] for row in control
    }


def test_owner50_objection_from_another_family_member_also_downgrades(tmp_path: Path) -> None:
    """Требование архитектора #7 (ужесточено по итогам ревью 25.07): раньше свежесть
    возражения проверялась только у ВЛАДЕЛЬЦА сигнала (objections_by_customer.get(customer_id)).
    Свежий негатив ДРУГОГО родителя той же семьи игнорировался. family["objections"] уже
    отсортирован DESC по всей семье в _owner50_snapshot -- берём самое свежее возражение
    ЛЮБОГО члена, не только текущего сигнала."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(
        db, tmp_path, family_id="family:two-parents", customer_id="customer:signal-owner",
        signal_type="callback_due",
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton", customer_id="customer:other-parent", identity_status=IdentityStatus.STRONG,
                display_name="Другой родитель", primary_phone="+79000008888",
                primary_email="other-parent@example.com", source_ref="amocrm:contact:customer:other-parent",
                metadata={"brands": ["foton"]},
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        con.execute(
            "INSERT OR REPLACE INTO family_members_v1 VALUES (?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "family:two-parents", "customer:other-parent", "confident", "high", "test",
                NOW.isoformat(), NOW.isoformat(), "hash-other-parent", "{}",
            ),
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_objections_v1 (
              tenant_id TEXT NOT NULL, customer_id TEXT NOT NULL, source_event_id TEXT NOT NULL,
              source_channel TEXT NOT NULL, objection_type TEXT NOT NULL, quote_preview TEXT NOT NULL,
              budget_hint_rub INTEGER, price_sensitivity TEXT NOT NULL, extracted_at TEXT NOT NULL,
              extractor_version TEXT NOT NULL, speaker TEXT NOT NULL DEFAULT 'unknown',
              direction TEXT NOT NULL DEFAULT 'unknown', confidence TEXT NOT NULL DEFAULT 'low',
              PRIMARY KEY (tenant_id, customer_id, source_event_id, objection_type)
            )
            """
        )
        # Возражение принадлежит customer:other-parent -- НЕ владельцу сигнала
        # (customer:signal-owner) -- и оно свежее оплаты семьи (2026-06-01, см.
        # _seed_owner50_member), поэтому обязано перекрыть READY в CANDIDATE.
        con.execute(
            "INSERT INTO customer_objections_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", "customer:other-parent", "event:objection-other", "email", "price",
                "Слишком дорого, надо подумать.", 30000, "high",
                (NOW - timedelta(days=1)).isoformat(),
                "test", "client", "inbound", "high",
            ),
        )
        con.commit()
        con.row_factory = sqlite3.Row
        candidates, control = _owner50_family_rows(con, tenant_id="foton", as_of=NOW)

    assert "family:two-parents" not in {row["family_id"] for row in candidates}
    assert ("family:two-parents", "candidate", "late_objection_no_fresh_positive") in {
        row[:3] for row in control
    }


# ---------------------------------------------------------------------------
# classify_family -- ported from codex_artifacts/TEST_owner50_classification.py
# (исполнитель прогнал 10/10 как отдельный файл против OWNER50_classifier.py; здесь тот же
# набор проверок против встроенной в manager_dossier.py версии). Синтетика: все имена, id и
# суммы придуманы для этого файла, реальных ПДн нет. score_family/select_owner50_families
# (и их тесты) убраны по итогам ревью 25.07 -- параллельный, невызываемый из рабочего пути
# путь ранжирования (см. rank_key в _owner50_family_rows -- единственная формула в проде).
#
# Отличие от исходного артефакта: тест graduate/grade-recompute адаптирован под решение
# владельца 25.07 -- "все дети выпускники" -> EXCLUDED с кодом "grade_11_or_graduate" (тот
# же код, что и у SQL-предфильтра), НЕ CANDIDATE, как было в более ранней правке Fable #4.
# ---------------------------------------------------------------------------

OWNER50_CLASSIFY_NOW = datetime(2026, 7, 24, 12, 0, tzinfo=timezone.utc)


def _owner50_golden_family(**overrides: Any) -> dict[str, Any]:
    """Минимальная синтетическая семья, гарантированно READY под classify_family.
    Каждый тест меняет РОВНО то, что проверяет, остальное остаётся золотым эталоном --
    так падение теста однозначно указывает на конкретное правило."""
    family: dict[str, Any] = {
        "family_id": "family:golden",
        "identity": {
            "customer_id": "customer:golden",
            "identity_status": "strong",
            "display_name": "Родитель Голден",
        },
        "brands": {"foton"},
        "unrecognized_brand_present": False,
        "family_conflict": False,
        "family_attribution_by_tag": False,
        "contact_missing": False,
        "children": [
            {"child_key": "child:1", "name": "Аня", "grade_current": 8},
        ],
        "payment": {
            "customer_id": "customer:golden",
            "total_in": 45000,
            "total_out": 0,
            "deals_cnt": 1,
            "last_purchase_at": OWNER50_CLASSIFY_NOW - timedelta(days=200),
            "period": "all_time",
        },
        "interest_quote": None,
        "signal": {
            "signal_type": "callback_due",
            "created_at": OWNER50_CLASSIFY_NOW - timedelta(days=2),
            "evidence_text": "Обещали перезвонить по подбору курса 22.07.",
            "event_id": "evt:signal-golden",
            "source_system": "amocrm_event",
        },
        "next_step": {
            "action": "Выполнить обещанный звонок от 22.07, срок сегодня-завтра",
            "due": "2026-07-25",
        },
        "product": None,  # callback_due -- продукт не обязателен (OWNER50_PRODUCT_OPTIONAL_SIGNALS)
        "last_objection": None,
        "events_by_id": {"evt:signal-golden": {"source_system": "amocrm_event"}},
        "open_p0": False,
        "opt_out": False,
        "identity_conflict": False,
        "recent_meaningful_outbound_after_evidence": False,
        "active_recent_manager_work": False,
        "active_risk_signals": (),
        "stale_data": False,
        "last_touch_at": None,
    }
    family.update(overrides)
    return family


def test_golden_family_is_actually_ready() -> None:
    """Санити-проверка самого эталона -- если это падает, остальные тесты ничего не
    доказывают (потому что каждый из них меняет золотую семью)."""
    result = classify_family(_owner50_golden_family(), as_of=OWNER50_CLASSIFY_NOW)
    assert result["status"] == "READY"
    assert result["reasons"] == ()
    assert result["missing"] == ()


def test_ready_requires_one_proven_child_key() -> None:
    missing_child = classify_family(
        _owner50_golden_family(children=[]), as_of=OWNER50_CLASSIFY_NOW,
    )
    assert missing_child["status"] == "CANDIDATE"
    assert "target_child_unproven" in missing_child["missing"]

    empty_key = classify_family(
        _owner50_golden_family(children=[{"child_key": "", "name": "Аня", "grade_current": 8}]),
        as_of=OWNER50_CLASSIFY_NOW,
    )
    assert empty_key["status"] == "CANDIDATE"
    assert "target_child_unproven" in empty_key["missing"]


def test_graduate_and_younger_sibling_not_excluded_but_all_graduates_are() -> None:
    mixed = _owner50_golden_family(
        children=[
            {"child_key": "child:older", "name": "Игорь", "grade_current": 11},
            {"child_key": "child:younger", "name": "Соня", "grade_current": 6},
        ]
    )
    mixed_result = classify_family(mixed, as_of=OWNER50_CLASSIFY_NOW)
    assert mixed_result["status"] == "READY"
    assert "grade_11_or_graduate" not in mixed_result["reasons"]

    only_graduate = _owner50_golden_family(
        children=[{"child_key": "child:only", "name": "Игорь", "grade_current": 11}]
    )
    graduate_result = classify_family(only_graduate, as_of=OWNER50_CLASSIFY_NOW)
    # решение владельца 25.07 (отменяет более раннюю правку Fable #4): все верифицированные
    # дети -- 11 класс/выпускники -> EXCLUDED, не CANDIDATE. Тот же код, что и в
    # SQL-предфильтре _owner50_family_rows -- один источник правды на обоих слоях.
    assert graduate_result["status"] == "EXCLUDED"
    assert "grade_11_or_graduate" in graduate_result["reasons"]


def test_unknown_brand_is_candidate_conflicting_brand_is_excluded() -> None:
    unknown_brand = _owner50_golden_family(brands=set())
    unknown_result = classify_family(unknown_brand, as_of=OWNER50_CLASSIFY_NOW)
    assert unknown_result["status"] == "CANDIDATE"
    assert "brand_unproven" in unknown_result["missing"]

    unrecognized_result = classify_family(
        _owner50_golden_family(brands={"неизвестный бренд"}), as_of=OWNER50_CLASSIFY_NOW,
    )
    assert unrecognized_result["status"] == "CANDIDATE"
    assert "brand_unproven" in unrecognized_result["missing"]

    conflicting_brand = _owner50_golden_family(brands={"foton", "unpk"})
    conflict_result = classify_family(conflicting_brand, as_of=OWNER50_CLASSIFY_NOW)
    assert conflict_result["status"] == "EXCLUDED"
    assert "brand_ambiguous" in conflict_result["reasons"]

    assert classify_family(_owner50_golden_family(brands={"МФТИ"}), as_of=OWNER50_CLASSIFY_NOW)["status"] == "READY"
    assert classify_family(_owner50_golden_family(brands={"ЦДПО"}), as_of=OWNER50_CLASSIFY_NOW)["status"] == "READY"
    alias_conflict = classify_family(
        _owner50_golden_family(brands={"МФТИ", "ЦДПО"}), as_of=OWNER50_CLASSIFY_NOW,
    )
    assert alias_conflict["status"] == "EXCLUDED"
    assert "brand_ambiguous" in alias_conflict["reasons"]


def test_objection_before_positive_is_ready_objection_after_is_candidate() -> None:
    # требование E3 (26.07, отменяет более раннюю версию этого теста): client_returned сам по
    # себе больше НЕ позитив -- см. test_client_returned_alone_does_not_override_a_late_objection
    # ниже, который проверяет именно этот гейт. Единственные позитивы, способные перекрыть
    # возражение, -- подтверждённая оплата или доказанная свежая interest_quote; здесь эталон --
    # interest_quote, привязанная к реальному событию золотой семьи (events_by_id уже несёт
    # evt:signal-golden), датирована NOW-2 дня -- как и позитив в исходном сценарии теста.
    positive_interest = {
        "text": "Хотим записаться в этом месяце.",
        "quoted_at": (OWNER50_CLASSIFY_NOW - timedelta(days=2)).isoformat(),
        "event_id": "evt:signal-golden",
        "source_system": "amocrm_event",
        "direction": "inbound",
    }

    # Возражение NOW-10 дней -- СТАРШЕ позитива.
    objection_before = _owner50_golden_family(
        interest_quote=dict(positive_interest),
        last_objection={"text": "Дорого, надо подумать.", "at": OWNER50_CLASSIFY_NOW - timedelta(days=10)},
    )
    before_result = classify_family(objection_before, as_of=OWNER50_CLASSIFY_NOW)
    assert before_result["status"] == "READY"
    assert "late_objection_no_fresh_positive" not in before_result["missing"]

    # Возражение NOW-1 день -- СВЕЖЕЕ позитива (интерес NOW-2 дня).
    objection_after = _owner50_golden_family(
        interest_quote=dict(positive_interest),
        last_objection={"text": "Дорого, надо подумать.", "at": OWNER50_CLASSIFY_NOW - timedelta(days=1)},
    )
    after_result = classify_family(objection_after, as_of=OWNER50_CLASSIFY_NOW)
    assert after_result["status"] == "CANDIDATE"
    assert "late_objection_no_fresh_positive" in after_result["missing"]


def test_client_returned_alone_does_not_override_a_late_objection() -> None:
    """Требование E3 (26.07): client_returned САМ ПО СЕБЕ не доказывает, что клиент передумал
    после возражения -- клиент мог написать снова и с новой претензией. Сигнал client_returned
    датирован NOW-1 день (свежее возражения NOW-2 дня), но золотая семья по умолчанию несёт
    только старую оплату (NOW-200 дней) и никакой interest_quote -- значит настоящего позитива,
    способного перекрыть возражение, нет вообще, статус обязан остаться CANDIDATE."""
    family = _owner50_golden_family(
        signal=dict(_owner50_golden_family()["signal"], signal_type="client_returned",
                    created_at=OWNER50_CLASSIFY_NOW - timedelta(days=1)),
        last_objection={"text": "Дорого, надо подумать.", "at": OWNER50_CLASSIFY_NOW - timedelta(days=2)},
    )
    result = classify_family(family, as_of=OWNER50_CLASSIFY_NOW)
    assert result["status"] == "CANDIDATE"
    assert "late_objection_no_fresh_positive" in result["missing"]


def test_stale_positive_signal_never_overrides_a_fresh_objection() -> None:
    """Требование архитектора #7: callback_due/deal_stalling -- НЕ новый позитив, способный
    перекрыть возражение (обещание перезвонить и зависшая сделка не доказывают, что клиент
    передумал). Возражение датировано NOW-5 дней -- строго МЕЖДУ сигналом (NOW-2 дня, позже
    возражения по времени) и оплатой (NOW-200 дней, раньше возражения). Если бы
    signal.created_at засчитывался как позитив для этих типов (старое, неверное поведение),
    возражение оказалось бы "до позитива" и статус остался бы READY; по требованию #7 сигнал
    НЕ считается, остаётся только старая оплата -- возражение свежее её, поэтому CANDIDATE."""
    for stale_signal_type in ("callback_due", "deal_stalling"):
        family = _owner50_golden_family(
            signal=dict(_owner50_golden_family()["signal"], signal_type=stale_signal_type),
            last_objection={"text": "Дорого, надо подумать.", "at": OWNER50_CLASSIFY_NOW - timedelta(days=5)},
        )
        result = classify_family(family, as_of=OWNER50_CLASSIFY_NOW)
        assert result["status"] == "CANDIDATE", stale_signal_type
        assert "late_objection_no_fresh_positive" in result["missing"], stale_signal_type


def test_product_must_be_kb_verified_not_from_history() -> None:
    # deal_stalling НЕ входит в OWNER50_PRODUCT_OPTIONAL_SIGNALS -- продукт решает исход.
    stalling_signal = {
        "signal_type": "deal_stalling",
        "created_at": OWNER50_CLASSIFY_NOW - timedelta(days=1),
        "evidence_text": "Сделка зависла на этапе счёта.",
        "event_id": "evt:deal",
        "source_system": "amocrm_event",
    }
    stalling_events = {"evt:deal": {"source_system": "amocrm_event"}}
    stalling_next_step = {"action": "Вернуться к зависшей сделке: счёт по курсу физики", "due": "2026-07-26"}

    from_history = _owner50_golden_family(
        signal=dict(stalling_signal),
        events_by_id=dict(stalling_events),
        next_step=dict(stalling_next_step),
        product={
            "name": "Курс физики (упомянут в переписке 2024 года)",
            "brand": "foton",
            "verified": False,
            "source": "amo_history_text",
        },
    )
    history_result = classify_family(from_history, as_of=OWNER50_CLASSIFY_NOW)
    assert history_result["status"] == "CANDIDATE"
    assert history_result["missing"] == ("product_not_confirmed_by_kb",)

    from_kb = _owner50_golden_family(
        signal=dict(stalling_signal),
        events_by_id=dict(stalling_events),
        next_step=dict(stalling_next_step),
        product={
            "name": "Курс физики, 8 класс",
            "brand": "foton",
            "verified": True,
            "source": "kb_price_axes_catalog:synthetic:foton:physics:8",
            "entry_id": "synthetic:foton:physics:8",
            "seats_available": True,
        },
    )
    kb_result = classify_family(from_kb, as_of=OWNER50_CLASSIFY_NOW)
    assert kb_result["status"] == "READY"
    assert kb_result["missing"] == ()

    # Ловушка §5.2 #1: подтверждённый продукт, но мест нет (например ЛВШ лето-2026).
    sold_out = _owner50_golden_family(
        signal=dict(stalling_signal),
        events_by_id=dict(stalling_events),
        next_step=dict(stalling_next_step),
        product={
            "name": "ЛВШ лето-2026",
            "brand": "foton",
            "verified": True,
            "source": "kb_price_axes_catalog",
            "seats_available": False,
        },
    )
    sold_out_result = classify_family(sold_out, as_of=OWNER50_CLASSIFY_NOW)
    assert "product_not_confirmed_by_kb" in sold_out_result["missing"]

    # Ловушка §5.2 #2: продукт верифицирован, но бренд не совпадает с брендом семьи.
    wrong_brand = _owner50_golden_family(
        signal=dict(stalling_signal),
        events_by_id=dict(stalling_events),
        next_step=dict(stalling_next_step),
        product={
            "name": "Смена Подлипки (август)",
            "brand": "unpk",
            "verified": True,
            "source": "kb_price_axes_catalog",
            "seats_available": True,
        },
    )
    wrong_brand_result = classify_family(wrong_brand, as_of=OWNER50_CLASSIFY_NOW)
    assert "product_not_confirmed_by_kb" in wrong_brand_result["missing"]


def test_payments_are_not_double_counted_fact_plus_all_time() -> None:
    rows = [
        {
            "customer_id": "customer:a", "period": "all_time", "money_kind": "fact",
            "total_in": 50000, "total_out": 0, "deals_cnt": 1, "last_purchase_at": OWNER50_CLASSIFY_NOW - timedelta(days=100),
        },
        {
            "customer_id": "customer:a", "period": "2025-26", "money_kind": "fact",
            "total_in": 20000, "total_out": 0, "deals_cnt": 1, "last_purchase_at": OWNER50_CLASSIFY_NOW - timedelta(days=400),
        },
        {
            "customer_id": "customer:b", "period": "all_time", "money_kind": "fact",
            "total_in": 15000, "total_out": 0, "deals_cnt": 1, "last_purchase_at": OWNER50_CLASSIFY_NOW - timedelta(days=50),
        },
    ]

    deduped = dedupe_family_payment_rows(rows)

    assert deduped["customer:a"]["total_in"] == 50000  # НЕ 70000 -- all_time уже включает узкий период
    assert deduped["customer:a"]["period_used"] == "all_time"
    assert deduped["customer:a"]["rows_used"] == 1
    assert deduped["customer:b"]["total_in"] == 15000

    # Суммирование ЧЕРЕЗ РАЗНЫХ членов семьи (a + b) -- это НЕ повторный счёт, так уже
    # считает существующий код (_owner50_family_rows).
    family_total_in = sum(row["total_in"] for row in deduped.values())
    assert family_total_in == 65000


def test_tallanto_balance_charge_does_not_cancel_confirmed_payment() -> None:
    family = _owner50_golden_family(
        payment={
            "customer_id": "customer:golden",
            "total_in": 45000,
            "total_out": 45000,
            "deals_cnt": 1,
            "last_purchase_at": OWNER50_CLASSIFY_NOW - timedelta(days=20),
            "period": "all_time",
        }
    )

    result = classify_family(family, as_of=OWNER50_CLASSIFY_NOW)

    assert result["status"] == "READY"
    assert "no_payment_or_interest_evidence" not in result["missing"]


def test_owner50_payment_still_requires_incoming_money_deal_and_date() -> None:
    base = dict(_owner50_golden_family()["payment"])
    for payment in (
        {**base, "total_in": 0, "total_out": 45000},
        {**base, "deals_cnt": 0, "total_out": 45000},
        {**base, "last_purchase_at": None, "total_out": 45000},
        {**base, "last_purchase_at": "not-a-date", "total_out": 45000},
    ):
        result = classify_family(
            _owner50_golden_family(payment=payment),
            as_of=OWNER50_CLASSIFY_NOW,
        )
        assert result["status"] == "CANDIDATE"
        assert "no_payment_or_interest_evidence" in result["missing"]


def test_evidence_resolves_to_its_source_system_and_flags_dangling_refs() -> None:
    events_by_id = {"evt:call-1": {"source_system": "tallanto_crm_call", "event_at": OWNER50_CLASSIFY_NOW.isoformat()}}
    known_records = {"evt:call-1": {"source_system": "tallanto_crm_call", "at": OWNER50_CLASSIFY_NOW.isoformat()}}

    # требование аудиторов BLOCKED #1 (fail-open доказательств): resolvable=True ТОЛЬКО когда
    # event_id реально найден в known_records с совпадающим source_system+датой -- не просто
    # потому что поля заполнены.
    record = build_evidence_record(
        kind="signal", text="Клиент просил перезвонить.", event_id="evt:call-1",
        source_system="tallanto_crm_call", at=OWNER50_CLASSIFY_NOW, known_records=known_records,
    )
    assert record["resolvable"] is True
    assert resolve_evidence_source("evt:call-1", events_by_id) == "tallanto_crm_call"
    assert resolve_evidence_source("evt:missing-entirely", events_by_id) is None

    # fail CLOSED: те же полностью заполненные поля, но БЕЗ known_records (нечем проверить) --
    # НЕ resolvable. Раньше (BLOCKED) это было бы True просто потому, что строки не пустые.
    unverifiable = build_evidence_record(
        kind="signal", text="Клиент просил перезвонить.", event_id="evt:call-1",
        source_system="tallanto_crm_call", at=OWNER50_CLASSIFY_NOW,
    )
    assert unverifiable["resolvable"] is False
    assert "not_found_in_database" in unverifiable["missing_fields"]

    # тот же event_id, но известная запись несёт ДРУГОЙ source_system -- не резолвится.
    wrong_source = build_evidence_record(
        kind="signal", text="Клиент просил перезвонить.", event_id="evt:call-1",
        source_system="amocrm_event", at=OWNER50_CLASSIFY_NOW, known_records=known_records,
    )
    assert wrong_source["resolvable"] is False

    # event_id, которого нет в known_records вовсе (выдуманный/повисший) -- не резолвится.
    fabricated = build_evidence_record(
        kind="signal", text="Клиент просил перезвонить.", event_id="evt:fabricated",
        source_system="tallanto_crm_call", at=OWNER50_CLASSIFY_NOW, known_records=known_records,
    )
    assert fabricated["resolvable"] is False

    incomplete = build_evidence_record(kind="signal", text="", event_id="", source_system="", at=None)
    assert incomplete["resolvable"] is False
    assert set(incomplete["missing_fields"]) == {"event_id", "source_system", "at", "text"}

    # Сигнал семьи ссылается на event_id, которого нет в её собственном индексе событий
    # ("повисшая" ссылка) -- ГОТОВАЯ строка обязана откатиться в КАНДИДАТ, а не остаться READY.
    dangling_reference = _owner50_golden_family(events_by_id={"evt:unrelated": {"source_system": "amocrm_event"}})
    dangling_result = classify_family(dangling_reference, as_of=OWNER50_CLASSIFY_NOW)
    assert dangling_result["status"] == "CANDIDATE"
    assert dangling_result["missing"] == ("no_active_outreach_signal",)


def test_grade_recompute_from_recorded_grade_and_fixed_year_reaches_graduate_threshold() -> None:
    # as_of = OWNER50_CLASSIFY_NOW = 2026-07-24 -> текущий учебный год (§5.5, старт 1
    # сентября) = 2025. Зафиксирован 9-классником 2023-10-01 -> учебный год фиксации = 2023.
    # Пересчёт: 9 + (2025 - 2023) = 11 -> уже выпускник по порогу classify_family.
    now_grown_up = _owner50_golden_family(
        children=[
            {
                "child_key": "child:recomputed",
                "name": "Игорь",
                "grade_recorded": 9,
                "grade_fixed_at": datetime(2023, 10, 1, tzinfo=timezone.utc),
            }
        ]
    )
    grown_up_result = classify_family(now_grown_up, as_of=OWNER50_CLASSIFY_NOW)
    # решение владельца 25.07 (отменяет более раннюю правку Fable #4): выпускник -- EXCLUDED,
    # не CANDIDATE, даже когда класс пересчитан (не указан явно в данных).
    assert grown_up_result["status"] == "EXCLUDED"
    assert "grade_11_or_graduate" in grown_up_result["reasons"]
    # E2: пересчитанный класс обязан быть помечен как предположение даже для EXCLUDED.
    assert any("пересчёт" in note for note in grown_up_result["assumptions"])

    # Тот же горизонт пересчёта (+2 года), но стартовая точка ниже -> ещё не выпускник.
    still_at_school = _owner50_golden_family(
        children=[
            {
                "child_key": "child:recomputed",
                "name": "Игорь",
                "grade_recorded": 6,
                "grade_fixed_at": datetime(2023, 10, 1, tzinfo=timezone.utc),
            }
        ]
    )
    still_result = classify_family(still_at_school, as_of=OWNER50_CLASSIFY_NOW)
    assert still_result["status"] == "READY"
    assert any("пересчёт" in note for note in still_result["assumptions"])


def test_signal_source_system_mismatch_is_not_resolved() -> None:
    """Требование архитектора #2: сигнал разрешается ТОЛЬКО если событие существует И его
    source_system совпадает с тем, что заявляет сам сигнал -- иначе это привязка к чужому
    событию (например, ошибка импорта), а не подтверждённое основание."""
    mismatched = _owner50_golden_family(
        events_by_id={"evt:signal-golden": {"source_system": "wappi_telegram"}},  # не amocrm_event
    )
    result = classify_family(mismatched, as_of=OWNER50_CLASSIFY_NOW)
    assert result["status"] == "CANDIDATE"
    assert result["missing"] == ("no_active_outreach_signal",)


def test_interest_quote_without_event_provenance_does_not_count_as_fresh() -> None:
    """Требование архитектора #2: интерес засчитывается ТОЛЬКО при event_id+source_system,
    разрешённых в реальное событие с совпадающим source_system -- голая цитата "из воздуха"
    (без привязки к событию) не заменяет ни оплату, ни доказуемый сигнал."""
    unprovenanced = _owner50_golden_family(
        payment=None,
        interest_quote={
            "text": "Интересует курс математики.",
            "quoted_at": OWNER50_CLASSIFY_NOW - timedelta(days=30),
            "event_id": "",
            "source_system": "",
        },
    )
    unprovenanced_result = classify_family(unprovenanced, as_of=OWNER50_CLASSIFY_NOW)
    assert unprovenanced_result["status"] == "CANDIDATE"
    assert "no_payment_or_interest_evidence" in unprovenanced_result["missing"]

    provenanced = _owner50_golden_family(
        payment=None,
        interest_quote={
            "text": "Интересует курс математики.",
            "quoted_at": OWNER50_CLASSIFY_NOW - timedelta(days=30),
            "event_id": "evt:interest-1",
            "source_system": "amocrm_event",
            "direction": "inbound",
        },
        events_by_id={
            "evt:signal-golden": {"source_system": "amocrm_event"},
            "evt:interest-1": {"source_system": "amocrm_event"},
        },
    )
    provenanced_result = classify_family(provenanced, as_of=OWNER50_CLASSIFY_NOW)
    assert provenanced_result["status"] == "READY"
    assert "no_payment_or_interest_evidence" not in provenanced_result["missing"]

    outbound = _owner50_golden_family(
        payment=None,
        interest_quote={**provenanced["interest_quote"], "direction": "outbound"},
        events_by_id=provenanced["events_by_id"],
    )
    outbound_result = classify_family(outbound, as_of=OWNER50_CLASSIFY_NOW)
    assert outbound_result["status"] == "CANDIDATE"
    assert "no_payment_or_interest_evidence" in outbound_result["missing"]

    unresolved = _owner50_golden_family(
        payment=None,
        interest_quote=provenanced["interest_quote"],
        events_by_id={"evt:signal-golden": {"source_system": "amocrm_event"}},
    )
    unresolved_result = classify_family(unresolved, as_of=OWNER50_CLASSIFY_NOW)
    assert unresolved_result["status"] == "CANDIDATE"
    assert "no_payment_or_interest_evidence" in unresolved_result["missing"]


def test_future_dated_payment_and_interest_are_not_fresh() -> None:
    """Требование архитектора #5: дата в будущем -- НЕ "свежо" (похоже на ошибку данных, не
    на доказательство), ни для оплаты, ни для цитаты интереса."""
    future_payment = _owner50_golden_family(
        payment={
            "customer_id": "customer:golden", "total_in": 45000, "total_out": 0, "deals_cnt": 1,
            "last_purchase_at": OWNER50_CLASSIFY_NOW + timedelta(days=5), "period": "all_time",
        },
        interest_quote=None,
    )
    future_payment_result = classify_family(future_payment, as_of=OWNER50_CLASSIFY_NOW)
    assert "no_payment_or_interest_evidence" in future_payment_result["missing"]

    future_interest = _owner50_golden_family(
        payment=None,
        interest_quote={
            "text": "Интересует курс математики.",
            "quoted_at": OWNER50_CLASSIFY_NOW + timedelta(days=5),
            "event_id": "evt:signal-golden",
            "source_system": "amocrm_event",
            "direction": "inbound",
        },
    )
    future_interest_result = classify_family(future_interest, as_of=OWNER50_CLASSIFY_NOW)
    assert "no_payment_or_interest_evidence" in future_interest_result["missing"]


def test_future_dated_signal_is_not_treated_as_freshest_possible() -> None:
    """Требование архитектора #5 (ужесточено по итогам ревью 25.07): _owner50_days_since
    клэмпит отрицательную разницу к 0.0 -- без явной проверки сигнал с датой создания в
    будущем читался бы как "0 дней назад", то есть МАКСИМАЛЬНО свежий, хотя это похоже на
    ошибку данных. golden-семья по умолчанию свежая (signal.created_at = NOW-2 дня);
    здесь та же семья, но created_at в будущем."""
    future_signal = _owner50_golden_family(
        signal=dict(_owner50_golden_family()["signal"], created_at=OWNER50_CLASSIFY_NOW + timedelta(days=5)),
    )
    result = classify_family(future_signal, as_of=OWNER50_CLASSIFY_NOW)
    assert result["status"] == "CANDIDATE"
    assert "signal_date_in_future" in result["missing"]
    assert "stale_signal" not in result["missing"]  # отдельная причина, не подменяет "старый"


def test_owner50_action_text_combines_action_who_and_deadline() -> None:
    """Требование архитектора #8: READY = одна фраза "действие + повод/продукт или
    уточнение + срок" (правка Fable #1, "тест 5 секунд")."""
    text = owner50_action_text(_owner50_golden_family())
    assert text.startswith("Выполнить обещанный звонок")
    assert "до 2026-07-25" in text
    assert "Родитель Голден" in text

    assert owner50_action_text(_owner50_golden_family(next_step=None)) == ""


def test_owner50_rebuild_is_deterministic_same_rows_and_order(tmp_path: Path) -> None:
    """A3 (ТЗ Owner50): повторная сборка очереди по НЕИЗМЕННОЙ БД обязана давать те же
    строки в том же порядке на READY_50/CANDIDATES/EXCLUDED -- иначе менеджер не может
    доверять рангу между запусками. family:det-a и family:det-c намеренно одинаковы по
    тиру/сигналу/оплате/дате основания -- различить их может только финальный tie-break
    rank_key по family_id, поэтому тест одновременно проверяет и повторяемость, и сам
    tie-break."""
    db = _timeline_db(tmp_path)
    _seed_owner50_member(db, tmp_path, family_id="family:det-a", customer_id="customer:det-a", signal_type="callback_due")
    _seed_owner50_member(db, tmp_path, family_id="family:det-b", customer_id="customer:det-b", signal_type="client_returned")
    _seed_owner50_member(db, tmp_path, family_id="family:det-c", customer_id="customer:det-c", signal_type="callback_due")
    _seed_owner50_member(db, tmp_path, family_id="family:det-stale", customer_id="customer:det-stale", signal_type="callback_due")
    _seed_owner50_member(
        db, tmp_path, family_id="family:det-excluded", customer_id="customer:det-excluded",
        signal_type="callback_due", event_summary="Хочу возврат денег за курс.",
    )
    with sqlite3.connect(db) as con:
        con.execute(
            "UPDATE derived_signals SET created_at=? WHERE customer_id='customer:det-stale'",
            ((NOW - timedelta(days=45)).isoformat(),),
        )
        con.commit()

    def build(name: str) -> tuple[list[tuple], list[tuple], list[tuple]]:
        out = tmp_path / ".codex_local" / name
        build_owner50_family_workbook(
            timeline_db=db, allowed_root=tmp_path, out_xlsx=out, as_of=NOW, enforce_freshness=False,
        )
        wb = load_workbook(out, read_only=True)
        return (
            list(wb["READY_50"].iter_rows(values_only=True)),
            list(wb["CANDIDATES"].iter_rows(values_only=True)),
            list(wb["EXCLUDED"].iter_rows(values_only=True)),
        )

    first_ready, first_candidates, first_excluded = build("owner50_det_1.xlsx")
    second_ready, second_candidates, second_excluded = build("owner50_det_2.xlsx")

    assert first_ready == second_ready
    assert first_candidates == second_candidates
    assert first_excluded == second_excluded
    # sanity: the run actually exercised more than a trivial single-row case, and the
    # tie-break is observable (det-a before det-c, both otherwise-identical callback_due).
    ready_family_order = [row[1] for row in first_ready[1:]]
    assert ready_family_order.index("family:det-a") < ready_family_order.index("family:det-c")
    assert len(first_ready) > 2
    assert any(row[0] == "family:det-stale" for row in first_candidates)
    assert any(row[0] == "family:det-excluded" for row in first_excluded)


def _timeline_db(tmp_path: Path) -> Path:
    db = tmp_path / "timeline.sqlite"
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    store.close()
    return db


def _seed_customer_with_call_and_opportunity(db: Path, tmp_path: Path) -> None:
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton",
                customer_id="customer:1",
                identity_status=IdentityStatus.STRONG,
                display_name="Тестовый клиент",
                primary_phone="+79000000000",
                primary_email="parent@example.com",
                summary={"products_of_interest": ["ОГЭ по физике"]},
                metadata={"brands": ["foton"]},
            )
        )
        store.upsert_opportunity(
            CustomerOpportunity(
                tenant_id="foton",
                customer_id="customer:1",
                opportunity_type=OpportunityType.AMO_DEAL,
                source_system="amo",
                source_id="lead-1",
                title="Летняя школа по математике",
                product_context={"products_of_interest": ["Летняя школа по математике"]},
                opened_at=NOW,
            )
        )
        store.upsert_opportunity(
            CustomerOpportunity(
                tenant_id="foton",
                customer_id="customer:1",
                opportunity_type=OpportunityType.AMO_DEAL,
                source_system="amo",
                source_id="lead-title-only",
                title="Служебная акция из title",
                product_context={},
                opened_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW,
                source_system="mango_call",
                source_id="call-client",
                direction=TimelineDirection.INBOUND,
                summary="Клиент интересуется летней школой.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.MANGO_CALL,
                event_at=NOW,
                source_system="mango_call",
                source_id="call-manager-only",
                direction=TimelineDirection.INBOUND,
                summary="Менеджер говорит: клиенту сложно оплатить.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
        store.upsert_event(
            TimelineEvent(
                tenant_id="foton",
                customer_id="customer:1",
                event_type=TimelineEventType.EMAIL_MESSAGE,
                event_at=NOW,
                source_system="mail_archive_stage2",
                source_id="mail-1",
                direction=TimelineDirection.INBOUND,
                summary="Письмо для попадания в сегмент звонки+письма.",
                match_status="strong_unique",
                created_at=NOW,
            )
        )
    finally:
        store.close()


def _seed_owner50_member(
    db: Path,
    tmp_path: Path,
    *,
    family_id: str,
    customer_id: str,
    signal_type: str,
    display_name: str | None = None,
    no_contact: bool = False,
    grade: str = "8",
    event_summary: str = "Клиент подтвердил интерес к курсу.",
    offer_title: str = "Курс математики 8 класс",
    # E5 (26.07): по умолчанию -- настоящий person-contact origin (структурный source_ref вида
    # "amocrm:contact:<id>", как реально штампует ingestion.py для AMO-контактов). Синтетические
    # регрессии E5 передают source_ref явно (например "amocrm:lead:..." или "" -- сделка/шаг без
    # происхождения), чтобы доказать, что такие записи НЕ становятся READY-клиентом.
    source_ref: str | None = None,
) -> None:
    number = ord(customer_id[-1].casefold()) - ord("a") + 1 if customer_id[-1].isalpha() else 9
    resolved_source_ref = f"amocrm:contact:{customer_id}" if source_ref is None else source_ref
    event = TimelineEvent(
        tenant_id="foton",
        customer_id=customer_id,
        event_type=TimelineEventType.EMAIL_MESSAGE,
        event_at=NOW - timedelta(days=10),
        source_system="mail_archive_stage2",
        source_id=f"event-{customer_id}",
        direction=TimelineDirection.INBOUND,
        summary=event_summary,
        match_status="strong_unique",
        created_at=NOW - timedelta(days=10),
    )
    opportunity = CustomerOpportunity(
        tenant_id="foton",
        customer_id=customer_id,
        opportunity_type=OpportunityType.AMO_DEAL,
        source_system="amo",
        source_id=f"lead-{customer_id}",
        title=offer_title,
        status="active",
        product_context={"products_of_interest": [offer_title]},
        opened_at=NOW - timedelta(days=30),
    )
    store = CustomerTimelineSQLiteStore(db, allowed_root=tmp_path)
    try:
        store.upsert_customer(
            CustomerIdentity(
                tenant_id="foton",
                customer_id=customer_id,
                identity_status=IdentityStatus.STRONG,
                display_name=display_name or f"Родитель {customer_id}",
                primary_phone=f"+790000000{number:02d}",
                primary_email=f"{customer_id.replace(':', '-')}@example.com",
                source_ref=resolved_source_ref or None,
                metadata={"brands": ["foton"], "no_contact": no_contact, "source_ref": resolved_source_ref},
            )
        )
        store.upsert_opportunity(opportunity)
        store.upsert_event(event)
        store.upsert_signal(
            DerivedSignal(
                tenant_id="foton",
                customer_id=customer_id,
                opportunity_id=opportunity.opportunity_id,
                event_id=event.event_id,
                signal_type=signal_type,
                severity=SignalSeverity.MEDIUM,
                evidence_text=f"Основание {signal_type} для {customer_id}.",
                recommended_action="Проверить историю и написать клиенту.",
                expires_at=NOW + timedelta(days=30),
                created_at=NOW - timedelta(days=10),
            )
        )
    finally:
        store.close()
    with sqlite3.connect(db) as con:
        signal_row = con.execute(
            "SELECT signal_id, record_json FROM derived_signals WHERE customer_id=?",
            (customer_id,),
        ).fetchone()
        signal_record = json.loads(signal_row[1])
        signal_record["follow_up_due_at"] = (NOW + timedelta(days=1)).isoformat()
        con.execute(
            "UPDATE derived_signals SET record_json=? WHERE signal_id=?",
            (json.dumps(signal_record, ensure_ascii=False), signal_row[0]),
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS family_links_v1 (
              tenant_id TEXT NOT NULL, family_id TEXT NOT NULL, customer_id TEXT NOT NULL,
              child_key TEXT NOT NULL, canonical_name TEXT NOT NULL, name_variants_json TEXT NOT NULL,
              grades_json TEXT NOT NULL, subjects_json TEXT NOT NULL, brand TEXT NOT NULL,
              status TEXT NOT NULL, confidence TEXT NOT NULL, reason TEXT NOT NULL,
              source_refs_json TEXT NOT NULL, evidence_count INTEGER NOT NULL, created_at TEXT NOT NULL,
              record_hash TEXT NOT NULL, record_json TEXT NOT NULL,
              PRIMARY KEY (tenant_id, family_id, customer_id, child_key)
            )
            """
        )
        con.execute(
            "INSERT OR REPLACE INTO family_members_v1 VALUES (?,?,?,?,?,?,?,?,?,?)",
            ("foton", family_id, customer_id, "confident", "high", "test", NOW.isoformat(), NOW.isoformat(), f"hash-{customer_id}", "{}"),
        )
        con.execute(
            "INSERT OR REPLACE INTO family_links_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (
                "foton", family_id, customer_id, f"child:{customer_id}", f"Ребёнок {customer_id}", "[]",
                json.dumps([grade]), json.dumps(["математика"], ensure_ascii=False), "foton", "confident", "high",
                "test", "[]", 1, NOW.isoformat(), f"hash-child-{customer_id}", "{}",
            ),
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_purchases_v1 (
              tenant_id TEXT NOT NULL, customer_id TEXT NOT NULL, period TEXT NOT NULL, money_kind TEXT NOT NULL,
              total_in REAL, total_out REAL, deals_cnt INTEGER NOT NULL DEFAULT 0, last_purchase_at TEXT,
              sources_json TEXT NOT NULL, computability TEXT NOT NULL, code_version TEXT NOT NULL,
              PRIMARY KEY (tenant_id, customer_id, period, money_kind)
            )
            """
        )
        con.execute(
            "INSERT OR REPLACE INTO customer_purchases_v1 VALUES (?,?,?,?,?,?,?,?,?,?,?)",
            ("foton", customer_id, "all_time", "fact", 50000, 0, 1, "2026-06-01", "[]", "computed", "test"),
        )
        con.commit()


def _canonical_calls_db(tmp_path: Path, transcripts: dict[str, str]) -> Path:
    db = tmp_path / "canonical_calls.sqlite"
    with sqlite3.connect(db) as con:
        con.execute("CREATE TABLE canonical_calls (canonical_call_id TEXT PRIMARY KEY, transcript_client TEXT)")
        con.executemany(
            "INSERT INTO canonical_calls (canonical_call_id, transcript_client) VALUES (?, ?)",
            sorted(transcripts.items()),
        )
        con.commit()
    return db


def _seed_full_dossier_tables(db: Path, *, signal_action: str = "Позвонить в понедельник по оплате") -> None:
    with sqlite3.connect(db) as con:
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS family_links_v1 (
              tenant_id TEXT NOT NULL,
              family_id TEXT NOT NULL,
              customer_id TEXT NOT NULL,
              child_key TEXT NOT NULL,
              canonical_name TEXT NOT NULL,
              name_variants_json TEXT NOT NULL,
              grades_json TEXT NOT NULL,
              subjects_json TEXT NOT NULL,
              brand TEXT NOT NULL,
              status TEXT NOT NULL,
              confidence TEXT NOT NULL,
              reason TEXT NOT NULL,
              source_refs_json TEXT NOT NULL,
              evidence_count INTEGER NOT NULL,
              created_at TEXT NOT NULL,
              record_hash TEXT NOT NULL,
              record_json TEXT NOT NULL,
              PRIMARY KEY (tenant_id, family_id, customer_id, child_key)
            )
            """
        )
        con.execute(
            """
            INSERT OR REPLACE INTO family_links_v1
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "foton",
                "family:1",
                "customer:1",
                "child:1",
                "Иван",
                json.dumps(["Иван"], ensure_ascii=False),
                json.dumps(["8"], ensure_ascii=False),
                json.dumps(["математика"], ensure_ascii=False),
                "foton",
                "confident",
                "high",
                "single_child_family",
                json.dumps(["test"], ensure_ascii=False),
                1,
                NOW.isoformat(),
                "hash-family",
                "{}",
            ),
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_purchases_v1 (
              tenant_id TEXT NOT NULL,
              customer_id TEXT NOT NULL,
              period TEXT NOT NULL,
              money_kind TEXT NOT NULL,
              total_in REAL,
              total_out REAL,
              deals_cnt INTEGER NOT NULL DEFAULT 0,
              last_purchase_at TEXT,
              sources_json TEXT NOT NULL,
              computability TEXT NOT NULL,
              code_version TEXT NOT NULL,
              PRIMARY KEY (tenant_id, customer_id, period, money_kind)
            )
            """
        )
        con.executemany(
            """
            INSERT OR REPLACE INTO customer_purchases_v1
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                ("foton", "customer:1", "all_time", "fact", 120000, 0, 1, "2026-06-01", "[]", "ok", "test"),
                ("foton", "customer:1", "all_time", "plan", 80000, 0, 1, None, "[]", "ok", "test"),
            ],
        )
        con.execute(
            """
            INSERT OR REPLACE INTO derived_signals
            (signal_id, tenant_id, customer_id, opportunity_id, event_id, signal_type, severity, status, expires_at,
             confidence, requires_manager_review, created_at, record_hash, record_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "signal:1",
                "foton",
                "customer:1",
                None,
                None,
                "deal_stalling",
                "high",
                "active",
                "2026-12-31T00:00:00+00:00",
                0.9,
                1,
                NOW.isoformat(),
                "hash-signal",
                json.dumps({"recommended_action": signal_action, "evidence_text": "нет движения 7 дней"}, ensure_ascii=False),
            ),
        )
        con.execute(
            """
            CREATE TABLE IF NOT EXISTS customer_objections_v1 (
              tenant_id TEXT NOT NULL,
              customer_id TEXT NOT NULL,
              source_event_id TEXT NOT NULL,
              source_channel TEXT NOT NULL,
              objection_type TEXT NOT NULL,
              quote_preview TEXT NOT NULL,
              budget_hint_rub INTEGER,
              price_sensitivity TEXT NOT NULL,
              extracted_at TEXT NOT NULL,
              extractor_version TEXT NOT NULL,
              speaker TEXT NOT NULL DEFAULT 'unknown',
              direction TEXT NOT NULL DEFAULT 'unknown',
              confidence TEXT NOT NULL DEFAULT 'low',
              PRIMARY KEY (tenant_id, customer_id, source_event_id, objection_type)
            )
            """
        )
        con.execute(
            """
            INSERT OR REPLACE INTO customer_objections_v1
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                "foton",
                "customer:1",
                "event:email",
                "email",
                "price",
                "дорого, просит рассрочку",
                60000,
                "high",
                NOW.isoformat(),
                "test",
                "client",
                "inbound",
                "high",
            ),
        )
        con.execute(
            """
            UPDATE timeline_events
            SET subject = 'Расписание',
                summary = 'Требуется ручная проверка модельной выжимки: текст короткий'
            WHERE tenant_id = 'foton' AND customer_id = 'customer:1' AND event_type = 'email_message'
            """
        )
        con.commit()
